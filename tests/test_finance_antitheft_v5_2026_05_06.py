"""金流 A 錢守衛 v5（2026-05-06）回歸測試。

涵蓋本批次修補:
1. P2 報表 dashboard 薪資只採 finalized + 非 stale,草稿另列 pending
2. P2 export_all_salaries 預設只匯出已封存且非 stale,且含草稿時寫 explicit audit
3. P2 download_anomaly_report 改為依 (year, month) 即時從 DB 重算
4. P1 leaves/overtimes 來源異動在 commit 前同時 acquire lock + pre-mark stale,
   封住 caller commit 與 engine 重新 acquire lock 之間 finalize 搶先封存的 race
5. P1 attendance/meetings/shifts 的 mark_salary_stale 路徑改用 lock_and_premark_stale
"""

from __future__ import annotations

import os
import sys
from datetime import date, datetime
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
import api.salary as salary_module
from api.salary import router as salary_router
from api.reports import router as reports_router
from api.attendance import router as attendance_router
from api.auth import router as auth_router, _account_failures, _ip_attempts
from models.database import (
    Base,
    Employee,
    User,
    SalaryRecord,
    Attendance,
)
from utils.auth import hash_password


@pytest.fixture
def client(tmp_path):
    db_path = tmp_path / "v5.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = Session

    # seed admin user + employees
    s = Session()
    admin_emp = Employee(
        name="Admin",
        employee_id="A001",
        position="admin",
        is_active=True,
        base_salary=50000,
        hire_date=date(2020, 1, 1),
    )
    s.add(admin_emp)
    s.flush()
    admin_user = User(
        username="admin",
        password_hash=hash_password("admin123"),
        role="admin",
        employee_id=admin_emp.id,
        is_active=True,
    )
    s.add(admin_user)

    emp1 = Employee(
        name="Alice",
        employee_id="E001",
        position="teacher",
        is_active=True,
        base_salary=30000,
        hire_date=date(2024, 1, 1),
    )
    emp2 = Employee(
        name="Bob",
        employee_id="E002",
        position="teacher",
        is_active=True,
        base_salary=32000,
        hire_date=date(2024, 1, 1),
    )
    s.add_all([emp1, emp2])
    s.flush()

    # 三筆薪資:封存非 stale / 封存但 stale / 草稿
    s.add_all(
        [
            SalaryRecord(
                employee_id=emp1.id,
                salary_year=2026,
                salary_month=4,
                base_salary=30000,
                gross_salary=30000,
                net_salary=27000,
                total_deduction=3000,
                is_finalized=True,
                needs_recalc=False,
            ),
            SalaryRecord(
                employee_id=emp2.id,
                salary_year=2026,
                salary_month=4,
                base_salary=32000,
                gross_salary=32000,
                net_salary=29000,
                total_deduction=3000,
                is_finalized=True,
                needs_recalc=True,  # stale
            ),
            SalaryRecord(
                employee_id=admin_emp.id,
                salary_year=2026,
                salary_month=4,
                base_salary=50000,
                gross_salary=50000,
                net_salary=45000,
                total_deduction=5000,
                is_finalized=False,  # 草稿
                needs_recalc=False,
            ),
        ]
    )
    s.commit()

    # 異常考勤
    s.add(
        Attendance(
            employee_id=emp1.id,
            attendance_date=date(2026, 4, 3),
            is_late=True,
            late_minutes=15,
            status="late",
        )
    )
    s.add(
        Attendance(
            employee_id=emp2.id,
            attendance_date=date(2026, 4, 5),
            is_missing_punch_in=True,
            status="missing_punch",
        )
    )
    s.commit()
    s.close()

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)
    app.include_router(reports_router)
    app.include_router(attendance_router)

    _account_failures.clear()
    _ip_attempts.clear()

    with TestClient(app) as c:
        res = c.post(
            "/api/auth/login", json={"username": "admin", "password": "admin123"}
        )
        assert res.status_code == 200, res.text
        # cookie-based auth：access_token 已存在 TestClient cookie jar
        yield c, Session

    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory


class TestReportsDashboardFinalizedOnly:
    def test_salary_monthly_only_counts_finalized_non_stale(self, client):
        c, _ = client
        res = c.get("/api/reports/dashboard?year=2026")
        assert res.status_code == 200, res.text
        body = res.json()
        salary_monthly = body["salary_monthly"]
        # 只有 emp1 (finalized + 非 stale) 進總額;emp2 stale + admin 草稿須排除
        apr = next((m for m in salary_monthly if m["month"] == 4), None)
        assert apr is not None, "4 月必須有資料"
        assert (
            apr["employee_count"] == 1
        ), f"封存非 stale 只有 1 人(emp1),實際 {apr['employee_count']}"
        assert apr["total_gross"] == 30000
        # pending = stale(emp2) + 草稿(admin) 共 2
        assert apr["employee_count_pending"] == 2


class TestExportAllSalariesFiltered:
    def test_default_excludes_pending(self, client):
        c, _ = client
        res = c.get("/api/salaries/export-all?year=2026&month=4&format=xlsx")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # header + 1 finalized non-stale
        body_text = "\n".join(str(r) for r in rows)
        assert "Alice" in body_text, "封存非 stale 員工應出現"
        assert "Bob" not in body_text, "stale 員工預設不應匯出"
        assert "Admin" not in body_text, "草稿薪資預設不應匯出"

    def test_include_pending_returns_all_and_audits(self, client, monkeypatch):
        c, _ = client
        captured = []

        def fake_audit(request, **kwargs):
            captured.append(kwargs)

        monkeypatch.setattr(
            "utils.audit.write_explicit_audit", fake_audit, raising=True
        )
        res = c.get(
            "/api/salaries/export-all?year=2026&month=4&format=xlsx&include_pending=true"
        )
        assert res.status_code == 200, res.text
        # write_explicit_audit 必須被呼叫
        assert any(
            kw.get("action") == "EXPORT" and kw.get("entity_type") == "salary"
            for kw in captured
        ), f"含 pending 必須留 audit,實際 {captured}"
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        body_text = "\n".join(str(r) for r in ws.iter_rows(values_only=True))
        assert "Alice" in body_text
        assert "Bob" in body_text
        assert "Admin" in body_text


class TestAnomalyReportPerRequest:
    def test_anomaly_report_filters_by_year_month(self, client):
        c, _ = client
        res = c.get("/api/attendance/anomaly-report?year=2026&month=4")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        body_text = "\n".join(str(r) for r in ws.iter_rows(values_only=True))
        # 4 月有 emp1 遲到 + emp2 缺打卡
        assert "Alice" in body_text
        assert "Bob" in body_text
        # 5 月當月應為空(只有表頭)
        res2 = c.get("/api/attendance/anomaly-report?year=2026&month=5")
        assert res2.status_code == 200
        wb2 = load_workbook(BytesIO(res2.content))
        ws2 = wb2.active
        rows2 = list(ws2.iter_rows(values_only=True))
        assert len(rows2) == 1, "5 月無異常,應只剩表頭"

    def test_anomaly_report_requires_year_month(self, client):
        c, _ = client
        res = c.get("/api/attendance/anomaly-report")
        assert res.status_code == 422, "year/month 為必填"


class TestLockAndPremarkStaleHelper:
    def test_marks_stale_for_unfinalized_record(self, client):
        c, Session = client
        s = Session()
        try:
            from services.salary.utils import lock_and_premark_stale

            admin_emp = s.query(Employee).filter(Employee.name == "Admin").first()
            assert admin_emp is not None
            lock_and_premark_stale(s, admin_emp.id, {(2026, 4)})
            s.commit()

            rec = (
                s.query(SalaryRecord)
                .filter(
                    SalaryRecord.employee_id == admin_emp.id,
                    SalaryRecord.salary_year == 2026,
                    SalaryRecord.salary_month == 4,
                )
                .first()
            )
            assert rec is not None
            assert rec.needs_recalc is True, "草稿薪資應被預標 stale"
        finally:
            s.close()

    def test_does_not_overwrite_finalized(self, client):
        c, Session = client
        s = Session()
        try:
            from services.salary.utils import lock_and_premark_stale

            emp1 = s.query(Employee).filter(Employee.name == "Alice").first()
            assert emp1 is not None
            lock_and_premark_stale(s, emp1.id, {(2026, 4)})
            s.commit()

            rec = (
                s.query(SalaryRecord)
                .filter(SalaryRecord.employee_id == emp1.id)
                .first()
            )
            # 已 finalized 的 record 不會被標 stale (mark_salary_stale 自然跳過)
            assert rec.is_finalized is True
            assert rec.needs_recalc is False
        finally:
            s.close()
