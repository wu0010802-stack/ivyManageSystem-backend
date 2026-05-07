"""
Reports router - aggregated statistics for dashboard charts
"""

import logging
from datetime import date
from collections import defaultdict
from typing import Optional

from fastapi import APIRouter, Depends, Query
from sqlalchemy import extract, func, Integer

from models.base import session_scope
from models.database import (
    get_session,
    Attendance,
    Employee,
    Classroom,
    LeaveRecord,
    SalaryRecord,
)
from services.finance_report_service import build_finance_detail, build_finance_summary
from services.report_cache_service import report_cache_service
from utils.auth import require_staff_permission
from utils.excel_utils import SafeWorksheet, xlsx_streaming_response
from utils.permissions import Permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports", tags=["reports"])
REPORT_DASHBOARD_CACHE_TTL_SECONDS = 1800


def _query_attendance_monthly(session, start: date, end: date) -> list:
    """查詢並整理年度每月考勤統計（正常/遲到/早退/漏打卡）。"""
    rows = (
        session.query(
            extract("month", Attendance.attendance_date).label("month"),
            func.count(Attendance.id).label("total"),
            func.sum(func.cast(Attendance.is_late, Integer)).label("late"),
            func.sum(func.cast(Attendance.is_early_leave, Integer)).label(
                "early_leave"
            ),
            func.sum(
                func.cast(Attendance.is_missing_punch_in, Integer)
                + func.cast(Attendance.is_missing_punch_out, Integer)
            ).label("missing"),
        )
        .filter(
            Attendance.attendance_date >= start,
            Attendance.attendance_date <= end,
        )
        .group_by("month")
        .order_by("month")
        .all()
    )

    result = []
    for row in rows:
        total = int(row.total)
        late = int(row.late or 0)
        early = int(row.early_leave or 0)
        missing = int(row.missing or 0)
        anomaly = late + early + missing
        rate = round((total - anomaly) / total * 100, 1) if total > 0 else 0
        result.append(
            {
                "month": int(row.month),
                "total_records": total,
                "normal": total - anomaly,
                "late": late,
                "early_leave": early,
                "missing": missing,
                "rate": rate,
            }
        )
    return result


def _query_attendance_by_classroom(session, start: date, end: date) -> list:
    """查詢並整理各班級年度考勤出勤率。"""
    rows = (
        session.query(
            Classroom.name.label("classroom"),
            func.count(Attendance.id).label("total"),
            func.sum(func.cast(Attendance.is_late, Integer)).label("late"),
            func.sum(func.cast(Attendance.is_early_leave, Integer)).label(
                "early_leave"
            ),
        )
        .join(
            Employee,
            Employee.classroom_id == Classroom.id,
        )
        .join(
            Attendance,
            Attendance.employee_id == Employee.id,
        )
        .filter(
            Attendance.attendance_date >= start,
            Attendance.attendance_date <= end,
            Classroom.is_active == True,
        )
        .group_by(Classroom.name)
        .order_by(Classroom.name)
        .all()
    )

    result = []
    for row in rows:
        total = int(row.total)
        late = int(row.late or 0)
        early = int(row.early_leave or 0)
        anomaly = late + early
        rate = round((total - anomaly) / total * 100, 1) if total > 0 else 0
        result.append(
            {
                "classroom": row.classroom,
                "total_records": total,
                "late": late,
                "early_leave": early,
                "rate": rate,
            }
        )
    return result


def _query_leave_monthly(session, start: date, end: date) -> list:
    """查詢並整理年度每月各假別請假統計（12 個月完整列表）。"""
    _EMPTY_MONTH = {
        "personal": 0,
        "sick": 0,
        "annual": 0,
        "menstrual": 0,
        "maternity": 0,
        "paternity": 0,
        "total_hours": 0,
    }
    rows = (
        session.query(
            extract("month", LeaveRecord.start_date).label("month"),
            LeaveRecord.leave_type,
            func.count(LeaveRecord.id).label("count"),
            func.sum(LeaveRecord.leave_hours).label("total_hours"),
        )
        .filter(
            LeaveRecord.start_date >= start,
            LeaveRecord.start_date <= end,
        )
        .group_by("month", LeaveRecord.leave_type)
        .order_by("month")
        .all()
    )

    leave_by_month = defaultdict(lambda: dict(_EMPTY_MONTH))
    for row in rows:
        month = int(row.month)
        lt = row.leave_type
        if lt in leave_by_month[month]:
            leave_by_month[month][lt] = int(row.count)
        leave_by_month[month]["total_hours"] += float(row.total_hours or 0)

    return [
        {"month": m, **leave_by_month.get(m, dict(_EMPTY_MONTH))} for m in range(1, 13)
    ]


def _query_salary_monthly(session, year: int) -> list:
    """查詢並整理年度每月薪資彙總（總應發、實發、扣款、獎金）。

    報表趨勢只認封存且非 stale 的薪資（is_finalized=True AND needs_recalc=False）。
    Why: 草稿/待重算薪資是中間態,讓會計用測試重算的草稿影響管理層看到的薪資趨勢
    形同 A 錢空間。pending 統計另列 (employee_count_pending),前端可選擇是否顯示提示。
    """
    finalized_rows = (
        session.query(
            SalaryRecord.salary_month.label("month"),
            func.count(SalaryRecord.id).label("employee_count"),
            func.sum(SalaryRecord.gross_salary).label("total_gross"),
            func.sum(SalaryRecord.net_salary).label("total_net"),
            func.sum(SalaryRecord.total_deduction).label("total_deductions"),
            func.sum(
                func.coalesce(SalaryRecord.festival_bonus, 0)
                + func.coalesce(SalaryRecord.overtime_bonus, 0)
                + func.coalesce(SalaryRecord.performance_bonus, 0)
                + func.coalesce(SalaryRecord.special_bonus, 0)
            ).label("total_bonus"),
            func.sum(func.coalesce(SalaryRecord.overtime_pay, 0)).label(
                "total_overtime_pay"
            ),
        )
        .filter(
            SalaryRecord.salary_year == year,
            SalaryRecord.is_finalized == True,  # noqa: E712
            SalaryRecord.needs_recalc == False,  # noqa: E712
        )
        .group_by(SalaryRecord.salary_month)
        .order_by(SalaryRecord.salary_month)
        .all()
    )

    pending_rows = (
        session.query(
            SalaryRecord.salary_month.label("month"),
            func.count(SalaryRecord.id).label("pending_count"),
        )
        .filter(
            SalaryRecord.salary_year == year,
            (SalaryRecord.is_finalized == False)  # noqa: E712
            | (SalaryRecord.needs_recalc == True),  # noqa: E712
        )
        .group_by(SalaryRecord.salary_month)
        .all()
    )
    pending_map = {int(row.month): int(row.pending_count) for row in pending_rows}

    return [
        {
            "month": int(row.month),
            "employee_count": int(row.employee_count),
            "total_gross": round(float(row.total_gross or 0)),
            "total_net": round(float(row.total_net or 0)),
            "total_deductions": round(float(row.total_deductions or 0)),
            "total_bonus": round(float(row.total_bonus or 0)),
            "total_overtime_pay": round(float(row.total_overtime_pay or 0)),
            "employee_count_pending": pending_map.get(int(row.month), 0),
        }
        for row in finalized_rows
    ]


def _build_report_dashboard_data(session, year: int) -> dict:
    start = date(year, 1, 1)
    end = date(year, 12, 31)

    return {
        "year": year,
        "attendance_monthly": _query_attendance_monthly(session, start, end),
        "attendance_by_classroom": _query_attendance_by_classroom(session, start, end),
        "leave_monthly": _query_leave_monthly(session, start, end),
        "salary_monthly": _query_salary_monthly(session, year),
    }


@router.get("/dashboard")
def get_report_dashboard(
    year: int = Query(...),
    current_user: dict = Depends(require_staff_permission(Permission.REPORTS)),
):
    """取得年度報表統計資料"""
    session = get_session()
    try:
        return report_cache_service.get_or_build(
            session,
            category="reports_dashboard",
            ttl_seconds=REPORT_DASHBOARD_CACHE_TTL_SECONDS,
            params={"year": year},
            builder=lambda: _build_report_dashboard_data(session, year),
        )
    finally:
        session.close()


FINANCE_SUMMARY_CACHE_TTL_SECONDS = 1800  # 30 分鐘


@router.get("/finance-summary")
def get_finance_summary(
    year: int = Query(..., ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
    current_user: dict = Depends(require_staff_permission(Permission.REPORTS)),
):
    """園務統計：整合學費、才藝、薪資的月度收支彙總。

    快取：sub-category 為 reports_finance_summary，TTL 30 分。
    學費繳費/才藝 POS/薪資計算端可呼叫
    `report_cache_service.invalidate_category(session, "reports_finance_summary")`
    主動讓快取失效。
    """
    session = get_session()
    try:
        return report_cache_service.get_or_build(
            session,
            category="reports_finance_summary",
            ttl_seconds=FINANCE_SUMMARY_CACHE_TTL_SECONDS,
            params={"year": year, "month": month},
            builder=lambda: build_finance_summary(session, year=year, month=month),
        )
    finally:
        session.close()


@router.get("/finance-summary/detail")
def get_finance_summary_detail(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    current_user: dict = Depends(require_staff_permission(Permission.REPORTS)),
):
    """單月收支明細：學費 / 才藝 / 薪資三來源原始交易列表（下鑽用）。

    F-031：薪資逐員金額（gross/net/festival/overtime/employer_benefit/real_cost）
    屬 SALARY_READ 範疇；REPORTS 持有者若非 admin/hr 不可下鑽到逐員實發/獎金，
    此處依角色層遮罩 salary[] 欄位。festival_bonus / overtime_bonus 同樣屬個薪敏感
    （可推回個人總所得），必須一併遮罩。
    """
    from utils.salary_access import has_full_salary_view, mask_dict_fields

    with session_scope() as session:
        result = build_finance_detail(session, year=year, month=month)
        if not has_full_salary_view(current_user):
            result = dict(result)
            result["salary"] = [
                mask_dict_fields(
                    r,
                    (
                        "gross_salary",
                        "net_salary",
                        "festival_bonus",
                        "overtime_bonus",
                        "employer_benefit",
                        "real_cost",
                    ),
                    placeholder=None,
                )
                for r in result.get("salary", [])
            ]
        return result


@router.get("/finance-summary/export")
def export_finance_summary(
    year: int = Query(..., ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
    current_user: dict = Depends(require_staff_permission(Permission.REPORTS)),
):
    """匯出收支彙總為 Excel。

    指定 month 時，額外附三張明細分頁（學費、才藝、薪資）；
    未指定 month 時只輸出月度彙總與分類統計兩張分頁。

    F-031：薪資明細 Sheet 5 在非 admin/hr 角色下需以「—」遮罩金額欄位，
    避免 supervisor / 自訂 REPORTS 角色透過匯出取得逐員實發名冊。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from utils.salary_access import has_full_salary_view

    can_see_salary = has_full_salary_view(current_user)

    with session_scope() as session:
        summary = build_finance_summary(session, year=year, month=month)
        detail = (
            build_finance_detail(session, year=year, month=month)
            if month is not None
            else None
        )

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    header_align = Alignment(horizontal="center")

    def _write_header(ws: SafeWorksheet, headers: list[str]) -> None:
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    # Sheet 1：月度彙總
    ws1 = SafeWorksheet(wb.active)
    ws1.title = "月度彙總"
    _write_header(ws1, ["月", "收入", "退款", "支出", "淨現金"])
    for i, row in enumerate(summary["monthly_trend"], start=2):
        ws1.cell(row=i, column=1, value=f"{row['month']} 月")
        ws1.cell(row=i, column=2, value=row["revenue"])
        ws1.cell(row=i, column=3, value=row["refund"])
        ws1.cell(row=i, column=4, value=row["expense"])
        ws1.cell(row=i, column=5, value=row["net"])
    # 合計列
    total_row = len(summary["monthly_trend"]) + 2
    ws1.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws1.cell(row=total_row, column=2, value=summary["summary"]["total_revenue"])
    ws1.cell(row=total_row, column=3, value=summary["summary"]["total_refund"])
    ws1.cell(row=total_row, column=4, value=summary["summary"]["total_expense"])
    ws1.cell(row=total_row, column=5, value=summary["summary"]["net_cashflow"])

    # Sheet 2：分類
    ws2 = SafeWorksheet(wb.create_sheet("分類統計"))
    ws2.cell(row=1, column=1, value="收入分類").font = Font(bold=True)
    _write_header_row = 2
    ws2.cell(row=_write_header_row, column=1, value="類別")
    ws2.cell(row=_write_header_row, column=2, value="金額")
    ws2.cell(row=_write_header_row, column=3, value="退款")
    for i, c in enumerate(summary["revenue_by_category"], start=3):
        ws2.cell(row=i, column=1, value=c["label"])
        ws2.cell(row=i, column=2, value=c["amount"])
        ws2.cell(row=i, column=3, value=c.get("refund", 0))
    offset = len(summary["revenue_by_category"]) + 5
    ws2.cell(row=offset, column=1, value="支出分類").font = Font(bold=True)
    ws2.cell(row=offset + 1, column=1, value="類別")
    ws2.cell(row=offset + 1, column=2, value="金額")
    for i, c in enumerate(summary["expense_by_category"], start=offset + 2):
        ws2.cell(row=i, column=1, value=c["label"])
        ws2.cell(row=i, column=2, value=c["amount"])

    if detail is not None:
        # Sheet 3：學費明細
        ws3 = SafeWorksheet(wb.create_sheet("學費明細"))
        _write_header(
            ws3,
            ["類型", "日期", "學生", "班級", "費用項目", "金額", "付款方式", "備註"],
        )
        for i, r in enumerate(detail["tuition"], start=2):
            ws3.cell(
                row=i, column=1, value="繳費" if r["kind"] == "payment" else "退款"
            )
            ws3.cell(row=i, column=2, value=r.get("date"))
            ws3.cell(row=i, column=3, value=r.get("student_name"))
            ws3.cell(row=i, column=4, value=r.get("classroom_name"))
            ws3.cell(row=i, column=5, value=r.get("fee_item_name"))
            ws3.cell(row=i, column=6, value=r["amount"])
            ws3.cell(row=i, column=7, value=r.get("payment_method"))
            ws3.cell(row=i, column=8, value=r.get("reason"))

        # Sheet 4：才藝明細
        ws4 = SafeWorksheet(wb.create_sheet("才藝明細"))
        _write_header(
            ws4,
            ["類型", "日期", "學生", "金額", "付款方式", "操作人", "收據號"],
        )
        for i, r in enumerate(detail["activity"], start=2):
            ws4.cell(
                row=i, column=1, value="繳費" if r["kind"] == "payment" else "退費"
            )
            ws4.cell(row=i, column=2, value=r.get("date"))
            ws4.cell(row=i, column=3, value=r.get("student_name"))
            ws4.cell(row=i, column=4, value=r["amount"])
            ws4.cell(row=i, column=5, value=r.get("payment_method"))
            ws4.cell(row=i, column=6, value=r.get("operator"))
            ws4.cell(row=i, column=7, value=r.get("receipt_no"))

        # Sheet 5：薪資明細（金額僅 admin/hr 可見；F-031）
        ws5 = SafeWorksheet(wb.create_sheet("薪資明細"))
        _write_header(
            ws5,
            ["員工", "應發", "實發", "雇主保費+勞退", "園方真實支出", "已封存"],
        )
        for i, r in enumerate(detail["salary"], start=2):
            ws5.cell(row=i, column=1, value=r.get("employee_name"))
            if can_see_salary:
                ws5.cell(row=i, column=2, value=r["gross_salary"])
                ws5.cell(row=i, column=3, value=r["net_salary"])
                ws5.cell(row=i, column=4, value=r["employer_benefit"])
                ws5.cell(row=i, column=5, value=r["real_cost"])
            else:
                ws5.cell(row=i, column=2, value="—")
                ws5.cell(row=i, column=3, value="—")
                ws5.cell(row=i, column=4, value="—")
                ws5.cell(row=i, column=5, value="—")
            ws5.cell(row=i, column=6, value="是" if r["is_finalized"] else "")

    suffix = f"{year}" + (f"-{month:02d}" if month else "-全年")
    return xlsx_streaming_response(wb, f"收支彙總_{suffix}.xlsx")
