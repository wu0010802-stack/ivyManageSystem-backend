"""銀行轉帳名冊匯出測試（services/transfer_roster + api/salary/transfer_roster）。"""

import os
import sys
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import router as auth_router
from api.auth import _account_failures, _ip_attempts
from api.salary import init_salary_services
from api.salary import router as salary_router
from models.database import Base, Employee, SalaryRecord, SystemConfig, User
from services.salary.engine import SalaryEngine
from utils.auth import hash_password
from utils.permissions import Permission


@pytest.fixture
def roster_client(tmp_path):
    db_path = tmp_path / "roster.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    init_salary_services(SalaryEngine(load_from_db=False), MagicMock())

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _add_employee(
    session,
    *,
    employee_id: str,
    name: str,
    bank_account: str | None = "0727-979-000001",
    employee_type: str = "regular",
    skip_payroll_transfer: bool = False,
) -> Employee:
    emp = Employee(
        employee_id=employee_id,
        name=name,
        base_salary=36000,
        is_active=True,
        bank_account=bank_account,
        bank_account_name=name,
        employee_type=employee_type,
        skip_payroll_transfer=skip_payroll_transfer,
    )
    session.add(emp)
    session.flush()
    return emp


def _add_record(
    session,
    employee_id: int,
    *,
    year: int = 2026,
    month: int = 4,
    net_salary: float = 30000,
    festival_bonus: float = 0,
    overtime_bonus: float = 0,
    supervisor_dividend: float = 0,
    bonus_separate: bool = False,
    bonus_amount: float = 0,
    is_finalized: bool = True,
) -> SalaryRecord:
    rec = SalaryRecord(
        employee_id=employee_id,
        salary_year=year,
        salary_month=month,
        net_salary=net_salary,
        festival_bonus=festival_bonus,
        overtime_bonus=overtime_bonus,
        supervisor_dividend=supervisor_dividend,
        bonus_separate=bonus_separate,
        bonus_amount=bonus_amount,
        is_finalized=is_finalized,
    )
    session.add(rec)
    session.flush()
    return rec


def _login_as_salary_admin(client: TestClient, session_factory):
    with session_factory() as session:
        session.add(
            User(
                username="roster_admin",
                password_hash=hash_password("TempPass123"),
                role="admin",
                permissions=int(Permission.SALARY_READ),
                is_active=True,
                must_change_password=False,
            )
        )
        session.commit()
    res = client.post(
        "/api/auth/login",
        json={"username": "roster_admin", "password": "TempPass123"},
    )
    assert res.status_code == 200


def _load_xlsx(content: bytes):
    wb = load_workbook(BytesIO(content))
    return wb.active


class TestBaseSalaryRoster:
    def test_base_roster_happy_path(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            emp1 = _add_employee(session, employee_id="E001", name="王小明")
            emp2 = _add_employee(session, employee_id="E002", name="李大華")
            _add_record(session, emp1.id, net_salary=30000)
            _add_record(session, emp2.id, net_salary=25000)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        assert res.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        ws = _load_xlsx(res.content)
        assert ws["B2"].value == "115年04月 薪資轉帳名冊"
        # Row 5+ 為員工資料
        assert ws.cell(row=5, column=2).value == "王小明"
        assert ws.cell(row=5, column=3).value == 30000
        assert ws.cell(row=6, column=2).value == "李大華"
        assert ws.cell(row=6, column=3).value == 25000
        # 合計
        assert ws.cell(row=7, column=2).value == "合計"
        assert ws.cell(row=7, column=3).value == 55000

    def test_bonus_separate_does_not_deduct_from_base(self, roster_client):
        """REGRESSION: base 名冊金額 = net_salary，即使 bonus_separate=True 也不扣 bonus_amount。

        net_salary = gross_salary - total_deduction，gross_salary 不含 festival_bonus /
        overtime_bonus（見 engine.py:1764-1770、totals.py:21-30），所以與 festival/surplus
        名冊不會重複入帳；舊邏輯多扣 bonus_amount 會讓有獨立獎金的員工被少付（特別是
        supervisor_dividend 漏付，festival/overtime 被重複扣一次）。
        """
        client, session_factory = roster_client
        with session_factory() as session:
            emp = _add_employee(session, employee_id="E001", name="王雅玲")
            # 模擬主管：net_salary 已含 supervisor_dividend (4000)；festival 8000 走獨立名冊
            _add_record(
                session,
                emp.id,
                net_salary=44044,
                festival_bonus=8000,
                supervisor_dividend=4000,
                bonus_separate=True,
                bonus_amount=12000,  # festival 8000 + supervisor 4000
            )
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        # base 名冊 = net_salary 全額（含 supervisor_dividend），不扣 bonus_amount
        assert ws.cell(row=5, column=3).value == 44044

    def test_supervisor_dividend_stays_in_base_roster(self, roster_client):
        """REGRESSION: 只有主管紅利、無 festival/overtime 的員工，base 名冊應拿到完整 net_salary。

        舊邏輯下 bonus_separate=True 但 festival 名冊不會列此員工（festival_bonus=0），
        員工會短少 supervisor_dividend 整額。
        """
        client, session_factory = roster_client
        with session_factory() as session:
            emp = _add_employee(session, employee_id="E001", name="主管 A")
            _add_record(
                session,
                emp.id,
                net_salary=50000,  # 已含 supervisor_dividend 5000
                supervisor_dividend=5000,
                bonus_separate=True,
                bonus_amount=5000,
            )
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws.cell(row=5, column=3).value == 50000

        # 同月也不應出現在 festival/surplus 名冊（festival_bonus=0、overtime_bonus=0）
        res_f = client.get("/api/salaries/2026/4/transfer-roster?type=festival")
        assert res_f.status_code == 200
        ws_f = _load_xlsx(res_f.content)
        # row 5 直接是合計（沒人入榜）
        assert ws_f.cell(row=5, column=2).value == "合計"

    def test_base_plus_independent_rosters_equal_total_payable(self, roster_client):
        """REGRESSION: base + festival + surplus 三張名冊合計 = 員工實際應收總額。

        員工應收 = net_salary + festival_bonus + overtime_bonus（festival/overtime 走獨立轉帳）。
        舊邏輯下 base 多扣 bonus_amount → 三張名冊合計 < 應收。
        """
        client, session_factory = roster_client
        with session_factory() as session:
            emp = _add_employee(session, employee_id="E001", name="完整案例")
            _add_record(
                session,
                emp.id,
                net_salary=44000,  # gross - 扣款，已含 supervisor 4000
                festival_bonus=15000,
                overtime_bonus=3000,
                supervisor_dividend=4000,
                bonus_separate=True,
                bonus_amount=22000,  # festival 15000 + overtime 3000 + supervisor 4000
            )
            session.commit()

        _login_as_salary_admin(client, session_factory)
        base_amt = (
            _load_xlsx(
                client.get("/api/salaries/2026/4/transfer-roster?type=base").content
            )
            .cell(row=5, column=3)
            .value
        )
        festival_amt = (
            _load_xlsx(
                client.get("/api/salaries/2026/4/transfer-roster?type=festival").content
            )
            .cell(row=5, column=3)
            .value
        )
        surplus_amt = (
            _load_xlsx(
                client.get("/api/salaries/2026/4/transfer-roster?type=surplus").content
            )
            .cell(row=5, column=3)
            .value
        )

        # 三張名冊加總 = 員工實際應收 = net_salary + festival + overtime
        assert base_amt + festival_amt + surplus_amt == 44000 + 15000 + 3000

    def test_zero_amount_employee_excluded(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            emp_paid = _add_employee(session, employee_id="E001", name="王小明")
            emp_unpaid = _add_employee(session, employee_id="E002", name="無薪假員工")
            _add_record(session, emp_paid.id, net_salary=30000)
            _add_record(session, emp_unpaid.id, net_salary=0)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws.cell(row=5, column=2).value == "王小明"
        assert ws.cell(row=6, column=2).value == "合計"  # 下一行就是合計
        assert ws.cell(row=6, column=3).value == 30000

    def test_no_bank_account_employee_skipped(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            emp_ok = _add_employee(session, employee_id="E001", name="有帳號")
            emp_no = _add_employee(
                session, employee_id="E002", name="無帳號", bank_account=None
            )
            _add_record(session, emp_ok.id, net_salary=30000)
            _add_record(session, emp_no.id, net_salary=25000)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws.cell(row=5, column=2).value == "有帳號"
        assert ws.cell(row=6, column=2).value == "合計"
        assert ws.cell(row=6, column=3).value == 30000

    def test_skip_payroll_transfer_employee_excluded(self, roster_client):
        """skip_payroll_transfer=True 員工不入轉帳名冊（薪資仍計算，但用現金/其他管道）。"""
        client, session_factory = roster_client
        with session_factory() as session:
            emp_normal = _add_employee(session, employee_id="E001", name="正常員工")
            emp_skip = _add_employee(
                session,
                employee_id="E002",
                name="不薪轉員工",
                skip_payroll_transfer=True,
            )
            _add_record(session, emp_normal.id, net_salary=30000)
            _add_record(session, emp_skip.id, net_salary=25000)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws.cell(row=5, column=2).value == "正常員工"
        assert ws.cell(row=6, column=2).value == "合計"
        assert ws.cell(row=6, column=3).value == 30000

    def test_non_finalized_record_excluded(self, roster_client):
        """草稿薪資不應出現在轉帳名冊（避免拿草稿給銀行）。"""
        client, session_factory = roster_client
        with session_factory() as session:
            emp_final = _add_employee(session, employee_id="E001", name="已封存")
            emp_draft = _add_employee(session, employee_id="E002", name="草稿")
            _add_record(session, emp_final.id, net_salary=30000, is_finalized=True)
            _add_record(session, emp_draft.id, net_salary=25000, is_finalized=False)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws.cell(row=5, column=2).value == "已封存"
        assert ws.cell(row=6, column=2).value == "合計"


class TestFestivalAndSurplusRoster:
    def test_festival_uses_festival_bonus_column(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            emp = _add_employee(session, employee_id="E001", name="王雅玲")
            _add_record(session, emp.id, net_salary=48044, festival_bonus=8830)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=festival")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws["B2"].value == "115年04月 節慶獎金轉帳名冊"
        assert ws.cell(row=5, column=3).value == 8830

    def test_surplus_uses_overtime_bonus_column(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            emp = _add_employee(session, employee_id="E001", name="林佳穎")
            _add_record(session, emp.id, net_salary=36624, overtime_bonus=7094)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=surplus")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws["B2"].value == "115年04月 超額獎金轉帳名冊"
        assert ws.cell(row=5, column=3).value == 7094


class TestArtTeacherRoster:
    def test_art_teacher_only_includes_hourly_employees(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            regular = _add_employee(session, employee_id="E001", name="正職")
            art = _add_employee(
                session, employee_id="E100", name="才藝老師", employee_type="hourly"
            )
            _add_record(session, regular.id, net_salary=30000)
            _add_record(session, art.id, net_salary=13750)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=art_teacher")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws["B2"].value == "115年04月 才藝老師轉帳名冊"
        assert ws.cell(row=5, column=2).value == "才藝老師"
        assert ws.cell(row=5, column=3).value == 13750
        assert ws.cell(row=6, column=2).value == "合計"


class TestConfigAndValidation:
    def test_payer_account_from_system_config(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            session.add(
                SystemConfig(
                    config_key="bank.payer_account", config_value="9999-888-77777"
                )
            )
            session.add(
                SystemConfig(config_key="bank.payer_name", config_value="測試園所")
            )
            emp = _add_employee(session, employee_id="E001", name="王小明")
            _add_record(session, emp.id, net_salary=30000)
            session.commit()

        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        ws = _load_xlsx(res.content)
        assert ws["B1"].value == "測試園所"
        assert ws["B3"].value == "帳號：9999-888-77777"

    def test_invalid_type_rejected(self, roster_client):
        client, session_factory = roster_client
        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/4/transfer-roster?type=foo")
        assert res.status_code == 400

    def test_invalid_month_rejected(self, roster_client):
        client, session_factory = roster_client
        _login_as_salary_admin(client, session_factory)
        res = client.get("/api/salaries/2026/13/transfer-roster?type=base")
        assert res.status_code == 400

    def test_requires_salary_read_permission(self, roster_client):
        client, session_factory = roster_client
        with session_factory() as session:
            session.add(
                User(
                    username="no_perm",
                    password_hash=hash_password("TempPass123"),
                    role="staff",
                    permissions=0,
                    is_active=True,
                    must_change_password=False,
                )
            )
            session.commit()
        client.post(
            "/api/auth/login",
            json={"username": "no_perm", "password": "TempPass123"},
        )
        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code in (401, 403)
