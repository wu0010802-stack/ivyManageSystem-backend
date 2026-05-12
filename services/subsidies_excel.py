"""特教加給 Excel 匯出（Phase 4B）。"""

from __future__ import annotations

import io
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font

from utils.excel_utils import sanitize_excel_value

HEADERS = [
    "申領類型",
    "員工",
    "起期",
    "迄期",
    "時數/費率",
    "申請金額",
    "核定金額",
    "狀態",
    "備註",
]

TYPE_LABEL = {"teacher_extra": "特教加給", "assistant_hourly": "助理鐘點費"}


def generate_subsidies_excel(
    rows: Iterable[Mapping], *, period_label: str = ""
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "特教加給"

    if period_label:
        ws.append([f"期間：{period_label}"])
        ws.row_dimensions[1].font = Font(bold=True)
        ws.append([])

    ws.append(HEADERS)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for r in rows:
        ws.append(
            [
                sanitize_excel_value(
                    TYPE_LABEL.get(r["subsidy_type"], r["subsidy_type"])
                ),
                sanitize_excel_value(r.get("employee_name", "")),
                r.get("period_start"),
                r.get("period_end"),
                r.get("hours_or_rate"),
                r.get("amount_requested"),
                r.get("amount_approved"),
                sanitize_excel_value(r.get("status", "")),
                sanitize_excel_value(r.get("notes") or ""),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
