"""
Data export router - Excel downloads for employees, students, attendance, calendar
"""

import io
import logging
import calendar as cal_module
from datetime import date, timedelta, time
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from utils.auth import require_staff_permission
from utils.constants import LEAVE_TYPE_LABELS, OVERTIME_TYPE_LABELS
from utils.error_messages import EMPLOYEE_DOES_NOT_EXIST
from utils.masking import mask_bank_account
from utils.permissions import Permission, has_permission
from utils.rate_limit import SlidingWindowLimiter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from models.database import (
    get_session,
    Employee,
    Student,
    Attendance,
    Classroom,
    SchoolEvent,
    JobTitle,
    LeaveRecord,
    OvertimeRecord,
    Holiday,
    ShiftAssignment,
    ShiftType,
)
from services.official_calendar import build_admin_calendar_feed

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/exports", tags=["exports"])

# 匯出端點限流：同一 IP 每分鐘最多 5 次（匯出屬於重資源操作，防 DoS 消耗）
_export_rate_limit = SlidingWindowLimiter(
    max_calls=5,
    window_seconds=60,
    name="export",
    error_detail="匯出過於頻繁，請稍後再試",
).as_dependency()

# ============ Shared Styles ============

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center")


def _write_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


# Excel 公式注入防護：統一由 utils.excel_utils 提供；此處 re-export 維持既有 import 路徑。
from utils.excel_utils import (
    SafeWorksheet,
    sanitize_excel_value as _sanitize_excel_value,
    safe_ws as _safe_ws,
)


def _write_data_row(ws, row, values):
    for col, value in enumerate(values, 1):
        sanitized_value = _sanitize_excel_value(value)
        cell = ws.cell(row=row, column=col, value=sanitized_value)
        cell.border = THIN_BORDER


def _auto_width(ws):
    from openpyxl.cell.cell import MergedCell

    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 8), 40)


def _to_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def _id_name_map(session, model):
    """建立 {id: name} 對照表"""
    return {obj.id: obj.name for obj in session.query(model).all()}


def _mask_bank_account(account: str | None) -> str:
    return mask_bank_account(account) or ""


# ============ Employees ============


@router.get("/employees")
def export_employees(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.EMPLOYEES_READ)),
):
    """匯出員工名冊 Excel"""
    session = get_session()
    try:
        employees = list(
            session.query(Employee).order_by(Employee.employee_id).yield_per(500)
        )
        classrooms = _id_name_map(session, Classroom)
        job_titles = _id_name_map(session, JobTitle)
        can_view_full_account = has_permission(
            current_user.get("permissions", 0), Permission.SALARY_WRITE
        )
        if can_view_full_account:
            logger.warning(
                "員工名冊含完整銀行帳號已匯出 by user=%s role=%s",
                current_user.get("username"),
                current_user.get("role"),
            )

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = "員工名冊"

        ws.merge_cells("A1:O1")
        ws["A1"] = "員工名冊"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "工號",
            "姓名",
            "職稱",
            "職務",
            "員工類型",
            "所屬班級",
            "到職日期",
            "聯絡電話",
            "地址",
            "緊急聯絡人",
            "緊急聯絡人電話",
            "銀行代碼",
            "銀行帳號",
            "戶名",
            "在職狀態",
        ]
        _write_header_row(ws, 3, headers)

        for idx, emp in enumerate(employees, 4):
            jt = job_titles.get(emp.job_title_id, emp.title or "")
            cr = classrooms.get(emp.classroom_id, "")
            emp_type = "正職" if emp.employee_type == "regular" else "時薪"
            _write_data_row(
                ws,
                idx,
                [
                    emp.employee_id,
                    emp.name,
                    jt,
                    emp.position or "",
                    emp_type,
                    cr,
                    emp.hire_date.isoformat() if emp.hire_date else "",
                    emp.phone or "",
                    emp.address or "",
                    emp.emergency_contact_name or "",
                    emp.emergency_contact_phone or "",
                    emp.bank_code or "",
                    (
                        emp.bank_account or ""
                        if can_view_full_account
                        else _mask_bank_account(emp.bank_account)
                    ),
                    emp.bank_account_name or "",
                    "在職" if emp.is_active else "離職",
                ],
            )

        _auto_width(ws)
        return _to_response(wb, "員工名冊.xlsx")
    finally:
        session.close()


# ============ Students ============


@router.get("/students")
def export_students(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """匯出學生名冊 Excel"""
    session = get_session()
    try:
        students = list(
            session.query(Student).order_by(Student.student_id).yield_per(500)
        )
        classrooms = _id_name_map(session, Classroom)

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = "學生名冊"

        ws.merge_cells("A1:K1")
        ws["A1"] = "學生名冊"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "學號",
            "姓名",
            "性別",
            "生日",
            "班級",
            "入學日期",
            "家長姓名",
            "家長電話",
            "地址",
            "狀態標籤",
            "在籍狀態",
        ]
        _write_header_row(ws, 3, headers)

        gender_map = {"M": "男", "F": "女"}
        for idx, stu in enumerate(students, 4):
            cr = classrooms.get(stu.classroom_id, "")
            _write_data_row(
                ws,
                idx,
                [
                    stu.student_id,
                    stu.name,
                    gender_map.get(stu.gender, stu.gender or ""),
                    stu.birthday.isoformat() if stu.birthday else "",
                    cr,
                    stu.enrollment_date.isoformat() if stu.enrollment_date else "",
                    stu.parent_name or "",
                    stu.parent_phone or "",
                    stu.address or "",
                    stu.status_tag or "",
                    "在籍" if stu.is_active else "離校",
                ],
            )

        _auto_width(ws)
        return _to_response(wb, "學生名冊.xlsx")
    finally:
        session.close()


# ============ Attendance ============


@router.get("/attendance")
def export_attendance(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.ATTENDANCE_READ)),
    year: int = Query(...),
    month: int = Query(...),
):
    """匯出出勤月報 Excel"""
    session = get_session()
    try:
        _, last_day = cal_module.monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)

        employees = list(
            session.query(Employee)
            .filter(Employee.is_active == True)
            .order_by(Employee.employee_id)
            .yield_per(500)
        )
        emp_map = {e.id: e for e in employees}

        # 1. 計算當月應出勤天數（排除週末與國定假日）
        holiday_dates = {
            h.date
            for h in session.query(Holiday)
            .filter(
                Holiday.date >= start,
                Holiday.date <= end,
                Holiday.is_active.is_(True),
            )
            .all()
        }
        expected_workdays: set = set()
        cur = start
        while cur <= end:
            if cur.weekday() < 5 and cur not in holiday_dates:
                expected_workdays.add(cur)
            cur += timedelta(days=1)

        # 2. 已核准請假日期，用來區分「合法缺席」與「曠職」
        approved_leaves = (
            session.query(LeaveRecord)
            .filter(
                LeaveRecord.start_date <= end,
                LeaveRecord.end_date >= start,
                LeaveRecord.is_approved == True,
            )
            .all()
        )
        leave_dates_by_emp: dict = {}
        for lv in approved_leaves:
            if lv.employee_id not in emp_map:
                continue
            d = max(lv.start_date, start)
            lv_end = min(lv.end_date, end)
            while d <= lv_end:
                leave_dates_by_emp.setdefault(lv.employee_id, set()).add(d)
                d += timedelta(days=1)

        # 3. 從打卡記錄彙整統計，同時記錄每位員工的實際出勤日期集合
        records = (
            session.query(Attendance)
            .filter(
                Attendance.attendance_date >= start,
                Attendance.attendance_date <= end,
            )
            .all()
        )

        stats: dict = {}
        att_dates_by_emp: dict = {}
        for att in records:
            if att.employee_id not in emp_map:
                continue
            if att.employee_id not in stats:
                stats[att.employee_id] = {
                    "total": 0,
                    "normal": 0,
                    "late": 0,
                    "early": 0,
                    "missing_in": 0,
                    "missing_out": 0,
                    "late_min": 0,
                }
                att_dates_by_emp[att.employee_id] = set()
            s = stats[att.employee_id]
            s["total"] += 1
            att_dates_by_emp[att.employee_id].add(att.attendance_date)
            is_clean_attendance = (
                not att.is_late
                and not att.is_early_leave
                and not att.is_missing_punch_in
                and not att.is_missing_punch_out
            )
            if is_clean_attendance:
                s["normal"] += 1
            if att.is_late:
                s["late"] += 1
                s["late_min"] += att.late_minutes or 0
            if att.is_early_leave:
                s["early"] += 1
            if att.is_missing_punch_in:
                s["missing_in"] += 1
            if att.is_missing_punch_out:
                s["missing_out"] += 1

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年{month}月出勤月報"

        ws.merge_cells("A1:L1")
        ws["A1"] = f"{year}年{month}月 出勤月報"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "工號",
            "姓名",
            "應出勤天數",
            "實際出勤天數",
            "正常天數",
            "遲到次數",
            "早退次數",
            "缺卡(上班)",
            "缺卡(下班)",
            "遲到總分鐘",
            "請假天數",
            "曠職天數",
        ]
        _write_header_row(ws, 3, headers)

        row_idx = 4
        for emp in employees:
            s = stats.get(
                emp.id,
                {
                    "total": 0,
                    "normal": 0,
                    "late": 0,
                    "early": 0,
                    "missing_in": 0,
                    "missing_out": 0,
                    "late_min": 0,
                },
            )
            att_dates = att_dates_by_emp.get(emp.id, set())
            leave_dates = leave_dates_by_emp.get(emp.id, set())
            # 請假天數：已核准請假日 ∩ 應出勤日（排除本來就是假日的天）
            leave_days = len(leave_dates & expected_workdays)
            # 曠職天數：應出勤日 - 有打卡日 - 已核准請假日
            absent_days = len(expected_workdays - att_dates - leave_dates)
            _write_data_row(
                ws,
                row_idx,
                [
                    emp.employee_id,
                    emp.name,
                    len(expected_workdays),
                    s["total"],
                    s["normal"],
                    s["late"],
                    s["early"],
                    s["missing_in"],
                    s["missing_out"],
                    s["late_min"],
                    leave_days,
                    absent_days,
                ],
            )
            row_idx += 1

        _auto_width(ws)
        return _to_response(wb, f"{year}年{month}月出勤月報.xlsx")
    finally:
        session.close()


# ============ Calendar ============

EVENT_TYPE_LABELS = {
    "meeting": "會議",
    "activity": "活動",
    "holiday": "假日",
    "makeup_workday": "補班日",
    "general": "一般",
}


@router.get("/calendar")
def export_calendar(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.CALENDAR)),
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
):
    """匯出行事曆 Excel"""
    session = get_session()
    try:
        feed = build_admin_calendar_feed(session, year, month)
        events = feed["events"]

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年{month}月行事曆"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"{year}年{month}月 學校行事曆"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = ["日期", "結束日期", "類型", "標題", "說明", "時間", "地點"]
        _write_header_row(ws, 3, headers)

        for idx, ev in enumerate(events, 4):
            time_str = ""
            if ev["is_all_day"]:
                time_str = "全天"
            elif ev["start_time"] and ev["end_time"]:
                time_str = f"{ev['start_time']} - {ev['end_time']}"

            _write_data_row(
                ws,
                idx,
                [
                    ev["event_date"],
                    ev["end_date"] or "",
                    EVENT_TYPE_LABELS.get(ev["event_type"], ev["event_type"]),
                    ev["title"],
                    ev["description"] or "",
                    time_str,
                    ev["location"] or "",
                ],
            )

        _auto_width(ws)
        return _to_response(wb, f"{year}年{month}月行事曆.xlsx")
    finally:
        session.close()


# ============ Leaves ============


def _approval_label(is_approved):
    if is_approved is True:
        return "已核准"
    if is_approved is False:
        return "已駁回"
    return "待審核"


@router.get("/leaves")
def export_leaves(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_READ)),
    year: int = Query(...),
    month: int = Query(...),
):
    """匯出請假記錄 Excel"""
    session = get_session()
    try:
        _, last_day = cal_module.monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)

        leaves = (
            session.query(LeaveRecord)
            .filter(LeaveRecord.start_date <= end, LeaveRecord.end_date >= start)
            .order_by(LeaveRecord.start_date)
            .all()
        )
        emp_map = _id_name_map(session, Employee)

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年{month}月請假記錄"

        ws.merge_cells("A1:H1")
        ws["A1"] = f"{year}年{month}月 請假記錄"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "員工姓名",
            "假別",
            "開始日期",
            "結束日期",
            "時數",
            "扣薪比例",
            "原因",
            "審核狀態",
        ]
        _write_header_row(ws, 3, headers)

        for idx, lv in enumerate(leaves, 4):
            _write_data_row(
                ws,
                idx,
                [
                    emp_map.get(lv.employee_id, ""),
                    LEAVE_TYPE_LABELS.get(lv.leave_type, lv.leave_type),
                    lv.start_date.isoformat(),
                    lv.end_date.isoformat(),
                    lv.leave_hours or 8,
                    lv.deduction_ratio,
                    lv.reason or "",
                    _approval_label(lv.is_approved),
                ],
            )

        _auto_width(ws)
        return _to_response(wb, f"{year}年{month}月請假記錄.xlsx")
    finally:
        session.close()


# ============ Overtimes ============


@router.get("/overtimes")
def export_overtimes(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_READ)),
    year: int = Query(...),
    month: int = Query(...),
):
    """匯出加班記錄 Excel"""
    session = get_session()
    try:
        _, last_day = cal_module.monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)

        overtimes = (
            session.query(OvertimeRecord)
            .filter(
                OvertimeRecord.overtime_date >= start,
                OvertimeRecord.overtime_date <= end,
            )
            .order_by(OvertimeRecord.overtime_date)
            .all()
        )
        emp_map = _id_name_map(session, Employee)

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年{month}月加班記錄"

        ws.merge_cells("A1:I1")
        ws["A1"] = f"{year}年{month}月 加班記錄"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "員工姓名",
            "日期",
            "類型",
            "開始時間",
            "結束時間",
            "時數",
            "加班費",
            "原因",
            "審核狀態",
        ]
        _write_header_row(ws, 3, headers)

        for idx, ot in enumerate(overtimes, 4):
            start_t = ot.start_time.strftime("%H:%M") if ot.start_time else ""
            end_t = ot.end_time.strftime("%H:%M") if ot.end_time else ""
            _write_data_row(
                ws,
                idx,
                [
                    emp_map.get(ot.employee_id, ""),
                    ot.overtime_date.isoformat(),
                    OVERTIME_TYPE_LABELS.get(ot.overtime_type, ot.overtime_type),
                    start_t,
                    end_t,
                    ot.hours or 0,
                    round(ot.overtime_pay or 0),
                    ot.reason or "",
                    _approval_label(ot.is_approved),
                ],
            )

        _auto_width(ws)
        return _to_response(wb, f"{year}年{month}月加班記錄.xlsx")
    finally:
        session.close()


# ============ Holidays ============


@router.get("/holidays")
def export_holidays(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.CALENDAR)),
    year: int = Query(..., description="要匯出的年份"),
):
    """匯出指定年份國定假日 Excel"""
    session = get_session()
    try:
        holidays = (
            session.query(Holiday)
            .filter(
                Holiday.date >= date(year, 1, 1),
                Holiday.date <= date(year, 12, 31),
                Holiday.is_active.is_(True),
            )
            .order_by(Holiday.date)
            .all()
        )

        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年國定假日"

        ws.merge_cells("A1:C1")
        ws["A1"] = f"{year} 年國定假日清單"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = ["日期", "假日名稱", "說明"]
        _write_header_row(ws, 3, headers)

        for idx, h in enumerate(holidays, 4):
            _write_data_row(
                ws,
                idx,
                [
                    h.date.isoformat(),
                    h.name,
                    h.description or "",
                ],
            )

        _auto_width(ws)
        return _to_response(wb, f"{year}年國定假日.xlsx")
    finally:
        session.close()


# ============ Shifts ============


@router.get("/shifts")
def export_shifts(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
    week_start: str = Query(..., description="週起始日 YYYY-MM-DD（週一）"),
):
    """匯出指定週排班 Excel"""
    session = get_session()
    try:
        try:
            week_date = date.fromisoformat(week_start)
            week_date = week_date - timedelta(days=week_date.weekday())
        except ValueError:
            from fastapi import HTTPException as FHTTPException

            raise FHTTPException(
                status_code=400, detail="week_start 格式錯誤，請使用 YYYY-MM-DD"
            )

        assignments = (
            session.query(ShiftAssignment, Employee, ShiftType)
            .join(Employee, ShiftAssignment.employee_id == Employee.id)
            .outerjoin(ShiftType, ShiftAssignment.shift_type_id == ShiftType.id)
            .filter(ShiftAssignment.week_start_date == week_date)
            .order_by(Employee.employee_id)
            .all()
        )

        week_end = week_date + timedelta(days=6)
        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"排班表"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"排班表：{week_date.isoformat()} ～ {week_end.isoformat()}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = ["工號", "姓名", "班別名稱", "上班時間", "下班時間", "備註"]
        _write_header_row(ws, 3, headers)

        for idx, (a, emp, st) in enumerate(assignments, 4):
            _write_data_row(
                ws,
                idx,
                [
                    emp.employee_id,
                    emp.name,
                    st.name if st else "",
                    st.work_start if st else "",
                    st.work_end if st else "",
                    a.notes or "",
                ],
            )

        _auto_width(ws)
        return _to_response(wb, f"排班表_{week_date.isoformat()}.xlsx")
    finally:
        session.close()


# ============ Employee Attendance (Personal Monthly Report) ============

WEEKDAY_LABELS = ["一", "二", "三", "四", "五", "六", "日"]

_ANOMALY_FILL = PatternFill(start_color="FFF0CC", end_color="FFF0CC", fill_type="solid")
_ABSENT_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
_WEEKEND_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_SUMMARY_FONT = Font(bold=True, size=11)


def _calc_work_hours(att) -> float | None:
    """依打卡時間計算工時（扣午休 1h），無完整打卡回傳 None"""
    if not (att and att.punch_in_time and att.punch_out_time):
        return None
    total = (att.punch_out_time - att.punch_in_time).seconds / 3600
    pi = (
        att.punch_in_time.time()
        if hasattr(att.punch_in_time, "time")
        else att.punch_in_time
    )
    po = (
        att.punch_out_time.time()
        if hasattr(att.punch_out_time, "time")
        else att.punch_out_time
    )
    lunch_s, lunch_e = time(12, 0), time(13, 0)
    if pi < lunch_e and po > lunch_s:
        overlap = (
            min(po.hour * 60 + po.minute, lunch_e.hour * 60 + lunch_e.minute)
            - max(pi.hour * 60 + pi.minute, lunch_s.hour * 60 + lunch_s.minute)
        ) / 60
        total -= max(0, overlap)
    return max(0, round(total, 1))


@router.get("/employee-attendance")
def export_employee_attendance(
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.ATTENDANCE_READ)),
    employee_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
):
    """匯出指定員工的個人出勤月報 Excel（逐日打卡明細）"""
    session = get_session()
    try:
        # 1. 員工基本資料
        emp = session.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail=EMPLOYEE_DOES_NOT_EXIST)

        job_title_name = ""
        if emp.job_title_id:
            jt = session.query(JobTitle).filter(JobTitle.id == emp.job_title_id).first()
            if jt:
                job_title_name = jt.name

        classroom_name = ""
        if emp.classroom_id:
            cr = (
                session.query(Classroom)
                .filter(Classroom.id == emp.classroom_id)
                .first()
            )
            if cr:
                classroom_name = cr.name

        _, last_day = cal_module.monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)

        # 2. 打卡記錄
        att_records = (
            session.query(Attendance)
            .filter(
                Attendance.employee_id == employee_id,
                Attendance.attendance_date >= start,
                Attendance.attendance_date <= end,
            )
            .all()
        )
        att_map = {att.attendance_date: att for att in att_records}

        # 3. 請假記錄（無論審核狀態）
        leave_records = (
            session.query(LeaveRecord)
            .filter(
                LeaveRecord.employee_id == employee_id,
                LeaveRecord.start_date <= end,
                LeaveRecord.end_date >= start,
            )
            .all()
        )
        leave_map: dict = {}
        for lv in leave_records:
            d = max(lv.start_date, start)
            lv_end = min(lv.end_date, end)
            while d <= lv_end:
                leave_map.setdefault(d, []).append(lv)
                d += timedelta(days=1)

        # 4. 加班記錄
        ot_records = (
            session.query(OvertimeRecord)
            .filter(
                OvertimeRecord.employee_id == employee_id,
                OvertimeRecord.overtime_date >= start,
                OvertimeRecord.overtime_date <= end,
            )
            .all()
        )
        ot_map: dict = {}
        for ot in ot_records:
            ot_map.setdefault(ot.overtime_date, []).append(ot)

        # 5. 國定假日
        holiday_set = {
            h.date
            for h in session.query(Holiday)
            .filter(
                Holiday.date >= start,
                Holiday.date <= end,
                Holiday.is_active.is_(True),
            )
            .all()
        }

        # ---- 建立 Excel ----
        wb = Workbook()
        ws = _safe_ws(wb)
        ws.title = f"{year}年{month}月出勤月報"

        NUM_COLS = 14
        last_col_letter = chr(ord("A") + NUM_COLS - 1)  # 'N'

        # 列1: 標題
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = f"{emp.name}  {year}年{month}月 出勤月報"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = CENTER_ALIGN

        # 列2: 員工資訊
        ws.merge_cells(f"A2:{last_col_letter}2")
        info_parts = [f"工號: {emp.employee_id}"]
        if job_title_name:
            info_parts.append(f"職稱: {job_title_name}")
        if classroom_name:
            info_parts.append(f"所屬班級: {classroom_name}")
        ws["A2"] = "  ".join(info_parts)
        ws["A2"].font = Font(size=11)
        ws["A2"].alignment = CENTER_ALIGN

        # 列3: 欄位標頭（藍底白字）
        headers = [
            "日期",
            "星期",
            "假日",
            "上班打卡",
            "下班打卡",
            "工時(h)",
            "狀態",
            "遲到(分)",
            "早退(分)",
            "請假種類",
            "假時",
            "加班類型",
            "加班時數",
            "備註",
        ]
        _write_header_row(ws, 3, headers)

        # ---- 逐日填資料 ----
        # 統計摘要用
        expected_workdays = 0
        actual_workdays = 0
        late_count = 0
        late_total_min = 0
        early_count = 0
        missing_punch_count = 0
        leave_day_count = 0
        absent_count = 0

        row_idx = 4
        cur = start
        while cur <= end:
            is_weekend = cur.weekday() >= 5
            is_holiday = cur in holiday_set
            is_non_work = is_weekend or is_holiday

            att = att_map.get(cur)
            day_leaves = leave_map.get(cur, [])
            day_ots = ot_map.get(cur, [])

            # 打卡時間
            punch_in_str = (
                att.punch_in_time.strftime("%H:%M") if att and att.punch_in_time else ""
            )
            punch_out_str = (
                att.punch_out_time.strftime("%H:%M")
                if att and att.punch_out_time
                else ""
            )

            # 工時
            wh = _calc_work_hours(att)
            work_hours_str = str(wh) if wh is not None else "-"

            # 狀態
            status_parts = []
            if att:
                if att.is_late:
                    status_parts.append("遲到")
                if att.is_early_leave:
                    status_parts.append("早退")
                if att.is_missing_punch_in:
                    status_parts.append("缺卡(上)")
                if att.is_missing_punch_out:
                    status_parts.append("缺卡(下)")
            status_str = (
                "、".join(status_parts) if status_parts else ("正常" if att else "")
            )

            late_min = att.late_minutes or 0 if att and att.is_late else 0
            early_min = (
                att.early_leave_minutes or 0 if att and att.is_early_leave else 0
            )

            # 請假
            leave_str = "、".join(
                LEAVE_TYPE_LABELS.get(lv.leave_type, lv.leave_type) for lv in day_leaves
            )
            leave_hours_val = sum(lv.leave_hours or 0 for lv in day_leaves)
            leave_hours_str = f"{leave_hours_val}h" if day_leaves else ""

            # 加班
            ot_str = "、".join(
                OVERTIME_TYPE_LABELS.get(ot.overtime_type, ot.overtime_type)
                for ot in day_ots
            )
            ot_hours_val = sum(ot.hours or 0 for ot in day_ots)
            ot_hours_str = str(ot_hours_val) if day_ots else ""

            # 備註
            remark = att.remark if att and hasattr(att, "remark") and att.remark else ""

            row_values = [
                cur.isoformat(),
                WEEKDAY_LABELS[cur.weekday()],
                "假日" if is_holiday else ("週末" if is_weekend else ""),
                punch_in_str,
                punch_out_str,
                work_hours_str,
                status_str,
                late_min if late_min else "",
                early_min if early_min else "",
                leave_str,
                leave_hours_str,
                ot_str,
                ot_hours_str,
                remark,
            ]

            for col, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = THIN_BORDER

            # 列背景色
            if is_non_work:
                fill = _WEEKEND_FILL
            elif att and (
                att.is_late
                or att.is_early_leave
                or att.is_missing_punch_in
                or att.is_missing_punch_out
            ):
                fill = _ANOMALY_FILL
            elif not att and not day_leaves and not is_non_work:
                fill = _ABSENT_FILL
            else:
                fill = None

            if fill:
                for col in range(1, NUM_COLS + 1):
                    ws._ws.cell(row=row_idx, column=col).fill = fill

            # 統計累計
            if not is_non_work:
                expected_workdays += 1
                if att:
                    actual_workdays += 1
                    if att.is_late:
                        late_count += 1
                        late_total_min += att.late_minutes or 0
                    if att.is_early_leave:
                        early_count += 1
                    if att.is_missing_punch_in or att.is_missing_punch_out:
                        missing_punch_count += 1
                if day_leaves and not att:
                    leave_day_count += 1
                elif day_leaves and att:
                    # 有請假也有出勤，計為請假
                    leave_day_count += 1
                    actual_workdays -= 1  # 避免重複計入
                if not att and not day_leaves:
                    absent_count += 1

            row_idx += 1
            cur += timedelta(days=1)

        # ---- 摘要列（綠底）----
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        summary_label_cell = ws._ws.cell(row=row_idx, column=1, value="摘要")
        summary_label_cell.font = _SUMMARY_FONT
        summary_label_cell.border = THIN_BORDER

        summary_values = [
            expected_workdays,
            actual_workdays,
            late_count,
            late_total_min,
            early_count,
            missing_punch_count,
            leave_day_count,
            absent_count,
        ]
        summary_labels = [
            "應出勤天數",
            "實際出勤",
            "遲到次數",
            "遲到總分鐘",
            "早退次數",
            "缺卡次數",
            "請假天數",
            "曠職天數",
        ]
        for col_offset, (label, val) in enumerate(zip(summary_labels, summary_values)):
            # 每個摘要佔一格（從 col 3 開始）
            col = 3 + col_offset
            if col > NUM_COLS:
                break
            cell_label = ws._ws.cell(row=row_idx, column=col, value=f"{label}: {val}")
            cell_label.font = _SUMMARY_FONT
            cell_label.border = THIN_BORDER
            cell_label.fill = _SUMMARY_FILL

        # 摘要前兩格也加綠底
        for col in range(1, 3):
            ws._ws.cell(row=row_idx, column=col).fill = _SUMMARY_FILL

        _auto_width(ws)

        logger.info(
            "個人出勤月報匯出：emp=%s(%s) year=%d month=%d operator=%s",
            emp.employee_id,
            emp.name,
            year,
            month,
            current_user.get("username", "?"),
        )
        return _to_response(wb, f"{year}年{month}月_{emp.name}_出勤月報.xlsx")
    finally:
        session.close()
