"""
api/activity/registrations_static.py — 才藝報名靜態路由

含三個端點：
- PUT /registrations/batch-payment （批次標記已繳費）
- GET /registrations/export       （Excel 報名名單匯出）
- GET /registrations/payment-report（Excel 繳費帳務報表）

⚠️ 必須在 __init__.py 內 include 於 /{registration_id}/... 動態路由之前，
   否則 batch-payment / export / payment-report 會被 /{id} 吞掉。
"""

import io
import logging
from collections import defaultdict
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from models.database import (
    get_session,
    ActivityRegistration,
    ActivityPaymentRecord,
)
from services.activity_service import activity_service
from utils.errors import raise_safe_500
from utils.excel_utils import SafeWorksheet
from utils.auth import require_staff_permission
from utils.permissions import Permission
from utils.rate_limit import SlidingWindowLimiter
from utils.finance_guards import require_finance_approve

from ._shared import (
    BatchPaymentUpdate,
    SYSTEM_RECONCILE_METHOD,
    _compute_is_paid,
    _derive_payment_status,
    _invalidate_activity_dashboard_caches,
    _invalidate_finance_summary_cache,
    _batch_calc_total_amounts,
    _build_registration_filter_query,
    _fetch_reg_course_names,
    _require_daily_close_unlocked,
    TAIPEI_TZ,
    today_taipei,
)

logger = logging.getLogger(__name__)
router = APIRouter()


_export_limiter = SlidingWindowLimiter(
    max_calls=5,
    window_seconds=60,
    name="activity_export",
    error_detail="匯出過於頻繁，請稍後再試",
).as_dependency()

# 批次繳費寫入每次影響多筆 registration + 繳費紀錄，放寬但仍需限流
_batch_payment_limiter = SlidingWindowLimiter(
    max_calls=10,
    window_seconds=60,
    name="activity_batch_payment",
    error_detail="批次繳費操作過於頻繁，請稍後再試",
).as_dependency()

# 匯出單次查詢上限，避免大報表造成記憶體爆炸或 timeout
MAX_EXPORT_ROWS = 5000


@router.put("/registrations/batch-payment")
async def batch_update_payment(
    body: BatchPaymentUpdate,
    _rl=Depends(_batch_payment_limiter),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """批次標記為已繳費（僅此方向；未繳費全額沖帳已禁用）

    Why: 批次把一串報名「標記未繳費」會一次寫多筆全額 refund，誤操作後果嚴重且
    不可部分回滾。schema 層已禁止 is_paid=False（進到這裡一律是 True）；
    若需把某筆退費，請改用 PUT /registrations/{id}/payment（帶
    confirm_refund_amount）或 DELETE /payments/{id} 軟刪對應繳費。

    併發保護：`.with_for_update()` 鎖住目標 registration，避免兩個客戶端同時
    呼叫造成重複寫 payment_record 或 lost update。
    """
    session = get_session()
    try:
        regs = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id.in_(body.ids),
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .all()
        )
        if not regs:
            raise HTTPException(status_code=404, detail="找不到指定報名資料")

        operator = current_user.get("username", "")
        today = today_taipei()

        # 批次補齊皆寫 today；若今日已簽核則拒絕，避免日結 snapshot 失準
        _require_daily_close_unlocked(session, today)

        # P3 N+1 修正：一次 GROUP BY 查詢所有應繳金額
        unpaid_reg_ids = [reg.id for reg in regs if not reg.is_paid]
        total_amount_map = (
            _batch_calc_total_amounts(session, unpaid_reg_ids) if unpaid_reg_ids else {}
        )

        # ── 整批 shortfall 累積簽核 ────────────────────────────────────
        # Why: 批次補齊把欠費直接寫成 system payment，會計可一鍵把報表變漂亮；
        # 整批 shortfall 合計超過閾值即整批需具備 ACTIVITY_PAYMENT_APPROVE 才能執行
        total_shortfall = 0
        for reg in regs:
            if not reg.is_paid:
                total_amount = total_amount_map.get(reg.id, 0)
                shortfall = total_amount - (reg.paid_amount or 0)
                if shortfall > 0:
                    total_shortfall += shortfall
        if total_shortfall > 0:
            require_finance_approve(
                total_shortfall,
                current_user,
                action_label=f"批次補齊整批合計 NT${total_shortfall:,}",
            )

        notes_text = f"（批次標記已繳費自動補齊：{body.reason}）"
        for reg in regs:
            if not reg.is_paid:
                total_amount = total_amount_map.get(reg.id, 0)
                shortfall = total_amount - (reg.paid_amount or 0)
                if shortfall > 0:
                    rec = ActivityPaymentRecord(
                        registration_id=reg.id,
                        type="payment",
                        amount=shortfall,
                        payment_date=today,
                        payment_method=SYSTEM_RECONCILE_METHOD,
                        notes=notes_text,
                        operator=operator,
                    )
                    session.add(rec)
                    reg.paid_amount = total_amount
                reg.is_paid = _compute_is_paid(reg.paid_amount or 0, total_amount)
            activity_service.log_change(
                session,
                reg.id,
                reg.student_name,
                "批次更新付款狀態",
                f"付款狀態批次更新為：已繳費（原因：{body.reason}）",
                operator,
            )

        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        _invalidate_finance_summary_cache()
        logger.warning(
            "批次付款狀態更新：筆數=%d is_paid=True operator=%s",
            len(regs),
            operator,
        )
        return {
            "message": f"已更新 {len(regs)} 筆報名為已繳費",
            "updated": len(regs),
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.get("/registrations/export")
def export_registrations(
    search: Optional[str] = None,
    payment_status: Optional[str] = None,
    course_id: Optional[int] = None,
    classroom_name: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """匯出報名名單為 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    session = get_session()
    try:
        q = _build_registration_filter_query(
            session,
            search=search,
            payment_status=payment_status,
            course_id=course_id,
            classroom_name=classroom_name,
        )
        total_count = q.count()
        if total_count > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"匯出筆數 {total_count} 超過上限 {MAX_EXPORT_ROWS}，"
                    "請加上條件（如課程/班級/繳費狀態）縮小範圍後再匯出"
                ),
            )
        regs = q.order_by(ActivityRegistration.created_at.desc()).all()
        reg_ids = [r.id for r in regs]
        course_name_map = _fetch_reg_course_names(session, reg_ids)

        wb = openpyxl.Workbook()
        ws = SafeWorksheet(wb.active)
        ws.title = "報名名單"

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
        )
        center = Alignment(horizontal="center", vertical="center")

        headers = ["序號", "學生姓名", "班級", "課程", "付款狀態", "備註", "報名時間"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for idx, reg in enumerate(regs, start=1):
            ws.append(
                [
                    idx,
                    reg.student_name,
                    reg.class_name or "",
                    "、".join(course_name_map.get(reg.id, [])),
                    "已繳費" if reg.is_paid else "未繳費",
                    reg.remark or "",
                    reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else "",
                ]
            )

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"activity_registrations_{datetime.now(TAIPEI_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()


@router.get("/registrations/payment-report")
def export_payment_report(
    search: Optional[str] = None,
    payment_status: Optional[str] = None,
    course_id: Optional[int] = None,
    classroom_name: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
    _: None = Depends(_export_limiter),
):
    """匯出繳費帳務報表（兩個工作表：繳費總覽 + 繳費明細）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    _HEADER_FONT = Font(bold=True)
    _HEADER_FILL = PatternFill(
        start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
    )
    _CENTER = Alignment(horizontal="center", vertical="center")

    session = get_session()
    try:
        q = _build_registration_filter_query(
            session,
            search=search,
            payment_status=payment_status,
            course_id=course_id,
            classroom_name=classroom_name,
        )
        total_count = q.count()
        if total_count > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"匯出筆數 {total_count} 超過上限 {MAX_EXPORT_ROWS}，"
                    "請加上條件（如課程/班級/繳費狀態）縮小範圍後再匯出"
                ),
            )
        regs = q.order_by(ActivityRegistration.created_at.desc()).all()
        reg_ids = [r.id for r in regs]

        # 批次計算應繳金額
        total_amount_map = (
            _batch_calc_total_amounts(session, reg_ids) if reg_ids else {}
        )

        # 批次查詢課程名稱
        course_name_map = _fetch_reg_course_names(session, reg_ids)

        # 批次查詢繳費明細
        payment_records = []
        payment_map: dict[int, list] = defaultdict(list)
        last_payment_date_map: dict[int, str] = {}
        if reg_ids:
            payment_records = (
                session.query(ActivityPaymentRecord)
                .filter(ActivityPaymentRecord.registration_id.in_(reg_ids))
                .order_by(
                    ActivityPaymentRecord.registration_id,
                    ActivityPaymentRecord.payment_date.asc(),
                )
                .all()
            )
            for pr in payment_records:
                payment_map[pr.registration_id].append(pr)
                if pr.payment_date:
                    date_str = pr.payment_date.isoformat()
                    existing = last_payment_date_map.get(pr.registration_id, "")
                    if date_str > existing:
                        last_payment_date_map[pr.registration_id] = date_str

        wb = openpyxl.Workbook()

        # ── 工作表一：繳費總覽 ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "繳費總覽"
        headers1 = [
            "序號",
            "學生",
            "班級",
            "報名課程",
            "應繳總額",
            "已繳金額",
            "差額",
            "狀態",
            "最後繳費日",
        ]
        for col, h in enumerate(headers1, start=1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        status_label_map = {
            "paid": "已繳清",
            "partial": "部分繳費",
            "unpaid": "未繳費",
            "overpaid": "超額繳費",
        }
        for idx, reg in enumerate(regs, start=1):
            total = total_amount_map.get(reg.id, 0)
            paid = reg.paid_amount or 0
            diff = paid - total
            status = _derive_payment_status(paid, total)
            ws1.append(
                [
                    idx,
                    reg.student_name,
                    reg.class_name or "",
                    "、".join(course_name_map.get(reg.id, [])),
                    total,
                    paid,
                    diff,
                    status_label_map.get(status, status),
                    last_payment_date_map.get(reg.id, ""),
                ]
            )

        for col in ws1.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── 工作表二：繳費明細 ──────────────────────────────────────────
        # 軟刪（voided）紀錄獨立標示：類型欄加「（已作廢）」、新增作廢欄位顯示作廢人/
        # 時間/原因，避免與有效流水混為一談；總覽工作表的金額計算已排除 voided。
        ws2 = SafeWorksheet(wb.create_sheet(title="繳費明細"))
        headers2 = [
            "學生",
            "班級",
            "類型",
            "金額",
            "方式",
            "日期",
            "操作人員",
            "備註",
            "作廢狀態",
            "作廢人",
            "作廢時間",
            "作廢原因",
        ]
        for col, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        reg_meta = {r.id: r for r in regs}
        for pr in payment_records:
            reg = reg_meta.get(pr.registration_id)
            is_voided = pr.voided_at is not None
            type_label = "繳費" if pr.type == "payment" else "退費"
            if is_voided:
                type_label = f"{type_label}（已作廢）"
            ws2.append(
                [
                    reg.student_name if reg else "",
                    reg.class_name if reg else "",
                    type_label,
                    pr.amount,
                    pr.payment_method or "",
                    pr.payment_date.isoformat() if pr.payment_date else "",
                    pr.operator or "",
                    pr.notes or "",
                    "已作廢" if is_voided else "",
                    pr.voided_by or "",
                    pr.voided_at.isoformat() if pr.voided_at else "",
                    pr.void_reason or "",
                ]
            )

        for col in ws2.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = (
            f"payment_report_{datetime.now(TAIPEI_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        logger.warning(
            "繳費帳務報表匯出：operator=%s 筆數=%d",
            current_user.get("username"),
            len(regs),
        )
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()
