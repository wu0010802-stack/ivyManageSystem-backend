"""
排班管理 API
- 班別模板 CRUD
- 每週排班指派
"""

import logging
from datetime import date, timedelta
from io import BytesIO
from typing import Optional, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pydantic import BaseModel

from cachetools import TTLCache
from sqlalchemy.orm import joinedload

from models.database import (
    get_session,
    ShiftType,
    ShiftAssignment,
    Employee,
    DailyShift,
    ShiftSwapRequest,
)
from utils.auth import require_staff_permission
from utils.errors import raise_safe_500
from utils.permissions import Permission
from utils.file_upload import read_upload_with_size_check, validate_file_signature
from utils.excel_utils import SafeWorksheet, xlsx_streaming_response
from utils.schedule_utils import (
    get_week_dates,
    get_employee_weekly_shift_hours,
    compute_weekly_hours,
    build_weekly_warning,
)

logger = logging.getLogger(__name__)

# ShiftType 很少變動，使用 TTLCache 減少重複 DB 查詢（5 分鐘 TTL）
_shift_type_cache: TTLCache = TTLCache(maxsize=3, ttl=300)


def _clear_shift_type_cache():
    _shift_type_cache.clear()


def _get_all_shift_types_cached(session) -> list:
    """回傳 list[dict]，快取 5 分鐘"""
    cached = _shift_type_cache.get("all")
    if cached is not None:
        return cached
    types = session.query(ShiftType).order_by(ShiftType.sort_order).all()
    result = [
        {
            "id": t.id,
            "name": t.name,
            "work_start": t.work_start,
            "work_end": t.work_end,
            "sort_order": t.sort_order,
            "is_active": t.is_active,
        }
        for t in types
    ]
    _shift_type_cache["all"] = result
    return result


def _get_shift_type_id_map_cached(session) -> dict:
    """回傳 {id: SimpleNamespace(work_start, work_end, name, is_active)}，快取 5 分鐘。
    用於需要 .work_start / .work_end 屬性存取的場景（如工時計算）。
    """
    from types import SimpleNamespace

    cached = _shift_type_cache.get("id_map")
    if cached is not None:
        return cached
    types = session.query(ShiftType).all()
    result = {
        t.id: SimpleNamespace(
            id=t.id,
            name=t.name,
            work_start=t.work_start,
            work_end=t.work_end,
            sort_order=t.sort_order,
            is_active=t.is_active,
        )
        for t in types
    }
    _shift_type_cache["id_map"] = result
    return result


router = APIRouter(prefix="/api/shifts", tags=["shifts"])


# ---------------------------------------------------------------------------
# Pydantic Schemas
# ---------------------------------------------------------------------------


class ShiftTypeCreate(BaseModel):
    name: str
    work_start: str
    work_end: str
    sort_order: int = 0


class ShiftTypeUpdate(BaseModel):
    name: Optional[str] = None
    work_start: Optional[str] = None
    work_end: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None


class AssignmentItem(BaseModel):
    employee_id: int
    shift_type_id: Optional[int] = None
    notes: Optional[str] = None


class BulkAssignmentRequest(BaseModel):
    week_start_date: str  # YYYY-MM-DD (must be a Monday)
    assignments: List[AssignmentItem]


class DailyShiftCreate(BaseModel):
    """每日排班（調班）請求"""

    employee_id: int
    shift_type_id: int
    date: str  # YYYY-MM-DD
    notes: Optional[str] = None


# ---------------------------------------------------------------------------
# 班別模板 CRUD
# ---------------------------------------------------------------------------


@router.get("/types")
def list_shift_types(
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    session = get_session()
    try:
        return _get_all_shift_types_cached(session)
    finally:
        session.close()


@router.post("/types", status_code=201)
def create_shift_type(
    data: ShiftTypeCreate,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    session = get_session()
    try:
        st = ShiftType(
            name=data.name,
            work_start=data.work_start,
            work_end=data.work_end,
            sort_order=data.sort_order,
        )
        session.add(st)
        session.commit()
        session.refresh(st)
        _clear_shift_type_cache()
        logger.info(f"Created shift type: {st.name}")
        return {
            "id": st.id,
            "name": st.name,
            "work_start": st.work_start,
            "work_end": st.work_end,
            "sort_order": st.sort_order,
            "is_active": st.is_active,
        }
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/types/{type_id}")
def update_shift_type(
    type_id: int,
    data: ShiftTypeUpdate,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    session = get_session()
    try:
        st = session.query(ShiftType).get(type_id)
        if not st:
            raise HTTPException(status_code=404, detail="班別不存在")
        for field, value in data.model_dump(exclude_unset=True).items():
            setattr(st, field, value)
        session.commit()
        _clear_shift_type_cache()
        logger.info(f"Updated shift type: {st.name}")
        return {
            "id": st.id,
            "name": st.name,
            "work_start": st.work_start,
            "work_end": st.work_end,
            "sort_order": st.sort_order,
            "is_active": st.is_active,
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


def _shift_type_in_use_message(
    assignment_count: int, daily_count: int, swap_count: int
) -> "str | None":
    """匯總三張子表的引用計數，回傳人可讀錯誤訊息；若可安全刪除則回傳 None。"""
    parts = []
    if assignment_count > 0:
        parts.append(f"每週排班 {assignment_count} 筆")
    if daily_count > 0:
        parts.append(f"每日調班 {daily_count} 筆")
    if swap_count > 0:
        parts.append(f"換班申請 {swap_count} 筆")
    if parts:
        return f"此班別已被使用（{'、'.join(parts)}），無法刪除"
    return None


@router.delete("/types/{type_id}")
def delete_shift_type(
    type_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    session = get_session()
    try:
        st = session.query(ShiftType).get(type_id)
        if not st:
            raise HTTPException(status_code=404, detail="班別不存在")

        assignment_count = (
            session.query(ShiftAssignment)
            .filter(ShiftAssignment.shift_type_id == type_id)
            .count()
        )
        daily_count = (
            session.query(DailyShift)
            .filter(DailyShift.shift_type_id == type_id)
            .count()
        )
        swap_count = (
            session.query(ShiftSwapRequest)
            .filter(
                (ShiftSwapRequest.requester_shift_type_id == type_id)
                | (ShiftSwapRequest.target_shift_type_id == type_id)
            )
            .count()
        )

        error_msg = _shift_type_in_use_message(
            assignment_count, daily_count, swap_count
        )
        if error_msg:
            raise HTTPException(status_code=400, detail=error_msg)

        session.delete(st)
        session.commit()
        _clear_shift_type_cache()
        logger.info(f"Deleted shift type: {st.name}")
        return {"message": "已刪除"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 每週排班
# ---------------------------------------------------------------------------


@router.get("/assignments")
def get_assignments(
    week_start: str,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """查詢某週排班。week_start 為該週週一日期 (YYYY-MM-DD)"""
    session = get_session()
    try:
        week_date = date.fromisoformat(week_start)
        # Align to Monday
        week_date = week_date - timedelta(days=week_date.weekday())

        assignments = (
            session.query(ShiftAssignment)
            .options(
                joinedload(ShiftAssignment.employee),
                joinedload(ShiftAssignment.shift_type),
            )
            .filter(ShiftAssignment.week_start_date == week_date)
            .all()
        )
        result = []
        for a in assignments:
            emp = a.employee
            st = a.shift_type
            result.append(
                {
                    "id": a.id,
                    "employee_id": a.employee_id,
                    "employee_name": emp.name if emp else "",
                    "shift_type_id": a.shift_type_id,
                    "shift_type_name": st.name if st else "",
                    "work_start": st.work_start if st else "",
                    "work_end": st.work_end if st else "",
                    "week_start_date": str(a.week_start_date),
                    "notes": a.notes or "",
                }
            )
        return result
    finally:
        session.close()


def _apply_employee_assignment_action(session, existing, item, week_date: str) -> str:
    """針對單一員工執行排班 upsert 或刪除。

    僅影響該員工自身的記錄，不動其他員工的資料。

    Returns: 'inserted' | 'updated' | 'deleted' | 'skipped'
    """
    if item.shift_type_id is None:
        if existing:
            session.delete(existing)
            return "deleted"
        return "skipped"

    if existing:
        existing.shift_type_id = item.shift_type_id
        existing.notes = item.notes
        return "updated"

    session.add(
        ShiftAssignment(
            employee_id=item.employee_id,
            shift_type_id=item.shift_type_id,
            week_start_date=week_date,
            notes=item.notes,
        )
    )
    return "inserted"


@router.post("/assignments", status_code=201)
def save_assignments(
    data: BulkAssignmentRequest,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """批次儲存某週排班（per-employee upsert，不影響清單外的員工）"""
    session = get_session()
    try:
        week_date = date.fromisoformat(data.week_start_date)
        # Align to Monday
        week_date = week_date - timedelta(days=week_date.weekday())

        saved = deleted = 0
        for item in data.assignments:
            existing = (
                session.query(ShiftAssignment)
                .filter(
                    ShiftAssignment.employee_id == item.employee_id,
                    ShiftAssignment.week_start_date == week_date,
                )
                .first()
            )
            action = _apply_employee_assignment_action(
                session, existing, item, str(week_date)
            )
            if action in ("inserted", "updated"):
                saved += 1
            elif action == "deleted":
                deleted += 1

        session.commit()
        logger.info(
            f"Saved {saved} / deleted {deleted} shift assignments for week {week_date}"
        )

        # ── 週工時超時預警（commit 後直接讀 DB 最新狀態，不需 overrides）──
        assigned_ids = {
            item.employee_id
            for item in data.assignments
            if item.shift_type_id is not None
        }
        warnings = []
        if assigned_ids:
            shift_type_map = _get_shift_type_id_map_cached(session)
            emp_map = {
                e.id: e.name
                for e in session.query(Employee)
                .filter(Employee.id.in_(assigned_ids))
                .all()
            }
            week_dates = get_week_dates(week_date)
            for emp_id in assigned_ids:
                shift_hours = get_employee_weekly_shift_hours(
                    session, emp_id, week_dates, shift_type_map
                )
                weekly_hours = compute_weekly_hours(shift_hours)
                w = build_weekly_warning(
                    emp_id,
                    emp_map.get(emp_id, str(emp_id)),
                    week_dates[0],
                    weekly_hours,
                )
                if w:
                    warnings.append(w)

        resp = {
            "message": f"已儲存 {saved} 筆、清除 {deleted} 筆排班",
            "week_start_date": str(week_date),
        }
        if warnings:
            resp["warnings"] = warnings
        return resp
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 每日排班（調班/換班）
# ---------------------------------------------------------------------------


@router.get("/daily")
def get_daily_shifts(
    start_date: str,
    end_date: str,
    employee_id: Optional[int] = None,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """查詢日期範圍內的排班調動/每日排班"""
    session = get_session()
    try:
        s_date = date.fromisoformat(start_date)
        e_date = date.fromisoformat(end_date)

        query = (
            session.query(DailyShift)
            .options(
                joinedload(DailyShift.employee),
                joinedload(DailyShift.shift_type),
            )
            .filter(DailyShift.date >= s_date, DailyShift.date <= e_date)
        )

        if employee_id:
            query = query.filter(DailyShift.employee_id == employee_id)

        daily_shifts = query.order_by(DailyShift.date).all()

        result = []
        for ds in daily_shifts:
            emp = ds.employee
            st = ds.shift_type
            result.append(
                {
                    "id": ds.id,
                    "employee_id": ds.employee_id,
                    "employee_name": emp.name if emp else "",
                    "shift_type_id": ds.shift_type_id,
                    "shift_type_name": st.name if st else "",
                    "work_start": st.work_start if st else "",
                    "work_end": st.work_end if st else "",
                    "date": str(ds.date),
                    "notes": ds.notes or "",
                }
            )
        return result
    finally:
        session.close()


@router.post("/daily", status_code=201)
def upsert_daily_shift(
    data: DailyShiftCreate,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """新增或更新每日排班（支援 UPSERT）"""
    session = get_session()
    try:
        target_date = date.fromisoformat(data.date)

        # 檢查是否已存在
        existing = (
            session.query(DailyShift)
            .filter(
                DailyShift.employee_id == data.employee_id,
                DailyShift.date == target_date,
            )
            .first()
        )

        if existing:
            existing.shift_type_id = data.shift_type_id
            existing.notes = data.notes
            msg = "Updated daily shift"
        else:
            new_shift = DailyShift(
                employee_id=data.employee_id,
                shift_type_id=data.shift_type_id,
                date=target_date,
                notes=data.notes,
            )
            session.add(new_shift)
            msg = "Created daily shift"

        session.commit()
        logger.info(f"{msg}: {data.employee_id} on {target_date}")
        return {"message": "已儲存"}

    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/daily/{shift_id}")
def delete_daily_shift(
    shift_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """刪除每日排班（恢復為週排班或預設）"""
    session = get_session()
    try:
        ds = session.query(DailyShift).get(shift_id)
        if not ds:
            raise HTTPException(status_code=404, detail="找不到該排班記錄")

        session.delete(ds)
        session.commit()
        return {"message": "已刪除"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 換班歷史（管理端）
# ---------------------------------------------------------------------------


@router.get("/swap-history")
def get_swap_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """查看換班歷史（管理端）"""
    session = get_session()
    try:
        query = session.query(ShiftSwapRequest).order_by(
            ShiftSwapRequest.created_at.desc()
        )

        if start_date:
            query = query.filter(
                ShiftSwapRequest.swap_date >= date.fromisoformat(start_date)
            )
        if end_date:
            query = query.filter(
                ShiftSwapRequest.swap_date <= date.fromisoformat(end_date)
            )
        if status:
            query = query.filter(ShiftSwapRequest.status == status)

        swaps = query.limit(100).all()

        # Pre-fetch employees and shift types
        emp_ids = set()
        for s in swaps:
            emp_ids.add(s.requester_id)
            emp_ids.add(s.target_id)
        emps = (
            {
                e.id: e.name
                for e in session.query(Employee).filter(Employee.id.in_(emp_ids)).all()
            }
            if emp_ids
            else {}
        )
        sts = {
            sid: ns.name for sid, ns in _get_shift_type_id_map_cached(session).items()
        }

        return [
            {
                "id": s.id,
                "requester_name": emps.get(s.requester_id, ""),
                "target_name": emps.get(s.target_id, ""),
                "swap_date": s.swap_date.isoformat(),
                "requester_shift": sts.get(s.requester_shift_type_id, "未排班"),
                "target_shift": sts.get(s.target_shift_type_id, "未排班"),
                "reason": s.reason,
                "status": s.status,
                "target_remark": s.target_remark,
                "target_responded_at": (
                    s.target_responded_at.isoformat() if s.target_responded_at else None
                ),
                "created_at": s.created_at.isoformat() if s.created_at else None,
            }
            for s in swaps
        ]
    finally:
        session.close()


# ---------------------------------------------------------------------------
# 排班匯入/匯出輔助
# ---------------------------------------------------------------------------


_SH_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_SH_HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
_SH_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_SH_CENTER_ALIGN = Alignment(horizontal="center")


def _sh_write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _SH_HEADER_FONT
        cell.fill = _SH_HEADER_FILL
        cell.border = _SH_THIN_BORDER
        cell.alignment = _SH_CENTER_ALIGN


@router.get("/import-template")
def get_shift_import_template(
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """下載排班批次匯入 Excel 範本"""
    session = get_session()
    try:
        shift_types = [
            ns for ns in _get_shift_type_id_map_cached(session).values() if ns.is_active
        ]
        shift_types.sort(key=lambda ns: ns.sort_order)

        wb = Workbook()
        ws = SafeWorksheet(wb.active)
        ws.title = "排班匯入範本"

        headers = ["員工編號", "員工姓名", "班別名稱", "備註(可空)"]
        _sh_write_header(ws, 1, headers)

        ws.cell(row=2, column=1, value="E001")
        ws.cell(row=2, column=2, value="王小明")
        ws.cell(row=2, column=3, value=shift_types[0].name if shift_types else "早班")
        ws.cell(row=2, column=4, value="")

        ws2 = SafeWorksheet(wb.create_sheet("班別說明"))
        ws2.cell(row=1, column=1, value="班別名稱")
        ws2.cell(row=1, column=2, value="上班時間")
        ws2.cell(row=1, column=3, value="下班時間")
        for idx, st in enumerate(shift_types, 2):
            ws2.cell(row=idx, column=1, value=st.name)
            ws2.cell(row=idx, column=2, value=st.work_start)
            ws2.cell(row=idx, column=3, value=st.work_end)

        note_ws = SafeWorksheet(wb.create_sheet("說明"))
        note_ws.cell(row=1, column=1, value="注意事項")
        note_ws.cell(
            row=2,
            column=1,
            value="1. 員工編號或員工姓名二擇一填寫即可（建議填員工編號）",
        )
        note_ws.cell(row=3, column=1, value="2. 班別名稱須完全符合「班別說明」頁的名稱")
        note_ws.cell(
            row=4,
            column=1,
            value="3. 上傳時需指定 week_start 參數（週一日期，格式 YYYY-MM-DD）",
        )

        return xlsx_streaming_response(wb, "排班匯入範本.xlsx")
    finally:
        session.close()


@router.post("/import")
async def import_shifts(
    file: UploadFile = File(...),
    week_start: str = Query(..., description="週起始日 YYYY-MM-DD（週一）"),
    current_user: dict = Depends(require_staff_permission(Permission.SCHEDULE)),
):
    """批次匯入排班（覆蓋指定週的排班，per-employee upsert）"""
    content = await read_upload_with_size_check(file)
    validate_file_signature(content, ".xlsx")
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{e}")

    try:
        week_date = date.fromisoformat(week_start)
        week_date = week_date - timedelta(days=week_date.weekday())
    except ValueError:
        raise HTTPException(
            status_code=400, detail="week_start 格式錯誤，請使用 YYYY-MM-DD"
        )

    results: dict = {"total": 0, "saved": 0, "failed": 0, "errors": []}
    session = get_session()
    try:
        employees = session.query(Employee).filter(Employee.is_active == True).all()
        emp_by_id = {str(e.employee_id): e for e in employees}
        emp_by_name = {e.name: e for e in employees}

        shift_types = [
            ns for ns in _get_shift_type_id_map_cached(session).values() if ns.is_active
        ]
        st_by_name = {ns.name: ns for ns in shift_types}

        for idx, row in df.iterrows():
            results["total"] += 1
            row_num = int(idx) + 2
            try:
                emp_id_str = str(row.get("員工編號", "")).strip()
                emp_name_str = str(row.get("員工姓名", "")).strip()
                emp = None
                if emp_id_str and emp_id_str not in ("nan", ""):
                    emp = emp_by_id.get(emp_id_str)
                if emp is None and emp_name_str and emp_name_str not in ("nan", ""):
                    emp = emp_by_name.get(emp_name_str)
                if emp is None:
                    raise ValueError(
                        f"找不到員工（編號:{emp_id_str}，姓名:{emp_name_str}）"
                    )

                st_name_raw = str(row.get("班別名稱", "")).strip()
                if st_name_raw in ("nan", ""):
                    raise ValueError("班別名稱不得為空")
                st = st_by_name.get(st_name_raw)
                if st is None:
                    raise ValueError(
                        f"找不到班別：{st_name_raw}（請參考「班別說明」頁）"
                    )

                notes_raw = row.get("備註(可空)")
                notes = (
                    str(notes_raw).strip()
                    if notes_raw is not None and not pd.isna(notes_raw)
                    else None
                )

                existing = (
                    session.query(ShiftAssignment)
                    .filter(
                        ShiftAssignment.employee_id == emp.id,
                        ShiftAssignment.week_start_date == week_date,
                    )
                    .first()
                )

                if existing:
                    existing.shift_type_id = st.id
                    existing.notes = notes
                else:
                    session.add(
                        ShiftAssignment(
                            employee_id=emp.id,
                            shift_type_id=st.id,
                            week_start_date=str(week_date),
                            notes=notes,
                        )
                    )

                session.flush()
                results["saved"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"第 {row_num} 行: {str(e)}")

        session.commit()
        logger.info(
            "排班批次匯入：使用者 %s，週 %s，共 %d 筆，成功 %d 筆，失敗 %d 筆",
            current_user.get("username"),
            week_date,
            results["total"],
            results["saved"],
            results["failed"],
        )
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e, context="匯入失敗")
    finally:
        session.close()

    return results
