"""才藝老師薪資明細測試。"""

import os
import sys
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.art_teacher_payroll import router as art_router
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from api.salary import init_salary_services
from api.salary import router as salary_router
from models.database import (
    ArtTeacherPayrollEntry,
    Base,
    Employee,
    SalaryRecord,
    User,
)
from services.art_teacher_payroll import (
    compute_total_for_month,
    generate_art_teacher_roster_xlsx,
)
from services.salary.engine import SalaryEngine
from utils.auth import hash_password
from utils.permissions import Permission


@pytest.fixture
def art_client(tmp_path):
    db_path = tmp_path / "art.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()
    init_salary_services(SalaryEngine(load_from_db=False), MagicMock())

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)
    app.include_router(art_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _add_hourly_emp(session, employee_id="A001", name="才藝老師"):
    emp = Employee(
        employee_id=employee_id,
        name=name,
        employee_type="hourly",
        hourly_rate=550,
        base_salary=0,
        is_active=True,
    )
    session.add(emp)
    session.flush()
    return emp


def _login(client, session_factory, perm=None):
    if perm is None:
        perm = ["SALARY_READ", "SALARY_WRITE"]
    with session_factory() as session:
        session.add(
            User(
                username="art_admin",
                password_hash=hash_password("TempPass123"),
                role="admin",
                permission_names=perm,
                is_active=True,
                must_change_password=False,
            )
        )
        session.commit()
    res = client.post(
        "/api/auth/login",
        json={"username": "art_admin", "password": "TempPass123"},
    )
    assert res.status_code == 200


# ── Service ─────────────────────────────────────────────────────────────────


class TestService:
    def test_compute_total_sums_entries(self, art_client):
        _, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            session.add_all(
                [
                    ArtTeacherPayrollEntry(
                        employee_id=emp.id,
                        salary_year=2026,
                        salary_month=4,
                        subject="美語",
                        hours=25,
                        hourly_rate=620,
                        base_amount=15500,
                        total_amount=16120,  # 含活動加給 620
                        activity_bonus=620,
                    ),
                    ArtTeacherPayrollEntry(
                        employee_id=emp.id,
                        salary_year=2026,
                        salary_month=4,
                        subject="課後美語",
                        classroom_label="(二)",
                        hours=4,
                        hourly_rate=620,
                        base_amount=2480,
                        total_amount=2480,
                    ),
                ]
            )
            session.commit()
            total = compute_total_for_month(session, emp.id, 2026, 4)
            assert total == 18600

    def test_compute_total_zero_when_no_entries(self, art_client):
        _, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            session.commit()
            assert compute_total_for_month(session, emp.id, 2026, 4) == 0

    def test_roster_xlsx_has_summary(self, art_client):
        _, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session, name="鍾馨瑤")
            session.add(
                ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    subject="舞蹈",
                    classroom_label="(二)",
                    hours=4,
                    hourly_rate=1000,
                    base_amount=4000,
                    excess_amount=200,
                    total_amount=4200,
                )
            )
            session.commit()
            filename, xlsx = generate_art_teacher_roster_xlsx(session, 2026, 4)
            assert filename.endswith(".xlsx")
            wb = load_workbook(BytesIO(xlsx))
            ws = wb.active
            assert "115.04" in ws["A1"].value
            assert ws.cell(row=2, column=1).value == "科目"
            # 第 3 列為資料
            assert ws.cell(row=3, column=2).value == "鍾馨瑤"
            assert ws.cell(row=3, column=5).value == 4000
            assert ws.cell(row=3, column=6).value == 200
            assert ws.cell(row=3, column=8).value == 4200
            # 第 4 列為合計
            assert ws.cell(row=4, column=1).value == "合計"
            assert ws.cell(row=4, column=8).value == 4200


# ── Engine 整合 ─────────────────────────────────────────────────────────────


class TestEngineIntegration:
    def test_hourly_uses_entries_total_when_present(self, art_client):
        """hourly 員工該月有 entries → calculate_salary 用 sum(entries) 覆寫 hourly_total。"""
        engine = SalaryEngine(load_from_db=False)
        emp_dict = {
            "employee_id": "A001",
            "name": "Vadim",
            "employee_type": "hourly",
            "hourly_rate": 620,
            "work_hours": 100,  # 應被忽略
            "base_salary": 0,
            "art_teacher_entries_total": 18600,  # 外師 16120 + 課後 2480
        }
        breakdown = engine.calculate_salary(emp_dict, 2026, 4)
        assert breakdown.hourly_total == 18600
        assert breakdown.gross_salary == 18600

    def test_hourly_fallback_when_no_entries(self, art_client):
        """無 entries → fallback 到 hourly_rate × work_hours。"""
        engine = SalaryEngine(load_from_db=False)
        emp_dict = {
            "employee_id": "A001",
            "name": "歐瑞煌",
            "employee_type": "hourly",
            "hourly_rate": 550,
            "work_hours": 25,
            "base_salary": 0,
        }
        breakdown = engine.calculate_salary(emp_dict, 2026, 4)
        assert breakdown.hourly_total == 13750
        assert breakdown.gross_salary == 13750


# ── API ─────────────────────────────────────────────────────────────────────


class TestApi:
    def test_create_recomputes_base_and_total(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            emp_id = emp.id
            session.commit()
        _login(client, session_factory)

        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "美語",
                "classroom_label": "向.滿",
                "hours": 25,
                "hourly_rate": 620,
                "excess_amount": 0,
                "activity_bonus": 0,
            },
        )
        assert res.status_code == 200
        body = res.json()
        assert body["base_amount"] == 15500  # 25 × 620
        assert body["total_amount"] == 15500

    def test_create_with_excess_and_bonus(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session, name="鍾馨瑤")
            emp_id = emp.id
            session.commit()
        _login(client, session_factory)

        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "舞蹈",
                "classroom_label": "(二)",
                "hours": 4,
                "hourly_rate": 1000,
                "excess_amount": 200,
                "activity_bonus": 0,
            },
        )
        body = res.json()
        assert body["base_amount"] == 4000
        assert body["total_amount"] == 4200

    def test_create_rejects_regular_employee(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = Employee(
                employee_id="R001",
                name="正職",
                employee_type="regular",
                is_active=True,
            )
            session.add(emp)
            session.commit()
            emp_id = emp.id
        _login(client, session_factory)

        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "test",
                "hours": 1,
                "hourly_rate": 100,
            },
        )
        assert res.status_code == 400

    def test_update_recomputes(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            entry = ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="美語",
                hours=10,
                hourly_rate=500,
                base_amount=5000,
                total_amount=5000,
            )
            session.add(entry)
            session.commit()
            eid = entry.id
        _login(client, session_factory)

        res = client.put(
            f"/api/art-teacher-payroll/{eid}",
            json={"hours": 20, "activity_bonus": 1000},
        )
        body = res.json()
        assert body["base_amount"] == 10000  # 20 × 500
        assert body["total_amount"] == 11000  # 10000 + 1000

    def test_delete(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            entry = ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="美語",
                hours=1,
                hourly_rate=100,
                base_amount=100,
                total_amount=100,
            )
            session.add(entry)
            session.commit()
            eid = entry.id
        _login(client, session_factory)

        res = client.delete(f"/api/art-teacher-payroll/{eid}")
        assert res.status_code == 200

    def test_list_orders_by_employee_id(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp1 = _add_hourly_emp(session, employee_id="A002", name="Vadim")
            emp2 = _add_hourly_emp(session, employee_id="A001", name="歐瑞煌")
            session.add_all(
                [
                    ArtTeacherPayrollEntry(
                        employee_id=emp1.id,
                        salary_year=2026,
                        salary_month=4,
                        subject="外師",
                        hours=25,
                        hourly_rate=620,
                        base_amount=15500,
                        total_amount=15500,
                    ),
                    ArtTeacherPayrollEntry(
                        employee_id=emp2.id,
                        salary_year=2026,
                        salary_month=4,
                        subject="美語",
                        hours=25,
                        hourly_rate=550,
                        base_amount=13750,
                        total_amount=13750,
                    ),
                ]
            )
            session.commit()
        _login(client, session_factory)

        res = client.get("/api/art-teacher-payroll", params={"year": 2026, "month": 4})
        items = res.json()["items"]
        assert len(items) == 2
        # A001 → A002
        assert items[0]["employee_name"] == "歐瑞煌"
        assert items[1]["employee_name"] == "Vadim"

    def test_roster_export_returns_xlsx(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            session.add(
                ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    subject="美語",
                    hours=25,
                    hourly_rate=550,
                    base_amount=13750,
                    total_amount=13750,
                )
            )
            session.commit()
        _login(client, session_factory)

        res = client.get("/api/art-teacher-payroll/2026/4/roster")
        assert res.status_code == 200
        assert res.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_import_template_returns_xlsx(self, art_client):
        client, session_factory = art_client
        _login(client, session_factory)
        res = client.get("/api/art-teacher-payroll/import-template")
        assert res.status_code == 200
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert "員工姓名" in headers
        assert "科目" in headers
        assert "時數" in headers
        assert "鐘點費" in headers

    def test_batch_import_happy_path(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            _add_hourly_emp(session, employee_id="A001", name="歐瑞煌")
            _add_hourly_emp(session, employee_id="A002", name="鍾馨瑤")
            session.commit()
        _login(client, session_factory)

        # 建構上傳 xlsx
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "員工姓名",
                "工號(選填)",
                "科目",
                "班級備註",
                "時數",
                "鐘點費",
                "超額",
                "加給活動",
                "備註",
            ]
        )
        ws.append(["歐瑞煌", "", "美語", "向.滿", 25, 550, 0, 0, ""])
        ws.append(["鍾馨瑤", "", "舞蹈", "(二)", 4, 1000, 200, 0, ""])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert res.status_code == 200
        body = res.json()
        assert body["total"] == 2
        assert body["imported"] == 2
        assert body["skipped"] == 0

        # 驗證 DB
        res2 = client.get("/api/art-teacher-payroll?year=2026&month=4")
        items = res2.json()["items"]
        assert len(items) == 2
        # 自動算 base/total
        oh = [i for i in items if i["employee_name"] == "歐瑞煌"][0]
        assert oh["base_amount"] == 13750
        assert oh["total_amount"] == 13750
        zh = [i for i in items if i["employee_name"] == "鍾馨瑤"][0]
        assert zh["base_amount"] == 4000
        assert zh["total_amount"] == 4200

    def test_batch_import_skips_missing_employee(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            _add_hourly_emp(session, employee_id="A001", name="歐瑞煌")
            session.commit()
        _login(client, session_factory)

        wb = Workbook()
        ws = wb.active
        ws.append(["員工姓名", "科目", "時數", "鐘點費"])
        ws.append(["歐瑞煌", "美語", 25, 550])
        ws.append(["不存在的人", "舞蹈", 4, 1000])
        buf = BytesIO()
        wb.save(buf)

        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        body = res.json()
        assert body["imported"] == 1
        assert body["skipped"] == 1
        assert len(body["errors"]) == 1
        assert body["errors"][0]["row"] == 3
        assert "找不到" in body["errors"][0]["message"]

    def test_batch_import_replace_existing(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            session.add(
                ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    subject="舊資料",
                    hours=10,
                    hourly_rate=100,
                    base_amount=1000,
                    total_amount=1000,
                )
            )
            session.commit()
        _login(client, session_factory)

        wb = Workbook()
        ws = wb.active
        ws.append(["員工姓名", "科目", "時數", "鐘點費"])
        ws.append(["才藝老師", "新資料", 20, 200])
        buf = BytesIO()
        wb.save(buf)

        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4&replace_existing=true",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert res.json()["imported"] == 1

        res2 = client.get("/api/art-teacher-payroll?year=2026&month=4")
        items = res2.json()["items"]
        assert len(items) == 1
        assert items[0]["subject"] == "新資料"

    def test_batch_import_missing_required_header(self, art_client):
        client, session_factory = art_client
        _login(client, session_factory)

        wb = Workbook()
        ws = wb.active
        ws.append(["員工姓名", "科目"])  # 缺時數/鐘點費
        ws.append(["X", "Y"])
        buf = BytesIO()
        wb.save(buf)

        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert res.status_code == 400
        assert "範本" in res.json()["detail"]

    def test_requires_salary_write_for_create(self, art_client):
        client, session_factory = art_client
        with session_factory() as session:
            emp = _add_hourly_emp(session)
            emp_id = emp.id
            session.commit()
        _login(client, session_factory, perm=["SALARY_READ"])
        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "x",
                "hours": 1,
                "hourly_rate": 1,
            },
        )
        assert res.status_code in (401, 403)


def _login_role(client, session_factory, *, role, perm):
    """以指定 role + perm 建 user 並登入（_login 寫死 admin，這裡測非 admin 角色）。"""
    uname = f"u_{role}"
    with session_factory() as session:
        session.add(
            User(
                username=uname,
                password_hash=hash_password("TempPass123"),
                role=role,
                permission_names=perm,
                is_active=True,
                must_change_password=False,
            )
        )
        session.commit()
    res = client.post(
        "/api/auth/login", json={"username": uname, "password": "TempPass123"}
    )
    assert res.status_code == 200


class TestArtPayrollViewerScope:
    """才藝薪資明細 list viewer 守衛（稽核 2026-06-03 P3-11）。

    只持 SALARY_READ 的角色（如園長 principal）不可越權看全所才藝老師逐筆給付；
    admin/hr（full salary view）或持 SALARY_WRITE（會計，本就建立/管理才藝薪資）可看全部。
    """

    def test_salary_read_only_role_cannot_list(self, art_client):
        client, sf = art_client
        _login_role(client, sf, role="principal", perm=["SALARY_READ"])
        res = client.get("/api/art-teacher-payroll", params={"year": 2026, "month": 4})
        assert res.status_code == 403, res.text

    def test_salary_write_role_can_list(self, art_client):
        client, sf = art_client
        _login_role(client, sf, role="accountant", perm=["SALARY_READ", "SALARY_WRITE"])
        res = client.get("/api/art-teacher-payroll", params={"year": 2026, "month": 4})
        assert res.status_code == 200, res.text

    def test_admin_can_list(self, art_client):
        client, sf = art_client
        _login_role(client, sf, role="admin", perm=["SALARY_READ"])
        res = client.get("/api/art-teacher-payroll", params={"year": 2026, "month": 4})
        assert res.status_code == 200, res.text


def _finalized_record(session, emp_id, year=2026, month=4):
    """建一筆封存（鎖定）的 SalaryRecord。"""
    rec = SalaryRecord(
        employee_id=emp_id,
        salary_year=year,
        salary_month=month,
        is_finalized=True,
        needs_recalc=False,
        finalized_by="acct",
    )
    session.add(rec)
    return rec


def _draft_record(session, emp_id, year=2026, month=4):
    """建一筆未封存草稿 SalaryRecord（needs_recalc=False）。"""
    rec = SalaryRecord(
        employee_id=emp_id,
        salary_year=year,
        salary_month=month,
        is_finalized=False,
        needs_recalc=False,
    )
    session.add(rec)
    return rec


class TestFinalizeGuard:
    """才藝鐘點明細是薪資引擎的 hourly_total 來源（engine.py:3104）。
    若該月薪資已封存仍可改來源，明細/薪資紀錄/轉帳清冊/財報會互相對不起來；
    未封存月份改動則必須標 needs_recalc 讓 finalize 完整性檢查攔下。
    """

    def test_create_blocked_when_month_finalized(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session)
            emp_id = emp.id
            session.flush()
            _finalized_record(session, emp_id)
            session.commit()
        _login(client, sf)
        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "美語",
                "hours": 25,
                "hourly_rate": 620,
            },
        )
        assert res.status_code == 409, res.text

    def test_update_blocked_when_month_finalized(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session)
            entry = ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="美語",
                hours=10,
                hourly_rate=500,
                base_amount=5000,
                total_amount=5000,
            )
            session.add(entry)
            session.flush()
            _finalized_record(session, emp.id)
            session.commit()
            eid = entry.id
        _login(client, sf)
        res = client.put(f"/api/art-teacher-payroll/{eid}", json={"hours": 20})
        assert res.status_code == 409, res.text

    def test_delete_blocked_when_month_finalized(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session)
            entry = ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="美語",
                hours=1,
                hourly_rate=100,
                base_amount=100,
                total_amount=100,
            )
            session.add(entry)
            session.flush()
            _finalized_record(session, emp.id)
            session.commit()
            eid = entry.id
        _login(client, sf)
        res = client.delete(f"/api/art-teacher-payroll/{eid}")
        assert res.status_code == 409, res.text

    def test_batch_import_blocked_when_month_finalized(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session, name="歐瑞煌")
            session.flush()
            _finalized_record(session, emp.id)
            session.commit()
        _login(client, sf)

        wb = Workbook()
        ws = wb.active
        ws.append(["員工姓名", "科目", "時數", "鐘點費"])
        ws.append(["歐瑞煌", "美語", 25, 550])
        buf = BytesIO()
        wb.save(buf)
        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert res.status_code == 409, res.text

    def test_create_marks_draft_salary_stale(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session)
            emp_id = emp.id
            session.flush()
            _draft_record(session, emp_id)
            session.commit()
        _login(client, sf)
        res = client.post(
            "/api/art-teacher-payroll",
            json={
                "employee_id": emp_id,
                "salary_year": 2026,
                "salary_month": 4,
                "subject": "美語",
                "hours": 25,
                "hourly_rate": 620,
            },
        )
        assert res.status_code == 200, res.text
        with sf() as session:
            rec = (
                session.query(SalaryRecord)
                .filter_by(employee_id=emp_id, salary_year=2026, salary_month=4)
                .first()
            )
            assert rec.needs_recalc is True

    def test_update_marks_draft_salary_stale(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session)
            emp_id = emp.id
            entry = ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="美語",
                hours=10,
                hourly_rate=500,
                base_amount=5000,
                total_amount=5000,
            )
            session.add(entry)
            session.flush()
            _draft_record(session, emp_id)
            session.commit()
            eid = entry.id
        _login(client, sf)
        res = client.put(f"/api/art-teacher-payroll/{eid}", json={"hours": 20})
        assert res.status_code == 200, res.text
        with sf() as session:
            rec = (
                session.query(SalaryRecord)
                .filter_by(employee_id=emp_id, salary_year=2026, salary_month=4)
                .first()
            )
            assert rec.needs_recalc is True

    def test_batch_import_marks_draft_salary_stale(self, art_client):
        client, sf = art_client
        with sf() as session:
            emp = _add_hourly_emp(session, name="歐瑞煌")
            emp_id = emp.id
            session.flush()
            _draft_record(session, emp_id)
            session.commit()
        _login(client, sf)

        wb = Workbook()
        ws = wb.active
        ws.append(["員工姓名", "科目", "時數", "鐘點費"])
        ws.append(["歐瑞煌", "美語", 25, 550])
        buf = BytesIO()
        wb.save(buf)
        res = client.post(
            "/api/art-teacher-payroll/batch-import?year=2026&month=4",
            files={
                "file": (
                    "in.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert res.status_code == 200, res.text
        assert res.json()["imported"] == 1
        with sf() as session:
            rec = (
                session.query(SalaryRecord)
                .filter_by(employee_id=emp_id, salary_year=2026, salary_month=4)
                .first()
            )
            assert rec.needs_recalc is True
