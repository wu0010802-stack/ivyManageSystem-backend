"""SystemConfig 通用 CRUD 測試。"""

import os
import sys
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from api.salary import init_salary_services
from api.salary import router as salary_router
from api.system_config import router as system_config_router
from models.database import Base, SystemConfig, User
from services.salary.engine import SalaryEngine
from utils.auth import hash_password
from utils.permissions import Permission


@pytest.fixture
def sc_client(tmp_path):
    db_path = tmp_path / "sc.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()
    init_salary_services(SalaryEngine(load_from_db=False), MagicMock())

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)
    app.include_router(system_config_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _login(client, session_factory, username="sc_admin", perm=None):
    if perm is None:
        perm = int(Permission.SETTINGS_READ | Permission.SETTINGS_WRITE)
    with session_factory() as session:
        session.add(
            User(
                username=username,
                password_hash=hash_password("TempPass123"),
                role="admin",
                permissions=perm,
                is_active=True,
                must_change_password=False,
            )
        )
        session.commit()
    res = client.post(
        "/api/auth/login",
        json={"username": username, "password": "TempPass123"},
    )
    assert res.status_code == 200


class TestSystemConfigApi:
    def test_list_returns_defaults_when_empty(self, sc_client):
        """DB 無資料時，list 仍要回 KNOWN_DEFAULTS 中的 key（is_default=True）。"""
        client, session_factory = sc_client
        _login(client, session_factory)
        res = client.get("/api/system-configs?prefix=bank")
        assert res.status_code == 200
        items = res.json()["items"]
        keys = {i["config_key"] for i in items}
        assert "bank.payer_name" in keys
        assert "bank.payer_account" in keys
        # 預設值
        payer_name = [i for i in items if i["config_key"] == "bank.payer_name"][0]
        assert payer_name["is_default"] is True
        assert payer_name["config_value"] == "高雄市私立常春藤幼兒園"

    def test_upsert_creates_new_record(self, sc_client):
        client, session_factory = sc_client
        _login(client, session_factory)

        res = client.put(
            "/api/system-configs/bank.payer_account",
            json={"config_value": "9999-888-777666"},
        )
        assert res.status_code == 200
        body = res.json()
        assert body["config_value"] == "9999-888-777666"
        assert body["is_default"] is False

        # 二次 GET 應取得 DB 值
        res2 = client.get("/api/system-configs/bank.payer_account")
        assert res2.json()["config_value"] == "9999-888-777666"
        assert res2.json()["is_default"] is False

    def test_upsert_updates_existing_record(self, sc_client):
        client, session_factory = sc_client
        with session_factory() as session:
            session.add(
                SystemConfig(
                    config_key="bank.payer_account",
                    config_value="1111-222-333",
                    config_type="bank",
                )
            )
            session.commit()
        _login(client, session_factory)

        res = client.put(
            "/api/system-configs/bank.payer_account",
            json={"config_value": "4444-555-666"},
        )
        assert res.status_code == 200
        assert res.json()["config_value"] == "4444-555-666"

        # DB 應只有一筆
        with session_factory() as session:
            count = (
                session.query(SystemConfig)
                .filter(SystemConfig.config_key == "bank.payer_account")
                .count()
            )
            assert count == 1

    def test_get_unknown_key_returns_404(self, sc_client):
        client, session_factory = sc_client
        _login(client, session_factory)
        res = client.get("/api/system-configs/totally.unknown")
        assert res.status_code == 404

    def test_get_known_default_when_empty(self, sc_client):
        """KNOWN_DEFAULTS 中的 key，即使 DB 無記錄也應回 default。"""
        client, session_factory = sc_client
        _login(client, session_factory)
        res = client.get("/api/system-configs/bank.payer_name")
        assert res.status_code == 200
        body = res.json()
        assert body["is_default"] is True
        assert body["config_value"] == "高雄市私立常春藤幼兒園"

    def test_prefix_filter(self, sc_client):
        client, session_factory = sc_client
        with session_factory() as session:
            session.add(
                SystemConfig(
                    config_key="other.foo",
                    config_value="bar",
                    config_type="general",
                )
            )
            session.commit()
        _login(client, session_factory)
        res = client.get("/api/system-configs?prefix=bank")
        items = res.json()["items"]
        keys = {i["config_key"] for i in items}
        assert "other.foo" not in keys
        assert "bank.payer_name" in keys

    def test_update_requires_settings_write(self, sc_client):
        client, session_factory = sc_client
        _login(
            client,
            session_factory,
            username="ro_user",
            perm=int(Permission.SETTINGS_READ),
        )
        res = client.put(
            "/api/system-configs/bank.payer_account",
            json={"config_value": "x"},
        )
        assert res.status_code in (401, 403)

    def test_transfer_roster_uses_updated_config(self, sc_client):
        """更新後，transfer_roster 匯出應使用新值（端到端整合）。"""
        from io import BytesIO

        from openpyxl import load_workbook

        from models.database import Employee, SalaryRecord

        client, session_factory = sc_client

        # 設定新付款帳號
        with session_factory() as session:
            session.add(
                SystemConfig(
                    config_key="bank.payer_account",
                    config_value="CUSTOM-ACC-9999",
                    config_type="bank",
                )
            )
            session.add(
                SystemConfig(
                    config_key="bank.payer_name",
                    config_value="自訂園所名",
                    config_type="bank",
                )
            )
            emp = Employee(
                employee_id="E001",
                name="王小明",
                employee_type="regular",
                base_salary=30000,
                is_active=True,
                bank_account="1234-5678",
                bank_account_name="王小明",
            )
            session.add(emp)
            session.flush()
            session.add(
                SalaryRecord(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    net_salary=30000,
                    is_finalized=True,
                )
            )
            session.commit()

        _login(
            client,
            session_factory,
            perm=int(
                Permission.SETTINGS_READ
                | Permission.SETTINGS_WRITE
                | Permission.SALARY_READ
            ),
        )

        res = client.get("/api/salaries/2026/4/transfer-roster?type=base")
        assert res.status_code == 200
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        assert ws["B1"].value == "自訂園所名"
        assert ws["B3"].value == "帳號：CUSTOM-ACC-9999"
