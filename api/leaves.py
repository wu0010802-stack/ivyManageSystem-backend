"""
Leave management router
"""

import asyncio
import json
import logging
import calendar as cal_module
from pathlib import Path
from datetime import date
from typing import Optional, List, Any
from io import BytesIO

from config import settings

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from utils.errors import raise_safe_500, safe_batch_reason
from utils.excel_utils import SafeWorksheet
from utils.rate_limit import create_limiter

# 批次核准為重 DB 操作；每分鐘 10 次緩衝正常工作量，封住批次濫用
_batch_approve_limiter = create_limiter(
    max_calls=10,
    window_seconds=60,
    name="leave_batch_approve",
    error_detail="批次審核操作過於頻繁，請稍後再試",
).as_dependency()
from pydantic import BaseModel, Field, field_validator, model_validator
from utils.leave_validators import validate_leave_hours_value, validate_leave_date_order
from utils.validators import validate_hhmm_format
from utils.leave_overtime_conflict import times_overlap
from sqlalchemy import or_, and_

from models.database import (
    get_session,
    Employee,
    LeaveRecord,
    LeaveQuota,
    OvertimeRecord,
    SalaryRecord,
    User,
)
from models.approval import ApprovalStatus
from schemas._common import MutationResultOut
from schemas.leaves import (
    LeaveApproveResultOut,
    LeaveDeleteResultOut,
    LeaveImportResultOut,
    LeaveUpdateResultOut,
)
from utils.auth import require_staff_permission
from utils.error_messages import EMPLOYEE_DOES_NOT_EXIST, LEAVE_RECORD_NOT_FOUND
from utils.permissions import Permission
from utils.approval_helpers import (
    _get_submitter_role,
    _check_approval_eligibility,
    _write_approval_log,
    assert_approver_eligible,
    collect_months_from_date_range,
    is_self_approval,
)
from services.salary.utils import lock_and_premark_stale, mark_salary_stale
from services.salary.finalize_guard import (
    collect_months_from_range,
    assert_months_not_finalized,
)
from utils.excel_utils import xlsx_streaming_response
from utils.excel_io import ExcelImportSchema, parse_excel
from utils.import_utils import build_employee_lookup, resolve_employee_from_row
from utils.file_upload import read_upload_with_size_check, validate_file_signature
from utils.storage import get_storage_path
from api.leaves_quota import (
    quota_router,
    LEAVE_TYPE_LABELS,
    LEAVE_DEDUCTION_RULES,
    _check_leave_limits,
    _check_quota,
    _check_compensatory_quota,
    _get_sick_committed_hours,
    assert_sick_leave_within_statutory_caps,
)
from api.leaves_workday import workday_router, validate_leave_hours_against_schedule
from services.leave_policy import requires_supporting_document
from services.approval.cross_type_offset import resolve_cross_type_offset

_UPLOAD_MODULE = "leave_attachments"


def _upload_base() -> Path:
    return get_storage_path(_UPLOAD_MODULE)


def _parse_paths(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        return json.loads(raw)
    except Exception:
        return []


def _safe_attach_path(leave_id: int, filename: str) -> Path:
    """解析附件路徑並確認落在 upload base 之內（路徑穿越防護）。"""
    base = _upload_base()
    resolved = (base / str(leave_id) / filename).resolve()
    try:
        resolved.relative_to(base.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="無效的附件路徑")
    return resolved


logger = logging.getLogger(__name__)


def _guard_leave_quota(
    session,
    employee_id: int,
    leave_type: str,
    year: int,
    leave_hours: float,
    is_hospitalized: bool,
    exclude_id: int = None,
    include_pending: bool = True,
) -> None:
    """sick 走勞工請假規則第 4 條雙配額；compensatory 走補休專用配額（quota 不存在=0）；
    其他假別走 _check_quota 單一配額。"""
    if leave_type == "sick":
        from datetime import date as _date  # local re-import safe

        out_used = _get_sick_committed_hours(
            session, employee_id, year, is_hospitalized=False, exclude_id=exclude_id
        )
        hosp_used = _get_sick_committed_hours(
            session, employee_id, year, is_hospitalized=True, exclude_id=exclude_id
        )
        assert_sick_leave_within_statutory_caps(
            out_used, hosp_used, leave_hours, bool(is_hospitalized)
        )
    elif leave_type == "compensatory":
        _check_compensatory_quota(
            session,
            employee_id,
            year,
            leave_hours,
            exclude_id=exclude_id,
            include_pending=include_pending,
        )
    else:
        _check_quota(
            session,
            employee_id,
            leave_type,
            year,
            leave_hours,
            include_pending=include_pending,
            exclude_id=exclude_id,
        )


def _consume_compensatory_grants_fifo(session, employee_id: int, hours: float) -> None:
    """FIFO 從最早 expires_at 的 active grant 扣 consumed_hours。

    若 active grant 總額不足，raise ValueError（不允許超扣，前端應已驗 quota）。
    """
    from models.overtime_comp_leave_grant import OvertimeCompLeaveGrant

    remaining = float(hours)
    grants = (
        session.query(OvertimeCompLeaveGrant)
        .filter(
            OvertimeCompLeaveGrant.employee_id == employee_id,
            OvertimeCompLeaveGrant.status == "active",
        )
        .order_by(OvertimeCompLeaveGrant.expires_at.asc())
        .all()
    )
    for g in grants:
        if remaining <= 0:
            break
        available = g.granted_hours - g.consumed_hours
        if available <= 0:
            continue
        take = min(available, remaining)
        g.consumed_hours += take
        remaining -= take
    if remaining > 0:
        raise ValueError(f"補休 grant 不足扣抵：尚缺 {remaining} 小時")


def _release_compensatory_grants_fifo(session, employee_id: int, hours: float) -> None:
    """退回 consumed_hours：FIFO 從最早 expires_at 的 grant 開始退。

    保守版：按最早 expires_at 順序退，維持總結餘正確。
    """
    from models.overtime_comp_leave_grant import OvertimeCompLeaveGrant

    remaining = float(hours)
    grants = (
        session.query(OvertimeCompLeaveGrant)
        .filter(
            OvertimeCompLeaveGrant.employee_id == employee_id,
            OvertimeCompLeaveGrant.status == "active",
            OvertimeCompLeaveGrant.consumed_hours > 0,
        )
        .order_by(OvertimeCompLeaveGrant.expires_at.asc())
        .all()
    )
    for g in grants:
        if remaining <= 0:
            break
        take = min(g.consumed_hours, remaining)
        g.consumed_hours -= take
        remaining -= take


def _apply_leave_update_and_revoke(leave, data, current_user, leave_id: int) -> None:
    """將 update 欄位套用到 leave，並在已核准時執行退審邏輯。

    此函式負責：
    1. 以 setattr 套用所有非 None 欄位
    2. 假別更換時重設 deduction_ratio / is_deductible
    3. 若原狀態為已核准，自動退回待審核並記錄警告

    Args:
        leave:        LeaveRecord ORM 物件（會被就地修改）
        data:         LeaveUpdate Pydantic schema
        current_user: 操作者資訊（供 audit log）
        leave_id:     假單 ID（供 audit log）
    """
    update_data = data.model_dump(exclude_unset=True)
    # 修補 2026-05-11 P1-7：start_time/end_time 允許明確傳 null 清空（半日↔全日場景）；
    # 其他欄位維持「不傳=不改、傳 null=不改」的舊行為避免破壞契約。
    _NULLABLE_FIELDS = {"start_time", "end_time"}
    for key, value in update_data.items():
        if value is None and key not in _NULLABLE_FIELDS:
            continue
        setattr(leave, key, value)

    # 假別更換時重設 deduction_ratio，但若本次同時明確傳入 deduction_ratio 則以傳入值為準
    if data.leave_type and data.leave_type in LEAVE_DEDUCTION_RULES:
        if data.deduction_ratio is None:
            leave.deduction_ratio = LEAVE_DEDUCTION_RULES[data.leave_type]
        leave.is_deductible = leave.deduction_ratio > 0

    # ── 稽核退審：已核准的記錄被修改，自動退回待審核 ──────────────────────
    was_approved = leave.status == ApprovalStatus.APPROVED.value
    if was_approved:
        leave.status = ApprovalStatus.PENDING.value
        leave.approved_by = None
        leave.rejection_reason = None
        logger.warning(
            "稽核警告：已核准請假記錄 #%d（員工 ID=%d, %s~%s, %s）被管理員「%s」修改，"
            "已自動退回待審核狀態，需重新核准",
            leave_id,
            leave.employee_id,
            leave.start_date,
            leave.end_date,
            leave.leave_type,
            current_user.get("username", "unknown"),
        )


router = APIRouter(prefix="/api", tags=["leaves"])
router.include_router(quota_router)
router.include_router(workday_router)

# ============ Service Injection ============

_salary_engine = None


def init_leaves_services(salary_engine_instance):
    global _salary_engine
    _salary_engine = salary_engine_instance


# ============ Pydantic Models ============


class LeaveCreate(BaseModel):
    employee_id: int
    leave_type: str
    start_date: date
    end_date: date
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    leave_hours: float = 8
    reason: Optional[str] = None
    is_hospitalized: bool = False  # 病假住院旗標（勞工請假規則第 4 條）
    deduction_ratio: Optional[float] = Field(
        None,
        ge=0.0,
        le=1.0,
        description="扣薪比例覆蓋（不提供則依假別預設值，0.0=全薪，1.0=全扣）",
    )

    @field_validator("leave_type")
    @classmethod
    def validate_leave_type(cls, v):
        if v not in LEAVE_DEDUCTION_RULES:
            raise ValueError(f"無效的假別: {v}")
        return v

    @field_validator("leave_hours")
    @classmethod
    def validate_leave_hours(cls, v):
        return validate_leave_hours_value(v)

    @field_validator("start_time", "end_time")
    @classmethod
    def _normalize_time(cls, v):
        # 統一補零成 HH:MM，避免 '9:00' 這類未補零字串在重疊偵測時走字典序誤判
        return validate_hhmm_format(v)

    @model_validator(mode="after")
    def validate_date_order(self):
        validate_leave_date_order(self.start_date, self.end_date)
        return self

    @model_validator(mode="after")
    def _validate_partial_leave_times(self) -> "LeaveCreate":
        if self.leave_hours is not None and self.leave_hours < 8:
            if not self.start_time or not self.end_time:
                raise ValueError(
                    "部分請假(leave_hours<8)必須提供 start_time 與 end_time"
                )
        return self


class LeaveUpdate(BaseModel):
    leave_type: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    leave_hours: Optional[float] = None
    reason: Optional[str] = None
    is_hospitalized: Optional[bool] = None
    deduction_ratio: Optional[float] = Field(
        None,
        ge=0.0,
        le=1.0,
        description="扣薪比例覆蓋（不提供則依假別預設值，0.0=全薪，1.0=全扣）",
    )

    @field_validator("leave_type")
    @classmethod
    def validate_leave_type(cls, v):
        if v is not None and v not in LEAVE_DEDUCTION_RULES:
            raise ValueError(f"無效的假別: {v}")
        return v

    @field_validator("leave_hours")
    @classmethod
    def validate_leave_hours(cls, v):
        if v is not None:
            if v < 0.5:
                raise ValueError("請假時數至少 0.5 小時")
            if v > 480:
                raise ValueError("請假時數不得超過 480 小時")
            if round(v * 2) != v * 2:
                raise ValueError("請假時數必須為 0.5 小時的倍數（如 0.5、1、1.5、2…）")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def _normalize_time(cls, v):
        # 統一補零成 HH:MM，避免 '9:00' 這類未補零字串在重疊偵測時走字典序誤判
        return validate_hhmm_format(v)

    @model_validator(mode="after")
    def validate_date_order(self):
        if self.start_date and self.end_date and self.end_date < self.start_date:
            raise ValueError("結束日期不得早於開始日期")
        if (
            self.start_date
            and self.end_date
            and (
                self.start_date.year != self.end_date.year
                or self.start_date.month != self.end_date.month
            )
        ):
            raise ValueError("請假區間不可跨月，若需跨越月底請拆成兩張假單分別申請")
        return self

    @model_validator(mode="after")
    def _validate_partial_leave_times(self) -> "LeaveUpdate":
        # 只在 leave_hours 出現且 <8 時檢查
        if self.leave_hours is not None and self.leave_hours < 8:
            if not self.start_time or not self.end_time:
                raise ValueError(
                    "部分請假(leave_hours<8)必須提供 start_time 與 end_time"
                )
        return self


# ============ Helpers ============


def _check_substitute_guard(leave, *, allow_without_substitute: bool = False) -> None:
    """序列式守衛：核准假單前確認代理人已接受。

    - pending  → 409（等待代理人回應）
    - rejected → 409（需重新指定代理人）
    - accepted / not_required → 放行
    """
    status = getattr(leave, "substitute_status", "not_required") or "not_required"
    if allow_without_substitute and status in {"pending", "rejected"}:
        return
    if status == "pending":
        raise HTTPException(
            status_code=409,
            detail="代理人尚未接受，請等待代理人回應後再核准",
        )
    if status == "rejected":
        raise HTTPException(
            status_code=409,
            detail="代理人已拒絕此代理請求，請要求員工重新指定代理人後再核准",
        )


# F1 第一波：_find_overlapping_leave / _check_overlap 抽到 services/leave_overlap_service.py
# 維持本檔既有 import surface（api/portal/leaves.py 仍 `from api.leaves import _check_overlap`）。
from services.leave_overlap_service import (
    find_overlapping_leave as _find_overlapping_leave,
    find_approved_overlapping_leave as _check_overlap,
)


def _check_employee_has_conflicting_overtime(
    session,
    employee_id: int,
    start_date: date,
    end_date: date,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
) -> None:
    """申請請假時檢查同員工同時段是否已有 approved/pending 加班。

    修補 2026-05-11 P1-5：請假與加班不互查重疊，導致同日扣款 + 加班費雙重溢付。

    時段比對規則：
    - leave 全日（start_time/end_time 為 None）→ 與 OT 同日就衝突
    - leave 半日（HH:MM）→ 與 OT 時段比對；OT 缺時段視為全日衝突

    NOTE: 目前只在 create 路徑使用。若未來在 update 路徑也呼叫此 helper，需新增
    exclude_overtime_id 參數避免自我衝突；同步調整 _check_employee_has_conflicting_leave。
    """
    from models.database import OvertimeRecord  # avoid circular import at module load

    candidates = (
        session.query(OvertimeRecord)
        .filter(
            OvertimeRecord.employee_id == employee_id,
            OvertimeRecord.status.in_(
                [ApprovalStatus.PENDING.value, ApprovalStatus.APPROVED.value]
            ),
            OvertimeRecord.overtime_date >= start_date,
            OvertimeRecord.overtime_date <= end_date,
        )
        .all()
    )
    for ot in candidates:
        # leave 全日 → 同日有 OT 即衝突
        if start_time is None or end_time is None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"員工於 {ot.overtime_date} 已有加班申請 #{ot.id}，"
                    "請假與加班時段重疊"
                ),
            )
        # leave 半日 → OT 缺時段視為全日 → 衝突
        if ot.start_time is None or ot.end_time is None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"員工於 {ot.overtime_date} 已有加班申請 #{ot.id}（全日），"
                    "請假時段與加班重疊"
                ),
            )
        # 時段精比：用 times_overlap（內部 to_time 正規化），避免未補零字串字典序誤判
        ot_start_str = ot.start_time.strftime("%H:%M")
        ot_end_str = ot.end_time.strftime("%H:%M")
        if times_overlap(start_time, end_time, ot.start_time, ot.end_time):
            raise HTTPException(
                status_code=409,
                detail=(
                    f"員工於 {ot.overtime_date} 已有加班申請 #{ot.id}"
                    f"（{ot_start_str}~{ot_end_str}），請假時段與其重疊"
                ),
            )


def _check_substitute_leave_conflict(
    session,
    substitute_employee_id: Optional[int],
    start_date: date,
    end_date: date,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
) -> None:
    """代理人不可在相同時段有待審或已核准的請假或加班記錄（V14）。

    F-005：detail 採 generic 訊息，不再揭露代理人的請假/加班區間與審核狀態。
    舊實作回傳「代理人在 {start_date} ~ {end_date} 已有{待審核|已核准}請假記錄」
    讓員工 A 可用 substitute_employee_id=B + 不同日期反覆探測 B 的請假/加班行事曆
    （二分搜尋還原整段排程）。改為單一中性訊息後，攻擊者無法從 409 detail 推回
    代理人的具體日期或審批狀態。管理端（api/leaves.py 核准路徑）也會走此 helper，
    一併收斂訊息（避免維護兩套 detail）。
    """
    if substitute_employee_id is None:
        return

    _GENERIC_DETAIL = "代理人於該期間已有其他請假/加班，無法擔任代理人，請改選其他人選"

    # ── 代理人 active 檢查（修補 2026-05-11 P1-10）──────────────────────
    # 離職或停用的代理人不可被指定；approve 時也要重驗（建立時是 active，
    # approve 時可能已離職）。
    sub_emp = (
        session.query(Employee).filter(Employee.id == substitute_employee_id).first()
    )
    if sub_emp is None or not sub_emp.is_active:
        raise HTTPException(
            status_code=400,
            detail="代理人不存在或已離職／停用，請改選其他人選",
        )

    # ── 檢查請假衝突 ────────────────────────────────────────────────────
    leave_conflict = _find_overlapping_leave(
        session,
        substitute_employee_id,
        start_date,
        end_date,
        start_time=start_time,
        end_time=end_time,
        include_pending=True,
    )
    if leave_conflict:
        raise HTTPException(status_code=409, detail=_GENERIC_DETAIL)

    # ── 檢查加班衝突（V14；修補 2026-05-11 P1-9 改時段精比）──────────────
    # 代理人 OT 與請假時段需精細比對：申請人半日假 08:00-12:00 vs 代理人
    # 同日 18:00-20:00 OT 不應誤判衝突；舊實作只比 overtime_date 區間。
    ot_candidates = (
        session.query(OvertimeRecord)
        .filter(
            OvertimeRecord.employee_id == substitute_employee_id,
            OvertimeRecord.status.in_(
                [ApprovalStatus.PENDING.value, ApprovalStatus.APPROVED.value]
            ),
            OvertimeRecord.overtime_date >= start_date,
            OvertimeRecord.overtime_date <= end_date,
        )
        .all()
    )
    for ot in ot_candidates:
        # 申請人請假是全日（無 start_time/end_time）→ 與 OT 同日即衝突
        if not start_time or not end_time:
            raise HTTPException(status_code=409, detail=_GENERIC_DETAIL)
        # OT 無時段（罕見）→ 視為全日 → 衝突
        if not ot.start_time or not ot.end_time:
            raise HTTPException(status_code=409, detail=_GENERIC_DETAIL)
        ot_start_str = ot.start_time.strftime("%H:%M")
        ot_end_str = ot.end_time.strftime("%H:%M")
        if max(start_time, ot_start_str) < min(end_time, ot_end_str):
            raise HTTPException(status_code=409, detail=_GENERIC_DETAIL)


# ============ Routes ============


@router.get("/leaves")
def get_leaves(
    employee_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_READ)),
):
    """查詢請假記錄"""
    session = get_session()
    try:
        q = session.query(LeaveRecord, Employee).join(
            Employee, LeaveRecord.employee_id == Employee.id
        )
        if employee_id:
            q = q.filter(LeaveRecord.employee_id == employee_id)
        if status == "pending":
            q = q.filter(LeaveRecord.status == ApprovalStatus.PENDING.value)
        elif status == "approved":
            q = q.filter(LeaveRecord.status == ApprovalStatus.APPROVED.value)
        elif status == "rejected":
            q = q.filter(LeaveRecord.status == ApprovalStatus.REJECTED.value)
        if year and month:
            _, last_day = cal_module.monthrange(year, month)
            start = date(year, month, 1)
            end = date(year, month, last_day)
            q = q.filter(LeaveRecord.start_date <= end, LeaveRecord.end_date >= start)
        elif year:
            q = q.filter(
                LeaveRecord.start_date >= date(year, 1, 1),
                LeaveRecord.start_date <= date(year, 12, 31),
            )

        records = q.order_by(LeaveRecord.start_date.desc()).limit(5000).all()

        # 預先載入員工角色映射（減少逐筆查 DB 的 N+1 問題）
        from models.database import User as UserModel

        employee_ids = list({leave.employee_id for leave, _ in records})
        user_roles = {}
        if employee_ids:
            users = (
                session.query(UserModel)
                .filter(
                    UserModel.employee_id.in_(employee_ids),
                    UserModel.is_active == True,
                )
                .all()
            )
            user_roles = {u.employee_id: u.role for u in users}

        # 預先載入代理人姓名（批次查詢，避免 N+1）
        substitute_ids = list(
            {
                leave.substitute_employee_id
                for leave, _ in records
                if leave.substitute_employee_id
            }
        )
        substitute_names: dict = {}
        if substitute_ids:
            subs = session.query(Employee).filter(Employee.id.in_(substitute_ids)).all()
            substitute_names = {s.id: s.name for s in subs}

        # 預先載入換班關聯（同員工、假單區間內有 pending/accepted 的換班申請）
        from models.database import ShiftSwapRequest

        # 只做一次聚合查詢，按 requester_id + swap_date 建立快速索引
        swap_map: dict = {}
        if records:
            involved_emp_ids = list({lv.employee_id for lv, _ in records})
            # 計算所有假單的整體日期範圍，在 DB 層過濾 swap_date，避免 Python 端大量掃描
            _all_starts = [lv.start_date for lv, _ in records]
            _all_ends = [lv.end_date for lv, _ in records]
            _swap_range_start = min(_all_starts)
            _swap_range_end = max(_all_ends)
            all_swaps = (
                session.query(ShiftSwapRequest)
                .filter(
                    ShiftSwapRequest.requester_id.in_(involved_emp_ids),
                    ShiftSwapRequest.status.in_(["pending", "accepted"]),
                    ShiftSwapRequest.swap_date >= _swap_range_start,
                    ShiftSwapRequest.swap_date <= _swap_range_end,
                )
                .all()
            )
            for sw in all_swaps:
                swap_map.setdefault(sw.requester_id, []).append(sw)

        results = []
        for leave, emp in records:
            # 換班關聯：查詢同員工、同期間的換班申請
            related_swap = None
            for sw in swap_map.get(leave.employee_id, []):
                if leave.start_date <= sw.swap_date <= leave.end_date:
                    related_swap = {
                        "id": sw.id,
                        "swap_date": sw.swap_date.isoformat(),
                        "status": sw.status,
                        "target_id": sw.target_id,
                    }
                    break

            results.append(
                {
                    "id": leave.id,
                    "employee_id": leave.employee_id,
                    "employee_name": emp.name,
                    "submitter_role": user_roles.get(leave.employee_id, "teacher"),
                    "leave_type": leave.leave_type,
                    "leave_type_label": LEAVE_TYPE_LABELS.get(
                        leave.leave_type, leave.leave_type
                    ),
                    "start_date": leave.start_date.isoformat(),
                    "end_date": leave.end_date.isoformat(),
                    "start_time": leave.start_time,
                    "end_time": leave.end_time,
                    "leave_hours": leave.leave_hours,
                    "deduction_ratio": leave.deduction_ratio,
                    "reason": leave.reason,
                    "status": leave.status,
                    "approved_by": leave.approved_by,
                    "rejection_reason": leave.rejection_reason,
                    "attachment_paths": _parse_paths(leave.attachment_paths),
                    "substitute_employee_id": leave.substitute_employee_id,
                    "substitute_employee_name": substitute_names.get(
                        leave.substitute_employee_id
                    ),
                    "substitute_status": leave.substitute_status or "not_required",
                    "substitute_responded_at": (
                        leave.substitute_responded_at.isoformat()
                        if leave.substitute_responded_at
                        else None
                    ),
                    "related_swap": related_swap,
                    "created_at": (
                        leave.created_at.isoformat() if leave.created_at else None
                    ),
                }
            )
        return results
    finally:
        session.close()


# ── 請假記錄 CRUD ──────────────────────────────────────────────


@router.post("/leaves", status_code=201, response_model=MutationResultOut)
def create_leave(
    data: LeaveCreate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """新增請假記錄"""
    session = get_session()
    try:
        emp = session.query(Employee).filter(Employee.id == data.employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail=EMPLOYEE_DOES_NOT_EXIST)

        overlap = _check_overlap(
            session,
            data.employee_id,
            data.start_date,
            data.end_date,
            data.start_time,
            data.end_time,
        )
        if overlap:
            raise HTTPException(
                status_code=409,
                detail=f"該員工在 {overlap.start_date} ~ {overlap.end_date} 已有已核准的請假記錄（ID: {overlap.id}），無法重複請假",
            )

        # 修補 2026-05-11 P1-5：跨類重疊檢查
        _check_employee_has_conflicting_overtime(
            session,
            data.employee_id,
            data.start_date,
            data.end_date,
            data.start_time,
            data.end_time,
        )

        validate_leave_hours_against_schedule(
            session,
            data.employee_id,
            data.start_date,
            data.end_date,
            data.leave_hours,
            data.start_time,
            data.end_time,
        )

        _check_leave_limits(
            session,
            data.employee_id,
            data.leave_type,
            data.start_date,
            data.leave_hours,
        )
        _guard_leave_quota(
            session,
            data.employee_id,
            data.leave_type,
            data.start_date.year,
            data.leave_hours,
            bool(data.is_hospitalized),
        )

        # 優先使用 API 傳入的覆蓋值；未提供則依假別預設規則
        effective_ratio = (
            data.deduction_ratio
            if data.deduction_ratio is not None
            else LEAVE_DEDUCTION_RULES[data.leave_type]
        )
        if data.deduction_ratio is not None:
            default_ratio = LEAVE_DEDUCTION_RULES.get(data.leave_type, -1)
            logger.warning(
                "假單扣薪比例被手動覆蓋 by user=%s employee_id=%s leave_type=%s "
                "default_ratio=%s overridden_ratio=%s",
                current_user.get("username"),
                data.employee_id,
                data.leave_type,
                default_ratio,
                data.deduction_ratio,
            )
        leave = LeaveRecord(
            employee_id=data.employee_id,
            leave_type=data.leave_type,
            start_date=data.start_date,
            end_date=data.end_date,
            start_time=data.start_time,
            end_time=data.end_time,
            leave_hours=data.leave_hours,
            is_deductible=effective_ratio > 0,
            deduction_ratio=effective_ratio,
            reason=data.reason,
            is_hospitalized=(
                bool(data.is_hospitalized) if data.leave_type == "sick" else False
            ),
        )
        session.add(leave)
        session.commit()

        request.state.audit_entity_id = str(leave.id)
        request.state.audit_summary = (
            f"管理端建立請假：employee_id={data.employee_id} "
            f"{data.leave_type} {data.start_date}~{data.end_date}（{data.leave_hours}h）"
        )
        request.state.audit_changes = {
            "action": "leave_create",
            "leave_id": leave.id,
            "employee_id": data.employee_id,
            "leave_type": data.leave_type,
            "start_date": data.start_date.isoformat(),
            "end_date": data.end_date.isoformat(),
            "start_time": str(data.start_time) if data.start_time else None,
            "end_time": str(data.end_time) if data.end_time else None,
            "leave_hours": data.leave_hours,
            "is_deductible": effective_ratio > 0,
            "deduction_ratio": effective_ratio,
            "deduction_ratio_overridden": data.deduction_ratio is not None,
            "is_hospitalized": bool(data.is_hospitalized),
            "reason": data.reason,
        }
        return {"message": "請假記錄已新增", "id": leave.id}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/leaves/{leave_id}", response_model=LeaveUpdateResultOut)
def update_leave(
    leave_id: int,
    data: LeaveUpdate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """更新請假記錄。若記錄已核准，修改後自動退回「待審核」狀態以符合稽核要求。"""
    session = get_session()
    try:
        # 列鎖（修補 2026-05-11 P0-3）：與 approve 路徑對齊，防止並發 update+approve
        # 造成 lost update（補休配額負數、deduction_ratio 不同步等）。
        leave = (
            session.query(LeaveRecord)
            .filter(LeaveRecord.id == leave_id)
            .with_for_update()
            .first()
        )
        if not leave:
            raise HTTPException(status_code=404, detail=LEAVE_RECORD_NOT_FOUND)

        # 在 setattr 套用新日期前，捕捉原始狀態（同時供 audit_changes 留下完整 before）
        was_approved = leave.status == ApprovalStatus.APPROVED.value
        orig_month = (leave.start_date.year, leave.start_date.month)
        # 用 getattr 容忍部分欄位缺失（SQLAlchemy 物件正常情況下都有；測試 fake 物件較寬鬆）
        before_snapshot = {
            "leave_type": getattr(leave, "leave_type", None),
            "start_date": leave.start_date.isoformat(),
            "end_date": leave.end_date.isoformat(),
            "start_time": (
                str(leave.start_time) if getattr(leave, "start_time", None) else None
            ),
            "end_time": (
                str(leave.end_time) if getattr(leave, "end_time", None) else None
            ),
            "leave_hours": getattr(leave, "leave_hours", None),
            "status": leave.status,
            "is_hospitalized": bool(getattr(leave, "is_hospitalized", False)),
            "deduction_ratio": getattr(leave, "deduction_ratio", None),
            "reason": getattr(leave, "reason", None),
        }

        # 計算更新後的日期 / 時間（未傳入的欄位沿用原值）
        new_start = data.start_date or leave.start_date
        new_end = data.end_date or leave.end_date
        new_start_time = (
            data.start_time if data.start_time is not None else leave.start_time
        )
        new_end_time = data.end_time if data.end_time is not None else leave.end_time

        # 跨月檢查：更新後的區間也不允許跨月
        if new_start.year != new_end.year or new_start.month != new_end.month:
            raise HTTPException(
                status_code=400,
                detail=(
                    "請假區間不可跨月，若需跨越月底請拆成兩張假單分別申請"
                    f"（更新後 {new_start.year}/{new_start.month:02d} 月 →"
                    f" {new_end.year}/{new_end.month:02d} 月）"
                ),
            )

        overlap = _check_overlap(
            session,
            leave.employee_id,
            new_start,
            new_end,
            new_start_time,
            new_end_time,
            exclude_id=leave_id,
        )
        if overlap:
            raise HTTPException(
                status_code=409,
                detail=f"修改後的日期與已核准的請假記錄重疊（{overlap.start_date} ~ {overlap.end_date}，ID: {overlap.id}）",
            )

        new_type = data.leave_type or leave.leave_type
        new_hours = (
            data.leave_hours if data.leave_hours is not None else leave.leave_hours
        )
        validate_leave_hours_against_schedule(
            session,
            leave.employee_id,
            new_start,
            new_end,
            new_hours,
            new_start_time,
            new_end_time,
        )
        # 已核准的假單退審後視同重新提交：重新過一次配額（排除自身）
        _check_leave_limits(
            session,
            leave.employee_id,
            new_type,
            new_start,
            new_hours,
            exclude_id=leave_id,
        )
        new_is_hosp = (
            data.is_hospitalized
            if data.is_hospitalized is not None
            else leave.is_hospitalized
        )
        _guard_leave_quota(
            session,
            leave.employee_id,
            new_type,
            new_start.year,
            new_hours,
            new_is_hosp,
            exclude_id=leave_id,
        )

        # 封存月薪保護：同時檢查原始月份與更新後月份
        # 跨月假單（理論上 schema 已擋，但歷史資料/直接 DB 寫入仍可能存在）的
        # 原區間每一個月份都會在 collect_months_from_range 重算迴圈中被觸及，
        # 必須涵蓋完整月份集合，與 line 935 的 months_to_recalc 對齊。
        if was_approved:
            _affected_months = collect_months_from_range(
                leave.start_date, leave.end_date
            ) | collect_months_from_range(new_start, new_end)
            assert_months_not_finalized(
                session,
                employee_id=leave.employee_id,
                months=_affected_months,
            )
            # commit→recalc 鎖延伸：取得 per-emp salary lock 並 pre-mark stale,
            # 確保 commit 釋放鎖後即使 finalize 搶先,看到的也是 needs_recalc=True 而被擋下。

            lock_and_premark_stale(session, leave.employee_id, _affected_months)

        # ── 在 model 寫回前 snapshot（reapply 需要舊範圍）─────────────────────
        old_sync_snapshot = {
            "start_date": leave.start_date,
            "end_date": leave.end_date,
            "start_time": leave.start_time,
            "end_time": leave.end_time,
            "leave_type": leave.leave_type,
            "leave_hours": leave.leave_hours,
            "status": leave.status,
        }

        _apply_leave_update_and_revoke(leave, data, current_user, leave_id)

        # ── 考勤同步 hook（Hook 3/4）────────────────────────────────────────────
        # Hook 3（退審路徑）：was_approved=True → _apply_leave_update_and_revoke 後
        #   leave.status='pending' → revert
        # Hook 4（改關鍵欄仍 approved，罕見）：reapply
        # LeaveAttendanceConflict / LeavePartialTimeMissing → 422
        from services import employee_leave_attendance_sync as sync

        _key_fields_changed = any(
            old_sync_snapshot[k] != getattr(leave, k)
            for k in (
                "start_date",
                "end_date",
                "start_time",
                "end_time",
                "leave_type",
                "leave_hours",
            )
        )
        try:
            if (
                old_sync_snapshot["status"] == ApprovalStatus.APPROVED.value
                and leave.status == ApprovalStatus.PENDING.value
            ):
                # 退審路徑（Hook 3）
                sync.revert(session, leave_id)
            elif (
                old_sync_snapshot["status"] == ApprovalStatus.APPROVED.value
                and leave.status == ApprovalStatus.APPROVED.value
                and _key_fields_changed
            ):
                # 改關鍵欄但仍 approved（Hook 4，罕見）
                sync.reapply(session, leave_id, old_snapshot=old_sync_snapshot)
        except (sync.LeaveAttendanceConflict, sync.LeavePartialTimeMissing) as e:
            raise HTTPException(status_code=422, detail=str(e))
        # ────────────────────────────────────────────────────────────────────────

        session.commit()

        after_snapshot = {
            "leave_type": getattr(leave, "leave_type", None),
            "start_date": leave.start_date.isoformat(),
            "end_date": leave.end_date.isoformat(),
            "start_time": (
                str(leave.start_time) if getattr(leave, "start_time", None) else None
            ),
            "end_time": (
                str(leave.end_time) if getattr(leave, "end_time", None) else None
            ),
            "leave_hours": getattr(leave, "leave_hours", None),
            "status": leave.status,
            "is_hospitalized": bool(getattr(leave, "is_hospitalized", False)),
            "deduction_ratio": getattr(leave, "deduction_ratio", None),
            "reason": getattr(leave, "reason", None),
        }
        diff = {
            k: {"before": before_snapshot[k], "after": after_snapshot[k]}
            for k in before_snapshot
            if before_snapshot[k] != after_snapshot[k]
        }
        request.state.audit_entity_id = str(leave_id)
        request.state.audit_summary = (
            f"管理端修改請假 #{leave_id}（employee_id={leave.employee_id}）"
            + ("；自動退回待審" if was_approved else "")
        )
        request.state.audit_changes = {
            "action": "leave_update",
            "leave_id": leave_id,
            "employee_id": leave.employee_id,
            "was_approved": was_approved,
            "reset_to_pending": was_approved,
            "diff": diff,
            "orig_month": f"{orig_month[0]}-{orig_month[1]:02d}",
        }

        result = {"message": "請假記錄已更新"}
        if was_approved:
            result["message"] += "；原核准狀態已自動退回「待審核」，請重新送審"
            result["reset_to_pending"] = True
            if _salary_engine is not None:
                # 跨月修改時,新區間只涵蓋新月份；舊月份的扣款須一併撤銷重算,
                # 否則 orig_month 的舊扣款仍存在於該月薪資,與新月份合計形成重複扣薪。
                months_to_recalc = collect_months_from_range(
                    leave.start_date, leave.end_date
                )
                months_to_recalc.add(orig_month)
                try:
                    for yr, mo in sorted(months_to_recalc):
                        _salary_engine.process_salary_calculation(
                            leave.employee_id, yr, mo
                        )
                    result["salary_recalculated"] = True
                except Exception as e:
                    result["salary_warning"] = (
                        "薪資重算失敗，請手動前往薪資頁面重新計算"
                    )
                    logger.error("請假修改退審後薪資重算失敗：%s", e)
                    # 重算失敗 → 標 stale,避免後續 finalize 把「假單已退審但薪資未更新」
                    # 的舊資料封存。對齊 approve_leave 降級樣板。
                    for yr, mo in sorted(months_to_recalc):
                        try:
                            mark_salary_stale(session, leave.employee_id, yr, mo)
                        except Exception:
                            logger.warning(
                                "請假修改降級時標記 SalaryRecord stale 失敗 emp=%d %d/%d",
                                leave.employee_id,
                                yr,
                                mo,
                                exc_info=True,
                            )
                    try:
                        session.commit()
                    except Exception:
                        logger.warning(
                            "請假修改降級時 commit stale 標記失敗", exc_info=True
                        )
                        session.rollback()

        return result
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/leaves/{leave_id}", response_model=LeaveDeleteResultOut)
def delete_leave(
    leave_id: int,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """刪除請假記錄"""
    session = get_session()
    try:
        # 列鎖（修補 2026-05-11 P0-3）：與 approve 路徑對齊，防止並發 delete+approve race。
        leave = (
            session.query(LeaveRecord)
            .filter(LeaveRecord.id == leave_id)
            .with_for_update()
            .first()
        )
        if not leave:
            raise HTTPException(status_code=404, detail=LEAVE_RECORD_NOT_FOUND)

        # ── 封存保護：已核准假單在封存月份不得刪除 ──────────────────────────
        was_approved = leave.status == ApprovalStatus.APPROVED.value
        # 跨月假單需涵蓋完整月份集合，與 lock_and_premark_stale 的範圍對齊。
        leave_months = collect_months_from_range(leave.start_date, leave.end_date)
        emp_id = leave.employee_id
        # 預先 snapshot：刪除後 leave 物件被 expunge，audit_changes 必須在這裡備份
        deleted_snapshot = {
            "leave_id": leave_id,
            "employee_id": emp_id,
            "leave_type": leave.leave_type,
            "start_date": leave.start_date.isoformat(),
            "end_date": leave.end_date.isoformat(),
            "leave_hours": leave.leave_hours,
            "status": leave.status,
            "deduction_ratio": leave.deduction_ratio,
            "reason": leave.reason,
        }
        if was_approved:
            assert_months_not_finalized(
                session, employee_id=emp_id, months=leave_months
            )
            # commit→recalc 鎖延伸：見 update_leave 同款註解（封 finalize race window）

            lock_and_premark_stale(session, emp_id, leave_months)

        # ── 考勤同步 hook（Hook 5: delete_leave revert）───────────────────────
        # approved leave 被刪 → 先 revert attendance，再 delete LeaveRecord。
        # FK ON DELETE SET NULL 是雙保險，但主路徑是 revert 主動清。
        if leave.status == ApprovalStatus.APPROVED.value:
            try:
                from services import employee_leave_attendance_sync as sync

                sync.revert(session, leave_id)
            except Exception as e:
                from services.employee_leave_attendance_sync import (
                    LeaveAttendanceConflict,
                )

                if isinstance(e, LeaveAttendanceConflict):
                    raise HTTPException(status_code=422, detail=str(e))
                raise

        # ── 補休 grant 退回（已核准補休假單被刪除）───────────────────────────
        if was_approved and leave.leave_type == "compensatory":
            _release_compensatory_grants_fifo(
                session, leave.employee_id, leave.leave_hours
            )
        # ─────────────────────────────────────────────────────────────────────

        session.delete(leave)
        session.commit()

        request.state.audit_entity_id = str(leave_id)
        request.state.audit_summary = (
            f"管理端刪除請假 #{leave_id}（employee_id={emp_id}）"
            + ("；已核准 → 觸發薪資重算" if was_approved else "")
        )
        request.state.audit_changes = {
            "action": "leave_delete",
            "deleted": deleted_snapshot,
            "was_approved": was_approved,
            "leave_months": sorted(f"{yr}-{mo:02d}" for yr, mo in leave_months),
            "triggered_salary_recalc": was_approved,
        }

        result = {"message": "請假記錄已刪除"}
        # 刪除已核准假單後補算薪資，撤銷原扣款；跨月假單需重算每一個月份。
        if was_approved and _salary_engine is not None:
            try:
                for yr, mo in sorted(leave_months):
                    _salary_engine.process_salary_calculation(emp_id, yr, mo)
                result["salary_recalculated"] = True
            except Exception as e:
                result["salary_warning"] = (
                    "假單已刪除，但薪資重算失敗，請手動前往薪資頁面重新計算"
                )
                logger.error("刪除假單後薪資重算失敗：%s", e)
                # 重算失敗 → 標 stale,避免後續 finalize 在「假單已刪除但扣款未撤銷」
                # 的狀態下誤封存舊薪資。對齊 approve_leave 降級樣板。
                for yr, mo in sorted(leave_months):
                    try:
                        mark_salary_stale(session, emp_id, yr, mo)
                    except Exception:
                        logger.warning(
                            "刪除假單降級時標記 SalaryRecord stale 失敗 emp=%d %d/%d",
                            emp_id,
                            yr,
                            mo,
                            exc_info=True,
                        )
                try:
                    session.commit()
                except Exception:
                    logger.warning(
                        "刪除假單降級時 commit stale 標記失敗", exc_info=True
                    )
                    session.rollback()

        return result
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


class ApproveRequest(BaseModel):
    approved: bool
    rejection_reason: Optional[str] = None
    force_without_substitute: bool = False
    # 同員工同時段已有其他已核准假單時，預設 409 阻擋；主管確認後可帶
    # force_overlap=True 強制過（記稽核日誌），對齊 force_without_substitute 模式。
    force_overlap: bool = False
    # force_overlap=True 時必填原因（≥10 字）；會寫進 ApprovalLog comment 與 logger，
    # 供日後稽核回溯為何允許重疊（避免重複扣薪/重複占配額被忽視）。
    force_overlap_reason: Optional[str] = Field(None, max_length=500)

    @model_validator(mode="after")
    def _force_overlap_requires_reason(self):
        if self.force_overlap:
            cleaned = (self.force_overlap_reason or "").strip()
            if len(cleaned) < 10:
                raise ValueError(
                    "force_overlap=True 時必須填寫 force_overlap_reason（至少 10 字）"
                )
            self.force_overlap_reason = cleaned
        return self


class LeaveBatchApproveRequest(BaseModel):
    ids: List[int]
    approved: bool
    rejection_reason: Optional[str] = None


@router.put("/leaves/{leave_id}/approve", response_model=LeaveApproveResultOut)
def approve_leave(
    leave_id: int,
    data: ApproveRequest,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """核准/駁回請假。駁回時 rejection_reason 為必填。"""
    if not data.approved and not (data.rejection_reason or "").strip():
        raise HTTPException(status_code=400, detail="駁回時必須填寫原因")
    session = get_session()
    try:
        leave = (
            session.query(LeaveRecord)
            .filter(LeaveRecord.id == leave_id)
            .with_for_update()
            .first()
        )
        if not leave:
            raise HTTPException(status_code=404, detail=LEAVE_RECORD_NOT_FOUND)

        # ── 自我核准防護 ─────────────────────────────────────────────────────────
        if is_self_approval(current_user, leave.employee_id):
            raise HTTPException(status_code=403, detail="不可自我核准請假單")

        # ── 角色資格檢查 ────────────────────────────────────────────────────────
        # 既有設計允許「已核准 → 駁回」的合法業務路徑（例如發現超時數需撤銷），
        # 由 risk_tags="reject_of_approved" 在 audit_summary 顯式打標記。撤回 audit
        # 2026-05-07 P1 hard-block flip-flop 提案：業主業務模型即是「approver 可改判」，
        # 硬擋會破壞合法 reject_of_approved 路徑（test_leave_reject_approved_regression
        # / TestApprovedOvertimeRollback 既有測試）。
        approver_role = current_user.get("role", "")
        submitter_role = assert_approver_eligible(
            session,
            doc_type="leave",
            doc_label="請假",
            submitter_employee_id=leave.employee_id,
            approver_role=approver_role,
        )

        # 核准/駁回狀態是否實質變動；True→False、False→True、None→True 都屬於
        # 會影響 SalaryRecord 扣款的轉換，必須觸發封存守衛與薪資重算。
        was_approved = leave.status == ApprovalStatus.APPROVED.value
        was_pending = leave.status == ApprovalStatus.PENDING.value
        was_rejected = leave.status == ApprovalStatus.REJECTED.value
        approval_changed = was_approved != data.approved
        # 整單 snapshot 用於 audit_changes 的 before/after
        leave_snapshot_before = {
            "status": leave.status,
            "leave_type": leave.leave_type,
            "leave_hours": leave.leave_hours,
            "deduction_ratio": leave.deduction_ratio,
            "approved_by": leave.approved_by,
            "rejection_reason": leave.rejection_reason,
            "substitute_status": leave.substitute_status,
        }
        # ⚠ 高風險事件偵測：對「已核准 → 修改」這條軌跡留下顯式旗標，
        # 方便前端 AuditLogView 篩出來作為高風險事件。
        is_reject_of_approved = was_approved and not data.approved
        # employee_id 也先抓出來；後面 commit + LINE 推播後 leave 物件還可用，
        # 但提前抓出減少耦合
        leave_employee_id = leave.employee_id

        # ── 提早取得薪資鎖 + pre-mark stale（封住兩個 race window）────────────
        # Why:
        # 1. 原流程「_check_finalized → commit leave → process_salary_calculation」
        #    中間沒有 lock,finalize 可能在 commit 之後 / recalc 之前搶到鎖,結果
        #    假單變更已落地但薪資沒重算就被封存。
        # 2. 即使 caller 在自己 session 上 acquire 了 advisory lock,commit 時鎖即
        #    釋放;engine 開新 session 才 acquire,中間有縫隙,finalize 仍可搶先。
        # 修法:在 caller session 同時做兩件事:
        #   - 取得 per-emp salary lock(關 race window 1)
        #   - mark_salary_stale 把該員工受影響月份預標 needs_recalc=True
        #     → commit 後即使 finalize 搶到鎖也會被 stale 守衛擋下(關 race window 2)
        if approval_changed:
            _months_for_lock = collect_months_from_date_range(
                leave.start_date, leave.end_date
            )
            lock_and_premark_stale(session, leave.employee_id, _months_for_lock)

        warning = None
        if data.approved:
            if requires_supporting_document(
                leave.start_date, leave.end_date
            ) and not _parse_paths(leave.attachment_paths):
                raise HTTPException(
                    status_code=400,
                    detail="請假超過 2 天需檢附證明附件後才能核准",
                )
            # ── 代理人序列式守衛 ──────────────────────────────────────────────
            _check_substitute_guard(
                leave, allow_without_substitute=data.force_without_substitute
            )
            if not data.force_without_substitute:
                _check_substitute_leave_conflict(
                    session,
                    leave.substitute_employee_id,
                    leave.start_date,
                    leave.end_date,
                    leave.start_time,
                    leave.end_time,
                )

            # 核准最後一道關卡：時數不得超過該區間排班工時。建立路徑（管理端 / portal）
            # 已檢查過，但匯入、舊資料或 DB 手改可能繞過，此處補一層 defense-in-depth。
            validate_leave_hours_against_schedule(
                session,
                leave.employee_id,
                leave.start_date,
                leave.end_date,
                leave.leave_hours,
                leave.start_time,
                leave.end_time,
            )

            # 重疊核准硬擋：同員工同時段已有其他已核准假單時，預設拒絕核准，
            # 避免重複扣薪或重複占配額。主管確認後仍要核准，需帶 force_overlap=True
            # 才會降級為 warning（會記稽核日誌）。
            conflict = _check_overlap(
                session,
                leave.employee_id,
                leave.start_date,
                leave.end_date,
                leave.start_time,
                leave.end_time,
                exclude_id=leave_id,
            )
            if conflict:
                if not data.force_overlap:
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"該員工在 {conflict.start_date} ~ {conflict.end_date} "
                            f"已有另一筆已核准的請假（ID: {conflict.id}），"
                            "若確認仍要核准請帶 force_overlap=true 並填寫 force_overlap_reason"
                        ),
                    )
                # ── force_overlap 三重守衛 ─────────────────────────────────
                # 重疊放行可能造成同時段重複扣薪/重複占配額；不該只憑 force=True 即過：
                # 1. force_overlap_reason ≥10 字（schema 已驗）
                # 2. ACTIVITY_PAYMENT_APPROVE 簽核（避免一線主管獨自決定）
                # 3. conflict 詳情寫進 ApprovalLog comment（永久稽核軌跡，不只是 logger）
                from utils.finance_guards import has_finance_approve

                if not has_finance_approve(current_user):
                    raise HTTPException(
                        status_code=403,
                        detail=(
                            "force_overlap=True 需具備『金流簽核』權限"
                            "（ACTIVITY_PAYMENT_APPROVE），請改由具該權限者執行"
                        ),
                    )
                conflict_summary = (
                    f"主管警告後核准（force_overlap）：與假單 #{conflict.id} "
                    f"({conflict.start_date}~{conflict.end_date}) 重疊；"
                    f"原因：{data.force_overlap_reason}"
                )
                warning = conflict_summary
                logger.warning(
                    "假單 #%d 主管以 force_overlap=true 強制核准，"
                    "與已核准假單 #%d (%s~%s) 重疊；操作者：%s；原因：%s",
                    leave_id,
                    conflict.id,
                    conflict.start_date,
                    conflict.end_date,
                    current_user.get("username", "unknown"),
                    data.force_overlap_reason,
                )

            # ── 配額硬檢查（核准動作）──────────────────────────────────────────
            # 使用 include_pending=True + exclude_id=leave_id：
            # 計算「所有已核准 + 其他待審（排除本張）+ 本次時數」是否超出年度配額。
            # 此策略防止主管同時批准多張待審假單造成配額超支（concurrent approval race）。
            _check_leave_limits(
                session,
                leave.employee_id,
                leave.leave_type,
                leave.start_date,
                leave.leave_hours,
                include_pending=True,
                exclude_id=leave_id,
            )
            _guard_leave_quota(
                session,
                leave.employee_id,
                leave.leave_type,
                leave.start_date.year,
                leave.leave_hours,
                bool(leave.is_hospitalized),
                exclude_id=leave_id,
                include_pending=True,
            )

        # ── 封存月薪保護（commit 前）────────────────────────────────────────
        # 核准假單或將已核准假單改為駁回都會改變 SalaryRecord；若該月薪資已封存，
        # 必須在 commit 前阻擋，否則假單狀態被翻面、薪資沒更新，DB 永遠處於矛盾狀態。
        # 跨月假單需檢查整個跨越區間（中間月份/end_date 月份），與上方 lock_and_premark_stale
        # 的範圍對齊，避免 check 漏放後 stale 標記被 finalize 守衛擋下、假單已 commit 的破口。
        if approval_changed:
            assert_months_not_finalized(
                session,
                employee_id=leave.employee_id,
                months=collect_months_from_range(leave.start_date, leave.end_date),
            )

        # ── V8：防禦性扣款比例同步 ───────────────────────────────────────────
        # 核准時強制確保 deduction_ratio 與假別標準一致，避免「假別變更後
        # ratio 未重設」造成薪資漏扣/誤扣。若 HR 需要自訂比例，應在核准前
        # 以假單編輯入口明確傳入 deduction_ratio（該路徑會設置 is_deductible
        # 並記錄來源）。
        if data.approved and leave.leave_type in LEAVE_DEDUCTION_RULES:
            standard_ratio = LEAVE_DEDUCTION_RULES[leave.leave_type]
            if leave.deduction_ratio is None:
                leave.deduction_ratio = standard_ratio
                logger.info(
                    "核准假單 #%d 時補全缺失的 deduction_ratio：%s → %.2f",
                    leave_id,
                    leave.leave_type,
                    standard_ratio,
                )
            elif leave.deduction_ratio != standard_ratio:
                logger.warning(
                    "核准假單 #%d deduction_ratio 與假別標準不符：%s 標準=%.2f 實際=%.2f，強制改回標準值",
                    leave_id,
                    leave.leave_type,
                    standard_ratio,
                    leave.deduction_ratio,
                )
                leave.deduction_ratio = standard_ratio
            leave.is_deductible = (leave.deduction_ratio or 0) > 0

        leave.status = (
            ApprovalStatus.APPROVED.value
            if data.approved
            else ApprovalStatus.REJECTED.value
        )
        leave.approved_by = (
            current_user.get("username", "管理員") if data.approved else None
        )
        leave.rejection_reason = (
            data.rejection_reason.strip()
            if not data.approved and data.rejection_reason
            else None
        )
        if data.approved and data.force_without_substitute:
            leave.substitute_status = "waived"

        action = "approved" if data.approved else "rejected"
        approval_comment = data.rejection_reason if not data.approved else None
        if data.approved and data.force_without_substitute:
            approval_comment = "主管警告後核准：未取得代理人接受"
        # 重疊放行：把 conflict 與原因寫進 ApprovalLog，留永久稽核痕跡
        if data.approved and warning:
            approval_comment = (
                f"{approval_comment}\n{warning}" if approval_comment else warning
            )
        approval_log_row = _write_approval_log(
            session=session,
            doc_type="leave",
            doc_id=leave_id,
            action=action,
            approver=current_user,
            comment=approval_comment,
        )

        # ── 考勤同步 hook（leave↔attendance sync）────────────────────────────
        # 在 ApprovalLog 寫入之後、cross_offset 及 commit 之前執行：
        # - approved=True 且本次為真實核准（was_approved=False/None）→ apply
        # - approved=False 且本次為撤銷已核准（was_approved=True）→ revert
        # LeaveAttendanceConflict / LeavePartialTimeMissing → 422；
        # 其他例外（如 RuntimeError）不被 catch，讓 FastAPI 500 + finally 的
        # session.close() 觸發隱式 rollback，確保 leave.status 不殘留。
        from services import employee_leave_attendance_sync as sync

        try:
            if data.approved is True and not was_approved:
                sync.apply(session, leave_id)
            elif data.approved is False and was_approved:
                sync.revert(session, leave_id)
        except (sync.LeaveAttendanceConflict, sync.LeavePartialTimeMissing) as e:
            raise HTTPException(status_code=422, detail=str(e))
        # ─────────────────────────────────────────────────────────────────────

        # ── 補休假單 grant ledger FIFO 扣抵 / 退回 ──────────────────────────
        # 僅在狀態真實變動時觸發，防止重複 approve 或重複 reject 雙重扣抵。
        # - pending/rejected → approved：消耗 grant consumed_hours（FIFO 最早 expires_at 優先）
        # - approved → rejected：退回 consumed_hours（對稱退回）
        # ValueError（grant 不足）理論上已被前端 _check_compensatory_quota 擋住；
        # 此處作為 defense-in-depth 讓 FastAPI 以 422 回應。
        if leave.leave_type == "compensatory":
            if data.approved is True and approval_changed:
                try:
                    _consume_compensatory_grants_fifo(
                        session, leave.employee_id, leave.leave_hours
                    )
                except ValueError as exc:
                    raise HTTPException(status_code=422, detail=str(exc))
            elif data.approved is False and was_approved:
                _release_compensatory_grants_fifo(
                    session, leave.employee_id, leave.leave_hours
                )
        # ─────────────────────────────────────────────────────────────────────

        # ── leave↔OT 跨類抵扣（feature flag gated, v1 metadata-only）─────────
        # 條件：(1) 由 pending/rejected → approved 的真實變動 (2) feature flag 開啟
        # v1 行為：偵測到同員工同日已核准 OT 時，僅在 ApprovalLog 留下 metadata 軌跡；
        # 不動 OvertimeRecord schema、不影響 salary engine。詳見 RELEASE_NOTES.md。
        cross_offset_ot_id = None
        if data.approved and approval_changed:
            try:
                offset_ot = resolve_cross_type_offset(session, leave)
                if offset_ot is not None:
                    cross_offset_ot_id = offset_ot.id
                    _write_approval_log(
                        session=session,
                        doc_type="overtime",
                        doc_id=offset_ot.id,
                        action="update",
                        approver=current_user,
                        comment="leave 跨類抵扣（auto, v1 metadata-only）",
                        metadata={
                            "offset_by_leave_id": leave_id,
                            "offset_date": str(leave.start_date),
                        },
                    )
                    logger.info(
                        "leave↔OT 跨類抵扣 v1：leave #%d 偵測到同日 OT #%d，已寫 ApprovalLog metadata",
                        leave_id,
                        offset_ot.id,
                    )
            except Exception as exc:
                # 跨類抵扣失敗不阻斷主流程；僅留 warning 由 audit 後續追查
                logger.warning(
                    "leave #%d 跨類抵扣偵測失敗：%s", leave_id, exc, exc_info=True
                )

        # 審核結果通知（dispatch，tx commit 前登錄；after_commit hook 自動 fan-out）
        # lazy import 避免 circular: api.leaves → dispatch → ws → contact_book_ws → portal → leaves
        from services.notification import dispatch

        _leave_owner_user = (
            session.query(User).filter(User.employee_id == leave.employee_id).first()
        )
        if _leave_owner_user is not None:
            dispatch.enqueue(
                session=session,
                event_type="leave.approved" if data.approved else "leave.rejected",
                recipient_user_id=_leave_owner_user.id,
                context={
                    "reviewer_name": current_user.get("name")
                    or current_user.get("username", ""),
                    "leave_type": leave.leave_type,
                    "start": (
                        leave.start_date.isoformat()
                        if hasattr(leave.start_date, "isoformat")
                        else str(leave.start_date)
                    ),
                    "end": (
                        leave.end_date.isoformat()
                        if hasattr(leave.end_date, "isoformat")
                        else str(leave.end_date)
                    ),
                    "leave_id": leave_id,
                    "rejection_reason": (
                        data.rejection_reason if not data.approved else None
                    ),
                },
                sender_id=current_user.get("user_id"),
                source_entity_type="leave_request",
                source_entity_id=leave_id,
            )

        session.commit()

        # AuditLog changes：留下完整的 before/after + ApprovalLog 連結
        # 並用 high_risk 旗標標出「force_overlap」「reject of approved」「force_without_substitute」
        request.state.audit_entity_id = str(leave_id)
        action_label = "核准" if data.approved else "駁回"
        risk_tags = []
        if data.force_overlap:
            risk_tags.append("force_overlap")
        if data.force_without_substitute:
            risk_tags.append("force_without_substitute")
        if is_reject_of_approved:
            risk_tags.append("reject_of_approved")
        request.state.audit_summary = (
            f"{action_label}請假 #{leave_id}（employee_id={leave_employee_id}）"
            + (f"｜⚠ {','.join(risk_tags)}" if risk_tags else "")
        )
        request.state.audit_changes = {
            "action": "leave_approve",
            "leave_id": leave_id,
            "employee_id": leave_employee_id,
            "decision": "approved" if data.approved else "rejected",
            "before": leave_snapshot_before,
            "after": {
                "status": leave.status,
                "approved_by": leave.approved_by,
                "rejection_reason": leave.rejection_reason,
                "deduction_ratio": leave.deduction_ratio,
                "substitute_status": leave.substitute_status,
            },
            "approval_log_id": approval_log_row.id if approval_log_row else None,
            "approval_changed": approval_changed,
            "is_reject_of_approved": is_reject_of_approved,
            "force_overlap": data.force_overlap,
            "force_overlap_reason": data.force_overlap_reason,
            "force_without_substitute": data.force_without_substitute,
            "rejection_reason": data.rejection_reason,
            "warning": warning,
            "risk_tags": risk_tags,
            "cross_offset_ot_id": cross_offset_ot_id,
        }

        result = {"message": "已核准" if data.approved else "已駁回"}
        if warning:
            result["warning"] = warning

        # 核准狀態變動（approve 或 reject-of-approved）後，自動重算該員工所有涉及月份的薪資
        if approval_changed and _salary_engine is not None:
            try:
                emp_id = leave.employee_id
                # 計算假單跨越的所有 (year, month)
                months_to_recalc = set()
                cur = date(leave.start_date.year, leave.start_date.month, 1)
                end = date(leave.end_date.year, leave.end_date.month, 1)
                while cur <= end:
                    months_to_recalc.add((cur.year, cur.month))
                    cur = (
                        date(cur.year + 1, 1, 1)
                        if cur.month == 12
                        else date(cur.year, cur.month + 1, 1)
                    )

                for year, month in sorted(months_to_recalc):
                    _salary_engine.process_salary_calculation(emp_id, year, month)
                    logger.info(
                        f"請假審核狀態變動後自動重算薪資：emp_id={emp_id}, {year}/{month}, approved={data.approved}"
                    )

                result["salary_recalculated"] = True
                if data.approved:
                    result["message"] = "已核准，薪資已自動重算"
                else:
                    result["message"] = "已駁回，薪資已自動重算"
            except Exception as e:
                result["salary_recalculated"] = False
                result["salary_warning"] = (
                    "操作成功，但薪資重算失敗，請手動前往薪資頁面重新計算"
                )
                logger.error(f"請假審核後薪資重算失敗：{e}")
                # 把所有應重算月份的 SalaryRecord 標 stale,避免後續 finalize
                # 在「假單已審核但薪資未更新」的狀態下誤封存舊薪資。
                for year, month in sorted(months_to_recalc):
                    try:
                        mark_salary_stale(session, emp_id, year, month)
                    except Exception:
                        logger.warning(
                            "請假審核降級時標記 SalaryRecord stale 失敗 emp=%d %d/%d",
                            emp_id,
                            year,
                            month,
                            exc_info=True,
                        )
                try:
                    session.commit()
                except Exception:
                    logger.warning(
                        "請假審核降級時 commit stale 標記失敗", exc_info=True
                    )
                    session.rollback()

        return result
    finally:
        session.close()


@router.post("/leaves/batch-approve")
def batch_approve_leaves(
    data: LeaveBatchApproveRequest,
    request: Request,
    _rl=Depends(_batch_approve_limiter),
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """批次核准/駁回請假。兩階段原子提交：先全部驗證，再統一 commit。"""
    if not data.approved and not (data.rejection_reason or "").strip():
        raise HTTPException(status_code=400, detail="批次駁回時必須填寫原因")

    succeeded = []
    failed = []

    session = get_session()
    try:
        # ── 預先批次載入，避免 N+1 ────────────────────────────────────────
        leave_map = {
            lv.id: lv
            for lv in session.query(LeaveRecord)
            .filter(LeaveRecord.id.in_(data.ids))
            .with_for_update()
            .all()
        }
        emp_ids = {lv.employee_id for lv in leave_map.values()}
        submitter_role_map: dict[int, str] = (
            {
                u.employee_id: u.role
                for u in session.query(User.employee_id, User.role)
                .filter(User.employee_id.in_(emp_ids), User.is_active == True)
                .all()
            }
            if emp_ids
            else {}
        )
        approver_role = current_user.get("role", "")
        _eligibility_cache: dict[str, bool] = {}

        # ── Pass 1：純驗證（不 touch ORM dirty state） ────────────────────
        # 修補 2026-05-11 P0-1：原本 setattr/_write_approval_log/lock_and_premark_stale
        # 都在 Phase 1 內，catch-all rollback 會抹掉同 batch 已通過條目的變更，造成
        # succeeded 與 DB 脫鉤（silent data loss）。改為「驗證→收集→Pass 2 統一套用」。
        changes = []  # list of (leave_id, leave)
        for leave_id in data.ids:
            try:
                leave = leave_map.get(leave_id)
                if not leave:
                    failed.append({"id": leave_id, "reason": LEAVE_RECORD_NOT_FOUND})
                    continue

                # 防止自我核准
                if is_self_approval(current_user, leave.employee_id):
                    failed.append({"id": leave_id, "reason": "不可自我核准"})
                    continue

                # 角色資格檢查（403 視為失敗條目，不中斷批次）
                submitter_role = submitter_role_map.get(leave.employee_id, "teacher")
                if submitter_role not in _eligibility_cache:
                    _eligibility_cache[submitter_role] = _check_approval_eligibility(
                        "leave", submitter_role, approver_role, session
                    )
                if not _eligibility_cache[submitter_role]:
                    failed.append(
                        {
                            "id": leave_id,
                            "reason": f"您的角色（{approver_role}）無權審核此員工（{submitter_role}）的請假申請",
                        }
                    )
                    continue

                # 核准/駁回狀態是否實質變動（影響 SalaryRecord）
                was_approved = leave.status == ApprovalStatus.APPROVED.value
                approval_changed = was_approved != data.approved
                # 高風險：對「已核准 → 駁回」批次操作要顯式標記
                is_reject_of_approved = was_approved and not data.approved

                if data.approved:
                    try:
                        if requires_supporting_document(
                            leave.start_date, leave.end_date
                        ) and not _parse_paths(leave.attachment_paths):
                            raise HTTPException(
                                status_code=400,
                                detail="請假超過 2 天需檢附證明附件後才能核准",
                            )
                        # 與單筆核准相同的防線：時數不得超過該區間排班工時，
                        # 防止 import、舊資料或 DB 手改繞過的超額假單進入薪資。
                        validate_leave_hours_against_schedule(
                            session,
                            leave.employee_id,
                            leave.start_date,
                            leave.end_date,
                            leave.leave_hours,
                            leave.start_time,
                            leave.end_time,
                        )
                        # 代理人序列式守衛
                        _check_substitute_guard(leave)
                        _check_substitute_leave_conflict(
                            session,
                            leave.substitute_employee_id,
                            leave.start_date,
                            leave.end_date,
                            leave.start_time,
                            leave.end_time,
                        )
                        _check_leave_limits(
                            session,
                            leave.employee_id,
                            leave.leave_type,
                            leave.start_date,
                            leave.leave_hours,
                            include_pending=True,
                            exclude_id=leave_id,
                        )
                        _guard_leave_quota(
                            session,
                            leave.employee_id,
                            leave.leave_type,
                            leave.start_date.year,
                            leave.leave_hours,
                            bool(leave.is_hospitalized),
                            exclude_id=leave_id,
                            include_pending=True,
                        )
                        # 重疊核准硬擋：批次無 force_overlap 旗標，一律拒絕。
                        # 借助 SQLAlchemy autoflush，前一輪已 set status='approved'
                        # 的記錄會在這個 _check_overlap 查詢時被視為已核准，
                        # 確保「同批兩張同員工同時段」的後者會被擋下。
                        conflict = _check_overlap(
                            session,
                            leave.employee_id,
                            leave.start_date,
                            leave.end_date,
                            leave.start_time,
                            leave.end_time,
                            exclude_id=leave_id,
                        )
                        if conflict:
                            raise HTTPException(
                                status_code=409,
                                detail=(
                                    f"與已核准假單 #{conflict.id}"
                                    f"（{conflict.start_date}~{conflict.end_date}）"
                                    "重疊；批次核准不支援強制過，請改用單筆核准並帶 force_overlap"
                                ),
                            )
                    except HTTPException as e:
                        failed.append({"id": leave_id, "reason": e.detail})
                        continue

                # 封存月薪保護：approve 或 reject-of-approved 都會改變 SalaryRecord
                # 跨月假單需涵蓋整個跨越區間（中間月份 / end_date 月份）。
                affected_months = None
                if approval_changed:
                    try:
                        affected_months = collect_months_from_range(
                            leave.start_date, leave.end_date
                        )
                        assert_months_not_finalized(
                            session,
                            employee_id=leave.employee_id,
                            months=affected_months,
                        )
                    except HTTPException as e:
                        failed.append({"id": leave_id, "reason": e.detail})
                        continue

                # Pass 1 只收集 validated metadata；setattr/log/lock 全部移到 Pass 2。
                changes.append(
                    (
                        leave_id,
                        leave,
                        approval_changed,
                        is_reject_of_approved,
                        affected_months,
                        None,  # approval_log_row.id placeholder (Pass 2 才填)
                    )
                )
            except Exception as e:
                # Pass 1 純查詢/驗證，幾乎不可能 dirty session；保留 rollback 防呆。
                session.rollback()
                failed.append(
                    {
                        "id": leave_id,
                        "reason": safe_batch_reason(e, context="批次請假核准"),
                    }
                )
                session.expire_all()

        # ── Pass 2：套用 setattr + lock + ApprovalLog（仍不 commit）──────
        applied = []
        for _i, (
            leave_id,
            leave,
            approval_changed,
            is_reject_of_approved,
            affected_months,
            _placeholder,
        ) in enumerate(list(changes)):
            try:
                if data.approved:
                    # 同批 in-batch overlap：autoflush 會把 Pass 2 前面已 setattr 的
                    # 條目視為 approved，達成「同員工同時段第二張被擋」效果。
                    conflict = _check_overlap(
                        session,
                        leave.employee_id,
                        leave.start_date,
                        leave.end_date,
                        leave.start_time,
                        leave.end_time,
                        exclude_id=leave_id,
                    )
                    if conflict:
                        failed.append(
                            {
                                "id": leave_id,
                                "reason": (
                                    f"與同批已核准假單 #{conflict.id}"
                                    f"（{conflict.start_date}~{conflict.end_date}）重疊"
                                ),
                            }
                        )
                        continue

                if approval_changed and affected_months:
                    # commit→recalc 鎖延伸 + pre-mark stale，封住 finalize race window。
                    lock_and_premark_stale(session, leave.employee_id, affected_months)

                # V8：批次核准時強制同步 deduction_ratio 到假別標準值
                if data.approved and leave.leave_type in LEAVE_DEDUCTION_RULES:
                    standard_ratio = LEAVE_DEDUCTION_RULES[leave.leave_type]
                    if leave.deduction_ratio is None:
                        leave.deduction_ratio = standard_ratio
                    elif leave.deduction_ratio != standard_ratio:
                        logger.warning(
                            "批次核准假單 #%d deduction_ratio 與假別標準不符：%s 標準=%.2f 實際=%.2f，強制改回標準值",
                            leave_id,
                            leave.leave_type,
                            standard_ratio,
                            leave.deduction_ratio,
                        )
                        leave.deduction_ratio = standard_ratio
                    leave.is_deductible = (leave.deduction_ratio or 0) > 0

                # ── 補休假單 grant ledger FIFO 扣抵 / 退回（批次版）──────────
                if leave.leave_type == "compensatory" and approval_changed:
                    if data.approved is True:
                        try:
                            _consume_compensatory_grants_fifo(
                                session, leave.employee_id, leave.leave_hours
                            )
                        except ValueError as exc:
                            failed.append({"id": leave_id, "reason": str(exc)})
                            continue
                    elif data.approved is False and is_reject_of_approved:
                        _release_compensatory_grants_fifo(
                            session, leave.employee_id, leave.leave_hours
                        )
                # ─────────────────────────────────────────────────────────────

                leave.status = (
                    ApprovalStatus.APPROVED.value
                    if data.approved
                    else ApprovalStatus.REJECTED.value
                )
                leave.approved_by = (
                    current_user.get("username", "管理員") if data.approved else None
                )
                leave.rejection_reason = (
                    data.rejection_reason.strip()
                    if not data.approved and data.rejection_reason
                    else None
                )
                action = "approved" if data.approved else "rejected"
                approval_log_row = _write_approval_log(
                    session=session,
                    doc_type="leave",
                    doc_id=leave_id,
                    action=action,
                    approver=current_user,
                    comment=data.rejection_reason if not data.approved else None,
                )
                applied.append(
                    (
                        leave_id,
                        leave,
                        approval_changed,
                        is_reject_of_approved,
                        approval_log_row.id if approval_log_row else None,
                    )
                )
            except Exception as e:
                # Pass 2 套用階段意外：整批中止；rollback 抹掉 applied 內已 setattr 條目。
                logger.exception(
                    "批次核准 Pass 2 套用階段意外失敗（leave #%d）", leave_id
                )
                session.rollback()
                for prev_id, *_ in applied:
                    failed.append(
                        {
                            "id": prev_id,
                            "reason": "同批後續條目失敗導致整批回滾",
                        }
                    )
                failed.append(
                    {
                        "id": leave_id,
                        "reason": safe_batch_reason(e, context="批次請假 Pass2"),
                    }
                )
                applied = []
                break

        # 用 applied 覆寫 changes，後續 Phase 2 (commit + LINE + 重算) 邏輯不變
        changes = applied

        # ── Phase 2：所有驗證通過後統一 commit ──────────────────────────────
        if changes:
            try:
                session.commit()

                # 批次預載通知所需的 User（1 查詢取代 N）
                _user_map: dict = {}
                change_emp_ids = list({lv.employee_id for _, lv, _, _, _ in changes})
                for u in (
                    session.query(User)
                    .filter(User.employee_id.in_(change_emp_ids))
                    .all()
                ):
                    _user_map[u.employee_id] = u

                # 通知 pending 列表：recalc 成功後收集，loop 結束後統一 enqueue+commit
                _pending_notifs: list = []

                for (
                    leave_id,
                    leave,
                    approval_changed,
                    _is_reject_of_approved,
                    _approval_log_id,
                ) in changes:
                    # approve 或 reject-of-approved 都需重算薪資
                    # Why: succeeded 必須等到「假單狀態 commit + 薪資重算成功」兩步都 OK
                    # 才寫入；不然會出現「假單已核准但薪資仍是舊值」的中間狀態，
                    # 呼叫端誤以為一切就緒並進 finalize，把錯薪資封存。
                    recalc_failed = False
                    if approval_changed and _salary_engine is not None:
                        emp_id = leave.employee_id
                        months: set = set()
                        cur = date(leave.start_date.year, leave.start_date.month, 1)
                        end_m = date(leave.end_date.year, leave.end_date.month, 1)
                        while cur <= end_m:
                            months.add((cur.year, cur.month))
                            cur = (
                                date(cur.year + 1, 1, 1)
                                if cur.month == 12
                                else date(cur.year, cur.month + 1, 1)
                            )
                        try:
                            for yr, mo in sorted(months):
                                _salary_engine.process_salary_calculation(
                                    emp_id, yr, mo
                                )
                        except Exception as se:
                            recalc_failed = True
                            logger.error(
                                "批次審核後薪資重算失敗（假單 #%d）：%s", leave_id, se
                            )
                            # 重算失敗 → 標 stale,避免後續 finalize 在「假單已審核
                            # 但薪資未更新」的狀態下誤封存舊薪資。
                            for yr, mo in sorted(months):
                                try:
                                    mark_salary_stale(session, emp_id, yr, mo)
                                except Exception:
                                    logger.warning(
                                        "批次審核降級時標記 SalaryRecord stale 失敗 emp=%d %d/%d",
                                        emp_id,
                                        yr,
                                        mo,
                                        exc_info=True,
                                    )
                            try:
                                session.commit()
                            except Exception:
                                logger.warning(
                                    "批次審核降級時 commit stale 標記失敗",
                                    exc_info=True,
                                )
                                session.rollback()
                            failed.append(
                                {
                                    "id": leave_id,
                                    "reason": (
                                        "假單已審核但薪資重算失敗，已標 stale，"
                                        "請手動前往薪資頁面重新計算"
                                    ),
                                }
                            )

                    if not recalc_failed:
                        # 修補 2026-05-11 P2-12：通知挪到 recalc 成功後才收集，
                        # 避免「重算失敗但員工已收到核准通知」與 DB 矛盾的場景。
                        # 通知統一在 loop 結束後 enqueue+commit，避免 stale-mark
                        # 中途 commit 提前觸發 after_commit fan-out。
                        _owner_user = _user_map.get(leave.employee_id)
                        if _owner_user is not None:
                            _pending_notifs.append(
                                dict(
                                    event_type=(
                                        "leave.approved"
                                        if data.approved
                                        else "leave.rejected"
                                    ),
                                    recipient_user_id=_owner_user.id,
                                    context={
                                        "reviewer_name": current_user.get("name")
                                        or current_user.get("username", ""),
                                        "leave_type": leave.leave_type,
                                        "start": (
                                            leave.start_date.isoformat()
                                            if hasattr(leave.start_date, "isoformat")
                                            else str(leave.start_date)
                                        ),
                                        "end": (
                                            leave.end_date.isoformat()
                                            if hasattr(leave.end_date, "isoformat")
                                            else str(leave.end_date)
                                        ),
                                        "leave_id": leave_id,
                                        "rejection_reason": (
                                            data.rejection_reason
                                            if not data.approved
                                            else None
                                        ),
                                    },
                                    sender_id=current_user.get("user_id"),
                                    source_entity_type="leave_request",
                                    source_entity_id=leave_id,
                                )
                            )
                        succeeded.append(leave_id)

                # 統一 enqueue 並 commit，讓 after_commit hook 觸發 fan-out
                # lazy import 避免 circular（同 single approval 路徑）
                if _pending_notifs:
                    from services.notification import dispatch

                    for _notif in _pending_notifs:
                        dispatch.enqueue(session=session, **_notif)
                    try:
                        session.commit()
                    except Exception:
                        logger.warning(
                            "批次審核通知 dispatch commit 失敗", exc_info=True
                        )
                        session.rollback()
            except Exception as e:
                session.rollback()
                reason = safe_batch_reason(
                    e,
                    context="批次請假統一提交",
                    fallback="統一提交失敗，請稍後重試或聯絡管理員",
                )
                for leave_id, *_ in changes:
                    failed.append({"id": leave_id, "reason": reason})

        # AuditLog changes：批次操作彙整為單筆 audit，列出受影響的 leave_id 與每筆 ApprovalLog 連結
        request.state.audit_summary = (
            f"批次{'核准' if data.approved else '駁回'}請假 "
            f"成功 {len(succeeded)}/失敗 {len(failed)}（請求 {len(data.ids)} 筆）"
        )
        request.state.audit_changes = {
            "action": "leave_batch_approve",
            "decision": "approved" if data.approved else "rejected",
            "rejection_reason": data.rejection_reason,
            "requested_ids": data.ids,
            "succeeded_ids": list(succeeded),
            "failed": failed,
            "approval_log_ids": [
                {
                    "leave_id": _lid,
                    "employee_id": _lv.employee_id,
                    "is_reject_of_approved": _ir,
                    "approval_log_id": _alog,
                }
                for _lid, _lv, _ac, _ir, _alog in changes
            ],
            "high_risk_count": sum(1 for _lid, _lv, _ac, _ir, _alog in changes if _ir),
        }
    finally:
        session.close()

    return {"succeeded": succeeded, "failed": failed}


from utils.excel_writer import write_header_row as _lv_write_header  # noqa: E402


@router.get("/leaves/import-template")
def get_leave_import_template(
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """下載請假批次匯入 Excel 範本"""
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "請假匯入範本"

    headers = [
        "員工編號",
        "員工姓名",
        "假別代碼",
        "開始日期",
        "結束日期",
        "時數(可空)",
        "原因(可空)",
    ]
    _lv_write_header(ws, 1, headers)

    ws.cell(row=2, column=1, value="E001")
    ws.cell(row=2, column=2, value="王小明")
    ws.cell(row=2, column=3, value="annual")
    ws.cell(row=2, column=4, value="2026-03-15")
    ws.cell(row=2, column=5, value="2026-03-15")
    ws.cell(row=2, column=6, value=8)
    ws.cell(row=2, column=7, value="年度特休")

    ws2 = SafeWorksheet(wb.create_sheet("假別代碼說明"))
    ws2.cell(row=1, column=1, value="假別代碼")
    ws2.cell(row=1, column=2, value="中文名稱")
    for idx, (code, label) in enumerate(LEAVE_TYPE_LABELS.items(), 2):
        ws2.cell(row=idx, column=1, value=code)
        ws2.cell(row=idx, column=2, value=label)

    return xlsx_streaming_response(wb, "請假匯入範本.xlsx")


class LeaveImportRow(ExcelImportSchema):
    """請假匯入單列 schema（excel header 用中文 alias）。

    僅負責「Excel 列 → 結構化資料」轉換；日期/時數/假別代碼的業務驗證
    仍留在 endpoint（pandas 之前怎麼處理、現在保留相同行為）。
    """

    # employee_code 與 employee_name 兩者擇一即可由業務層 resolve_employee_from_row 判斷
    employee_code: Optional[str] = Field(default=None, alias="員工編號")
    employee_name: Optional[str] = Field(default=None, alias="員工姓名")
    leave_type: Optional[str] = Field(default=None, alias="假別代碼")
    start_date: Any = Field(default=None, alias="開始日期")
    end_date: Any = Field(default=None, alias="結束日期")
    leave_hours: Any = Field(default=None, alias="時數(可空)")
    reason: Any = Field(default=None, alias="原因(可空)")


@router.post("/leaves/import", response_model=LeaveImportResultOut)
async def import_leaves(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_WRITE)),
):
    """批次匯入請假申請（建立草稿假單，status='pending'，需後續人工審核）"""
    content = await read_upload_with_size_check(file)
    # parse + DB 迴圈為同步 CPU/IO，卸載到 executor 避免阻塞 event loop（行為不變）
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _import_leaves_sync, content)


def _import_leaves_sync(content: bytes) -> dict:
    validate_file_signature(content, ".xlsx")

    parse_result = parse_excel(BytesIO(content), schema=LeaveImportRow)
    # 整檔級錯誤（INVALID_FILE / EMPTY_FILE / MISSING_COLUMN）→ 400，與既有行為一致
    file_level_codes = {"INVALID_FILE", "EMPTY_FILE", "MISSING_COLUMN"}
    file_level_errors = [
        e for e in parse_result.errors if e["error_code"] in file_level_codes
    ]
    if file_level_errors:
        # 取第一筆訊息維持訊息可讀性
        msg = file_level_errors[0]["message"]
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{msg}")

    label_to_code = {v: k for k, v in LEAVE_TYPE_LABELS.items()}

    results: dict = {"total": 0, "created": 0, "failed": 0, "errors": []}
    # 將 parse_excel 的 row-level errors（型別/缺欄）先映射回舊格式 "第 N 行: msg"
    for err in parse_result.errors:
        if err["error_code"] in file_level_codes:
            continue
        results["total"] += 1
        results["failed"] += 1
        results["errors"].append(f"第 {err['row']} 行: {err['message']}")

    session = get_session()
    try:
        emp_by_id, emp_by_name = build_employee_lookup(session)

        for row_idx, row in enumerate(parse_result.rows, start=2):
            results["total"] += 1
            row_num = row_idx
            try:
                # resolve_employee_from_row 接受 dict-like / pandas Series；
                # pydantic BaseModel 用 model_dump() 取得 dict 並補上原中文 keys。
                row_dict = {
                    "員工編號": row.employee_code,
                    "員工姓名": row.employee_name,
                }
                emp = resolve_employee_from_row(row_dict, emp_by_id, emp_by_name)

                leave_type_raw = (
                    str(row.leave_type).strip() if row.leave_type is not None else ""
                )
                if leave_type_raw in LEAVE_TYPE_LABELS:
                    leave_type = leave_type_raw
                elif leave_type_raw in label_to_code:
                    leave_type = label_to_code[leave_type_raw]
                else:
                    raise ValueError(
                        f"無效的假別代碼：{leave_type_raw}（請參考「假別代碼說明」頁）"
                    )

                start_raw = row.start_date
                end_raw = row.end_date
                if start_raw is None or end_raw is None:
                    raise ValueError("開始日期或結束日期不得為空")
                try:
                    # pd.to_datetime 對 str / datetime / Timestamp 均能 normalize，
                    # 維持與舊 endpoint 完全相同的日期解析行為
                    start_date = pd.to_datetime(start_raw).date()
                    end_date = pd.to_datetime(end_raw).date()
                except Exception:
                    raise ValueError("日期格式錯誤，建議使用 YYYY-MM-DD")

                if end_date < start_date:
                    raise ValueError("結束日期不得早於開始日期")
                if (
                    start_date.year != end_date.year
                    or start_date.month != end_date.month
                ):
                    raise ValueError("請假區間不可跨月，請拆成多筆分別匯入")

                hours_raw = row.leave_hours
                if hours_raw is None:
                    leave_hours = 8.0
                else:
                    try:
                        leave_hours = float(hours_raw)
                    except (ValueError, TypeError):
                        raise ValueError("時數格式錯誤，必須為數字（如 4、4.5、8）")
                    # 與管理端/portal 建立假單一致：最小 0.5、最大 480、0.5 倍數。
                    # 不可在驗證失敗時 fallback 成 8h（會把非法時數靜默變全日扣薪假）。
                    leave_hours = validate_leave_hours_value(leave_hours)

                reason_raw = row.reason
                reason = str(reason_raw).strip() if reason_raw is not None else None

                # 與管理端 / portal 建立假單相同：時數不得超過該區間排班工時，
                # 且不得超過假別年度配額（含待審）。匯入跳過會讓單日 100h 這種
                # 超額假單進入待審，主管核准後直接依超額時數扣薪。
                validate_leave_hours_against_schedule(
                    session,
                    emp.id,
                    start_date,
                    end_date,
                    leave_hours,
                )
                # 修補 2026-05-11 P1-4：原本不查 overlap，可一次 import 兩筆同員工
                # 同日 pending；逐筆主管核准時看到的 approved 集合都不包含對方
                # （_check_overlap include_pending=False），兩張都通過進薪資扣款。
                # 用 include_pending=True 涵蓋 pending 與 approved。
                conflict = _find_overlapping_leave(
                    session,
                    emp.id,
                    start_date,
                    end_date,
                    include_pending=True,
                )
                if conflict:
                    raise ValueError(
                        f"與既有假單 #{conflict.id}"
                        f"（{conflict.start_date}~{conflict.end_date}）重疊"
                    )
                _check_leave_limits(
                    session,
                    emp.id,
                    leave_type,
                    start_date,
                    leave_hours,
                )
                # 補休配額由加班核准動態累積；_check_quota 對非 QUOTA_LEAVE_TYPES
                # 直接 return，會放過超額補休（HR 可用 Excel 大量匯入繞過）。
                if leave_type == "compensatory":
                    _check_compensatory_quota(
                        session,
                        emp.id,
                        start_date.year,
                        leave_hours,
                    )
                else:
                    _check_quota(
                        session,
                        emp.id,
                        leave_type,
                        start_date.year,
                        leave_hours,
                    )

                effective_ratio = LEAVE_DEDUCTION_RULES[leave_type]
                leave = LeaveRecord(
                    employee_id=emp.id,
                    leave_type=leave_type,
                    start_date=start_date,
                    end_date=end_date,
                    leave_hours=leave_hours,
                    is_deductible=effective_ratio > 0,
                    deduction_ratio=effective_ratio,
                    reason=reason,
                    status=ApprovalStatus.PENDING.value,
                )
                session.add(leave)
                session.flush()
                results["created"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"第 {row_num} 行: {str(e)}")

        session.commit()
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e, context="匯入失敗")
    finally:
        session.close()

    return results


@router.get("/leaves/{leave_id}/attachments/{filename}")
def get_leave_attachment(
    leave_id: int,
    filename: str,
    current_user: dict = Depends(require_staff_permission(Permission.LEAVES_READ)),
):
    """取得假單附件（管理後台）。

    backend 為 local：直接 stream bytes（既有行為）
    backend 為 supabase：302 redirect 到 signed URL（TTL 預設 1 小時）
    """
    from fastapi.responses import RedirectResponse, Response as _Response
    from utils.storage import LocalStorage, get_backend

    session = get_session()
    try:
        leave = session.query(LeaveRecord).filter(LeaveRecord.id == leave_id).first()
        if not leave:
            raise HTTPException(status_code=404, detail="找不到請假記錄")

        paths = _parse_paths(leave.attachment_paths)
        if filename not in paths:
            raise HTTPException(status_code=404, detail="找不到附件")

        backend = get_backend()
        key = f"{leave_id}/{filename}"
        if not backend.exists(_UPLOAD_MODULE, key):
            raise HTTPException(status_code=404, detail="檔案不存在")

        if isinstance(backend, LocalStorage):
            data = backend.read(_UPLOAD_MODULE, key)
            return _Response(content=data, media_type="application/octet-stream")

        ttl = settings.storage.supabase_signed_url_ttl
        url = backend.signed_url(_UPLOAD_MODULE, key, ttl)
        return RedirectResponse(url, status_code=302)
    finally:
        session.close()
