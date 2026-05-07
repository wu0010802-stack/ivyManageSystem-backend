"""
api/activity/stats.py — 統計儀表板端點（5 個）
"""

from typing import Optional

from fastapi import APIRouter, Depends, Query

from models.database import get_session
from utils.academic import resolve_academic_term_filters
from utils.auth import require_staff_permission
from utils.excel_utils import SafeWorksheet
from utils.permissions import Permission
from services.activity_service import activity_service

router = APIRouter()


@router.get("/stats")
async def get_stats(
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得儀表板統計資料（相容舊版：summary + charts）。"""
    session = get_session()
    try:
        return activity_service.get_stats(session)
    finally:
        session.close()


@router.get("/stats-summary")
async def get_stats_summary(
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得儀表板摘要統計資料。"""
    session = get_session()
    try:
        return activity_service.get_stats_summary(session)
    finally:
        session.close()


@router.get("/stats-charts")
async def get_stats_charts(
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得儀表板圖表資料。"""
    session = get_session()
    try:
        return activity_service.get_stats_charts(session)
    finally:
        session.close()


@router.get("/dashboard-table")
async def get_dashboard_table(
    school_year: Optional[int] = Query(None, ge=100, le=200),
    semester: Optional[int] = Query(None, ge=1, le=2),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得儀表板統計表格資料。school_year/semester 不傳時使用當前學期。"""
    resolved_year, resolved_semester = resolve_academic_term_filters(
        school_year, semester
    )
    session = get_session()
    try:
        return activity_service.get_dashboard_table(
            session, school_year=resolved_year, semester=resolved_semester
        )
    finally:
        session.close()


@router.get("/dashboard-table/export")
def export_dashboard_table(
    school_year: Optional[int] = Query(None, ge=100, le=200),
    semester: Optional[int] = Query(None, ge=1, le=2),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """匯出統計表為 Excel（多 sheet）"""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from ._shared import TAIPEI_TZ

    resolved_year, resolved_semester = resolve_academic_term_filters(
        school_year, semester
    )
    session = get_session()
    try:
        data = activity_service.get_dashboard_table(
            session, school_year=resolved_year, semester=resolved_semester
        )

        wb = openpyxl.Workbook()

        # --- Sheet 1: 總覽 ---
        ws_overview = SafeWorksheet(wb.active)
        ws_overview.title = "總覽"
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
        )

        ws_overview.append(["項目", "數值"])
        ws_overview.cell(1, 1).font = header_font
        ws_overview.cell(1, 2).font = header_font

        grand = data.get("grand_total", {})
        ws_overview.append(["全園在籍學生", grand.get("student_count", 0)])
        ws_overview.append(["全園報名人次", grand.get("total_enrollments", 0)])
        ws_overview.append(["全園報名比率", f"{grand.get('ratio', 0)}%"])

        for course in data.get("courses", []):
            ws_overview.append(
                [
                    f"課程「{course['name']}」報名",
                    grand.get("courses", {}).get(str(course["id"]), 0),
                ]
            )

        for col in ws_overview.columns:
            ws_overview.column_dimensions[col[0].column_letter].width = 30

        # --- Sheet 2: 班級統計 ---
        ws_detail = SafeWorksheet(wb.create_sheet("班級統計"))
        course_list = data.get("courses", [])
        detail_headers = (
            ["年級", "班級", "班導師", "在籍人數"]
            + [c["name"] for c in course_list]
            + ["總報名", "比率(%)"]
        )
        ws_detail.append(detail_headers)
        for cell in ws_detail[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for grade in data.get("grades", []):
            for cls in grade.get("classrooms", []):
                row = [
                    grade["grade_name"],
                    cls["classroom_name"],
                    cls["teacher_name"],
                    cls["student_count"],
                ]
                for c in course_list:
                    row.append(cls["courses"].get(str(c["id"]), 0))
                row += [cls["total_enrollments"], cls["ratio"]]
                ws_detail.append(row)

            sub = grade.get("subtotal", {})
            sub_row = [
                f"【{grade['grade_name']}小計】",
                "",
                "",
                sub.get("student_count", 0),
            ]
            for c in course_list:
                sub_row.append(sub.get("courses", {}).get(str(c["id"]), 0))
            sub_row += [sub.get("total_enrollments", 0), sub.get("ratio", 0)]
            ws_detail.append(sub_row)
            for cell in ws_detail[ws_detail.max_row]:
                cell.font = Font(bold=True)

        for col in ws_detail.columns:
            ws_detail.column_dimensions[col[0].column_letter].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"activity_dashboard_{datetime.now(TAIPEI_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()
