"""finance_report_service 與 /api/reports/finance-summary 端點測試。"""

import os
import sys
from datetime import date, datetime
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import router as auth_router, _account_failures, _ip_attempts
from api.reports import router as reports_router
from models.database import (
    Base,
    Employee,
    User,
    ActivityPaymentRecord,
    SalaryRecord,
)
from models.fees import StudentFeePayment, StudentFeeRecord, StudentFeeRefund
from models.vendor_payment import VendorPayment
from services import finance_report_service as svc
from sqlalchemy import event as sa_event
from utils.auth import hash_password


def _register_auto_mirror_payment(session_factory):
    """測試 fixture 輔助：StudentFeeRecord.amount_paid > 0 時自動 mirror
    一筆 StudentFeePayment，讓舊測試資料配合新的 append-only 流水邏輯。

    Why: 實作 P1-B 後，財務月報改讀 StudentFeePayment。既有測試直接建
    StudentFeeRecord(amount_paid=X, status='paid', payment_date=D) 的 fixture
    若不同步補 payment 流水，聚合結果會變 0 並失去覆蓋意義。
    """

    @sa_event.listens_for(session_factory, "after_flush")
    def _mirror(session, flush_context):
        for obj in list(session.new):
            if not isinstance(obj, StudentFeeRecord):
                continue
            if (obj.amount_paid or 0) <= 0 or not obj.payment_date:
                continue
            # 避免同次 flush 已手動建 payment 時重複寫入
            has_existing = any(
                isinstance(o, StudentFeePayment) and o.record_id == obj.id
                for o in list(session.new)
            )
            if has_existing:
                continue
            session.add(
                StudentFeePayment(
                    record_id=obj.id,
                    amount=obj.amount_paid,
                    payment_date=obj.payment_date,
                    payment_method=obj.payment_method or "現金",
                    notes="（測試 auto-mirror）",
                    operator="test",
                )
            )


@pytest.fixture
def fin_client(tmp_path):
    db_path = tmp_path / "fin.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}", connect_args={"check_same_thread": False}
    )
    session_factory = sessionmaker(bind=engine)
    _register_auto_mirror_payment(session_factory)
    old_e, old_sf = base_module._engine, base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory
    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()
    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(reports_router)
    with TestClient(app) as client:
        yield client, session_factory
    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_e
    base_module._SessionFactory = old_sf
    engine.dispose()


def _seed_admin(sf, client):
    with sf() as s:
        s.add(
            User(
                username="fin_admin",
                password_hash=hash_password("FinPass123"),
                role="admin",
                permission_names=["*"],
                is_active=True,
                must_change_password=False,
            )
        )
        s.commit()
    r = client.post(
        "/api/auth/login", json={"username": "fin_admin", "password": "FinPass123"}
    )
    assert r.status_code == 200


# ── Provider 單元測試 ──────────────────────────────────────────────────────


class TestTuitionProvider:
    def test_only_paid_counted(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add_all(
                [
                    StudentFeeRecord(
                        student_id=1,
                        period="2026-1",
                        amount_due=5000,
                        amount_paid=5000,
                        status="paid",
                        payment_date=date(2026, 3, 10),
                        student_name="A",
                        fee_item_name="月費",
                    ),
                    StudentFeeRecord(
                        student_id=2,
                        period="2026-1",
                        amount_due=5000,
                        amount_paid=0,
                        status="unpaid",
                        payment_date=None,
                        student_name="B",
                        fee_item_name="月費",
                    ),
                ]
            )
            s.commit()
        with sf() as s:
            out = svc.get_tuition_revenue_by_month(s, 2026)
        assert out == {3: 5000}

    def test_refund_by_month(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add(
                StudentFeeRefund(
                    record_id=1,
                    amount=1000,
                    reason="退學",
                    refunded_at=datetime(2026, 4, 5, 10, 0),
                    refunded_by="admin",
                    idempotency_key="k1",
                )
            )
            s.commit()
        with sf() as s:
            out = svc.get_tuition_refund_by_month(s, 2026)
        assert out == {4: 1000}


class TestActivityProvider:
    def test_payment_and_refund_split(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add_all(
                [
                    ActivityPaymentRecord(
                        registration_id=1,
                        type="payment",
                        amount=2000,
                        payment_date=date(2026, 3, 10),
                    ),
                    ActivityPaymentRecord(
                        registration_id=1,
                        type="refund",
                        amount=500,
                        payment_date=date(2026, 4, 2),
                    ),
                ]
            )
            s.commit()
        with sf() as s:
            rev = svc.get_activity_revenue_by_month(s, 2026)
            ref = svc.get_activity_refund_by_month(s, 2026)
        assert rev == {3: 2000}
        assert ref == {4: 500}


class TestSalaryExpense:
    def test_gross_plus_employer_split(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            emp = Employee(
                employee_id="E1",
                name="T",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            s.add(
                SalaryRecord(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=3,
                    gross_salary=30000,
                    labor_insurance_employer=2500,
                    health_insurance_employer=1400,
                    pension_employer=1800,
                    net_salary=25000,
                    total_deduction=5000,
                    is_finalized=True,
                )
            )
            s.commit()
        with sf() as s:
            out = svc.get_salary_expense_by_month(s, 2026)
        assert out == {3: {"employee_gross": 30000, "employer_benefit": 5700}}

    def test_unused_leave_payout_included_in_expense(self, fin_client):
        """特休未休折現（unused_leave_payout）不進 gross_salary，但屬園方實際現金流出，
        get_salary_expense_by_month 的 employee_gross 必須加入，否則離職月薪資支出低估。"""
        _, sf = fin_client
        with sf() as s:
            emp = Employee(
                employee_id="E_ULP1",
                name="特休測試員",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            s.add(
                SalaryRecord(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=6,
                    gross_salary=30000,
                    unused_leave_payout=5000,  # 特休未休折現，獨立欄位不進 gross
                    labor_insurance_employer=2500,
                    health_insurance_employer=1400,
                    pension_employer=1800,
                    net_salary=25000,
                    total_deduction=5000,
                    is_finalized=True,
                )
            )
            s.commit()
        with sf() as s:
            out = svc.get_salary_expense_by_month(s, 2026)
        # employee_gross 應含 unused_leave_payout(5000)：30000 + 5000 = 35000
        assert (
            out[6]["employee_gross"] == 35000
        ), f"employee_gross 應含 unused_leave_payout 5000，實際={out[6]['employee_gross']}"

    def test_unused_leave_payout_included_in_detail_real_cost(self, fin_client):
        """get_salary_detail 的 real_cost 同樣必須含 unused_leave_payout，
        否則明細加總與月摘要不一致。"""
        _, sf = fin_client
        with sf() as s:
            emp = Employee(
                employee_id="E_ULP2",
                name="特休明細員",
                base_salary=28000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            s.add(
                SalaryRecord(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=7,
                    gross_salary=28000,
                    unused_leave_payout=3000,  # 特休折現
                    labor_insurance_employer=2500,
                    health_insurance_employer=1400,
                    pension_employer=1800,
                    net_salary=23000,
                    total_deduction=5000,
                    is_finalized=True,
                )
            )
            s.commit()
        with sf() as s:
            rows = svc.get_salary_detail(s, 2026, 7)
        assert len(rows) == 1
        # real_cost = gross(28000) + unused_leave_payout(3000) + employer(5700) = 36700
        assert (
            rows[0]["real_cost"] == 36700
        ), f"real_cost 應含 unused_leave_payout 3000，實際={rows[0]['real_cost']}"


class TestVendorPaymentExpense:
    def test_aggregated_by_month_regardless_of_status(self, fin_client):
        """pending 與 signed 都計入支出總額（園所現金已付出）。"""
        _, sf = fin_client
        with sf() as s:
            s.add_all(
                [
                    VendorPayment(
                        payment_date=date(2026, 3, 5),
                        vendor_name="A 清潔",
                        amount=1200,
                        payment_method="cash",
                        status="signed",
                    ),
                    VendorPayment(
                        payment_date=date(2026, 3, 20),
                        vendor_name="B 教具",
                        amount=800,
                        payment_method="bank_transfer",
                        status="pending",
                    ),
                    VendorPayment(
                        payment_date=date(2026, 4, 1),
                        vendor_name="C 食材",
                        amount=5000,
                        payment_method="check",
                        status="signed",
                    ),
                ]
            )
            s.commit()
        with sf() as s:
            out = svc.get_vendor_payment_expense_by_month(s, 2026)
        assert out == {3: 2000, 4: 5000}

    def test_vendor_expense_aggregates_into_summary(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add(
                VendorPayment(
                    payment_date=date(2026, 5, 10),
                    vendor_name="Z 公司",
                    amount=3000,
                    payment_method="cash",
                    status="pending",
                )
            )
            s.commit()
        with sf() as s:
            data = svc.build_finance_summary(s, 2026, month=5)
        row = data["monthly_trend"][0]
        assert row["expense"] == 3000
        vendor_cat = next(
            c for c in data["expense_by_category"] if c["category"] == "vendor_payment"
        )
        assert vendor_cat["amount"] == 3000
        assert vendor_cat["label"] == "廠商付款"

    def test_vendor_detail_in_drilldown(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add(
                VendorPayment(
                    payment_date=date(2026, 5, 15),
                    vendor_name="D 紙箱行",
                    amount=750,
                    payment_method="cash",
                    description="搬家用紙箱",
                    invoice_number="X-9",
                    status="signed",
                )
            )
            s.commit()
        with sf() as s:
            detail = svc.build_finance_detail(s, 2026, 5)
        rows = detail["vendor_payment"]
        assert len(rows) == 1
        assert rows[0]["vendor_name"] == "D 紙箱行"
        assert rows[0]["amount"] == 750
        assert rows[0]["status"] == "signed"


# ── Aggregator ─────────────────────────────────────────────────────────────


class TestAggregator:
    def test_full_year_contains_12_months(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            data = svc.build_finance_summary(s, 2026)
        assert len(data["monthly_trend"]) == 12
        assert data["period"] == {"year": 2026, "month": None}
        assert data["summary"]["total_revenue"] == 0
        assert data["summary"]["net_cashflow"] == 0

    def test_single_month_contains_1_row(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            data = svc.build_finance_summary(s, 2026, month=3)
        assert len(data["monthly_trend"]) == 1
        assert data["monthly_trend"][0]["month"] == 3

    def test_net_cashflow_accounts_refund(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add_all(
                [
                    StudentFeeRecord(
                        student_id=1,
                        period="2026-1",
                        amount_due=10000,
                        amount_paid=10000,
                        status="paid",
                        payment_date=date(2026, 3, 10),
                        student_name="A",
                        fee_item_name="月費",
                    ),
                    StudentFeeRefund(
                        record_id=1,
                        amount=2000,
                        reason="x",
                        refunded_at=datetime(2026, 3, 15, 9, 0),
                        refunded_by="admin",
                        idempotency_key="r1",
                    ),
                ]
            )
            s.commit()
        with sf() as s:
            data = svc.build_finance_summary(s, 2026, month=3)
        row = data["monthly_trend"][0]
        assert row["revenue"] == 10000
        assert row["refund"] == 2000
        assert row["net"] == 8000

    def test_categories_have_expected_shape(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            data = svc.build_finance_summary(s, 2026)
        rev_cats = {c["category"] for c in data["revenue_by_category"]}
        exp_cats = {c["category"] for c in data["expense_by_category"]}
        assert rev_cats == {"tuition", "activity"}
        assert exp_cats == {"salary_gross", "employer_benefit", "vendor_payment"}

    def test_cross_year_isolation(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add(
                StudentFeeRecord(
                    student_id=1,
                    period="2025-2",
                    amount_due=5000,
                    amount_paid=5000,
                    status="paid",
                    payment_date=date(2025, 12, 10),
                    student_name="A",
                    fee_item_name="月費",
                )
            )
            s.commit()
        with sf() as s:
            assert svc.get_tuition_revenue_by_month(s, 2026) == {}
            assert svc.get_tuition_revenue_by_month(s, 2025) == {12: 5000}


# ── 端點整合測試 ───────────────────────────────────────────────────────────


class TestFinanceSummaryEndpoint:
    def test_year_only(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary?year=2026")
        assert r.status_code == 200
        body = r.json()
        assert body["period"]["year"] == 2026
        assert body["period"]["month"] is None
        assert len(body["monthly_trend"]) == 12
        assert "revenue_by_category" in body
        assert "expense_by_category" in body

    def test_with_month(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary?year=2026&month=3")
        assert r.status_code == 200
        body = r.json()
        assert body["period"]["month"] == 3
        assert len(body["monthly_trend"]) == 1
        assert body["monthly_trend"][0]["month"] == 3

    def test_requires_auth(self, fin_client):
        client, _ = fin_client
        r = client.get("/api/reports/finance-summary?year=2026")
        assert r.status_code in (401, 403)

    def test_year_out_of_range_rejected(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary?year=1999")
        assert r.status_code == 422

    def test_with_actual_data(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        with sf() as s:
            emp = Employee(
                employee_id="E1",
                name="T",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            s.add_all(
                [
                    StudentFeeRecord(
                        student_id=1,
                        period="2026-1",
                        amount_due=5000,
                        amount_paid=5000,
                        status="paid",
                        payment_date=date(2026, 4, 5),
                        student_name="A",
                        fee_item_name="月費",
                    ),
                    ActivityPaymentRecord(
                        registration_id=1,
                        type="payment",
                        amount=1200,
                        payment_date=date(2026, 4, 6),
                    ),
                    SalaryRecord(
                        employee_id=emp.id,
                        salary_year=2026,
                        salary_month=4,
                        gross_salary=30000,
                        labor_insurance_employer=2500,
                        health_insurance_employer=1400,
                        pension_employer=1800,
                        net_salary=25000,
                        total_deduction=5000,
                        is_finalized=True,
                    ),
                ]
            )
            s.commit()
        r = client.get("/api/reports/finance-summary?year=2026&month=4")
        assert r.status_code == 200
        body = r.json()
        row = body["monthly_trend"][0]
        assert row["revenue"] == 6200  # 5000 + 1200
        assert row["refund"] == 0
        assert row["expense"] == 35700  # 30000 + 5700
        assert row["net"] == 6200 - 35700


# ── 快取行為 ───────────────────────────────────────────────────────────────


class TestSummaryCaching:
    def test_second_call_serves_from_cache(self, fin_client):
        """第二次呼叫應從 ReportSnapshot 讀，不重跑 builder。"""
        client, sf = fin_client
        _seed_admin(sf, client)
        r1 = client.get("/api/reports/finance-summary?year=2026").json()
        # 在快取期間新增一筆繳費記錄
        with sf() as s:
            s.add(
                StudentFeeRecord(
                    student_id=1,
                    period="2026-1",
                    amount_due=9999,
                    amount_paid=9999,
                    status="paid",
                    payment_date=date(2026, 3, 1),
                    student_name="Z",
                    fee_item_name="月費",
                )
            )
            s.commit()
        r2 = client.get("/api/reports/finance-summary?year=2026").json()
        # cache 命中 → 金額與第一次一致（未看到新資料）
        assert r1["summary"]["total_revenue"] == r2["summary"]["total_revenue"]

    def test_different_params_have_independent_cache(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r1 = client.get("/api/reports/finance-summary?year=2026").json()
        r2 = client.get("/api/reports/finance-summary?year=2026&month=3").json()
        assert len(r1["monthly_trend"]) == 12
        assert len(r2["monthly_trend"]) == 1


# ── Detail 端點 ────────────────────────────────────────────────────────────


class TestDetailEndpoint:
    def test_empty(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary/detail?year=2026&month=3")
        assert r.status_code == 200
        body = r.json()
        assert body["tuition"] == []
        assert body["activity"] == []
        assert body["salary"] == []
        assert body["period"] == {"year": 2026, "month": 3}

    def test_includes_tuition_activity_salary(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        with sf() as s:
            emp = Employee(
                employee_id="E1",
                name="Teacher",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            s.add_all(
                [
                    StudentFeeRecord(
                        student_id=1,
                        period="2026-1",
                        amount_due=5000,
                        amount_paid=5000,
                        status="paid",
                        payment_date=date(2026, 3, 5),
                        student_name="Alice",
                        classroom_name="大班",
                        fee_item_name="月費",
                        payment_method="現金",
                    ),
                    StudentFeeRefund(
                        record_id=1,
                        amount=500,
                        reason="請假退款",
                        refunded_at=datetime(2026, 3, 10, 10, 0),
                        refunded_by="admin",
                        idempotency_key="refund1",
                    ),
                    ActivityPaymentRecord(
                        registration_id=1,
                        type="payment",
                        amount=1500,
                        payment_date=date(2026, 3, 6),
                        payment_method="轉帳",
                        operator="staff",
                    ),
                    SalaryRecord(
                        employee_id=emp.id,
                        salary_year=2026,
                        salary_month=3,
                        gross_salary=30000,
                        labor_insurance_employer=2500,
                        health_insurance_employer=1400,
                        pension_employer=1800,
                        net_salary=25000,
                        total_deduction=5000,
                        is_finalized=True,
                    ),
                ]
            )
            s.commit()
        r = client.get("/api/reports/finance-summary/detail?year=2026&month=3").json()
        assert len(r["tuition"]) == 2  # 1 繳費 + 1 退款
        assert {x["kind"] for x in r["tuition"]} == {"payment", "refund"}
        assert len(r["activity"]) == 1
        assert r["activity"][0]["kind"] == "payment"
        assert len(r["salary"]) == 1
        salary_row = r["salary"][0]
        assert salary_row["employee_name"] == "Teacher"
        assert salary_row["gross_salary"] == 30000
        assert salary_row["employer_benefit"] == 5700
        assert salary_row["real_cost"] == 35700


# ── Export 端點 ────────────────────────────────────────────────────────────


class TestExportEndpoint:
    def test_year_only_returns_xlsx(self, fin_client):
        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary/export?year=2026")
        assert r.status_code == 200
        assert (
            r.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert r.content[:2] == b"PK"  # xlsx 是 zip，以 PK 開頭

    def test_with_month_adds_detail_sheets(self, fin_client):
        from io import BytesIO
        from openpyxl import load_workbook

        client, sf = fin_client
        _seed_admin(sf, client)
        r = client.get("/api/reports/finance-summary/export?year=2026&month=3")
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert "月度彙總" in wb.sheetnames
        assert "分類統計" in wb.sheetnames
        assert "學費明細" in wb.sheetnames
        assert "才藝明細" in wb.sheetnames
        assert "薪資明細" in wb.sheetnames


# ── 薪資只認封存且非 stale（actual expenditure）─────────────────────────────
#
# Bug: 四個薪資 provider（expense / breakdown / breakdown_with_role / detail）
# 原本只篩 salary_year/month，把草稿（is_finalized=False）與 stale
# （needs_recalc=True）薪資當實際支出。對齊 api/reports.py _query_salary_monthly
# 的「只認 is_finalized=True AND needs_recalc=False」。草稿/待重算是中間態，
# 讓會計用測試重算的草稿影響財務總覽/月損益/匯出形同 A 錢空間。
class TestSalaryFinalizedOnly:
    def _seed_three_states(self, sf):
        """5 月三筆薪資：封存非 stale（計入）/ 草稿（排除）/ 封存但 stale（排除）。"""
        with sf() as s:
            emps = []
            for i in range(3):
                e = Employee(
                    employee_id=f"SF{i}",
                    name=f"SF{i}",
                    base_salary=30000,
                    employee_type="regular",
                    is_active=True,
                )
                s.add(e)
                emps.append(e)
            s.flush()
            # 0: 封存非 stale → 唯一計入
            s.add(
                SalaryRecord(
                    employee_id=emps[0].id,
                    salary_year=2026,
                    salary_month=5,
                    gross_salary=30000,
                    festival_bonus=1000,
                    overtime_bonus=500,
                    labor_insurance_employer=2500,
                    health_insurance_employer=1400,
                    pension_employer=1800,
                    net_salary=25000,
                    total_deduction=5000,
                    is_finalized=True,
                    needs_recalc=False,
                )
            )
            # 1: 草稿（未封存）→ 排除
            s.add(
                SalaryRecord(
                    employee_id=emps[1].id,
                    salary_year=2026,
                    salary_month=5,
                    gross_salary=99999,
                    festival_bonus=99999,
                    overtime_bonus=99999,
                    labor_insurance_employer=99999,
                    health_insurance_employer=99999,
                    pension_employer=99999,
                    net_salary=99999,
                    total_deduction=0,
                    is_finalized=False,
                    needs_recalc=False,
                )
            )
            # 2: 封存但 needs_recalc=True（異常 stale）→ 排除
            s.add(
                SalaryRecord(
                    employee_id=emps[2].id,
                    salary_year=2026,
                    salary_month=5,
                    gross_salary=88888,
                    festival_bonus=88888,
                    overtime_bonus=88888,
                    labor_insurance_employer=88888,
                    health_insurance_employer=88888,
                    pension_employer=88888,
                    net_salary=88888,
                    total_deduction=0,
                    is_finalized=True,
                    needs_recalc=True,
                )
            )
            s.commit()

    def test_expense_excludes_draft_and_stale(self, fin_client):
        _, sf = fin_client
        self._seed_three_states(sf)
        with sf() as s:
            out = svc.get_salary_expense_by_month(s, 2026)
        # employee_gross = 30000 + 1000 + 500；employer_benefit = 2500+1400+1800
        assert out == {5: {"employee_gross": 31500, "employer_benefit": 5700}}

    def test_breakdown_excludes_draft_and_stale(self, fin_client):
        _, sf = fin_client
        self._seed_three_states(sf)
        with sf() as s:
            out = svc.get_salary_breakdown_by_month(s, 2026)
        assert out[5]["gross_salary"] == 30000
        assert out[5]["festival_bonus"] == 1000

    def test_breakdown_with_role_excludes_draft_and_stale(self, fin_client):
        _, sf = fin_client
        self._seed_three_states(sf)
        with sf() as s:
            out = svc.get_salary_breakdown_by_month_with_role(s, 2026)
        assert out[5]["regular"]["gross_salary"] == 30000

    def test_detail_excludes_draft_and_stale(self, fin_client):
        _, sf = fin_client
        self._seed_three_states(sf)
        with sf() as s:
            rows = svc.get_salary_detail(s, 2026, 5)
        assert len(rows) == 1
        assert rows[0]["gross_salary"] == 30000


class TestInsuredEmployeeCountByMonth:
    """get_insured_employee_count_by_month：單次撈列在記憶體分月，取代 12 次 COUNT。

    characterization 測試——refactor 前後皆須 GREEN，保證行為等價
    （hire/resign 邊界與 labor_insured_salary>0 從嚴條件不變）。
    """

    def test_counts_match_hire_resign_and_labor_insured(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            s.add_all(
                [
                    # A：2026-01-01 入職、無離職、有投保 → 全年 12 月計入
                    Employee(
                        employee_id="A",
                        name="A",
                        base_salary=30000,
                        employee_type="regular",
                        is_active=True,
                        hire_date=date(2026, 1, 1),
                        labor_insured_salary=30000,
                    ),
                    # B：2025-12-31 入職、2026-06-15 離職 → 1-6 月計入
                    Employee(
                        employee_id="B",
                        name="B",
                        base_salary=30000,
                        employee_type="regular",
                        is_active=False,
                        hire_date=date(2025, 12, 31),
                        resign_date=date(2026, 6, 15),
                        labor_insured_salary=30000,
                    ),
                    # C：labor_insured_salary=NULL → 不計入
                    Employee(
                        employee_id="C",
                        name="C",
                        base_salary=30000,
                        employee_type="regular",
                        is_active=True,
                        hire_date=date(2026, 1, 1),
                        labor_insured_salary=None,
                    ),
                    # D：2026-05-01 入職 → 5-12 月計入（hire_date < month_end_exclusive）
                    Employee(
                        employee_id="D",
                        name="D",
                        base_salary=25000,
                        employee_type="regular",
                        is_active=True,
                        hire_date=date(2026, 5, 1),
                        labor_insured_salary=25000,
                    ),
                    # E：labor_insured_salary=0 → 不計入（從嚴，0 不算投保）
                    Employee(
                        employee_id="E",
                        name="E",
                        base_salary=30000,
                        employee_type="regular",
                        is_active=True,
                        hire_date=date(2026, 1, 1),
                        labor_insured_salary=0,
                    ),
                ]
            )
            s.commit()
        with sf() as s:
            out = svc.get_insured_employee_count_by_month(s, 2026)
        # 1-4 月：A + B = 2
        assert [out[m] for m in range(1, 5)] == [2, 2, 2, 2]
        # 5-6 月：A + B + D = 3（B 6/15 離職，resign_date > 6/1 仍計入）
        assert out[5] == 3
        assert out[6] == 3
        # 7-12 月：A + D = 2
        assert [out[m] for m in range(7, 13)] == [2] * 6

    def test_resign_on_first_of_month_included_that_month(self, fin_client):
        """resign_date == 月初當天 → 該月仍計入（qa-loop #11：resign_date 為最後在職日，
        1 號離職者當月投保≥1 天，條件 resign_date >= month_first，對齊 gov_reports/salary）。"""
        _, sf = fin_client
        with sf() as s:
            s.add(
                Employee(
                    employee_id="R",
                    name="R",
                    base_salary=30000,
                    employee_type="regular",
                    is_active=False,
                    hire_date=date(2026, 1, 1),
                    resign_date=date(2026, 3, 1),
                    labor_insured_salary=30000,
                )
            )
            s.commit()
        with sf() as s:
            out = svc.get_insured_employee_count_by_month(s, 2026)
        assert out[1] == 1
        assert out[2] == 1
        assert out[3] == 1  # 3/1 離職、最後在職日當月、投保≥1 天 → 計入（qa-loop #11）
        assert out[4] == 0  # 3/1 < 4/1 → 4 月已不在職

    def test_empty_returns_all_zero(self, fin_client):
        _, sf = fin_client
        with sf() as s:
            out = svc.get_insured_employee_count_by_month(s, 2026)
        assert out == {m: 0 for m in range(1, 13)}
