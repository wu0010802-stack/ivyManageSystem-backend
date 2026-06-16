"""MOE Phase 2 月報 Excel 寫入器（openpyxl）。

3 sheet：班級總表 / 幼生明細 / 統計摘要
不需 embed 字型（XLSX 使用 OS 系統字型）。
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.excel_utils import SafeWorksheet

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT = Font(bold=True)


def _apply_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autofit_columns(ws, min_w: int = 10, max_w: int = 30) -> None:
    for col in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                # 中文字寬約英文 1.5 倍
                if any("一" <= ch <= "鿿" for ch in str(cell.value)):
                    length = int(length * 1.5)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)


def _dash_if_none(value):
    # 用全形破折號「—」（非 Excel 公式前綴）作為缺值占位，與總表合計列一致；
    # 半形「-」會被 SafeWorksheet 視為公式前綴而被加上單引號（SEC-004）。
    return "—" if value is None or value == "" else value


def _format_rate(rate_int: int) -> str:
    """rate_int = 9542 → '95.42%'（萬分比整數 → 百分比字串）"""
    return f"{rate_int / 100:.2f}%"


def _build_sheet1(wb: Workbook, rows: list[dict]) -> None:
    ws = SafeWorksheet(wb.create_sheet("班級總表"))
    headers = [
        "班級",
        "教師",
        "年齡層",
        "應到人日",
        "實到人日",
        "出席率",
        "男",
        "女",
        "弱勢",
        "身障",
        "原民",
        "外籍",
    ]
    ws.append(headers)
    _apply_header(ws)

    total_exp = total_act = 0
    total_m = total_f = total_dis = total_dis2 = total_ind = total_for = 0

    for r in rows:
        ws.append(
            [
                _dash_if_none(r.get("classroom_name")),
                _dash_if_none(r.get("teacher_names")),
                r.get("age_group") or "未知",
                r["expected_attendance_days"],
                r["actual_attendance_days"],
                _format_rate(r["attendance_rate"]),
                r["male_count"],
                r["female_count"],
                r["disadvantaged_count"],
                r["disability_count"],
                r["indigenous_count"],
                r["foreign_count"],
            ]
        )
        total_exp += r["expected_attendance_days"]
        total_act += r["actual_attendance_days"]
        total_m += r["male_count"]
        total_f += r["female_count"]
        total_dis += r["disadvantaged_count"]
        total_dis2 += r["disability_count"]
        total_ind += r["indigenous_count"]
        total_for += r["foreign_count"]

    total_rate = round(total_act / total_exp * 10000) if total_exp else 0
    total_row_idx = ws.max_row + 1
    ws.append(
        [
            "合計",
            "—",
            "—",
            total_exp,
            total_act,
            _format_rate(total_rate),
            total_m,
            total_f,
            total_dis,
            total_dis2,
            total_ind,
            total_for,
        ]
    )
    for cell in ws[total_row_idx]:
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _build_sheet2(wb: Workbook, details: list[dict]) -> None:
    ws = SafeWorksheet(wb.create_sheet("幼生明細"))
    headers = [
        "學號",
        "姓名",
        "身分證",
        "班級",
        "年齡層",
        "應到日數",
        "實到日數",
        "出席率",
        "弱勢標記",
    ]
    ws.append(headers)
    _apply_header(ws)

    for d in details:
        ws.append(
            [
                _dash_if_none(d.get("student_no")),
                _dash_if_none(d.get("name")),
                _dash_if_none(d.get("id_number")),
                _dash_if_none(d.get("classroom_name")),
                d.get("age_group") or "未知",
                d["expected_days"],
                d["actual_days"],
                f"{d['attendance_rate_pct']:.2f}%",
                "是" if d.get("is_disadvantaged") else "否",
            ]
        )

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _build_sheet3(wb: Workbook, overview: dict) -> None:
    ws = SafeWorksheet(wb.create_sheet("統計摘要"))
    lines = [
        ("總人數", overview.get("total_students", 0)),
        ("", ""),
        ("年齡層分布", ""),
    ]
    for ag in ["2-3", "3-4", "4-5", "5-6"]:
        cnt = overview.get("by_age_group", {}).get(ag, 0)
        lines.append((f"  {ag} 歲", cnt))
    lines.extend(
        [
            ("", ""),
            ("特殊屬性占比", ""),
            ("  弱勢", f"{overview.get('disadvantaged_pct', 0):.2f}%"),
            ("  身障", f"{overview.get('disability_pct', 0):.2f}%"),
            ("  原住民", f"{overview.get('indigenous_pct', 0):.2f}%"),
            ("  外籍", f"{overview.get('foreign_pct', 0):.2f}%"),
            ("", ""),
            ("出席統計", ""),
            ("  全園應到人日", overview.get("total_expected_days", 0)),
            ("  全園實到人日", overview.get("total_actual_days", 0)),
            ("  全園出席率", f"{overview.get('total_attendance_rate_pct', 0):.2f}%"),
            ("", ""),
            ("產生資訊", ""),
            ("  快照日期", str(overview.get("snapshot_date", ""))),
            (
                "  產生時間",
                (
                    overview.get("generated_at").strftime("%Y-%m-%d %H:%M")
                    if overview.get("generated_at")
                    else "—"
                ),
            ),
            ("  產生人", overview.get("generated_by", "—")),
        ]
    )
    for label, value in lines:
        ws.append([label, value])

    # 第一欄主標題粗體（沒縮排的 label）
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value and not cell.value.startswith(" "):
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def build_monthly_xlsx_bytes(
    snapshot_rows: list[dict],
    student_details: list[dict],
    overview: dict,
) -> bytes:
    """產 xlsx bytes（3 sheets：班級總表 / 幼生明細 / 統計摘要）。"""
    wb = Workbook()
    wb.remove(wb.active)  # remove default Sheet
    _build_sheet1(wb, snapshot_rows)
    _build_sheet2(wb, student_details)
    _build_sheet3(wb, overview)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
