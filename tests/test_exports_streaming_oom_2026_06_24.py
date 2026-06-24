"""崩潰防護 P1：全校學生 / 員工名冊匯出改 write_only streaming，避免大表 OOM worker。

問題：export_students / export_employees 用 `list(session.query(...).yield_per(500))`
—— 外層 list() 抵消 streaming，把整張表 ORM row materialize；再寫進**非** write_only
的 Workbook（全部 Cell 物件常駐 RAM）。學生終態 row 不 hard-delete、逐年單調成長，
幾年後一次匯出可把單 uvicorn worker 記憶體吃爆 → OOM-kill → 重啟（全站中斷）。

修法：改 openpyxl `Workbook(write_only=True)` + 直接迭代 yield_per cursor（不 list()）。
write_only 把列 streaming 寫到暫存檔、不保留全部 Cell 物件，記憶體不隨資料量爆。
（write_only 不支援 merge_cells / 讀 cell 算 auto-width，故標題不合併、欄寬用固定值
——屬可接受的外觀微調，資料內容與列結構不變、仍經 _sanitize_excel_value 防公式注入。）
"""

import io

import inspect

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import models.base as base_module
import api.exports as exports_module
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from api.exports import _export_rate_limit
from api.exports import router as exports_router
from models.base import Base
from models.database import Student, User
from utils.auth import hash_password


@pytest.fixture
def exports_client(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'exp-stream.sqlite'}",
        connect_args={"check_same_thread": False},
    )
    sf = sessionmaker(bind=engine)
    old_e, old_s = base_module._engine, base_module._SessionFactory
    base_module._engine, base_module._SessionFactory = engine, sf
    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    with sf() as s:
        admin = User(
            username="exp_admin",
            password_hash=hash_password("Temp123456"),
            role="admin",
            permission_names=["*"],  # wildcard → assert_all_scope 通過
            employee_id=None,
            is_active=True,
            must_change_password=False,
        )
        s.add(admin)
        for i in range(7):
            s.add(
                Student(
                    student_id=f"S{i:03d}",
                    name=f"學生{i}",
                    gender="M" if i % 2 == 0 else "F",
                    is_active=True,
                )
            )
        s.commit()

    app = FastAPI()
    from utils.exception_handlers import register_exception_handlers

    register_exception_handlers(app)
    app.include_router(auth_router)
    app.include_router(exports_router)
    app.dependency_overrides[_export_rate_limit] = lambda: None

    with TestClient(app) as client:
        login = client.post(
            "/api/auth/login",
            json={"username": "exp_admin", "password": "Temp123456"},
        )
        assert login.status_code == 200, login.text
        yield client

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine, base_module._SessionFactory = old_e, old_s
    engine.dispose()


def _data_rows(content: bytes) -> int:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # 結構：row1=標題、row2=空白、row3=表頭、row4+=資料
    return max(ws.max_row - 3, 0)


def test_students_export_correctness(exports_client):
    """重寫後行為不變：7 名學生 → 7 列資料，學號/姓名正確。"""
    res = exports_client.get("/api/exports/students")
    assert res.status_code == 200, res.text
    assert _data_rows(res.content) == 7
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    first_data = [c.value for c in ws[4]]
    assert first_data[0] == "S000"  # 學號
    assert first_data[1] == "學生0"  # 姓名


def test_students_export_uses_write_only_streaming():
    src = inspect.getsource(exports_module.export_students)
    # 走 streaming helper + 直接迭代 cursor（不把整張表 materialize 成 list）
    assert "_stream_export_response" in src, "export_students 未走 streaming 匯出路徑"
    assert (
        "yield_per" in src
    ), "export_students 未以 yield_per 逐批迭代 → 仍可能整表載入"


def test_employees_export_uses_write_only_streaming():
    src = inspect.getsource(exports_module.export_employees)
    assert "_stream_export_response" in src, "export_employees 未走 streaming 匯出路徑"
    assert (
        "yield_per" in src
    ), "export_employees 未以 yield_per 逐批迭代 → 仍可能整表載入"


@pytest.mark.parametrize(
    "path",
    [
        "/api/exports/attendance",
        "/api/exports/leaves",
        "/api/exports/overtimes",
    ],
)
def test_export_month_endpoints_reject_out_of_range_month(exports_client, path):
    """這些月報匯出把 month 餵進 monthrange()/date()，month=13 須被 422 擋下而非 500。"""
    res = exports_client.get(f"{path}?year=2026&month=13")
    assert res.status_code == 422, f"{path} month=13 應 422，實得 {res.status_code}"


def test_export_holidays_rejects_out_of_range_year(exports_client):
    res = exports_client.get("/api/exports/holidays?year=0")
    assert res.status_code == 422, res.text


def test_module_adopts_write_only_workbook():
    # streaming 路徑採用 write_only Workbook（不保留全部 Cell 物件）
    src = inspect.getsource(exports_module)
    assert (
        "write_only=True" in src
    ), "大表匯出未採用 Workbook(write_only=True) streaming"
