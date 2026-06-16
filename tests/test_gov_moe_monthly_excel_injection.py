"""教育部月報 Excel 匯出須過 SafeWorksheet 防公式注入（SEC-004 / 資安掃描 2026-06-15 P2）。

build_monthly_xlsx_bytes 原以裸 wb.create_sheet() + ws.append() 寫入幼生姓名 /
學號 / 身分證 / 班級名 / 教師名等員工可編輯欄位，若含 =cmd|'/C calc'!A0 會在
教育部 / 財會承辦人端 Excel 開啟時被當公式執行（DDE / HYPERLINK），且月報含全園
幼生完整身分證，整列 PII 可經 cell-reference 串接外滲。

修法：三個 _build_sheetN 改用 SafeWorksheet(wb.create_sheet(...))，與 appraisal /
year_end / gov_reports 等既有匯出對齊。
"""

from io import BytesIO

from openpyxl import load_workbook

from services.gov_moe.monthly_excel_writer import build_monthly_xlsx_bytes

EVIL = "=cmd|'/C calc'!A0"


def _all_values(payload: bytes) -> list:
    wb = load_workbook(BytesIO(payload))
    vals = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    vals.append(c.value)
    return vals


def _snapshot_row(**overrides) -> dict:
    base = dict(
        classroom_name="小班",
        teacher_names="王老師",
        age_group="3-4",
        expected_attendance_days=20,
        actual_attendance_days=18,
        attendance_rate=9000,
        male_count=5,
        female_count=5,
        disadvantaged_count=0,
        disability_count=0,
        indigenous_count=0,
        foreign_count=0,
    )
    base.update(overrides)
    return base


def _student_detail(**overrides) -> dict:
    base = dict(
        student_no="S001",
        name="陳小明",
        id_number="A123456789",
        classroom_name="小班",
        age_group="3-4",
        expected_days=20,
        actual_days=18,
        attendance_rate_pct=90.0,
        is_disadvantaged=False,
    )
    base.update(overrides)
    return base


def _overview(**overrides) -> dict:
    base = dict(
        total_students=1,
        by_age_group={"3-4": 1},
        disadvantaged_pct=0.0,
        disability_pct=0.0,
        indigenous_pct=0.0,
        foreign_pct=0.0,
        total_expected_days=20,
        total_actual_days=18,
        total_attendance_rate_pct=90.0,
        snapshot_date="2026-06-01",
        generated_at=None,
        generated_by="管理員",
    )
    base.update(overrides)
    return base


def test_sheet2_sanitizes_student_name_injection():
    payload = build_monthly_xlsx_bytes(
        [_snapshot_row()], [_student_detail(name=EVIL)], _overview()
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "幼生姓名公式注入未被清理（原樣寫入 → 承辦人端可被執行）"
    assert ("'" + EVIL) in vals, "幼生姓名應被 sanitize 為 ' 前綴純字串"


def test_sheet2_sanitizes_id_number_injection():
    payload = build_monthly_xlsx_bytes(
        [_snapshot_row()], [_student_detail(id_number=EVIL)], _overview()
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "幼生身分證欄公式注入未被清理"
    assert ("'" + EVIL) in vals


def test_sheet1_sanitizes_classroom_and_teacher_injection():
    payload = build_monthly_xlsx_bytes(
        [_snapshot_row(classroom_name=EVIL, teacher_names=EVIL)],
        [_student_detail()],
        _overview(),
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "班級名 / 教師名公式注入未被清理"
    assert ("'" + EVIL) in vals


def test_sheet3_sanitizes_generated_by_injection():
    payload = build_monthly_xlsx_bytes(
        [_snapshot_row()], [_student_detail()], _overview(generated_by=EVIL)
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "產生人欄公式注入未被清理"
    assert ("'" + EVIL) in vals
