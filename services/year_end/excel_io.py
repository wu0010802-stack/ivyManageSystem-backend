"""年終獎金 Excel I/O（M3）— 雙向轉換。

對應 Excel「114年年終經營績效」22 sheets 中的 9 大核心區塊：
  - 「114.01.15」               每員工平均績效輸入（step1 來源）
  - 「年終獎金」                 每員工 6-step 計算明細
  - 「年終獎金總表」              特別獎金 8 欄 + 合計（最核心，importer 主表）
  - 「轉帳名冊~P6」              帳號 + 金額
  - 「班級經營績效114.01.15」    班級每月人數 + 編制 + 經營績效
  - 「114上節慶獎金比例差額」    FESTIVAL_DIFF 特別獎金
  - 「114上超額獎金」            EXCESS_ENROLLMENT 特別獎金
  - 「鼓勵課後才藝統計表」        AFTER_CLASS_AWARD 特別獎金
  - 「113上.113下學期紅利獎金」  SEMESTER_DIVIDEND_FIRST/SECOND 特別獎金
  - 「教課教師獎勵金」            TEACHING_EXTRA 特別獎金（複雜分頁，先抓合計）

主要 API：
  parse_year_end_excel(path) → ParsedYearEndExcel
  import_year_end_to_db(parsed, session, ...) → YearEndImportResult
  export_year_end_summary_xlsx(...) → bytes  # 年終獎金總表
  export_year_end_transfer_xlsx(...) → bytes # 轉帳名冊
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable, Optional

from sqlalchemy.orm import Session

from models.year_end import (
    ClassEnrollmentTarget,
    EmployeeYearEndSnapshot,
    OrgYearSettings,
    SpecialBonusItem,
    SpecialBonusType,
    YearEndCycle,
    YearEndCycleStatus,
    YearEndSettlement,
    YearEndSettlementStatus,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class ParsedSettlementRow:
    """對應「年終獎金」sheet 每行（一位員工）的 6-step 計算明細。"""

    excel_row: int
    name: str
    base_salary: Decimal
    festival_total: Decimal
    avg_performance_rate: Decimal
    gross_amount: Decimal
    org_achievement_rate: Decimal
    subtotal: Decimal
    deduction_leave_late: Decimal = Decimal("0")
    deduction_disciplinary: Decimal = Decimal("0")
    deduction_meeting: Decimal = Decimal("0")
    deduction_personal_leave: Decimal = Decimal("0")
    deduction_sick_leave: Decimal = Decimal("0")
    deduction_late: Decimal = Decimal("0")
    total_in_year: Decimal = Decimal("12")
    payable: Decimal = Decimal("0")
    remark: Optional[str] = None


@dataclass
class ParsedSpecialBonus:
    name: str
    bonus_type: SpecialBonusType
    period_label: str
    amount: Decimal
    calc_meta: dict = field(default_factory=dict)


@dataclass
class ParsedClassEnrollmentTarget:
    semester_first: bool
    class_name: str
    head_teacher_name: Optional[str]
    assistant_name: Optional[str]
    head_count_target: int
    avg_monthly_enrollment: Decimal
    class_performance_rate: Decimal
    returning_student_rate: Decimal


@dataclass
class ParsedYearEndExcel:
    academic_year: int
    settlements: list[ParsedSettlementRow] = field(default_factory=list)
    special_bonuses: list[ParsedSpecialBonus] = field(default_factory=list)
    class_targets: list[ParsedClassEnrollmentTarget] = field(default_factory=list)


@dataclass
class YearEndImportResult:
    cycle_id: int
    settlements_upserted: int
    special_bonuses_upserted: int
    class_targets_upserted: int
    skipped_unresolved_names: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _to_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


def _to_decimal_opt(value) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_summary_data_row(name: str) -> bool:
    """跳過標題、表頭、空、規則說明等 row。"""
    if not name:
        return False
    skip_starts = ("常春藤", "姓名", "姓　名", "*", "李麗珍 114", "1月", "備註")
    if any(name.startswith(s) for s in skip_starts):
        return False
    if "年度" in name or "獎金" in name:
        # 子標題如「常春藤幼兒園 114年平均績效及年終獎金」
        if name.startswith(("常春藤", "高雄市")):
            return False
    return True


def _open_xls_book(path: Path):
    import xlrd  # type: ignore[import-untyped]

    return xlrd.open_workbook(str(path))


def _sheet_rows(sheet) -> list[list]:
    return [
        [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
    ]


# ---------------------------------------------------------------------------
# Parsers — 各 sheet 拆開
# ---------------------------------------------------------------------------


def _parse_year_end_main_sheet(rows: list[list]) -> list[ParsedSettlementRow]:
    """解析「年終獎金」sheet（每行一位員工的 6-step 明細）。

    欄位順序（r01 標頭）：
      0 姓名 | 1 基本薪俸 | 2 節慶獎金 | 3 合計 | 4 平均績效 | 5 年終獎金 |
      6 達成比率 | 7 小計 | 8 114.02-114.12 請假遲到 | 9 115.01 請假遲到 |
      10 奬懲 | 11 自強活動/機構會議 | 12 事假 | 13 病假/育嬰假 |
      14 遲到/早退 | 15 合計 | 16 到職(月) | 17 應領小計 | 18 備註
    """
    results = []
    for r_idx, row in enumerate(rows):
        if not row:
            continue
        name = _normalize(row[0]) if len(row) > 0 else ""
        if not _is_summary_data_row(name):
            continue
        # 跳過 header / 子標題重複 row（姓名為空欄但有「常春藤」/「姓名」）
        # 跳過「常春藤幼兒園 114年...」子標題行
        # base_salary 欄位若為空跳過 — 例如「呂麗珍」第一行帶數字、第二行空欄為續行
        base_salary = _to_decimal_opt(row[1]) if len(row) > 1 else None
        if base_salary is None or base_salary == 0:
            # 例外：兼職如「李麗珍」的列只填合計，base/festival 為空
            # 此情況下年終獎金欄(col 5)直接給合計；payable 是 col 5
            payable = _to_decimal_opt(row[5]) if len(row) > 5 else None
            if payable is None or payable == 0:
                continue
            results.append(
                ParsedSettlementRow(
                    excel_row=r_idx,
                    name=name,
                    base_salary=Decimal("0"),
                    festival_total=Decimal("0"),
                    avg_performance_rate=Decimal("0"),
                    gross_amount=Decimal("0"),
                    org_achievement_rate=Decimal("0"),
                    subtotal=Decimal("0"),
                    payable=payable,
                    remark=_normalize(row[18]) if len(row) > 18 else None,
                )
            )
            continue

        festival = _to_decimal(row[2]) if len(row) > 2 else Decimal("0")
        avg_perf = _to_decimal(row[4]) if len(row) > 4 else Decimal("0")
        gross = _to_decimal(row[5]) if len(row) > 5 else Decimal("0")
        org_rate = _to_decimal(row[6]) if len(row) > 6 else Decimal("0")
        subtotal = _to_decimal(row[7]) if len(row) > 7 else Decimal("0")
        leave_late = _to_decimal(row[8]) if len(row) > 8 else Decimal("0")
        disciplinary = _to_decimal(row[10]) if len(row) > 10 else Decimal("0")
        meeting = _to_decimal(row[11]) if len(row) > 11 else Decimal("0")
        personal_leave = _to_decimal(row[12]) if len(row) > 12 else Decimal("0")
        sick_leave = _to_decimal(row[13]) if len(row) > 13 else Decimal("0")
        late = _to_decimal(row[14]) if len(row) > 14 else Decimal("0")
        hire_months = _to_decimal(row[16], default=Decimal("12")) if len(row) > 16 else Decimal("12")
        payable = _to_decimal(row[17]) if len(row) > 17 else Decimal("0")
        remark = _normalize(row[18]) if len(row) > 18 else None
        # Excel 中 hire_months 欄空白代表滿 12 個月
        if hire_months == 0:
            hire_months = Decimal("12")

        results.append(
            ParsedSettlementRow(
                excel_row=r_idx,
                name=name,
                base_salary=base_salary,
                festival_total=festival,
                avg_performance_rate=avg_perf,
                gross_amount=gross,
                org_achievement_rate=org_rate,
                subtotal=subtotal,
                deduction_leave_late=leave_late,
                deduction_disciplinary=disciplinary,
                deduction_meeting=meeting,
                deduction_personal_leave=personal_leave,
                deduction_sick_leave=sick_leave,
                deduction_late=late,
                total_in_year=hire_months,
                payable=payable,
                remark=remark or None,
            )
        )
    return results


def _parse_year_end_summary_sheet(
    rows: list[list],
) -> list[ParsedSpecialBonus]:
    """解析「年終獎金總表」sheet。

    欄位（r03 標頭）：
      0 姓名 | 1 年終獎金 | 2 113上考核 | 3 113下考核 | 4 113上紅利 |
      5 113下紅利 | 6 114上鼓勵才藝 | 7 114上教課 | 8 114上超額 |
      9 114上節慶差額 | 10 合計

    回傳各員工 8 種特別獎金（金額 > 0 才產出 record）。
    """
    type_map: list[tuple[int, SpecialBonusType, str]] = [
        (2, SpecialBonusType.APPRAISAL_HALF_BONUS_FIRST, "113上"),
        (3, SpecialBonusType.APPRAISAL_HALF_BONUS_SECOND, "113下"),
        (4, SpecialBonusType.SEMESTER_DIVIDEND_FIRST, "113上"),
        (5, SpecialBonusType.SEMESTER_DIVIDEND_SECOND, "113下"),
        (6, SpecialBonusType.AFTER_CLASS_AWARD, "114上"),
        (7, SpecialBonusType.TEACHING_EXTRA, "114上"),
        (8, SpecialBonusType.EXCESS_ENROLLMENT, "114上"),
        (9, SpecialBonusType.FESTIVAL_DIFF, "114.8-115.01"),
    ]

    results: list[ParsedSpecialBonus] = []
    for r_idx, row in enumerate(rows):
        if not row:
            continue
        name = _normalize(row[0]) if len(row) > 0 else ""
        if not name or name in ("姓名", "學期", "核算  日期"):
            continue
        # 跳過子標題與規則行
        if name.startswith(("常春藤", "高雄市", "*")):
            continue
        if "獎金" in name and "考核" in name:
            continue

        for col_idx, btype, period in type_map:
            if col_idx >= len(row):
                continue
            amount = _to_decimal_opt(row[col_idx])
            if amount is None or amount == 0:
                continue
            results.append(
                ParsedSpecialBonus(
                    name=name,
                    bonus_type=btype,
                    period_label=period,
                    amount=amount,
                    calc_meta={"excel_row": r_idx, "excel_col": col_idx},
                )
            )
    return results


def _parse_class_performance_sheet(
    rows: list[list],
) -> list[ParsedClassEnrollmentTarget]:
    """解析「班級經營績效114.01.15」sheet — 上下學期各一段，
    每段以「(113學年度下學期)」「(114學年度上學期)」標題行起始。
    """
    results: list[ParsedClassEnrollmentTarget] = []
    semester_first = True
    for r_idx, row in enumerate(rows):
        if not row:
            continue
        first_cell = _normalize(row[0]) if len(row) > 0 else ""
        # 標題行（含「上學期」「下學期」）切換 semester_first
        full_line = " ".join(_normalize(c) for c in row[:3])
        if "下學期" in full_line and "班級經營" in full_line:
            semester_first = False
            continue
        if "上學期" in full_line and "班級經營" in full_line:
            semester_first = True
            continue
        if first_cell in ("", "班級", "合計"):
            continue
        if first_cell in ("天堂鳥", "百合", "櫻花", "茉莉", "薔薇", "芙蓉", "牡丹", "向日葵", "滿天星"):
            head_name = _normalize(row[1]) if len(row) > 1 else None
            assist = _normalize(row[2]) if len(row) > 2 else None
            head_count = int(_to_decimal(row[11], Decimal("0"))) if len(row) > 11 else 0
            class_perf = _to_decimal(row[12]) if len(row) > 12 else Decimal("0")
            returning = _to_decimal(row[13]) if len(row) > 13 else Decimal("0")
            avg_enrol = _to_decimal(row[10]) if len(row) > 10 else Decimal("0")
            results.append(
                ParsedClassEnrollmentTarget(
                    semester_first=semester_first,
                    class_name=first_cell,
                    head_teacher_name=head_name if head_name and head_name != "已離職" else None,
                    assistant_name=assist if assist and assist != "已離職" else None,
                    head_count_target=head_count,
                    avg_monthly_enrollment=avg_enrol,
                    class_performance_rate=class_perf,
                    returning_student_rate=returning,
                )
            )
    return results


def parse_year_end_excel(path: Path | str) -> ParsedYearEndExcel:
    """解析年終經營績效 .xls 主要 sheets。

    回傳合併三大來源：年終獎金（settlements）+ 年終獎金總表（special_bonuses）+
    班級經營績效（class_targets）。學年從 sheet 名「114」開頭推斷。
    """
    p = Path(path)
    if p.suffix.lower() != ".xls":
        raise ValueError("年終經營績效目前只支援 .xls (Excel 97-2003)")
    book = _open_xls_book(p)

    # 推斷學年（從 sheet 名「114上節慶...」「班級經營績效114.01.15」抓 3 位數）
    import re

    year_pattern = re.compile(r"\b(\d{3})\b")
    academic_year = 114  # default
    for sname in book.sheet_names():
        m = year_pattern.search(sname)
        if m:
            academic_year = int(m.group(1))
            break

    settlements: list[ParsedSettlementRow] = []
    special_bonuses: list[ParsedSpecialBonus] = []
    class_targets: list[ParsedClassEnrollmentTarget] = []

    if "年終獎金" in book.sheet_names():
        rows = _sheet_rows(book.sheet_by_name("年終獎金"))
        settlements = _parse_year_end_main_sheet(rows)

    if "年終獎金總表" in book.sheet_names():
        rows = _sheet_rows(book.sheet_by_name("年終獎金總表"))
        special_bonuses = _parse_year_end_summary_sheet(rows)

    if "班級經營績效114.01.15" in book.sheet_names():
        rows = _sheet_rows(book.sheet_by_name("班級經營績效114.01.15"))
        class_targets = _parse_class_performance_sheet(rows)

    return ParsedYearEndExcel(
        academic_year=academic_year,
        settlements=settlements,
        special_bonuses=special_bonuses,
        class_targets=class_targets,
    )


# ---------------------------------------------------------------------------
# Importer
# ---------------------------------------------------------------------------


def import_year_end_to_db(
    parsed: ParsedYearEndExcel,
    session: Session,
    *,
    employee_resolver: Callable[[str], Optional[int]],
    classroom_resolver: Optional[Callable[[str], Optional[int]]] = None,
    cycle_dates: tuple[date, date, date],
    org_achievement_rate_first: Decimal,
    org_achievement_rate_second: Decimal,
    enrollment_target: int = 160,
) -> YearEndImportResult:
    """把解析結果 upsert 到 DB（year_end_cycles / org_year_settings /
    employee_year_end_snapshot / year_end_settlements / special_bonus_items /
    class_enrollment_targets）。

    Args:
        parsed: parse_year_end_excel 的回傳
        session: SQLAlchemy session（呼叫方 commit）
        employee_resolver: 姓名 → employee_id
        classroom_resolver: 班級名 → classroom_id（用於班級績效）
        cycle_dates: (start_date, end_date, bonus_calc_date)
        org_achievement_rate_first/second: 上/下學期機構達成比率
        enrollment_target: 預設招生目標（160）
    """
    start_date, end_date, bonus_calc_date = cycle_dates

    # cycle upsert
    cycle = (
        session.query(YearEndCycle)
        .filter_by(academic_year=parsed.academic_year)
        .one_or_none()
    )
    if cycle is None:
        cycle = YearEndCycle(
            academic_year=parsed.academic_year,
            start_date=start_date,
            end_date=end_date,
            bonus_calc_date=bonus_calc_date,
            status=YearEndCycleStatus.OPEN,
        )
        session.add(cycle)
        session.flush()

    # org_year_settings — 上下學期各一筆
    for sem_first, rate in (
        (True, org_achievement_rate_first),
        (False, org_achievement_rate_second),
    ):
        existing = (
            session.query(OrgYearSettings)
            .filter_by(year_end_cycle_id=cycle.id, semester_first=sem_first)
            .one_or_none()
        )
        if existing is None:
            session.add(
                OrgYearSettings(
                    year_end_cycle_id=cycle.id,
                    semester_first=sem_first,
                    enrollment_target=enrollment_target,
                    org_achievement_rate=rate,
                )
            )
        else:
            existing.org_achievement_rate = rate

    # class_enrollment_targets upsert
    class_targets_count = 0
    for ct in parsed.class_targets:
        classroom_id = classroom_resolver(ct.class_name) if classroom_resolver else None
        if classroom_id is None:
            logger.warning(
                "import_year_end: 找不到班級 %r，跳過 class_enrollment_targets",
                ct.class_name,
            )
            continue
        head_id = (
            employee_resolver(ct.head_teacher_name) if ct.head_teacher_name else None
        )
        assist_id = (
            employee_resolver(ct.assistant_name) if ct.assistant_name else None
        )
        existing = (
            session.query(ClassEnrollmentTarget)
            .filter_by(
                year_end_cycle_id=cycle.id,
                semester_first=ct.semester_first,
                classroom_id=classroom_id,
            )
            .one_or_none()
        )
        if existing is None:
            session.add(
                ClassEnrollmentTarget(
                    year_end_cycle_id=cycle.id,
                    semester_first=ct.semester_first,
                    classroom_id=classroom_id,
                    head_teacher_employee_id=head_id,
                    assistant_employee_id=assist_id,
                    head_count_target=ct.head_count_target,
                    avg_monthly_enrollment=ct.avg_monthly_enrollment,
                    class_performance_rate=ct.class_performance_rate,
                    returning_student_rate=ct.returning_student_rate,
                )
            )
        else:
            existing.head_teacher_employee_id = head_id
            existing.assistant_employee_id = assist_id
            existing.head_count_target = ct.head_count_target
            existing.avg_monthly_enrollment = ct.avg_monthly_enrollment
            existing.class_performance_rate = ct.class_performance_rate
            existing.returning_student_rate = ct.returning_student_rate
        class_targets_count += 1

    # employee_year_end_snapshot + year_end_settlements upsert
    skipped: list[str] = []
    settlements_count = 0
    for s in parsed.settlements:
        emp_id = employee_resolver(s.name)
        if emp_id is None:
            skipped.append(s.name)
            continue
        # snapshot upsert
        snapshot = (
            session.query(EmployeeYearEndSnapshot)
            .filter_by(year_end_cycle_id=cycle.id, employee_id=emp_id)
            .one_or_none()
        )
        if snapshot is None:
            snapshot = EmployeeYearEndSnapshot(
                year_end_cycle_id=cycle.id,
                employee_id=emp_id,
                base_salary=s.base_salary,
                festival_total=s.festival_total,
                hire_months=s.total_in_year,
                is_contracted=True,
            )
            session.add(snapshot)
            session.flush()
        else:
            snapshot.base_salary = s.base_salary
            snapshot.festival_total = s.festival_total
            snapshot.hire_months = s.total_in_year

        # settlement upsert
        deduction_total = (
            s.deduction_leave_late
            + s.deduction_disciplinary
            + s.deduction_meeting
            + s.deduction_personal_leave
            + s.deduction_sick_leave
            + s.deduction_late
        )
        proration = (s.total_in_year / Decimal("12")).quantize(Decimal("0.0001"))
        settlement = (
            session.query(YearEndSettlement)
            .filter_by(year_end_cycle_id=cycle.id, employee_id=emp_id)
            .one_or_none()
        )
        if settlement is None:
            settlement = YearEndSettlement(
                year_end_cycle_id=cycle.id,
                employee_id=emp_id,
                snapshot_id=snapshot.id,
                avg_performance_rate=s.avg_performance_rate,
                base_salary=s.base_salary,
                festival_total=s.festival_total,
                gross_amount=s.gross_amount,
                org_achievement_rate=s.org_achievement_rate,
                subtotal_amount=s.subtotal,
                deduction_leave_late=s.deduction_leave_late,
                deduction_meeting=s.deduction_meeting,
                deduction_personal_leave=s.deduction_personal_leave,
                deduction_sick_leave=s.deduction_sick_leave,
                deduction_late=s.deduction_late,
                deduction_disciplinary=s.deduction_disciplinary,
                deduction_total=deduction_total,
                hire_months=s.total_in_year,
                proration_rate=proration,
                payable_amount=s.payable,
                total_amount=s.payable,  # 加 special_bonus_total 後續更新
                remark=s.remark,
                status=YearEndSettlementStatus.DRAFT,
            )
            session.add(settlement)
        else:
            settlement.snapshot_id = snapshot.id
            settlement.avg_performance_rate = s.avg_performance_rate
            settlement.base_salary = s.base_salary
            settlement.festival_total = s.festival_total
            settlement.gross_amount = s.gross_amount
            settlement.org_achievement_rate = s.org_achievement_rate
            settlement.subtotal_amount = s.subtotal
            settlement.deduction_leave_late = s.deduction_leave_late
            settlement.deduction_meeting = s.deduction_meeting
            settlement.deduction_personal_leave = s.deduction_personal_leave
            settlement.deduction_sick_leave = s.deduction_sick_leave
            settlement.deduction_late = s.deduction_late
            settlement.deduction_disciplinary = s.deduction_disciplinary
            settlement.deduction_total = deduction_total
            settlement.hire_months = s.total_in_year
            settlement.proration_rate = proration
            settlement.payable_amount = s.payable
            settlement.remark = s.remark
        settlements_count += 1

    # special_bonus_items upsert + 更新 settlement.special_bonus_total / total_amount
    special_count = 0
    employee_special_totals: dict[int, Decimal] = {}
    for sb in parsed.special_bonuses:
        emp_id = employee_resolver(sb.name)
        if emp_id is None:
            skipped.append(sb.name)
            continue
        existing = (
            session.query(SpecialBonusItem)
            .filter_by(
                year_end_cycle_id=cycle.id,
                employee_id=emp_id,
                bonus_type=sb.bonus_type,
                period_label=sb.period_label,
            )
            .one_or_none()
        )
        if existing is None:
            session.add(
                SpecialBonusItem(
                    year_end_cycle_id=cycle.id,
                    employee_id=emp_id,
                    bonus_type=sb.bonus_type,
                    period_label=sb.period_label,
                    amount=sb.amount,
                    calc_meta=sb.calc_meta,
                    source_ref="年終獎金總表",
                )
            )
        else:
            existing.amount = sb.amount
            existing.calc_meta = sb.calc_meta
        employee_special_totals[emp_id] = (
            employee_special_totals.get(emp_id, Decimal("0")) + sb.amount
        )
        special_count += 1

    session.flush()
    # 把 special_bonus_total 同步到 settlements
    for emp_id, total in employee_special_totals.items():
        settlement = (
            session.query(YearEndSettlement)
            .filter_by(year_end_cycle_id=cycle.id, employee_id=emp_id)
            .one_or_none()
        )
        if settlement is not None:
            settlement.special_bonus_total = total
            settlement.total_amount = settlement.payable_amount + total

    session.flush()
    return YearEndImportResult(
        cycle_id=cycle.id,
        settlements_upserted=settlements_count,
        special_bonuses_upserted=special_count,
        class_targets_upserted=class_targets_count,
        skipped_unresolved_names=sorted(set(skipped)),
    )


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------


@dataclass
class TransferRow:
    bank_account: str
    name: str
    amount: Decimal


def export_year_end_transfer_xlsx(
    *,
    rows: Iterable[TransferRow],
    school_name: str = "高雄市私立常春藤幼兒園",
    title: str = "年終獎金 轉帳名冊",
    org_bank_account: str = "0727-940-008106",
) -> bytes:
    """產出年終轉帳名冊 .xlsx，對應 Excel「轉帳名冊~P6」。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "年終轉帳名冊"

    ws.cell(row=1, column=2, value=school_name).font = Font(bold=True, size=14)
    ws.cell(row=2, column=2, value=title).font = Font(bold=True, size=12)
    ws.cell(row=3, column=2, value=f"帳號：{org_bank_account}")
    for c, h in enumerate(("帳號", "戶名", "金額"), start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    total = Decimal("0")
    r = 5
    for row in rows:
        if row.amount <= 0:
            continue
        ws.cell(row=r, column=1, value=row.bank_account)
        ws.cell(row=r, column=2, value=row.name)
        ws.cell(row=r, column=3, value=float(row.amount))
        total += row.amount
        r += 1
    ws.cell(row=r + 1, column=2, value="合計").font = Font(bold=True)
    ws.cell(row=r + 1, column=3, value=float(total)).font = Font(bold=True)
    ws.cell(row=r + 2, column=2, value="經辦人:")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class SummaryExportRow:
    name: str
    year_end_amount: Decimal  # payable_amount
    bonus_by_type: dict[SpecialBonusType, Decimal]  # 8 種獎金
    total: Decimal


def export_year_end_summary_xlsx(
    *,
    rows: Iterable[SummaryExportRow],
    academic_year: int = 114,
    school_name: str = "高雄市私立常春藤幼兒園",
) -> bytes:
    """產出年終獎金總表 .xlsx — 與 Excel「年終獎金總表」結構相同。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "年終獎金總表"

    ws.cell(row=1, column=1, value=f"{academic_year}年度年終分紅獎金").font = Font(
        bold=True, size=14
    )
    headers = (
        "姓名",
        "年終獎金",
        "113上考核獎金",
        "113下考核獎金",
        "113上學期紅利",
        "113下學期紅利",
        "114上鼓勵才藝獎金",
        "114上教課教師獎勵金",
        "114上超額獎金",
        "114上節慶獎金差額",
        "合計",
    )
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    type_order = [
        SpecialBonusType.APPRAISAL_HALF_BONUS_FIRST,
        SpecialBonusType.APPRAISAL_HALF_BONUS_SECOND,
        SpecialBonusType.SEMESTER_DIVIDEND_FIRST,
        SpecialBonusType.SEMESTER_DIVIDEND_SECOND,
        SpecialBonusType.AFTER_CLASS_AWARD,
        SpecialBonusType.TEACHING_EXTRA,
        SpecialBonusType.EXCESS_ENROLLMENT,
        SpecialBonusType.FESTIVAL_DIFF,
    ]
    r = 3
    for row in rows:
        ws.cell(row=r, column=1, value=row.name)
        ws.cell(row=r, column=2, value=float(row.year_end_amount))
        for col_offset, t in enumerate(type_order, start=3):
            v = row.bonus_by_type.get(t)
            if v is not None and v != 0:
                ws.cell(row=r, column=col_offset, value=float(v))
        ws.cell(row=r, column=11, value=float(row.total))
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
