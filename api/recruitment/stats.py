"""招生統計查詢、Excel 匯出、未預繳名單分析。"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, case, cast, func, String

from models.base import session_scope
from models.recruitment import RecruitmentVisit
from utils.auth import require_staff_permission
from utils.excel_utils import xlsx_streaming_response
from utils.permissions import Permission

from api.recruitment.shared import (
    COLD_LEAD_DAYS,
    DATASET_SCOPE_ALL,
    DEFAULT_OVERDUE_DAYS,
    HIGH_PRIORITY_NO_DEPOSIT_REASONS,
    NO_DEPOSIT_PRIORITY_REASON_MAP,
    TOP_SOURCES_COUNT,
    _aggregate_snapshot,
    _build_action_queue,
    _build_alerts,
    _build_month_over_month,
    _build_scoped_query,
    _build_ytd_snapshot,
    _chuannian_sql_cond,
    _dataset_scope_filters,
    _empty_snapshot,
    _find_source_imbalance,
    _metric_snapshot,
    _normalize_dataset_scope,
    _roc_month_sort_key,
    _safe_normalize_roc_month,
    _select_reference_month,
    _shift_roc_month,
    _source_group_sql_expr,
    _to_dict,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/recruitment", tags=["recruitment-stats"])


# ---------------------------------------------------------------------------
# 核心統計計算（SQL GROUP BY）
# ---------------------------------------------------------------------------


def _query_stats(
    session,
    reference_month: Optional[str] = None,
    dataset_scope: Optional[str] = None,
) -> dict:
    """執行招生統計所有 SQL 查詢，回傳統計字典（供 /stats 與 /stats/export 共用）。"""

    def _pct_value(num: int, den: int) -> float:
        return round(num / den * 100, 1) if den else 0

    base_filters = _dataset_scope_filters(dataset_scope)
    ch_cond = _chuannian_sql_cond()
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    enrolled_case = case((RecruitmentVisit.enrolled == True, 1), else_=0)
    transfer_case = case((RecruitmentVisit.transfer_term == True, 1), else_=0)
    pending_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.enrolled == False,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )
    effective_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )
    ch_case = case((ch_cond, 1), else_=0)
    ch_dep_case = case(
        (and_(ch_cond, RecruitmentVisit.has_deposit == True), 1), else_=0
    )

    # ── 1. 整體 KPI（單次查詢）──────────────────────────────────
    kpi_query = session.query(
        func.count(RecruitmentVisit.id),
        func.sum(dep_case),
        func.sum(enrolled_case),
        func.sum(transfer_case),
        func.sum(pending_dep_case),
        func.sum(effective_dep_case),
        func.sum(ch_case),
        func.sum(ch_dep_case),
    )
    if base_filters:
        kpi_query = kpi_query.filter(*base_filters)
    kpi = kpi_query.one()
    (
        total_visit,
        total_deposit,
        total_enrolled,
        total_transfer_term,
        total_pending_deposit,
        total_effective_deposit,
        chuannian_visit,
        chuannian_deposit,
    ) = (
        kpi[0] or 0,
        kpi[1] or 0,
        kpi[2] or 0,
        kpi[3] or 0,
        kpi[4] or 0,
        kpi[5] or 0,
        kpi[6] or 0,
        kpi[7] or 0,
    )

    # ── 2. 唯一幼生 ─────────────────────────────────────────────
    unique_key = (
        func.coalesce(RecruitmentVisit.child_name, "")
        + "|"
        + func.coalesce(cast(RecruitmentVisit.birthday, String), "")
    )
    dep_unique_key = case(
        (RecruitmentVisit.has_deposit == True, unique_key), else_=None
    )
    uq_query = session.query(
        func.count(func.distinct(unique_key)),
        func.count(func.distinct(dep_unique_key)),
    )
    if base_filters:
        uq_query = uq_query.filter(*base_filters)
    uq_row = uq_query.one()
    unique_visit = uq_row[0] or 0
    unique_deposit = uq_row[1] or 0

    # ── 3. 月度統計 ─────────────────────────────────────────────
    monthly_query = session.query(
        func.coalesce(RecruitmentVisit.month, "未知").label("month"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
        func.sum(enrolled_case).label("enrolled"),
        func.sum(transfer_case).label("transfer_term"),
        func.sum(pending_dep_case).label("pending_deposit"),
        func.sum(effective_dep_case).label("effective_deposit"),
        func.sum(ch_case).label("chuannian_visit"),
        func.sum(ch_dep_case).label("chuannian_deposit"),
    )
    if base_filters:
        monthly_query = monthly_query.filter(*base_filters)
    monthly_rows = monthly_query.group_by(RecruitmentVisit.month).all()

    monthly = sorted(
        [
            {
                "month": r.month,
                "visit": r.visit or 0,
                "deposit": r.deposit or 0,
                "enrolled": r.enrolled or 0,
                "transfer_term": r.transfer_term or 0,
                "pending_deposit": r.pending_deposit or 0,
                "effective_deposit": r.effective_deposit or 0,
                "visit_to_deposit_rate": _pct_value(r.deposit or 0, r.visit or 0),
                "visit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.visit or 0),
                "deposit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.deposit or 0),
                "effective_to_enrolled_rate": _pct_value(
                    r.enrolled or 0, r.effective_deposit or 0
                ),
                "chuannian_visit": r.chuannian_visit or 0,
                "chuannian_deposit": r.chuannian_deposit or 0,
            }
            for r in monthly_rows
        ],
        key=lambda item: _roc_month_sort_key(item["month"]),
    )

    # ── 3b. 年度統計 ────────────────────────────────────────────
    yearly_map: dict[str, dict] = {}
    for row in monthly:
        month_label = row["month"]
        if month_label in (None, "", "未知") or "." not in month_label:
            continue
        year = month_label.split(".", 1)[0]
        bucket = yearly_map.setdefault(
            year,
            {
                "year": year,
                "visit": 0,
                "deposit": 0,
                "enrolled": 0,
                "transfer_term": 0,
                "pending_deposit": 0,
                "effective_deposit": 0,
                "chuannian_visit": 0,
                "chuannian_deposit": 0,
            },
        )
        for key in (
            "visit",
            "deposit",
            "enrolled",
            "transfer_term",
            "pending_deposit",
            "effective_deposit",
            "chuannian_visit",
            "chuannian_deposit",
        ):
            bucket[key] += row[key]

    by_year = []
    for year in sorted(
        yearly_map.keys(), key=lambda value: int(value) if value.isdigit() else 999999
    ):
        bucket = yearly_map[year]
        by_year.append(
            {
                **bucket,
                "visit_to_deposit_rate": _pct_value(bucket["deposit"], bucket["visit"]),
                "visit_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["visit"]
                ),
                "deposit_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["deposit"]
                ),
                "effective_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["effective_deposit"]
                ),
            }
        )

    # ── 4. 班別統計 ─────────────────────────────────────────────
    grade_query = session.query(
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
        func.sum(enrolled_case).label("enrolled"),
    )
    if base_filters:
        grade_query = grade_query.filter(*base_filters)
    grade_rows = (
        grade_query.group_by(RecruitmentVisit.grade)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_grade = [
        {
            "grade": r.grade,
            "visit": r.visit or 0,
            "deposit": r.deposit or 0,
            "enrolled": r.enrolled or 0,
            "visit_to_deposit_rate": _pct_value(r.deposit or 0, r.visit or 0),
            "visit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.visit or 0),
            "deposit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.deposit or 0),
        }
        for r in grade_rows
    ]

    # ── 5. 月份 × 班別 ───────────────────────────────────────────
    mg_query = session.query(
        func.coalesce(RecruitmentVisit.month, "未知").label("month"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        mg_query = mg_query.filter(*base_filters)
    mg_rows = mg_query.group_by(RecruitmentVisit.month, RecruitmentVisit.grade).all()

    month_grade: dict = {}
    for r in mg_rows:
        m = r.month
        if m not in month_grade:
            month_grade[m] = {}
        month_grade[m][r.grade] = r.cnt
        month_grade[m]["合計"] = month_grade[m].get("合計", 0) + r.cnt

    # ── 6. 來源統計 ──────────────────────────────────────────────
    source_group_expr = _source_group_sql_expr().label("source")
    source_query = session.query(
        source_group_expr,
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        source_query = source_query.filter(*base_filters)
    source_rows = (
        source_query.group_by(source_group_expr)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_source = sorted(
        [
            {
                "source": row.source,
                "visit": row.visit or 0,
                "deposit": row.deposit or 0,
            }
            for row in source_rows
        ],
        key=lambda item: (-item["visit"], -item["deposit"], item["source"]),
    )

    # ── 7. 接待人員 × 各年級 ─────────────────────────────────────
    ref_grade_query = session.query(
        func.coalesce(RecruitmentVisit.referrer, "未填寫").label("referrer"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ref_grade_query = ref_grade_query.filter(*base_filters)
    ref_grade_rows = ref_grade_query.group_by(
        RecruitmentVisit.referrer, RecruitmentVisit.grade
    ).all()

    by_referrer: dict = {}
    for r in ref_grade_rows:
        ref = r.referrer
        if ref not in by_referrer:
            by_referrer[ref] = {
                "referrer": ref,
                "visit": 0,
                "deposit": 0,
                "by_grade": {},
            }
        by_referrer[ref]["visit"] += r.visit or 0
        by_referrer[ref]["deposit"] += r.deposit or 0
        by_referrer[ref]["by_grade"][r.grade] = {
            "visit": r.visit or 0,
            "deposit": r.deposit or 0,
        }

    by_referrer_list = sorted(by_referrer.values(), key=lambda x: -x["visit"])

    # ── 8. 接待者 × 來源 交叉表 ──────────────────────────────────
    cross_source_expr = _source_group_sql_expr().label("source")
    cross_query = session.query(
        func.coalesce(RecruitmentVisit.referrer, "未填寫").label("referrer"),
        cross_source_expr,
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        cross_query = cross_query.filter(*base_filters)
    cross_qrows = cross_query.group_by(
        RecruitmentVisit.referrer, cross_source_expr
    ).all()

    _cross_raw: dict = {}
    for r in cross_qrows:
        if r.referrer not in _cross_raw:
            _cross_raw[r.referrer] = {}
        _cross_raw[r.referrer][r.source] = _cross_raw[r.referrer].get(r.source, 0) + (
            r.cnt or 0
        )

    top_source_names = [item["source"] for item in by_source[:TOP_SOURCES_COUNT]]

    cross_rows_out = sorted(
        [
            {
                "referrer": ref,
                "sources": {s: _cross_raw[ref].get(s, 0) for s in top_source_names},
                "total": sum(_cross_raw[ref].values()),
            }
            for ref in _cross_raw
        ],
        key=lambda x: -x["total"],
    )
    referrer_source_cross = {"referrers": cross_rows_out, "sources": top_source_names}

    # ── 9. 行政區統計 ────────────────────────────────────────────
    district_query = session.query(
        func.coalesce(RecruitmentVisit.district, "未填寫").label("district"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        district_query = district_query.filter(*base_filters)
    district_rows = (
        district_query.group_by(RecruitmentVisit.district)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_district = [
        {"district": r.district, "visit": r.visit or 0, "deposit": r.deposit or 0}
        for r in district_rows
    ]

    # ── 10. 未預繳原因 ──────────────────────────────────────────
    reason_query = session.query(
        func.coalesce(RecruitmentVisit.no_deposit_reason, "未分類").label("reason"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        reason_query = reason_query.filter(*base_filters)
    reason_rows = (
        reason_query.filter(RecruitmentVisit.has_deposit == False)
        .group_by(RecruitmentVisit.no_deposit_reason, RecruitmentVisit.grade)
        .all()
    )

    no_deposit_total_query = session.query(func.count(RecruitmentVisit.id))
    if base_filters:
        no_deposit_total_query = no_deposit_total_query.filter(*base_filters)
    no_deposit_total = (
        no_deposit_total_query.filter(RecruitmentVisit.has_deposit == False).scalar()
        or 0
    )

    reason_stats: dict = {}
    for r in reason_rows:
        if r.reason not in reason_stats:
            reason_stats[r.reason] = {"reason": r.reason, "count": 0, "by_grade": {}}
        reason_stats[r.reason]["count"] += r.cnt
        reason_stats[r.reason]["by_grade"][r.grade] = r.cnt

    no_deposit_reasons = sorted(reason_stats.values(), key=lambda x: -x["count"])

    def _expected_sort_key(x: dict):
        label = x["expected_month"]
        return (1, "") if label == "未知" else (0, label)

    # ── 11. 童年綠地 by expected label ──────────────────────────
    ch_expected_query = session.query(
        func.coalesce(RecruitmentVisit.expected_start_label, "未知").label(
            "expected_month"
        ),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ch_expected_query = ch_expected_query.filter(*base_filters)
    ch_expected_rows = (
        ch_expected_query.filter(ch_cond)
        .group_by(RecruitmentVisit.expected_start_label)
        .all()
    )

    chuannian_by_expected_list = sorted(
        [
            {
                "expected_month": r.expected_month,
                "visit": r.visit or 0,
                "deposit": r.deposit or 0,
            }
            for r in ch_expected_rows
        ],
        key=_expected_sort_key,
    )

    # ── 12. 童年綠地各班別 ──────────────────────────────────────
    ch_grade_query = session.query(
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ch_grade_query = ch_grade_query.filter(*base_filters)
    ch_grade_rows = (
        ch_grade_query.filter(ch_cond).group_by(RecruitmentVisit.grade).all()
    )

    chuannian_by_grade = sorted(
        [
            {"grade": r.grade, "visit": r.visit or 0, "deposit": r.deposit or 0}
            for r in ch_grade_rows
        ],
        key=lambda x: -x["visit"],
    )

    monthly_map = {
        _safe_normalize_roc_month(item["month"]) or item["month"]: item
        for item in monthly
    }
    resolved_reference_month = _select_reference_month(monthly, reference_month)
    previous_month = (
        _shift_roc_month(resolved_reference_month, -1)
        if resolved_reference_month
        else None
    )
    rolling_30d = _aggregate_snapshot(
        session,
        *base_filters,
        RecruitmentVisit.created_at >= datetime.now() - timedelta(days=30),
    )
    rolling_90d = _aggregate_snapshot(
        session,
        *base_filters,
        RecruitmentVisit.created_at >= datetime.now() - timedelta(days=90),
    )
    current_month_snapshot = (
        _metric_snapshot(
            visit=monthly_map.get(resolved_reference_month, {}).get("visit", 0),
            deposit=monthly_map.get(resolved_reference_month, {}).get("deposit", 0),
            enrolled=monthly_map.get(resolved_reference_month, {}).get("enrolled", 0),
            transfer_term=monthly_map.get(resolved_reference_month, {}).get(
                "transfer_term", 0
            ),
            pending_deposit=monthly_map.get(resolved_reference_month, {}).get(
                "pending_deposit", 0
            ),
            effective_deposit=monthly_map.get(resolved_reference_month, {}).get(
                "effective_deposit", 0
            ),
        )
        if resolved_reference_month
        else _empty_snapshot()
    )
    month_over_month = _build_month_over_month(
        resolved_reference_month, previous_month, monthly_map
    )

    overdue_cutoff = datetime.now() - timedelta(days=DEFAULT_OVERDUE_DAYS)
    high_potential_backlog_count = (
        session.query(func.count(RecruitmentVisit.id))
        .filter(
            *base_filters,
            RecruitmentVisit.has_deposit == False,
            RecruitmentVisit.no_deposit_reason.in_(
                tuple(HIGH_PRIORITY_NO_DEPOSIT_REASONS)
            ),
            RecruitmentVisit.created_at <= overdue_cutoff,
        )
        .scalar()
        or 0
    )
    source_imbalance = _find_source_imbalance(
        session, datetime.now() - timedelta(days=90), *base_filters
    )

    dominant_district_row = (
        session.query(
            func.coalesce(RecruitmentVisit.district, "未填寫").label("district"),
            func.count(RecruitmentVisit.id).label("visit"),
        )
        .filter(*base_filters, RecruitmentVisit.month == resolved_reference_month)
        .group_by(RecruitmentVisit.district)
        .order_by(
            func.count(RecruitmentVisit.id).desc(),
            func.coalesce(RecruitmentVisit.district, "未填寫"),
        )
        .first()
        if resolved_reference_month
        else None
    )
    dominant_district = (
        dominant_district_row.district if dominant_district_row else None
    )
    alerts = _build_alerts(
        month_over_month=month_over_month,
        high_potential_backlog_count=high_potential_backlog_count,
        source_imbalance=source_imbalance,
        reference_month=resolved_reference_month,
    )
    top_action_queue = _build_action_queue(
        current_month=resolved_reference_month,
        high_potential_backlog_count=high_potential_backlog_count,
        dominant_district=dominant_district,
        source_imbalance=source_imbalance,
    )

    return {
        "total_visit": total_visit,
        "total_deposit": total_deposit,
        "total_enrolled": total_enrolled,
        "total_transfer_term": total_transfer_term,
        "total_pending_deposit": total_pending_deposit,
        "total_effective_deposit": total_effective_deposit,
        "unique_visit": unique_visit,
        "unique_deposit": unique_deposit,
        "visit_to_deposit_rate": _pct_value(total_deposit, total_visit),
        "visit_to_enrolled_rate": _pct_value(total_enrolled, total_visit),
        "deposit_to_enrolled_rate": _pct_value(total_enrolled, total_deposit),
        "effective_to_enrolled_rate": _pct_value(
            total_enrolled, total_effective_deposit
        ),
        "chuannian_visit": chuannian_visit,
        "chuannian_deposit": chuannian_deposit,
        "monthly": monthly,
        "by_grade": by_grade,
        "month_grade": month_grade,
        "by_source": by_source,
        "by_referrer": by_referrer_list,
        "referrer_source_cross": referrer_source_cross,
        "top_source_names": top_source_names,
        "by_district": by_district,
        "no_deposit_reasons": no_deposit_reasons,
        "no_deposit_total": no_deposit_total,
        "chuannian_by_expected": chuannian_by_expected_list,
        "chuannian_by_grade": chuannian_by_grade,
        "by_year": by_year,
        "reference_month": resolved_reference_month,
        "decision_summary": {
            "current_month": current_month_snapshot,
            "rolling_30d": rolling_30d,
            "rolling_90d": rolling_90d,
            "ytd": _build_ytd_snapshot(resolved_reference_month, monthly_map),
        },
        "funnel_snapshot": {
            field: current_month_snapshot[field]
            for field in (
                "visit",
                "deposit",
                "enrolled",
                "transfer_term",
                "effective_deposit",
                "pending_deposit",
            )
        },
        "month_over_month": month_over_month,
        "alerts": alerts,
        "top_action_queue": top_action_queue,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/stats")
def get_recruitment_stats(
    reference_month: Optional[str] = None,
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """完整統計匯總（全 SQL GROUP BY，效能最佳化版）"""
    with session_scope() as session:
        return _query_stats(
            session, reference_month=reference_month, dataset_scope=dataset_scope
        )


# Excel 匯出樣式
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=13)
_CENTER = Alignment(horizontal="center")


def _hrow(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER


def _pct(num: int, den: int) -> str:
    return f"{num / den * 100:.1f}%" if den else "—"


@router.get("/stats/export")
def export_recruitment_stats(
    reference_month: Optional[str] = None,
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """匯出招生統計 Excel（多頁簽）"""
    with session_scope() as session:
        normalized_scope = _normalize_dataset_scope(dataset_scope)
        s = _query_stats(
            session, reference_month=reference_month, dataset_scope=normalized_scope
        )

    wb = Workbook()

    # ── Sheet 1：決策摘要 ─────────────────────────────────────────
    ws = wb.active
    ws.title = "決策摘要"
    ws.append(["招生決策摘要"])
    ws["A1"].font = _TITLE_FONT
    ws.append([f"參考月份：{s['reference_month'] or '無資料'}"])
    ws.append([])
    _hrow(ws, 4, ["區塊", "指標", "數值"])
    summary_rows = []
    summary_label_map = {
        "current_month": "本月",
        "rolling_30d": "近 30 天",
        "rolling_90d": "近 90 天",
        "ytd": "年度累計",
    }
    for key, label in summary_label_map.items():
        snapshot = s["decision_summary"][key]
        summary_rows.extend(
            [
                (label, "參觀", snapshot["visit"]),
                (label, "預繳", snapshot["deposit"]),
                (label, "註冊", snapshot["enrolled"]),
                (label, "參觀→預繳率", _pct(snapshot["deposit"], snapshot["visit"])),
                (label, "參觀→註冊率", _pct(snapshot["enrolled"], snapshot["visit"])),
            ]
        )
    for row in summary_rows:
        ws.append(list(row))

    ws.append([])
    month_over_month = s["month_over_month"]
    ws.append(["月比觀察", "本月", month_over_month["current_month"] or "—"])
    ws.append(["月比觀察", "上月", month_over_month["previous_month"] or "—"])
    ws.append(
        [
            "月比觀察",
            "參觀→預繳率變化",
            f"{month_over_month['visit_to_deposit_rate']['delta']:.1f} 個百分點",
        ]
    )
    ws.append(
        [
            "月比觀察",
            "參觀→註冊率變化",
            f"{month_over_month['visit_to_enrolled_rate']['delta']:.1f} 個百分點",
        ]
    )

    ws.append([])
    _hrow(ws, ws.max_row + 1, ["警示代碼", "等級", "標題", "說明"])
    if s["alerts"]:
        for alert in s["alerts"]:
            ws.append([alert["code"], alert["level"], alert["title"], alert["message"]])
    else:
        ws.append(["—", "info", "目前無警示", "本期未觸發決策警示。"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48

    # ── Sheet 2：總覽 KPI ─────────────────────────────────────────
    ws = wb.create_sheet("總覽")
    ws.append(["招生統計總覽"])
    ws["A1"].font = _TITLE_FONT
    ws.append([])
    _hrow(ws, 3, ["指標", "數值"])
    kpi_rows = [
        ("總參觀紀錄", s["total_visit"]),
        ("唯一幼生數", s["unique_visit"]),
        ("總預繳人數", s["total_deposit"]),
        ("總註冊人數", s["total_enrolled"]),
        ("轉其他學期", s["total_transfer_term"]),
        ("預繳未註冊", s["total_pending_deposit"]),
        ("有效預繳", s["total_effective_deposit"]),
        ("唯一幼生預繳數", s["unique_deposit"]),
        ("參觀→預繳率", _pct(s["total_deposit"], s["total_visit"])),
        ("參觀→註冊率", _pct(s["total_enrolled"], s["total_visit"])),
        ("預繳→註冊率", _pct(s["total_enrolled"], s["total_deposit"])),
        ("排除轉期→註冊率", _pct(s["total_enrolled"], s["total_effective_deposit"])),
        ("唯一幼生預繳率", _pct(s["unique_deposit"], s["unique_visit"])),
        ("童年綠地參觀人數", s["chuannian_visit"]),
        ("童年綠地預繳人數", s["chuannian_deposit"]),
        ("童年綠地預繳率", _pct(s["chuannian_deposit"], s["chuannian_visit"])),
    ]
    for row in kpi_rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # ── Sheet 3：月度明細 ─────────────────────────────────────────
    ws2 = wb.create_sheet("月度明細")
    _hrow(
        ws2,
        1,
        [
            "月份",
            "參觀人數",
            "預繳人數",
            "註冊人數",
            "轉其他學期",
            "有效預繳",
            "預繳未註冊",
            "參觀→預繳率",
            "參觀→註冊率",
            "預繳→註冊率",
            "排除轉期→註冊率",
            "童年綠地參觀",
            "童年綠地預繳",
            "童年綠地預繳率",
        ],
    )
    for r in s["monthly"]:
        ws2.append(
            [
                r["month"],
                r["visit"],
                r["deposit"],
                r["enrolled"],
                r["transfer_term"],
                r["effective_deposit"],
                r["pending_deposit"],
                _pct(r["deposit"], r["visit"]),
                _pct(r["enrolled"], r["visit"]),
                _pct(r["enrolled"], r["deposit"]),
                _pct(r["enrolled"], r["effective_deposit"]),
                r["chuannian_visit"],
                r["chuannian_deposit"],
                _pct(r["chuannian_deposit"], r["chuannian_visit"]),
            ]
        )
    for col_letter, width in zip(
        "ABCDEFGHIJKLMN", [10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 14, 12, 12, 14]
    ):
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 4：班別分析 ─────────────────────────────────────────
    ws3 = wb.create_sheet("班別分析")
    _hrow(ws3, 1, ["班別", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_grade"]:
        ws3.append(
            [r["grade"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    for col_letter, width in zip("ABCD", [10, 10, 10, 10]):
        ws3.column_dimensions[col_letter].width = width

    # ── Sheet 5：來源分析 ─────────────────────────────────────────
    ws4 = wb.create_sheet("來源分析")
    _hrow(ws4, 1, ["來源", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_source"]:
        ws4.append(
            [r["source"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws4.column_dimensions["A"].width = 20
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws4.column_dimensions[col_letter].width = width

    # ── Sheet 6：接待人員 ─────────────────────────────────────────
    ws5 = wb.create_sheet("接待人員")
    _hrow(ws5, 1, ["接待人員", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_referrer"]:
        ws5.append(
            [r["referrer"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws5.column_dimensions["A"].width = 16
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws5.column_dimensions[col_letter].width = width

    # ── Sheet 7：行政區 ───────────────────────────────────────────
    ws6 = wb.create_sheet("行政區")
    _hrow(ws6, 1, ["行政區", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_district"]:
        ws6.append(
            [r["district"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws6.column_dimensions["A"].width = 14
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws6.column_dimensions[col_letter].width = width

    # ── Sheet 8：未預繳原因 ───────────────────────────────────────
    ws7 = wb.create_sheet("未預繳原因")
    _hrow(ws7, 1, ["原因", "人數"])
    for r in s["no_deposit_reasons"]:
        ws7.append([r["reason"], r["count"]])
    ws7.append(["（合計）", s["no_deposit_total"]])
    ws7.column_dimensions["A"].width = 28
    ws7.column_dimensions["B"].width = 10

    # ── Sheet 9：年度統計 ─────────────────────────────────────────
    ws8 = wb.create_sheet("年度統計")
    _hrow(
        ws8,
        1,
        [
            "年份",
            "參觀人數",
            "預繳人數",
            "註冊人數",
            "轉其他學期",
            "有效預繳",
            "預繳未註冊",
            "參觀→預繳率",
            "參觀→註冊率",
            "預繳→註冊率",
            "排除轉期→註冊率",
            "童年綠地參觀",
            "童年綠地預繳",
        ],
    )
    for r in s["by_year"]:
        ws8.append(
            [
                f"{r['year']}年",
                r["visit"],
                r["deposit"],
                r["enrolled"],
                r["transfer_term"],
                r["effective_deposit"],
                r["pending_deposit"],
                _pct(r["deposit"], r["visit"]),
                _pct(r["enrolled"], r["visit"]),
                _pct(r["enrolled"], r["deposit"]),
                _pct(r["enrolled"], r["effective_deposit"]),
                r["chuannian_visit"],
                r["chuannian_deposit"],
            ]
        )
    for col_letter, width in zip(
        "ABCDEFGHIJKLM", [10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 14, 12, 12]
    ):
        ws8.column_dimensions[col_letter].width = width

    filename = "招生統計.xlsx"
    return xlsx_streaming_response(wb, filename)


@router.get("/no-deposit-analysis")
def get_no_deposit_analysis(
    reason: Optional[str] = Query(None),
    grade: Optional[str] = Query(None),
    priority: Optional[str] = Query(None, pattern="^(high|medium|low)$"),
    overdue_days: Optional[int] = Query(None, ge=1, le=365),
    cold_only: Optional[bool] = Query(None),
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """未預繳名單明細（含原因分類篩選，支援分頁）"""
    with session_scope() as session:
        q = _build_scoped_query(session, dataset_scope).filter(
            RecruitmentVisit.has_deposit == False
        )
        if reason:
            q = q.filter(RecruitmentVisit.no_deposit_reason == reason)
        if grade:
            q = q.filter(RecruitmentVisit.grade == grade)
        base_query = q
        effective_overdue_days = overdue_days or DEFAULT_OVERDUE_DAYS
        overdue_cutoff = datetime.now() - timedelta(days=effective_overdue_days)
        cold_cutoff = datetime.now() - timedelta(days=COLD_LEAD_DAYS)
        summary = {
            "high_potential_count": (
                base_query.filter(
                    RecruitmentVisit.no_deposit_reason.in_(
                        tuple(HIGH_PRIORITY_NO_DEPOSIT_REASONS)
                    )
                ).count()
            ),
            "overdue_followup_count": base_query.filter(
                RecruitmentVisit.created_at <= overdue_cutoff
            ).count(),
            "cold_count": base_query.filter(
                RecruitmentVisit.created_at <= cold_cutoff
            ).count(),
        }
        if priority:
            q = q.filter(
                RecruitmentVisit.no_deposit_reason.in_(
                    tuple(NO_DEPOSIT_PRIORITY_REASON_MAP[priority])
                )
            )
        if overdue_days is not None:
            q = q.filter(RecruitmentVisit.created_at <= overdue_cutoff)
        if cold_only is True:
            q = q.filter(RecruitmentVisit.created_at <= cold_cutoff)
        total = q.count()
        records = (
            q.order_by(RecruitmentVisit.month.desc(), RecruitmentVisit.seq_no)
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "summary": summary,
            "records": [_to_dict(r) for r in records],
        }
