"""tests/test_activity_security_fixes.py — 本次安全審查修復的回歸測試。

涵蓋：
- AddPaymentRequest.amount 上限（le=999999）
- add_registration_payment 退費超過已繳金額 → 400（原本靜默 max(0,...)）
- batch_update_attendance 過濾 is_active=False / match_status='rejected'
- /public/inquiries rate limit（3/60s）
- export_payment_report / export_registrations 筆數超過 MAX_EXPORT_ROWS → 413
"""

import os
import sys
from datetime import date

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.activity import router as activity_router
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from models.database import (
    ActivityAttendance,
    ActivityCourse,
    ActivityPaymentRecord,
    ActivityRegistration,
    ActivitySession,
    ActivitySupply,
    Base,
    Classroom,  # noqa: F401 metadata
    RegistrationCourse,
    RegistrationSupply,
    User,
)
from utils.auth import hash_password
from utils.permissions import Permission


@pytest.fixture
def client(tmp_path):
    db_path = tmp_path / "sec.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    # 清空 public / registrations / pos 模組的 limiter 計數，避免跨測試污染
    from api.activity import public as public_mod
    from api.activity import registrations as reg_mod
    from api.activity import pos as pos_mod

    for mod in (public_mod, reg_mod, pos_mod):
        for attr in dir(mod):
            obj = getattr(mod, attr)
            if hasattr(obj, "_timestamps"):
                obj._timestamps.clear()

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(activity_router)

    with TestClient(app) as c:
        yield c, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _admin(session, username="sec_admin"):
    # 補 ACTIVITY_PAYMENT_APPROVE：新 guard 3 (diff verify) 在 _make_reg sessions=NULL 時
    # fallback 為全退建議；TestRefundExceedsPaid 用部分退費，diff > 100 會被 guard 3 先擋；
    # 本 test 目的是測 guard 1（超額 400）與 guard 4（精確退費 200），故授予 approver 略過 guard 3。
    user = User(
        username=username,
        password_hash=hash_password("Temp123456"),
        role="admin",
        permission_names=[
            "ACTIVITY_READ",
            "ACTIVITY_WRITE",
            "ACTIVITY_PAYMENT_APPROVE",
        ],
        is_active=True,
    )
    session.add(user)
    session.flush()
    return user


def _login(c: TestClient, username="sec_admin"):
    return c.post(
        "/api/auth/login", json={"username": username, "password": "Temp123456"}
    )


_course_counter = {"n": 0}


def _make_reg(
    session,
    *,
    paid_amount=0,
    match_status="matched",
    is_active=True,
    student_name="王小明",
    class_name="大班",
):
    from utils.academic import resolve_current_academic_term

    sy, sem = resolve_current_academic_term()
    _course_counter["n"] += 1
    course = ActivityCourse(
        name=f"課程_{_course_counter['n']}",
        price=1000,
        capacity=30,
        school_year=sy,
        semester=sem,
    )
    session.add(course)
    session.flush()
    reg = ActivityRegistration(
        student_name=student_name,
        birthday="2020-01-01",
        class_name=class_name,
        paid_amount=paid_amount,
        is_active=is_active,
        match_status=match_status,
        school_year=sy,
        semester=sem,
    )
    session.add(reg)
    session.flush()
    session.add(
        RegistrationCourse(
            registration_id=reg.id,
            course_id=course.id,
            status="enrolled",
            price_snapshot=1000,
        )
    )
    session.flush()
    return reg, course


# ── 1. AddPaymentRequest.amount 上限（Pydantic le=999999） ─────────────────


class TestPaymentAmountUpperBound:
    def test_amount_above_limit_rejected(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            reg, _ = _make_reg(s)
            s.commit()
            reg_id = reg.id

        assert _login(c).status_code == 200
        res = c.post(
            f"/api/activity/registrations/{reg_id}/payments",
            json={
                "type": "payment",
                "amount": 1_000_000,  # 超過上限 999999
                "payment_date": date.today().isoformat(),
                "payment_method": "現金",
                "notes": "",
            },
        )
        assert res.status_code == 422

    def test_amount_at_limit_accepted(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            reg, _ = _make_reg(s)
            s.commit()
            reg_id = reg.id

        assert _login(c).status_code == 200
        res = c.post(
            f"/api/activity/registrations/{reg_id}/payments",
            json={
                "type": "payment",
                "amount": 999_999,
                "payment_date": date.today().isoformat(),
                "payment_method": "現金",
            },
        )
        assert res.status_code in (200, 201), res.text


# ── 2. add_registration_payment 退費超額檢驗 ───────────────────────────────


class TestRefundExceedsPaid:
    def test_refund_over_paid_returns_400(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            reg, _ = _make_reg(s, paid_amount=500)
            s.commit()
            reg_id = reg.id

        assert _login(c).status_code == 200
        res = c.post(
            f"/api/activity/registrations/{reg_id}/payments",
            json={
                "type": "refund",
                "amount": 600,
                "payment_date": date.today().isoformat(),
                "payment_method": "現金",
                "notes": "退費原因：超額測試需家長同意辦理",
            },
        )
        assert res.status_code == 400
        assert "超過已繳金額" in res.json()["detail"]

        # 確認 paid_amount 未被改動（原本 bug 會靜默置 0）
        with sf() as s:
            reg = s.query(ActivityRegistration).filter_by(id=reg_id).one()
            assert reg.paid_amount == 500
            assert s.query(ActivityPaymentRecord).count() == 0

    def test_refund_exact_paid_amount_ok(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            reg, _ = _make_reg(s, paid_amount=500)
            s.commit()
            reg_id = reg.id

        assert _login(c).status_code == 200
        res = c.post(
            f"/api/activity/registrations/{reg_id}/payments",
            json={
                "type": "refund",
                "amount": 500,
                "payment_date": date.today().isoformat(),
                "payment_method": "現金",
                "notes": "客戶取消報名退回（家長申請辦理）",
            },
        )
        assert res.status_code in (200, 201), res.text
        with sf() as s:
            reg = s.query(ActivityRegistration).filter_by(id=reg_id).one()
            assert reg.paid_amount == 0


# ── 3. batch_update_attendance 過濾無效報名 ────────────────────────────────


class TestAttendanceFiltersInvalidRegs:
    def test_skips_inactive_and_rejected(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            reg_ok, course = _make_reg(s, match_status="matched", is_active=True)
            reg_inactive, _ = _make_reg(s, is_active=False)
            reg_rejected, _ = _make_reg(s, match_status="rejected", is_active=True)
            sess = ActivitySession(
                course_id=course.id,
                session_date=date.today(),
                created_by="test",
            )
            s.add(sess)
            s.flush()
            s.commit()
            session_id = sess.id
            ok_id = reg_ok.id
            inactive_id = reg_inactive.id
            rejected_id = reg_rejected.id

        assert _login(c).status_code == 200
        res = c.put(
            f"/api/activity/attendance/sessions/{session_id}/records",
            json={
                "records": [
                    {"registration_id": ok_id, "is_present": True, "notes": ""},
                    {"registration_id": inactive_id, "is_present": True, "notes": ""},
                    {"registration_id": rejected_id, "is_present": True, "notes": ""},
                ]
            },
        )
        assert res.status_code in (200, 201), res.text
        data = res.json()
        assert data["updated"] == 1
        assert data["skipped"] == 2

        with sf() as s:
            atts = s.query(ActivityAttendance).all()
            assert len(atts) == 1
            assert atts[0].registration_id == ok_id


# ── 4. /public/inquiries rate limit ────────────────────────────────────────


class TestInquiryRateLimit:
    def test_fourth_request_within_window_returns_429(self, client):
        c, _ = client
        payload = {
            "name": "家長",
            "phone": "0912345678",
            "question": "請問有什麼課程？",
        }
        for _ in range(3):
            res = c.post("/api/activity/public/inquiries", json=payload)
            assert res.status_code == 201, res.text
        res = c.post("/api/activity/public/inquiries", json=payload)
        assert res.status_code == 429


# ── 5. export 筆數上限 ─────────────────────────────────────────────────────


class TestExportRowLimit:
    def test_export_rejects_when_exceeding_cap(self, client, monkeypatch):
        c, sf = client
        # 把 MAX_EXPORT_ROWS 暫時調小以避免要建 5001 筆
        import api.activity.registrations_static as reg_module

        monkeypatch.setattr(reg_module, "MAX_EXPORT_ROWS", 2)

        with sf() as s:
            _admin(s)
            _make_reg(s)
            _make_reg(s)
            _make_reg(s)  # 共 3 筆 > 上限 2
            s.commit()

        assert _login(c).status_code == 200
        res = c.get("/api/activity/registrations/payment-report")
        assert res.status_code == 413
        assert "超過上限" in res.json()["detail"]

    def test_export_allowed_within_cap(self, client, monkeypatch):
        c, sf = client
        import api.activity.registrations_static as reg_module

        monkeypatch.setattr(reg_module, "MAX_EXPORT_ROWS", 10)

        with sf() as s:
            _admin(s)
            _make_reg(s)
            s.commit()

        assert _login(c).status_code == 200
        res = c.get("/api/activity/registrations/payment-report")
        assert res.status_code == 200


# ── 6. payment-report ws1（繳費總覽）Excel 公式注入防護 ────────────────────


class TestPaymentReportFormulaInjection:
    """S1：ws1「繳費總覽」過去裸用 wb.active 未包 SafeWorksheet，
    家長自填 student_name/class_name 帶公式前綴可注入財會端 Excel。"""

    def _export_workbook(self, c):
        import io

        import openpyxl

        res = c.get("/api/activity/registrations/payment-report")
        assert res.status_code == 200, res.text
        return openpyxl.load_workbook(io.BytesIO(res.content))

    def test_ws1_student_name_formula_quoted(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            _make_reg(s, student_name="=2+2", class_name="+CMD()")
            s.commit()

        assert _login(c).status_code == 200
        wb = self._export_workbook(c)
        ws1 = wb["繳費總覽"]
        # row 2 = 第一筆資料列；B=學生、C=班級
        assert ws1.cell(row=2, column=2).value == "'=2+2"
        assert ws1.cell(row=2, column=3).value == "'+CMD()"

    def test_ws1_normal_name_unchanged(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            _make_reg(s, student_name="王小明")
            s.commit()

        assert _login(c).status_code == 200
        wb = self._export_workbook(c)
        ws1 = wb["繳費總覽"]
        assert ws1.cell(row=2, column=2).value == "王小明"


# ── 7. ILIKE 萬用字元跳脫（搜尋 % / _ 不再全表匹配） ───────────────────────


def _staff_with_students_read(session, username="sec_staff"):
    user = User(
        username=username,
        password_hash=hash_password("Temp123456"),
        role="admin",
        permission_names=[
            "ACTIVITY_READ",
            "ACTIVITY_WRITE",
            "STUDENTS_READ",
            "GUARDIANS_READ",
        ],
        is_active=True,
    )
    session.add(user)
    session.flush()
    return user


class TestIlikeWildcardEscape:
    """S2：raw f"%{search}%" 未跳脫 % / _，搜尋 '%' 會全表匹配
    （registrations 列表 / pending 列表 / students/search 三處）。"""

    def test_registrations_list_percent_literal(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            _make_reg(s, student_name="王小明")
            _make_reg(s, student_name="折扣50%生")
            s.commit()

        assert _login(c).status_code == 200
        res = c.get("/api/activity/registrations", params={"search": "%"})
        assert res.status_code == 200, res.text
        names = [it["student_name"] for it in res.json()["items"]]
        assert names == ["折扣50%生"]

    def test_registrations_list_underscore_literal(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            _make_reg(s, student_name="王小明")
            _make_reg(s, student_name="A_B")
            s.commit()

        assert _login(c).status_code == 200
        res = c.get("/api/activity/registrations", params={"search": "_"})
        assert res.status_code == 200, res.text
        names = [it["student_name"] for it in res.json()["items"]]
        assert names == ["A_B"]

    def test_pending_list_percent_literal(self, client):
        c, sf = client
        with sf() as s:
            _admin(s)
            _make_reg(s, student_name="王小明", match_status="pending")
            _make_reg(s, student_name="折扣50%生", match_status="pending")
            s.commit()
        with sf() as s:
            for reg in s.query(ActivityRegistration).all():
                reg.pending_review = True
            s.commit()

        assert _login(c).status_code == 200
        res = c.get(
            "/api/activity/registrations/pending",
            params={"search": "%", "status": "pending"},
        )
        assert res.status_code == 200, res.text
        names = [it["student_name"] for it in res.json()["items"]]
        assert names == ["折扣50%生"]

    def test_admin_search_students_percent_literal(self, client):
        c, sf = client
        from models.database import Student

        with sf() as s:
            _staff_with_students_read(s)
            s.add(
                Student(
                    student_id="S001",
                    name="王小明",
                    is_active=True,
                )
            )
            s.add(
                Student(
                    student_id="S002",
                    name="百分%童",
                    is_active=True,
                )
            )
            s.commit()

        assert _login(c, username="sec_staff").status_code == 200
        res = c.get("/api/activity/students/search", params={"q": "%"})
        assert res.status_code == 200, res.text
        names = [it["name"] for it in res.json()["items"]]
        assert names == ["百分%童"]
