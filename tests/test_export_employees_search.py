"""tests/test_export_employees_search.py — 員工匯出支援 search 篩選測試。

涵蓋：
- GET /api/exports/employees 不帶 search → 匯出全部員工
- GET /api/exports/employees?search=王 → 只匯出名稱含「王」的員工
"""

import io

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import models.base as base_module
from api.auth import router as auth_router
from api.employees import router as employees_router
from api.exports import _export_rate_limit
from api.exports import router as exports_router
from models.base import Base
from tests.test_employees import _login_admin


@pytest.fixture
def exports_client(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'exp.sqlite'}",
        connect_args={"check_same_thread": False},
    )
    sf = sessionmaker(bind=engine)
    old_e, old_s = base_module._engine, base_module._SessionFactory
    base_module._engine, base_module._SessionFactory = engine, sf
    Base.metadata.create_all(engine)

    app = FastAPI()
    from utils.exception_handlers import register_exception_handlers

    register_exception_handlers(app)
    app.include_router(auth_router)
    app.include_router(employees_router)
    app.include_router(exports_router)
    app.dependency_overrides[_export_rate_limit] = lambda: None

    with TestClient(app) as client:
        yield client, sf

    base_module._engine, base_module._SessionFactory = old_e, old_s
    engine.dispose()


def _rows(content: bytes) -> int:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # 員工名冊 Excel 結構：row1=標題、row2=空白、row3=表頭、row4+=資料
    return max(ws.max_row - 3, 0)  # 扣標題+空白+表頭共 3 列


def test_export_employees_search_narrows(exports_client):
    client, sf = exports_client
    _login_admin(client, sf)

    for name in ("王小明", "陳大文"):
        r = client.post(
            "/api/employees",
            json={"name": name, "employee_type": "regular"},
        )
        assert r.status_code == 201, r.text

    full = client.get("/api/exports/employees")
    assert full.status_code == 200

    narrowed = client.get("/api/exports/employees", params={"search": "王"})
    assert narrowed.status_code == 200

    assert _rows(narrowed.content) == 1
    assert _rows(full.content) == 2
