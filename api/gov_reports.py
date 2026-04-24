"""
政府申報格式匯出 API

端點：
    GET /api/gov-reports/labor-insurance   勞保月份投保薪資申報（Excel / TXT）
    GET /api/gov-reports/health-insurance  健保被保險人名冊（Excel）
    GET /api/gov-reports/withholding       國稅局年度薪資所得扣繳憑單（Excel）
    GET /api/gov-reports/pension           勞退月提繳明細（Excel）
"""

import logging
import calendar as cal_module
from datetime import date
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sqlalchemy import or_, and_

from models.database import get_session, Employee, SalaryRecord
from services.salary.insurance_salary import resolve_insurance_salary_raw
from utils.auth import require_staff_permission
from utils.permissions import Permission
from utils.rate_limit import SlidingWindowLimiter
from utils.excel_utils import SafeWorksheet, xlsx_streaming_response

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/gov-reports", tags=["gov-reports"])

_rate_limit = SlidingWindowLimiter(
    max_calls=5,
    window_seconds=60,
    name="gov_report",
    error_detail="申報匯出過於頻繁，請稍後再試",
).as_dependency()

_insurance_service = None


def init_gov_report_services(insurance_service) -> None:
    global _insurance_service
    _insurance_service = insurance_service


# ─────────────────────────────────────────────────────────────────────────────
# Shared Excel Styles
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
_HDR_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=13)
_BOLD = Font(bold=True)
_ITALIC_RED = Font(italic=True, color="FF0000")
_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_FORMULA_PREFIXES = ("=", "+", "-", "@", "|")


def _sanitize(v):
    if not isinstance(v, str):
        return v
    clean = v.lstrip("\t\r\n")
    return ("'" + clean) if clean.startswith(_FORMULA_PREFIXES) else clean


def _hdr(ws, row: int, headers: list) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.border = _THIN
        c.alignment = _CENTER


def _cell(ws, row: int, col: int, value, align=None):
    c = ws.cell(row=row, column=col, value=_sanitize(value))
    c.border = _THIN
    c.alignment = align or _CENTER
    return c


def _total_row(ws, row: int, count: int, totals: dict) -> None:
    """寫入合計列（row=行號，totals={col: value}）"""
    c = ws.cell(row=row, column=1, value="合計")
    c.font = _BOLD
    c.border = _THIN
    c2 = ws.cell(row=row, column=2, value=count)
    c2.border = _THIN
    for col, val in totals.items():
        cx = ws.cell(row=row, column=col, value=val)
        cx.border = _THIN
        cx.font = _BOLD
        cx.alignment = _RIGHT


def _auto_width(ws) -> None:
    from openpyxl.cell.cell import MergedCell

    for col in ws.columns:
        real_cells = [c for c in col if not isinstance(c, MergedCell)]
        if not real_cells:
            continue
        max_len = max((len(str(c.value or "")) for c in real_cells), default=8)
        ws.column_dimensions[real_cells[0].column_letter].width = max(max_len + 2, 10)


def _title_row(ws, text: str, span_end: str) -> None:
    ws.merge_cells(f"A1:{span_end}1")
    tc = ws["A1"]
    tc.value = text
    tc.font = _TITLE_FONT
    tc.alignment = _CENTER


# ─────────────────────────────────────────────────────────────────────────────
# Shared DB Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _active_employees(session, year: int, month: int) -> list:
    """取「當月在職」員工：hire_date <= month_end 且 (resign_date IS NULL OR resign_date >= month_start)。

    Why: 舊版只用 `is_active` 判斷，會造成兩個方向的錯誤申報：
    - 今天在職但當月尚未到職 → 被列入（舊版：is_active=True 分支放過）
    - 歷史當月在職但今天已離職 → 被漏掉（舊版：只有「當月離職」可進第二分支，
      若離職月份晚於計算月份就沒機會被列入）
    改用 hire/resign 日期的時間區間判斷，與 is_active 無關。
    """
    last_day = cal_module.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, last_day)
    return (
        session.query(Employee)
        .filter(
            or_(Employee.hire_date.is_(None), Employee.hire_date <= month_end),
            or_(
                Employee.resign_date.is_(None),
                Employee.resign_date >= month_start,
            ),
        )
        .order_by(Employee.name)
        .all()
    )


def _salary_map(session, emp_ids: list, year: int, month: int) -> dict:
    records = (
        session.query(SalaryRecord)
        .filter(
            SalaryRecord.employee_id.in_(emp_ids),
            SalaryRecord.salary_year == year,
            SalaryRecord.salary_month == month,
        )
        .all()
    )
    return {r.employee_id: r for r in records}


def _resolve_insured(emp: Employee) -> int:
    """用 resolve_insurance_salary_raw 取合法投保基準（時薪員工會套 hourly × 176）。

    舊實作 `insurance_salary_level or base_salary` 對時薪員工 base=0 會短報。
    """
    raw = resolve_insurance_salary_raw(
        employee_type=getattr(emp, "employee_type", "regular") or "regular",
        base_salary=emp.base_salary or 0,
        insurance_salary_level=emp.insurance_salary_level or 0,
        hourly_rate=getattr(emp, "hourly_rate", 0) or 0,
    )
    return int(raw or 0)


def _ins_calc(emp: Employee):
    """使用 InsuranceService fallback 計算（當月無 SalaryRecord 時）"""
    if _insurance_service is None:
        return None
    salary = _resolve_insured(emp)
    return _insurance_service.calculate(
        salary,
        dependents=emp.dependents or 0,
        pension_self_rate=emp.pension_self_rate or 0,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 1. 勞保月份投保薪資申報清單
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/labor-insurance")
def export_labor_insurance(
    year: int = Query(..., ge=2000, le=2100, description="申報年份"),
    month: int = Query(..., ge=1, le=12, description="申報月份"),
    fmt: str = Query("xlsx", description="xlsx（Excel）或 txt（純文字）"),
    employer_name: str = Query("（請填入投保單位名稱）", description="投保單位名稱"),
    employer_code: str = Query("（請填入投保單位代號）", description="投保單位代號"),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
    _: None = Depends(_rate_limit),
):
    """勞工保險月份投保薪資申報清單（Excel 或 TXT）"""
    session = get_session()
    try:
        employees = _active_employees(session, year, month)
        smap = _salary_map(session, [e.id for e in employees], year, month)

        rows = []
        for emp in employees:
            sr = smap.get(emp.id)
            insured = _resolve_insured(emp)

            # 必須「員工端 + 雇主端皆有值」才採用 record；否則走 fallback 重算。
            # Why: 舊版 SalaryRecord 曾漏寫雇主端三欄（labor/health/pension
            # _employer 長期為 0），條件寬鬆只要「員工端 or 雇主端」任一有值就
            # 採用 record → 雇主端就寫成 0 匯出給勞保局。P1-A 修復後新紀錄
            # 雇主端都有值，舊紀錄則透過下方 _ins_calc fallback 重算補齊。
            if (
                sr
                and (sr.labor_insurance_employee or 0) > 0
                and (sr.labor_insurance_employer or 0) > 0
            ):
                labor_emp = round(sr.labor_insurance_employee or 0)
                labor_er = round(sr.labor_insurance_employer or 0)
                # 政府補助 ≈ 總保費 * 10%；員工 20%，雇主 70%，故 gov = employee * 0.5
                labor_gov = round(labor_emp * 0.5)
            else:
                calc = _ins_calc(emp)
                if calc:
                    labor_emp = round(calc.labor_employee)
                    labor_er = round(calc.labor_employer)
                    labor_gov = round(calc.labor_government)
                    insured = int(calc.insured_amount)
                else:
                    labor_emp = labor_er = labor_gov = 0

            rows.append(
                {
                    "name": emp.name,
                    "id_number": emp.id_number or "",
                    "insured": insured,
                    "labor_emp": labor_emp,
                    "labor_er": labor_er,
                    "labor_gov": labor_gov,
                    "hire_date": emp.hire_date.isoformat() if emp.hire_date else "",
                    "resign_date": (
                        emp.resign_date.isoformat() if emp.resign_date else ""
                    ),
                }
            )
    finally:
        session.close()

    logger.warning(
        "勞保申報匯出：%s年%s月，共 %d 人，操作人：%s",
        year,
        month,
        len(rows),
        current_user.get("username", ""),
    )

    if fmt == "txt":
        return _labor_txt(rows, year, month, employer_name, employer_code)
    return _labor_xlsx(rows, year, month, employer_name, employer_code)


def _labor_xlsx(rows, year, month, employer_name, employer_code):
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = f"勞保{year}{month:02d}"

    _title_row(ws, f"勞工保險月份投保薪資申報清單　{year}年{month:02d}月", "I")
    ws["A2"] = f"投保單位：{employer_name}"
    ws["A2"].font = _BOLD
    ws.merge_cells("A2:D2")
    ws["E2"] = f"單位代號：{employer_code}"
    ws["E2"].font = _BOLD
    ws["H2"] = f"申報年月：{year}/{month:02d}"
    ws["H2"].font = _BOLD
    ws.row_dimensions[2].height = 18

    headers = [
        "序號",
        "姓名",
        "身分證字號",
        "月投保薪資",
        "員工自付\n(20%)",
        "雇主負擔\n(70%)",
        "政府補助\n(10%)",
        "到職日期",
        "離職日期",
    ]
    _hdr(ws, 3, headers)
    ws.row_dimensions[3].height = 30

    for i, row in enumerate(rows, 1):
        r = i + 3
        _cell(ws, r, 1, i)
        _cell(ws, r, 2, row["name"], _LEFT)
        _cell(ws, r, 3, row["id_number"])
        _cell(ws, r, 4, row["insured"], _RIGHT)
        _cell(ws, r, 5, row["labor_emp"], _RIGHT)
        _cell(ws, r, 6, row["labor_er"], _RIGHT)
        _cell(ws, r, 7, row["labor_gov"], _RIGHT)
        _cell(ws, r, 8, row["hire_date"])
        _cell(ws, r, 9, row["resign_date"])

    _total_row(
        ws,
        len(rows) + 4,
        len(rows),
        {
            4: sum(r["insured"] for r in rows),
            5: sum(r["labor_emp"] for r in rows),
            6: sum(r["labor_er"] for r in rows),
            7: sum(r["labor_gov"] for r in rows),
        },
    )
    _auto_width(ws)
    return xlsx_streaming_response(wb, f"勞保申報_{year}{month:02d}.xlsx")


def _labor_txt(rows, year, month, employer_name, employer_code):
    """勞保 CSV-style TXT（可上傳至勞保局 e申報系統前參考）"""
    lines = [
        "# 勞工保險月份投保薪資申報清單",
        f"# 投保單位：{employer_name}　代號：{employer_code}",
        f"# 申報年月：{year}年{month:02d}月",
        "# 提醒：本檔案供核對用，實際申報請至勞保局 e申報系統 (eservice.bli.gov.tw)",
        "# 欄位：序號,身分證字號,姓名,月投保薪資,員工自付(20%),雇主負擔(70%),政府補助(10%),到職日,離職日",
        "",
    ]
    for i, row in enumerate(rows, 1):
        lines.append(
            f"{i:04d},"
            f"{row['id_number']:<10},"
            f"{row['name']:<12},"
            f"{row['insured']:>6},"
            f"{row['labor_emp']:>5},"
            f"{row['labor_er']:>5},"
            f"{row['labor_gov']:>5},"
            f"{row['hire_date']},"
            f"{row['resign_date']}"
        )

    content = "\n".join(lines).encode("utf-8-sig")
    filename = f"勞保申報_{year}{month:02d}.txt"
    return StreamingResponse(
        BytesIO(content),
        media_type="text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 2. 健保被保險人名冊
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/health-insurance")
def export_health_insurance(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    employer_name: str = Query("（請填入投保單位名稱）"),
    employer_code: str = Query("（請填入投保單位代號）"),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
    _: None = Depends(_rate_limit),
):
    """全民健康保險被保險人名冊（Excel）"""
    session = get_session()
    try:
        employees = _active_employees(session, year, month)
        smap = _salary_map(session, [e.id for e in employees], year, month)

        rows = []
        for emp in employees:
            sr = smap.get(emp.id)
            insured = _resolve_insured(emp)

            # 同勞保：需員工端 + 雇主端皆有值才採用 record（避免舊資料雇主 0 被信任）
            if (
                sr
                and (sr.health_insurance_employee or 0) > 0
                and (sr.health_insurance_employer or 0) > 0
            ):
                health_emp = round(sr.health_insurance_employee or 0)
                health_er = round(sr.health_insurance_employer or 0)
                calc = _ins_calc(emp)
                insured_amt = int(calc.insured_amount) if calc else insured
            else:
                calc = _ins_calc(emp)
                if calc:
                    health_emp = round(calc.health_employee)
                    health_er = round(calc.health_employer)
                    insured_amt = int(calc.insured_amount)
                else:
                    health_emp = health_er = 0
                    insured_amt = insured

            rows.append(
                {
                    "name": emp.name,
                    "id_number": emp.id_number or "",
                    "birthday": emp.birthday.isoformat() if emp.birthday else "",
                    "dependents": emp.dependents or 0,
                    "insured_amt": insured_amt,
                    "health_emp": health_emp,
                    "health_er": health_er,
                    "hire_date": emp.hire_date.isoformat() if emp.hire_date else "",
                    "resign_date": (
                        emp.resign_date.isoformat() if emp.resign_date else ""
                    ),
                }
            )
    finally:
        session.close()

    logger.warning(
        "健保名冊匯出：%s年%s月，共 %d 人，操作人：%s",
        year,
        month,
        len(rows),
        current_user.get("username", ""),
    )

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = f"健保{year}{month:02d}"

    _title_row(ws, f"全民健康保險被保險人名冊　{year}年{month:02d}月", "I")
    ws["A2"] = f"投保單位：{employer_name}"
    ws["A2"].font = _BOLD
    ws.merge_cells("A2:D2")
    ws["E2"] = f"投保單位代號：{employer_code}"
    ws["E2"].font = _BOLD
    ws.row_dimensions[2].height = 18

    headers = [
        "序號",
        "姓名",
        "身分證字號",
        "出生日期",
        "月投保金額",
        "員工自付\n(30%)",
        "雇主負擔\n(60%)",
        "眷屬人數",
        "在職狀況",
    ]
    _hdr(ws, 3, headers)
    ws.row_dimensions[3].height = 30

    for i, row in enumerate(rows, 1):
        r = i + 3
        _cell(ws, r, 1, i)
        _cell(ws, r, 2, row["name"], _LEFT)
        _cell(ws, r, 3, row["id_number"])
        _cell(ws, r, 4, row["birthday"])
        _cell(ws, r, 5, row["insured_amt"], _RIGHT)
        _cell(ws, r, 6, row["health_emp"], _RIGHT)
        _cell(ws, r, 7, row["health_er"], _RIGHT)
        _cell(ws, r, 8, row["dependents"])
        status = "在職" if not row["resign_date"] else f"離職 {row['resign_date']}"
        _cell(ws, r, 9, status)

    _total_row(
        ws,
        len(rows) + 4,
        len(rows),
        {
            5: sum(r["insured_amt"] for r in rows),
            6: sum(r["health_emp"] for r in rows),
            7: sum(r["health_er"] for r in rows),
            8: sum(r["dependents"] for r in rows),
        },
    )
    _auto_width(ws)
    return xlsx_streaming_response(wb, f"健保被保險人名冊_{year}{month:02d}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 3. 國稅局年度薪資所得扣繳憑單
# ─────────────────────────────────────────────────────────────────────────────

# 2026年(115年)扣繳標準估算常數
# 免稅額(97,000) + 薪資特別扣除額(207,000) + 標準扣除額(124,000) = 428,000
_WITHHOLDING_DEDUCTION = 428_000
_WITHHOLDING_RATE = 0.05
_WITHHOLDING_TAXABLE_CAP = 560_000  # 5%稅率所得上限


def _estimate_withholding(annual_gross: float) -> int:
    """估算年度薪資所得扣繳稅額。

    計算基礎：
        扣繳稅額 = max(0, min(全年薪資 - 428,000, 560,000) * 5%)
        免稅基數 = 免稅額(97,000) + 薪資特別扣除額(207,000) + 標準扣除額(124,000)

    ⚠ 此為估算值，實際扣繳須依個人申報情形（眷屬、身心障礙、其他扣除等）調整。
    """
    taxable = annual_gross - _WITHHOLDING_DEDUCTION
    if taxable <= 0:
        return 0
    return round(min(taxable, _WITHHOLDING_TAXABLE_CAP) * _WITHHOLDING_RATE)


@router.get("/withholding")
def export_withholding(
    year: int = Query(..., ge=2000, le=2100),
    employer_name: str = Query("（請填入扣繳義務人名稱）"),
    employer_id: str = Query("（請填入統一編號）"),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
    _: None = Depends(_rate_limit),
):
    """國稅局年度薪資所得扣繳憑單（所得類別50，Excel）"""
    session = get_session()
    try:
        records = (
            session.query(SalaryRecord, Employee)
            .join(Employee, SalaryRecord.employee_id == Employee.id)
            .filter(SalaryRecord.salary_year == year)
            .all()
        )

        agg: dict[int, dict] = {}
        for sr, emp in records:
            if sr.employee_id not in agg:
                agg[sr.employee_id] = {
                    "name": emp.name,
                    "id_number": emp.id_number or "",
                    "gross": 0.0,
                    "festival_bonus": 0.0,
                    "overtime_bonus": 0.0,
                    "labor_emp": 0.0,
                    "health_emp": 0.0,
                    "pension_emp": 0.0,
                }
            d = agg[sr.employee_id]
            d["gross"] += sr.gross_salary or 0
            d["festival_bonus"] += sr.festival_bonus or 0
            d["overtime_bonus"] += sr.overtime_bonus or 0
            d["labor_emp"] += sr.labor_insurance_employee or 0
            d["health_emp"] += sr.health_insurance_employee or 0
            d["pension_emp"] += sr.pension_employee or 0

        rows = []
        for _emp_id, data in sorted(agg.items(), key=lambda x: x[1]["name"]):
            # 全年所得 = 月薪合計 + 節慶獎金 + 超額獎金（均屬薪資所得）
            annual_income = round(
                data["gross"] + data["festival_bonus"] + data["overtime_bonus"]
            )
            withholding = _estimate_withholding(annual_income)
            rows.append(
                {
                    "name": data["name"],
                    "id_number": data["id_number"],
                    "annual_income": annual_income,
                    "labor_emp": round(data["labor_emp"]),
                    "health_emp": round(data["health_emp"]),
                    "pension_emp": round(data["pension_emp"]),
                    "withholding": withholding,
                    "note": "估算值" if withholding > 0 else "低於起徵點",
                }
            )
    finally:
        session.close()

    logger.warning(
        "扣繳憑單匯出：%s年，共 %d 人，操作人：%s",
        year,
        len(rows),
        current_user.get("username", ""),
    )

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = f"扣繳憑單{year}"

    _title_row(ws, f"薪資所得扣繳憑單（所得類別 50）　{year}年度", "I")
    ws["A2"] = f"扣繳義務人：{employer_name}"
    ws["A2"].font = _BOLD
    ws.merge_cells("A2:C2")
    ws["D2"] = f"統一編號：{employer_id}"
    ws["D2"].font = _BOLD
    ws["G2"] = "⚠ 扣繳稅額為估算值，請依個人申報情形調整"
    ws["G2"].font = _ITALIC_RED
    ws.merge_cells("G2:I2")
    ws.row_dimensions[2].height = 18

    headers = [
        "序號",
        "受領人姓名",
        "身分證字號",
        "所得\n類別",
        "全年給付\n總額",
        "全年\n勞保費",
        "全年\n健保費",
        "估計\n扣繳稅額",
        "備註",
    ]
    _hdr(ws, 3, headers)
    ws.row_dimensions[3].height = 30

    for i, row in enumerate(rows, 1):
        r = i + 3
        _cell(ws, r, 1, i)
        _cell(ws, r, 2, row["name"], _LEFT)
        _cell(ws, r, 3, row["id_number"])
        _cell(ws, r, 4, "50")
        _cell(ws, r, 5, row["annual_income"], _RIGHT)
        _cell(ws, r, 6, row["labor_emp"], _RIGHT)
        _cell(ws, r, 7, row["health_emp"], _RIGHT)
        _cell(ws, r, 8, row["withholding"], _RIGHT)
        _cell(ws, r, 9, row["note"])

    _total_row(
        ws,
        len(rows) + 4,
        len(rows),
        {
            5: sum(r["annual_income"] for r in rows),
            6: sum(r["labor_emp"] for r in rows),
            7: sum(r["health_emp"] for r in rows),
            8: sum(r["withholding"] for r in rows),
        },
    )
    _auto_width(ws)
    return xlsx_streaming_response(wb, f"扣繳憑單_{year}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 4. 勞退月提繳明細
# ─────────────────────────────────────────────────────────────────────────────


@router.get("/pension")
def export_pension(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    employer_name: str = Query("（請填入雇主名稱）"),
    employer_code: str = Query("（請填入雇主代號）"),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
    _: None = Depends(_rate_limit),
):
    """勞工退休金月提繳明細（Excel）"""
    session = get_session()
    try:
        employees = _active_employees(session, year, month)
        smap = _salary_map(session, [e.id for e in employees], year, month)

        rows = []
        for emp in employees:
            sr = smap.get(emp.id)
            insured = _resolve_insured(emp)

            # 勞退：雇主提撥 > 0 才採用 record；= 0 或 None 一律走 fallback
            # （舊資料 pension_employer 為 0，過去 `is not None` 會誤採用）
            if sr and (sr.pension_employer or 0) > 0:
                pension_er = round(sr.pension_employer or 0)
                pension_self = round(sr.pension_employee or 0)
            else:
                calc = _ins_calc(emp)
                if calc:
                    pension_er = round(calc.pension_employer)
                    pension_self = round(calc.pension_employee)
                    insured = int(calc.insured_amount)
                else:
                    pension_er = pension_self = 0

            self_rate_pct = f"{(emp.pension_self_rate or 0) * 100:.0f}%"
            rows.append(
                {
                    "name": emp.name,
                    "id_number": emp.id_number or "",
                    "insured": insured,
                    "pension_er": pension_er,
                    "pension_self": pension_self,
                    "total": pension_er + pension_self,
                    "self_rate": self_rate_pct,
                }
            )
    finally:
        session.close()

    logger.warning(
        "勞退提繳匯出：%s年%s月，共 %d 人，操作人：%s",
        year,
        month,
        len(rows),
        current_user.get("username", ""),
    )

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = f"勞退{year}{month:02d}"

    _title_row(ws, f"勞工退休金月提繳明細　{year}年{month:02d}月", "H")
    ws["A2"] = f"雇主：{employer_name}"
    ws["A2"].font = _BOLD
    ws.merge_cells("A2:C2")
    ws["D2"] = f"代號：{employer_code}"
    ws["D2"].font = _BOLD
    ws["F2"] = "法定雇主提繳率：6%"
    ws["F2"].font = _BOLD
    ws.row_dimensions[2].height = 18

    headers = [
        "序號",
        "姓名",
        "身分證字號",
        "月提繳工資",
        "雇主提繳\n(6%)",
        "員工自提",
        "自提比率",
        "合計提繳",
    ]
    _hdr(ws, 3, headers)
    ws.row_dimensions[3].height = 30

    for i, row in enumerate(rows, 1):
        r = i + 3
        _cell(ws, r, 1, i)
        _cell(ws, r, 2, row["name"], _LEFT)
        _cell(ws, r, 3, row["id_number"])
        _cell(ws, r, 4, row["insured"], _RIGHT)
        _cell(ws, r, 5, row["pension_er"], _RIGHT)
        _cell(ws, r, 6, row["pension_self"], _RIGHT)
        _cell(ws, r, 7, row["self_rate"])
        _cell(ws, r, 8, row["total"], _RIGHT)

    _total_row(
        ws,
        len(rows) + 4,
        len(rows),
        {
            4: sum(r["insured"] for r in rows),
            5: sum(r["pension_er"] for r in rows),
            6: sum(r["pension_self"] for r in rows),
            8: sum(r["total"] for r in rows),
        },
    )
    _auto_width(ws)
    return xlsx_streaming_response(wb, f"勞退提繳_{year}{month:02d}.xlsx")
