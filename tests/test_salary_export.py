"""薪資匯出回歸測試。"""

import os
import sys
from datetime import date
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import router as auth_router
from api.auth import _account_failures, _ip_attempts
from api.salary import init_salary_services
from api.salary import router as salary_router
from models.database import Base, Employee, SalaryRecord, User
from services.salary.engine import SalaryEngine
from utils.auth import hash_password
from utils.permissions import Permission


@pytest.fixture
def salary_export_client(tmp_path):
    """建立隔離的 sqlite 測試 app。"""
    db_path = tmp_path / "salary-export.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    init_salary_services(SalaryEngine(load_from_db=False), MagicMock())

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _create_employee(session, employee_id: str, name: str) -> Employee:
    employee = Employee(
        employee_id=employee_id,
        name=name,
        base_salary=36000,
        is_active=True,
    )
    session.add(employee)
    session.flush()
    return employee


def _create_user(session, username: str, permissions: int) -> User:
    user = User(
        username=username,
        password_hash=hash_password("TempPass123"),
        role="admin",
        permissions=permissions,
        is_active=True,
        must_change_password=False,
    )
    session.add(user)
    session.flush()
    return user


def _login(client: TestClient, username: str, password: str = "TempPass123"):
    return client.post("/api/auth/login", json={"username": username, "password": password})


class TestSalaryExcelExport:
    def test_xlsx_export_does_not_require_reportlab(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E001", "王小明")
            _create_user(session, "salary_admin", int(Permission.SALARY_READ))
            session.add(
                SalaryRecord(
                    employee_id=employee.id,
                    salary_year=2026,
                    salary_month=3,
                    base_salary=30000,
                    gross_salary=32000,
                    total_deduction=1000,
                    net_salary=31000,
                )
            )
            session.commit()

        login_res = _login(client, "salary_admin")
        assert login_res.status_code == 200

        response = client.get("/api/salaries/export-all?year=2026&month=3&format=xlsx")

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(response.content) > 0

    def test_xlsx_export_includes_remark_column(self, salary_export_client):
        from io import BytesIO
        from openpyxl import load_workbook

        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E011", "備註老師")
            _create_user(session, "salary_excel_admin", int(Permission.SALARY_READ))
            session.add(
                SalaryRecord(
                    employee_id=employee.id,
                    salary_year=2026,
                    salary_month=3,
                    gross_salary=32000,
                    total_deduction=1000,
                    net_salary=31000,
                    remark="[2026-03-14 10:00] 手動編輯：節慶獎金 2000→1800",
                )
            )
            session.commit()

        login_res = _login(client, "salary_excel_admin")
        assert login_res.status_code == 200

        response = client.get("/api/salaries/export-all?year=2026&month=3&format=xlsx")

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
        assert headers[-1] == "編輯紀錄"
        assert "手動編輯" in str(ws.cell(row=4, column=ws.max_column).value)


class TestSalaryFieldBreakdownApi:
    def test_returns_field_breakdown_for_record(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E002", "林老師")
            employee.birthday = date(2020, 3, 10)
            _create_user(session, "salary_breakdown_admin", int(Permission.SALARY_READ))
            record = SalaryRecord(
                employee_id=employee.id,
                salary_year=2026,
                salary_month=3,
                base_salary=30000,
                festival_bonus=2000,
                birthday_bonus=500,
            )
            session.add(record)
            session.commit()
            record_id = record.id

        login_res = _login(client, "salary_breakdown_admin")
        assert login_res.status_code == 200

        response = client.get(f"/api/salaries/{record_id}/field-breakdown?field=festival_bonus")

        assert response.status_code == 200
        payload = response.json()
        assert payload["employee"]["record_id"] == record_id
        assert payload["employee"]["employee_name"] == "林老師"
        assert payload["employee"]["employee_code"] == "E002"
        assert payload["employee"]["year"] == 2026
        assert payload["employee"]["month"] == 3
        assert payload["field"] == "festival_bonus"
        assert payload["columns"][0]["label"] == "姓名"
        assert payload["summary"]["amount"] == 2000
        assert payload["rows"][0]["name"] == "林老師"

    def test_birthday_bonus_zero_still_returns_reason(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E003", "王老師")
            employee.birthday = date(2020, 5, 10)
            _create_user(session, "salary_birthday_admin", int(Permission.SALARY_READ))
            record = SalaryRecord(
                employee_id=employee.id,
                salary_year=2026,
                salary_month=3,
                birthday_bonus=0,
            )
            session.add(record)
            session.commit()
            record_id = record.id

        login_res = _login(client, "salary_birthday_admin")
        assert login_res.status_code == 200

        response = client.get(f"/api/salaries/{record_id}/field-breakdown?field=birthday_bonus")

        assert response.status_code == 200
        payload = response.json()
        assert payload["rows"][0]["matched"] == "否"
        assert payload["summary"]["amount"] == 0
        assert payload["note"] == "只看生日月份，不看日期。"

    def test_returns_400_when_field_is_invalid(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E004", "陳老師")
            _create_user(session, "salary_breakdown_invalid", int(Permission.SALARY_READ))
            session.add(SalaryRecord(employee_id=employee.id, salary_year=2026, salary_month=3))
            session.commit()

        login_res = _login(client, "salary_breakdown_invalid")
        assert login_res.status_code == 200

        response = client.get("/api/salaries/1/field-breakdown?field=not_supported")

        assert response.status_code == 400

    def test_returns_404_when_record_not_found(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            _create_user(session, "salary_breakdown_missing", int(Permission.SALARY_READ))
            session.commit()

        login_res = _login(client, "salary_breakdown_missing")
        assert login_res.status_code == 200

        response = client.get("/api/salaries/999999/field-breakdown?field=festival_bonus")

        assert response.status_code == 404


class TestSalaryManualAdjustApi:
    def test_updates_record_and_appends_edit_remark(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E012", "可調整老師")
            _create_user(session, "salary_write_admin", int(Permission.SALARY_WRITE | Permission.SALARY_READ))
            record = SalaryRecord(
                employee_id=employee.id,
                salary_year=2026,
                salary_month=3,
                base_salary=30000,
                meal_allowance=2400,
                festival_bonus=2000,
                overtime_bonus=500,
                supervisor_dividend=1000,
                labor_insurance_employee=700,
                health_insurance_employee=500,
                pension_employee=1800,
                late_deduction=100,
                leave_deduction=300,
                gross_salary=33400,
                total_deduction=3400,
                net_salary=30000,
                bonus_amount=3500,
                bonus_separate=True,
            )
            session.add(record)
            session.commit()
            record_id = record.id

        login_res = _login(client, "salary_write_admin")
        assert login_res.status_code == 200

        response = client.put(
            f"/api/salaries/{record_id}/manual-adjust",
            json={
                "festival_bonus": 1800,
                "leave_deduction": 500,
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["record"]["festival_bonus"] == 1800
        assert payload["record"]["leave_deduction"] == 500
        assert payload["record"]["bonus_amount"] == 3300
        assert payload["record"]["total_deduction"] == 3600
        assert payload["record"]["net_salary"] == 29800
        assert "手動編輯" in payload["record"]["remark"]
        assert "節慶獎金 2000→1800" in payload["record"]["remark"]
        assert "請假扣款 300→500" in payload["record"]["remark"]

    def test_rejects_adjustment_for_finalized_record(self, salary_export_client):
        client, session_factory = salary_export_client

        with session_factory() as session:
            employee = _create_employee(session, "E013", "封存老師")
            _create_user(session, "salary_write_finalized", int(Permission.SALARY_WRITE | Permission.SALARY_READ))
            record = SalaryRecord(
                employee_id=employee.id,
                salary_year=2026,
                salary_month=3,
                gross_salary=30000,
                total_deduction=1000,
                net_salary=29000,
                is_finalized=True,
            )
            session.add(record)
            session.commit()
            record_id = record.id

        login_res = _login(client, "salary_write_finalized")
        assert login_res.status_code == 200

        response = client.put(
            f"/api/salaries/{record_id}/manual-adjust",
            json={"festival_bonus": 1000},
        )

        assert response.status_code == 409
