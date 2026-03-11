"""
School Events (Calendar) router - CRUD for school calendar events
"""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from utils.errors import raise_safe_500
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from models.database import get_session, SchoolEvent, Holiday
from utils.auth import require_permission
from utils.permissions import Permission
from utils.file_upload import read_upload_with_size_check

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["events"])

EVENT_TYPE_LABELS = {
    "meeting": "會議",
    "activity": "活動",
    "holiday": "假日",
    "general": "一般",
}


# ============ Pydantic Models ============

class EventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    event_date: date
    end_date: Optional[date] = None
    event_type: str = "general"
    is_all_day: bool = True
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None


class EventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    event_date: Optional[date] = None
    end_date: Optional[date] = None
    event_type: Optional[str] = None
    is_all_day: Optional[bool] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None


# ============ Endpoints ============

def _event_to_dict(ev: SchoolEvent) -> dict:
    return {
        "id": ev.id,
        "title": ev.title,
        "description": ev.description,
        "event_date": ev.event_date.isoformat(),
        "end_date": ev.end_date.isoformat() if ev.end_date else None,
        "event_type": ev.event_type,
        "event_type_label": EVENT_TYPE_LABELS.get(ev.event_type, ev.event_type),
        "is_all_day": ev.is_all_day,
        "start_time": ev.start_time,
        "end_time": ev.end_time,
        "location": ev.location,
        "created_at": ev.created_at.isoformat() if ev.created_at else None,
        "updated_at": ev.updated_at.isoformat() if ev.updated_at else None,
    }


@router.get("/events")
def get_events(
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    event_type: Optional[str] = Query(None),
    current_user: dict = Depends(require_permission(Permission.CALENDAR)),
):
    """取得行事曆事件列表"""
    session = get_session()
    try:
        q = session.query(SchoolEvent).filter(SchoolEvent.is_active == True)

        if year:
            start = date(year, month or 1, 1)
            if month:
                import calendar as cal_module
                _, last_day = cal_module.monthrange(year, month)
                end = date(year, month, last_day)
            else:
                end = date(year, 12, 31)
            # Include events that overlap with the range
            q = q.filter(
                SchoolEvent.event_date <= end,
                (SchoolEvent.end_date >= start) | (SchoolEvent.end_date.is_(None) & (SchoolEvent.event_date >= start)),
            )

        if event_type:
            q = q.filter(SchoolEvent.event_type == event_type)

        events = q.order_by(SchoolEvent.event_date).all()
        return [_event_to_dict(ev) for ev in events]
    finally:
        session.close()


@router.get("/events/{event_id}")
def get_event(event_id: int, current_user: dict = Depends(require_permission(Permission.CALENDAR))):
    """取得單一事件"""
    session = get_session()
    try:
        ev = session.query(SchoolEvent).filter(
            SchoolEvent.id == event_id,
            SchoolEvent.is_active == True,
        ).first()
        if not ev:
            raise HTTPException(status_code=404, detail="找不到該事件")
        return _event_to_dict(ev)
    finally:
        session.close()


@router.post("/events", status_code=201)
def create_event(data: EventCreate, current_user: dict = Depends(require_permission(Permission.CALENDAR))):
    """新增行事曆事件"""
    session = get_session()
    try:
        if data.event_type not in EVENT_TYPE_LABELS:
            raise HTTPException(status_code=400, detail=f"無效的事件類型: {data.event_type}")
        if data.end_date and data.end_date < data.event_date:
            raise HTTPException(status_code=400, detail="結束日期不可早於開始日期")

        ev = SchoolEvent(
            title=data.title,
            description=data.description,
            event_date=data.event_date,
            end_date=data.end_date,
            event_type=data.event_type,
            is_all_day=data.is_all_day,
            start_time=data.start_time,
            end_time=data.end_time,
            location=data.location,
        )
        session.add(ev)
        session.commit()
        return {"message": "事件已建立", "id": ev.id}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/events/{event_id}")
def update_event(event_id: int, data: EventUpdate, current_user: dict = Depends(require_permission(Permission.CALENDAR))):
    """更新行事曆事件"""
    session = get_session()
    try:
        ev = session.query(SchoolEvent).filter(
            SchoolEvent.id == event_id,
            SchoolEvent.is_active == True,
        ).first()
        if not ev:
            raise HTTPException(status_code=404, detail="找不到該事件")

        update_data = data.dict(exclude_unset=True)
        if "event_type" in update_data and update_data["event_type"] not in EVENT_TYPE_LABELS:
            raise HTTPException(status_code=400, detail=f"無效的事件類型: {update_data['event_type']}")

        for key, value in update_data.items():
            setattr(ev, key, value)

        # Validate end_date
        if ev.end_date and ev.end_date < ev.event_date:
            raise HTTPException(status_code=400, detail="結束日期不可早於開始日期")

        session.commit()
        return {"message": "事件已更新"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/events/{event_id}")
def delete_event(event_id: int, current_user: dict = Depends(require_permission(Permission.CALENDAR))):
    """刪除行事曆事件（軟刪除）"""
    session = get_session()
    try:
        ev = session.query(SchoolEvent).filter(
            SchoolEvent.id == event_id,
            SchoolEvent.is_active == True,
        ).first()
        if not ev:
            raise HTTPException(status_code=404, detail="找不到該事件")
        ev.is_active = False
        session.commit()
        return {"message": "事件已刪除"}
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


# ============ 假日批次匯入（Holiday 表） ============

def _ev_xlsx_response(wb, filename: str):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


_EV_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_EV_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_EV_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_EV_CENTER_ALIGN = Alignment(horizontal="center")


def _ev_write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _EV_HEADER_FONT
        cell.fill = _EV_HEADER_FILL
        cell.border = _EV_THIN_BORDER
        cell.alignment = _EV_CENTER_ALIGN


@router.get("/events/holidays/import-template")
def get_holiday_import_template(
    current_user: dict = Depends(require_permission(Permission.CALENDAR)),
):
    """下載國定假日批次匯入 Excel 範本"""
    wb = Workbook()
    ws = wb.active
    ws.title = "假日匯入範本"

    headers = ["日期", "假日名稱", "說明(可空)"]
    _ev_write_header(ws, 1, headers)

    ws.cell(row=2, column=1, value="2026-01-01")
    ws.cell(row=2, column=2, value="元旦")
    ws.cell(row=2, column=3, value="新年第一天")
    ws.cell(row=3, column=1, value="2026-02-17")
    ws.cell(row=3, column=2, value="農曆春節")

    note_ws = wb.create_sheet("說明")
    note_ws.cell(row=1, column=1, value="注意事項")
    note_ws.cell(row=2, column=1, value="1. 日期格式建議使用 YYYY-MM-DD")
    note_ws.cell(row=3, column=1, value="2. 同日期若已存在則更新，否則新增（UPSERT）")
    note_ws.cell(row=4, column=1, value="3. 匯入後考勤計算將自動排除這些假日")

    return _ev_xlsx_response(wb, "假日匯入範本.xlsx")


@router.post("/events/holidays/import")
async def import_holidays(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission(Permission.CALENDAR)),
):
    """批次匯入國定假日（UPSERT by date，同日期若已存在則更新）"""
    content = await read_upload_with_size_check(file)
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{e}")

    results: dict = {"total": 0, "upserted": 0, "failed": 0, "errors": []}
    session = get_session()
    try:
        for idx, row in df.iterrows():
            results["total"] += 1
            row_num = int(idx) + 2
            try:
                date_raw = row.get("日期")
                if date_raw is None or pd.isna(date_raw):
                    raise ValueError("日期不得為空")
                try:
                    holiday_date = pd.to_datetime(date_raw).date()
                except Exception:
                    raise ValueError("日期格式錯誤，建議使用 YYYY-MM-DD")

                name_raw = row.get("假日名稱")
                if name_raw is None or pd.isna(name_raw):
                    raise ValueError("假日名稱不得為空")
                name = str(name_raw).strip()
                if not name:
                    raise ValueError("假日名稱不得為空")

                desc_raw = row.get("說明(可空)")
                description = (
                    str(desc_raw).strip()
                    if desc_raw is not None and not pd.isna(desc_raw)
                    else None
                )

                existing = session.query(Holiday).filter(Holiday.date == holiday_date).first()
                if existing:
                    existing.name = name
                    existing.description = description
                    existing.is_active = True
                    existing.updated_at = datetime.now()
                else:
                    session.add(Holiday(
                        date=holiday_date,
                        name=name,
                        description=description,
                        is_active=True,
                    ))

                session.flush()
                results["upserted"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"第 {row_num} 行: {str(e)}")

        session.commit()
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"匯入失敗：{e}")
    finally:
        session.close()

    logger.info(
        "假日批次匯入：使用者 %s，共 %d 筆，成功 %d 筆，失敗 %d 筆",
        current_user.get("username"), results["total"], results["upserted"], results["failed"],
    )
    return results
