"""Endpoint 整合測試 — POST /generate, GET /, GET /export, audit, lock。

認證方式：HttpOnly cookie（`access_token`）。
TestClient 在同一 instance 內自動帶 cookie；切換使用者時手動替換 cookie jar。

Fixture 設計：
  monthly_ctx dict:
    client      TestClient（cookie jar 在同一個 app session）
    sf          sessionmaker
    export_cookie   {"access_token": "..."}
    view_only_cookie
    none_cookie
"""

import os
import sys
from datetime import date
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from api.gov_moe import router as gov_moe_router
from models.audit import AuditLog
from models.base import Base
from models.classroom import Classroom, Student, StudentAttendance
from models.database import User
from models.gov_moe import MonthlyEnrollmentSnapshot  # noqa: F401 — registers table
from utils.auth import hash_password
from utils.permissions import Permission

# ---------------------------------------------------------------------------
# 主 Fixture
# ---------------------------------------------------------------------------


@pytest.fixture
def monthly_ctx(tmp_path):
    """一個 SQLite DB + FastAPI app + 三組 cookie。

    回傳 dict:
        client            TestClient
        sf                sessionmaker
        export_cookie     dict  {"access_token": "..."}
        view_only_cookie  dict
        none_cookie       dict
    """
    db_path = tmp_path / "gov_moe_monthly.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}", connect_args={"check_same_thread": False}
    )
    sf = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_sf = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = sf

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(gov_moe_router, prefix="/api")

    export_perms = int(Permission.GOV_REPORTS_EXPORT | Permission.GOV_REPORTS_VIEW)
    view_perms = int(Permission.GOV_REPORTS_VIEW)

    with sf() as s:
        s.add(
            User(
                username="export_admin",
                password_hash=hash_password("ExportPass1"),
                role="admin",
                permissions=export_perms,
                is_active=True,
            )
        )
        s.add(
            User(
                username="view_admin",
                password_hash=hash_password("ViewPass1"),
                role="admin",
                permissions=view_perms,
                is_active=True,
            )
        )
        s.add(
            User(
                username="no_perm_admin",
                password_hash=hash_password("NopermPass1"),
                role="admin",
                permissions=0,
                is_active=True,
            )
        )
        s.commit()

    with TestClient(app) as client:

        def _login(username, password):
            res = client.post(
                "/api/auth/login",
                json={"username": username, "password": password},
            )
            assert res.status_code == 200, res.text
            token = res.cookies.get("access_token")
            assert token, f"no access_token cookie for {username}"
            return {"access_token": token}

        export_cookie = _login("export_admin", "ExportPass1")
        view_only_cookie = _login("view_admin", "ViewPass1")
        none_cookie = _login("no_perm_admin", "NopermPass1")

        yield {
            "client": client,
            "sf": sf,
            "export_cookie": export_cookie,
            "view_only_cookie": view_only_cookie,
            "none_cookie": none_cookie,
        }

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_sf
    engine.dispose()


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _with_cookie(ctx, cookie_key: str, method: str, path: str, **kwargs):
    """呼叫 client.<method>，並以指定 cookie 替換 client cookie jar。"""
    client: TestClient = ctx["client"]
    cookies = ctx[cookie_key]  # {"access_token": "..."}
    return getattr(client, method)(path, cookies=cookies, **kwargs)


def _generate(ctx, cookie_key="export_cookie", year=2026, month=5):
    return _with_cookie(
        ctx,
        cookie_key,
        "post",
        "/api/gov-moe/monthly/generate",
        json={"year": year, "month": month},
    )


def _seed_student(ctx, year=2026, month=5):
    """在測試 DB 建 Classroom + Student + StudentAttendance（month 有出席紀錄）。"""
    with ctx["sf"]() as s:
        classroom = Classroom(name="草莓班", capacity=25, school_year=2025, semester=2)
        s.add(classroom)
        s.flush()

        student = Student(
            student_id="T001",
            name="測試生",
            gender="男",
            birthday=date(2021, 6, 1),
            classroom_id=classroom.id,
            enrollment_date=date(2025, 1, 1),
            nationality="本國",
            lifecycle_status="active",
        )
        s.add(student)
        s.flush()

        for day in [1, 2, 5, 6, 7]:
            s.add(
                StudentAttendance(
                    student_id=student.id,
                    date=date(year, month, day),
                    status="出席",
                )
            )
        s.commit()
        return {"classroom_id": classroom.id, "student_id": student.id}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_generate_creates_snapshot_rows(monthly_ctx):
    _seed_student(monthly_ctx)
    resp = _generate(monthly_ctx)
    assert resp.status_code == 200
    body = resp.json()
    assert body["year"] == 2026
    assert body["month"] == 5
    assert body["rows_generated"] >= 1


def test_generate_writes_audit_log(monthly_ctx):
    _seed_student(monthly_ctx)
    resp = _generate(monthly_ctx)
    assert resp.status_code == 200

    with monthly_ctx["sf"]() as s:
        audit = (
            s.query(AuditLog)
            .filter(
                AuditLog.action == "GENERATE",
                AuditLog.entity_type == "monthly_enrollment_snapshot",
            )
            .first()
        )
    assert audit is not None
    assert audit.entity_id == "2026-05"


def test_regenerate_overwrites_and_audits(monthly_ctx):
    _seed_student(monthly_ctx)
    _generate(monthly_ctx)

    with monthly_ctx["sf"]() as s:
        first_count = (
            s.query(MonthlyEnrollmentSnapshot).filter_by(year=2026, month=5).count()
        )

    resp = _generate(monthly_ctx)
    assert resp.status_code == 200

    with monthly_ctx["sf"]() as s:
        second_count = (
            s.query(MonthlyEnrollmentSnapshot).filter_by(year=2026, month=5).count()
        )
        assert second_count == first_count  # 覆寫不 double

        audits = (
            s.query(AuditLog)
            .filter(
                AuditLog.entity_type == "monthly_enrollment_snapshot",
                AuditLog.entity_id == "2026-05",
            )
            .all()
        )
    assert len(audits) >= 2
    assert "REGENERATE" in [a.action for a in audits]


def test_generate_invalid_year_400(monthly_ctx):
    resp = _generate(monthly_ctx, year=1999)
    assert resp.status_code == 400


def test_generate_invalid_month_422(monthly_ctx):
    resp = _with_cookie(
        monthly_ctx,
        "export_cookie",
        "post",
        "/api/gov-moe/monthly/generate",
        json={"year": 2026, "month": 13},
    )
    assert resp.status_code == 422


def test_generate_requires_export_permission(monthly_ctx):
    resp = _with_cookie(
        monthly_ctx,
        "view_only_cookie",
        "post",
        "/api/gov-moe/monthly/generate",
        json={"year": 2026, "month": 5},
    )
    assert resp.status_code == 403


def test_get_returns_404_before_generate(monthly_ctx):
    resp = _with_cookie(
        monthly_ctx,
        "view_only_cookie",
        "get",
        "/api/gov-moe/monthly?year=2026&month=5",
    )
    assert resp.status_code == 404


def test_get_returns_three_dimensions(monthly_ctx):
    _seed_student(monthly_ctx)
    _generate(monthly_ctx)  # export user 先產生

    # view_only user 讀取
    resp = _with_cookie(
        monthly_ctx,
        "view_only_cookie",
        "get",
        "/api/gov-moe/monthly?year=2026&month=5",
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "classroom_summary" in body
    assert "student_detail" in body
    assert "overview" in body
    assert body["year"] == 2026


def test_export_returns_xlsx_bytes(monthly_ctx):
    _seed_student(monthly_ctx)
    _generate(monthly_ctx)

    resp = _with_cookie(
        monthly_ctx,
        "export_cookie",
        "get",
        "/api/gov-moe/monthly/export?year=2026&month=5&format=xlsx",
    )
    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # RFC 5987 encoded: 義華幼兒園_月報 → %E7%BE%A9%E8%8F%AF...；年月保留 ASCII
    cd = resp.headers["content-disposition"]
    assert "2026-05" in cd
    assert "UTF-8''" in cd  # RFC 5987 prefix
    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["班級總表", "幼生明細", "統計摘要"]


def test_export_returns_404_before_generate(monthly_ctx):
    resp = _with_cookie(
        monthly_ctx,
        "export_cookie",
        "get",
        "/api/gov-moe/monthly/export?year=2026&month=5",
    )
    assert resp.status_code == 404


def test_get_requires_view_permission(monthly_ctx):
    resp = _with_cookie(
        monthly_ctx,
        "none_cookie",
        "get",
        "/api/gov-moe/monthly?year=2026&month=5",
    )
    assert resp.status_code == 403


def test_generate_409_when_lock_contention(monthly_ctx, monkeypatch):
    """Advisory lock 被佔用時回 409。"""
    from api.gov_moe import monthly as monthly_module

    monkeypatch.setattr(monthly_module, "_try_advisory_lock", lambda db, y, m: False)
    resp = monthly_ctx["client"].post(
        "/api/gov-moe/monthly/generate",
        json={"year": 2026, "month": 5},
        cookies=monthly_ctx["export_cookie"],
    )
    assert resp.status_code == 409


def test_export_requires_export_permission(monthly_ctx):
    """view_only 不能 export（403）。"""
    # 先 generate 拿到資料
    _seed_student(monthly_ctx)
    _generate(monthly_ctx)
    resp = _with_cookie(
        monthly_ctx,
        "view_only_cookie",
        "get",
        "/api/gov-moe/monthly/export?year=2026&month=5&format=xlsx",
    )
    assert resp.status_code == 403
