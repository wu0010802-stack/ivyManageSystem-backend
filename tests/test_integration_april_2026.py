"""端到端整合驗證：用《義華薪資 115.04》Excel 才藝老師 sheet 的 14 筆資料
建 hourly 員工 + 才藝 entries，跑 SalaryEngine 並驗證輸出對齊。

不啟動 backend server，純 Python in-memory 驗證計算層。
"""

import os
import sys
from unittest.mock import MagicMock

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from models.database import ArtTeacherPayrollEntry, Base, Employee
from services.art_teacher_payroll import (
    compute_total_for_month,
    generate_art_teacher_roster_xlsx,
)
from services.salary.engine import SalaryEngine

# Excel《義華薪資 115.04》才藝老師 sheet 完整 14 筆資料
# (科目, 班級備註, 姓名, 時數, 鐘點費, 超額, 加給活動, 預期總計)
ART_TEACHER_DATA = [
    ("美語", "向.滿", "歐瑞煌Johnny", 25, 550, 0, 0, 13750),
    ("美語", "天.玫", "黃雅絹Jessica", 37.5, 550, 0, 0, 20625),
    ("美語", "櫻.牡", "簡鳳儀Tiffany", 30.84, 550, 0, 0, 16962),
    ("美語", "茉.薔", "黃毓慧Ivy", 36.5, 530, 0, 530, 19875),
    ("美語", "百.芙", "楊秀敏Nicole", 37.5, 500, 0, 0, 18750),
    ("外師", "", "Vadim", 26, 620, 0, 0, 16120),
    ("課後美語", "(二)", "Vadim", 4, 620, 0, 0, 2480),
    ("課後美語", "(五)", "鐘綩菱", 3, 500, 0, 0, 1500),
    ("舞蹈", "(二)", "鍾馨瑶", 4, 1000, 200, 0, 4200),
    ("舞蹈", "(四)", "鄭奷容", 4, 1200, 300, 0, 5100),
    ("音樂", "(四)", "李紋娟", 3, 800, 0, 0, 2400),
    ("管家", "", "李麗珍", 208, 220, 0, 6000, 51760),
    ("體能", "(三)", "陳安星Patrick", 7.5, 1000, 0, 0, 7500),
    ("足球", "(一)", "吳政軒", 62, 100, 0, 0, 6200),
    ("足球", "(五)", "董老爺(阿貴)", 42, 100, 0, 0, 4200),
    ("美術", "(三.五)", "謝奇玲", 8, 550, 0, 0, 4400),
    ("感統", "(三)", "陳博明", 70, 280, 0, 0, 19600),
]


@pytest.fixture
def integration_session(tmp_path):
    """獨立 sqlite，內含 schema，不影響其他測試。"""
    db_path = tmp_path / "integration.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    yield session_factory()

    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _create_hourly_employee(session, name, index):
    emp = Employee(
        employee_id=f"ART{index:03d}",
        name=name,
        employee_type="hourly",
        base_salary=0,
        hourly_rate=0,
        insurance_salary_level=0,  # 才藝老師多半不投保（業主慣例）
        is_active=True,
        bank_account=f"0727-979-{index:06d}",
        bank_account_name=name,
    )
    session.add(emp)
    session.flush()
    return emp


class TestApril2026ArtTeacherIntegration:
    """跑完《義華薪資 115.04》Excel 才藝 sheet 14 筆，驗證每筆與 Excel 一致。"""

    def test_seed_and_total_matches_excel(self, integration_session):
        session = integration_session

        # 建員工（有些老師多筆，共 14 個唯一姓名）
        unique_names = []
        seen = set()
        for row in ART_TEACHER_DATA:
            name = row[2]
            if name not in seen:
                seen.add(name)
                unique_names.append(name)

        emp_by_name = {}
        for idx, name in enumerate(unique_names, start=1):
            emp_by_name[name] = _create_hourly_employee(session, name, idx)

        # 建 entries
        for (
            subject,
            classroom,
            name,
            hours,
            rate,
            excess,
            activity,
            expected,
        ) in ART_TEACHER_DATA:
            emp = emp_by_name[name]
            base = round(hours * rate)
            total = base + excess + activity
            session.add(
                ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    subject=subject,
                    classroom_label=classroom or None,
                    hours=hours,
                    hourly_rate=rate,
                    base_amount=base,
                    excess_amount=excess,
                    activity_bonus=activity,
                    total_amount=total,
                )
            )
            assert (
                total == expected
            ), f"{name} {subject}{classroom}: 計算 {total} ≠ Excel 期望 {expected}"

        session.commit()

        # 驗證每位老師合計（Vadim 有兩筆要相加）
        expected_per_teacher = {
            "歐瑞煌Johnny": 13750,
            "黃雅絹Jessica": 20625,
            "簡鳳儀Tiffany": 16962,
            "黃毓慧Ivy": 19875,
            "楊秀敏Nicole": 18750,
            "Vadim": 16120 + 2480,  # 外師 + 課後美語
            "鐘綩菱": 1500,
            "鍾馨瑶": 4200,
            "鄭奷容": 5100,
            "李紋娟": 2400,
            "李麗珍": 51760,
            "陳安星Patrick": 7500,
            "吳政軒": 6200,
            "董老爺(阿貴)": 4200,
            "謝奇玲": 4400,
            "陳博明": 19600,
        }
        for name, expected_total in expected_per_teacher.items():
            actual = compute_total_for_month(session, emp_by_name[name].id, 2026, 4)
            assert (
                int(actual) == expected_total
            ), f"{name} 合計：系統 {actual} ≠ Excel {expected_total}"

        # 驗證 Excel 整體合計 215422（從 Excel 才藝老師轉帳名冊讀出）
        grand_total = sum(expected_per_teacher.values())
        assert grand_total == 215422, f"全月合計 {grand_total} ≠ Excel 215422"

    def test_engine_uses_entries_overriding_hourly_rate(self, integration_session):
        """engine 在發現 entries 時，hourly_total 應 = sum(entries)，不走 rate×hours。"""
        session = integration_session
        emp = _create_hourly_employee(session, "Vadim", 1)
        # 故意把 employee.hourly_rate 設 0（如果系統錯誤用 hourly_rate × work_hours 會=0）
        session.add(
            ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="外師",
                hours=26,
                hourly_rate=620,
                base_amount=16120,
                total_amount=16120,
            )
        )
        session.add(
            ArtTeacherPayrollEntry(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=4,
                subject="課後美語",
                classroom_label="(二)",
                hours=4,
                hourly_rate=620,
                base_amount=2480,
                total_amount=2480,
            )
        )
        session.commit()

        engine = SalaryEngine(load_from_db=False)
        emp_dict = {
            "employee_id": emp.employee_id,
            "name": emp.name,
            "employee_type": "hourly",
            "hourly_rate": 0,  # 故意 0
            "work_hours": 0,
            "base_salary": 0,
            "art_teacher_entries_total": float(
                compute_total_for_month(session, emp.id, 2026, 4)
            ),
        }
        breakdown = engine.calculate_salary(emp_dict, 2026, 4)
        # 即使 hourly_rate=0，entries 提供的 18600 應生效
        assert breakdown.hourly_total == 18600
        assert breakdown.gross_salary == 18600

    def test_excel_roster_export_matches_excel_total(self, integration_session):
        """匯出才藝清冊 Excel，合計列應 = 215422。"""
        from io import BytesIO

        from openpyxl import load_workbook

        session = integration_session

        # seed 同上
        unique_names = []
        seen = set()
        for row in ART_TEACHER_DATA:
            if row[2] not in seen:
                seen.add(row[2])
                unique_names.append(row[2])
        emp_by_name = {}
        for idx, name in enumerate(unique_names, start=1):
            emp_by_name[name] = _create_hourly_employee(session, name, idx)
        for (
            subject,
            classroom,
            name,
            hours,
            rate,
            excess,
            activity,
            _,
        ) in ART_TEACHER_DATA:
            emp = emp_by_name[name]
            base = round(hours * rate)
            session.add(
                ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=2026,
                    salary_month=4,
                    subject=subject,
                    classroom_label=classroom or None,
                    hours=hours,
                    hourly_rate=rate,
                    base_amount=base,
                    excess_amount=excess,
                    activity_bonus=activity,
                    total_amount=base + excess + activity,
                )
            )
        session.commit()

        _, xlsx = generate_art_teacher_roster_xlsx(session, 2026, 4)
        wb = load_workbook(BytesIO(xlsx))
        ws = wb.active

        # 找合計列（第一欄=「合計」）
        grand_total = None
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == "合計":
                grand_total = ws.cell(row=row_idx, column=8).value
                break
        assert grand_total is not None, "匯出 xlsx 沒有合計列"
        assert int(grand_total) == 215422, f"清冊合計 {grand_total} ≠ Excel 215422"

    def test_supplementary_health_for_hourly_at_excel_threshold(self):
        """Excel 註記：未達 29500 不扣補充保費。驗證 Excel 14 筆中哪些會扣。

        歐瑞煌 13750 < 29500 → 不扣 ✓（Excel 一致）
        李麗珍 51760 ≥ 29500 → 扣 51760 × 2.11% = 1092
        陳博明 19600 < 29500 → 不扣 ✓
        Vadim 個別 entry 16120+2480 ≠ 18600（合計），合計 < 29500 → 不扣

        但業主實務上補充保費是「單筆給付」≥ 門檻才扣，所以是看「合計給付」。
        系統實作以 hourly_total（= entries 合計）判斷，符合此語意。
        """
        engine = SalaryEngine(load_from_db=False)

        # 李麗珍 51760 → 應扣 51760 × 0.0211 = 1092.14 → 1092
        emp_dict = {
            "employee_id": "ART011",
            "name": "李麗珍",
            "employee_type": "hourly",
            "hourly_rate": 0,
            "work_hours": 0,
            "base_salary": 0,
            "art_teacher_entries_total": 51760,
        }
        breakdown = engine.calculate_salary(emp_dict, 2026, 4)
        assert breakdown.supplementary_health_employee == 1092

        # 歐瑞煌 13750 → 不扣
        emp_dict2 = dict(emp_dict, name="歐瑞煌", art_teacher_entries_total=13750)
        b2 = engine.calculate_salary(emp_dict2, 2026, 4)
        assert b2.supplementary_health_employee == 0
