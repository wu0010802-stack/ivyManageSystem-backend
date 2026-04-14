"""招生統計 API / schema 回歸測試。"""

import asyncio
import os
import sys
import threading
from datetime import date, datetime, timedelta
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api import recruitment as recruitment_api
from api.recruitment import (
    CampusSettingPayload,
    ImportRecord,
    MonthCreate,
    RecruitmentVisitCreate,
    RecruitmentVisitUpdate,
    get_recruitment_campus_setting,
    get_recruitment_address_hotspots,
    get_recruitment_market_intelligence,
    get_recruitment_options,
    get_nearby_kindergartens,
    get_recruitment_stats,
    get_periods_summary,
    import_recruitment_records,
    list_recruitment_records,
    normalize_existing_months,
    export_recruitment_stats,
    sync_recruitment_market_intelligence,
    sync_recruitment_address_hotspots,
    update_recruitment_campus_setting,
)
from models.base import Base
from models.recruitment import (
    RecruitmentAreaInsightCache,
    RecruitmentCampusSetting,
    RecruitmentGeocodeCache,
    RecruitmentPeriod,
    RecruitmentVisit,
)


@pytest.fixture
def recruitment_session_factory(tmp_path):
    db_path = tmp_path / "recruitment-api.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}",
        connect_args={"check_same_thread": False},
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)

    try:
        yield session_factory
    finally:
        base_module._engine = old_engine
        base_module._SessionFactory = old_session_factory
        engine.dispose()


@pytest.fixture
def recruitment_client(recruitment_session_factory):
    app = FastAPI()
    app.include_router(recruitment_api.router)

    nearby_route = next(
        route
        for route in app.routes
        if getattr(route, "path", "") == "/api/recruitment/nearby-kindergartens"
    )
    for dependency in nearby_route.dependant.dependencies:
        app.dependency_overrides[dependency.call] = lambda: None

    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


class TestRecruitmentMonthNormalization:
    def test_schema_normalizes_month_to_zero_padded_format(self):
        create_payload = RecruitmentVisitCreate(month="115.4", child_name="小安")
        update_payload = RecruitmentVisitUpdate(month="115.4")
        month_payload = MonthCreate(month="115.4")

        assert create_payload.month == "115.04"
        assert update_payload.month == "115.04"
        assert month_payload.month == "115.04"

    def test_import_normalizes_month_before_persisting(self, recruitment_session_factory):
        result = import_recruitment_records(
            [ImportRecord(**{"月份": "115.4", "幼生姓名": "小安"})],
            _=None,
        )

        assert result == {"inserted": 1, "skipped": 0}

        with recruitment_session_factory() as session:
            record = session.query(RecruitmentVisit).one()
            assert record.month == "115.04"

    def test_import_preserves_multiple_visits_for_same_child_in_same_month(self, recruitment_session_factory):
        result = import_recruitment_records(
            [
                ImportRecord(**{
                    "月份": "114.12",
                    "序號": "60",
                    "日期": "114.12.19媽咪參觀",
                    "幼生姓名": "范廖翊程",
                    "幼生來源": "自行蒞園",
                    "是否預繳": "否",
                    "備註": "115.07 讀中班",
                    "電訪後家長回應": "童年綠地,情緒障礙,已告知無名額",
                }),
                ImportRecord(**{
                    "月份": "114.12",
                    "序號": "69",
                    "日期": "114.12.19媽咪參觀",
                    "幼生姓名": "范廖翊程",
                    "幼生來源": "童年綠地 19-2",
                    "是否預繳": "是",
                    "備註": "115.08 讀大班",
                    "電訪後家長回應": "班導-雅婷",
                }),
            ],
            _=None,
        )

        assert result == {"inserted": 2, "skipped": 0}

        with recruitment_session_factory() as session:
            rows = (
                session.query(RecruitmentVisit)
                .filter(RecruitmentVisit.child_name == "范廖翊程")
                .order_by(RecruitmentVisit.seq_no)
                .all()
            )
            assert len(rows) == 2
            assert [row.seq_no for row in rows] == ["60", "69"]
            assert [row.has_deposit for row in rows] == [False, True]

    def test_import_uses_visit_date_month_when_month_column_conflicts(self, recruitment_session_factory):
        result = import_recruitment_records(
            [
                ImportRecord(**{
                    "月份": "114.11",
                    "日期": "114.10.23爸媽參觀",
                    "幼生姓名": "洪苡真",
                }),
            ],
            _=None,
        )

        assert result == {"inserted": 1, "skipped": 0}

        with recruitment_session_factory() as session:
            record = session.query(RecruitmentVisit).one()
            assert record.month == "114.10"
            assert record.visit_date == "114.10.23爸媽參觀"

    def test_normalize_existing_months_repairs_records_using_visit_date(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentVisit(
                    month="114.11",
                    visit_date="114.10.23爸媽參觀",
                    child_name="洪苡真",
                    has_deposit=False,
                )
            )
            session.commit()

        updated = normalize_existing_months()

        assert updated == 1

        with recruitment_session_factory() as session:
            record = session.query(RecruitmentVisit).one()
            assert record.month == "114.10"


class TestPeriodsSummary:
    def test_by_grade_only_counts_visits_within_defined_periods(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentPeriod(
                    period_name="114.09.16~115.03.15",
                    visit_count=2,
                    deposit_count=1,
                    enrolled_count=1,
                    effective_deposit_count=1,
                    sort_order=1,
                )
            )
            session.add_all([
                RecruitmentVisit(
                    month="114.09",
                    child_name="小安",
                    grade="小班",
                    has_deposit=True,
                    enrolled=True,
                ),
                RecruitmentVisit(
                    month="115.03",
                    child_name="小寶",
                    grade="中班",
                    has_deposit=False,
                    enrolled=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小晴",
                    grade="大班",
                    has_deposit=True,
                    enrolled=True,
                ),
            ])
            session.commit()

        summary = get_periods_summary(_=None)

        assert summary["total_visit"] == 2
        assert summary["by_grade"] == [
            {
                "grade": "小班",
                "visit": 1,
                "deposit": 1,
                "enrolled": 1,
                "visit_to_deposit_rate": 100.0,
                "visit_to_enrolled_rate": 100.0,
                "deposit_to_enrolled_rate": 100.0,
            },
            {
                "grade": "中班",
                "visit": 1,
                "deposit": 0,
                "enrolled": 0,
                "visit_to_deposit_rate": 0.0,
                "visit_to_enrolled_rate": 0.0,
                "deposit_to_enrolled_rate": 0,
            },
        ]

class TestRecruitmentStats:
    def test_live_stats_include_enrollment_funnel_counts_and_rates(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    has_deposit=True,
                    enrolled=True,
                    transfer_term=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    has_deposit=True,
                    enrolled=False,
                    transfer_term=False,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="小晴",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="小樂",
                    has_deposit=True,
                    enrolled=False,
                    transfer_term=True,
                ),
            ])
            session.commit()

        stats = get_recruitment_stats(_=None)

        assert stats["total_visit"] == 4
        assert stats["total_deposit"] == 3
        assert stats["total_enrolled"] == 1
        assert stats["total_transfer_term"] == 1
        assert stats["total_pending_deposit"] == 1
        assert stats["total_effective_deposit"] == 2
        assert stats["visit_to_deposit_rate"] == 75.0
        assert stats["visit_to_enrolled_rate"] == 25.0
        assert stats["deposit_to_enrolled_rate"] == pytest.approx(33.3, abs=0.1)
        assert stats["effective_to_enrolled_rate"] == 50.0
        assert stats["monthly"] == [
            {
                "month": "115.04",
                "visit": 2,
                "deposit": 2,
                "enrolled": 1,
                "transfer_term": 0,
                "pending_deposit": 1,
                "effective_deposit": 2,
                "visit_to_deposit_rate": 100.0,
                "visit_to_enrolled_rate": 50.0,
                "deposit_to_enrolled_rate": 50.0,
                "effective_to_enrolled_rate": 50.0,
                "chuannian_visit": 0,
                "chuannian_deposit": 0,
            },
            {
                "month": "115.05",
                "visit": 2,
                "deposit": 1,
                "enrolled": 0,
                "transfer_term": 1,
                "pending_deposit": 0,
                "effective_deposit": 0,
                "visit_to_deposit_rate": 50.0,
                "visit_to_enrolled_rate": 0.0,
                "deposit_to_enrolled_rate": 0.0,
                "effective_to_enrolled_rate": 0,
                "chuannian_visit": 0,
                "chuannian_deposit": 0,
            },
        ]
        assert stats["by_year"] == [
            {
                "year": "115",
                "visit": 4,
                "deposit": 3,
                "enrolled": 1,
                "transfer_term": 1,
                "pending_deposit": 1,
                "effective_deposit": 2,
                "visit_to_deposit_rate": 75.0,
                "visit_to_enrolled_rate": 25.0,
                "deposit_to_enrolled_rate": pytest.approx(33.3, abs=0.1),
                "effective_to_enrolled_rate": 50.0,
                "chuannian_visit": 0,
                "chuannian_deposit": 0,
            },
        ]

    def test_source_analysis_groups_configured_aliases_and_supports_grouped_filter(
        self,
        recruitment_session_factory,
    ):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.05",
                    child_name="童年甲",
                    source="童年綠地 19-1",
                    referrer="Ruby",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="童年乙",
                    source="二人同行",
                    referrer="Ruby",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="童年丙",
                    source="Ruby老師",
                    referrer="Amy",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="分校甲",
                    source="國際校介紹",
                    referrer="Amy",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="分校乙",
                    source="仁武校介紹",
                    referrer="Amy",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="分校丙",
                    source="崇德校介紹",
                    referrer="Amy",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="網路甲",
                    source="網路",
                    referrer="Amy",
                    has_deposit=False,
                ),
            ])
            session.commit()

        stats = get_recruitment_stats(_=None)

        by_source_map = {
            item["source"]: {"visit": item["visit"], "deposit": item["deposit"]}
            for item in stats["by_source"]
        }
        assert by_source_map == {
            "童年綠地": {"visit": 1, "deposit": 1},
            "二人同行": {"visit": 1, "deposit": 0},
            "Ruby老師": {"visit": 1, "deposit": 1},
            "分校介紹": {"visit": 3, "deposit": 1},
            "網路": {"visit": 1, "deposit": 0},
        }

        options = get_recruitment_options(_=None)
        assert "二人同行" in options["sources"]
        assert "分校介紹" in options["sources"]
        assert "崇德校介紹" not in options["sources"]

        branch_result = list_recruitment_records(
            month=None,
            grade=None,
            source="分校介紹",
            referrer=None,
            has_deposit=None,
            no_deposit_reason=None,
            keyword=None,
            page=1,
            page_size=50,
            _=None,
        )
        assert branch_result["total"] == 3
        assert {row["source"] for row in branch_result["records"]} == {
            "國際校介紹",
            "仁武校介紹",
            "崇德校介紹",
        }

        grouped_result = list_recruitment_records(
            month=None,
            grade=None,
            source="童年綠地",
            referrer=None,
            has_deposit=None,
            no_deposit_reason=None,
            keyword=None,
            page=1,
            page_size=50,
            _=None,
        )
        assert grouped_result["total"] == 1
        assert grouped_result["records"][0]["source"] == "童年綠地 19-1"

        exact_result = list_recruitment_records(
            month=None,
            grade=None,
            source="二人同行",
            referrer=None,
            has_deposit=None,
            no_deposit_reason=None,
            keyword=None,
            page=1,
            page_size=50,
            _=None,
        )
        assert exact_result["total"] == 1
        assert exact_result["records"][0]["child_name"] == "童年乙"

    def test_source_analysis_reclassifies_chuannian_keyword_hits_and_detail_filter_matches(
        self,
        recruitment_session_factory,
    ):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.05",
                    child_name="關鍵字甲",
                    source="朋友介紹",
                    referrer="Ruby",
                    has_deposit=True,
                    parent_response="童年綠地家長介紹",
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="關鍵字乙",
                    source="自行蒞園",
                    referrer="Amy",
                    has_deposit=False,
                    notes="班導-雅婷",
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="一般來源",
                    source="朋友介紹",
                    referrer="Amy",
                    has_deposit=False,
                ),
            ])
            session.commit()

        stats = get_recruitment_stats(_=None)

        assert stats["chuannian_visit"] == 2
        assert stats["chuannian_deposit"] == 1
        assert stats["by_source"] == [
            {"source": "童年綠地", "visit": 2, "deposit": 1},
            {"source": "朋友介紹", "visit": 1, "deposit": 0},
        ]

        grouped_result = list_recruitment_records(
            month=None,
            grade=None,
            source="童年綠地",
            referrer=None,
            has_deposit=None,
            no_deposit_reason=None,
            keyword=None,
            page=1,
            page_size=50,
            _=None,
        )
        assert grouped_result["total"] == 2
        assert {row["child_name"] for row in grouped_result["records"]} == {
            "關鍵字甲",
            "關鍵字乙",
        }

    def test_stats_include_decision_summary_alerts_action_queue_and_reference_month(
        self,
        recruitment_session_factory,
    ):
        now = datetime.now()
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="前月已註冊1",
                    district="三民區",
                    source="介紹",
                    has_deposit=True,
                    enrolled=True,
                    transfer_term=False,
                    created_at=now - timedelta(days=40),
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="前月已註冊2",
                    district="三民區",
                    source="介紹",
                    has_deposit=True,
                    enrolled=True,
                    transfer_term=False,
                    created_at=now - timedelta(days=39),
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="前月已預繳3",
                    district="左營區",
                    source="網路",
                    has_deposit=True,
                    enrolled=False,
                    transfer_term=False,
                    created_at=now - timedelta(days=38),
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="前月已預繳4",
                    district="左營區",
                    source="網路",
                    has_deposit=True,
                    enrolled=False,
                    transfer_term=False,
                    created_at=now - timedelta(days=37),
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="前月未預繳",
                    district="左營區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    created_at=now - timedelta(days=36),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月高潛力1",
                    district="三民區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    no_deposit_reason="時程未到／仍在觀望",
                    created_at=now - timedelta(days=20),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月高潛力2",
                    district="三民區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    no_deposit_reason="時程未到／仍在觀望",
                    created_at=now - timedelta(days=19),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月高潛力3",
                    district="三民區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    no_deposit_reason="課程／環境仍在評估",
                    created_at=now - timedelta(days=18),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月高潛力4",
                    district="三民區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    no_deposit_reason="課程／環境仍在評估",
                    created_at=now - timedelta(days=17),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月高潛力5",
                    district="三民區",
                    source="網路",
                    has_deposit=False,
                    enrolled=False,
                    transfer_term=False,
                    no_deposit_reason="時程未到／仍在觀望",
                    created_at=now - timedelta(days=16),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月其他來源1",
                    district="鳳山區",
                    source="介紹",
                    has_deposit=True,
                    enrolled=True,
                    transfer_term=False,
                    created_at=now - timedelta(days=12),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="本月其他來源2",
                    district="鳳山區",
                    source="介紹",
                    has_deposit=True,
                    enrolled=False,
                    transfer_term=False,
                    created_at=now - timedelta(days=10),
                ),
            ])
            session.commit()

        stats = get_recruitment_stats(reference_month="115.05", _=None)

        assert stats["reference_month"] == "115.05"
        assert stats["decision_summary"]["current_month"]["visit"] == 7
        assert stats["decision_summary"]["current_month"]["deposit"] == 2
        assert stats["decision_summary"]["rolling_30d"]["visit"] == 7
        assert stats["decision_summary"]["rolling_90d"]["visit"] == 12
        assert stats["decision_summary"]["ytd"]["visit"] == 12
        assert stats["funnel_snapshot"] == {
            "visit": 7,
            "deposit": 2,
            "enrolled": 1,
            "transfer_term": 0,
            "effective_deposit": 2,
            "pending_deposit": 1,
        }
        assert stats["month_over_month"]["current_month"] == "115.05"
        assert stats["month_over_month"]["previous_month"] == "115.04"
        assert stats["month_over_month"]["visit_to_deposit_rate"]["delta"] == -51.4
        assert stats["month_over_month"]["visit_to_enrolled_rate"]["delta"] == -25.7

        alert_codes = {item["code"] for item in stats["alerts"]}
        assert {"FUNNEL_DROP", "HIGH_POTENTIAL_BACKLOG", "SOURCE_IMBALANCE"} <= alert_codes
        assert any(item["target_tab"] == "nodeposit" for item in stats["alerts"])
        assert any(item["target_tab"] == "detail" for item in stats["alerts"])
        assert any(item["target_tab"] == "area" for item in stats["top_action_queue"])
        assert any(item["target_tab"] == "detail" for item in stats["top_action_queue"])
        assert any(item["target_tab"] == "nodeposit" for item in stats["top_action_queue"])

    def test_no_deposit_analysis_summary_and_priority_overdue_filters(
        self,
        recruitment_session_factory,
    ):
        now = datetime.now()
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.05",
                    child_name="高潛力逾期1",
                    grade="小班",
                    has_deposit=False,
                    no_deposit_reason="時程未到／仍在觀望",
                    created_at=now - timedelta(days=20),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="高潛力逾期2",
                    grade="小班",
                    has_deposit=False,
                    no_deposit_reason="課程／環境仍在評估",
                    created_at=now - timedelta(days=18),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="高潛力未逾期",
                    grade="中班",
                    has_deposit=False,
                    no_deposit_reason="課程／環境仍在評估",
                    created_at=now - timedelta(days=5),
                ),
                RecruitmentVisit(
                    month="115.05",
                    child_name="冷名單",
                    grade="大班",
                    has_deposit=False,
                    no_deposit_reason="已有其他就學選項／比較他校",
                    created_at=now - timedelta(days=95),
                ),
            ])
            session.commit()

        result = recruitment_api.get_no_deposit_analysis(
            priority="high",
            overdue_days=14,
            cold_only=None,
            reason=None,
            grade=None,
            page=1,
            page_size=50,
            _=None,
        )

        assert result["summary"] == {
            "high_potential_count": 3,
            "overdue_followup_count": 3,
            "cold_count": 1,
        }
        assert result["total"] == 2
        assert {row["child_name"] for row in result["records"]} == {"高潛力逾期1", "高潛力逾期2"}

    def test_stats_export_adds_decision_summary_sheet(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(month="115.04", child_name="小安", has_deposit=True, enrolled=True),
                RecruitmentVisit(month="115.05", child_name="小寶", has_deposit=False, enrolled=False),
            ])
            session.commit()

        response = export_recruitment_stats(reference_month="115.05", _=None)

        async def _collect_body():
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        body = asyncio.run(_collect_body())
        workbook = load_workbook(BytesIO(body))

        assert workbook.sheetnames[0] == "決策摘要"
        sheet = workbook["決策摘要"]
        assert sheet["A1"].value == "招生決策摘要"
        assert any("115.05" in str(sheet.cell(row=row, column=1).value or "") for row in range(1, 12))


class TestAddressHotspots:
    def test_groups_records_by_full_address_and_falls_back_to_district_from_address(
        self,
        recruitment_session_factory,
    ):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district=None,
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小晴",
                    address="高雄市鳳山區光遠路50號",
                    district="鳳山區",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小樂",
                    address="   ",
                    district="左營區",
                    has_deposit=False,
                ),
            ])
            session.commit()

        result = get_recruitment_address_hotspots(limit=10, _=None)

        assert result["records_with_address"] == 3
        assert result["total_hotspots"] == 2
        assert result["provider_available"] is True
        assert result["provider_name"] in {"nominatim", "google", "tgos"}
        assert result["geocoded_hotspots"] == 0
        assert result["pending_hotspots"] == 2
        assert result["remaining_hotspots"] == 2
        assert result["failed_hotspots"] == 0
        assert result["stale_hotspots"] == 0
        assert result["hotspots"] == [
            {
                "address": "高雄市三民區民族一路100號",
                "district": "三民區",
                "visit": 2,
                "deposit": 1,
                "lat": None,
                "lng": None,
                "geocode_status": "pending",
                "provider": None,
                "formatted_address": None,
                "matched_address": None,
                "google_place_id": None,
                "town_code": None,
                "town_name": None,
                "county_name": None,
                "land_use_label": None,
                "travel_minutes": None,
                "travel_distance_km": None,
                "data_quality": "partial",
            },
            {
                "address": "高雄市鳳山區光遠路50號",
                "district": "鳳山區",
                "visit": 1,
                "deposit": 1,
                "lat": None,
                "lng": None,
                "geocode_status": "pending",
                "provider": None,
                "formatted_address": None,
                "matched_address": None,
                "google_place_id": None,
                "town_code": None,
                "town_name": None,
                "county_name": None,
                "land_use_label": None,
                "travel_minutes": None,
                "travel_distance_km": None,
                "data_quality": "partial",
            },
        ]

    def test_address_hotspots_counts_all_cached_rows_beyond_display_limit(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    address="高雄市鳳山區光遠路50號",
                    district="鳳山區",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小晴",
                    address="高雄市左營區自由路1號",
                    district="左營區",
                    has_deposit=False,
                ),
            ])
            session.add_all([
                RecruitmentGeocodeCache(
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    provider="google",
                    status="resolved",
                    lat=22.6461,
                    lng=120.3209,
                    google_place_id="google-place-1",
                ),
                RecruitmentGeocodeCache(
                    address="高雄市鳳山區光遠路50號",
                    district="鳳山區",
                    provider="google",
                    status="resolved",
                    lat=22.6321,
                    lng=120.3565,
                    google_place_id="google-place-2",
                ),
                RecruitmentGeocodeCache(
                    address="高雄市左營區自由路1號",
                    district="左營區",
                    provider="google",
                    status="resolved",
                    lat=22.6782,
                    lng=120.3081,
                    google_place_id="google-place-3",
                ),
            ])
            session.commit()

        result = get_recruitment_address_hotspots(limit=2, _=None)

        assert len(result["hotspots"]) == 2
        assert result["total_hotspots"] == 3
        assert result["geocoded_hotspots"] == 3
        assert result["pending_hotspots"] == 0
        assert result["remaining_hotspots"] == 0
        assert result["failed_hotspots"] == 0

    def test_sync_address_hotspots_persists_geocode_cache(self, recruitment_session_factory, monkeypatch):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=True,
                )
            )
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "market_provider_available", lambda: True)
        monkeypatch.setattr(recruitment_api.market_service, "current_market_provider", lambda: "google")
        monkeypatch.setattr(
            recruitment_api.market_service,
            "resolve_address_metadata",
            lambda address, campus=None: {
                "provider": "google",
                "lat": 22.6461,
                "lng": 120.3209,
                "formatted_address": f"TW {address}",
                "matched_address": f"TW {address}",
                "google_place_id": "google-place-1",
                "town_code": "64000010",
                "town_name": "三民區",
                "county_name": "高雄市",
                "land_use_label": "住宅區",
                "travel_minutes": 8.5,
                "travel_distance_km": 3.2,
                "data_quality": "estimated",
            },
        )

        result = sync_recruitment_address_hotspots(batch_size=5, limit=10, _=None)

        assert result["sync_mode"] == "incremental"
        assert result["attempted"] == 1
        assert result["synced"] == 1
        assert result["failed"] == 0
        assert result["skipped"] == 0
        assert result["geocoded_hotspots"] == 1
        assert result["pending_hotspots"] == 0
        assert result["remaining_hotspots"] == 0
        assert result["stale_hotspots"] == 0
        assert result["hotspots"][0]["lat"] == 22.6461
        assert result["hotspots"][0]["provider"] == "google"
        assert result["hotspots"][0]["google_place_id"] == "google-place-1"
        assert result["hotspots"][0]["town_code"] == "64000010"
        assert result["hotspots"][0]["travel_minutes"] == 8.5

        with recruitment_session_factory() as session:
            cached = session.query(RecruitmentGeocodeCache).one()
            assert cached.address == "高雄市三民區民族一路100號"
            assert cached.status == "resolved"
            assert cached.lat == 22.6461
            assert cached.lng == 120.3209
            assert cached.google_place_id == "google-place-1"
            assert cached.town_code == "64000010"
            assert cached.travel_distance_km == 3.2

    def test_sync_address_hotspots_processes_all_hotspots_beyond_display_limit(
        self,
        recruitment_session_factory,
        monkeypatch,
    ):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    address="高雄市鳳山區光遠路50號",
                    district="鳳山區",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小晴",
                    address="高雄市左營區自由路1號",
                    district="左營區",
                    has_deposit=False,
                ),
            ])
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "market_provider_available", lambda: True)
        monkeypatch.setattr(recruitment_api.market_service, "current_market_provider", lambda: "google")
        monkeypatch.setattr(
            recruitment_api.market_service,
            "resolve_address_metadata",
            lambda address, campus=None: {
                "provider": "google",
                "lat": 22.6,
                "lng": 120.3,
                "formatted_address": f"Google {address}",
                "matched_address": f"Google {address}",
                "google_place_id": f"place:{address}",
                "town_code": "64000010",
                "town_name": "測試行政區",
                "county_name": "高雄市",
                "land_use_label": "住宅區",
                "travel_minutes": 9.5,
                "travel_distance_km": 4.2,
                "data_quality": "complete",
            },
        )

        result = sync_recruitment_address_hotspots(batch_size=5, limit=1, _=None)

        assert len(result["hotspots"]) == 1
        assert result["attempted"] == 3
        assert result["synced"] == 3
        assert result["failed"] == 0
        assert result["geocoded_hotspots"] == 3
        assert result["pending_hotspots"] == 0
        assert result["remaining_hotspots"] == 0

        with recruitment_session_factory() as session:
            assert session.query(RecruitmentGeocodeCache).count() == 3

    def test_address_hotspots_reports_stale_google_upgrade_count(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=True,
                )
            )
            session.add(
                RecruitmentGeocodeCache(
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    provider="nominatim",
                    status="resolved",
                    lat=22.6461,
                    lng=120.3209,
                )
            )
            session.commit()

        result = get_recruitment_address_hotspots(limit=10, _=None)

        assert result["geocoded_hotspots"] == 1
        assert result["stale_hotspots"] == 1
        assert result["hotspots"][0]["provider"] == "nominatim"
        assert result["hotspots"][0]["google_place_id"] is None

    def test_resync_google_only_targets_non_google_or_failed_cache(self, recruitment_session_factory, monkeypatch):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    address="高雄市三民區澄清路88號",
                    district="三民區",
                    has_deposit=False,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小晴",
                    address="高雄市左營區自由路1號",
                    district="左營區",
                    has_deposit=False,
                ),
            ])
            session.add_all([
                RecruitmentGeocodeCache(
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    provider="google",
                    status="resolved",
                    lat=22.6461,
                    lng=120.3209,
                    google_place_id="google-place-current",
                ),
                RecruitmentGeocodeCache(
                    address="高雄市三民區澄清路88號",
                    district="三民區",
                    provider="nominatim",
                    status="resolved",
                    lat=22.6401,
                    lng=120.3301,
                ),
                RecruitmentGeocodeCache(
                    address="高雄市左營區自由路1號",
                    district="左營區",
                    provider=None,
                    status="failed",
                ),
            ])
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "market_provider_available", lambda: True)
        monkeypatch.setattr(recruitment_api.market_service, "current_market_provider", lambda: "google")

        resolved_addresses = []

        def fake_resolve(address, campus=None):
            resolved_addresses.append(address)
            return {
                "provider": "google",
                "lat": 22.6,
                "lng": 120.3,
                "formatted_address": f"Google {address}",
                "matched_address": f"Google {address}",
                "google_place_id": f"place:{address}",
                "town_code": "64000010",
                "town_name": "三民區",
                "county_name": "高雄市",
                "land_use_label": "住宅區",
                "travel_minutes": 7.5,
                "travel_distance_km": 2.8,
                "data_quality": "complete",
            }

        monkeypatch.setattr(recruitment_api.market_service, "resolve_address_metadata", fake_resolve)

        result = sync_recruitment_address_hotspots(
            batch_size=10,
            limit=10,
            sync_mode="resync_google",
            _=None,
        )

        assert result["sync_mode"] == "resync_google"
        assert result["attempted"] == 2
        assert result["synced"] == 2
        assert result["failed"] == 0
        assert result["skipped"] == 1
        assert result["stale_hotspots"] == 0
        assert resolved_addresses == [
            "高雄市三民區澄清路88號",
            "高雄市左營區自由路1號",
        ]

        with recruitment_session_factory() as session:
            current = {
                row.address: row
                for row in session.query(RecruitmentGeocodeCache).all()
            }
            assert current["高雄市三民區民族一路100號"].google_place_id == "google-place-current"
            assert current["高雄市三民區澄清路88號"].provider == "google"
            assert current["高雄市三民區澄清路88號"].google_place_id == "place:高雄市三民區澄清路88號"
            assert current["高雄市左營區自由路1號"].provider == "google"


class TestMarketIntelligence:
    def test_campus_setting_upsert_and_query(self, recruitment_session_factory):
        payload = CampusSettingPayload(
            campus_name="本園",
            campus_address="高雄市三民區民族一路100號",
            campus_lat=22.6461,
            campus_lng=120.3209,
            travel_mode="driving",
        )

        update_recruitment_campus_setting(payload, _=None)
        result = get_recruitment_campus_setting(_=None)

        assert result["campus_name"] == "本園"
        assert result["travel_mode"] == "driving"
        assert result["campus_lat"] == 22.6461

    def test_market_intelligence_returns_raw_metrics(self, recruitment_session_factory):
        with recruitment_session_factory() as session:
            session.add_all([
                RecruitmentVisit(
                    month="115.04",
                    child_name="小安",
                    district="三民區",
                    address="高雄市三民區民族一路100號",
                    has_deposit=True,
                ),
                RecruitmentVisit(
                    month="115.04",
                    child_name="小寶",
                    district="三民區",
                    address="高雄市三民區澄清路88號",
                    has_deposit=False,
                ),
                RecruitmentGeocodeCache(
                    address="高雄市三民區民族一路100號",
                    district="三民區",
                    town_code="64000010",
                    travel_minutes=9.0,
                    status="resolved",
                ),
                RecruitmentAreaInsightCache(
                    district="三民區",
                    town_code="64000010",
                    population_density=1234.5,
                    population_0_6=456,
                    data_completeness="partial",
                ),
                RecruitmentCampusSetting(
                    campus_name="本園",
                    campus_address="高雄市三民區民族一路100號",
                    campus_lat=22.6461,
                    campus_lng=120.3209,
                    travel_mode="driving",
                ),
            ])
            session.commit()

        snapshot = get_recruitment_market_intelligence(_=None)

        assert snapshot["campus"]["campus_name"] == "本園"
        assert snapshot["districts"][0]["district"] == "三民區"
        assert snapshot["districts"][0]["lead_count_90d"] == 2
        assert snapshot["districts"][0]["deposit_rate_90d"] == 50.0
        assert snapshot["districts"][0]["population_density"] == 1234.5

    def test_market_sync_keeps_cached_snapshot_when_no_external_indexes(self, recruitment_session_factory, monkeypatch):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentAreaInsightCache(
                    district="三民區",
                    town_code="64000010",
                    population_density=1234.5,
                    data_completeness="cached",
                )
            )
            session.add(
                RecruitmentCampusSetting(
                    campus_name="本園",
                    campus_address="高雄市三民區民族一路100號",
                    campus_lat=22.6461,
                    campus_lng=120.3209,
                    travel_mode="driving",
                )
            )
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "load_population_density_index", lambda: {})
        monkeypatch.setattr(recruitment_api.market_service, "load_population_age_index", lambda: {})

        result = sync_recruitment_market_intelligence(hotspot_limit=200, _=None)

        assert "snapshot" in result
        assert result["snapshot"]["districts"][0]["data_completeness"] == "cached"


class TestNearbyKindergartens:
    def test_nearby_kindergartens_returns_deduped_places_and_query_bounds(
        self,
        recruitment_session_factory,
        monkeypatch,
    ):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentCampusSetting(
                    campus_name="本園",
                    campus_address="高雄市三民區民族一路100號",
                    campus_lat=22.6461,
                    campus_lng=120.3209,
                    travel_mode="driving",
                )
            )
            session.commit()

        requests = []

        def fake_query(payload, *, field_mask):
            requests.append((payload, field_mask))
            if payload.get("pageToken") == "token-2":
                return {
                    "places": [
                        {
                            "id": "place-2",
                            "displayName": {"text": "重複幼兒園"},
                            "formattedAddress": "高雄市三民區澄清路88號",
                            "location": {"latitude": 22.6501, "longitude": 120.3309},
                            "primaryType": "preschool",
                            "types": ["preschool", "school"],
                            "businessStatus": "OPERATIONAL",
                            "googleMapsUri": "https://maps.google.com/?cid=2",
                        },
                        {
                            "id": "place-3",
                            "displayName": {"text": "第三間幼兒園"},
                            "formattedAddress": "高雄市左營區自由一路1號",
                            "location": {"latitude": 22.6782, "longitude": 120.3081},
                            "primaryType": "preschool",
                            "types": ["preschool", "school"],
                            "businessStatus": "OPERATIONAL",
                            "googleMapsUri": "https://maps.google.com/?cid=3",
                        },
                    ]
                }

            return {
                "places": [
                    {
                        "id": "place-1",
                        "displayName": {"text": "本園旁幼兒園"},
                        "formattedAddress": "高雄市三民區民族一路100號",
                        "location": {"latitude": 22.6461, "longitude": 120.3209},
                        "primaryType": "preschool",
                        "types": ["preschool", "school"],
                        "businessStatus": "OPERATIONAL",
                        "googleMapsUri": "https://maps.google.com/?cid=1",
                    },
                    {
                        "id": "place-2",
                        "displayName": {"text": "重複幼兒園"},
                        "formattedAddress": "高雄市三民區澄清路88號",
                        "location": {"latitude": 22.6501, "longitude": 120.3309},
                        "primaryType": "preschool",
                        "types": ["preschool", "school"],
                        "businessStatus": "OPERATIONAL",
                        "googleMapsUri": "https://maps.google.com/?cid=2",
                    },
                ],
                "nextPageToken": "token-2",
            }

        monkeypatch.setattr(recruitment_api.market_service, "_google_places_api_available", lambda: True)
        monkeypatch.setattr(recruitment_api.market_service, "_query_google_places_text", fake_query)

        result = get_nearby_kindergartens(
            south=22.62,
            west=120.29,
            north=22.69,
            east=120.35,
            zoom=13,
            _=None,
        )

        assert result["provider_available"] is True
        assert result["provider_name"] == "google"
        assert result["query_bounds"] == {
            "south": 22.62,
            "west": 120.29,
            "north": 22.69,
            "east": 120.35,
            "zoom": 13,
        }
        assert result["total"] == 3
        assert [school["place_id"] for school in result["schools"]] == [
            "place-1",
            "place-2",
            "place-3",
        ]
        assert result["schools"][0]["distance_km"] == pytest.approx(0.0, abs=0.01)
        assert result["schools"][1]["distance_km"] > result["schools"][0]["distance_km"]
        assert len(requests) == 2
        assert requests[0][0]["textQuery"] == "幼兒園"
        assert requests[0][0]["includedType"] == "preschool"
        assert requests[0][0]["strictTypeFiltering"] is True
        assert requests[0][0]["locationRestriction"]["rectangle"] == {
            "low": {"latitude": 22.62, "longitude": 120.29},
            "high": {"latitude": 22.69, "longitude": 120.35},
        }
        assert requests[1][0]["pageToken"] == "token-2"

    def test_nearby_kindergartens_reports_provider_unavailable(
        self,
        recruitment_session_factory,
        monkeypatch,
    ):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentCampusSetting(
                    campus_name="本園",
                    campus_address="高雄市三民區民族一路100號",
                    campus_lat=22.6461,
                    campus_lng=120.3209,
                    travel_mode="driving",
                )
            )
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "_google_places_api_available", lambda: False)

        result = get_nearby_kindergartens(
            south=22.62,
            west=120.29,
            north=22.69,
            east=120.35,
            zoom=13,
            _=None,
        )

        assert result["provider_available"] is False
        assert result["provider_name"] == "google"
        assert result["total"] == 0
        assert result["schools"] == []
        assert "Places API" in result["message"]

    def test_nearby_kindergartens_rejects_invalid_query_bounds(self, recruitment_client):
        response = recruitment_client.get(
            "/api/recruitment/nearby-kindergartens",
            params={
                "south": 95,
                "west": 120.29,
                "north": 22.69,
                "east": 120.35,
                "zoom": 13,
            },
        )

        assert response.status_code == 422

    def test_nearby_kindergartens_distance_uses_campus_center(
        self,
        recruitment_session_factory,
        monkeypatch,
    ):
        with recruitment_session_factory() as session:
            session.add(
                RecruitmentCampusSetting(
                    campus_name="本園",
                    campus_address="高雄市三民區民族一路100號",
                    campus_lat=22.6461,
                    campus_lng=120.3209,
                    travel_mode="driving",
                )
            )
            session.commit()

        monkeypatch.setattr(recruitment_api.market_service, "_google_places_api_available", lambda: True)
        monkeypatch.setattr(
            recruitment_api.market_service,
            "_query_google_places_text",
            lambda payload, *, field_mask: {
                "places": [
                    {
                        "id": "place-1",
                        "displayName": {"text": "同位置幼兒園"},
                        "formattedAddress": "高雄市三民區民族一路100號",
                        "location": {"latitude": 22.6461, "longitude": 120.3209},
                        "primaryType": "preschool",
                        "types": ["preschool"],
                        "businessStatus": "OPERATIONAL",
                        "googleMapsUri": "https://maps.google.com/?cid=1",
                    },
                    {
                        "id": "place-2",
                        "displayName": {"text": "另一間幼兒園"},
                        "formattedAddress": "高雄市三民區建國一路10號",
                        "location": {"latitude": 22.6401, "longitude": 120.3159},
                        "primaryType": "preschool",
                        "types": ["preschool"],
                        "businessStatus": "OPERATIONAL",
                        "googleMapsUri": "https://maps.google.com/?cid=2",
                    },
                ]
            },
        )

        result = get_nearby_kindergartens(
            south=22.62,
            west=120.29,
            north=22.69,
            east=120.35,
            zoom=13,
            _=None,
        )

        assert result["schools"][0]["distance_km"] == pytest.approx(0.0, abs=0.01)
        assert result["schools"][1]["distance_km"] > 0
