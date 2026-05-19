"""Excel writer 整合測試（讀回 xlsx 驗證內容）。"""

from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from services.gov_moe.monthly_excel_writer import build_monthly_xlsx_bytes


@pytest.fixture
def sample_rows():
    return [
        {
            "classroom_id": 1,
            "classroom_name": "蘋果班",
            "teacher_names": "張老師",
            "age_group": "4-5",
            "total_count": 5,
            "male_count": 3,
            "female_count": 2,
            "disadvantaged_count": 1,
            "disability_count": 0,
            "indigenous_count": 0,
            "foreign_count": 0,
            "expected_attendance_days": 100,
            "actual_attendance_days": 95,
            "attendance_rate": 9500,
        },
    ]


@pytest.fixture
def sample_student_details():
    return [
        {
            "student_id": 1,
            "student_no": "S001",
            "name": "王小明",
            "id_number": "A123456789",
            "classroom_name": "蘋果班",
            "age_group": "4-5",
            "expected_days": 22,
            "actual_days": 20,
            "attendance_rate_pct": 90.91,
            "is_disadvantaged": False,
        },
    ]


@pytest.fixture
def sample_overview():
    return {
        "year": 2026,
        "month": 5,
        "snapshot_date": date(2026, 5, 31),
        "generated_at": datetime(2026, 6, 1, 10, 23),
        "generated_by": "test@example.com",
        "total_students": 28,
        "by_age_group": {"2-3": 0, "3-4": 8, "4-5": 12, "5-6": 8},
        "disadvantaged_pct": 7.14,
        "disability_pct": 3.57,
        "indigenous_pct": 0.0,
        "foreign_pct": 0.0,
        "total_expected_days": 1300,
        "total_actual_days": 1238,
        "total_attendance_rate_pct": 95.23,
    }


def test_xlsx_has_three_sheets(sample_rows, sample_student_details, sample_overview):
    data = build_monthly_xlsx_bytes(
        sample_rows, sample_student_details, sample_overview
    )
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["班級總表", "幼生明細", "統計摘要"]


def test_sheet1_classroom_summary_headers(
    sample_rows, sample_student_details, sample_overview
):
    data = build_monthly_xlsx_bytes(
        sample_rows, sample_student_details, sample_overview
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["班級總表"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "班級",
        "教師",
        "年齡層",
        "應到人日",
        "實到人日",
        "出席率",
        "男",
        "女",
        "弱勢",
        "身障",
        "原民",
        "外籍",
    ]


def test_sheet1_has_total_row(sample_rows, sample_student_details, sample_overview):
    data = build_monthly_xlsx_bytes(
        sample_rows, sample_student_details, sample_overview
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["班級總表"]
    last_row = list(ws.rows)[-1]
    assert last_row[0].value == "合計"


def test_sheet2_student_headers(sample_rows, sample_student_details, sample_overview):
    data = build_monthly_xlsx_bytes(
        sample_rows, sample_student_details, sample_overview
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["幼生明細"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "學號",
        "姓名",
        "身分證",
        "班級",
        "年齡層",
        "應到日數",
        "實到日數",
        "出席率",
        "弱勢標記",
    ]


def test_sheet3_overview_contains_total_students(
    sample_rows, sample_student_details, sample_overview
):
    data = build_monthly_xlsx_bytes(
        sample_rows, sample_student_details, sample_overview
    )
    wb = load_workbook(BytesIO(data))
    ws = wb["統計摘要"]
    cells = [c.value for row in ws.rows for c in row]
    assert "總人數" in cells
    assert 28 in cells


def test_empty_rows_does_not_raise(sample_overview):
    data = build_monthly_xlsx_bytes([], [], sample_overview)
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["班級總表", "幼生明細", "統計摘要"]


def test_id_number_none_displayed_as_dash(sample_rows, sample_overview):
    details = [
        {
            "student_id": 1,
            "student_no": "S001",
            "name": "王小明",
            "id_number": None,
            "classroom_name": "蘋果班",
            "age_group": "4-5",
            "expected_days": 22,
            "actual_days": 20,
            "attendance_rate_pct": 90.91,
            "is_disadvantaged": False,
        }
    ]
    data = build_monthly_xlsx_bytes(sample_rows, details, sample_overview)
    wb = load_workbook(BytesIO(data))
    ws = wb["幼生明細"]
    row2 = list(ws.rows)[1]
    assert row2[2].value == "-"  # 身分證欄位
