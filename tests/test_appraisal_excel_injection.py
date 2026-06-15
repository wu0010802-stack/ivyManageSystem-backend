"""考核 Excel 匯出須過 SafeWorksheet 防公式注入（bug hunt 2026-06-15 P1#1）。

export_half_year_xlsx / export_transfer_roster_xlsx 原直接用 wb.active 寫入，
員工姓名 / 戶名 / 事假病假備註 / 不計算考核原因若含 =cmd|'/C calc'!A0，
會在財會端 Excel 開啟時被當公式執行（DDE / HYPERLINK）。轉帳名冊更與撥款直接相關。

修法：改用 SafeWorksheet(wb.active)，與 year_end/excel_io.py、transfer_roster.py、
gov_reports 等既有匯出對齊（appraisal/excel_io.py 原為全 codebase 唯一漏網的 xlsx 匯出器）。
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from models.appraisal import Grade, Semester
from services.appraisal.excel_io import (
    ExportRow,
    TransferRow,
    export_half_year_xlsx,
    export_transfer_roster_xlsx,
)

EVIL = "=cmd|'/C calc'!A0"


def _all_values(payload: bytes) -> list:
    ws = load_workbook(BytesIO(payload)).active
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


def _make_export_row(**overrides) -> ExportRow:
    base = dict(
        name="王測試",
        score_items={},
        total_score=Decimal("100"),
        grade=Grade.GOOD,
        bonus_amount=Decimal("5000"),
        leave_note=None,
        is_excluded=False,
        exclude_reason=None,
    )
    base.update(overrides)
    return ExportRow(**base)


def _export_half_year(rows):
    return export_half_year_xlsx(
        title="114(上)年度考核統計表",
        academic_year=114,
        semester=Semester.FIRST,
        base_score=Decimal("80"),
        rows=rows,
    )


def test_half_year_xlsx_sanitizes_name_injection():
    payload = _export_half_year([_make_export_row(name=EVIL)])
    vals = _all_values(payload)
    assert EVIL not in vals, "姓名公式注入未被清理（原樣寫入 → 財會端可被執行）"
    assert ("'" + EVIL) in vals, "姓名應被 sanitize 為 ' 前綴純字串"


def test_half_year_xlsx_sanitizes_leave_note_injection():
    payload = _export_half_year([_make_export_row(leave_note=EVIL)])
    vals = _all_values(payload)
    assert EVIL not in vals, "事假/病假備註公式注入未被清理"
    assert ("'" + EVIL) in vals


def test_half_year_xlsx_sanitizes_exclude_reason_injection():
    payload = _export_half_year(
        [_make_export_row(is_excluded=True, exclude_reason=EVIL)]
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "不計算考核原因公式注入未被清理"
    assert ("'" + EVIL) in vals


def test_transfer_roster_xlsx_sanitizes_name_injection():
    payload = export_transfer_roster_xlsx(
        rows=[TransferRow(bank_account="0123456789", name=EVIL, amount=Decimal("1000"))]
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "戶名公式注入未被清理（轉帳名冊 → 財會端可被執行）"
    assert ("'" + EVIL) in vals, "戶名應被 sanitize 為 ' 前綴純字串"
