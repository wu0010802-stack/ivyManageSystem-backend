"""
Attendance anomalies - admin batch view, batch confirm, and export endpoints.
"""

import io
import calendar as cal_module
import logging
from datetime import date, datetime
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from utils.errors import raise_safe_500
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pydantic import BaseModel

from sqlalchemy import or_

from models.database import get_session, Employee, Attendance, SalaryRecord
from services.salary.utils import calc_daily_salary
from utils.auth import require_staff_permission
from utils.permissions import Permission

logger = logging.getLogger(__name__)

router = APIRouter()

# ============ Excel 樣式（複用 exports.py 的定義） ============

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center")

# Excel 公式注入防護：統一由 utils.excel_utils 提供
from utils.excel_utils import (
    SafeWorksheet,
    sanitize_excel_value as _sanitize_excel_value,
)

# ============ 確認動作標籤映射 ============

ACTION_LABELS = {
    "accept": "接受扣款",
    "admin_accept": "接受扣款",
    "use_pto": "特休抵銷",
    "dispute": "申訴中",
    "admin_waive": "管理員豁免",
}

WEEKDAY_NAMES = ["一", "二", "三", "四", "五", "六", "日"]

# ============ Pydantic Models ============


class BatchConfirmRequest(BaseModel):
    attendance_ids: List[int]
    action: str  # "admin_accept" | "admin_waive"
    remark: Optional[str] = None


# ============ 共用輔助：建立異常列表 ============


def _build_anomaly_rows(session, year: int, month: int, status_filter: str):
    """查詢指定月份所有員工的異常記錄，回傳 list[dict]"""
    _, last_day = cal_module.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, last_day)

    # 在 SQL 層過濾：只取有異常的記錄，大幅減少資料傳輸
    query = (
        session.query(Attendance, Employee)
        .join(Employee, Attendance.employee_id == Employee.id)
        .filter(
            Attendance.attendance_date >= start,
            Attendance.attendance_date <= end,
            Employee.is_active == True,
            or_(
                Attendance.is_late == True,
                Attendance.is_early_leave == True,
                Attendance.is_missing_punch_in == True,
                Attendance.is_missing_punch_out == True,
            ),
        )
    )

    # 狀態篩選也在 SQL 層完成
    if status_filter == "pending":
        query = query.filter(Attendance.confirmed_action == None)
    elif status_filter == "confirmed":
        query = query.filter(Attendance.confirmed_action != None)

    records = query.all()

    rows = []
    for att, emp in records:
        daily_salary = calc_daily_salary(emp.base_salary)

        items = []
        if att.is_late and att.late_minutes and att.late_minutes > 0:
            deduction = round(daily_salary / 8 / 60 * att.late_minutes)
            items.append(
                {
                    "type": "late",
                    "type_label": "遲到",
                    "detail": f"遲到 {att.late_minutes} 分鐘",
                    "estimated_deduction": deduction,
                }
            )
        if att.is_early_leave:
            items.append(
                {
                    "type": "early_leave",
                    "type_label": "早退",
                    "detail": "早退",
                    "estimated_deduction": 50,
                }
            )
        if att.is_missing_punch_in:
            items.append(
                {
                    "type": "missing_punch",
                    "type_label": "未打卡(上班)",
                    "detail": "上班未打卡（不扣款，僅記錄）",
                    "estimated_deduction": 0,
                }
            )
        if att.is_missing_punch_out:
            items.append(
                {
                    "type": "missing_punch",
                    "type_label": "未打卡(下班)",
                    "detail": "下班未打卡（不扣款，僅記錄）",
                    "estimated_deduction": 0,
                }
            )

        for item in items:
            rows.append(
                {
                    "id": att.id,
                    "employee_name": emp.name,
                    "employee_number": emp.employee_number or "",
                    "date": att.attendance_date.isoformat(),
                    "weekday": WEEKDAY_NAMES[att.attendance_date.weekday()],
                    "confirmed_action": att.confirmed_action,
                    "confirmed_by": att.confirmed_by,
                    "confirmed_at": (
                        att.confirmed_at.isoformat() if att.confirmed_at else None
                    ),
                    **item,
                }
            )

    return rows


# ============ Endpoints ============


@router.get("/anomalies")
def get_attendance_anomalies(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    status: str = Query("all", pattern="^(all|pending|confirmed)$"),
    current_user: dict = Depends(require_staff_permission(Permission.ATTENDANCE_READ)),
):
    """查詢月份異常清單（所有員工）"""
    session = get_session()
    try:
        rows = _build_anomaly_rows(session, year, month, status)
        total = len(rows)
        pending = sum(1 for r in rows if r["confirmed_action"] is None)
        return {
            "total": total,
            "pending": pending,
            "confirmed": total - pending,
            "items": rows,
        }
    finally:
        session.close()


@router.post("/anomalies/batch-confirm")
def batch_confirm_anomalies(
    data: BatchConfirmRequest,
    current_user: dict = Depends(require_staff_permission(Permission.ATTENDANCE_WRITE)),
):
    """批次確認異常處理方式（管理員代確認）"""
    if data.action not in ("admin_accept", "admin_waive"):
        raise HTTPException(
            status_code=400,
            detail="無效的批次確認動作，允許值：admin_accept、admin_waive",
        )

    operator = current_user.get("username", "admin")
    session = get_session()
    try:
        processed = 0
        now = datetime.now()
        att_map = {
            a.id: a
            for a in session.query(Attendance)
            .filter(Attendance.id.in_(data.attendance_ids))
            .all()
        }

        # 封存守衛：admin_waive 會改變薪資扣款結果（薪資端視為不扣），
        # 已封存月份不可被改寫；admin_accept 雖不改薪資結果，但仍代表
        # 對該月 attendance 狀態的事後變更，封存後若需修正應先解封。
        # 任一目標 (emp, year, month) 已封存 → 整批 409，避免部份成功部份阻擋。
        target_months = {
            (a.employee_id, a.attendance_date.year, a.attendance_date.month)
            for a in att_map.values()
            if a.attendance_date
        }
        if target_months:
            from sqlalchemy import and_, or_

            finalized_rows = (
                session.query(
                    SalaryRecord.employee_id,
                    SalaryRecord.salary_year,
                    SalaryRecord.salary_month,
                    SalaryRecord.finalized_by,
                )
                .filter(
                    SalaryRecord.is_finalized == True,
                    or_(
                        *(
                            and_(
                                SalaryRecord.employee_id == eid,
                                SalaryRecord.salary_year == y,
                                SalaryRecord.salary_month == m,
                            )
                            for eid, y, m in target_months
                        )
                    ),
                )
                .all()
            )
            if finalized_rows:
                detail = ", ".join(
                    f"員工#{r.employee_id} {r.salary_year}/{r.salary_month:02d}"
                    f"（結算人：{r.finalized_by or '系統'}）"
                    for r in finalized_rows
                )
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"下列月份薪資已封存，無法批次處理考勤異常：{detail}。"
                        "請先至薪資管理頁面解除封存後再操作。"
                    ),
                )

        # 收集需重算薪資的 (employee_id, year, month)：admin_waive 改變薪資扣款結果
        # （薪資端會把 waive 視為不扣），未封存的薪資需標 stale 觸發重算
        salary_recalc_keys: set = set()
        for att_id in data.attendance_ids:
            att = att_map.get(att_id)
            if not att:
                continue
            att.confirmed_action = data.action
            att.confirmed_by = operator
            att.confirmed_at = now
            if data.remark:
                att.remark = (
                    att.remark or ""
                ) + f" [批次{ACTION_LABELS[data.action]}: {data.remark}]"
            processed += 1
            if data.action == "admin_waive" and att.attendance_date:
                salary_recalc_keys.add(
                    (
                        att.employee_id,
                        att.attendance_date.year,
                        att.attendance_date.month,
                    )
                )

        if salary_recalc_keys:
            from services.salary.utils import mark_salary_stale

            for emp_id, year, month in salary_recalc_keys:
                mark_salary_stale(session, emp_id, year, month)

        session.commit()
        logger.warning(
            "批次確認考勤異常：操作者=%s action=%s 筆數=%d",
            operator,
            data.action,
            processed,
        )
        return {"processed": processed}
    except HTTPException:
        # 封存守衛、無效 action 等業務錯誤需保留原 status code,別被 500 蓋掉
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.get("/anomalies/export")
def export_attendance_anomalies(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    status: str = Query("all", pattern="^(all|pending|confirmed)$"),
    current_user: dict = Depends(require_staff_permission(Permission.ATTENDANCE_READ)),
):
    """匯出考勤異常 Excel"""
    session = get_session()
    try:
        rows = _build_anomaly_rows(session, year, month, status)
    finally:
        session.close()

    wb = Workbook()
    ws_raw = wb.active
    ws = SafeWorksheet(ws_raw)
    ws.title = f"{year}年{month}月考勤異常"

    headers = [
        "員工編號",
        "姓名",
        "日期",
        "星期",
        "異常類型",
        "明細",
        "預估扣款",
        "確認狀態",
        "確認動作",
        "確認人員",
        "確認時間",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    for row_idx, item in enumerate(rows, 2):
        is_confirmed = item["confirmed_action"] is not None
        action_label = (
            ACTION_LABELS.get(item["confirmed_action"] or "", "")
            if is_confirmed
            else ""
        )
        ws.cell(row=row_idx, column=1, value=item["employee_number"])
        ws.cell(row=row_idx, column=2, value=item["employee_name"])
        ws.cell(row=row_idx, column=3, value=item["date"])
        ws.cell(row=row_idx, column=4, value=item["weekday"])
        ws.cell(row=row_idx, column=5, value=item["type_label"])
        ws.cell(row=row_idx, column=6, value=item["detail"])
        ws.cell(row=row_idx, column=7, value=item["estimated_deduction"])
        ws.cell(row=row_idx, column=8, value="已處理" if is_confirmed else "待處理")
        ws.cell(row=row_idx, column=9, value=action_label)
        ws.cell(row=row_idx, column=10, value=item["confirmed_by"] or "")
        ws.cell(row=row_idx, column=11, value=item["confirmed_at"] or "")

    # 自動欄寬
    for col in ws_raw.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_raw.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{year}年{month}月考勤異常.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
