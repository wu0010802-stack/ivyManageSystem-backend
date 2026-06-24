"""才藝老師薪資 service：加總、清冊匯出。

清冊 Excel 對齊《義華薪資》才藝老師 sheet 排版：
    Row1: 「常春藤 {民國年}.{月:02d}月才藝」
    Row2: 表頭（科目/姓名/時數/每鐘點費/小計/超額/加給活動/總計）
    Row3+: 每 entry 一行（依員工編號 → entry id 排序）
    末: 合計
"""

from __future__ import annotations

import io
import logging
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from models.database import ArtTeacherPayrollEntry, Employee
from utils.excel_utils import SafeWorksheet
from utils.rounding import round_half_up

logger = logging.getLogger(__name__)


def compute_total_for_month(
    session: Session, employee_id: int, year: int, month: int
) -> float:
    """加總某員工某月所有 entry 的 total_amount。"""
    result = (
        session.query(func.coalesce(func.sum(ArtTeacherPayrollEntry.total_amount), 0))
        .filter(
            ArtTeacherPayrollEntry.employee_id == employee_id,
            ArtTeacherPayrollEntry.salary_year == year,
            ArtTeacherPayrollEntry.salary_month == month,
        )
        .scalar()
    )
    return float(result or 0)


def compute_totals_by_emp_for_month(
    session: Session, year: int, month: int
) -> dict[int, float]:
    """批次版 compute_total_for_month：一次查回 {employee_id: sum(total_amount)}。

    供批次薪資（process_bulk_salary_calculation）預載使用，避免 per-employee N+1。
    與單筆 compute_total_for_month 同口徑（同 filter、同 coalesce）。
    """
    rows = (
        session.query(
            ArtTeacherPayrollEntry.employee_id,
            func.coalesce(func.sum(ArtTeacherPayrollEntry.total_amount), 0),
        )
        .filter(
            ArtTeacherPayrollEntry.salary_year == year,
            ArtTeacherPayrollEntry.salary_month == month,
        )
        .group_by(ArtTeacherPayrollEntry.employee_id)
        .all()
    )
    return {emp_id: float(total or 0) for emp_id, total in rows}


def list_entries_for_month(
    session: Session, year: int, month: int
) -> list[tuple[Employee, ArtTeacherPayrollEntry]]:
    """按工號 → entry id 排序，回傳 (employee, entry) 列表。"""
    rows = (
        session.query(Employee, ArtTeacherPayrollEntry)
        .join(ArtTeacherPayrollEntry, ArtTeacherPayrollEntry.employee_id == Employee.id)
        .filter(
            ArtTeacherPayrollEntry.salary_year == year,
            ArtTeacherPayrollEntry.salary_month == month,
        )
        .order_by(Employee.employee_id, ArtTeacherPayrollEntry.id)
        .all()
    )
    return rows


def recompute_entry_amounts(entry: ArtTeacherPayrollEntry) -> None:
    """根據 hours × rate + excess + activity 重算 base_amount 與 total_amount。

    讓 API 端統一呼叫，避免前端送錯 total。
    """
    base = round_half_up(float(entry.hours or 0) * float(entry.hourly_rate or 0))
    entry.base_amount = base
    entry.total_amount = (
        base + float(entry.excess_amount or 0) + float(entry.activity_bonus or 0)
    )


def generate_art_teacher_roster_xlsx(
    session: Session, year: int, month: int
) -> tuple[str, bytes]:
    """產生才藝老師薪資清冊 xlsx，回傳 (filename, bytes)。"""
    roc_year = year - 1911
    rows = list_entries_for_month(session, year, month)

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "才藝老師薪資"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws["A1"] = f"常春藤 {roc_year}.{month:02d} 月才藝"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "科目",
        "姓名",
        "時數",
        "每鐘點費",
        "小計",
        "超額",
        "加給活動",
        "總計",
        "備註",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = bold
        c.alignment = center

    row_idx = 3
    total_base = 0
    total_excess = 0
    total_activity = 0
    total_all = 0
    for emp, entry in rows:
        subject_label = entry.subject
        if entry.classroom_label:
            subject_label = f"{entry.subject}{entry.classroom_label}"
        ws.cell(row=row_idx, column=1, value=subject_label)
        ws.cell(row=row_idx, column=2, value=emp.name)
        ws.cell(row=row_idx, column=3, value=float(entry.hours or 0))
        ws.cell(row=row_idx, column=4, value=float(entry.hourly_rate or 0))
        ws.cell(row=row_idx, column=5, value=float(entry.base_amount or 0))
        excess = float(entry.excess_amount or 0)
        activity = float(entry.activity_bonus or 0)
        ws.cell(row=row_idx, column=6, value=excess if excess else None)
        ws.cell(row=row_idx, column=7, value=activity if activity else None)
        ws.cell(row=row_idx, column=8, value=float(entry.total_amount or 0))
        ws.cell(row=row_idx, column=9, value=entry.note or "")
        total_base += float(entry.base_amount or 0)
        total_excess += excess
        total_activity += activity
        total_all += float(entry.total_amount or 0)
        row_idx += 1

    # 合計行
    ws.cell(row=row_idx, column=1, value="合計").font = bold
    ws.cell(row=row_idx, column=5, value=total_base).font = bold
    if total_excess:
        ws.cell(row=row_idx, column=6, value=total_excess).font = bold
    if total_activity:
        ws.cell(row=row_idx, column=7, value=total_activity).font = bold
    ws.cell(row=row_idx, column=8, value=total_all).font = bold

    # 欄寬
    raw_ws = wb.active
    raw_ws.column_dimensions["A"].width = 18
    raw_ws.column_dimensions["B"].width = 14
    for col in ("C", "D", "E", "F", "G", "H"):
        raw_ws.column_dimensions[col].width = 12
    raw_ws.column_dimensions["I"].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = f"art_teacher_payroll_{year}_{month:02d}.xlsx"
    return filename, buffer.getvalue()
