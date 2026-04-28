"""IDOR audit Phase 2：薪資欄位級洩漏修補（F-012/13/14/31/36）。

涵蓋 5 個 finding：
- F-012：GET /employees/{id}/final-salary-preview — 缺 self_or_full_salary
- F-013：GET /salaries/festival-bonus、/period-accrual — 跨員工彙總端點
- F-014：GET /employees/{id}/contracts — salary_at_contract 未遮罩
- F-031：GET /reports/finance-summary/detail|export — 逐員薪資對 supervisor 外洩
- F-036：GET /exports/overtimes — overtime_pay 對 OVERTIME_READ 持有者外洩

Adversarial 案例：自訂角色（非 admin/hr）即使持 SALARY_READ + REPORTS / OVERTIME_READ，
仍應被遮罩 — 對應守衛是「角色」而非單純權限位元。
"""

import os
import sys
from datetime import date, datetime, time, timedelta
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.auth import router as auth_router, _account_failures, _ip_attempts
import api.salary as salary_module
from api.salary import router as salary_router
from api.employees import router as employees_router
from api.employees import init_employee_services
from api.employees_docs import router as employees_docs_router
from api.reports import router as reports_router
from api.exports import router as exports_router
from models.database import (
    Base,
    Employee,
    EmployeeContract,
    OvertimeRecord,
    SalaryRecord,
    User,
)
from utils.auth import hash_password
from utils.permissions import Permission

# ─────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────


@pytest.fixture
def field_leak_client(tmp_path):
    """建立隔離的 sqlite 測試 app，包含所有 5 個受影響 router。"""
    db_path = tmp_path / "field-leak.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}", connect_args={"check_same_thread": False}
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    fake_salary_engine = MagicMock()
    # final-salary-preview 會呼叫 preview_salary_calculation
    breakdown = MagicMock()
    breakdown.base_salary = 30000
    breakdown.festival_bonus = 0
    breakdown.gross_salary = 30000
    breakdown.total_deduction = 0
    breakdown.labor_insurance = 0
    breakdown.health_insurance = 0
    breakdown.pension_self = 0
    breakdown.net_salary = 30000
    fake_salary_engine.preview_salary_calculation.return_value = breakdown
    fake_insurance_service = MagicMock()
    salary_module.init_salary_services(fake_salary_engine, fake_insurance_service)
    init_employee_services(fake_salary_engine)

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)
    app.include_router(employees_router)
    app.include_router(employees_docs_router)
    app.include_router(reports_router)
    app.include_router(exports_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _create_user(
    session,
    *,
    username,
    password="Pass1234",
    role,
    permissions,
    employee_id=None,
):
    user = User(
        employee_id=employee_id,
        username=username,
        password_hash=hash_password(password),
        role=role,
        permissions=int(permissions),
        is_active=True,
        must_change_password=False,
    )
    session.add(user)
    session.flush()
    return user


def _login(client, username, password="Pass1234"):
    res = client.post(
        "/api/auth/login", json={"username": username, "password": password}
    )
    assert res.status_code == 200, res.text


def _create_employee(
    session, employee_id_str: str, name: str, base_salary: int = 30000
) -> Employee:
    emp = Employee(
        employee_id=employee_id_str,
        name=name,
        base_salary=base_salary,
        hire_date=date(2024, 1, 1),
        is_active=True,
    )
    session.add(emp)
    session.flush()
    return emp


# ─────────────────────────────────────────────────────────────────────────
# F-012：GET /employees/{id}/final-salary-preview
# ─────────────────────────────────────────────────────────────────────────


class TestF012_FinalSalaryPreview:
    """F-012：非 admin/hr 不可看他人 final-salary-preview。"""

    def test_non_admin_calls_other_employee_403(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            self_emp = _create_employee(s, "T_self", "本人")
            other_emp = _create_employee(s, "T_other", "他人")
            _create_user(
                s,
                username="sv_self1",
                role="supervisor",
                permissions=int(Permission.SALARY_READ),
                employee_id=self_emp.id,
            )
            s.commit()
            other_id = other_emp.id

        _login(client, "sv_self1")
        res = client.get(
            f"/api/employees/{other_id}/final-salary-preview?year=2026&month=4"
        )
        assert res.status_code == 403, res.text

    def test_non_admin_calls_self_200(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            self_emp = _create_employee(s, "T_self2", "本人2")
            # 使用 supervisor 角色（非 admin/hr 但能透過 require_staff_permission 檢查）
            # 模擬主管被臨時授予 SALARY_READ 用於查自己歷史薪資的場景。
            _create_user(
                s,
                username="sv_self2",
                role="supervisor",
                permissions=int(Permission.SALARY_READ),
                employee_id=self_emp.id,
            )
            s.commit()
            self_id = self_emp.id

        _login(client, "sv_self2")
        res = client.get(
            f"/api/employees/{self_id}/final-salary-preview?year=2026&month=4"
        )
        assert res.status_code == 200, res.text

    def test_admin_calls_any_200(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            target = _create_employee(s, "T_anyA", "目標A")
            _create_user(
                s,
                username="admin_a",
                role="admin",
                permissions=-1,
            )
            s.commit()
            target_id = target.id

        _login(client, "admin_a")
        res = client.get(
            f"/api/employees/{target_id}/final-salary-preview?year=2026&month=4"
        )
        assert res.status_code == 200, res.text

    def test_hr_calls_any_200(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            target = _create_employee(s, "T_anyH", "目標H")
            _create_user(
                s,
                username="hr_a",
                role="hr",
                permissions=int(Permission.SALARY_READ),
            )
            s.commit()
            target_id = target.id

        _login(client, "hr_a")
        res = client.get(
            f"/api/employees/{target_id}/final-salary-preview?year=2026&month=4"
        )
        assert res.status_code == 200, res.text


# ─────────────────────────────────────────────────────────────────────────
# F-013：GET /salaries/festival-bonus[/period-accrual]
# ─────────────────────────────────────────────────────────────────────────


class TestF013_FestivalBonus:
    """F-013：跨員工彙總端點限縮為 admin/hr。"""

    def test_non_admin_with_salary_read_403_on_festival_bonus(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            emp = _create_employee(s, "T_fb1", "員工FB1")
            # supervisor + SALARY_READ：通過 staff gate，但仍應被 enforce_full_salary_view 擋下
            _create_user(
                s,
                username="sv_fb",
                role="supervisor",
                permissions=int(Permission.SALARY_READ),
                employee_id=emp.id,
            )
            s.commit()

        _login(client, "sv_fb")
        res = client.get("/api/salaries/festival-bonus?year=2026&month=4")
        assert res.status_code == 403, res.text

    def test_non_admin_with_salary_read_403_on_period_accrual(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            emp = _create_employee(s, "T_fb2", "員工FB2")
            _create_user(
                s,
                username="sv_fb2",
                role="supervisor",
                permissions=int(Permission.SALARY_READ),
                employee_id=emp.id,
            )
            s.commit()

        _login(client, "sv_fb2")
        res = client.get(
            "/api/salaries/festival-bonus/period-accrual?year=2026&month=4"
        )
        assert res.status_code == 403, res.text

    def test_hr_200_on_both(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _create_user(
                s,
                username="hr_fb",
                role="hr",
                permissions=int(Permission.SALARY_READ),
            )
            s.commit()

        _login(client, "hr_fb")
        res1 = client.get("/api/salaries/festival-bonus?year=2026&month=4")
        assert res1.status_code == 200, res1.text
        res2 = client.get(
            "/api/salaries/festival-bonus/period-accrual?year=2026&month=4"
        )
        assert res2.status_code == 200, res2.text

    def test_admin_200_on_both(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _create_user(s, username="admin_fb", role="admin", permissions=-1)
            s.commit()

        _login(client, "admin_fb")
        res1 = client.get("/api/salaries/festival-bonus?year=2026&month=4")
        assert res1.status_code == 200, res1.text
        res2 = client.get(
            "/api/salaries/festival-bonus/period-accrual?year=2026&month=4"
        )
        assert res2.status_code == 200, res2.text


# ─────────────────────────────────────────────────────────────────────────
# F-014：GET /employees/{id}/contracts
# ─────────────────────────────────────────────────────────────────────────


class TestF014_Contracts:
    """F-014：非 admin/hr 且非 self 看他人合約時 salary_at_contract 應遮罩。"""

    def _seed(self, s):
        self_emp = _create_employee(s, "C_self", "本人C")
        other_emp = _create_employee(s, "C_other", "他人C")
        # 給 other_emp 建一筆合約
        s.add(
            EmployeeContract(
                employee_id=other_emp.id,
                contract_type="正式",
                start_date=date(2024, 1, 1),
                salary_at_contract=42000.0,
            )
        )
        # 給 self_emp 也建一筆合約
        s.add(
            EmployeeContract(
                employee_id=self_emp.id,
                contract_type="正式",
                start_date=date(2024, 1, 1),
                salary_at_contract=35000.0,
            )
        )
        s.flush()
        return self_emp, other_emp

    def test_non_admin_lists_other_contracts_salary_masked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            self_emp, other_emp = self._seed(s)
            # 用 supervisor（非 admin/hr）模擬「人資助理」角色，須透過 require_staff_permission gate
            _create_user(
                s,
                username="sv_emp_viewer",
                role="supervisor",
                permissions=int(Permission.EMPLOYEES_READ),
                employee_id=self_emp.id,
            )
            s.commit()
            other_id = other_emp.id

        _login(client, "sv_emp_viewer")
        res = client.get(f"/api/employees/{other_id}/contracts")
        assert res.status_code == 200, res.text
        rows = res.json()
        assert len(rows) == 1
        assert rows[0]["salary_at_contract"] is None

    def test_self_lists_own_contracts_salary_present(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            self_emp, _ = self._seed(s)
            _create_user(
                s,
                username="sv_emp_self",
                role="supervisor",
                permissions=int(Permission.EMPLOYEES_READ),
                employee_id=self_emp.id,
            )
            s.commit()
            self_id = self_emp.id

        _login(client, "sv_emp_self")
        res = client.get(f"/api/employees/{self_id}/contracts")
        assert res.status_code == 200, res.text
        rows = res.json()
        assert len(rows) == 1
        assert rows[0]["salary_at_contract"] == 35000.0

    def test_admin_lists_any_contracts_salary_present(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _, other_emp = self._seed(s)
            _create_user(s, username="adm_c", role="admin", permissions=-1)
            s.commit()
            other_id = other_emp.id

        _login(client, "adm_c")
        res = client.get(f"/api/employees/{other_id}/contracts")
        assert res.status_code == 200, res.text
        rows = res.json()
        assert len(rows) == 1
        assert rows[0]["salary_at_contract"] == 42000.0


# ─────────────────────────────────────────────────────────────────────────
# F-031：GET /reports/finance-summary/detail|export
# ─────────────────────────────────────────────────────────────────────────


def _seed_salary_record(s):
    emp = _create_employee(s, "F031", "薪資員工")
    rec = SalaryRecord(
        employee_id=emp.id,
        salary_year=2026,
        salary_month=4,
        gross_salary=42000,
        labor_insurance_employer=2500,
        health_insurance_employer=1400,
        pension_employer=1800,
        net_salary=37000,
        total_deduction=5000,
        is_finalized=False,
    )
    s.add(rec)
    s.flush()
    return emp


class TestF031_FinanceSummary:
    """F-031：跨員工薪資明細只給 admin/hr；其餘 REPORTS 持有者欄位遮罩。"""

    def test_supervisor_detail_salary_section_masked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _seed_salary_record(s)
            _create_user(
                s,
                username="sv1",
                role="supervisor",
                permissions=int(Permission.REPORTS),
            )
            s.commit()

        _login(client, "sv1")
        res = client.get("/api/reports/finance-summary/detail?year=2026&month=4")
        assert res.status_code == 200, res.text
        data = res.json()
        assert "salary" in data
        assert len(data["salary"]) >= 1
        for row in data["salary"]:
            for k in ("gross_salary", "net_salary", "employer_benefit", "real_cost"):
                assert row.get(k) is None, f"{k} 應被遮罩，實際 {row.get(k)}"

    def test_supervisor_export_sheet5_masked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            emp = _seed_salary_record(s)
            _create_user(
                s,
                username="sv2",
                role="supervisor",
                permissions=int(Permission.REPORTS),
            )
            s.commit()
            emp_name = emp.name

        _login(client, "sv2")
        res = client.get("/api/reports/finance-summary/export?year=2026&month=4")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        assert "薪資明細" in wb.sheetnames
        ws = wb["薪資明細"]
        # row 2 是第一筆資料（row 1 為 header）
        assert ws.cell(row=2, column=1).value == emp_name
        # 應發/實發/雇主保費/真實支出皆遮罩為 "—"
        for col in (2, 3, 4, 5):
            assert (
                ws.cell(row=2, column=col).value == "—"
            ), f"col {col} 應被遮罩，實際 {ws.cell(row=2, column=col).value}"

    def test_custom_role_with_reports_and_salary_read_still_masked(
        self, field_leak_client
    ):
        """關鍵 adversarial：role=custom（非 admin/hr）即使 REPORTS+SALARY_READ 也應遮罩。"""
        client, sf = field_leak_client
        with sf() as s:
            _seed_salary_record(s)
            _create_user(
                s,
                username="custom_lite",
                role="accountant_lite",
                permissions=int(Permission.REPORTS) | int(Permission.SALARY_READ),
            )
            s.commit()

        _login(client, "custom_lite")
        res = client.get("/api/reports/finance-summary/detail?year=2026&month=4")
        assert res.status_code == 200, res.text
        data = res.json()
        for row in data["salary"]:
            for k in ("gross_salary", "net_salary", "employer_benefit", "real_cost"):
                assert row.get(k) is None, f"自訂角色仍應被遮罩，{k}={row.get(k)}"

    def test_hr_detail_unmasked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _seed_salary_record(s)
            _create_user(
                s,
                username="hr_fd",
                role="hr",
                permissions=int(Permission.REPORTS) | int(Permission.SALARY_READ),
            )
            s.commit()

        _login(client, "hr_fd")
        res = client.get("/api/reports/finance-summary/detail?year=2026&month=4")
        assert res.status_code == 200, res.text
        data = res.json()
        assert any(r.get("gross_salary") == 42000 for r in data["salary"])
        assert any(r.get("net_salary") == 37000 for r in data["salary"])

    def test_admin_detail_unmasked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _seed_salary_record(s)
            _create_user(s, username="adm_fd", role="admin", permissions=-1)
            s.commit()

        _login(client, "adm_fd")
        res = client.get("/api/reports/finance-summary/detail?year=2026&month=4")
        assert res.status_code == 200, res.text
        data = res.json()
        assert any(r.get("gross_salary") == 42000 for r in data["salary"])


# ─────────────────────────────────────────────────────────────────────────
# F-036：GET /exports/overtimes
# ─────────────────────────────────────────────────────────────────────────


def _seed_overtime(s):
    emp = _create_employee(s, "OT_emp", "加班員")
    ot = OvertimeRecord(
        employee_id=emp.id,
        overtime_date=date(2026, 4, 5),
        overtime_type="weekday",
        start_time=datetime(2026, 4, 5, 18, 0),
        end_time=datetime(2026, 4, 5, 20, 0),
        hours=2.0,
        overtime_pay=536,
        reason="加班測試",
        is_approved=True,
    )
    s.add(ot)
    s.flush()
    return emp


class TestF036_OvertimesExport:
    """F-036：overtime_pay 欄位需 admin/hr 才能看到實際金額。"""

    def test_supervisor_overtime_pay_column_masked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _seed_overtime(s)
            _create_user(
                s,
                username="sv_ot",
                role="supervisor",
                permissions=int(Permission.OVERTIME_READ),
            )
            s.commit()

        _login(client, "sv_ot")
        res = client.get("/api/exports/overtimes?year=2026&month=4")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        # row 4 是第一筆資料（row 1 為 title、row 3 為 header）
        # 欄位順序：員工/日期/類型/開始/結束/時數/加班費/原因/狀態
        assert (
            ws.cell(row=4, column=7).value == "—"
        ), f"加班費應遮罩，實際 {ws.cell(row=4, column=7).value}"

    def test_custom_role_with_overtime_read_and_salary_read_still_masked(
        self, field_leak_client
    ):
        """關鍵 adversarial：custom role（非 admin/hr）即使持 OVERTIME_READ+SALARY_READ
        仍應遮罩 overtime_pay。"""
        client, sf = field_leak_client
        with sf() as s:
            _seed_overtime(s)
            _create_user(
                s,
                username="custom_ot",
                role="hr_lite_assist",
                permissions=int(Permission.OVERTIME_READ) | int(Permission.SALARY_READ),
            )
            s.commit()

        _login(client, "custom_ot")
        res = client.get("/api/exports/overtimes?year=2026&month=4")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        assert ws.cell(row=4, column=7).value == "—"

    def test_hr_overtime_pay_unmasked(self, field_leak_client):
        client, sf = field_leak_client
        with sf() as s:
            _seed_overtime(s)
            _create_user(
                s,
                username="hr_ot",
                role="hr",
                permissions=int(Permission.OVERTIME_READ),
            )
            s.commit()

        _login(client, "hr_ot")
        res = client.get("/api/exports/overtimes?year=2026&month=4")
        assert res.status_code == 200, res.text
        wb = load_workbook(BytesIO(res.content))
        ws = wb.active
        # hr 應看到實際金額 536
        assert (
            ws.cell(row=4, column=7).value == 536
        ), f"hr 應看到加班費，實際 {ws.cell(row=4, column=7).value}"
