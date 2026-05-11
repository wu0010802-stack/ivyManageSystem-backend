"""appraisal reports 匯出測試。依賴 conftest fixtures（T17）。"""

import pytest
from io import BytesIO


@pytest.mark.skipif(True, reason="conftest fixtures 將在 T17 加入")
class TestAppraisalReports:

    def test_cycle_report_json(self, client, admin_headers, finalized_cycle):
        resp = client.get(
            f"/api/appraisal/cycles/{finalized_cycle.id}/report",
            headers=admin_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "cycle" in data
        assert "summaries" in data

    def test_cycle_report_xlsx(self, client, admin_headers, finalized_cycle):
        from openpyxl import load_workbook

        resp = client.get(
            f"/api/appraisal/cycles/{finalized_cycle.id}/report.xlsx",
            headers=admin_headers,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
        wb = load_workbook(BytesIO(resp.content))
        assert "考核總表" in wb.sheetnames

    def test_penalty_log_xlsx(self, client, admin_headers, finalized_cycle):
        from openpyxl import load_workbook

        resp = client.get(
            f"/api/appraisal/cycles/{finalized_cycle.id}/penalty_log.xlsx",
            headers=admin_headers,
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        assert "懲處日誌" in wb.sheetnames

    def test_participant_sheet_xlsx(self, client, admin_headers, finalized_summary):
        from openpyxl import load_workbook

        pid = finalized_summary.participant_id
        resp = client.get(
            f"/api/appraisal/participants/{pid}/sheet.xlsx",
            headers=admin_headers,
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        assert "個人考核表" in wb.sheetnames
