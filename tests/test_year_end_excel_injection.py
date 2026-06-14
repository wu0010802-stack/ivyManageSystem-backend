"""年終 Excel 匯出須過 SafeWorksheet 防公式注入（C11，bug hunt money-auth 2026-06-14）。

export_year_end_transfer_xlsx / export_year_end_summary_xlsx 原直接用 wb.active 寫入，
員工姓名/戶名若含 =cmd|'/C calc'!A0 會在財會端 Excel 開啟時被當公式執行。
修法：改用 SafeWorksheet(wb.active)，與 transfer_roster / gov_reports 等既有匯出對齊。
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from services.year_end.excel_io import (
    SummaryExportRow,
    TransferRow,
    export_year_end_summary_xlsx,
    export_year_end_transfer_xlsx,
)

EVIL = "=cmd|'/C calc'!A0"


def _all_values(payload: bytes) -> list:
    ws = load_workbook(BytesIO(payload)).active
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


def test_transfer_xlsx_sanitizes_formula_injection():
    payload = export_year_end_transfer_xlsx(
        rows=[
            TransferRow(bank_account="0123456789", name=EVIL, amount=Decimal("1000"))
        ],
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "戶名公式注入未被清理（原樣寫入 → 財會端可被執行）"
    assert ("'" + EVIL) in vals, "戶名應被 sanitize 為 ' 前綴純字串"


def test_summary_xlsx_sanitizes_formula_injection():
    payload = export_year_end_summary_xlsx(
        rows=[
            SummaryExportRow(
                name=EVIL,
                year_end_amount=Decimal("1000"),
                bonus_by_type={},
                total=Decimal("1000"),
            )
        ],
        academic_year=114,
    )
    vals = _all_values(payload)
    assert EVIL not in vals, "姓名公式注入未被清理（原樣寫入）"
    assert ("'" + EVIL) in vals, "姓名應被 sanitize 為 ' 前綴純字串"
