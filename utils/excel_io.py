"""Excel 匯入骨架 — 統一錯誤格式 {row, col, value, error_code, message}。

Why centralize: leaves/overtimes/shifts 等各自寫 openpyxl/pandas 解析，欄位驗證與
錯誤回報格式不一致。本骨架提供 schema 宣告（pydantic v2）+ 統一 ImportResult。

設計重點：
- `ExcelImportSchema` 子類別宣告 columns，支援 Chinese alias（`Field(alias="員工編號")`）。
- 內部使用 `model_validate`（非 `**kwargs`）以解析 alias。
- header 欄位驗證 = schema fields 的 alias（若有）或 field name。

採用端：
- api/leaves.py:import_leaves（2026-05-13）
- api/overtimes.py:import_overtimes（2026-05-19）
- api/punch_corrections.py 目前無 Excel 匯入 endpoint，無需接入。
"""

from dataclasses import dataclass, field
from typing import IO, Any

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError


class ExcelImportSchema(BaseModel):
    """匯入列 schema 基底。

    子類別範例：
        class LeaveImportRow(ExcelImportSchema):
            employee_name: str = Field(alias="員工姓名")
            start_date: str = Field(alias="開始日期")

    說明：
    - `extra="forbid"`：header 含未宣告欄位時，該欄位忽略（parse_excel 只填 expected_cols）。
    - `populate_by_name=True`：同時允許 field name 與 alias，方便英文 schema 直接套。
    """

    model_config = {"extra": "forbid", "populate_by_name": True}


@dataclass
class ImportResult:
    """統一匯入結果。

    rows: 成功 validate 的 schema instances（順序同 Excel 列順序）。
    errors: 結構化錯誤列表，每筆 {row, col, value, error_code, message}。
        - row: Excel 列號（header=1，第一筆資料=2）。
        - col: 出錯的欄位名（field name 或 alias），整體錯誤（如 MISSING_COLUMN）為 None。
        - value: 該 cell 的原始值（若有）。
        - error_code: 大寫常數（INVALID_FILE / EMPTY_FILE / MISSING_COLUMN / pydantic 錯誤類型大寫）。
        - message: 人類可讀訊息（含 pydantic 原文）。
    """

    rows: list[Any] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)


def _expected_header_names(schema: type[ExcelImportSchema]) -> set[str]:
    """回傳 schema 期望的 header 集合：優先用 alias，沒有的退回 field name。"""
    names: set[str] = set()
    for name, info in schema.model_fields.items():
        alias = getattr(info, "alias", None)
        names.add(alias if alias else name)
    return names


def _required_header_names(schema: type[ExcelImportSchema]) -> set[str]:
    """回傳必填的 header 集合（is_required=True 的欄位）。"""
    names: set[str] = set()
    for name, info in schema.model_fields.items():
        if not info.is_required():
            continue
        alias = getattr(info, "alias", None)
        names.add(alias if alias else name)
    return names


def parse_excel(
    file_or_buffer: IO[bytes],
    *,
    schema: type[ExcelImportSchema],
) -> ImportResult:
    """解析 Excel 第一個 sheet，每列轉為 schema instance。

    錯誤統一收進 result.errors，不 raise（除了無法讀檔本身）。
    呼叫端負責檔案大小檢查 / 簽章驗證（utils/file_upload 那層）。
    """
    result = ImportResult()
    try:
        wb = load_workbook(file_or_buffer, read_only=True, data_only=True)
    except Exception as exc:
        result.errors.append(
            {
                "row": 0,
                "col": None,
                "value": None,
                "error_code": "INVALID_FILE",
                "message": f"無法讀取 Excel：{exc}",
            }
        )
        return result

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        result.errors.append(
            {
                "row": 0,
                "col": None,
                "value": None,
                "error_code": "EMPTY_FILE",
                "message": "Excel 為空",
            }
        )
        return result

    header = [str(c) if c is not None else "" for c in header_row]
    if not any(h.strip() for h in header):
        result.errors.append(
            {
                "row": 0,
                "col": None,
                "value": None,
                "error_code": "EMPTY_FILE",
                "message": "Excel 為空（header 列無內容）",
            }
        )
        return result

    expected_cols = _expected_header_names(schema)
    required_cols = _required_header_names(schema)
    actual_cols = set(header)
    missing = required_cols - actual_cols
    if missing:
        result.errors.append(
            {
                "row": 1,
                "col": None,
                "value": None,
                "error_code": "MISSING_COLUMN",
                "message": f"缺欄位：{', '.join(sorted(missing))}",
            }
        )
        return result

    for row_idx, raw in enumerate(rows_iter, start=2):
        # 整列空白跳過（避免 trailing empty rows 觸發 validation 噪音）
        if raw is None or all(v is None for v in raw):
            continue
        record: dict[str, Any] = {}
        for i, col_name in enumerate(header):
            if col_name not in expected_cols:
                continue
            if i < len(raw):
                record[col_name] = raw[i]
        try:
            # 使用 model_validate（非 schema(**record)）以正確解析 alias。
            result.rows.append(schema.model_validate(record))
        except ValidationError as exc:
            for err in exc.errors():
                # loc 對應 schema 內部 field name；對外回報用 alias 較直觀。
                field_name = err["loc"][0] if err["loc"] else None
                info = schema.model_fields.get(field_name) if field_name else None
                col_label = (
                    getattr(info, "alias", None) or field_name if info else field_name
                )
                value = record.get(col_label) if col_label else None
                result.errors.append(
                    {
                        "row": row_idx,
                        "col": col_label,
                        "value": value,
                        "error_code": str(err.get("type", "VALIDATION_ERROR")).upper(),
                        "message": err.get("msg", "驗證錯誤"),
                    }
                )
    return result
