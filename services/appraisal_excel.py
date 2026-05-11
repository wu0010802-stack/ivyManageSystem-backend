"""考核報表 Excel 產出（對齊文件附表八 + 月考核表結構）。"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from models.appraisal import (
    AppraisalCycle,
    AppraisalEvent,
    AppraisalParticipant,
    AppraisalSummary,
    Grade,
    RoleGroup,
)
from models.employee import Employee


def _role_label(rg: RoleGroup) -> str:
    return {
        RoleGroup.SUPERVISOR: "主管",
        RoleGroup.HEAD_TEACHER: "班導/會計",
        RoleGroup.ASSISTANT: "副班導/廚/司機/儲備",
    }.get(rg, rg.value)


def _grade_label(g: Grade) -> str:
    return {
        Grade.OUTSTANDING: "優",
        Grade.GOOD: "甲",
        Grade.PASS: "乙",
        Grade.WARN: "丙",
        Grade.FAIL: "丁",
    }[g]


def _emp_name(emp: Employee) -> str:
    return getattr(emp, "name", "") or ""


def _emp_job_title_name(emp: Employee) -> str:
    """取 employee 的職稱名稱；對齊 Employee.job_title_rel 關係。"""
    rel = getattr(emp, "job_title_rel", None)
    if rel is not None and hasattr(rel, "name"):
        return rel.name or ""
    return ""


def build_cycle_report(db: Session, cycle: AppraisalCycle) -> bytes:
    """全校學期總表 Excel：每位 participant 一列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考核總表"

    title_text = (
        f"考核總表 {cycle.academic_year} 學年度"
        f" {'第一' if cycle.semester.value == 'FIRST' else '第二'}學期"
    )
    ws.append([title_text])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    headers = [
        "員工",
        "職務",
        "Role Group",
        "基本分",
        "事件加減分",
        "總分",
        "等級",
        "獎金",
        "簽核狀態",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    rows = db.execute(
        select(AppraisalSummary, AppraisalParticipant, Employee)
        .join(
            AppraisalParticipant,
            AppraisalParticipant.id == AppraisalSummary.participant_id,
        )
        .join(Employee, Employee.id == AppraisalParticipant.employee_id)
        .where(AppraisalSummary.cycle_id == cycle.id)
        .order_by(AppraisalParticipant.role_group, Employee.id)
    ).all()

    for s, p, emp in rows:
        ws.append(
            [
                _emp_name(emp),
                _emp_job_title_name(emp),
                _role_label(p.role_group),
                float(s.base_score),
                float(s.event_score_sum),
                float(s.total_score),
                _grade_label(s.grade),
                float(s.bonus_amount),
                s.status.value,
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_penalty_log(db: Session, cycle: AppraisalCycle) -> bytes:
    """懲處日誌：列出該 cycle 全部 penalty 類事件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "懲處日誌"
    ws.append(
        [
            "事件日期",
            "員工",
            "事由代碼",
            "類別",
            "事件類型",
            "扣加分",
            "階梯",
            "家長反應",
            "標題",
            "詳述",
            "登錄者",
            "已撤銷",
            "附件數",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = db.execute(
        select(AppraisalEvent, AppraisalParticipant, Employee)
        .join(
            AppraisalParticipant,
            AppraisalParticipant.id == AppraisalEvent.participant_id,
        )
        .join(Employee, Employee.id == AppraisalParticipant.employee_id)
        .where(AppraisalEvent.cycle_id == cycle.id)
        .order_by(AppraisalEvent.event_date)
    ).all()

    for ev, p, emp in rows:
        cat_code = ev.catalog_item.code if ev.catalog_item else ""
        cat_category = ev.catalog_item.category.value if ev.catalog_item else ""
        ws.append(
            [
                ev.event_date.isoformat(),
                _emp_name(emp),
                cat_code,
                cat_category,
                ev.event_type.value,
                float(ev.score_delta),
                ev.severity_level,
                ev.parent_reaction.value if ev.parent_reaction else "",
                ev.title,
                ev.detail,
                ev.created_by,
                "是" if ev.reverted_at else "",
                len(ev.attachments or []),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_participant_sheet(db: Session, participant: AppraisalParticipant) -> bytes:
    """個人考核表（對齊月考核表紙本，但 v1 簡化版）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "個人考核表"

    emp = db.get(Employee, participant.employee_id)
    cycle = db.get(AppraisalCycle, participant.cycle_id)
    summary = db.execute(
        select(AppraisalSummary).where(
            AppraisalSummary.participant_id == participant.id
        )
    ).scalar_one_or_none()

    ws["A1"] = "教職員考核表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"姓名：{_emp_name(emp)}"
    ws["B2"] = f"職務：{_emp_job_title_name(emp)}"
    ws["A3"] = (
        f"學年期：{cycle.academic_year} 學年度"
        f" {'上' if cycle.semester.value == 'FIRST' else '下'}學期"
    )
    ws["A4"] = f"考核期間：{cycle.start_date} ~ {cycle.end_date}"

    ws["A6"] = "基本分數"
    ws["B6"] = (
        float(participant.base_score) if summary is None else float(summary.base_score)
    )
    ws["A7"] = "事件加減分合計"
    ws["B7"] = 0 if summary is None else float(summary.event_score_sum)
    ws["A8"] = "總分"
    ws["B8"] = 0 if summary is None else float(summary.total_score)
    ws["A9"] = "等級"
    ws["B9"] = "未結算" if summary is None else _grade_label(summary.grade)
    ws["A10"] = "考核獎金（試算）"
    ws["B10"] = 0 if summary is None else float(summary.bonus_amount)

    ws["A12"] = "事件紀錄"
    ws["A12"].font = Font(bold=True)
    ws.append([])
    ws.append(["日期", "類型", "扣加分", "標題", "詳述", "已撤銷"])
    for ev in (
        db.execute(
            select(AppraisalEvent)
            .where(AppraisalEvent.participant_id == participant.id)
            .order_by(AppraisalEvent.event_date)
        )
        .scalars()
        .all()
    ):
        ws.append(
            [
                ev.event_date.isoformat(),
                ev.event_type.value,
                float(ev.score_delta),
                ev.title,
                ev.detail,
                "是" if ev.reverted_at else "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
