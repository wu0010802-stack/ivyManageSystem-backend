"""
Overtime management router
"""

import asyncio
import logging
import calendar as cal_module
from datetime import date, datetime, time as dt_time, timedelta
from io import BytesIO
from typing import Any, Optional, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from fastapi import APIRouter, Body, Depends, HTTPException, Request, UploadFile, File
from utils.errors import raise_safe_500, safe_batch_reason
from utils.excel_utils import SafeWorksheet
from utils.rate_limit import create_limiter

# 批次核准為重 DB 操作（多筆 FOR UPDATE + 狀態變更 + LINE 推播），防內部濫用或帳號被盜後的放大
_batch_approve_limiter = create_limiter(
    max_calls=10,
    window_seconds=60,
    name="overtime_batch_approve",
    error_detail="批次審核操作過於頻繁，請稍後再試",
).as_dependency()
from utils.file_upload import read_upload_with_size_check, validate_file_signature
from utils.excel_io import ExcelImportSchema, parse_excel
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy import func, or_, and_

from models.database import (
    get_session,
    Employee,
    OvertimeRecord,
    LeaveQuota,
    User,
    LeaveRecord,
    SalaryRecord,
)
from models.approval import ApprovalStatus
from models.event import Holiday
from models.overtime_comp_leave_grant import OvertimeCompLeaveGrant
from schemas._common import BatchApproveResultOut
from schemas.overtimes import (
    BatchOvertimeCreateResultOut,
    OvertimeApproveResultOut,
    OvertimeCreateResultOut,
    OvertimeDeleteResultOut,
    OvertimeImportResultOut,
    OvertimeUpdateResultOut,
)
from utils.auth import require_staff_permission
from utils.constants import (
    OVERTIME_TYPE_LABELS,
    MAX_OVERTIME_HOURS,
    MAX_MONTHLY_OVERTIME_HOURS,
    DAILY_WORK_HOURS,
    WEEKDAY_FIRST_2H_RATE,
    WEEKDAY_AFTER_2H_RATE,
    WEEKDAY_THRESHOLD_HOURS,
    HOLIDAY_RATE,
    RESTDAY_FIRST_2H_RATE,
    RESTDAY_MID_RATE,
    RESTDAY_AFTER_8H_RATE,
    RESTDAY_FIRST_SEGMENT,
    RESTDAY_SECOND_SEGMENT,
    RESTDAY_MIN_HOURS,
)
from utils.validators import validate_hhmm_format
from utils.error_messages import EMPLOYEE_DOES_NOT_EXIST, OVERTIME_RECORD_NOT_FOUND
from utils.permissions import Permission
from services.salary.utils import (
    lock_and_premark_stale,
    mark_salary_stale as _mark_salary_stale,
)
from services.salary.finalize_guard import (
    collect_months_from_dates,
    assert_months_not_finalized,
)
from utils.approval_helpers import (
    _get_submitter_role,
    _check_approval_eligibility,
    _write_approval_log,
    assert_approver_eligible,
    is_self_approval,
)
from utils.excel_utils import xlsx_streaming_response
from utils.import_utils import build_employee_lookup, resolve_employee_from_row

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["overtimes"])

# ============ Service Injection ============

_salary_engine = None


def init_overtimes_services(salary_engine_instance):
    global _salary_engine
    _salary_engine = salary_engine_instance


MONTHLY_BASE_DAYS = 30  # 勞基法時薪計算基準日數（月薪 ÷ 30 ÷ 8）


# ============ Helper Functions ============


def _revoke_comp_leave_grant(
    session,
    ot: OvertimeRecord,
    current_user: Optional[dict] = None,
) -> None:
    """撤銷已發放的補休配額。

    1. 若有與此加班記錄明確關聯（source_overtime_id）的補休假單：
       - 已核准（已使用）→ 拋出 409，必須人工撤銷
       - 待審核 → 自動駁回，釋放配額占用
    2. 全域檢查：確認撤銷後配額不低於仍在使用/申請中的補休總量（含無關聯舊資料）
    """
    if not ot.use_comp_leave or not ot.comp_leave_granted:
        return

    year = ot.overtime_date.year
    quota = (
        session.query(LeaveQuota)
        .filter(
            LeaveQuota.employee_id == ot.employee_id,
            LeaveQuota.year == year,
            LeaveQuota.leave_type == "compensatory",
        )
        .first()
    )

    if not quota:
        ot.comp_leave_granted = False
        logger.warning(
            "補休回滾時找不到配額紀錄：員工 ID=%d, %d 年（加班 #%d）",
            ot.employee_id,
            year,
            ot.id,
        )
        return

    # ── 步驟 1：處理有明確關聯的補休假單 ──────────────────────────────────
    linked_leaves = (
        session.query(LeaveRecord)
        .filter(
            LeaveRecord.source_overtime_id == ot.id,
        )
        .all()
    )

    linked_approved = [
        lv for lv in linked_leaves if lv.status == ApprovalStatus.APPROVED.value
    ]
    linked_pending = [
        lv for lv in linked_leaves if lv.status == ApprovalStatus.PENDING.value
    ]

    if linked_approved:
        approved_h = sum(lv.leave_hours for lv in linked_approved)
        ids = ", ".join(str(lv.id) for lv in linked_approved)
        raise HTTPException(
            status_code=409,
            detail=(
                f"此筆加班已發放的補休有 {approved_h:.1f} 小時已核准使用"
                f"（假單 ID：{ids}），請先撤銷相關補休假單後再操作"
            ),
        )

    # 自動駁回待審核的關聯補休假單
    # 修補 2026-05-11 P1-8：補寫 ApprovalLog 確保稽核軌跡完整。current_user 由
    # caller 傳入；無 caller 傳入時 fallback "system_auto" 確保函式仍可用。
    _audit_actor = current_user or {
        "username": "system_auto",
        "role": "system",
    }
    for lv in linked_pending:
        lv.status = ApprovalStatus.REJECTED.value
        lv.rejection_reason = (
            f"來源加班申請（#{ot.id}，{ot.overtime_date}）已被撤銷，補休資格取消"
        )
        _write_approval_log(
            session=session,
            doc_type="leave",
            doc_id=lv.id,
            action="rejected",
            approver=_audit_actor,
            comment=f"auto_revoked_by_overtime_rollback (#{ot.id})",
        )
        logger.info(
            "補休假單 #%d 因來源加班 #%d 被撤銷而自動駁回（員工 ID=%d）",
            lv.id,
            ot.id,
            ot.employee_id,
        )

    # ── 步驟 2：全域 committed 檢查（含舊資料或其他來源的補休） ─────────────
    # autoflush 確保上方的狀態變更在下方查詢前已寫入 session，
    # 避免自動駁回的假單仍被計入 pending。
    new_total = float(quota.total_hours or 0) - float(ot.hours or 0)

    approved_committed = float(
        session.query(func.coalesce(func.sum(LeaveRecord.leave_hours), 0))
        .filter(
            LeaveRecord.employee_id == ot.employee_id,
            LeaveRecord.leave_type == "compensatory",
            LeaveRecord.status == ApprovalStatus.APPROVED.value,
            LeaveRecord.start_date >= date(year, 1, 1),
            LeaveRecord.start_date < date(year + 1, 1, 1),
        )
        .scalar()
    )
    pending_committed = float(
        session.query(func.coalesce(func.sum(LeaveRecord.leave_hours), 0))
        .filter(
            LeaveRecord.employee_id == ot.employee_id,
            LeaveRecord.leave_type == "compensatory",
            LeaveRecord.status == ApprovalStatus.PENDING.value,
            LeaveRecord.start_date >= date(year, 1, 1),
            LeaveRecord.start_date < date(year + 1, 1, 1),
        )
        .scalar()
    )
    committed = approved_committed + pending_committed

    if new_total + 1e-9 < committed:
        raise HTTPException(
            status_code=409,
            detail=(
                f"此筆加班已發放的補休 {ot.hours:.1f} 小時已有 {committed:.1f} 小時被使用或申請，"
                "請先撤銷相關補休假單後再修改、刪除或駁回此加班申請"
            ),
        )

    # ── 新增：mark grant ledger row 為 revoked（不刪除留 audit） ──
    grant = (
        session.query(OvertimeCompLeaveGrant)
        .filter(OvertimeCompLeaveGrant.overtime_record_id == ot.id)
        .first()
    )
    if grant is not None:
        grant.status = "revoked"

    quota.total_hours = max(0.0, new_total)
    ot.comp_leave_granted = False


def _recalculate_salary_for_overtime_months(
    employee_id: int, months: set[tuple[int, int]]
) -> None:
    """重算受影響月份的薪資。"""
    if _salary_engine is None:
        return
    for year, month in sorted(months):
        _salary_engine.process_salary_calculation(employee_id, year, month)


def _grant_comp_leave_quota(session, ot: OvertimeRecord, result: dict) -> None:
    """核准補休模式加班時，upsert 補休配額並標記 comp_leave_granted。

    防止重複發放：只在 ot.comp_leave_granted 為 False 時才執行。
    發放後將小時數寫入 result['comp_leave_hours_granted'] 供 API 回傳。
    """
    if not (ot.use_comp_leave and not ot.comp_leave_granted):
        return
    year = ot.overtime_date.year
    quota = (
        session.query(LeaveQuota)
        .filter(
            LeaveQuota.employee_id == ot.employee_id,
            LeaveQuota.year == year,
            LeaveQuota.leave_type == "compensatory",
        )
        .with_for_update()
        .first()
    )
    if quota:
        quota.total_hours += ot.hours
    else:
        quota = LeaveQuota(
            employee_id=ot.employee_id,
            year=year,
            leave_type="compensatory",
            total_hours=ot.hours,
            note="由加班補休累積",
        )
        session.add(quota)
    # ── 新增 grant ledger row（per-OT 帳本，T9 scheduler 將從此撈到期 grant 結算）──
    grant = OvertimeCompLeaveGrant(
        overtime_record_id=ot.id,
        employee_id=ot.employee_id,
        granted_hours=ot.hours,
        granted_at=ot.overtime_date,
        expires_at=ot.overtime_date + timedelta(days=365),
        status="active",
    )
    session.add(grant)
    ot.comp_leave_granted = True
    result["comp_leave_hours_granted"] = ot.hours
    logger.info(
        "補休配額已發放：員工 ID=%d, %d 年度 +%.1f 小時（加班記錄 #%d）",
        ot.employee_id,
        year,
        ot.hours,
        ot.id,
    )


def _notify_and_recalc_overtime(
    session, ot: OvertimeRecord, approved: bool, was_approved: bool, result: dict
) -> None:
    """核准/駁回後的 LINE 推播與薪資重算。

    LINE 通知失敗不影響主流程（僅記錄 warning）。
    薪資重算失敗則在 result 中寫入 warning 訊息，由 API 回傳給前端。
    """
    # 核准或撤銷已核准狀態後都需要重算薪資
    if (approved or was_approved) and _salary_engine is not None:
        try:
            _salary_engine.process_salary_calculation(
                ot.employee_id, ot.overtime_date.year, ot.overtime_date.month
            )
            result["salary_recalculated"] = True
            result["message"] = (
                "已核准，薪資已自動重算" if approved else "已駁回，薪資已自動重算"
            )
            logger.info(
                "加班審核後自動重算薪資：emp_id=%d, %d/%d",
                ot.employee_id,
                ot.overtime_date.year,
                ot.overtime_date.month,
            )
        except Exception as e:
            result["salary_recalculated"] = False
            result["warning"] = (
                "已核准，但薪資重算失敗，請手動前往薪資頁面重新計算"
                if approved
                else "已駁回，但薪資重算失敗，請手動前往薪資頁面重新計算"
            )
            logger.error("加班審核後薪資重算失敗：%s", e)
            try:
                _mark_salary_stale(
                    session,
                    ot.employee_id,
                    ot.overtime_date.year,
                    ot.overtime_date.month,
                )
                session.commit()
            except Exception:
                logger.warning(
                    "加班審核降級時標記 SalaryRecord stale 失敗", exc_info=True
                )
                session.rollback()


# F5: 抽到 utils/leave_overtime_conflict.py 共用（leaves.py 與 overtimes.py
# 原本各自有一份）。保留本檔 alias 維持既有 import surface。
from utils.leave_overtime_conflict import (
    to_time as _to_time,
    times_overlap as _times_overlap,
)

# F1 第二/三波：calculate_overtime_pay + 4 衝突檢查 helper 全部抽到 services。
# 本檔保留 alias 維持既有 import surface 不變（api/portal/overtimes.py 與
# 其他模組已透過 `from api.overtimes import _check_*` 取得；改用 alias 不需動）。
from services.overtime_pay_calculator import calculate_overtime_pay  # noqa: E402,F401
from services.overtime_conflict_service import (  # noqa: E402,F401
    _assert_within_monthly_cap,
    _validate_overtime_type_matches_calendar,
    check_employee_has_conflicting_leave as _check_employee_has_conflicting_leave,
    check_monthly_overtime_cap as _check_monthly_overtime_cap,
    check_overtime_overlap as _check_overtime_overlap,
    check_overtime_type_calendar as _check_overtime_type_calendar,
    check_quarterly_overtime_cap as _check_quarterly_overtime_cap,
)

# ============ Pydantic Models ============


def _assert_hours_within_span(
    hours: float, start_time: "Optional[str]", end_time: "Optional[str]"
) -> None:
    """P1-2：加班 hours 不可超過 start~end 時段差（防超報溢付）。

    允許 hours <= 時段差（容許中間休息不計薪，業主 2026-06-05 定案 ≤ 而非嚴格相等）。
    start_time/end_time 為已過 validate_hhmm_format 的 HH:MM 字串；任一為 None 不檢查。
    """
    if not start_time or not end_time:
        return
    start_min = int(start_time[:2]) * 60 + int(start_time[3:5])
    end_min = int(end_time[:2]) * 60 + int(end_time[3:5])
    span_hours = (end_min - start_min) / 60
    if hours > span_hours + 1e-6:
        raise ValueError(f"加班時數（{hours}）不可超過起迄時段差（{span_hours} 小時）")


class OvertimeCreate(BaseModel):
    employee_id: int
    overtime_date: date
    overtime_type: str  # weekday / weekend / holiday
    start_time: Optional[str] = None  # HH:MM
    end_time: Optional[str] = None  # HH:MM
    hours: float
    reason: Optional[str] = None
    use_comp_leave: bool = False  # 以補休代替加班費

    @field_validator("overtime_type")
    @classmethod
    def validate_overtime_type(cls, v):
        if v not in OVERTIME_TYPE_LABELS:
            allowed = ", ".join(OVERTIME_TYPE_LABELS.keys())
            raise ValueError(f"無效的加班類型，允許值：{allowed}")
        return v

    @field_validator("hours")
    @classmethod
    def validate_hours(cls, v):
        if v <= 0:
            raise ValueError("加班時數必須大於 0")
        if v > MAX_OVERTIME_HOURS:
            raise ValueError(f"單筆加班時數不得超過 {MAX_OVERTIME_HOURS} 小時")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_time_format(cls, v):
        return validate_hhmm_format(v)

    @model_validator(mode="after")
    def validate_time_order(self):
        if self.start_time and self.end_time:
            if self.start_time >= self.end_time:
                raise ValueError("start_time 必須早於 end_time（不支援跨日加班）")
            _assert_hours_within_span(self.hours, self.start_time, self.end_time)
        return self


class BatchOvertimeEmployeeItem(BaseModel):
    employee_id: int
    hours: float

    @field_validator("hours")
    @classmethod
    def validate_hours(cls, v):
        if v <= 0:
            raise ValueError("加班時數必須大於 0")
        if v > MAX_OVERTIME_HOURS:
            raise ValueError(f"單筆加班時數不得超過 {MAX_OVERTIME_HOURS} 小時")
        return v


class BatchOvertimeCreate(BaseModel):
    overtime_date: date
    overtime_type: str  # weekday / weekend / holiday
    start_time: Optional[str] = None  # HH:MM，共用，選填
    end_time: Optional[str] = None  # HH:MM，共用，選填
    reason: Optional[str] = None
    use_comp_leave: bool = False
    employees: List[BatchOvertimeEmployeeItem] = Field(
        ..., min_length=1, max_length=500
    )

    @field_validator("overtime_type")
    @classmethod
    def validate_overtime_type(cls, v):
        if v not in OVERTIME_TYPE_LABELS:
            allowed = ", ".join(OVERTIME_TYPE_LABELS.keys())
            raise ValueError(f"無效的加班類型，允許值：{allowed}")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_time_format(cls, v):
        return validate_hhmm_format(v)

    @model_validator(mode="after")
    def validate_time_order(self):
        if self.start_time and self.end_time:
            if self.start_time >= self.end_time:
                raise ValueError("start_time 必須早於 end_time（不支援跨日加班）")
            for _emp in self.employees:
                _assert_hours_within_span(_emp.hours, self.start_time, self.end_time)
        return self


class OvertimeUpdate(BaseModel):
    overtime_date: Optional[date] = None
    overtime_type: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    hours: Optional[float] = None
    reason: Optional[str] = None
    # 注意：刻意不列 use_comp_leave；翻轉模式必須走 reject + recreate 流程，
    # 避免在 update 過程中切換補休/加班費而與 _grant_comp_leave_quota /
    # _revoke_comp_leave_grant 的狀態機脫節（修補 2026-05-11 P2-14）。

    @model_validator(mode="before")
    @classmethod
    def _reject_use_comp_leave_flip(cls, values):
        if isinstance(values, dict) and "use_comp_leave" in values:
            raise ValueError(
                "不允許在 update 中翻轉 use_comp_leave；如需切換補休/加班費模式，"
                "請先 reject 該加班，然後以新 use_comp_leave 值重新申請"
            )
        return values

    @field_validator("overtime_type")
    @classmethod
    def validate_overtime_type(cls, v):
        if v is not None and v not in OVERTIME_TYPE_LABELS:
            allowed = ", ".join(OVERTIME_TYPE_LABELS.keys())
            raise ValueError(f"無效的加班類型，允許值：{allowed}")
        return v

    @field_validator("hours")
    @classmethod
    def validate_hours(cls, v):
        if v is None:
            return v
        if v <= 0:
            raise ValueError("加班時數必須大於 0")
        if v > MAX_OVERTIME_HOURS:
            raise ValueError(f"單筆加班時數不得超過 {MAX_OVERTIME_HOURS} 小時")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_time_format(cls, v):
        return validate_hhmm_format(v)

    @model_validator(mode="after")
    def validate_time_order(self):
        if self.start_time and self.end_time:
            if self.start_time >= self.end_time:
                raise ValueError("start_time 必須早於 end_time（不支援跨日加班）")
        return self


# ============ Batch Approve Request Model ============


class OvertimeApproveRequest(BaseModel):
    """單筆加班核准/駁回 body schema（修補 2026-05-11 P1-6）。

    原本 approve_overtime 用 query parameter 接 rejection_reason，會把駁回原因
    寫進 proxy/CDN/access log；approved_by 也可被外部覆寫。新增 body schema 後
    新前端應改用 body；查詢字串保留作向後相容 fallback。
    """

    approved: bool = True
    rejection_reason: Optional[str] = None


class OvertimeBatchApproveRequest(BaseModel):
    ids: List[int]
    approved: bool
    rejection_reason: Optional[str] = None


# ============ Excel Helpers (local) ============


from utils.excel_writer import write_header_row as _ot_write_header  # noqa: E402

# ============ Routes ============


@router.get("/overtimes")
def get_overtimes(
    employee_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    status: Optional[str] = None,  # pending, approved, rejected
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_READ)),
):
    """查詢加班記錄"""
    session = get_session()
    try:
        q = session.query(OvertimeRecord, Employee).join(
            Employee, OvertimeRecord.employee_id == Employee.id
        )
        if employee_id:
            q = q.filter(OvertimeRecord.employee_id == employee_id)
        if year and month:
            _, last_day = cal_module.monthrange(year, month)
            start = date(year, month, 1)
            end = date(year, month, last_day)
            q = q.filter(
                OvertimeRecord.overtime_date >= start,
                OvertimeRecord.overtime_date <= end,
            )
        elif year:
            q = q.filter(
                OvertimeRecord.overtime_date >= date(year, 1, 1),
                OvertimeRecord.overtime_date <= date(year, 12, 31),
            )

        if status == "pending":
            q = q.filter(OvertimeRecord.status == ApprovalStatus.PENDING.value)
        elif status == "approved":
            q = q.filter(OvertimeRecord.status == ApprovalStatus.APPROVED.value)
        elif status == "rejected":
            q = q.filter(OvertimeRecord.status == ApprovalStatus.REJECTED.value)

        records = q.order_by(OvertimeRecord.overtime_date.desc()).limit(5000).all()

        # 預先載入員工角色映射
        employee_ids = list({ot.employee_id for ot, _ in records})
        user_roles = {}
        if employee_ids:
            users = (
                session.query(User)
                .filter(
                    User.employee_id.in_(employee_ids),
                    User.is_active == True,
                )
                .all()
            )
            user_roles = {u.employee_id: u.role for u in users}

        results = []
        for ot, emp in records:
            results.append(
                {
                    "id": ot.id,
                    "employee_id": ot.employee_id,
                    "employee_name": emp.name,
                    "submitter_role": user_roles.get(ot.employee_id, "teacher"),
                    "overtime_date": ot.overtime_date.isoformat(),
                    "overtime_type": ot.overtime_type,
                    "overtime_type_label": OVERTIME_TYPE_LABELS.get(
                        ot.overtime_type, ot.overtime_type
                    ),
                    "start_time": (
                        ot.start_time.strftime("%H:%M") if ot.start_time else None
                    ),
                    "end_time": ot.end_time.strftime("%H:%M") if ot.end_time else None,
                    "hours": ot.hours,
                    "overtime_pay": ot.overtime_pay,
                    "use_comp_leave": ot.use_comp_leave,
                    "comp_leave_granted": ot.comp_leave_granted,
                    "status": ot.status,
                    "approved_by": ot.approved_by,
                    "reason": ot.reason,
                    "created_at": ot.created_at.isoformat() if ot.created_at else None,
                }
            )
        return results
    finally:
        session.close()


def _parse_hhmm_on_date(overtime_date: date, hhmm: Optional[str]) -> Optional[datetime]:
    """將 'HH:MM' 字串組成指定日期的 datetime；None 原樣回傳。"""
    if not hhmm:
        return None
    h, m = map(int, hhmm.split(":"))
    return datetime.combine(
        overtime_date, datetime.min.time().replace(hour=h, minute=m)
    )


def _validate_overtime_for_employee(
    session,
    employee_id: int,
    overtime_date: date,
    overtime_type: str,
    start_dt: Optional[datetime],
    end_dt: Optional[datetime],
    hours: float,
) -> None:
    """單筆建立與批次建立共用的完整驗證鏈（避免驗證漂移）。

    任一不通過即 raise HTTPException（overlap→409，其餘各檢查自行 raise 400/409）。
    刻意不含 assert_months_not_finalized：與單筆建立對齊（封存守衛在 approve 路徑）。
    """
    overlap = _check_overtime_overlap(
        session, employee_id, overtime_date, start_dt, end_dt
    )
    if overlap:
        st = overlap.start_time.strftime("%H:%M") if overlap.start_time else "未指定"
        et = overlap.end_time.strftime("%H:%M") if overlap.end_time else "未指定"
        raise HTTPException(
            status_code=409,
            detail=(
                f"該員工在 {overlap.overtime_date} 已有時間重疊的加班申請"
                f"（ID: {overlap.id}，{st}～{et}），請勿重複申請"
            ),
        )
    _check_employee_has_conflicting_leave(
        session, employee_id, overtime_date, start_dt, end_dt
    )
    _check_monthly_overtime_cap(session, employee_id, overtime_date, hours)
    _check_quarterly_overtime_cap(session, employee_id, overtime_date, hours)
    _check_overtime_type_calendar(session, overtime_date, overtime_type)


@router.post("/overtimes", status_code=201, response_model=OvertimeCreateResultOut)
def create_overtime(
    data: OvertimeCreate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """新增加班記錄（自動計算加班費）"""
    session = get_session()
    try:
        emp = session.query(Employee).filter(Employee.id == data.employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail=EMPLOYEE_DOES_NOT_EXIST)

        pay = (
            0.0
            if data.use_comp_leave
            else calculate_overtime_pay(emp.base_salary, data.hours, data.overtime_type)
        )

        start_dt = _parse_hhmm_on_date(data.overtime_date, data.start_time)
        end_dt = _parse_hhmm_on_date(data.overtime_date, data.end_time)

        _validate_overtime_for_employee(
            session,
            data.employee_id,
            data.overtime_date,
            data.overtime_type,
            start_dt,
            end_dt,
            data.hours,
        )

        ot = OvertimeRecord(
            employee_id=data.employee_id,
            overtime_date=data.overtime_date,
            overtime_type=data.overtime_type,
            start_time=start_dt,
            end_time=end_dt,
            hours=data.hours,
            overtime_pay=pay,
            use_comp_leave=data.use_comp_leave,
            reason=data.reason,
            status=ApprovalStatus.PENDING.value,  # Explicitly set to Pending
        )
        session.add(ot)
        session.commit()

        request.state.audit_entity_id = str(ot.id)
        request.state.audit_summary = (
            f"管理端建立加班：employee_id={data.employee_id} "
            f"{data.overtime_type} {data.overtime_date}（{data.hours}h）"
        )
        request.state.audit_changes = {
            "action": "overtime_create",
            "overtime_id": ot.id,
            "employee_id": data.employee_id,
            "overtime_date": data.overtime_date.isoformat(),
            "overtime_type": data.overtime_type,
            "hours": data.hours,
            "start_time": data.start_time,
            "end_time": data.end_time,
            "use_comp_leave": data.use_comp_leave,
            "overtime_pay": pay,
            "reason": data.reason,
        }
        return {"message": "加班記錄已新增", "id": ot.id, "overtime_pay": pay}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.post(
    "/overtimes/batch-create",
    status_code=200,
    response_model=BatchOvertimeCreateResultOut,
    dependencies=[Depends(_batch_approve_limiter)],
)
def batch_create_overtimes(
    data: BatchOvertimeCreate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """一次為多位員工建立加班記錄（學校活動多人出席）。

    全部或全無：Phase 1 對每位員工跑完整驗證並蒐集所有失敗；
    任一失敗 → 422 整批不寫入。Phase 2 全通過才一次 commit。
    每筆狀態 pending，不觸發薪資重算（與單筆建立一致）。
    """
    session = get_session()
    try:
        start_dt = _parse_hhmm_on_date(data.overtime_date, data.start_time)
        end_dt = _parse_hhmm_on_date(data.overtime_date, data.end_time)

        # ── Phase 1：全員驗證（不寫 DB），蒐集所有失敗 ──
        errors: list[dict] = []
        validated: list[tuple[Employee, float]] = []
        seen: set[int] = set()

        for item in data.employees:
            if item.employee_id in seen:
                errors.append(
                    {
                        "employee_id": item.employee_id,
                        "name": None,
                        "reason": "員工在批次清單中重複出現",
                    }
                )
                continue
            seen.add(item.employee_id)

            emp = (
                session.query(Employee).filter(Employee.id == item.employee_id).first()
            )
            if not emp:
                errors.append(
                    {
                        "employee_id": item.employee_id,
                        "name": None,
                        "reason": EMPLOYEE_DOES_NOT_EXIST,
                    }
                )
                continue

            try:
                _validate_overtime_for_employee(
                    session,
                    item.employee_id,
                    data.overtime_date,
                    data.overtime_type,
                    start_dt,
                    end_dt,
                    item.hours,
                )
            except HTTPException as exc:
                errors.append(
                    {
                        "employee_id": item.employee_id,
                        "name": emp.name,
                        "reason": str(exc.detail),
                    }
                )
                continue

            validated.append((emp, item.hours))

        if errors:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "批次建立失敗，請修正下列項目後重送",
                    "errors": errors,
                },
            )

        # ── Phase 2：全通過 → 一次建立 + 單次 commit ──
        records: list[OvertimeRecord] = []
        for emp, hours in validated:
            pay = (
                0.0
                if data.use_comp_leave
                else calculate_overtime_pay(emp.base_salary, hours, data.overtime_type)
            )
            records.append(
                OvertimeRecord(
                    employee_id=emp.id,
                    overtime_date=data.overtime_date,
                    overtime_type=data.overtime_type,
                    start_time=start_dt,
                    end_time=end_dt,
                    hours=hours,
                    overtime_pay=pay,
                    use_comp_leave=data.use_comp_leave,
                    reason=data.reason,
                    status=ApprovalStatus.PENDING.value,
                    comp_leave_granted=False,
                )
            )
        session.add_all(records)
        session.commit()

        created_ids = [r.id for r in records]
        # NOTE: 批次建立涉及多筆，無單一 entity_id，audit_entity_id 刻意略過
        request.state.audit_summary = (
            f"管理端批次建立加班：{len(created_ids)} 筆 "
            f"{data.overtime_type} {data.overtime_date}"
        )
        request.state.audit_changes = {
            "action": "overtime_batch_create",
            "overtime_date": data.overtime_date.isoformat(),
            "overtime_type": data.overtime_type,
            "use_comp_leave": data.use_comp_leave,
            "employee_ids": [emp.id for emp, _ in validated],
            "created_ids": created_ids,
        }
        return {
            "message": f"已建立 {len(created_ids)} 筆加班記錄",
            "created_ids": created_ids,
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/overtimes/{overtime_id}", response_model=OvertimeUpdateResultOut)
def update_overtime(
    overtime_id: int,
    data: OvertimeUpdate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """更新加班記錄。若記錄已核准，修改後自動退回「待審核」狀態以符合稽核要求。"""
    session = get_session()
    try:
        # 列鎖（修補 2026-05-11 P0-3）：與 approve 路徑對齊，防止並發 update+approve
        # 在補休配額 / overtime_pay 重算間產生 lost update。
        ot = (
            session.query(OvertimeRecord)
            .filter(OvertimeRecord.id == overtime_id)
            .with_for_update()
            .first()
        )
        if not ot:
            raise HTTPException(status_code=404, detail=OVERTIME_RECORD_NOT_FOUND)

        # 記錄修改前的核准狀態（供後續稽核退審判斷）+ before snapshot 供 audit_changes
        was_approved = ot.status == ApprovalStatus.APPROVED.value
        original_month = (ot.overtime_date.year, ot.overtime_date.month)
        before_snapshot = {
            "overtime_date": ot.overtime_date.isoformat(),
            "overtime_type": getattr(ot, "overtime_type", None),
            "hours": getattr(ot, "hours", None),
            "start_time": (
                str(ot.start_time) if getattr(ot, "start_time", None) else None
            ),
            "end_time": str(ot.end_time) if getattr(ot, "end_time", None) else None,
            "use_comp_leave": getattr(ot, "use_comp_leave", None),
            "overtime_pay": getattr(ot, "overtime_pay", None),
            "status": ot.status,
            "reason": getattr(ot, "reason", None),
        }

        # 先計算更新後的日期與時間（供重疊檢查使用）
        check_date = data.overtime_date or ot.overtime_date
        date_changed = (
            data.overtime_date is not None and data.overtime_date != ot.overtime_date
        )

        # 改日期但沒重新指定時間時，須以「新日期 + 舊時間」重組 datetime，
        # 否則 datetime 會留在舊日期，造成 overtime_date 與 start/end 日期欄不一致，
        # 且 _check_overtime_overlap 在 SQL 端會用舊日期 datetime 做比較而誤判。
        if data.start_time:
            h, m = map(int, data.start_time.split(":"))
            new_start_dt = datetime.combine(
                check_date, datetime.min.time().replace(hour=h, minute=m)
            )
        elif date_changed and ot.start_time is not None:
            new_start_dt = datetime.combine(check_date, ot.start_time.time())
        else:
            new_start_dt = ot.start_time
        if data.end_time:
            h, m = map(int, data.end_time.split(":"))
            new_end_dt = datetime.combine(
                check_date, datetime.min.time().replace(hour=h, minute=m)
            )
        elif date_changed and ot.end_time is not None:
            new_end_dt = datetime.combine(check_date, ot.end_time.time())
        else:
            new_end_dt = ot.end_time

        if (
            new_start_dt is not None
            and new_end_dt is not None
            and new_start_dt >= new_end_dt
        ):
            raise HTTPException(
                status_code=400,
                detail="start_time 必須早於 end_time（不支援跨日加班）",
            )

        overlap = _check_overtime_overlap(
            session,
            ot.employee_id,
            check_date,
            new_start_dt,
            new_end_dt,
            exclude_id=overtime_id,
        )
        if overlap:
            st = (
                overlap.start_time.strftime("%H:%M") if overlap.start_time else "未指定"
            )
            et = overlap.end_time.strftime("%H:%M") if overlap.end_time else "未指定"
            raise HTTPException(
                status_code=409,
                detail=(
                    f"修改後的時段與已存在的加班申請重疊"
                    f"（ID: {overlap.id}，{overlap.overtime_date} {st}～{et}），請調整時段"
                ),
            )

        # P1-2：跨類重疊檢查——與 create_overtime 對齊。update 原本只查 OT↔OT
        # overlap，可把加班移到已有 approved/pending 假的日子，繞過 2026-05-11
        # P1-5 跨類守衛 → 雙重給付。helper 查的是 LeaveRecord，被改的 OT 不在
        # 該表故無自我衝突、無需 exclude。
        _check_employee_has_conflicting_leave(
            session,
            ot.employee_id,
            check_date,
            new_start_dt,
            new_end_dt,
        )

        # 修改後的時數驗證月上限（排除自己）
        new_hours_val = data.hours if data.hours is not None else ot.hours
        # P1-2：最終 hours 不可超過最終起迄時段差（涵蓋改時段 / 僅改時數對既存時段兩路徑）。
        if new_start_dt is not None and new_end_dt is not None:
            _span_h = (new_end_dt - new_start_dt).total_seconds() / 3600
            if new_hours_val > _span_h + 1e-6:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        f"加班時數（{new_hours_val}）不可超過起迄時段差"
                        f"（{_span_h} 小時）"
                    ),
                )
        _check_monthly_overtime_cap(
            session,
            ot.employee_id,
            check_date,
            new_hours_val,
            exclude_id=overtime_id,
        )
        _check_quarterly_overtime_cap(
            session,
            ot.employee_id,
            check_date,
            new_hours_val,
            exclude_id=overtime_id,
        )

        new_type = (
            data.overtime_type if data.overtime_type is not None else ot.overtime_type
        )
        _check_overtime_type_calendar(session, check_date, new_type)

        recalculation_months = {original_month, (check_date.year, check_date.month)}
        if was_approved:
            assert_months_not_finalized(
                session, employee_id=ot.employee_id, months=recalculation_months
            )
            # commit→recalc 鎖延伸：取 per-emp salary lock 並 pre-mark stale,
            # 封住 caller commit 與 engine 重新 acquire lock 之間 finalize 搶先封存舊薪資的 race。

            lock_and_premark_stale(session, ot.employee_id, recalculation_months)
            _revoke_comp_leave_grant(session, ot, current_user=current_user)

        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            if value is not None and key not in ("start_time", "end_time"):
                setattr(ot, key, value)

        # 改日期或改時間都需同步寫回 ORM，避免 overtime_date 與 start/end 日期欄
        # 形成不一致狀態。
        if data.start_time or (date_changed and ot.start_time is not None):
            ot.start_time = new_start_dt
        if data.end_time or (date_changed and ot.end_time is not None):
            ot.end_time = new_end_dt

        # Recalculate pay（補休模式加班費固定為 0）
        emp = session.query(Employee).filter(Employee.id == ot.employee_id).first()
        if emp:
            ot.overtime_pay = (
                0.0
                if ot.use_comp_leave
                else calculate_overtime_pay(emp.base_salary, ot.hours, ot.overtime_type)
            )

        # ── 稽核退審：已核准的記錄被修改，自動退回待審核 ──────────────────────
        # 防止管理員靜默修改已核准加班時數，導致薪資異常（財務防呆）
        if was_approved:
            ot.status = ApprovalStatus.PENDING.value
            ot.approved_by = None
            logger.warning(
                "稽核警告：已核准加班記錄 #%d（員工 ID=%d, %s）被管理員「%s」修改，"
                "已自動退回待審核狀態，需重新核准",
                overtime_id,
                ot.employee_id,
                ot.overtime_date,
                current_user.get("username", "unknown"),
            )

        session.commit()

        after_snapshot = {
            "overtime_date": ot.overtime_date.isoformat(),
            "overtime_type": getattr(ot, "overtime_type", None),
            "hours": getattr(ot, "hours", None),
            "start_time": (
                str(ot.start_time) if getattr(ot, "start_time", None) else None
            ),
            "end_time": str(ot.end_time) if getattr(ot, "end_time", None) else None,
            "use_comp_leave": getattr(ot, "use_comp_leave", None),
            "overtime_pay": getattr(ot, "overtime_pay", None),
            "status": ot.status,
            "reason": getattr(ot, "reason", None),
        }
        diff = {
            k: {"before": before_snapshot[k], "after": after_snapshot[k]}
            for k in before_snapshot
            if before_snapshot[k] != after_snapshot[k]
        }
        request.state.audit_entity_id = str(overtime_id)
        request.state.audit_summary = (
            f"管理端修改加班 #{overtime_id}（employee_id={ot.employee_id}）"
            + ("；自動退回待審" if was_approved else "")
        )
        request.state.audit_changes = {
            "action": "overtime_update",
            "overtime_id": overtime_id,
            "employee_id": ot.employee_id,
            "was_approved": was_approved,
            "reset_to_pending": was_approved,
            "diff": diff,
            "recalculation_months": [
                f"{y}-{m:02d}" for (y, m) in sorted(recalculation_months)
            ],
        }

        msg = "加班記錄已更新"
        if was_approved:
            msg += "；原核准狀態已自動退回「待審核」，請重新送審"
        result = {
            "message": msg,
            "overtime_pay": ot.overtime_pay,
            "reset_to_pending": was_approved,
        }
        if was_approved:
            try:
                _recalculate_salary_for_overtime_months(
                    ot.employee_id, recalculation_months
                )
                result["salary_recalculated"] = True
            except Exception as e:
                result["salary_recalculated"] = False
                result["warning"] = (
                    "加班記錄已更新，但薪資重算失敗，請手動前往薪資頁面重新計算"
                )
                logger.error("加班修改退審後薪資重算失敗：%s", e)
                try:
                    for year, month in sorted(recalculation_months):
                        _mark_salary_stale(session, ot.employee_id, year, month)
                    session.commit()
                except Exception:
                    logger.warning("加班修改退審降級時標記 stale 失敗", exc_info=True)
                    session.rollback()
        return result
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/overtimes/{overtime_id}", response_model=OvertimeDeleteResultOut)
def delete_overtime(
    overtime_id: int,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """刪除加班記錄"""
    session = get_session()
    try:
        # 列鎖（修補 2026-05-11 P0-3）：與 approve 路徑對齊，防止並發 delete+approve race
        # 在補休配額退還階段 lost update。
        ot = (
            session.query(OvertimeRecord)
            .filter(OvertimeRecord.id == overtime_id)
            .with_for_update()
            .first()
        )
        if not ot:
            raise HTTPException(status_code=404, detail=OVERTIME_RECORD_NOT_FOUND)
        was_approved = ot.status == ApprovalStatus.APPROVED.value
        overtime_month = (ot.overtime_date.year, ot.overtime_date.month)
        employee_id = ot.employee_id
        # 預先 snapshot：刪除後 ot 物件被 expunge，audit 必須在這裡備份
        deleted_snapshot = {
            "overtime_id": overtime_id,
            "employee_id": employee_id,
            "overtime_date": ot.overtime_date.isoformat(),
            "overtime_type": getattr(ot, "overtime_type", None),
            "hours": getattr(ot, "hours", None),
            "use_comp_leave": getattr(ot, "use_comp_leave", None),
            "overtime_pay": getattr(ot, "overtime_pay", None),
            "status": ot.status,
            "reason": getattr(ot, "reason", None),
        }
        if was_approved:
            assert_months_not_finalized(
                session,
                employee_id=employee_id,
                months={overtime_month},
            )
            # commit→recalc 鎖延伸 + pre-mark stale（同 update_overtime 註解）

            lock_and_premark_stale(session, employee_id, {overtime_month})
            _revoke_comp_leave_grant(session, ot, current_user=current_user)
        session.delete(ot)
        session.commit()

        request.state.audit_entity_id = str(overtime_id)
        request.state.audit_summary = (
            f"管理端刪除加班 #{overtime_id}（employee_id={employee_id}）"
            + ("；已核准 → 觸發補休/薪資重算" if was_approved else "")
        )
        request.state.audit_changes = {
            "action": "overtime_delete",
            "deleted": deleted_snapshot,
            "was_approved": was_approved,
            "overtime_month": f"{overtime_month[0]}-{overtime_month[1]:02d}",
            "triggered_salary_recalc": was_approved,
        }

        result = {"message": "加班記錄已刪除"}
        if was_approved:
            try:
                _recalculate_salary_for_overtime_months(employee_id, {overtime_month})
                result["salary_recalculated"] = True
            except Exception as e:
                result["salary_recalculated"] = False
                result["warning"] = (
                    "加班記錄已刪除，但薪資重算失敗，請手動前往薪資頁面重新計算"
                )
                logger.error("刪除加班後薪資重算失敗：%s", e)
                try:
                    _mark_salary_stale(
                        session, employee_id, overtime_month[0], overtime_month[1]
                    )
                    session.commit()
                except Exception:
                    logger.warning("刪除加班降級時標記 stale 失敗", exc_info=True)
                    session.rollback()
        return result
    finally:
        session.close()


@router.put("/overtimes/{overtime_id}/approve", response_model=OvertimeApproveResultOut)
def approve_overtime(
    overtime_id: int,
    request: Request,
    data: Optional[OvertimeApproveRequest] = Body(None),
    # 以下為向後相容 query parameter；新前端應改用 body（修補 2026-05-11 P1-6）
    approved: bool = True,
    approved_by: str = "管理員",
    rejection_reason: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """核准/駁回加班；核准後自動重算該員工當月薪資，補休模式核准後自動累積配額。

    audit P1（2026-05-07）：駁回（approved=False）必填 rejection_reason
    （≥3 字），對齊 leaves / punch_corrections 既有要求；避免管理員惡意
    零原因駁回他人加班費。reason 落 ApprovalLog.comment（OvertimeRecord
    無 rejection_reason 欄位）。

    P1-6 修補（2026-05-11）：body 優先；無 body 才回退 query parameter。
    """
    # body 優先：新前端走 body，rejection_reason 不再寫進 URL log
    if data is not None:
        approved = data.approved
        rejection_reason = data.rejection_reason
    # 駁回必填原因（schema 也驗一次，雙保險）
    if not approved:
        cleaned = (rejection_reason or "").strip()
        if len(cleaned) < 3:
            raise HTTPException(
                status_code=400,
                detail="駁回時必須填寫原因（至少 3 個字）",
            )

    session = get_session()
    try:
        # with_for_update() 鎖定加班記錄，防止並發核准觸發補休配額重複發放（Race Condition）
        ot = (
            session.query(OvertimeRecord)
            .filter(OvertimeRecord.id == overtime_id)
            .with_for_update()
            .first()
        )
        if not ot:
            raise HTTPException(status_code=404, detail=OVERTIME_RECORD_NOT_FOUND)
        was_approved = ot.status == ApprovalStatus.APPROVED.value

        # 既有設計允許「已核准 → 駁回」的合法業務路徑（例如發現超時數需撤銷），
        # 由 risk_tags="reject_of_approved" 在 audit_summary 打標記。撤回 audit
        # 2026-05-07 P1 hard-block flip-flop 提案：業主業務模型即是「approver
        # 可改判」，硬擋會破壞合法 TestApprovedOvertimeRollback 路徑。

        # ── 自我核准防護 ─────────────────────────────────────────────────────────
        if is_self_approval(current_user, ot.employee_id):
            raise HTTPException(status_code=403, detail="不可自我核准加班單")

        # ── 角色資格檢查 ──────────────────────────────────────────────────────
        approver_role = current_user.get("role", "")
        submitter_role = assert_approver_eligible(
            session,
            doc_type="overtime",
            doc_label="加班",
            submitter_employee_id=ot.employee_id,
            approver_role=approver_role,
        )

        if approved or was_approved:
            # 提早取得薪資鎖 + pre-mark stale,封住 commit→recalc 的兩個 race window。
            # caller commit 釋放鎖後,即使 finalize 搶到鎖也會看到 needs_recalc=True 而被擋下;
            # engine 之後在新 session 取鎖重算成功會把 stale 旗標清掉。
            assert_months_not_finalized(
                session,
                employee_id=ot.employee_id,
                months=collect_months_from_dates([ot.overtime_date]),
            )

            lock_and_premark_stale(
                session,
                ot.employee_id,
                {(ot.overtime_date.year, ot.overtime_date.month)},
            )
        if not approved and was_approved:
            _revoke_comp_leave_grant(session, ot, current_user=current_user)

        # ── 核准最後一致性驗證 ───────────────────────────────────────────────
        # 建立路徑（管理端 create、portal、import）都有這些檢查，但舊資料、import
        # 過去版本、或直接改 DB 的紀錄可能違規。核准會進薪資，必須在最後守門。
        if approved and not was_approved:
            if ot.start_time and ot.end_time and ot.start_time >= ot.end_time:
                raise HTTPException(
                    status_code=400,
                    detail="start_time 必須早於 end_time（不支援跨日加班）",
                )
            overlap = _check_overtime_overlap(
                session,
                ot.employee_id,
                ot.overtime_date,
                ot.start_time,
                ot.end_time,
                exclude_id=overtime_id,
            )
            if overlap:
                st = (
                    overlap.start_time.strftime("%H:%M")
                    if overlap.start_time
                    else "未指定"
                )
                et = (
                    overlap.end_time.strftime("%H:%M") if overlap.end_time else "未指定"
                )
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"該員工在 {overlap.overtime_date} 已有時間重疊的加班申請"
                        f"（ID: {overlap.id}，{st}～{et}），請先處理衝突記錄再核准"
                    ),
                )
            _check_monthly_overtime_cap(
                session,
                ot.employee_id,
                ot.overtime_date,
                ot.hours,
                exclude_id=overtime_id,
            )
            _check_quarterly_overtime_cap(
                session,
                ot.employee_id,
                ot.overtime_date,
                ot.hours,
                exclude_id=overtime_id,
            )
            _check_overtime_type_calendar(session, ot.overtime_date, ot.overtime_type)

        ot.status = (
            ApprovalStatus.APPROVED.value if approved else ApprovalStatus.REJECTED.value
        )
        ot.approved_by = current_user.get("username", approved_by) if approved else None
        # OvertimeRecord 無 rejection_reason 欄位（LeaveRecord 才有）；
        # 駁回原因落 ApprovalLog.comment，與 _write_approval_log 同 transaction。
        cleaned_reason: Optional[str] = None
        if not approved:
            cleaned_reason = (rejection_reason or "").strip() or None

        result = {"message": "已核准" if approved else "已駁回"}

        # 補休配額發放（核准時才執行，且防止重複發放）
        if approved:
            _grant_comp_leave_quota(session, ot, result)

        action = "approved" if approved else "rejected"
        approval_log_row = _write_approval_log(
            session=session,
            doc_type="overtime",
            doc_id=overtime_id,
            action=action,
            approver=current_user,
            comment=cleaned_reason,
        )

        # 審核結果通知（dispatch，tx commit 前登錄；after_commit hook 自動 fan-out）
        # lazy import 避免 circular: api.overtimes → dispatch → ws → contact_book_ws → ...
        from services.notification import dispatch

        _overtime_owner_user = (
            session.query(User).filter(User.employee_id == ot.employee_id).first()
        )
        if _overtime_owner_user is not None:
            dispatch.enqueue(
                session=session,
                event_type="overtime.approved" if approved else "overtime.rejected",
                recipient_user_id=_overtime_owner_user.id,
                context={
                    "reviewer_name": current_user.get("name")
                    or current_user.get("username", ""),
                    "ot_date": (
                        ot.overtime_date.isoformat()
                        if hasattr(ot.overtime_date, "isoformat")
                        else str(ot.overtime_date)
                    ),
                    "ot_type": OVERTIME_TYPE_LABELS.get(
                        ot.overtime_type, ot.overtime_type
                    ),
                    "overtime_id": overtime_id,
                },
                sender_id=current_user.get("user_id"),
                source_entity_type="overtime",
                source_entity_id=overtime_id,
            )

        session.commit()

        # 高風險：已核准 → 駁回 顯式標記，AuditLogView 可篩
        is_reject_of_approved = was_approved and not approved
        risk_tags = []
        if is_reject_of_approved:
            risk_tags.append("reject_of_approved")
        request.state.audit_entity_id = str(overtime_id)
        request.state.audit_summary = (
            f"{'核准' if approved else '駁回'}加班 #{overtime_id}"
            f"（employee_id={ot.employee_id}）"
            + (f"｜⚠ {','.join(risk_tags)}" if risk_tags else "")
        )
        request.state.audit_changes = {
            "action": "overtime_approve",
            "overtime_id": overtime_id,
            "employee_id": ot.employee_id,
            "decision": "approved" if approved else "rejected",
            "before": {"status": ("approved" if was_approved else "pending")},
            "after": {
                "status": ot.status,
                "approved_by": ot.approved_by,
            },
            "approval_log_id": approval_log_row.id if approval_log_row else None,
            "is_reject_of_approved": is_reject_of_approved,
            "use_comp_leave": getattr(ot, "use_comp_leave", None),
            "hours": getattr(ot, "hours", None),
            "overtime_pay": getattr(ot, "overtime_pay", None),
            "comp_leave_granted": getattr(ot, "comp_leave_granted", None),
            "risk_tags": risk_tags,
        }

        _notify_and_recalc_overtime(session, ot, approved, was_approved, result)

        return result
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/overtimes/batch-approve", response_model=BatchApproveResultOut)
def batch_approve_overtimes(
    data: OvertimeBatchApproveRequest,
    request: Request,
    _rl=Depends(_batch_approve_limiter),
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """批次核准/駁回加班。兩階段原子提交：先全部驗證，再統一 commit。"""
    succeeded = []
    failed = []

    session = get_session()
    try:
        # ── 預先批次載入，避免 N+1 ────────────────────────────────────────
        ot_map = {
            ot.id: ot
            for ot in session.query(OvertimeRecord)
            .filter(OvertimeRecord.id.in_(data.ids))
            .with_for_update()
            .all()
        }
        emp_ids = {ot.employee_id for ot in ot_map.values()}
        submitter_role_map: dict[int, str] = (
            {
                u.employee_id: u.role
                for u in session.query(User.employee_id, User.role)
                .filter(User.employee_id.in_(emp_ids), User.is_active == True)
                .all()
            }
            if emp_ids
            else {}
        )
        approver_role = current_user.get("role", "")
        _eligibility_cache: dict[str, bool] = {}

        # ── Phase 1：驗證 + 準備所有異動（不 commit）────────────────────────
        changes = []  # list of (ot_id, ot, was_approved)
        for ot_id in data.ids:
            try:
                ot = ot_map.get(ot_id)
                if not ot:
                    failed.append({"id": ot_id, "reason": OVERTIME_RECORD_NOT_FOUND})
                    continue
                was_approved = ot.status == ApprovalStatus.APPROVED.value

                # 防止自我核准
                if is_self_approval(current_user, ot.employee_id):
                    failed.append({"id": ot_id, "reason": "不可自我核准"})
                    continue

                submitter_role = submitter_role_map.get(ot.employee_id, "teacher")
                if submitter_role not in _eligibility_cache:
                    _eligibility_cache[submitter_role] = _check_approval_eligibility(
                        "overtime", submitter_role, approver_role, session
                    )
                if not _eligibility_cache[submitter_role]:
                    failed.append(
                        {
                            "id": ot_id,
                            "reason": f"您的角色（{approver_role}）無權審核此員工（{submitter_role}）的加班申請",
                        }
                    )
                    continue

                if data.approved or was_approved:
                    assert_months_not_finalized(
                        session,
                        employee_id=ot.employee_id,
                        months=collect_months_from_dates([ot.overtime_date]),
                    )

                # 核准最後一致性驗證，防止舊資料 / import 舊版遺留壞紀錄進入薪資
                if data.approved and not was_approved:
                    if ot.start_time and ot.end_time and ot.start_time >= ot.end_time:
                        raise HTTPException(
                            status_code=400,
                            detail="start_time 必須早於 end_time（不支援跨日加班）",
                        )
                    overlap = _check_overtime_overlap(
                        session,
                        ot.employee_id,
                        ot.overtime_date,
                        ot.start_time,
                        ot.end_time,
                        exclude_id=ot_id,
                    )
                    if overlap:
                        st = (
                            overlap.start_time.strftime("%H:%M")
                            if overlap.start_time
                            else "未指定"
                        )
                        et = (
                            overlap.end_time.strftime("%H:%M")
                            if overlap.end_time
                            else "未指定"
                        )
                        raise HTTPException(
                            status_code=409,
                            detail=(
                                f"該員工在 {overlap.overtime_date} 已有時間重疊的加班申請"
                                f"（ID: {overlap.id}，{st}～{et}）"
                            ),
                        )
                    _check_monthly_overtime_cap(
                        session,
                        ot.employee_id,
                        ot.overtime_date,
                        ot.hours,
                        exclude_id=ot_id,
                    )
                    _check_quarterly_overtime_cap(
                        session,
                        ot.employee_id,
                        ot.overtime_date,
                        ot.hours,
                        exclude_id=ot_id,
                    )
                    _check_overtime_type_calendar(
                        session, ot.overtime_date, ot.overtime_type
                    )

                is_reject_of_approved = was_approved and not data.approved
                # Pass 1 純驗證收集；setattr/lock/log/grant 全部移到 Pass 2。
                changes.append(
                    (
                        ot_id,
                        ot,
                        was_approved,
                        is_reject_of_approved,
                        None,  # approval_log_row.id placeholder
                    )
                )
            except HTTPException as he:
                failed.append({"id": ot_id, "reason": he.detail})
            except Exception as e:
                session.rollback()
                failed.append(
                    {
                        "id": ot_id,
                        "reason": safe_batch_reason(e, context="批次加班核准"),
                    }
                )
                session.expire_all()

        # ── Pass 2：套用 setattr + lock + grant + ApprovalLog ────────────
        applied = []
        for ot_id, ot, was_approved, is_reject_of_approved, _ph in list(changes):
            try:
                if data.approved or was_approved:
                    # commit→recalc 鎖延伸：取 per-emp salary lock 並 pre-mark stale。
                    lock_and_premark_stale(
                        session,
                        ot.employee_id,
                        {(ot.overtime_date.year, ot.overtime_date.month)},
                    )
                if not data.approved and was_approved:
                    _revoke_comp_leave_grant(session, ot, current_user=current_user)

                ot.status = (
                    ApprovalStatus.APPROVED.value
                    if data.approved
                    else ApprovalStatus.REJECTED.value
                )
                ot.approved_by = (
                    current_user.get("username", "管理員") if data.approved else None
                )

                if data.approved:
                    # 使用共用 helper 統一配額發放邏輯（含 with_for_update() 列鎖）。
                    _grant_comp_leave_quota(session, ot, {})

                action = "approved" if data.approved else "rejected"
                approval_log_row = _write_approval_log(
                    session=session,
                    doc_type="overtime",
                    doc_id=ot_id,
                    action=action,
                    approver=current_user,
                )
                applied.append(
                    (
                        ot_id,
                        ot,
                        was_approved,
                        is_reject_of_approved,
                        approval_log_row.id if approval_log_row else None,
                    )
                )
            except Exception as e:
                logger.exception("批次核准 Pass 2 套用階段意外失敗（加班 #%d）", ot_id)
                session.rollback()
                for prev_id, *_rest in applied:
                    failed.append(
                        {
                            "id": prev_id,
                            "reason": "同批後續條目失敗導致整批回滾",
                        }
                    )
                failed.append(
                    {
                        "id": ot_id,
                        "reason": safe_batch_reason(e, context="批次加班 Pass2"),
                    }
                )
                applied = []
                break

        changes = applied

        # ── Phase 2：所有驗證通過後統一 commit ──────────────────────────────
        if changes:
            try:
                session.commit()
                for ot_id, ot, was_approved, _ir, _alog in changes:
                    succeeded.append(ot_id)
                    if (data.approved or was_approved) and _salary_engine is not None:
                        try:
                            _salary_engine.process_salary_calculation(
                                ot.employee_id,
                                ot.overtime_date.year,
                                ot.overtime_date.month,
                            )
                        except Exception as se:
                            logger.error(
                                "批次審核後薪資重算失敗（加班 #%d）：%s", ot_id, se
                            )
                            try:
                                _mark_salary_stale(
                                    session,
                                    ot.employee_id,
                                    ot.overtime_date.year,
                                    ot.overtime_date.month,
                                )
                                session.commit()
                            except Exception:
                                logger.warning(
                                    "批次審核降級時標記 stale 失敗（加班 #%d）",
                                    ot_id,
                                    exc_info=True,
                                )
                                session.rollback()
            except Exception as e:
                session.rollback()
                reason = safe_batch_reason(
                    e,
                    context="批次加班統一提交",
                    fallback="統一提交失敗，請稍後重試或聯絡管理員",
                )
                for ot_id, *_ in changes:
                    failed.append({"id": ot_id, "reason": reason})

        # AuditLog changes：批次操作彙整成單筆 audit 摘要
        request.state.audit_summary = (
            f"批次{'核准' if data.approved else '駁回'}加班 "
            f"成功 {len(succeeded)}/失敗 {len(failed)}（請求 {len(data.ids)} 筆）"
        )
        request.state.audit_changes = {
            "action": "overtime_batch_approve",
            "decision": "approved" if data.approved else "rejected",
            "requested_ids": data.ids,
            "succeeded_ids": list(succeeded),
            "failed": failed,
            "approval_log_ids": [
                {
                    "overtime_id": _id,
                    "employee_id": _ot.employee_id,
                    "is_reject_of_approved": _ir,
                    "approval_log_id": _alog,
                }
                for _id, _ot, _wa, _ir, _alog in changes
            ],
            "high_risk_count": sum(1 for _id, _ot, _wa, _ir, _alog in changes if _ir),
        }
    finally:
        session.close()

    return {"succeeded": succeeded, "failed": failed}


@router.get("/overtimes/import-template")
def get_overtime_import_template(
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """下載加班批次匯入 Excel 範本"""
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "加班匯入範本"

    headers = [
        "員工編號",
        "員工姓名",
        "加班日期",
        "加班類型",
        "時數",
        "開始時間(可空)",
        "結束時間(可空)",
        "原因(可空)",
        "補休(是/否,可空)",
    ]
    _ot_write_header(ws, 1, headers)

    ws.cell(row=2, column=1, value="E001")
    ws.cell(row=2, column=2, value="王小明")
    ws.cell(row=2, column=3, value="2026-03-15")
    ws.cell(row=2, column=4, value="weekday")
    ws.cell(row=2, column=5, value=2)
    ws.cell(row=2, column=6, value="18:00")
    ws.cell(row=2, column=7, value="20:00")
    ws.cell(row=2, column=8, value="開學準備")
    ws.cell(row=2, column=9, value="否")

    ws2 = SafeWorksheet(wb.create_sheet("加班類型說明"))
    ws2.cell(row=1, column=1, value="類型代碼")
    ws2.cell(row=1, column=2, value="說明")
    ws2.cell(row=1, column=3, value="加班費倍率")
    ws2.cell(row=2, column=1, value="weekday")
    ws2.cell(row=2, column=2, value="平日加班")
    ws2.cell(row=2, column=3, value="前2h×1.34，後2h×1.67")
    ws2.cell(row=3, column=1, value="weekend")
    ws2.cell(row=3, column=2, value="假日加班")
    ws2.cell(row=3, column=3, value="×2.0")
    ws2.cell(row=4, column=1, value="holiday")
    ws2.cell(row=4, column=2, value="國定假日加班")
    ws2.cell(row=4, column=3, value="×2.0")

    return xlsx_streaming_response(wb, "加班匯入範本.xlsx")


class OvertimeImportRow(ExcelImportSchema):
    """加班匯入單列 schema（excel header 用中文 alias）。

    僅負責「Excel 列 → 結構化資料」轉換；日期/時數/時間/類型的業務驗證
    仍留在 endpoint（pandas 之前怎麼處理、現在保留相同行為）。
    """

    employee_code: Optional[str] = Field(default=None, alias="員工編號")
    employee_name: Optional[str] = Field(default=None, alias="員工姓名")
    overtime_date: Any = Field(default=None, alias="加班日期")
    overtime_type: Optional[str] = Field(default=None, alias="加班類型")
    hours: Any = Field(default=None, alias="時數")
    start_time: Any = Field(default=None, alias="開始時間(可空)")
    end_time: Any = Field(default=None, alias="結束時間(可空)")
    reason: Any = Field(default=None, alias="原因(可空)")
    use_comp_leave: Any = Field(default=None, alias="補休(是/否,可空)")


@router.post("/overtimes/import", response_model=OvertimeImportResultOut)
async def import_overtimes(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_staff_permission(Permission.OVERTIME_WRITE)),
):
    """批次匯入加班申請（建立草稿加班單，status='pending'，需後續人工審核）"""
    content = await read_upload_with_size_check(file)
    # parse + DB 迴圈為同步 CPU/IO，卸載到 executor 避免阻塞 event loop（行為不變）
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _import_overtimes_sync, content)


def _import_overtimes_sync(content: bytes) -> dict:
    validate_file_signature(content, ".xlsx")

    parse_result = parse_excel(BytesIO(content), schema=OvertimeImportRow)
    # 整檔級錯誤（INVALID_FILE / EMPTY_FILE / MISSING_COLUMN）→ 400，與既有行為一致
    file_level_codes = {"INVALID_FILE", "EMPTY_FILE", "MISSING_COLUMN"}
    file_level_errors = [
        e for e in parse_result.errors if e["error_code"] in file_level_codes
    ]
    if file_level_errors:
        msg = file_level_errors[0]["message"]
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{msg}")

    results: dict = {"total": 0, "created": 0, "failed": 0, "errors": []}
    # parse_excel 的 row-level errors（型別/缺欄）先映射回舊格式 "第 N 行: msg"
    for err in parse_result.errors:
        if err["error_code"] in file_level_codes:
            continue
        results["total"] += 1
        results["failed"] += 1
        results["errors"].append(f"第 {err['row']} 行: {err['message']}")

    session = get_session()
    try:
        emp_by_id, emp_by_name = build_employee_lookup(session)

        for row_idx, row in enumerate(parse_result.rows, start=2):
            results["total"] += 1
            row_num = row_idx
            try:
                # resolve_employee_from_row 接受 dict-like / pandas Series；
                # pydantic BaseModel 補上原中文 keys 維持既有 lookup 行為。
                row_dict = {
                    "員工編號": row.employee_code,
                    "員工姓名": row.employee_name,
                }
                emp = resolve_employee_from_row(row_dict, emp_by_id, emp_by_name)

                ot_date_raw = row.overtime_date
                if ot_date_raw is None:
                    raise ValueError("加班日期不得為空")
                try:
                    # pd.to_datetime 對 str / datetime / Timestamp 均能 normalize，
                    # 維持與舊 endpoint 相同的日期解析行為（包含 Excel 數值序號）
                    overtime_date = pd.to_datetime(ot_date_raw).date()
                except Exception:
                    raise ValueError("加班日期格式錯誤，建議使用 YYYY-MM-DD")

                ot_type_raw = (
                    str(row.overtime_type).strip()
                    if row.overtime_type is not None
                    else ""
                )
                if ot_type_raw not in OVERTIME_TYPE_LABELS:
                    raise ValueError(
                        f"無效的加班類型：{ot_type_raw}（可用：weekday/weekend/holiday）"
                    )

                hours_raw = row.hours
                if hours_raw is None:
                    raise ValueError("時數不得為空")
                try:
                    hours = float(hours_raw)
                except (TypeError, ValueError):
                    raise ValueError("時數必須為數字")
                if hours <= 0:
                    raise ValueError("時數必須大於 0")
                if hours > MAX_OVERTIME_HOURS:
                    raise ValueError(f"時數不得超過 {MAX_OVERTIME_HOURS} 小時")

                start_dt = None
                end_dt = None
                for col_name, raw_val, is_start in [
                    ("開始時間(可空)", row.start_time, True),
                    ("結束時間(可空)", row.end_time, False),
                ]:
                    if raw_val is not None:
                        val_str = str(raw_val).strip()
                        if val_str and val_str not in ("nan", ""):
                            try:
                                h, m = map(int, val_str.split(":")[:2])
                                if not (0 <= h <= 23 and 0 <= m <= 59):
                                    raise ValueError(
                                        f"{col_name} 時間值超出範圍（小時 0-23，分鐘 0-59）"
                                    )
                                dt = datetime.combine(
                                    overtime_date,
                                    datetime.min.time().replace(hour=h, minute=m),
                                )
                                if is_start:
                                    start_dt = dt
                                else:
                                    end_dt = dt
                            except ValueError:
                                raise
                            except Exception:
                                raise ValueError(f"{col_name} 格式錯誤，應為 HH:MM")

                if start_dt is not None and end_dt is not None and start_dt >= end_dt:
                    raise ValueError("開始時間必須早於結束時間（不支援跨日加班）")

                # P1-2：匯入路徑 hours 不可超過起迄時段差（防超報溢付，業主定案 ≤ 時段差）
                if start_dt is not None and end_dt is not None:
                    _span_h = (end_dt - start_dt).total_seconds() / 3600
                    if hours > _span_h + 1e-6:
                        raise ValueError(
                            f"加班時數（{hours}）不可超過起迄時段差（{_span_h} 小時）"
                        )

                comp_raw = row.use_comp_leave
                use_comp_leave = False
                if comp_raw is not None:
                    use_comp_leave = str(comp_raw).strip() in (
                        "是",
                        "yes",
                        "Yes",
                        "YES",
                        "true",
                        "True",
                        "1",
                    )

                pay = (
                    0.0
                    if use_comp_leave
                    else calculate_overtime_pay(emp.base_salary, hours, ot_type_raw)
                )

                overlap = _check_overtime_overlap(
                    session, emp.id, overtime_date, start_dt, end_dt
                )
                if overlap:
                    st = (
                        overlap.start_time.strftime("%H:%M")
                        if overlap.start_time
                        else "未指定"
                    )
                    et = (
                        overlap.end_time.strftime("%H:%M")
                        if overlap.end_time
                        else "未指定"
                    )
                    raise ValueError(
                        f"該員工在 {overlap.overtime_date} 已有時間重疊的加班申請"
                        f"（ID: {overlap.id}，{st}～{et}），請勿重複匯入"
                    )

                _check_monthly_overtime_cap(session, emp.id, overtime_date, hours)
                _check_quarterly_overtime_cap(session, emp.id, overtime_date, hours)
                _check_overtime_type_calendar(session, overtime_date, ot_type_raw)

                reason_raw = row.reason
                reason = str(reason_raw).strip() if reason_raw is not None else None

                ot = OvertimeRecord(
                    employee_id=emp.id,
                    overtime_date=overtime_date,
                    overtime_type=ot_type_raw,
                    start_time=start_dt,
                    end_time=end_dt,
                    hours=hours,
                    overtime_pay=pay,
                    use_comp_leave=use_comp_leave,
                    reason=reason,
                    status=ApprovalStatus.PENDING.value,
                )
                session.add(ot)
                session.flush()
                results["created"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"第 {row_num} 行: {str(e)}")

        session.commit()
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e, context="匯入失敗")
    finally:
        session.close()

    return results
