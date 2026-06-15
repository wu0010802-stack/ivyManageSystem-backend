"""半年考核 Excel I/O（M2）— 雙向轉換。

支援格式：
- 讀入：`.xls`（Excel 97-2003，xlrd）+ `.xlsx`（openpyxl）
- 匯出：`.xlsx`（openpyxl）— 對應 Excel 原始三 sheets 結構

主要功能：
  parse_half_year_excel(path) → ParsedHalfYearExcel
  import_half_year_to_db(parsed, session, ...) → ImportResult
  export_half_year_xlsx(cycle, participants, summaries, ...) → bytes
  export_transfer_roster_xlsx(participants_with_bonus, bank_info) → bytes

對映 Excel「114(上)年度考核統計表」結構：
- r00..r02 header（學校名稱、標題、報表編號）
- r03 欄位編號 1-16
- r04 欄位名稱
- r05 空
- r06+ 員工資料；遇到「分數=...」說明行結束
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable, Optional

from sqlalchemy.orm import Session

from models.appraisal import (
    AppraisalCycle,
    AppraisalParticipant,
    AppraisalScoreItem,
    AppraisalScoreItemCatalog,
    AppraisalSummary,
    CycleStatus,
    Grade,
    RoleGroup,
    Semester,
)
from utils.excel_utils import SafeWorksheet

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 對映：Excel 欄位 index → item_code
# ---------------------------------------------------------------------------

# 對應 plan §3 半年表 16 項；index 為 r04 欄位列的 column index
EXCEL_COL_TO_ITEM_CODE: dict[int, str] = {
    2: "LEAVE",
    3: "LATE_EARLY",
    4: "NO_CLOCK",
    5: "MISS_PRESCHOOL_MEETING",
    6: "ORG_MEETING_0913",
    7: "ORG_MEETING_1115",
    8: "TEAM_ACTIVITY_1115",
    9: "DROPOUT_0915",
    10: "DROPOUT_0315",
    11: "CHILD_INCIDENT",
    12: "RETURNING_RATE_0315",
    13: "CLASS_SIZE",
    14: "AFTER_CLASS_RATE",
    15: "SPED",
    16: "REWARD_PUNISH",
}

# 合計 / 等第 / 獎金 / 事假備註 / 病假備註 欄
COL_TOTAL_SCORE = 17
COL_GRADE = 18
COL_BONUS = 19
COL_LEAVE_NOTE_1 = 20  # 事假
COL_LEAVE_NOTE_2 = 21  # 病假

# 等第中文對應
GRADE_LABEL_TO_ENUM: dict[str, Grade] = {
    "優等": Grade.OUTSTANDING,
    "甲等": Grade.GOOD,
    "乙等": Grade.PASS,
    "丙等": Grade.WARN,
    "丁等": Grade.FAIL,
}
GRADE_ENUM_TO_LABEL: dict[Grade, str] = {v: k for k, v in GRADE_LABEL_TO_ENUM.items()}

EXCLUSION_KEYWORDS = ("不計算考核", "未簽約", "簽約")
TITLE_PATTERN = re.compile(r"(\d+)\((上|下)\)年度考核")


# ---------------------------------------------------------------------------
# Parser 結果 dataclass
# ---------------------------------------------------------------------------


@dataclass
class ParsedScoreItem:
    item_code: str
    score_delta: Decimal
    raw_value: Optional[Decimal] = None
    note: Optional[str] = None


@dataclass
class ParsedParticipantRow:
    excel_row: int
    name: str
    score_items: list[ParsedScoreItem] = field(default_factory=list)
    total_score: Optional[Decimal] = None
    grade: Optional[Grade] = None
    bonus_amount: Optional[Decimal] = None
    leave_note: Optional[str] = None
    is_excluded: bool = False
    exclude_reason: Optional[str] = None


@dataclass
class ParsedHalfYearExcel:
    academic_year: int
    semester: Semester
    base_score: Decimal
    title: str
    sheet_name: str
    participants: list[ParsedParticipantRow] = field(default_factory=list)


@dataclass
class ImportResult:
    cycle_id: int
    participants_created: int
    participants_updated: int
    score_items_upserted: int
    summaries_upserted: int
    skipped_unresolved_names: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 1. Parser — 純函式，無 DB 依賴
# ---------------------------------------------------------------------------


def _to_decimal(value, default: Optional[Decimal] = None) -> Optional[Decimal]:
    """Excel cell → Decimal；空字串/None → default。"""
    if value is None or value == "" or value == 0:
        if value == 0:
            return Decimal("0")
        return default
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return default


def _normalize_cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_exclusion_marker(text: str) -> bool:
    """偵測「114.12.22到職,115.01.19簽約不計算考核」這類字串。"""
    return any(kw in text for kw in EXCLUSION_KEYWORDS)


def _parse_title(title: str) -> tuple[int, Semester]:
    """從「114(上)年度考核統計表(...)」抽出 (academic_year=114, FIRST)。"""
    m = TITLE_PATTERN.search(title)
    if not m:
        raise ValueError(f"無法從標題解析學年/學期：{title!r}")
    year = int(m.group(1))
    semester = Semester.FIRST if m.group(2) == "上" else Semester.SECOND
    return year, semester


def _read_sheet_rows(path: Path | str) -> tuple[str, list[list]]:
    """讀取主表 Sheet 0 → (sheet_name, [row[col], ...])。

    支援 .xls 與 .xlsx；統一回傳 list of list of raw cell values。
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".xls":
        import xlrd  # type: ignore[import-untyped]

        book = xlrd.open_workbook(str(p))
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            rows.append(row)
        return sheet.name, rows
    elif ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(str(p), data_only=True, read_only=True)
        sheet = wb.worksheets[0]
        rows = []
        for r_idx, raw_row in enumerate(sheet.iter_rows(values_only=True)):
            rows.append(list(raw_row))
        return sheet.title, rows
    else:
        raise ValueError(f"不支援的副檔名：{ext}，僅支援 .xls / .xlsx")


def parse_half_year_excel(path: Path | str) -> ParsedHalfYearExcel:
    """解析半年考核 Excel → ParsedHalfYearExcel（純資料，可單元測試）。"""
    sheet_name, rows = _read_sheet_rows(path)
    if len(rows) < 6:
        raise ValueError(f"Sheet '{sheet_name}' 列數過少 ({len(rows)})，疑似格式錯誤")

    # 從第 1-2 行找標題 — Excel 範例 r02 是「114(上)年度考核統計表...」
    title = ""
    for r in range(min(6, len(rows))):
        for c in range(min(5, len(rows[r]))):
            text = _normalize_cell_text(rows[r][c])
            if "年度考核" in text:
                title = text
                break
        if title:
            break
    if not title:
        raise ValueError(f"Sheet '{sheet_name}' 找不到「年度考核」標題行")
    academic_year, semester = _parse_title(title)

    # 員工資料從 r06 開始；遇到「分數=...」「優等：...」「執行長：」等說明行結束
    participants: list[ParsedParticipantRow] = []
    base_score: Optional[Decimal] = None
    for r in range(5, len(rows)):
        row = rows[r]
        if not row:
            continue
        name = _normalize_cell_text(row[0]) if len(row) > 0 else ""
        if not name:
            continue
        if any(name.startswith(prefix) for prefix in ("分數=", "優等", "甲等", "乙等", "丙等", "丁等", "考察", "連續", "休學人數=", "*", "執行長", "主任", "行政")):
            # 說明行 / 簽核行 → 結束資料區
            if any(name.startswith(p) for p in ("執行長", "主任", "行政", "分數=")):
                break
            continue

        # 偵測「不計算考核」標記行（姓名後接該字串於下一欄）
        col_1_text = _normalize_cell_text(row[1]) if len(row) > 1 else ""
        if _is_exclusion_marker(col_1_text):
            participants.append(
                ParsedParticipantRow(
                    excel_row=r,
                    name=name,
                    is_excluded=True,
                    exclude_reason=col_1_text,
                )
            )
            continue

        # base_score 從第一筆「正常員工」row 抓 col 1
        score_items: list[ParsedScoreItem] = []
        if base_score is None:
            bs = _to_decimal(col_1_text)
            if bs is not None and bs > 0:
                base_score = bs.quantize(Decimal("0.1"))

        # 16 項加減分（col 2-16）
        for col_idx, item_code in EXCEL_COL_TO_ITEM_CODE.items():
            if col_idx >= len(row):
                break
            delta = _to_decimal(row[col_idx], default=None)
            if delta is None or delta == Decimal("0"):
                continue
            score_items.append(
                ParsedScoreItem(item_code=item_code, score_delta=delta)
            )

        total_score = (
            _to_decimal(row[COL_TOTAL_SCORE]) if len(row) > COL_TOTAL_SCORE else None
        )
        grade_label = (
            _normalize_cell_text(row[COL_GRADE]) if len(row) > COL_GRADE else ""
        )
        grade = GRADE_LABEL_TO_ENUM.get(grade_label)
        bonus = (
            _to_decimal(row[COL_BONUS]) if len(row) > COL_BONUS else None
        )
        leave_parts: list[str] = []
        for col in (COL_LEAVE_NOTE_1, COL_LEAVE_NOTE_2):
            if col < len(row):
                t = _normalize_cell_text(row[col])
                if t:
                    leave_parts.append(t)
        leave_note = " / ".join(leave_parts) if leave_parts else None

        participants.append(
            ParsedParticipantRow(
                excel_row=r,
                name=name,
                score_items=score_items,
                total_score=total_score,
                grade=grade,
                bonus_amount=bonus,
                leave_note=leave_note,
            )
        )

    if base_score is None:
        base_score = Decimal("0")
    return ParsedHalfYearExcel(
        academic_year=academic_year,
        semester=semester,
        base_score=base_score,
        title=title,
        sheet_name=sheet_name,
        participants=participants,
    )


# ---------------------------------------------------------------------------
# 2. Importer — 寫入 DB（upsert）
# ---------------------------------------------------------------------------


def _build_catalog_index(session: Session) -> dict[str, AppraisalScoreItemCatalog]:
    items = session.query(AppraisalScoreItemCatalog).all()
    return {item.code: item for item in items}


def import_half_year_to_db(
    parsed: ParsedHalfYearExcel,
    session: Session,
    employee_resolver: Callable[[str], Optional[int]],
    *,
    role_group_resolver: Callable[[int], RoleGroup],
    cycle_dates: tuple[date, date, date],
    classroom_resolver: Optional[Callable[[int], Optional[int]]] = None,
) -> ImportResult:
    """將 ParsedHalfYearExcel 寫入 DB（upsert）。

    Args:
        parsed: parse_half_year_excel 的回傳值
        session: SQLAlchemy session（呼叫方負責 commit / rollback）
        employee_resolver: name → employee_id（None=找不到該員工，跳過）
        role_group_resolver: employee_id → RoleGroup
        cycle_dates: (start_date, end_date, base_score_calc_date)
        classroom_resolver: 可選；employee_id → classroom_id

    Returns:
        ImportResult，含 upsert 統計與未匹配名單。
    """
    start_date, end_date, base_calc_date = cycle_dates

    # cycle upsert
    cycle = (
        session.query(AppraisalCycle)
        .filter_by(academic_year=parsed.academic_year, semester=parsed.semester)
        .one_or_none()
    )
    if cycle is None:
        cycle = AppraisalCycle(
            academic_year=parsed.academic_year,
            semester=parsed.semester,
            start_date=start_date,
            end_date=end_date,
            base_score_calc_date=base_calc_date,
            base_score=parsed.base_score,
            status=CycleStatus.OPEN,
        )
        session.add(cycle)
        session.flush()
    else:
        cycle.base_score = parsed.base_score
        cycle.base_score_calc_date = base_calc_date

    catalog_idx = _build_catalog_index(session)

    participants_created = 0
    participants_updated = 0
    score_items_upserted = 0
    summaries_upserted = 0
    skipped: list[str] = []

    for prow in parsed.participants:
        employee_id = employee_resolver(prow.name)
        if employee_id is None:
            skipped.append(prow.name)
            logger.warning(
                "import_half_year: 找不到員工 %r（excel_row=%s），跳過",
                prow.name,
                prow.excel_row,
            )
            continue

        role_group = role_group_resolver(employee_id)
        classroom_id = (
            classroom_resolver(employee_id) if classroom_resolver else None
        )

        # participant upsert
        participant = (
            session.query(AppraisalParticipant)
            .filter_by(cycle_id=cycle.id, employee_id=employee_id)
            .one_or_none()
        )
        if participant is None:
            participant = AppraisalParticipant(
                cycle_id=cycle.id,
                employee_id=employee_id,
                role_group=role_group,
                classroom_id=classroom_id,
                hire_months_in_cycle=Decimal("6"),
                is_excluded=prow.is_excluded,
                exclude_reason=prow.exclude_reason,
            )
            session.add(participant)
            session.flush()
            participants_created += 1
        else:
            participant.role_group = role_group
            participant.classroom_id = classroom_id
            participant.is_excluded = prow.is_excluded
            participant.exclude_reason = prow.exclude_reason
            participants_updated += 1

        if prow.is_excluded:
            # 跳過 score_items 與 summary 寫入；excluded participant 不入分
            continue

        # score_items upsert（依 (participant_id, item_code, sequence_no=1) 對應）
        for sitem in prow.score_items:
            catalog = catalog_idx.get(sitem.item_code)
            existing = (
                session.query(AppraisalScoreItem)
                .filter_by(
                    participant_id=participant.id,
                    item_code=sitem.item_code,
                    sequence_no=1,
                )
                .one_or_none()
            )
            if existing is None:
                session.add(
                    AppraisalScoreItem(
                        participant_id=participant.id,
                        cycle_id=cycle.id,
                        catalog_id=catalog.id if catalog else None,
                        item_code=sitem.item_code,
                        sequence_no=1,
                        score_delta=sitem.score_delta,
                        raw_value=sitem.raw_value,
                        note=sitem.note,
                        source_ref=f"excel_row={prow.excel_row}",
                    )
                )
            else:
                existing.catalog_id = catalog.id if catalog else None
                existing.score_delta = sitem.score_delta
                existing.raw_value = sitem.raw_value
                existing.note = sitem.note
                existing.source_ref = f"excel_row={prow.excel_row}"
            score_items_upserted += 1

        # summary upsert（base + sum 直接用 Excel 既有合計，但仍保留欄位以便 engine 重算）
        event_sum = sum(
            (Decimal(s.score_delta) for s in prow.score_items), Decimal("0")
        )
        summary = (
            session.query(AppraisalSummary)
            .filter_by(participant_id=participant.id)
            .one_or_none()
        )
        if summary is None:
            summary = AppraisalSummary(
                participant_id=participant.id,
                cycle_id=cycle.id,
                base_score=parsed.base_score,
                event_score_sum=event_sum,
                total_score=prow.total_score or (parsed.base_score + event_sum),
                grade=prow.grade or Grade.FAIL,
                bonus_amount=prow.bonus_amount or Decimal("0"),
                leave_note=prow.leave_note,
            )
            session.add(summary)
        else:
            summary.base_score = parsed.base_score
            summary.event_score_sum = event_sum
            summary.total_score = prow.total_score or (parsed.base_score + event_sum)
            summary.grade = prow.grade or summary.grade
            summary.bonus_amount = prow.bonus_amount or Decimal("0")
            summary.leave_note = prow.leave_note
        summaries_upserted += 1

    session.flush()
    return ImportResult(
        cycle_id=cycle.id,
        participants_created=participants_created,
        participants_updated=participants_updated,
        score_items_upserted=score_items_upserted,
        summaries_upserted=summaries_upserted,
        skipped_unresolved_names=skipped,
    )


# ---------------------------------------------------------------------------
# 3. Exporter — 重生與 Excel 同欄位的 .xlsx
# ---------------------------------------------------------------------------


@dataclass
class ExportRow:
    name: str
    score_items: dict[str, Decimal]  # item_code → score_delta
    total_score: Decimal
    grade: Grade
    bonus_amount: Decimal
    leave_note: Optional[str] = None
    is_excluded: bool = False
    exclude_reason: Optional[str] = None


def export_half_year_xlsx(
    *,
    title: str,
    academic_year: int,
    semester: Semester,
    base_score: Decimal,
    rows: Iterable[ExportRow],
    school_name: str = "高雄市私立常春藤幼兒園",
    report_no: str = "",
) -> bytes:
    """產出與 Excel 原始版相同欄位順序的 .xlsx bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    sem_label = "上" if semester == Semester.FIRST else "下"
    ws.title = f"常春藤{academic_year}{sem_label}"

    # r00 報表編號
    if report_no:
        ws.cell(row=1, column=19, value=report_no)
    # r01 學校
    ws.cell(row=2, column=1, value=school_name).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=22)
    # r02 標題
    ws.cell(row=3, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=22)
    # r03 欄位編號 1-16
    for i in range(1, 17):
        ws.cell(row=4, column=i + 1, value=i)
    # r04 欄位名稱（中文）
    header_labels = [
        "姓名",
        f"{'9/15' if semester == Semester.FIRST else '3/15'} 分數",
        "請休假",
        "遲到\n早退",
        "未\n打\n卡",
        "園務會議未參加",
        "9/13\n機構會議研習",
        "11/15\n機構會議研習",
        "11/15自強活動",
        "9/15\n休學人數",
        "3/15\n休學人數",
        "幼兒意外",
        "3/15\n舊生註冊率",
        "帶班人數",
        "才藝班參加率",
        "特別辦法\n特教生",
        "獎懲",
        "合計",
        "等第",
        "獎金",
        "事假",
        "病假",
    ]
    for c, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=5, column=c, value=label)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.font = Font(bold=True)

    # r06+ 員工資料
    item_code_order = list(EXCEL_COL_TO_ITEM_CODE.values())  # 對應 col 2-16
    for idx, row_data in enumerate(rows, start=1):
        excel_row = 5 + idx  # r06 起
        ws.cell(row=excel_row, column=1, value=row_data.name)
        if row_data.is_excluded:
            ws.cell(row=excel_row, column=2, value=row_data.exclude_reason or "不計算考核")
            continue
        ws.cell(row=excel_row, column=2, value=float(base_score))
        for col_offset, item_code in enumerate(item_code_order, start=3):
            v = row_data.score_items.get(item_code)
            if v is not None and v != 0:
                ws.cell(row=excel_row, column=col_offset, value=float(v))
        ws.cell(row=excel_row, column=COL_TOTAL_SCORE + 1, value=float(row_data.total_score))
        ws.cell(
            row=excel_row,
            column=COL_GRADE + 1,
            value=GRADE_ENUM_TO_LABEL.get(row_data.grade, ""),
        )
        if row_data.bonus_amount > 0:
            ws.cell(row=excel_row, column=COL_BONUS + 1, value=float(row_data.bonus_amount))
        if row_data.leave_note:
            ws.cell(row=excel_row, column=COL_LEAVE_NOTE_1 + 1, value=row_data.leave_note)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class TransferRow:
    bank_account: str
    name: str
    amount: Decimal


def export_transfer_roster_xlsx(
    *,
    rows: Iterable[TransferRow],
    school_name: str = "高雄市私立常春藤幼兒園",
    title: str = "考核轉帳名冊",
    org_bank_account: str = "",
) -> bytes:
    """產出轉帳名冊 .xlsx — 對應 Excel「114上轉帳名冊」結構。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "轉帳名冊"

    ws.cell(row=1, column=2, value=school_name).font = Font(bold=True, size=14)
    ws.cell(row=2, column=2, value=title).font = Font(bold=True, size=12)
    if org_bank_account:
        ws.cell(row=3, column=2, value=f"帳號：{org_bank_account}")
    headers = ("帳號", "戶名", "金額")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    total = Decimal("0")
    r = 5
    for row in rows:
        if row.amount <= 0:
            continue
        ws.cell(row=r, column=1, value=row.bank_account)
        ws.cell(row=r, column=2, value=row.name)
        ws.cell(row=r, column=3, value=float(row.amount))
        total += row.amount
        r += 1
    ws.cell(row=r + 1, column=2, value="合計").font = Font(bold=True)
    ws.cell(row=r + 1, column=3, value=float(total)).font = Font(bold=True)
    ws.cell(row=r + 2, column=2, value="經辦人：")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
