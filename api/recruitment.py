"""
api/recruitment.py — 招生統計 API endpoints
"""

import logging
import re
from datetime import date, datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import func, or_, and_, case, cast, String

from models.base import session_scope
from models.recruitment import (
    RecruitmentVisit,
    RecruitmentPeriod,
    RecruitmentMonth,
    RecruitmentGeocodeCache,
)
from services import recruitment_market_intelligence as market_service
from utils.auth import require_staff_permission
from utils.excel_utils import xlsx_streaming_response
from utils.permissions import Permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/recruitment", tags=["recruitment"])

# ---------------------------------------------------------------------------
# 常數
# ---------------------------------------------------------------------------

NO_DEPOSIT_REASONS = [
    "時程未到／仍在觀望",
    "已有其他就學選項／比較他校",
    "未註明／待追蹤",
    "距離／地點因素",
    "家庭照顧安排考量",
    "特殊需求／名額限制",
    "課程／環境仍在評估",
    "費用考量",
]

TOP_SOURCES_COUNT = 10  # 接待×來源交叉表顯示最大來源數
ADDRESS_SYNC_MODES = {"incremental", "resync_google"}
HOTSPOT_CACHE_LOOKUP_CHUNK_SIZE = 200
DATASET_SCOPE_ALL = "all"
HIGH_PRIORITY_NO_DEPOSIT_REASONS = {
    "時程未到／仍在觀望",
    "課程／環境仍在評估",
}
MEDIUM_PRIORITY_NO_DEPOSIT_REASONS = {
    "距離／地點因素",
    "費用考量",
    "家庭照顧安排考量",
}
LOW_PRIORITY_NO_DEPOSIT_REASONS = {
    "已有其他就學選項／比較他校",
    "特殊需求／名額限制",
}
NO_DEPOSIT_PRIORITY_REASON_MAP = {
    "high": HIGH_PRIORITY_NO_DEPOSIT_REASONS,
    "medium": MEDIUM_PRIORITY_NO_DEPOSIT_REASONS,
    "low": LOW_PRIORITY_NO_DEPOSIT_REASONS,
}
FUNNEL_DROP_THRESHOLD = 10.0
HIGH_POTENTIAL_BACKLOG_THRESHOLD = 5
DEFAULT_OVERDUE_DAYS = 14
COLD_LEAD_DAYS = 90
SOURCE_IMBALANCE_SHARE_THRESHOLD = 40.0
ACTION_QUEUE_LIMIT = 3

# 童年綠地判定關鍵字
_CHUANNIAN_KW = "童年綠地"
_YAOTING_KW = "班導-雅婷"
_BRANCH_INTRO_KW = "分校介紹"
_SOURCE_GROUP_ALIASES = {
    _CHUANNIAN_KW: {"二人同行", "七人同行", "9人同行", "九人同行", "Ruby老師"},
    _BRANCH_INTRO_KW: {
        "國際校介紹",
        "仁武校介紹",
        "明華校介紹",
        "崇德校介紹",
        _BRANCH_INTRO_KW,
    },
}

# 就讀月份 / 班別 regex
_EXPECTED_MONTH_RE = re.compile(r"(1\d\d)\.(\d{1,2})")
_GRADE_RE = re.compile(r"(幼幼班|小班|中班|大班)")
_VISIT_DATE_MONTH_RE = re.compile(r"(?<!\d)(\d{3})[./-](\d{1,2})[./-]\d{1,2}")

# 期間名稱解析 regex："114.09.16~115.03.15" 或 "114.09.16-115.03.15"
_PERIOD_RANGE_RE = re.compile(r"(\d{3}\.\d{2})\.\d{2}[~\-](\d{3}\.\d{2})\.\d{2}")

# ---------------------------------------------------------------------------
# 純函式 helpers
# ---------------------------------------------------------------------------


def _extract_expected_label_from_text(
    notes: Optional[str],
    parent_response: Optional[str],
    grade: Optional[str],
) -> str:
    """從 notes / parent_response 解析「預計就讀月份＋班別」。
    取最後一個民國年月匹配（通常為最終確認），後向 30 字找班別。
    """
    text = (notes or "") + " " + (parent_response or "")
    matches = list(_EXPECTED_MONTH_RE.finditer(text))
    if not matches:
        return "未知"
    m = matches[-1]
    month_num = int(m.group(2))
    if not (1 <= month_num <= 12):
        return "未知"
    label = f"{m.group(1)}.{month_num:02d}"
    after = text[m.end() : m.end() + 30]
    gm = _GRADE_RE.search(after)
    if gm:
        return f"{label} 讀{gm.group(1)}"
    return f"{label} 讀{grade}" if grade else label


def _extract_expected_label(r: RecruitmentVisit) -> str:
    return _extract_expected_label_from_text(r.notes, r.parent_response, r.grade)


def _clean_source_label(source: Optional[str]) -> str:
    text = re.sub(r"\s+", " ", (source or "").strip())
    return text or "未填寫"


def _normalize_source_label(source: Optional[str]) -> str:
    label = _clean_source_label(source)
    if label.startswith(_CHUANNIAN_KW):
        return _CHUANNIAN_KW
    if label == _BRANCH_INTRO_KW or label in _SOURCE_GROUP_ALIASES[_BRANCH_INTRO_KW]:
        return _BRANCH_INTRO_KW
    return label


def _branch_intro_sql_cond():
    return RecruitmentVisit.source.in_(
        tuple(sorted(_SOURCE_GROUP_ALIASES[_BRANCH_INTRO_KW]))
    )


def _source_group_sql_expr():
    return case(
        (_chuannian_sql_cond(), _CHUANNIAN_KW),
        (_branch_intro_sql_cond(), _BRANCH_INTRO_KW),
        else_=func.coalesce(RecruitmentVisit.source, "未填寫"),
    )


def _normalize_dataset_scope(dataset_scope: Optional[str]) -> str:
    return DATASET_SCOPE_ALL


def _dataset_scope_filters(dataset_scope: Optional[str]) -> list:
    return []


def _build_source_filter_condition(source: Optional[str]):
    label = _clean_source_label(source)
    if label == _CHUANNIAN_KW:
        return _chuannian_sql_cond()
    if label == _BRANCH_INTRO_KW:
        return _branch_intro_sql_cond()
    return RecruitmentVisit.source == label


def _parse_period_range(period_name: str) -> Optional[tuple]:
    """解析期間名稱，回傳 (start_ym, end_ym)，如 ('114.09', '115.03')。"""
    m = _PERIOD_RANGE_RE.search(period_name.strip())
    if not m:
        return None
    return m.group(1), m.group(2)


def _normalize_roc_month(value: Optional[str]) -> Optional[str]:
    """正規化民國月份字串為 YYY.MM。"""
    if value is None:
        return None

    text = value.strip()
    if not text:
        return None

    parts = text.split(".")
    if len(parts) != 2:
        raise ValueError("月份格式應為 民國年.月，如 115.03")

    try:
        year_num = int(parts[0])
        month_num = int(parts[1])
    except ValueError as exc:
        raise ValueError("月份格式錯誤") from exc

    if year_num <= 0:
        raise ValueError(f"年份須為正整數，收到 {parts[0]}")
    if not (1 <= month_num <= 12):
        raise ValueError(f"月份須在 1-12 之間，收到 {month_num}")

    return f"{year_num}.{month_num:02d}"


def _extract_roc_month_from_visit_date(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None

    text = value.strip()
    if not text:
        return None

    match = _VISIT_DATE_MONTH_RE.search(text)
    if not match:
        return None

    year_num = int(match.group(1))
    month_num = int(match.group(2))
    if year_num <= 0 or not (1 <= month_num <= 12):
        return None
    return f"{year_num}.{month_num:02d}"


def _safe_normalize_roc_month(value: Optional[str]) -> Optional[str]:
    """盡量正規化月份，若既有資料異常則保留原值。"""
    if value is None:
        return None
    try:
        return _normalize_roc_month(value)
    except ValueError:
        stripped = value.strip()
        return stripped or None


def _roc_month_sort_key(value: Optional[str]) -> tuple:
    normalized = _safe_normalize_roc_month(value)
    if normalized in (None, "", "未知"):
        return (999999, 99, normalized or "")

    parts = normalized.split(".")
    if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
        return (999998, 99, normalized)

    return (int(parts[0]), int(parts[1]), normalized)


def _parse_roc_month_parts(value: Optional[str]) -> Optional[tuple[int, int]]:
    normalized = _normalize_roc_month(value)
    year_text, month_text = normalized.split(".")
    return int(year_text), int(month_text)


def _shift_roc_month(value: Optional[str], delta_months: int) -> Optional[str]:
    if value in (None, ""):
        return None

    year_num, month_num = _parse_roc_month_parts(value)
    total_months = year_num * 12 + (month_num - 1) + delta_months
    if total_months < 0:
        return None

    shifted_year = total_months // 12
    shifted_month = total_months % 12 + 1
    return f"{shifted_year}.{shifted_month:02d}"


def _roc_month_start(value: Optional[str]) -> Optional[datetime]:
    if value in (None, ""):
        return None
    year_num, month_num = _parse_roc_month_parts(value)
    return datetime(year_num + 1911, month_num, 1)


def _metric_snapshot(
    visit: int = 0,
    deposit: int = 0,
    enrolled: int = 0,
    transfer_term: int = 0,
    pending_deposit: int = 0,
    effective_deposit: int = 0,
) -> dict:
    def _pct_value(num: int, den: int) -> float:
        return round(num / den * 100, 1) if den else 0

    return {
        "visit": visit,
        "deposit": deposit,
        "enrolled": enrolled,
        "transfer_term": transfer_term,
        "pending_deposit": pending_deposit,
        "effective_deposit": effective_deposit,
        "visit_to_deposit_rate": _pct_value(deposit, visit),
        "visit_to_enrolled_rate": _pct_value(enrolled, visit),
        "deposit_to_enrolled_rate": _pct_value(enrolled, deposit),
        "effective_to_enrolled_rate": _pct_value(enrolled, effective_deposit),
    }


def _empty_snapshot() -> dict:
    return _metric_snapshot()


def _aggregate_snapshot(session, *filters) -> dict:
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    enrolled_case = case((RecruitmentVisit.enrolled == True, 1), else_=0)
    transfer_case = case((RecruitmentVisit.transfer_term == True, 1), else_=0)
    pending_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.enrolled == False,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )
    effective_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )

    query = session.query(
        func.count(RecruitmentVisit.id),
        func.sum(dep_case),
        func.sum(enrolled_case),
        func.sum(transfer_case),
        func.sum(pending_dep_case),
        func.sum(effective_dep_case),
    )
    applied_filters = [item for item in filters if item is not None]
    if applied_filters:
        query = query.filter(*applied_filters)

    row = query.one()
    return _metric_snapshot(
        visit=row[0] or 0,
        deposit=row[1] or 0,
        enrolled=row[2] or 0,
        transfer_term=row[3] or 0,
        pending_deposit=row[4] or 0,
        effective_deposit=row[5] or 0,
    )


def _select_reference_month(
    monthly: list[dict], requested: Optional[str]
) -> Optional[str]:
    if requested:
        return _normalize_roc_month(requested)
    if not monthly:
        return None
    return _safe_normalize_roc_month(monthly[-1]["month"])


def _build_month_over_month(
    current_month: Optional[str],
    previous_month: Optional[str],
    monthly_map: dict[str, dict],
) -> dict:
    current_row = (
        monthly_map.get(current_month, _empty_snapshot())
        if current_month
        else _empty_snapshot()
    )
    previous_row = (
        monthly_map.get(previous_month, _empty_snapshot())
        if previous_month
        else _empty_snapshot()
    )
    tracked_fields = [
        "visit",
        "deposit",
        "enrolled",
        "effective_deposit",
        "pending_deposit",
        "visit_to_deposit_rate",
        "visit_to_enrolled_rate",
        "deposit_to_enrolled_rate",
        "effective_to_enrolled_rate",
    ]
    diff = {
        field: {
            "current": current_row.get(field, 0),
            "previous": previous_row.get(field, 0),
            "delta": round(
                (current_row.get(field, 0) or 0) - (previous_row.get(field, 0) or 0), 1
            ),
        }
        for field in tracked_fields
    }
    return {
        "current_month": current_month,
        "previous_month": previous_month,
        **diff,
    }


def _build_ytd_snapshot(
    reference_month: Optional[str], monthly_map: dict[str, dict]
) -> dict:
    if not reference_month:
        return _empty_snapshot()

    ref_year, ref_month = _parse_roc_month_parts(reference_month)
    totals = _empty_snapshot()
    for label, snapshot in monthly_map.items():
        try:
            year_num, month_num = _parse_roc_month_parts(label)
        except ValueError:
            continue
        if year_num != ref_year or month_num > ref_month:
            continue
        for field in (
            "visit",
            "deposit",
            "enrolled",
            "transfer_term",
            "pending_deposit",
            "effective_deposit",
        ):
            totals[field] += snapshot.get(field, 0) or 0

    return _metric_snapshot(
        visit=totals["visit"],
        deposit=totals["deposit"],
        enrolled=totals["enrolled"],
        transfer_term=totals["transfer_term"],
        pending_deposit=totals["pending_deposit"],
        effective_deposit=totals["effective_deposit"],
    )


def _build_alerts(
    month_over_month: dict,
    high_potential_backlog_count: int,
    source_imbalance: Optional[dict],
    reference_month: Optional[str],
) -> list[dict]:
    alerts: list[dict] = []

    visit_to_deposit_delta = month_over_month.get("visit_to_deposit_rate", {}).get(
        "delta", 0
    )
    visit_to_enrolled_delta = month_over_month.get("visit_to_enrolled_rate", {}).get(
        "delta", 0
    )
    if (
        visit_to_deposit_delta <= -FUNNEL_DROP_THRESHOLD
        or visit_to_enrolled_delta <= -FUNNEL_DROP_THRESHOLD
    ):
        alerts.append(
            {
                "code": "FUNNEL_DROP",
                "level": "warning",
                "title": "本月漏斗轉換下滑",
                "message": (
                    f"{reference_month or '當期'} 參觀轉預繳 {visit_to_deposit_delta:.1f} 個百分點，"
                    f"參觀轉註冊 {visit_to_enrolled_delta:.1f} 個百分點。"
                ),
                "target_tab": "detail",
                "target_filter": {"month": reference_month},
            }
        )

    if high_potential_backlog_count >= HIGH_POTENTIAL_BACKLOG_THRESHOLD:
        alerts.append(
            {
                "code": "HIGH_POTENTIAL_BACKLOG",
                "level": "danger",
                "title": "高潛力未預繳名單堆積",
                "message": f"超過 {DEFAULT_OVERDUE_DAYS} 天仍未預繳的高潛力名單有 {high_potential_backlog_count} 筆。",
                "target_tab": "nodeposit",
                "target_filter": {
                    "priority": "high",
                    "overdue_days": DEFAULT_OVERDUE_DAYS,
                },
            }
        )

    if source_imbalance:
        alerts.append(
            {
                "code": "SOURCE_IMBALANCE",
                "level": "info",
                "title": "來源結構失衡",
                "message": (
                    f"{source_imbalance['source']} 近 90 天占比 {source_imbalance['share']:.1f}% ，"
                    f"預繳率 {source_imbalance['deposit_rate']:.1f}% 低於整體 {source_imbalance['overall_rate']:.1f}%。"
                ),
                "target_tab": "area",
                "target_filter": {"source": source_imbalance["source"]},
            }
        )

    return alerts


def _build_action_queue(
    current_month: Optional[str],
    high_potential_backlog_count: int,
    dominant_district: Optional[str],
    source_imbalance: Optional[dict],
) -> list[dict]:
    actions: list[dict] = []

    if high_potential_backlog_count:
        actions.append(
            {
                "code": "FOLLOW_HIGH_POTENTIAL",
                "title": "查看高風險未預繳",
                "description": f"目前有 {high_potential_backlog_count} 筆高潛力名單逾期未追。",
                "target_tab": "nodeposit",
                "target_filter": {
                    "priority": "high",
                    "overdue_days": DEFAULT_OVERDUE_DAYS,
                },
            }
        )

    if current_month:
        actions.append(
            {
                "code": "REVIEW_CURRENT_MONTH",
                "title": "查看本月明細",
                "description": f"切換到 {current_month} 明細，檢查本月漏斗掉點。",
                "target_tab": "detail",
                "target_filter": {"month": current_month},
            }
        )

    if dominant_district or source_imbalance:
        target_filter = {}
        if dominant_district:
            target_filter["district"] = dominant_district
        if source_imbalance:
            target_filter["source"] = source_imbalance["source"]
        actions.append(
            {
                "code": "AREA_OPPORTUNITY",
                "title": "查看區域機會",
                "description": f"優先檢查 {dominant_district or '重點行政區'} 的來源分布與通勤熱區。",
                "target_tab": "area",
                "target_filter": target_filter,
            }
        )

    return actions[:ACTION_QUEUE_LIMIT]


def _find_source_imbalance(session, window_start: datetime, *filters) -> Optional[dict]:
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    source_group_expr = _source_group_sql_expr().label("source")
    rows = (
        session.query(
            source_group_expr,
            func.count(RecruitmentVisit.id).label("visit"),
            func.sum(dep_case).label("deposit"),
        )
        .filter(*filters, RecruitmentVisit.created_at >= window_start)
        .group_by(source_group_expr)
        .all()
    )
    normalized_rows = [
        {"source": row.source, "visit": row.visit or 0, "deposit": row.deposit or 0}
        for row in rows
    ]

    total_visit = sum((row["visit"] or 0) for row in normalized_rows)
    total_deposit = sum((row["deposit"] or 0) for row in normalized_rows)
    overall_rate = round(total_deposit / total_visit * 100, 1) if total_visit else 0
    candidate: Optional[dict] = None
    for row in normalized_rows:
        visit = row["visit"] or 0
        if not visit or not total_visit:
            continue
        share = round(visit / total_visit * 100, 1)
        deposit = row["deposit"] or 0
        deposit_rate = round(deposit / visit * 100, 1) if visit else 0
        if share >= SOURCE_IMBALANCE_SHARE_THRESHOLD and deposit_rate < overall_rate:
            current = {
                "source": row["source"],
                "visit": visit,
                "deposit": deposit,
                "share": share,
                "deposit_rate": deposit_rate,
                "overall_rate": overall_rate,
            }
            if candidate is None or current["share"] > candidate["share"]:
                candidate = current
    return candidate


def _extract_district_from_address(address: Optional[str]) -> Optional[str]:
    """從完整地址盡量提取行政區，例如「高雄市三民區民族一路...」→「三民區」"""
    if not address:
        return None

    text = address.strip()
    if not text:
        return None

    match = re.search(r"[縣市]([一-龥]{1,4}區)", text)
    if match:
        return match.group(1)

    fallback = re.search(r"([一-龥]{1,4}區)", text)
    return fallback.group(1) if fallback else None


def _expand_roc_month_range(start_ym: str, end_ym: str) -> set[str]:
    start_normalized = _normalize_roc_month(start_ym)
    end_normalized = _normalize_roc_month(end_ym)
    start_year, start_month = map(int, start_normalized.split("."))
    end_year, end_month = map(int, end_normalized.split("."))

    cursor_year, cursor_month = start_year, start_month
    labels: set[str] = set()
    while (cursor_year, cursor_month) <= (end_year, end_month):
        labels.add(f"{cursor_year}.{cursor_month:02d}")
        labels.add(f"{cursor_year}.{cursor_month}")
        if cursor_month == 12:
            cursor_year += 1
            cursor_month = 1
        else:
            cursor_month += 1
    return labels


def _build_period_month_labels(periods: list[RecruitmentPeriod]) -> list[str]:
    labels: set[str] = set()
    for period in periods:
        period_range = _parse_period_range(period.period_name)
        if not period_range:
            logger.warning("略過無法解析的招生期間：%s", period.period_name)
            continue
        labels |= _expand_roc_month_range(*period_range)
    return sorted(labels, key=_roc_month_sort_key)


def _normalize_hotspot_sync_mode(sync_mode: str) -> str:
    normalized = (sync_mode or "").strip().lower() or "incremental"
    if normalized not in ADDRESS_SYNC_MODES:
        raise HTTPException(
            status_code=400, detail="sync_mode 僅支援 incremental 或 resync_google"
        )
    return normalized


def _is_google_stale_cache(cached: Optional[RecruitmentGeocodeCache]) -> bool:
    return market_service.is_google_stale_cache_row(cached)


def _needs_incremental_sync(cached: Optional[RecruitmentGeocodeCache]) -> bool:
    if not cached:
        return True
    if cached.status in {"pending", "failed"}:
        return True
    return cached.status == "resolved" and (cached.lat is None or cached.lng is None)


def _load_hotspot_cache_rows(
    session, addresses: list[str]
) -> dict[str, RecruitmentGeocodeCache]:
    cache_rows: dict[str, RecruitmentGeocodeCache] = {}
    deduped_addresses = list(dict.fromkeys(addresses))
    if not deduped_addresses:
        return cache_rows

    for index in range(0, len(deduped_addresses), HOTSPOT_CACHE_LOOKUP_CHUNK_SIZE):
        chunk = deduped_addresses[index : index + HOTSPOT_CACHE_LOOKUP_CHUNK_SIZE]
        for row in (
            session.query(RecruitmentGeocodeCache)
            .filter(RecruitmentGeocodeCache.address.in_(chunk))
            .all()
        ):
            cache_rows[row.address] = row
    return cache_rows


def _build_scoped_query(session, dataset_scope: Optional[str] = None):
    q = session.query(RecruitmentVisit)
    scope_filters = _dataset_scope_filters(dataset_scope)
    if scope_filters:
        q = q.filter(*scope_filters)
    return q


# ---------------------------------------------------------------------------
# Pydantic Schemas
# ---------------------------------------------------------------------------


class RecruitmentVisitCreate(BaseModel):
    month: str = Field(..., min_length=1, max_length=10)
    seq_no: Optional[str] = Field(None, max_length=10)
    visit_date: Optional[str] = Field(None, max_length=50)
    child_name: str = Field(..., min_length=1, max_length=50)
    birthday: Optional[date] = None
    grade: Optional[str] = Field(None, max_length=20)
    phone: Optional[str] = Field(None, max_length=100)
    address: Optional[str] = Field(None, max_length=200)
    district: Optional[str] = Field(None, max_length=30)
    source: Optional[str] = Field(None, max_length=50)
    referrer: Optional[str] = Field(None, max_length=50)
    deposit_collector: Optional[str] = Field(None, max_length=50)
    has_deposit: bool = False
    notes: Optional[str] = None
    parent_response: Optional[str] = None
    no_deposit_reason: Optional[str] = Field(None, max_length=60)
    no_deposit_reason_detail: Optional[str] = None
    enrolled: bool = False
    transfer_term: bool = False

    @field_validator("month")
    @classmethod
    def validate_month_format(cls, v: str) -> str:
        return _normalize_roc_month(v)


class RecruitmentVisitUpdate(BaseModel):
    month: Optional[str] = Field(None, min_length=1, max_length=10)
    seq_no: Optional[str] = Field(None, max_length=10)
    visit_date: Optional[str] = Field(None, max_length=50)
    child_name: Optional[str] = Field(None, min_length=1, max_length=50)
    birthday: Optional[date] = None
    grade: Optional[str] = Field(None, max_length=20)
    phone: Optional[str] = Field(None, max_length=100)
    address: Optional[str] = Field(None, max_length=200)
    district: Optional[str] = Field(None, max_length=30)
    source: Optional[str] = Field(None, max_length=50)
    referrer: Optional[str] = Field(None, max_length=50)
    deposit_collector: Optional[str] = Field(None, max_length=50)
    has_deposit: Optional[bool] = None
    notes: Optional[str] = None
    parent_response: Optional[str] = None
    no_deposit_reason: Optional[str] = Field(None, max_length=60)
    no_deposit_reason_detail: Optional[str] = None
    enrolled: Optional[bool] = None
    transfer_term: Optional[bool] = None

    @field_validator("month")
    @classmethod
    def validate_month_format(cls, v: Optional[str]) -> Optional[str]:
        if v is None:
            return v
        return _normalize_roc_month(v)


class ImportRecord(BaseModel):
    月份: Optional[str] = None
    序號: Optional[str] = None
    日期: Optional[str] = None
    幼生姓名: Optional[str] = None
    生日: Optional[str] = None
    適讀班級: Optional[str] = None
    電話: Optional[str] = None
    地址: Optional[str] = None
    行政區: Optional[str] = None
    幼生來源: Optional[str] = None
    介紹者: Optional[str] = None
    收預繳人員: Optional[str] = None
    是否預繳: Optional[str] = None
    備註: Optional[str] = None
    電訪後家長回應: Optional[str] = None


class PeriodCreate(BaseModel):
    period_name: str = Field(..., min_length=1, max_length=50)
    visit_count: int = Field(0, ge=0)
    deposit_count: int = Field(0, ge=0)
    enrolled_count: int = Field(0, ge=0)
    transfer_term_count: int = Field(0, ge=0)
    effective_deposit_count: int = Field(0, ge=0)
    not_enrolled_deposit: int = Field(0, ge=0)
    enrolled_after_school: int = Field(0, ge=0)
    notes: Optional[str] = None
    sort_order: int = 0


class PeriodUpdate(BaseModel):
    period_name: Optional[str] = Field(None, min_length=1, max_length=50)
    visit_count: Optional[int] = Field(None, ge=0)
    deposit_count: Optional[int] = Field(None, ge=0)
    enrolled_count: Optional[int] = Field(None, ge=0)
    transfer_term_count: Optional[int] = Field(None, ge=0)
    effective_deposit_count: Optional[int] = Field(None, ge=0)
    not_enrolled_deposit: Optional[int] = Field(None, ge=0)
    enrolled_after_school: Optional[int] = Field(None, ge=0)
    notes: Optional[str] = None
    sort_order: Optional[int] = None


class CampusSettingPayload(BaseModel):
    campus_name: str = Field(..., min_length=1, max_length=100)
    campus_address: str = Field("", max_length=255)
    campus_lat: Optional[float] = Field(None, ge=-90, le=90)
    campus_lng: Optional[float] = Field(None, ge=-180, le=180)
    travel_mode: str = Field("driving", pattern="^(driving|walking|cycling)$")


def _parse_roc_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        parts = s.strip().split(".")
        if len(parts) == 3:
            year = int(parts[0]) + 1911
            return date(year, int(parts[1]), int(parts[2]))
    except (ValueError, AttributeError):
        pass
    return None


# ---------------------------------------------------------------------------
# 基本 CRUD
# ---------------------------------------------------------------------------


@router.get("/records")
def list_recruitment_records(
    month: Optional[str] = Query(None),
    grade: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    referrer: Optional[str] = Query(None),
    has_deposit: Optional[bool] = Query(None),
    no_deposit_reason: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    with session_scope() as session:
        q = _build_scoped_query(session, dataset_scope)
        if month:
            q = q.filter(RecruitmentVisit.month == month)
        if grade:
            q = q.filter(RecruitmentVisit.grade == grade)
        if source:
            q = q.filter(_build_source_filter_condition(source))
        if referrer:
            q = q.filter(RecruitmentVisit.referrer == referrer)
        if has_deposit is not None:
            q = q.filter(RecruitmentVisit.has_deposit == has_deposit)
        if no_deposit_reason:
            q = q.filter(RecruitmentVisit.no_deposit_reason == no_deposit_reason)
        if keyword:
            kw = f"%{keyword}%"
            q = q.filter(
                RecruitmentVisit.child_name.ilike(kw)
                | RecruitmentVisit.address.ilike(kw)
                | RecruitmentVisit.notes.ilike(kw)
                | RecruitmentVisit.parent_response.ilike(kw)
            )
        total = q.count()
        records = (
            q.order_by(
                RecruitmentVisit.visit_date.desc().nulls_last(),
                RecruitmentVisit.month.desc(),
                RecruitmentVisit.seq_no,
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "records": [_to_dict(r) for r in records],
        }


def normalize_existing_months() -> int:
    """將 DB 中月份格式從 115.3 統一正規化為 115.03（幂等，啟動時呼叫）"""
    with session_scope() as session:
        updated = 0
        for r in (
            session.query(RecruitmentVisit)
            .filter(RecruitmentVisit.month.isnot(None), RecruitmentVisit.month != "")
            .all()
        ):
            normalized = _extract_roc_month_from_visit_date(
                r.visit_date
            ) or _safe_normalize_roc_month(r.month)
            if normalized and normalized != r.month:
                r.month = normalized
                updated += 1
        for r in (
            session.query(RecruitmentMonth)
            .filter(RecruitmentMonth.month.isnot(None), RecruitmentMonth.month != "")
            .all()
        ):
            normalized = _safe_normalize_roc_month(r.month)
            if normalized and normalized != r.month:
                r.month = normalized
                updated += 1
    if updated:
        logger.info("normalize_existing_months: 正規化 %d 筆月份格式", updated)
    return updated


def _auto_sync_periods_for_months(session, months: set) -> None:
    """CRUD 後自動重算受影響月份所在期間的統計數字。"""
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    for p in session.query(RecruitmentPeriod).all():
        period_range = _parse_period_range(p.period_name)
        if not period_range:
            continue
        period_labels = _expand_roc_month_range(*period_range)
        if not (months & period_labels):
            continue
        row = (
            session.query(
                func.count(RecruitmentVisit.id).label("visit_count"),
                func.sum(dep_case).label("deposit_count"),
                func.sum(case((RecruitmentVisit.enrolled == True, 1), else_=0)).label(
                    "enrolled_count"
                ),
                func.sum(
                    case((RecruitmentVisit.transfer_term == True, 1), else_=0)
                ).label("transfer_term_count"),
            )
            .filter(RecruitmentVisit.month.in_(period_labels))
            .one()
        )
        p.visit_count = row.visit_count or 0
        p.deposit_count = row.deposit_count or 0
        p.enrolled_count = row.enrolled_count or 0
        p.transfer_term_count = row.transfer_term_count or 0
        p.effective_deposit_count = max(
            (row.deposit_count or 0) - (row.transfer_term_count or 0), 0
        )
        p.updated_at = datetime.now()
        logger.info("自動同步期間 [%s] 完成", p.period_name)


@router.post("/records", status_code=201)
def create_recruitment_record(
    payload: RecruitmentVisitCreate,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        record = RecruitmentVisit(**payload.model_dump())
        record.expected_start_label = _extract_expected_label_from_text(
            record.notes, record.parent_response, record.grade
        )
        session.add(record)
        session.flush()
        _auto_sync_periods_for_months(session, {record.month})
        return _to_dict(record)


@router.put("/records/{record_id}")
def update_recruitment_record(
    record_id: int,
    payload: RecruitmentVisitUpdate,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        record = session.query(RecruitmentVisit).get(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="紀錄不存在")
        old_month = record.month
        for field, value in payload.model_dump(exclude_unset=True).items():
            setattr(record, field, value)
        record.expected_start_label = _extract_expected_label_from_text(
            record.notes, record.parent_response, record.grade
        )
        record.updated_at = datetime.now()
        session.flush()
        _auto_sync_periods_for_months(session, {old_month, record.month})
        return _to_dict(record)


@router.delete("/records/{record_id}", status_code=204)
def delete_recruitment_record(
    record_id: int,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        record = session.query(RecruitmentVisit).get(record_id)
        if not record:
            raise HTTPException(status_code=404, detail="紀錄不存在")
        month = record.month
        session.delete(record)
        session.flush()
        _auto_sync_periods_for_months(session, {month})


@router.get("/campus-setting")
def get_recruitment_campus_setting(
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    with session_scope() as session:
        setting = market_service.get_or_create_campus_setting(session)
        return market_service.serialize_campus_setting(setting)


@router.put("/campus-setting")
def update_recruitment_campus_setting(
    payload: CampusSettingPayload,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        setting = market_service.upsert_campus_setting(session, payload.model_dump())
        return market_service.serialize_campus_setting(setting)


@router.get("/nearby-kindergartens")
def get_nearby_kindergartens(
    south: float = Query(None, ge=-90, le=90, description="視野南界緯度"),
    west: float = Query(None, ge=-180, le=180, description="視野西界經度"),
    north: float = Query(None, ge=-90, le=90, description="視野北界緯度"),
    east: float = Query(None, ge=-180, le=180, description="視野東界經度"),
    zoom: int = Query(None, ge=1, le=22, description="地圖縮放等級"),
    radius_km: float = Query(
        None, ge=0.5, le=50.0, description="以本園為圓心的查詢半徑（向下相容）"
    ),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    bounds = None
    if all(v is not None for v in (south, west, north, east)):
        bounds = {
            "south": south,
            "west": west,
            "north": north,
            "east": east,
            "zoom": zoom,
        }
    with session_scope() as session:
        return market_service.search_nearby_kindergartens(
            session,
            radius_km=radius_km or 10.0,
            bounds=bounds,
        )


@router.post("/competitor-schools/geocode")
def geocode_competitor_schools(
    limit: int = Query(100, ge=1, le=500, description="本次最多 geocode 的學校筆數"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    """批量 geocode 尚無座標的 competitor_school 記錄，結果存回 DB。
    建議在教育部資料同步完成後執行一次，後續 nearby-kindergartens 即可使用 MOE gap-fill。
    """
    with session_scope() as session:
        return market_service.geocode_all_competitor_schools(session, limit=limit)


@router.get("/address-hotspots")
def get_recruitment_address_hotspots(
    limit: int = Query(200, ge=1, le=500),
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """依完整地址聚合的熱點資料，供區域分析簡易地圖使用。"""
    with session_scope() as session:
        return _build_address_hotspots_response(
            session, limit, dataset_scope=dataset_scope
        )


@router.post("/address-hotspots/sync")
def sync_recruitment_address_hotspots(
    batch_size: int = Query(10, ge=1, le=20),
    limit: int = Query(200, ge=1, le=500),
    sync_mode: str = "incremental",
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    """同步一小批地址座標到快取，避免每次前端渲染時重複 geocode。"""
    if not market_service.market_provider_available():
        raise HTTPException(
            status_code=400, detail="尚未設定 Google / TGOS / geocoding provider"
        )
    normalized_sync_mode = _normalize_hotspot_sync_mode(sync_mode)

    with session_scope() as session:
        hotspots, _records_with_address, _total_hotspots = _query_address_hotspots(
            session,
            dataset_scope=dataset_scope,
        )
        addresses = [hotspot["address"] for hotspot in hotspots]
        cached_rows = _load_hotspot_cache_rows(session, addresses)

        campus = market_service.serialize_campus_setting(
            market_service.get_or_create_campus_setting(session)
        )
        eligible_targets: list[tuple[dict, Optional[RecruitmentGeocodeCache]]] = []
        skipped = 0
        for hotspot in hotspots:
            cached = cached_rows.get(hotspot["address"])
            should_sync = (
                _needs_incremental_sync(cached)
                if normalized_sync_mode == "incremental"
                else _is_google_stale_cache(cached)
            )
            if should_sync:
                eligible_targets.append((hotspot, cached))
            else:
                skipped += 1

        sync_targets = eligible_targets[:batch_size]
        attempted = len(sync_targets)
        synced = 0
        failed = 0
        for hotspot, cached in sync_targets:
            result = market_service.resolve_address_metadata(
                hotspot["address"], campus=campus
            )
            if not cached:
                cached = RecruitmentGeocodeCache(address=hotspot["address"])
                session.add(cached)
                cached_rows[cached.address] = cached

            market_service._apply_metadata_to_geocode_cache(
                cached,
                result or {},
                district=hotspot["district"],
            )
            if cached.status == "resolved":
                synced += 1
            else:
                failed += 1

        session.flush()
        response = _build_address_hotspots_response(
            session, limit, dataset_scope=dataset_scope
        )
        response["sync_mode"] = normalized_sync_mode
        response["attempted"] = attempted
        response["synced"] = synced
        response["failed"] = failed
        response["skipped"] = skipped
        return response


@router.post("/market-intelligence/sync")
def sync_recruitment_market_intelligence(
    hotspot_limit: int = Query(200, ge=50, le=500),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        result = market_service.sync_market_intelligence(
            session, hotspot_limit=hotspot_limit
        )
        snapshot = market_service.build_market_intelligence_snapshot(session)
        return {
            **result,
            "snapshot": snapshot,
        }


@router.get("/market-intelligence")
def get_recruitment_market_intelligence(
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    with session_scope() as session:
        return market_service.build_market_intelligence_snapshot(
            session, dataset_scope=dataset_scope
        )


def _query_address_hotspots(
    session,
    limit: Optional[int] = None,
    dataset_scope: Optional[str] = None,
) -> tuple[list[dict], int, int]:
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    normalized_address = func.trim(RecruitmentVisit.address)
    rows_query = session.query(
        normalized_address.label("address"),
        RecruitmentVisit.district.label("district"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    scope_filters = _dataset_scope_filters(dataset_scope)
    if scope_filters:
        rows_query = rows_query.filter(*scope_filters)
    rows = (
        rows_query.filter(
            RecruitmentVisit.address.isnot(None),
            func.length(normalized_address) > 0,
        )
        .group_by(normalized_address, RecruitmentVisit.district)
        .all()
    )

    merged: dict[str, dict] = {}
    records_with_address = 0
    for row in rows:
        address = (row.address or "").strip()
        if not address:
            continue

        district = (
            (row.district or "").strip()
            or _extract_district_from_address(address)
            or "未填寫"
        )
        visit = row.visit or 0
        deposit = row.deposit or 0
        records_with_address += visit

        hotspot = merged.setdefault(
            address,
            {
                "address": address,
                "district": district,
                "visit": 0,
                "deposit": 0,
            },
        )
        hotspot["visit"] += visit
        hotspot["deposit"] += deposit
        if hotspot["district"] == "未填寫" and district != "未填寫":
            hotspot["district"] = district

    hotspots = sorted(
        merged.values(),
        key=lambda item: (-item["visit"], item["address"]),
    )
    if limit is not None:
        hotspots = hotspots[:limit]
    return hotspots, records_with_address, len(merged)


def _build_address_hotspots_response(
    session, limit: int, dataset_scope: Optional[str] = None
) -> dict:
    all_hotspots, records_with_address, total_hotspots = _query_address_hotspots(
        session,
        dataset_scope=dataset_scope,
    )
    hotspots = all_hotspots[:limit]
    cache_rows = _load_hotspot_cache_rows(
        session,
        [hotspot["address"] for hotspot in all_hotspots],
    )

    geocoded_hotspots = 0
    failed_hotspots = 0
    stale_hotspots = 0
    for hotspot in all_hotspots:
        cached = cache_rows.get(hotspot["address"])
        status = cached.status if cached else "pending"
        lat = cached.lat if cached and cached.status == "resolved" else None
        lng = cached.lng if cached and cached.status == "resolved" else None
        if lat is not None and lng is not None:
            geocoded_hotspots += 1
        elif status == "failed":
            failed_hotspots += 1
        if _is_google_stale_cache(cached):
            stale_hotspots += 1

    enriched_hotspots = []
    for hotspot in hotspots:
        cached = cache_rows.get(hotspot["address"])
        status = cached.status if cached else "pending"
        lat = cached.lat if cached and cached.status == "resolved" else None
        lng = cached.lng if cached and cached.status == "resolved" else None
        enriched_hotspots.append(
            {
                **hotspot,
                "lat": lat,
                "lng": lng,
                "geocode_status": status,
                "provider": cached.provider if cached else None,
                "formatted_address": cached.formatted_address if cached else None,
                "matched_address": cached.matched_address if cached else None,
                "google_place_id": cached.google_place_id if cached else None,
                "town_code": cached.town_code if cached else None,
                "town_name": cached.town_name if cached else None,
                "county_name": cached.county_name if cached else None,
                "land_use_label": cached.land_use_label if cached else None,
                "travel_minutes": cached.travel_minutes if cached else None,
                "travel_distance_km": cached.travel_distance_km if cached else None,
                "data_quality": cached.data_quality if cached else "partial",
            }
        )

    pending_hotspots = max(total_hotspots - geocoded_hotspots - failed_hotspots, 0)
    provider_name = market_service.current_market_provider()
    return {
        "records_with_address": records_with_address,
        "total_hotspots": total_hotspots,
        "geocoded_hotspots": geocoded_hotspots,
        "pending_hotspots": pending_hotspots,
        "remaining_hotspots": pending_hotspots,
        "failed_hotspots": failed_hotspots,
        "stale_hotspots": stale_hotspots,
        "provider_available": provider_name is not None,
        "provider_name": provider_name,
        "hotspots": enriched_hotspots,
    }


# ---------------------------------------------------------------------------
# 統計（SQL GROUP BY，避免全表 in-memory 聚合）
# ---------------------------------------------------------------------------


def _chuannian_sql_cond():
    """童年綠地 SQL 判定條件（與 Python _is_chuannian 邏輯一致）"""
    return or_(
        RecruitmentVisit.source.contains(_CHUANNIAN_KW),
        RecruitmentVisit.notes.contains(_CHUANNIAN_KW),
        RecruitmentVisit.parent_response.contains(_CHUANNIAN_KW),
        RecruitmentVisit.notes.contains(_YAOTING_KW),
        RecruitmentVisit.parent_response.contains(_YAOTING_KW),
    )


def _query_stats(
    session,
    reference_month: Optional[str] = None,
    dataset_scope: Optional[str] = None,
) -> dict:
    """執行招生統計所有 SQL 查詢，回傳統計字典（供 /stats 與 /stats/export 共用）。"""

    def _pct_value(num: int, den: int) -> float:
        return round(num / den * 100, 1) if den else 0

    base_filters = _dataset_scope_filters(dataset_scope)
    ch_cond = _chuannian_sql_cond()
    dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
    enrolled_case = case((RecruitmentVisit.enrolled == True, 1), else_=0)
    transfer_case = case((RecruitmentVisit.transfer_term == True, 1), else_=0)
    pending_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.enrolled == False,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )
    effective_dep_case = case(
        (
            and_(
                RecruitmentVisit.has_deposit == True,
                RecruitmentVisit.transfer_term == False,
            ),
            1,
        ),
        else_=0,
    )
    ch_case = case((ch_cond, 1), else_=0)
    ch_dep_case = case(
        (and_(ch_cond, RecruitmentVisit.has_deposit == True), 1), else_=0
    )

    # ── 1. 整體 KPI（單次查詢）──────────────────────────────────
    kpi_query = session.query(
        func.count(RecruitmentVisit.id),
        func.sum(dep_case),
        func.sum(enrolled_case),
        func.sum(transfer_case),
        func.sum(pending_dep_case),
        func.sum(effective_dep_case),
        func.sum(ch_case),
        func.sum(ch_dep_case),
    )
    if base_filters:
        kpi_query = kpi_query.filter(*base_filters)
    kpi = kpi_query.one()
    (
        total_visit,
        total_deposit,
        total_enrolled,
        total_transfer_term,
        total_pending_deposit,
        total_effective_deposit,
        chuannian_visit,
        chuannian_deposit,
    ) = (
        kpi[0] or 0,
        kpi[1] or 0,
        kpi[2] or 0,
        kpi[3] or 0,
        kpi[4] or 0,
        kpi[5] or 0,
        kpi[6] or 0,
        kpi[7] or 0,
    )

    # ── 2. 唯一幼生（child_name + birthday 組合去重，1 次查詢）──
    unique_key = (
        func.coalesce(RecruitmentVisit.child_name, "")
        + "|"
        + func.coalesce(cast(RecruitmentVisit.birthday, String), "")
    )
    dep_unique_key = case(
        (RecruitmentVisit.has_deposit == True, unique_key), else_=None
    )
    uq_query = session.query(
        func.count(func.distinct(unique_key)),
        func.count(func.distinct(dep_unique_key)),
    )
    if base_filters:
        uq_query = uq_query.filter(*base_filters)
    uq_row = uq_query.one()
    unique_visit = uq_row[0] or 0
    unique_deposit = uq_row[1] or 0

    # ── 3. 月度統計 ─────────────────────────────────────────────
    monthly_query = session.query(
        func.coalesce(RecruitmentVisit.month, "未知").label("month"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
        func.sum(enrolled_case).label("enrolled"),
        func.sum(transfer_case).label("transfer_term"),
        func.sum(pending_dep_case).label("pending_deposit"),
        func.sum(effective_dep_case).label("effective_deposit"),
        func.sum(ch_case).label("chuannian_visit"),
        func.sum(ch_dep_case).label("chuannian_deposit"),
    )
    if base_filters:
        monthly_query = monthly_query.filter(*base_filters)
    monthly_rows = monthly_query.group_by(RecruitmentVisit.month).all()

    monthly = sorted(
        [
            {
                "month": r.month,
                "visit": r.visit or 0,
                "deposit": r.deposit or 0,
                "enrolled": r.enrolled or 0,
                "transfer_term": r.transfer_term or 0,
                "pending_deposit": r.pending_deposit or 0,
                "effective_deposit": r.effective_deposit or 0,
                "visit_to_deposit_rate": _pct_value(r.deposit or 0, r.visit or 0),
                "visit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.visit or 0),
                "deposit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.deposit or 0),
                "effective_to_enrolled_rate": _pct_value(
                    r.enrolled or 0, r.effective_deposit or 0
                ),
                "chuannian_visit": r.chuannian_visit or 0,
                "chuannian_deposit": r.chuannian_deposit or 0,
            }
            for r in monthly_rows
        ],
        key=lambda item: _roc_month_sort_key(item["month"]),
    )

    # ── 3b. 年度統計（由月度聚合推回年度，避免 DB 方言差異）────────
    yearly_map: dict[str, dict] = {}
    for row in monthly:
        month_label = row["month"]
        if month_label in (None, "", "未知") or "." not in month_label:
            continue
        year = month_label.split(".", 1)[0]
        bucket = yearly_map.setdefault(
            year,
            {
                "year": year,
                "visit": 0,
                "deposit": 0,
                "enrolled": 0,
                "transfer_term": 0,
                "pending_deposit": 0,
                "effective_deposit": 0,
                "chuannian_visit": 0,
                "chuannian_deposit": 0,
            },
        )
        for key in (
            "visit",
            "deposit",
            "enrolled",
            "transfer_term",
            "pending_deposit",
            "effective_deposit",
            "chuannian_visit",
            "chuannian_deposit",
        ):
            bucket[key] += row[key]

    by_year = []
    for year in sorted(
        yearly_map.keys(), key=lambda value: int(value) if value.isdigit() else 999999
    ):
        bucket = yearly_map[year]
        by_year.append(
            {
                **bucket,
                "visit_to_deposit_rate": _pct_value(bucket["deposit"], bucket["visit"]),
                "visit_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["visit"]
                ),
                "deposit_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["deposit"]
                ),
                "effective_to_enrolled_rate": _pct_value(
                    bucket["enrolled"], bucket["effective_deposit"]
                ),
            }
        )

    # ── 4. 班別統計 ─────────────────────────────────────────────
    grade_query = session.query(
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
        func.sum(enrolled_case).label("enrolled"),
    )
    if base_filters:
        grade_query = grade_query.filter(*base_filters)
    grade_rows = (
        grade_query.group_by(RecruitmentVisit.grade)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_grade = [
        {
            "grade": r.grade,
            "visit": r.visit or 0,
            "deposit": r.deposit or 0,
            "enrolled": r.enrolled or 0,
            "visit_to_deposit_rate": _pct_value(r.deposit or 0, r.visit or 0),
            "visit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.visit or 0),
            "deposit_to_enrolled_rate": _pct_value(r.enrolled or 0, r.deposit or 0),
        }
        for r in grade_rows
    ]

    # ── 5. 月份 × 班別 ───────────────────────────────────────────
    mg_query = session.query(
        func.coalesce(RecruitmentVisit.month, "未知").label("month"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        mg_query = mg_query.filter(*base_filters)
    mg_rows = mg_query.group_by(RecruitmentVisit.month, RecruitmentVisit.grade).all()

    month_grade: dict = {}
    for r in mg_rows:
        m = r.month
        if m not in month_grade:
            month_grade[m] = {}
        month_grade[m][r.grade] = r.cnt
        month_grade[m]["合計"] = month_grade[m].get("合計", 0) + r.cnt

    # ── 6. 來源統計 ──────────────────────────────────────────────
    source_group_expr = _source_group_sql_expr().label("source")
    source_query = session.query(
        source_group_expr,
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        source_query = source_query.filter(*base_filters)
    source_rows = (
        source_query.group_by(source_group_expr)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_source = sorted(
        [
            {
                "source": row.source,
                "visit": row.visit or 0,
                "deposit": row.deposit or 0,
            }
            for row in source_rows
        ],
        key=lambda item: (-item["visit"], -item["deposit"], item["source"]),
    )

    # ── 7. 接待人員 × 各年級（GROUP BY referrer + grade）─────────
    ref_grade_query = session.query(
        func.coalesce(RecruitmentVisit.referrer, "未填寫").label("referrer"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ref_grade_query = ref_grade_query.filter(*base_filters)
    ref_grade_rows = ref_grade_query.group_by(
        RecruitmentVisit.referrer, RecruitmentVisit.grade
    ).all()

    by_referrer: dict = {}
    for r in ref_grade_rows:
        ref = r.referrer
        if ref not in by_referrer:
            by_referrer[ref] = {
                "referrer": ref,
                "visit": 0,
                "deposit": 0,
                "by_grade": {},
            }
        by_referrer[ref]["visit"] += r.visit or 0
        by_referrer[ref]["deposit"] += r.deposit or 0
        by_referrer[ref]["by_grade"][r.grade] = {
            "visit": r.visit or 0,
            "deposit": r.deposit or 0,
        }

    by_referrer_list = sorted(by_referrer.values(), key=lambda x: -x["visit"])

    # ── 8. 接待者 × 來源 交叉表 ──────────────────────────────────
    cross_source_expr = _source_group_sql_expr().label("source")
    cross_query = session.query(
        func.coalesce(RecruitmentVisit.referrer, "未填寫").label("referrer"),
        cross_source_expr,
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        cross_query = cross_query.filter(*base_filters)
    cross_qrows = cross_query.group_by(
        RecruitmentVisit.referrer, cross_source_expr
    ).all()

    _cross_raw: dict = {}
    for r in cross_qrows:
        if r.referrer not in _cross_raw:
            _cross_raw[r.referrer] = {}
        _cross_raw[r.referrer][r.source] = _cross_raw[r.referrer].get(r.source, 0) + (
            r.cnt or 0
        )

    top_source_names = [item["source"] for item in by_source[:TOP_SOURCES_COUNT]]

    cross_rows_out = sorted(
        [
            {
                "referrer": ref,
                "sources": {s: _cross_raw[ref].get(s, 0) for s in top_source_names},
                "total": sum(_cross_raw[ref].values()),
            }
            for ref in _cross_raw
        ],
        key=lambda x: -x["total"],
    )
    referrer_source_cross = {"referrers": cross_rows_out, "sources": top_source_names}

    # ── 9. 行政區統計 ────────────────────────────────────────────
    district_query = session.query(
        func.coalesce(RecruitmentVisit.district, "未填寫").label("district"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        district_query = district_query.filter(*base_filters)
    district_rows = (
        district_query.group_by(RecruitmentVisit.district)
        .order_by(func.count(RecruitmentVisit.id).desc())
        .all()
    )

    by_district = [
        {"district": r.district, "visit": r.visit or 0, "deposit": r.deposit or 0}
        for r in district_rows
    ]

    # ── 10. 未預繳原因（GROUP BY reason + grade）─────────────────
    reason_query = session.query(
        func.coalesce(RecruitmentVisit.no_deposit_reason, "未分類").label("reason"),
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("cnt"),
    )
    if base_filters:
        reason_query = reason_query.filter(*base_filters)
    reason_rows = (
        reason_query.filter(RecruitmentVisit.has_deposit == False)
        .group_by(RecruitmentVisit.no_deposit_reason, RecruitmentVisit.grade)
        .all()
    )

    no_deposit_total_query = session.query(func.count(RecruitmentVisit.id))
    if base_filters:
        no_deposit_total_query = no_deposit_total_query.filter(*base_filters)
    no_deposit_total = (
        no_deposit_total_query.filter(RecruitmentVisit.has_deposit == False).scalar()
        or 0
    )

    reason_stats: dict = {}
    for r in reason_rows:
        if r.reason not in reason_stats:
            reason_stats[r.reason] = {"reason": r.reason, "count": 0, "by_grade": {}}
        reason_stats[r.reason]["count"] += r.cnt
        reason_stats[r.reason]["by_grade"][r.grade] = r.cnt

    no_deposit_reasons = sorted(reason_stats.values(), key=lambda x: -x["count"])

    def _expected_sort_key(x: dict):
        label = x["expected_month"]
        return (1, "") if label == "未知" else (0, label)

    # ── 11. 童年綠地 by expected label（SQL GROUP BY expected_start_label）────
    ch_expected_query = session.query(
        func.coalesce(RecruitmentVisit.expected_start_label, "未知").label(
            "expected_month"
        ),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ch_expected_query = ch_expected_query.filter(*base_filters)
    ch_expected_rows = (
        ch_expected_query.filter(ch_cond)
        .group_by(RecruitmentVisit.expected_start_label)
        .all()
    )

    chuannian_by_expected_list = sorted(
        [
            {
                "expected_month": r.expected_month,
                "visit": r.visit or 0,
                "deposit": r.deposit or 0,
            }
            for r in ch_expected_rows
        ],
        key=_expected_sort_key,
    )

    # ── 12. 童年綠地各班別（SQL GROUP BY）───────────────────────
    ch_grade_query = session.query(
        func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
        func.count(RecruitmentVisit.id).label("visit"),
        func.sum(dep_case).label("deposit"),
    )
    if base_filters:
        ch_grade_query = ch_grade_query.filter(*base_filters)
    ch_grade_rows = (
        ch_grade_query.filter(ch_cond).group_by(RecruitmentVisit.grade).all()
    )

    chuannian_by_grade = sorted(
        [
            {"grade": r.grade, "visit": r.visit or 0, "deposit": r.deposit or 0}
            for r in ch_grade_rows
        ],
        key=lambda x: -x["visit"],
    )

    monthly_map = {
        _safe_normalize_roc_month(item["month"]) or item["month"]: item
        for item in monthly
    }
    resolved_reference_month = _select_reference_month(monthly, reference_month)
    previous_month = (
        _shift_roc_month(resolved_reference_month, -1)
        if resolved_reference_month
        else None
    )
    rolling_30d = _aggregate_snapshot(
        session,
        *base_filters,
        RecruitmentVisit.created_at >= datetime.now() - timedelta(days=30),
    )
    rolling_90d = _aggregate_snapshot(
        session,
        *base_filters,
        RecruitmentVisit.created_at >= datetime.now() - timedelta(days=90),
    )
    current_month_snapshot = (
        _metric_snapshot(
            visit=monthly_map.get(resolved_reference_month, {}).get("visit", 0),
            deposit=monthly_map.get(resolved_reference_month, {}).get("deposit", 0),
            enrolled=monthly_map.get(resolved_reference_month, {}).get("enrolled", 0),
            transfer_term=monthly_map.get(resolved_reference_month, {}).get(
                "transfer_term", 0
            ),
            pending_deposit=monthly_map.get(resolved_reference_month, {}).get(
                "pending_deposit", 0
            ),
            effective_deposit=monthly_map.get(resolved_reference_month, {}).get(
                "effective_deposit", 0
            ),
        )
        if resolved_reference_month
        else _empty_snapshot()
    )
    month_over_month = _build_month_over_month(
        resolved_reference_month, previous_month, monthly_map
    )

    overdue_cutoff = datetime.now() - timedelta(days=DEFAULT_OVERDUE_DAYS)
    high_potential_backlog_count = (
        session.query(func.count(RecruitmentVisit.id))
        .filter(
            *base_filters,
            RecruitmentVisit.has_deposit == False,
            RecruitmentVisit.no_deposit_reason.in_(
                tuple(HIGH_PRIORITY_NO_DEPOSIT_REASONS)
            ),
            RecruitmentVisit.created_at <= overdue_cutoff,
        )
        .scalar()
        or 0
    )
    source_imbalance = _find_source_imbalance(
        session, datetime.now() - timedelta(days=90), *base_filters
    )

    dominant_district_row = (
        session.query(
            func.coalesce(RecruitmentVisit.district, "未填寫").label("district"),
            func.count(RecruitmentVisit.id).label("visit"),
        )
        .filter(*base_filters, RecruitmentVisit.month == resolved_reference_month)
        .group_by(RecruitmentVisit.district)
        .order_by(
            func.count(RecruitmentVisit.id).desc(),
            func.coalesce(RecruitmentVisit.district, "未填寫"),
        )
        .first()
        if resolved_reference_month
        else None
    )
    dominant_district = (
        dominant_district_row.district if dominant_district_row else None
    )
    alerts = _build_alerts(
        month_over_month=month_over_month,
        high_potential_backlog_count=high_potential_backlog_count,
        source_imbalance=source_imbalance,
        reference_month=resolved_reference_month,
    )
    top_action_queue = _build_action_queue(
        current_month=resolved_reference_month,
        high_potential_backlog_count=high_potential_backlog_count,
        dominant_district=dominant_district,
        source_imbalance=source_imbalance,
    )

    return {
        "total_visit": total_visit,
        "total_deposit": total_deposit,
        "total_enrolled": total_enrolled,
        "total_transfer_term": total_transfer_term,
        "total_pending_deposit": total_pending_deposit,
        "total_effective_deposit": total_effective_deposit,
        "unique_visit": unique_visit,
        "unique_deposit": unique_deposit,
        "visit_to_deposit_rate": _pct_value(total_deposit, total_visit),
        "visit_to_enrolled_rate": _pct_value(total_enrolled, total_visit),
        "deposit_to_enrolled_rate": _pct_value(total_enrolled, total_deposit),
        "effective_to_enrolled_rate": _pct_value(
            total_enrolled, total_effective_deposit
        ),
        "chuannian_visit": chuannian_visit,
        "chuannian_deposit": chuannian_deposit,
        "monthly": monthly,
        "by_grade": by_grade,
        "month_grade": month_grade,
        "by_source": by_source,
        "by_referrer": by_referrer_list,
        "referrer_source_cross": referrer_source_cross,
        "top_source_names": top_source_names,
        "by_district": by_district,
        "no_deposit_reasons": no_deposit_reasons,
        "no_deposit_total": no_deposit_total,
        "chuannian_by_expected": chuannian_by_expected_list,
        "chuannian_by_grade": chuannian_by_grade,
        "by_year": by_year,
        "reference_month": resolved_reference_month,
        "decision_summary": {
            "current_month": current_month_snapshot,
            "rolling_30d": rolling_30d,
            "rolling_90d": rolling_90d,
            "ytd": _build_ytd_snapshot(resolved_reference_month, monthly_map),
        },
        "funnel_snapshot": {
            field: current_month_snapshot[field]
            for field in (
                "visit",
                "deposit",
                "enrolled",
                "transfer_term",
                "effective_deposit",
                "pending_deposit",
            )
        },
        "month_over_month": month_over_month,
        "alerts": alerts,
        "top_action_queue": top_action_queue,
    }


# ---------------------------------------------------------------------------
# 統計 API endpoints
# ---------------------------------------------------------------------------


@router.get("/stats")
def get_recruitment_stats(
    reference_month: Optional[str] = None,
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """完整統計匯總（全 SQL GROUP BY，效能最佳化版）"""
    with session_scope() as session:
        return _query_stats(
            session, reference_month=reference_month, dataset_scope=dataset_scope
        )


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=13)
_CENTER = Alignment(horizontal="center")


def _hrow(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.alignment = _HEADER_FONT, _HEADER_FILL, _CENTER


def _pct(num: int, den: int) -> str:
    return f"{num / den * 100:.1f}%" if den else "—"


@router.get("/stats/export")
def export_recruitment_stats(
    reference_month: Optional[str] = None,
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """匯出招生統計 Excel（多頁簽）"""
    with session_scope() as session:
        normalized_scope = _normalize_dataset_scope(dataset_scope)
        s = _query_stats(
            session, reference_month=reference_month, dataset_scope=normalized_scope
        )

    wb = Workbook()

    # ── Sheet 1：決策摘要 ─────────────────────────────────────────
    ws = wb.active
    ws.title = "決策摘要"
    ws.append(["招生決策摘要"])
    ws["A1"].font = _TITLE_FONT
    ws.append([f"參考月份：{s['reference_month'] or '無資料'}"])
    ws.append([])
    _hrow(ws, 4, ["區塊", "指標", "數值"])
    summary_rows = []
    summary_label_map = {
        "current_month": "本月",
        "rolling_30d": "近 30 天",
        "rolling_90d": "近 90 天",
        "ytd": "年度累計",
    }
    for key, label in summary_label_map.items():
        snapshot = s["decision_summary"][key]
        summary_rows.extend(
            [
                (label, "參觀", snapshot["visit"]),
                (label, "預繳", snapshot["deposit"]),
                (label, "註冊", snapshot["enrolled"]),
                (label, "參觀→預繳率", _pct(snapshot["deposit"], snapshot["visit"])),
                (label, "參觀→註冊率", _pct(snapshot["enrolled"], snapshot["visit"])),
            ]
        )
    for row in summary_rows:
        ws.append(list(row))

    ws.append([])
    month_over_month = s["month_over_month"]
    ws.append(["月比觀察", "本月", month_over_month["current_month"] or "—"])
    ws.append(["月比觀察", "上月", month_over_month["previous_month"] or "—"])
    ws.append(
        [
            "月比觀察",
            "參觀→預繳率變化",
            f"{month_over_month['visit_to_deposit_rate']['delta']:.1f} 個百分點",
        ]
    )
    ws.append(
        [
            "月比觀察",
            "參觀→註冊率變化",
            f"{month_over_month['visit_to_enrolled_rate']['delta']:.1f} 個百分點",
        ]
    )

    ws.append([])
    _hrow(ws, ws.max_row + 1, ["警示代碼", "等級", "標題", "說明"])
    if s["alerts"]:
        for alert in s["alerts"]:
            ws.append([alert["code"], alert["level"], alert["title"], alert["message"]])
    else:
        ws.append(["—", "info", "目前無警示", "本期未觸發決策警示。"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48

    # ── Sheet 2：總覽 KPI ─────────────────────────────────────────
    ws = wb.create_sheet("總覽")
    ws.append(["招生統計總覽"])
    ws["A1"].font = _TITLE_FONT
    ws.append([])
    _hrow(ws, 3, ["指標", "數值"])
    kpi_rows = [
        ("總參觀紀錄", s["total_visit"]),
        ("唯一幼生數", s["unique_visit"]),
        ("總預繳人數", s["total_deposit"]),
        ("總註冊人數", s["total_enrolled"]),
        ("轉其他學期", s["total_transfer_term"]),
        ("預繳未註冊", s["total_pending_deposit"]),
        ("有效預繳", s["total_effective_deposit"]),
        ("唯一幼生預繳數", s["unique_deposit"]),
        ("參觀→預繳率", _pct(s["total_deposit"], s["total_visit"])),
        ("參觀→註冊率", _pct(s["total_enrolled"], s["total_visit"])),
        ("預繳→註冊率", _pct(s["total_enrolled"], s["total_deposit"])),
        ("排除轉期→註冊率", _pct(s["total_enrolled"], s["total_effective_deposit"])),
        ("唯一幼生預繳率", _pct(s["unique_deposit"], s["unique_visit"])),
        ("童年綠地參觀人數", s["chuannian_visit"]),
        ("童年綠地預繳人數", s["chuannian_deposit"]),
        ("童年綠地預繳率", _pct(s["chuannian_deposit"], s["chuannian_visit"])),
    ]
    for row in kpi_rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # ── Sheet 3：月度明細 ─────────────────────────────────────────
    ws2 = wb.create_sheet("月度明細")
    _hrow(
        ws2,
        1,
        [
            "月份",
            "參觀人數",
            "預繳人數",
            "註冊人數",
            "轉其他學期",
            "有效預繳",
            "預繳未註冊",
            "參觀→預繳率",
            "參觀→註冊率",
            "預繳→註冊率",
            "排除轉期→註冊率",
            "童年綠地參觀",
            "童年綠地預繳",
            "童年綠地預繳率",
        ],
    )
    for r in s["monthly"]:
        ws2.append(
            [
                r["month"],
                r["visit"],
                r["deposit"],
                r["enrolled"],
                r["transfer_term"],
                r["effective_deposit"],
                r["pending_deposit"],
                _pct(r["deposit"], r["visit"]),
                _pct(r["enrolled"], r["visit"]),
                _pct(r["enrolled"], r["deposit"]),
                _pct(r["enrolled"], r["effective_deposit"]),
                r["chuannian_visit"],
                r["chuannian_deposit"],
                _pct(r["chuannian_deposit"], r["chuannian_visit"]),
            ]
        )
    for col_letter, width in zip(
        "ABCDEFGHIJKLMN", [10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 14, 12, 12, 14]
    ):
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 4：班別分析 ─────────────────────────────────────────
    ws3 = wb.create_sheet("班別分析")
    _hrow(ws3, 1, ["班別", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_grade"]:
        ws3.append(
            [r["grade"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    for col_letter, width in zip("ABCD", [10, 10, 10, 10]):
        ws3.column_dimensions[col_letter].width = width

    # ── Sheet 5：來源分析 ─────────────────────────────────────────
    ws4 = wb.create_sheet("來源分析")
    _hrow(ws4, 1, ["來源", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_source"]:
        ws4.append(
            [r["source"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws4.column_dimensions["A"].width = 20
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws4.column_dimensions[col_letter].width = width

    # ── Sheet 6：接待人員 ─────────────────────────────────────────
    ws5 = wb.create_sheet("接待人員")
    _hrow(ws5, 1, ["接待人員", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_referrer"]:
        ws5.append(
            [r["referrer"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws5.column_dimensions["A"].width = 16
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws5.column_dimensions[col_letter].width = width

    # ── Sheet 7：行政區 ───────────────────────────────────────────
    ws6 = wb.create_sheet("行政區")
    _hrow(ws6, 1, ["行政區", "參觀人數", "預繳人數", "預繳率"])
    for r in s["by_district"]:
        ws6.append(
            [r["district"], r["visit"], r["deposit"], _pct(r["deposit"], r["visit"])]
        )
    ws6.column_dimensions["A"].width = 14
    for col_letter, width in zip("BCD", [10, 10, 10]):
        ws6.column_dimensions[col_letter].width = width

    # ── Sheet 8：未預繳原因 ───────────────────────────────────────
    ws7 = wb.create_sheet("未預繳原因")
    _hrow(ws7, 1, ["原因", "人數"])
    for r in s["no_deposit_reasons"]:
        ws7.append([r["reason"], r["count"]])
    ws7.append(["（合計）", s["no_deposit_total"]])
    ws7.column_dimensions["A"].width = 28
    ws7.column_dimensions["B"].width = 10

    # ── Sheet 9：年度統計 ─────────────────────────────────────────
    ws8 = wb.create_sheet("年度統計")
    _hrow(
        ws8,
        1,
        [
            "年份",
            "參觀人數",
            "預繳人數",
            "註冊人數",
            "轉其他學期",
            "有效預繳",
            "預繳未註冊",
            "參觀→預繳率",
            "參觀→註冊率",
            "預繳→註冊率",
            "排除轉期→註冊率",
            "童年綠地參觀",
            "童年綠地預繳",
        ],
    )
    for r in s["by_year"]:
        ws8.append(
            [
                f"{r['year']}年",
                r["visit"],
                r["deposit"],
                r["enrolled"],
                r["transfer_term"],
                r["effective_deposit"],
                r["pending_deposit"],
                _pct(r["deposit"], r["visit"]),
                _pct(r["enrolled"], r["visit"]),
                _pct(r["enrolled"], r["deposit"]),
                _pct(r["enrolled"], r["effective_deposit"]),
                r["chuannian_visit"],
                r["chuannian_deposit"],
            ]
        )
    for col_letter, width in zip(
        "ABCDEFGHIJKLM", [10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 14, 12, 12]
    ):
        ws8.column_dimensions[col_letter].width = width

    filename = "招生統計.xlsx"
    return xlsx_streaming_response(wb, filename)


@router.get("/no-deposit-analysis")
def get_no_deposit_analysis(
    reason: Optional[str] = Query(None),
    grade: Optional[str] = Query(None),
    priority: Optional[str] = Query(None, pattern="^(high|medium|low)$"),
    overdue_days: Optional[int] = Query(None, ge=1, le=365),
    cold_only: Optional[bool] = Query(None),
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """未預繳名單明細（含原因分類篩選，支援分頁）"""
    with session_scope() as session:
        q = _build_scoped_query(session, dataset_scope).filter(
            RecruitmentVisit.has_deposit == False
        )
        if reason:
            q = q.filter(RecruitmentVisit.no_deposit_reason == reason)
        if grade:
            q = q.filter(RecruitmentVisit.grade == grade)
        base_query = q
        effective_overdue_days = overdue_days or DEFAULT_OVERDUE_DAYS
        overdue_cutoff = datetime.now() - timedelta(days=effective_overdue_days)
        cold_cutoff = datetime.now() - timedelta(days=COLD_LEAD_DAYS)
        summary = {
            "high_potential_count": (
                base_query.filter(
                    RecruitmentVisit.no_deposit_reason.in_(
                        tuple(HIGH_PRIORITY_NO_DEPOSIT_REASONS)
                    )
                ).count()
            ),
            "overdue_followup_count": base_query.filter(
                RecruitmentVisit.created_at <= overdue_cutoff
            ).count(),
            "cold_count": base_query.filter(
                RecruitmentVisit.created_at <= cold_cutoff
            ).count(),
        }
        if priority:
            q = q.filter(
                RecruitmentVisit.no_deposit_reason.in_(
                    tuple(NO_DEPOSIT_PRIORITY_REASON_MAP[priority])
                )
            )
        if overdue_days is not None:
            q = q.filter(RecruitmentVisit.created_at <= overdue_cutoff)
        if cold_only is True:
            q = q.filter(RecruitmentVisit.created_at <= cold_cutoff)
        total = q.count()
        records = (
            q.order_by(RecruitmentVisit.month.desc(), RecruitmentVisit.seq_no)
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "summary": summary,
            "records": [_to_dict(r) for r in records],
        }


# ---------------------------------------------------------------------------
# 近五年期間轉換整合
# ---------------------------------------------------------------------------


@router.get("/periods/summary")
def get_periods_summary(
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """近五年整體量體 KPI + 班別轉換分析"""
    with session_scope() as session:
        periods = (
            session.query(RecruitmentPeriod)
            .order_by(RecruitmentPeriod.sort_order)
            .all()
        )
        if not periods:
            return {
                "total_visit": 0,
                "total_deposit": 0,
                "total_enrolled": 0,
                "total_effective": 0,
                "total_transfer_term": 0,
                "total_not_enrolled_deposit": 0,
                "total_enrolled_after_school": 0,
                "total_net_enrolled": 0,
                "visit_to_deposit_rate": 0,
                "visit_to_enrolled_rate": 0,
                "deposit_to_enrolled_rate": 0,
                "effective_to_enrolled_rate": 0,
                "period_count": 0,
                "best_visit_to_enrolled": None,
                "worst_visit_to_enrolled": None,
                "best_deposit_to_enrolled": None,
                "worst_deposit_to_enrolled": None,
                "trend": [],
                "by_grade": [],
            }

        def _pct(num, den):
            return round(num / den * 100, 1) if den else 0

        normalized_scope = _normalize_dataset_scope(dataset_scope)
        dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)
        enrolled_case = case((RecruitmentVisit.enrolled == True, 1), else_=0)
        transfer_case = case((RecruitmentVisit.transfer_term == True, 1), else_=0)
        base_filters = _dataset_scope_filters(normalized_scope)

        if normalized_scope == DATASET_SCOPE_ALL:
            tv = sum(p.visit_count or 0 for p in periods)
            td = sum(p.deposit_count or 0 for p in periods)
            te = sum(p.enrolled_count or 0 for p in periods)
            teff = sum(p.effective_deposit_count or 0 for p in periods)
            ttr = sum(p.transfer_term_count or 0 for p in periods)

            trend = [
                {
                    "period_name": p.period_name,
                    "visit_count": p.visit_count or 0,
                    "deposit_count": p.deposit_count or 0,
                    "enrolled_count": p.enrolled_count or 0,
                    "not_enrolled_deposit": p.not_enrolled_deposit or 0,
                    "enrolled_after_school": p.enrolled_after_school or 0,
                    "net_enrolled_count": (p.enrolled_count or 0)
                    - (p.enrolled_after_school or 0),
                    "visit_to_deposit_rate": _pct(
                        p.deposit_count or 0, p.visit_count or 0
                    ),
                    "visit_to_enrolled_rate": _pct(
                        p.enrolled_count or 0, p.visit_count or 0
                    ),
                    "deposit_to_enrolled_rate": _pct(
                        p.enrolled_count or 0, p.deposit_count or 0
                    ),
                    "effective_to_enrolled_rate": _pct(
                        p.enrolled_count or 0, p.effective_deposit_count or 0
                    ),
                }
                for p in periods
            ]
        else:
            trend = []
            tv = td = te = teff = ttr = 0
            for period in periods:
                period_range = _parse_period_range(period.period_name)
                period_month_labels = (
                    _expand_roc_month_range(*period_range) if period_range else set()
                )
                row = None
                if period_month_labels:
                    query = session.query(
                        func.count(RecruitmentVisit.id).label("visit_count"),
                        func.sum(dep_case).label("deposit_count"),
                        func.sum(enrolled_case).label("enrolled_count"),
                        func.sum(transfer_case).label("transfer_term_count"),
                    )
                    if base_filters:
                        query = query.filter(*base_filters)
                    row = query.filter(
                        RecruitmentVisit.month.in_(period_month_labels)
                    ).one()

                visit_count = row.visit_count or 0 if row else 0
                deposit_count = row.deposit_count or 0 if row else 0
                enrolled_count = row.enrolled_count or 0 if row else 0
                transfer_term_count = row.transfer_term_count or 0 if row else 0
                effective_deposit_count = max(deposit_count - transfer_term_count, 0)

                tv += visit_count
                td += deposit_count
                te += enrolled_count
                teff += effective_deposit_count
                ttr += transfer_term_count
                trend.append(
                    {
                        "period_name": period.period_name,
                        "visit_count": visit_count,
                        "deposit_count": deposit_count,
                        "enrolled_count": enrolled_count,
                        "not_enrolled_deposit": 0,
                        "enrolled_after_school": 0,
                        "net_enrolled_count": enrolled_count,
                        "visit_to_deposit_rate": _pct(deposit_count, visit_count),
                        "visit_to_enrolled_rate": _pct(enrolled_count, visit_count),
                        "deposit_to_enrolled_rate": _pct(enrolled_count, deposit_count),
                        "effective_to_enrolled_rate": _pct(
                            enrolled_count, effective_deposit_count
                        ),
                    }
                )

        active = [d for d in trend if d["visit_count"] > 0]
        best_v2e = (
            max(active, key=lambda x: x["visit_to_enrolled_rate"]) if active else None
        )
        worst_v2e = (
            min(active, key=lambda x: x["visit_to_enrolled_rate"]) if active else None
        )
        best_d2e = (
            max(active, key=lambda x: x["deposit_to_enrolled_rate"]) if active else None
        )
        worst_d2e = (
            min(active, key=lambda x: x["deposit_to_enrolled_rate"]) if active else None
        )

        # 班別轉換（僅統計落在已定義期間內的 RecruitmentVisit）
        period_month_labels = _build_period_month_labels(periods)
        grade_rows = []
        if period_month_labels:
            grade_query = session.query(
                func.coalesce(RecruitmentVisit.grade, "未填寫").label("grade"),
                func.count(RecruitmentVisit.id).label("visit"),
                func.sum(dep_case).label("deposit"),
                func.sum(case((RecruitmentVisit.enrolled == True, 1), else_=0)).label(
                    "enrolled"
                ),
            )
            if base_filters:
                grade_query = grade_query.filter(*base_filters)
            grade_rows = (
                grade_query.filter(RecruitmentVisit.month.in_(period_month_labels))
                .group_by(RecruitmentVisit.grade)
                .all()
            )

        grade_order = ["幼幼班", "小班", "中班", "大班"]

        def _grade_rates(r) -> dict:
            v, dep, enr = r.visit or 0, r.deposit or 0, r.enrolled or 0
            return {
                "grade": r.grade,
                "visit": v,
                "deposit": dep,
                "enrolled": enr,
                "visit_to_deposit_rate": _pct(dep, v),
                "visit_to_enrolled_rate": _pct(enr, v),
                "deposit_to_enrolled_rate": _pct(enr, dep),
            }

        by_grade_list = sorted(
            [_grade_rates(r) for r in grade_rows],
            key=lambda x: (
                grade_order.index(x["grade"]) if x["grade"] in grade_order else 99
            ),
        )

        return {
            "total_visit": tv,
            "total_deposit": td,
            "total_enrolled": te,
            "total_effective": teff,
            "total_transfer_term": ttr,
            "total_not_enrolled_deposit": (
                sum(p.not_enrolled_deposit or 0 for p in periods)
                if normalized_scope == DATASET_SCOPE_ALL
                else 0
            ),
            "total_enrolled_after_school": (
                sum(p.enrolled_after_school or 0 for p in periods)
                if normalized_scope == DATASET_SCOPE_ALL
                else 0
            ),
            "total_net_enrolled": (
                te - sum(p.enrolled_after_school or 0 for p in periods)
                if normalized_scope == DATASET_SCOPE_ALL
                else te
            ),
            "visit_to_deposit_rate": _pct(td, tv),
            "visit_to_enrolled_rate": _pct(te, tv),
            "deposit_to_enrolled_rate": _pct(te, td),
            "effective_to_enrolled_rate": _pct(te, teff),
            "period_count": len(periods),
            "best_visit_to_enrolled": (
                {
                    "period": best_v2e["period_name"],
                    "rate": best_v2e["visit_to_enrolled_rate"],
                }
                if best_v2e
                else None
            ),
            "worst_visit_to_enrolled": (
                {
                    "period": worst_v2e["period_name"],
                    "rate": worst_v2e["visit_to_enrolled_rate"],
                }
                if worst_v2e
                else None
            ),
            "best_deposit_to_enrolled": (
                {
                    "period": best_d2e["period_name"],
                    "rate": best_d2e["deposit_to_enrolled_rate"],
                }
                if best_d2e
                else None
            ),
            "worst_deposit_to_enrolled": (
                {
                    "period": worst_d2e["period_name"],
                    "rate": worst_d2e["deposit_to_enrolled_rate"],
                }
                if worst_d2e
                else None
            ),
            "trend": trend,
            "by_grade": by_grade_list,
        }


@router.get("/periods")
def list_periods(
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    with session_scope() as session:
        periods = (
            session.query(RecruitmentPeriod)
            .order_by(RecruitmentPeriod.sort_order)
            .all()
        )
        return [_period_to_dict(p) for p in periods]


@router.post("/periods", status_code=201)
def create_period(
    payload: PeriodCreate,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        existing = (
            session.query(RecruitmentPeriod)
            .filter_by(period_name=payload.period_name)
            .first()
        )
        if existing:
            raise HTTPException(status_code=409, detail="期間名稱已存在")
        p = RecruitmentPeriod(
            **payload.model_dump(), created_at=datetime.now(), updated_at=datetime.now()
        )
        session.add(p)
        session.flush()
        return _period_to_dict(p)


@router.put("/periods/{period_id}")
def update_period(
    period_id: int,
    payload: PeriodUpdate,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        p = session.query(RecruitmentPeriod).get(period_id)
        if not p:
            raise HTTPException(status_code=404, detail="期間不存在")
        for field, value in payload.model_dump(exclude_unset=True).items():
            setattr(p, field, value)
        p.updated_at = datetime.now()
        session.flush()
        return _period_to_dict(p)


@router.delete("/periods/{period_id}", status_code=204)
def delete_period(
    period_id: int,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        p = session.query(RecruitmentPeriod).get(period_id)
        if not p:
            raise HTTPException(status_code=404, detail="期間不存在")
        session.delete(p)


@router.post("/periods/{period_id}/sync")
def sync_period_from_visits(
    period_id: int,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    """從訪視明細自動計算並更新指定期間的統計數字（依期間名稱解析月份範圍）"""
    with session_scope() as session:
        p = session.query(RecruitmentPeriod).get(period_id)
        if not p:
            raise HTTPException(status_code=404, detail="期間不存在")

        period_range = _parse_period_range(p.period_name)
        if not period_range:
            raise HTTPException(
                status_code=400,
                detail=f"無法從期間名稱解析日期範圍：{p.period_name}，格式應為 111.03.16~111.09.15",
            )

        start_ym, end_ym = period_range
        period_month_labels = _expand_roc_month_range(start_ym, end_ym)
        dep_case = case((RecruitmentVisit.has_deposit == True, 1), else_=0)

        row = (
            session.query(
                func.count(RecruitmentVisit.id).label("visit_count"),
                func.sum(dep_case).label("deposit_count"),
                func.sum(case((RecruitmentVisit.enrolled == True, 1), else_=0)).label(
                    "enrolled_count"
                ),
                func.sum(
                    case((RecruitmentVisit.transfer_term == True, 1), else_=0)
                ).label("transfer_term_count"),
            )
            .filter(
                RecruitmentVisit.month.in_(period_month_labels),
            )
            .one()
        )

        visit = row.visit_count or 0
        deposit = row.deposit_count or 0
        enrolled = row.enrolled_count or 0
        transfer = row.transfer_term_count or 0
        effective = max(deposit - transfer, 0)

        p.visit_count = visit
        p.deposit_count = deposit
        p.enrolled_count = enrolled
        p.transfer_term_count = transfer
        p.effective_deposit_count = effective
        p.updated_at = datetime.now()

        logger.info(
            f"期間 [{p.period_name}] 已同步：參觀={visit} 預繳={deposit} "
            f"註冊={enrolled} 轉期={transfer} 有效預繳={effective}"
        )
        return _period_to_dict(p)


# ---------------------------------------------------------------------------
# 選項 & 批次匯入
# ---------------------------------------------------------------------------


@router.get("/options")
def get_recruitment_options(
    dataset_scope: str = Query(DATASET_SCOPE_ALL, pattern="^(all)$"),
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """篩選用選項（以 distinct 查詢避免全表掃描回傳到 Python）。"""
    with session_scope() as session:
        scope_filters = _dataset_scope_filters(dataset_scope)

        def _distinct_values(column):
            query = session.query(column)
            if scope_filters:
                query = query.filter(*scope_filters)
            return [
                row[0]
                for row in query.filter(column.isnot(None), column != "")
                .distinct()
                .all()
            ]

        months_set = {
            normalized
            for normalized in (
                _safe_normalize_roc_month(v)
                for v in _distinct_values(RecruitmentVisit.month)
            )
            if normalized
        }
        grades_set = set(_distinct_values(RecruitmentVisit.grade))
        sources_set = {
            _normalize_source_label(value)
            for value in _distinct_values(RecruitmentVisit.source)
        }
        referrers_set = set(_distinct_values(RecruitmentVisit.referrer))

        # 合併手動登記月份
        if _normalize_dataset_scope(dataset_scope) == DATASET_SCOPE_ALL:
            registered = {
                normalized
                for normalized in (
                    _safe_normalize_roc_month(r.month)
                    for r in session.query(RecruitmentMonth.month).all()
                )
                if normalized
            }
            months_set |= registered

        return {
            "months": sorted(months_set, key=_roc_month_sort_key),
            "grades": sorted(grades_set),
            "sources": sorted(sources_set),
            "referrers": sorted(referrers_set),
            "no_deposit_reasons": NO_DEPOSIT_REASONS,
        }


# ---------------------------------------------------------------------------
# 月份管理
# ---------------------------------------------------------------------------


class MonthCreate(BaseModel):
    month: str = Field(..., min_length=1, max_length=10)

    @field_validator("month")
    @classmethod
    def validate_month_format(cls, v: str) -> str:
        return _normalize_roc_month(v)


@router.get("/months")
def list_months(
    _=Depends(require_staff_permission(Permission.RECRUITMENT_READ)),
):
    """列出所有已手動登記的月份"""
    with session_scope() as session:
        rows = session.query(RecruitmentMonth).all()
        return [
            {"id": r.id, "month": _safe_normalize_roc_month(r.month) or r.month}
            for r in sorted(rows, key=lambda item: _roc_month_sort_key(item.month))
        ]


@router.post("/months", status_code=201)
def create_month(
    payload: MonthCreate,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    """手動登記一個月份（若已有訪視記錄仍可登記，無重複效果）"""
    with session_scope() as session:
        existing = (
            session.query(RecruitmentMonth).filter_by(month=payload.month).first()
        )
        if existing:
            raise HTTPException(status_code=409, detail=f"月份 {payload.month} 已存在")
        rec = RecruitmentMonth(month=payload.month)
        session.add(rec)
        session.flush()
        logger.info(f"手動登記月份：{payload.month}")
        return {"id": rec.id, "month": rec.month}


@router.delete("/months/{month}")
def delete_month(
    month: str,
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    """刪除手動登記月份（不影響該月份的訪視記錄）"""
    with session_scope() as session:
        rec = session.query(RecruitmentMonth).filter_by(month=month).first()
        if not rec:
            raise HTTPException(status_code=404, detail=f"登記月份 {month} 不存在")
        session.delete(rec)
        logger.info(f"刪除登記月份：{month}")
        return {"deleted": month}


@router.post("/import", status_code=201)
def import_recruitment_records(
    records: List[ImportRecord],
    _=Depends(require_staff_permission(Permission.RECRUITMENT_WRITE)),
):
    with session_scope() as session:
        existing = set(
            (
                (r.child_name or "").strip(),
                (r.month or "").strip(),
                (r.seq_no or "").strip(),
                (r.visit_date or "").strip(),
            )
            for r in session.query(
                RecruitmentVisit.child_name,
                RecruitmentVisit.month,
                RecruitmentVisit.seq_no,
                RecruitmentVisit.visit_date,
            ).all()
        )
        inserted = 0
        skipped = 0
        for rec in records:
            name = (rec.幼生姓名 or "").strip()
            raw_month = (rec.月份 or "").strip()
            seq_no = (rec.序號 or "").strip()
            visit_date = (rec.日期 or "").strip()
            try:
                month = _normalize_roc_month(raw_month) if raw_month else ""
            except ValueError:
                skipped += 1
                continue
            month = _extract_roc_month_from_visit_date(visit_date) or month
            if not name or not month:
                skipped += 1
                continue
            dedup_key = (name, month, seq_no, visit_date)
            if dedup_key in existing:
                skipped += 1
                continue
            visit = RecruitmentVisit(
                month=month,
                seq_no=rec.序號,
                visit_date=rec.日期,
                child_name=name,
                birthday=_parse_roc_date(rec.生日),
                grade=rec.適讀班級,
                phone=rec.電話,
                address=rec.地址,
                district=rec.行政區,
                source=rec.幼生來源,
                referrer=rec.介紹者,
                deposit_collector=rec.收預繳人員,
                has_deposit=(rec.是否預繳 == "是"),
                notes=rec.備註,
                parent_response=rec.電訪後家長回應,
            )
            visit.expected_start_label = _extract_expected_label_from_text(
                visit.notes, visit.parent_response, visit.grade
            )
            session.add(visit)
            existing.add(dedup_key)
            inserted += 1
        logger.info(f"招生資料匯入：插入 {inserted} 筆，跳過 {skipped} 筆")
        return {"inserted": inserted, "skipped": skipped}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _to_dict(r: RecruitmentVisit) -> dict:
    return {
        "id": r.id,
        "month": r.month,
        "seq_no": r.seq_no,
        "visit_date": r.visit_date,
        "child_name": r.child_name,
        "birthday": r.birthday.isoformat() if r.birthday else None,
        "grade": r.grade,
        "phone": r.phone,
        "address": r.address,
        "district": r.district,
        "source": r.source,
        "referrer": r.referrer,
        "deposit_collector": r.deposit_collector,
        "has_deposit": r.has_deposit,
        "notes": r.notes,
        "parent_response": r.parent_response,
        "no_deposit_reason": r.no_deposit_reason,
        "no_deposit_reason_detail": r.no_deposit_reason_detail,
        "enrolled": r.enrolled,
        "transfer_term": r.transfer_term,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }


def _period_to_dict(p: RecruitmentPeriod) -> dict:
    visit = p.visit_count or 0
    deposit = p.deposit_count or 0
    enrolled = p.enrolled_count or 0
    effective = p.effective_deposit_count or 0

    def _pct(n, d):
        return round(n / d * 100, 1) if d else 0

    return {
        "id": p.id,
        "period_name": p.period_name,
        "visit_count": visit,
        "deposit_count": deposit,
        "enrolled_count": enrolled,
        "transfer_term_count": p.transfer_term_count or 0,
        "effective_deposit_count": effective,
        "not_enrolled_deposit": p.not_enrolled_deposit or 0,
        "enrolled_after_school": p.enrolled_after_school or 0,
        "notes": p.notes,
        "sort_order": p.sort_order or 0,
        # 回傳原始數值（百分比）供前端直接顯示加 %
        "visit_to_deposit_rate": _pct(deposit, visit),
        "visit_to_enrolled_rate": _pct(enrolled, visit),
        "deposit_to_enrolled_rate": _pct(enrolled, deposit),
        "effective_to_enrolled_rate": _pct(enrolled, effective),
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }
