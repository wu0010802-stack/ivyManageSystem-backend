"""
政府申報格式匯出 — 後端單元測試

測試重點：
  1. _estimate_withholding 扣繳估算邏輯
  2. 免稅額以下回傳 0
  3. 上限封頂（560_000 × 5% = 28_000）
  4. 勞保端點基本呼叫（mock session / insurance）
  5. 健保、勞退、扣繳端點 smoke test
"""

from io import BytesIO
from types import SimpleNamespace

import pytest
from unittest.mock import MagicMock, patch
from fastapi.testclient import TestClient
from fastapi import FastAPI
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# 直接測試純函式 _estimate_withholding
# ──────────────────────────────────────────────────────────────────────────────

from api.gov_reports import _estimate_withholding


class TestEstimateWithholding:
    def test_below_deduction_returns_zero(self):
        """全年收入低於免稅額（428,000）→ 扣繳為 0"""
        assert _estimate_withholding(400_000) == 0

    def test_exactly_at_deduction_returns_zero(self):
        assert _estimate_withholding(428_000) == 0

    def test_one_nt_above_deduction(self):
        """428,001 → taxable=1 → round(1 × 0.05) = 0"""
        assert _estimate_withholding(428_001) == 0

    def test_typical_salary(self):
        """600,000 → taxable=172,000 → round(172,000×0.05)=8,600"""
        assert _estimate_withholding(600_000) == 8_600

    def test_cap_at_560k_taxable(self):
        """超過上限 560,000 taxable → 最多 28,000"""
        # annual=1,100,000 → taxable=672,000 > 560,000 → capped 28,000
        result = _estimate_withholding(1_100_000)
        assert result == 28_000

    def test_exactly_at_cap(self):
        """428,000 + 560,000 = 988,000 → exactly 28,000"""
        assert _estimate_withholding(988_000) == 28_000

    def test_above_cap_same_result(self):
        """超過封頂線後結果不再增加"""
        r1 = _estimate_withholding(988_000)
        r2 = _estimate_withholding(2_000_000)
        assert r1 == r2 == 28_000


# ──────────────────────────────────────────────────────────────────────────────
# Smoke tests：端點在空資料庫回傳 200
# ──────────────────────────────────────────────────────────────────────────────


def _make_test_app():
    from api.gov_reports import router, init_gov_report_services

    ins_svc = MagicMock()
    ins_svc.calculate.return_value = MagicMock(
        employee_share=300,
        employer_share=600,
        govt_share=100,
        health_employee=400,
        health_employer=800,
        health_dependents=0,
        pension_employer=500,
        pension_employee=0,
        pension_employee_rate=0.0,
    )
    init_gov_report_services(ins_svc)

    app = FastAPI()

    # 覆蓋基礎 auth dependency，讓測試繞過 JWT 驗證
    from utils.auth import get_current_user

    async def _mock_user():
        return {"id": 1, "username": "test", "role": "admin", "permission_names": ["*"]}

    app.dependency_overrides[get_current_user] = _mock_user
    app.include_router(router)
    return app


@pytest.fixture(scope="module")
def gov_client():
    app = _make_test_app()
    with patch("api.gov_reports.get_session") as mock_gs:
        mock_sess = MagicMock()
        mock_sess.__enter__ = MagicMock(return_value=mock_sess)
        mock_sess.__exit__ = MagicMock(return_value=False)
        # employees query returns empty list
        mock_sess.query.return_value.filter.return_value.all.return_value = []
        mock_sess.query.return_value.filter.return_value.filter.return_value.all.return_value = (
            []
        )
        mock_gs.return_value = mock_sess
        yield TestClient(app, raise_server_exceptions=False)


class TestGovReportEndpoints:
    def test_labor_insurance_xlsx(self, gov_client):
        resp = gov_client.get(
            "/api/gov-reports/labor-insurance?year=2026&month=3&fmt=xlsx"
        )
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_labor_insurance_txt(self, gov_client):
        resp = gov_client.get(
            "/api/gov-reports/labor-insurance?year=2026&month=3&fmt=txt"
        )
        assert resp.status_code == 200
        assert resp.headers.get("content-type", "").startswith("text/plain")

    def test_health_insurance(self, gov_client):
        resp = gov_client.get("/api/gov-reports/health-insurance?year=2026&month=3")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_pension(self, gov_client):
        resp = gov_client.get("/api/gov-reports/pension?year=2026&month=3")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_withholding(self, gov_client):
        resp = gov_client.get("/api/gov-reports/withholding?year=2026")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")


# ──────────────────────────────────────────────────────────────────────────────
# Bug #4：健保名冊「員工自付(一般保費)」欄須扣除二代健保補充保費
# ──────────────────────────────────────────────────────────────────────────────


def _read_health_emp_cell(content: bytes):
    """讀回健保名冊 xlsx 第一筆資料列的「員工自付(30%)」欄（F4）。

    版面：row1 標題 / row2 投保單位 / row3 表頭 / row4 起為資料。
    員工自付為第 6 欄（F）。
    """
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    return ws.cell(row=4, column=6).value


class TestHealthInsuranceExcludesSupplementaryPremium:
    """record 路徑：health_insurance_employee 含二代補充保費，
    名冊一般保費欄須報乾淨 base（= 實扣健保 − 補充保費）。"""

    def _build_app_with_record(self, supplementary: float):
        from api import gov_reports
        from api.gov_reports import router, init_gov_report_services

        init_gov_report_services(MagicMock())

        app = FastAPI()
        from utils.auth import get_current_user

        async def _mock_user():
            return {
                "id": 1,
                "username": "test",
                "role": "admin",
                "permission_names": ["*"],
            }

        # 覆蓋限流依賴為 no-op，避免與同檔其他匯出 smoke test 共用 IP key 觸發 429
        app.dependency_overrides[get_current_user] = _mock_user
        app.dependency_overrides[gov_reports._rate_limit] = lambda: None
        app.include_router(router)

        emp = SimpleNamespace(
            id=1,
            name="王小明",
            id_number="A123456789",
            birthday=None,
            dependents=0,
            hire_date=None,
            resign_date=None,
        )
        # 當月實扣健保 1000，其中 150 為二代補充保費 → 名冊應報 850
        sr = SimpleNamespace(
            health_insurance_employee=1000,
            health_insurance_employer=2000,
            supplementary_health_employee=supplementary,
        )
        return app, emp, sr

    def test_record_path_excludes_supplementary(self):
        app, emp, sr = self._build_app_with_record(supplementary=150)

        with (
            patch("api.gov_reports.get_session", MagicMock()),
            patch("api.gov_reports._active_employees", return_value=[emp]),
            patch("api.gov_reports._assert_salary_period_finalized", return_value=None),
            patch("api.gov_reports._salary_map", return_value={emp.id: sr}),
            patch("api.gov_reports._resolve_insured", return_value=30000),
            patch(
                "api.gov_reports._ins_calc",
                return_value=SimpleNamespace(
                    health_insured_amount=30000, insured_amount=30000
                ),
            ),
        ):
            client = TestClient(app, raise_server_exceptions=False)
            resp = client.get("/api/gov-reports/health-insurance?year=2026&month=3")

        assert resp.status_code == 200, resp.text
        # 1000（實扣，含補充保費）− 150（補充保費）= 850（乾淨一般保費）
        assert _read_health_emp_cell(resp.content) == 850

    def test_record_path_zero_supplementary_unchanged(self):
        """無補充保費時，名冊金額不變（= 實扣健保）。"""
        app, emp, sr = self._build_app_with_record(supplementary=0)

        with (
            patch("api.gov_reports.get_session", MagicMock()),
            patch("api.gov_reports._active_employees", return_value=[emp]),
            patch("api.gov_reports._assert_salary_period_finalized", return_value=None),
            patch("api.gov_reports._salary_map", return_value={emp.id: sr}),
            patch("api.gov_reports._resolve_insured", return_value=30000),
            patch(
                "api.gov_reports._ins_calc",
                return_value=SimpleNamespace(
                    health_insured_amount=30000, insured_amount=30000
                ),
            ),
        ):
            client = TestClient(app, raise_server_exceptions=False)
            resp = client.get("/api/gov-reports/health-insurance?year=2026&month=3")

        assert resp.status_code == 200, resp.text
        assert _read_health_emp_cell(resp.content) == 1000


# ──────────────────────────────────────────────────────────────────────────────
# P1：扣繳憑單「全年給付總額」須納入特休未休折現（unused_leave_payout，屬§38 薪資所得）
# ──────────────────────────────────────────────────────────────────────────────


def _read_withholding_annual_income_cell(content: bytes):
    """讀回扣繳憑單 xlsx 第一筆資料列的「全年給付總額」欄（E4）。

    版面：row1 標題 / row2 扣繳義務人 / row3 表頭 / row4 起為資料。
    全年給付總額為第 5 欄（E，header「全年給付\\n總額」）。
    """
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    return ws.cell(row=4, column=5).value


class TestWithholdingIncludesUnusedLeavePayout:
    """扣繳憑單逐 SalaryRecord 聚合年度所得時，須納入 unused_leave_payout
    （特休未休折現，屬薪資所得§38）。漏報即對國稅局少報員工所得。"""

    def _build_app_and_records(self, *, unused_leave_payout: float):
        from api import gov_reports
        from api.gov_reports import router, init_gov_report_services

        init_gov_report_services(MagicMock())

        app = FastAPI()
        from utils.auth import get_current_user

        async def _mock_user():
            return {
                "id": 1,
                "username": "test",
                "role": "admin",
                "permission_names": ["*"],
            }

        app.dependency_overrides[get_current_user] = _mock_user
        app.dependency_overrides[gov_reports._rate_limit] = lambda: None
        app.include_router(router)

        emp = SimpleNamespace(
            id=1,
            name="王小明",
            id_number="A123456789",
            unreported_for_tax=False,
        )
        # 一名員工、一筆 record：底薪所得 500_000，加上特休未休折現
        sr = SimpleNamespace(
            employee_id=1,
            gross_salary=500_000,
            festival_bonus=0,
            overtime_bonus=0,
            unused_leave_payout=unused_leave_payout,
            labor_insurance_employee=0,
            health_insurance_employee=0,
            pension_employee=0,
        )
        return app, [(sr, emp)]

    def _mock_session_with(self, records):
        """建出讓 query(SalaryRecord, Employee).join(...).filter(...).all()
        回傳指定 (sr, emp) tuple list 的 mock session。"""
        sess = MagicMock()
        sess.query.return_value.join.return_value.filter.return_value.all.return_value = (
            records
        )
        return sess

    def test_annual_income_includes_unused_leave_payout(self):
        """RED 證明：修前「全年給付總額」漏算 unused_leave_payout。

        gross 500_000 + 特休未休折現 80_000 = 580_000。
        修前公式只算 gross + festival + overtime = 500_000（少報 80_000）。
        """
        app, records = self._build_app_and_records(unused_leave_payout=80_000)
        sess = self._mock_session_with(records)

        with (
            patch("api.gov_reports.get_session", return_value=sess),
            patch("api.gov_reports._assert_salary_period_finalized", return_value=None),
        ):
            client = TestClient(app, raise_server_exceptions=False)
            resp = client.get("/api/gov-reports/withholding?year=2026")

        assert resp.status_code == 200, resp.text
        # 500_000（gross）+ 80_000（特休未休折現）= 580_000
        assert _read_withholding_annual_income_cell(resp.content) == 580_000

    def test_zero_unused_leave_payout_unchanged(self):
        """無特休未休折現時，全年給付總額 = gross（行為不變）。"""
        app, records = self._build_app_and_records(unused_leave_payout=0)
        sess = self._mock_session_with(records)

        with (
            patch("api.gov_reports.get_session", return_value=sess),
            patch("api.gov_reports._assert_salary_period_finalized", return_value=None),
        ):
            client = TestClient(app, raise_server_exceptions=False)
            resp = client.get("/api/gov-reports/withholding?year=2026")

        assert resp.status_code == 200, resp.text
        assert _read_withholding_annual_income_cell(resp.content) == 500_000
