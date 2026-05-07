"""
api/activity/registrations.py — 報名管理端點（含 batch-payment、export）

⚠️ 注意：batch-payment 和 export 為靜態路由，必須定義在 /{registration_id}/... 之前。
"""

import io
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import func
from sqlalchemy.exc import CompileError, IntegrityError, OperationalError

from models.database import (
    get_session,
    ActivityCourse,
    ActivityRegistration,
    RegistrationCourse,
    RegistrationSupply,
    ActivityPaymentRecord,
    RegistrationChange,
    ActivitySupply,
    ActivitySession,
    ActivityAttendance,
)
from services.activity_service import activity_service
from services.report_cache_service import report_cache_service
from utils.errors import raise_safe_500
from utils.excel_utils import SafeWorksheet
from utils.auth import require_staff_permission
from utils.permissions import Permission
from utils.portfolio_access import can_view_guardian_pii, can_view_student_pii
from utils.rate_limit import SlidingWindowLimiter
from utils.finance_guards import (
    FINANCE_APPROVAL_THRESHOLD,
    require_finance_approve,
)

from ._shared import (
    PaymentUpdate,
    RemarkUpdate,
    BatchPaymentUpdate,
    AddPaymentRequest,
    AdminRegistrationPayload,
    AdminRegistrationBasicUpdate,
    AddCourseRequest,
    AddSupplyRequest,
    VoidPaymentRequest,
    SYSTEM_RECONCILE_METHOD,
    MIN_REFUND_REASON_LENGTH,
    _not_found,
    _derive_payment_status,
    _compute_is_paid,
    _calc_total_amount,
    _invalidate_activity_dashboard_caches,
    _batch_calc_total_amounts,
    _build_registration_filter_query,
    _fetch_reg_course_names,
    _require_active_classroom,
    _require_daily_close_unlocked,
    _attach_courses,
    _attach_supplies,
    _match_student_id,
    has_payment_approve,
    require_refund_reason,
    require_approve_for_large_refund,
    require_approve_for_cumulative_refund,
    TAIPEI_TZ,
    get_line_service,
    today_taipei,
)
from utils.academic import resolve_academic_term_filters

logger = logging.getLogger(__name__)
router = APIRouter()


def _invalidate_finance_summary_cache() -> None:
    """金流寫入後失效 /finance-summary 快取（TTL 30 分，否則看到舊值）。"""
    try:
        report_cache_service.invalidate_category(None, "reports_finance_summary")
    except Exception:
        logger.warning("invalidate finance_summary cache failed", exc_info=True)


_export_limiter = SlidingWindowLimiter(
    max_calls=5,
    window_seconds=60,
    name="activity_export",
    error_detail="匯出過於頻繁，請稍後再試",
).as_dependency()

# 批次繳費寫入每次影響多筆 registration + 繳費紀錄，放寬但仍需限流
_batch_payment_limiter = SlidingWindowLimiter(
    max_calls=10,
    window_seconds=60,
    name="activity_batch_payment",
    error_detail="批次繳費操作過於頻繁，請稍後再試",
).as_dependency()

# 匯出單次查詢上限，避免大報表造成記憶體爆炸或 timeout
MAX_EXPORT_ROWS = 5000


# ── 靜態路由（必須優先定義，在 /{id}/... 動態路由之前）─────────────────────


@router.put("/registrations/batch-payment")
async def batch_update_payment(
    body: BatchPaymentUpdate,
    _rl=Depends(_batch_payment_limiter),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """批次標記為已繳費（僅此方向；未繳費全額沖帳已禁用）

    Why: 批次把一串報名「標記未繳費」會一次寫多筆全額 refund，誤操作後果嚴重且
    不可部分回滾。schema 層已禁止 is_paid=False（進到這裡一律是 True）；
    若需把某筆退費，請改用 PUT /registrations/{id}/payment（帶
    confirm_refund_amount）或 DELETE /payments/{id} 軟刪對應繳費。

    併發保護：`.with_for_update()` 鎖住目標 registration，避免兩個客戶端同時
    呼叫造成重複寫 payment_record 或 lost update。
    """
    session = get_session()
    try:
        regs = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id.in_(body.ids),
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .all()
        )
        if not regs:
            raise HTTPException(status_code=404, detail="找不到指定報名資料")

        operator = current_user.get("username", "")
        today = today_taipei()

        # 批次補齊皆寫 today；若今日已簽核則拒絕，避免日結 snapshot 失準
        _require_daily_close_unlocked(session, today)

        # P3 N+1 修正：一次 GROUP BY 查詢所有應繳金額
        unpaid_reg_ids = [reg.id for reg in regs if not reg.is_paid]
        total_amount_map = (
            _batch_calc_total_amounts(session, unpaid_reg_ids) if unpaid_reg_ids else {}
        )

        # ── 整批 shortfall 累積簽核 ────────────────────────────────────
        # Why: 批次補齊把欠費直接寫成 system payment，會計可一鍵把報表變漂亮；
        # 整批 shortfall 合計超過閾值即整批需具備 ACTIVITY_PAYMENT_APPROVE 才能執行
        total_shortfall = 0
        for reg in regs:
            if not reg.is_paid:
                total_amount = total_amount_map.get(reg.id, 0)
                shortfall = total_amount - (reg.paid_amount or 0)
                if shortfall > 0:
                    total_shortfall += shortfall
        if total_shortfall > 0:
            require_finance_approve(
                total_shortfall,
                current_user,
                action_label=f"批次補齊整批合計 NT${total_shortfall:,}",
            )

        notes_text = f"（批次標記已繳費自動補齊：{body.reason}）"
        for reg in regs:
            if not reg.is_paid:
                total_amount = total_amount_map.get(reg.id, 0)
                shortfall = total_amount - (reg.paid_amount or 0)
                if shortfall > 0:
                    rec = ActivityPaymentRecord(
                        registration_id=reg.id,
                        type="payment",
                        amount=shortfall,
                        payment_date=today,
                        payment_method=SYSTEM_RECONCILE_METHOD,
                        notes=notes_text,
                        operator=operator,
                    )
                    session.add(rec)
                    reg.paid_amount = total_amount
                reg.is_paid = _compute_is_paid(reg.paid_amount or 0, total_amount)
            activity_service.log_change(
                session,
                reg.id,
                reg.student_name,
                "批次更新付款狀態",
                f"付款狀態批次更新為：已繳費（原因：{body.reason}）",
                operator,
            )

        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        _invalidate_finance_summary_cache()
        logger.warning(
            "批次付款狀態更新：筆數=%d is_paid=True operator=%s",
            len(regs),
            operator,
        )
        return {
            "message": f"已更新 {len(regs)} 筆報名為已繳費",
            "updated": len(regs),
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.get("/registrations/export")
def export_registrations(
    search: Optional[str] = None,
    payment_status: Optional[str] = None,
    course_id: Optional[int] = None,
    classroom_name: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """匯出報名名單為 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    session = get_session()
    try:
        q = _build_registration_filter_query(
            session,
            search=search,
            payment_status=payment_status,
            course_id=course_id,
            classroom_name=classroom_name,
        )
        total_count = q.count()
        if total_count > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"匯出筆數 {total_count} 超過上限 {MAX_EXPORT_ROWS}，"
                    "請加上條件（如課程/班級/繳費狀態）縮小範圍後再匯出"
                ),
            )
        regs = q.order_by(ActivityRegistration.created_at.desc()).all()
        reg_ids = [r.id for r in regs]
        course_name_map = _fetch_reg_course_names(session, reg_ids)

        wb = openpyxl.Workbook()
        ws = SafeWorksheet(wb.active)
        ws.title = "報名名單"

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
        )
        center = Alignment(horizontal="center", vertical="center")

        headers = ["序號", "學生姓名", "班級", "課程", "付款狀態", "備註", "報名時間"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for idx, reg in enumerate(regs, start=1):
            ws.append(
                [
                    idx,
                    reg.student_name,
                    reg.class_name or "",
                    "、".join(course_name_map.get(reg.id, [])),
                    "已繳費" if reg.is_paid else "未繳費",
                    reg.remark or "",
                    reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else "",
                ]
            )

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"activity_registrations_{datetime.now(TAIPEI_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()


@router.get("/registrations/payment-report")
def export_payment_report(
    search: Optional[str] = None,
    payment_status: Optional[str] = None,
    course_id: Optional[int] = None,
    classroom_name: Optional[str] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
    _: None = Depends(_export_limiter),
):
    """匯出繳費帳務報表（兩個工作表：繳費總覽 + 繳費明細）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    _HEADER_FONT = Font(bold=True)
    _HEADER_FILL = PatternFill(
        start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
    )
    _CENTER = Alignment(horizontal="center", vertical="center")

    session = get_session()
    try:
        q = _build_registration_filter_query(
            session,
            search=search,
            payment_status=payment_status,
            course_id=course_id,
            classroom_name=classroom_name,
        )
        total_count = q.count()
        if total_count > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"匯出筆數 {total_count} 超過上限 {MAX_EXPORT_ROWS}，"
                    "請加上條件（如課程/班級/繳費狀態）縮小範圍後再匯出"
                ),
            )
        regs = q.order_by(ActivityRegistration.created_at.desc()).all()
        reg_ids = [r.id for r in regs]

        # 批次計算應繳金額
        total_amount_map = (
            _batch_calc_total_amounts(session, reg_ids) if reg_ids else {}
        )

        # 批次查詢課程名稱
        course_name_map = _fetch_reg_course_names(session, reg_ids)

        # 批次查詢繳費明細
        payment_records = []
        payment_map: dict[int, list] = defaultdict(list)
        last_payment_date_map: dict[int, str] = {}
        if reg_ids:
            payment_records = (
                session.query(ActivityPaymentRecord)
                .filter(ActivityPaymentRecord.registration_id.in_(reg_ids))
                .order_by(
                    ActivityPaymentRecord.registration_id,
                    ActivityPaymentRecord.payment_date.asc(),
                )
                .all()
            )
            for pr in payment_records:
                payment_map[pr.registration_id].append(pr)
                if pr.payment_date:
                    date_str = pr.payment_date.isoformat()
                    existing = last_payment_date_map.get(pr.registration_id, "")
                    if date_str > existing:
                        last_payment_date_map[pr.registration_id] = date_str

        wb = openpyxl.Workbook()

        # ── 工作表一：繳費總覽 ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "繳費總覽"
        headers1 = [
            "序號",
            "學生",
            "班級",
            "報名課程",
            "應繳總額",
            "已繳金額",
            "差額",
            "狀態",
            "最後繳費日",
        ]
        for col, h in enumerate(headers1, start=1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        status_label_map = {
            "paid": "已繳清",
            "partial": "部分繳費",
            "unpaid": "未繳費",
            "overpaid": "超額繳費",
        }
        for idx, reg in enumerate(regs, start=1):
            total = total_amount_map.get(reg.id, 0)
            paid = reg.paid_amount or 0
            diff = paid - total
            status = _derive_payment_status(paid, total)
            ws1.append(
                [
                    idx,
                    reg.student_name,
                    reg.class_name or "",
                    "、".join(course_name_map.get(reg.id, [])),
                    total,
                    paid,
                    diff,
                    status_label_map.get(status, status),
                    last_payment_date_map.get(reg.id, ""),
                ]
            )

        for col in ws1.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── 工作表二：繳費明細 ──────────────────────────────────────────
        # 軟刪（voided）紀錄獨立標示：類型欄加「（已作廢）」、新增作廢欄位顯示作廢人/
        # 時間/原因，避免與有效流水混為一談；總覽工作表的金額計算已排除 voided。
        ws2 = SafeWorksheet(wb.create_sheet(title="繳費明細"))
        headers2 = [
            "學生",
            "班級",
            "類型",
            "金額",
            "方式",
            "日期",
            "操作人員",
            "備註",
            "作廢狀態",
            "作廢人",
            "作廢時間",
            "作廢原因",
        ]
        for col, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        reg_meta = {r.id: r for r in regs}
        for pr in payment_records:
            reg = reg_meta.get(pr.registration_id)
            is_voided = pr.voided_at is not None
            type_label = "繳費" if pr.type == "payment" else "退費"
            if is_voided:
                type_label = f"{type_label}（已作廢）"
            ws2.append(
                [
                    reg.student_name if reg else "",
                    reg.class_name if reg else "",
                    type_label,
                    pr.amount,
                    pr.payment_method or "",
                    pr.payment_date.isoformat() if pr.payment_date else "",
                    pr.operator or "",
                    pr.notes or "",
                    "已作廢" if is_voided else "",
                    pr.voided_by or "",
                    pr.voided_at.isoformat() if pr.voided_at else "",
                    pr.void_reason or "",
                ]
            )

        for col in ws2.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = (
            f"payment_report_{datetime.now(TAIPEI_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        logger.warning(
            "繳費帳務報表匯出：operator=%s 筆數=%d",
            current_user.get("username"),
            len(regs),
        )
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()


# ── 後台手動新增報名 ─────────────────────────────────────────────────────────


@router.post("/registrations", status_code=201)
async def admin_create_registration(
    body: AdminRegistrationPayload,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台手動新增報名（不受報名開放時間限制，需 ACTIVITY_WRITE 權限）"""
    # 空報名守衛：至少要選 1 門課程，避免產生 total_amount=0 的殼子後又被 POS 誤收款
    if not body.courses:
        raise HTTPException(status_code=400, detail="請至少選擇一門課程再新增報名")

    session = get_session()
    try:
        sy, sem = resolve_academic_term_filters(body.school_year, body.semester)

        existing = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.student_name == body.name,
                ActivityRegistration.birthday == body.birthday,
                ActivityRegistration.is_active.is_(True),
                ActivityRegistration.school_year == sy,
                ActivityRegistration.semester == sem,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=400, detail="此學生本學期已有有效報名，請改用編輯功能"
            )

        classroom = _require_active_classroom(session, body.class_)

        course_names = [item.name for item in body.courses]
        if len(course_names) != len(set(course_names)):
            raise HTTPException(status_code=400, detail="課程清單中有重複項目")
        supply_names = [item.name for item in body.supplies]
        if len(supply_names) != len(set(supply_names)):
            raise HTTPException(status_code=400, detail="用品清單中有重複項目")

        courses_by_name = (
            {
                c.name: c
                for c in session.query(ActivityCourse)
                .filter(
                    ActivityCourse.name.in_(course_names),
                    ActivityCourse.is_active.is_(True),
                    ActivityCourse.school_year == sy,
                    ActivityCourse.semester == sem,
                )
                .with_for_update()
                .all()
            }
            if course_names
            else {}
        )

        supplies_by_name = (
            {
                s.name: s
                for s in session.query(ActivitySupply)
                .filter(
                    ActivitySupply.name.in_(supply_names),
                    ActivitySupply.is_active.is_(True),
                    ActivitySupply.school_year == sy,
                    ActivitySupply.semester == sem,
                )
                .all()
            }
            if supply_names
            else {}
        )

        _reg_course_ids = [c.id for c in courses_by_name.values()]
        enrolled_count_map = (
            dict(
                session.query(
                    RegistrationCourse.course_id, func.count(RegistrationCourse.id)
                )
                .join(
                    ActivityRegistration,
                    RegistrationCourse.registration_id == ActivityRegistration.id,
                )
                .filter(
                    RegistrationCourse.course_id.in_(_reg_course_ids),
                    RegistrationCourse.status.in_(["enrolled", "promoted_pending"]),
                    ActivityRegistration.is_active.is_(True),
                )
                .group_by(RegistrationCourse.course_id)
                .all()
            )
            if _reg_course_ids
            else {}
        )

        matched_student_id = _match_student_id(session, body.name, body.birthday)

        operator = current_user.get("username", "")
        reg = ActivityRegistration(
            student_name=body.name,
            birthday=body.birthday,
            class_name=classroom.name,
            email=body.email or None,
            remark=body.remark or None,
            school_year=sy,
            semester=sem,
            student_id=matched_student_id,
        )
        session.add(reg)
        session.flush()

        has_waitlist, waitlist_course_names = _attach_courses(
            session, reg.id, body.courses, courses_by_name, enrolled_count_map
        )
        _attach_supplies(session, reg.id, body.supplies, supplies_by_name)

        activity_service.log_change(
            session,
            reg.id,
            reg.student_name,
            "後台新增報名",
            f"班級：{classroom.name}，課程：{'、'.join(course_names) or '無'}，用品：{'、'.join(supply_names) or '無'}",
            operator,
        )

        session.commit()
        _invalidate_activity_dashboard_caches(session)
        logger.warning(
            "後台新增報名：id=%s student=%s operator=%s",
            reg.id,
            reg.student_name,
            operator,
        )

        return {
            "message": ("新增成功（部分課程進入候補）" if has_waitlist else "新增成功"),
            "id": reg.id,
            "waitlisted": has_waitlist,
            "waitlist_courses": waitlist_course_names,
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("後台新增報名失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


# ── 審核工作流（pending / match / reject / rematch / students-search）─────


class RegistrationMatchRequest(BaseModel):
    student_id: int = Field(..., gt=0)


class RegistrationRejectRequest(BaseModel):
    reason: str = Field(..., min_length=2, max_length=200)

    @field_validator("reason", mode="before")
    @classmethod
    def _strip_reason(cls, v):
        if isinstance(v, str):
            stripped = v.strip()
            if len(stripped) < 2:
                raise ValueError("拒絕原因至少需 2 個字，方便事後追溯")
            return stripped
        return v


class RegistrationRematchRequest(BaseModel):
    """重新比對可選欄位：校方可即時修正家長打錯的 name/birthday/parent_phone。

    三欄皆可選——未提供時沿用 registration 原值。提供的欄位會在比對前寫回 reg，
    即使比對仍失敗也保留修改內容，避免校方白打一次字。
    """

    model_config = ConfigDict(populate_by_name=True)

    name: Optional[str] = Field(None, min_length=1, max_length=50)
    birthday: Optional[str] = None
    parent_phone: Optional[str] = Field(None, min_length=8, max_length=30)

    @field_validator("name", mode="before")
    @classmethod
    def _strip_name(cls, v):
        if isinstance(v, str):
            stripped = v.strip()
            return stripped or None
        return v

    @field_validator("birthday")
    @classmethod
    def _validate_birthday(cls, v):
        if v is None or v == "":
            return None
        from datetime import date as _d

        try:
            _d.fromisoformat(v)
        except ValueError:
            raise ValueError("生日格式必須為 YYYY-MM-DD")
        return v

    @field_validator("parent_phone", mode="before")
    @classmethod
    def _normalize_phone(cls, v):
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        from ._shared import _validate_tw_mobile

        return _validate_tw_mobile(v)


def _serialize_pending_item(
    r: ActivityRegistration,
    *,
    can_see_student_pii: bool = True,
    can_see_guardian_pii: bool = True,
) -> dict:
    """F-026：對缺 STUDENTS_READ 遮罩 birthday / classroom_id；對缺 GUARDIANS_READ
    遮罩 parent_phone / email。"""
    return {
        "id": r.id,
        "student_name": r.student_name,
        "birthday": r.birthday if can_see_student_pii else None,
        "class_name": r.class_name,
        "classroom_id": r.classroom_id if can_see_student_pii else None,
        "parent_phone": r.parent_phone if can_see_guardian_pii else None,
        "match_status": r.match_status,
        "pending_review": r.pending_review,
        "email": r.email if can_see_guardian_pii else None,
        "school_year": r.school_year,
        "semester": r.semester,
        "remark": r.remark or "",
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "reviewed_by": r.reviewed_by,
        "reviewed_at": r.reviewed_at.isoformat() if r.reviewed_at else None,
    }


@router.get("/registrations/pending")
async def list_pending_registrations(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    school_year: Optional[int] = Query(None, ge=100, le=200),
    semester: Optional[int] = Query(None, ge=1, le=2),
    status: str = Query("all", pattern="^(pending|rejected|all)$"),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得待審核 / 已拒絕報名清單（合併於同一頁）。

    status=pending：pending_review=true、is_active=true
    status=rejected：match_status='rejected'、is_active=false
    status=all（預設）：兩者聯集，前端以 match_status / is_active 判斷顯示
    """
    from utils.academic import resolve_academic_term_filters
    from sqlalchemy import and_, or_

    session = get_session()
    try:
        sy, sem = resolve_academic_term_filters(school_year, semester)
        q = session.query(ActivityRegistration).filter(
            ActivityRegistration.school_year == sy,
            ActivityRegistration.semester == sem,
        )
        pending_cond = and_(
            ActivityRegistration.pending_review.is_(True),
            ActivityRegistration.is_active.is_(True),
        )
        rejected_cond = and_(
            ActivityRegistration.match_status == "rejected",
            ActivityRegistration.is_active.is_(False),
        )
        if status == "pending":
            q = q.filter(pending_cond)
        elif status == "rejected":
            q = q.filter(rejected_cond)
        else:
            q = q.filter(or_(pending_cond, rejected_cond))
        if search:
            like = f"%{search}%"
            q = q.filter(
                or_(
                    ActivityRegistration.student_name.ilike(like),
                    ActivityRegistration.class_name.ilike(like),
                    ActivityRegistration.parent_phone.ilike(like),
                )
            )
        total = q.count()
        # 合併頁：待審核排前（created_at 倒序），已拒絕排後（reviewed_at 倒序）
        rows = (
            q.order_by(
                ActivityRegistration.is_active.desc(),
                ActivityRegistration.created_at.desc(),
            )
            .offset(skip)
            .limit(limit)
            .all()
        )
        # F-026：缺 STUDENTS_READ / GUARDIANS_READ 時遮罩對應 PII 欄位
        can_see_student = can_view_student_pii(current_user)
        can_see_guardian = can_view_guardian_pii(current_user)
        return {
            "items": [
                _serialize_pending_item(
                    r,
                    can_see_student_pii=can_see_student,
                    can_see_guardian_pii=can_see_guardian,
                )
                for r in rows
            ],
            "total": total,
            "skip": skip,
            "limit": limit,
            "school_year": sy,
            "semester": sem,
            "status": status,
        }
    finally:
        session.close()


@router.get("/students/search")
async def admin_search_students(
    q: str = Query(..., min_length=1, max_length=50),
    limit: int = Query(20, ge=1, le=50),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台審核用：依姓名/學號/家長手機模糊搜尋在籍學生。

    F-027：搜尋結果含 student_id（學號）/ birthday / parent_phone 等學生 PII，
    必須額外要求 STUDENTS_READ 權限；缺則 403（不採欄位遮罩，因搜尋結果無
    PII 即無辨識力）。
    """
    # F-027：缺 STUDENTS_READ 直接 403（避免「ACTIVITY_WRITE 拉學生目錄」側信道）
    if not can_view_student_pii(current_user):
        raise HTTPException(status_code=403, detail="缺少學生資料讀取權限")

    from models.database import Student, Classroom
    from sqlalchemy import or_

    session = get_session()
    try:
        like = f"%{q.strip()}%"
        rows = (
            session.query(Student, Classroom)
            .outerjoin(Classroom, Classroom.id == Student.classroom_id)
            .filter(
                Student.is_active.is_(True),
                or_(
                    Student.name.ilike(like),
                    Student.student_id.ilike(like),
                    Student.parent_phone.ilike(like),
                    Student.emergency_contact_phone.ilike(like),
                ),
            )
            .limit(limit)
            .all()
        )
        return {
            "items": [
                {
                    "id": s.id,
                    "student_id": s.student_id,
                    "name": s.name,
                    "birthday": s.birthday.isoformat() if s.birthday else None,
                    "classroom_id": s.classroom_id,
                    "classroom_name": c.name if c else None,
                    "parent_phone": s.parent_phone,
                }
                for s, c in rows
            ]
        }
    finally:
        session.close()


@router.post("/registrations/{registration_id}/match")
async def match_registration(
    registration_id: int,
    body: RegistrationMatchRequest,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台手動將待審核 registration 綁定到指定 student_id。"""
    from models.database import Student, Classroom

    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
                ActivityRegistration.pending_review.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise HTTPException(
                status_code=409,
                detail="該筆報名已不在待審核佇列（可能已被其他人處理）",
            )

        student = (
            session.query(Student)
            .filter(Student.id == body.student_id, Student.is_active.is_(True))
            .first()
        )
        if not student:
            raise HTTPException(status_code=400, detail="找不到啟用中的學生")

        classroom = None
        if student.classroom_id:
            classroom = (
                session.query(Classroom)
                .filter(Classroom.id == student.classroom_id)
                .first()
            )

        reg.student_id = student.id
        reg.classroom_id = student.classroom_id
        if classroom:
            reg.class_name = classroom.name
        reg.pending_review = False
        reg.match_status = "manual"
        reg.reviewed_by = current_user.get("username")
        reg.reviewed_at = datetime.now()
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.info(
            "後台手動匹配報名：reg_id=%s → student_id=%s by %s",
            reg.id,
            student.id,
            current_user.get("username"),
        )
        return {"message": "已完成手動匹配", "registration_id": reg.id}
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("手動匹配失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/reject")
async def reject_registration(
    registration_id: int,
    body: RegistrationRejectRequest,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台將待審核 registration 視為校外生/資料不符拒絕。

    軟刪除（is_active=False）+ match_status='rejected' + remark 加註原因。
    """
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise HTTPException(
                status_code=409,
                detail="該筆報名已不存在或已被處理",
            )

        reg.is_active = False
        reg.match_status = "rejected"
        reg.pending_review = False
        # Phase 3：null 掉查詢碼 hash — 即使後續有人手動把 is_active 改回 True,
        # 舊 token 也無法用來打 /public/query-by-token（hash 比對不上 None）。
        # rejected 的 reg 沒有新 token 要發給誰，直接 invalidate 即可。
        reg.query_token_hash = None
        reg.reviewed_by = current_user.get("username")
        reg.reviewed_at = datetime.now()
        reason = body.reason  # validator 已保證非空且已 strip
        prefix = (reg.remark or "").strip()
        note = f"[已拒絕 by {reg.reviewed_by}] {reason}"
        reg.remark = (prefix + "\n" + note).strip() if prefix else note
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "拒絕報名",
            f"拒絕原因：{reason}",
            reg.reviewed_by or "",
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.warning(
            "後台拒絕報名：reg_id=%s by %s reason=%s",
            reg.id,
            current_user.get("username"),
            reason,
        )
        return {"message": "已拒絕該筆報名", "registration_id": reg.id}
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("拒絕報名失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/rematch")
async def rematch_registration(
    registration_id: int,
    body: Optional[RegistrationRematchRequest] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台重跑三欄比對（可同時修正 name/birthday/parent_phone）。

    body 任一欄位非 None 時先寫回 registration，再用新值跑比對。
    即使比對仍失敗，編輯的欄位也會保留，避免校方白打一次。
    """
    from models.database import Classroom
    from ._shared import _match_student_with_parent_phone

    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        new_name = reg.student_name
        new_birthday = reg.birthday
        new_phone = reg.parent_phone
        field_changed = False
        if body is not None:
            if body.name is not None and body.name != reg.student_name:
                new_name = body.name
                field_changed = True
            if body.birthday is not None and body.birthday != reg.birthday:
                new_birthday = body.birthday
                field_changed = True
            if body.parent_phone is not None and body.parent_phone != reg.parent_phone:
                new_phone = body.parent_phone
                field_changed = True

        # 若 name/birthday 有變，檢查同學期是否已有另一筆有效報名會重複
        if field_changed and (
            new_name != reg.student_name or new_birthday != reg.birthday
        ):
            dup = (
                session.query(ActivityRegistration)
                .filter(
                    ActivityRegistration.id != reg.id,
                    ActivityRegistration.student_name == new_name,
                    ActivityRegistration.birthday == new_birthday,
                    ActivityRegistration.school_year == reg.school_year,
                    ActivityRegistration.semester == reg.semester,
                    ActivityRegistration.is_active.is_(True),
                )
                .first()
            )
            if dup:
                raise HTTPException(
                    status_code=400,
                    detail="修改後的姓名/生日與本學期另一筆有效報名重複",
                )

        reg.student_name = new_name
        reg.birthday = new_birthday
        reg.parent_phone = new_phone

        sid, cid = _match_student_with_parent_phone(
            session, reg.student_name, reg.birthday, reg.parent_phone
        )
        matched = False
        if sid and cid:
            classroom = (
                session.query(Classroom)
                .filter(
                    Classroom.id == cid,
                    Classroom.is_active.is_(True),
                )
                .first()
            )
            if classroom:
                reg.student_id = sid
                reg.classroom_id = cid
                reg.class_name = classroom.name
                reg.pending_review = False
                reg.match_status = "matched"
                reg.reviewed_by = current_user.get("username")
                reg.reviewed_at = datetime.now()
                matched = True

        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.info(
            "後台重新比對：reg_id=%s matched=%s fields_edited=%s by %s",
            reg.id,
            matched,
            field_changed,
            current_user.get("username"),
        )
        if matched:
            msg = "重新比對成功"
        elif field_changed:
            msg = "仍無符合的在校生，已保留修改後的資料"
        else:
            msg = "仍無符合的在校生，請手動處理"
        return {
            "message": msg,
            "matched": matched,
            "field_changed": field_changed,
            "registration_id": reg.id,
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("重新比對失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/force-accept")
async def force_accept_registration(
    registration_id: int,
    body: Optional[RegistrationRematchRequest] = None,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """跳過三欄比對，強行將報名插入正式課後才藝報名管理並加上 `forced` 標記。

    body 與 rematch 相同三欄可選：校方可同時修正家長打錯的 name/birthday/phone。
    用途：家長是校外生或資料永遠比對不上，但校方決定收這筆報名。
    """
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        new_name = reg.student_name
        new_birthday = reg.birthday
        new_phone = reg.parent_phone
        field_changed = False
        if body is not None:
            if body.name is not None and body.name != reg.student_name:
                new_name = body.name
                field_changed = True
            if body.birthday is not None and body.birthday != reg.birthday:
                new_birthday = body.birthday
                field_changed = True
            if body.parent_phone is not None and body.parent_phone != reg.parent_phone:
                new_phone = body.parent_phone
                field_changed = True

        if field_changed and (
            new_name != reg.student_name or new_birthday != reg.birthday
        ):
            dup = (
                session.query(ActivityRegistration)
                .filter(
                    ActivityRegistration.id != reg.id,
                    ActivityRegistration.student_name == new_name,
                    ActivityRegistration.birthday == new_birthday,
                    ActivityRegistration.school_year == reg.school_year,
                    ActivityRegistration.semester == reg.semester,
                    ActivityRegistration.is_active.is_(True),
                )
                .first()
            )
            if dup:
                raise HTTPException(
                    status_code=400,
                    detail="修改後的姓名/生日與本學期另一筆有效報名重複",
                )

        reg.student_name = new_name
        reg.birthday = new_birthday
        reg.parent_phone = new_phone
        reg.pending_review = False
        reg.match_status = "forced"
        reg.reviewed_by = current_user.get("username")
        reg.reviewed_at = datetime.now()
        prefix = (reg.remark or "").strip()
        note = f"[強行收件 by {reg.reviewed_by}]"
        if prefix and "[強行收件" not in prefix:
            reg.remark = prefix + "\n" + note
        elif not prefix:
            reg.remark = note
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.warning(
            "後台強行收件報名：reg_id=%s by %s field_changed=%s",
            reg.id,
            current_user.get("username"),
            field_changed,
        )
        return {
            "message": "已強行收件並標記 forced",
            "matched": False,
            "forced": True,
            "field_changed": field_changed,
            "registration_id": reg.id,
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("強行收件失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/restore")
async def restore_registration(
    registration_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台將已拒絕（軟刪除）的報名復原回待審核狀態。

    僅限 match_status='rejected' 且 is_active=False 的報名。
    復原後 is_active=True、match_status='pending'、pending_review=True，
    保留原拒絕人/時間於 remark 作為歷史軌跡。
    """
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(ActivityRegistration.id == registration_id)
            .first()
        )
        if not reg:
            raise _not_found("報名資料")
        if reg.match_status != "rejected" or reg.is_active:
            raise HTTPException(
                status_code=400, detail="此筆報名非已拒絕狀態，無法復原"
            )

        # 若本學期已有同姓名/生日的有效報名，擋下避免唯一性衝突
        dup = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id != reg.id,
                ActivityRegistration.student_name == reg.student_name,
                ActivityRegistration.birthday == reg.birthday,
                ActivityRegistration.school_year == reg.school_year,
                ActivityRegistration.semester == reg.semester,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if dup:
            raise HTTPException(
                status_code=400,
                detail="本學期已有同姓名/生日的有效報名，無法復原此筆",
            )

        reg.is_active = True
        reg.match_status = "pending"
        reg.pending_review = True
        prefix = (reg.remark or "").strip()
        note = f"[已還原 by {current_user.get('username')}]"
        reg.remark = (prefix + "\n" + note).strip() if prefix else note
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.info(
            "後台還原拒絕報名：reg_id=%s by %s",
            reg.id,
            current_user.get("username"),
        )
        return {"message": "已還原報名至待審核", "registration_id": reg.id}
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        logger.error("還原報名失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


# ── 動態路由 /{registration_id}/... ─────────────────────────────────────────


@router.get("/registrations")
async def get_registrations(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    payment_status: Optional[str] = None,
    course_id: Optional[int] = None,
    classroom_name: Optional[str] = None,
    school_year: Optional[int] = Query(None, ge=100, le=200),
    semester: Optional[int] = Query(None, ge=1, le=2),
    match_status: Optional[str] = Query(
        None, pattern="^(matched|pending|manual|rejected|unmatched)$"
    ),
    include_inactive: bool = Query(False),
    student_id: Optional[int] = Query(
        None, gt=0, description="指定在校學生 ID，查詢其歷史報名紀錄"
    ),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得報名列表（分頁、搜尋、付款狀態、課程、班級、學期、匹配狀態篩選）。

    school_year 與 semester：可同時給或同時不給（不給則預設當前學期）。
    match_status：篩選自動匹配/手動綁定/待審核/拒絕等狀態。
    include_inactive：列出 rejected（軟刪除）的 registration 時需設 True。
    student_id：查詢單一學生的歷史報名紀錄（跨學期）；提供時通常搭配
      school_year=None、semester=None 才能看全部學期。
    """
    from ._shared import _build_registration_filter_query as _q_builder
    from utils.academic import resolve_academic_term_filters

    session = get_session()
    try:
        # 指定學生查全部學期時，顯式傳 None 避免被預設學期覆蓋
        if student_id is not None and school_year is None and semester is None:
            sy, sem = None, None
        else:
            sy, sem = resolve_academic_term_filters(school_year, semester)
        q = _q_builder(
            session,
            search=search,
            payment_status=payment_status,
            course_id=course_id,
            classroom_name=classroom_name,
            school_year=sy,
            semester=sem,
            match_status=match_status,
            include_inactive=include_inactive,
            student_id=student_id,
        )
        total = q.count()
        regs = (
            q.order_by(ActivityRegistration.created_at.desc())
            .offset(skip)
            .limit(limit)
            .all()
        )
        reg_ids = [r.id for r in regs]

        course_count_map = {}
        supply_count_map = {}
        course_name_map: dict[int, list[str]] = defaultdict(list)
        course_amount_map = {}
        supply_amount_map = {}

        if reg_ids:
            # 查詢一：course_stats — 一次撈出所有課程資訊，Python 端同時建立 3 個 map
            course_stats = (
                session.query(
                    RegistrationCourse.registration_id,
                    RegistrationCourse.status,
                    ActivityCourse.name,
                    RegistrationCourse.price_snapshot,
                )
                .join(ActivityCourse, RegistrationCourse.course_id == ActivityCourse.id)
                .filter(RegistrationCourse.registration_id.in_(reg_ids))
                .all()
            )
            _course_count: dict = defaultdict(int)
            _course_amount: dict = defaultdict(int)
            for registration_id, status, course_name, price_snapshot in course_stats:
                _course_count[registration_id] += 1
                course_name_map[registration_id].append(
                    f"{course_name}（候補）" if status == "waitlist" else course_name
                )
                if status == "enrolled":
                    _course_amount[registration_id] += price_snapshot or 0
            course_count_map = dict(_course_count)
            course_amount_map = dict(_course_amount)

            # 查詢二：supply_stats — 一次撈出所有用品資訊，Python 端同時建立 2 個 map
            supply_stats = (
                session.query(
                    RegistrationSupply.registration_id,
                    RegistrationSupply.price_snapshot,
                )
                .filter(RegistrationSupply.registration_id.in_(reg_ids))
                .all()
            )
            _supply_count: dict = defaultdict(int)
            _supply_amount: dict = defaultdict(int)
            for registration_id, price_snapshot in supply_stats:
                _supply_count[registration_id] += 1
                _supply_amount[registration_id] += price_snapshot or 0
            supply_count_map = dict(_supply_count)
            supply_amount_map = dict(_supply_amount)

        # F-026：缺 STUDENTS_READ / GUARDIANS_READ 時遮罩對應 PII
        can_see_student = can_view_student_pii(current_user)
        can_see_guardian = can_view_guardian_pii(current_user)

        items = []
        for r in regs:
            paid_amount = r.paid_amount or 0
            total_amount = (course_amount_map.get(r.id, 0) or 0) + (
                supply_amount_map.get(r.id, 0) or 0
            )
            items.append(
                {
                    "id": r.id,
                    "student_name": r.student_name,
                    "student_id": r.student_id if can_see_student else None,
                    "birthday": r.birthday if can_see_student else None,
                    "class_name": r.class_name,
                    "classroom_id": r.classroom_id if can_see_student else None,
                    "parent_phone": r.parent_phone if can_see_guardian else None,
                    "match_status": r.match_status,
                    "pending_review": r.pending_review,
                    "is_active": r.is_active,
                    "email": r.email if can_see_guardian else None,
                    "is_paid": r.is_paid,
                    "paid_amount": paid_amount,
                    "total_amount": total_amount,
                    "payment_status": _derive_payment_status(paid_amount, total_amount),
                    "remark": r.remark or "",
                    "school_year": r.school_year,
                    "semester": r.semester,
                    "course_count": course_count_map.get(r.id, 0),
                    "supply_count": supply_count_map.get(r.id, 0),
                    "course_names": "、".join(course_name_map.get(r.id, [])),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                    "reviewed_by": r.reviewed_by,
                    "reviewed_at": (
                        r.reviewed_at.isoformat() if r.reviewed_at else None
                    ),
                }
            )
        return {
            "items": items,
            "total": total,
            "skip": skip,
            "limit": limit,
            "school_year": sy,
            "semester": sem,
        }
    finally:
        session.close()


@router.get("/registrations/{registration_id}")
async def get_registration_detail(
    registration_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得報名詳情（含課程/用品/修改紀錄）"""
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        rc_rows = (
            session.query(RegistrationCourse, ActivityCourse)
            .join(ActivityCourse, RegistrationCourse.course_id == ActivityCourse.id)
            .filter(RegistrationCourse.registration_id == registration_id)
            .all()
        )
        courses = [
            {
                "id": rc.id,
                "course_id": ac.id,
                "name": ac.name,
                "price": rc.price_snapshot,
                "status": rc.status,
                "confirm_deadline": (
                    rc.confirm_deadline.isoformat()
                    if rc.status == "promoted_pending" and rc.confirm_deadline
                    else None
                ),
            }
            for rc, ac in rc_rows
        ]

        rs_rows = (
            session.query(RegistrationSupply, ActivitySupply)
            .join(ActivitySupply, RegistrationSupply.supply_id == ActivitySupply.id)
            .filter(RegistrationSupply.registration_id == registration_id)
            .all()
        )
        supplies = [
            {
                "id": rs.id,
                "supply_id": sp.id,
                "name": sp.name,
                "price": rs.price_snapshot,
            }
            for rs, sp in rs_rows
        ]

        changes = (
            session.query(RegistrationChange)
            .filter(RegistrationChange.registration_id == registration_id)
            .order_by(RegistrationChange.created_at.desc())
            .limit(20)
            .all()
        )
        change_list = [
            {
                "id": ch.id,
                "change_type": ch.change_type,
                "description": ch.description,
                "changed_by": ch.changed_by,
                "created_at": ch.created_at.isoformat() if ch.created_at else None,
            }
            for ch in changes
        ]

        total_amount = sum(c["price"] for c in courses if c["status"] == "enrolled")
        total_amount += sum(s["price"] for s in supplies)
        paid_amount = reg.paid_amount or 0

        # F-026：缺 STUDENTS_READ / GUARDIANS_READ 時遮罩對應 PII
        can_see_student = can_view_student_pii(current_user)
        can_see_guardian = can_view_guardian_pii(current_user)
        return {
            "id": reg.id,
            "student_name": reg.student_name,
            "student_id": reg.student_id if can_see_student else None,
            "birthday": reg.birthday if can_see_student else None,
            "class_name": reg.class_name,
            "classroom_id": reg.classroom_id if can_see_student else None,
            "parent_phone": reg.parent_phone if can_see_guardian else None,
            "match_status": reg.match_status,
            "pending_review": reg.pending_review,
            "reviewed_by": reg.reviewed_by,
            "reviewed_at": reg.reviewed_at.isoformat() if reg.reviewed_at else None,
            "email": reg.email if can_see_guardian else None,
            "is_paid": reg.is_paid,
            "paid_amount": paid_amount,
            "payment_status": _derive_payment_status(paid_amount, total_amount),
            "remark": reg.remark or "",
            "courses": courses,
            "supplies": supplies,
            "changes": change_list,
            "total_amount": total_amount,
            "created_at": reg.created_at.isoformat() if reg.created_at else None,
            "updated_at": reg.updated_at.isoformat() if reg.updated_at else None,
        }
    finally:
        session.close()


@router.put("/registrations/{registration_id}/payment")
async def update_payment(
    registration_id: int,
    body: PaymentUpdate,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """更新付款狀態

    併發保護：鎖 reg 行，避免與 POS checkout / add_registration_payment 並發
    造成 lost update（POS 寫入的 paid_amount 可能被此處 set 覆寫）。
    """
    session = get_session()
    try:
        reg = _lock_registration(session, registration_id)
        if not reg:
            raise _not_found("報名資料")

        total_amount = _calc_total_amount(session, registration_id)
        operator = current_user.get("username", "")

        today = today_taipei()
        # 今日若已簽核，任何補齊/沖帳都會讓 snapshot 失準，先擋
        _require_daily_close_unlocked(session, today)
        if body.is_paid:
            if not reg.is_paid:
                shortfall = total_amount - (reg.paid_amount or 0)
                if shortfall > 0:
                    # ── 補齊欠費守衛 ──────────────────────────────────────
                    # 原設計直接寫「系統補齊」payment 補上欠費，無 method/原因/簽核，
                    # 會計可逐筆把欠費轉成收入流水。對齊 is_paid=False 嚴格度：
                    # 1. 必填人工 payment_method（拒絕 SYSTEM_RECONCILE_METHOD）
                    # 2. 必填 ≥5 字 payment_reason
                    # 3. shortfall 過 FINANCE_APPROVAL_THRESHOLD 需金流簽核
                    method_cleaned = (body.payment_method or "").strip()
                    if not method_cleaned:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"標記已繳費需補齊 NT${shortfall} 欠費，"
                                "請於 payment_method 填入「現金」"
                                "（目前才藝僅收現金），不接受系統補齊"
                            ),
                        )
                    if method_cleaned == SYSTEM_RECONCILE_METHOD:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"payment_method 不可填入「{SYSTEM_RECONCILE_METHOD}」，"
                                "請填寫實際收款方式以利稽核"
                            ),
                        )
                    reason_cleaned = (body.payment_reason or "").strip()
                    if len(reason_cleaned) < MIN_REFUND_REASON_LENGTH:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"標記已繳費需補齊 NT${shortfall}，"
                                f"請於 payment_reason 填寫原因（≥ {MIN_REFUND_REASON_LENGTH} 字）"
                            ),
                        )
                    require_finance_approve(
                        shortfall,
                        current_user,
                        threshold=FINANCE_APPROVAL_THRESHOLD,
                        action_label="補齊欠費金額",
                    )
                    rec = ActivityPaymentRecord(
                        registration_id=registration_id,
                        type="payment",
                        amount=shortfall,
                        payment_date=today,
                        payment_method=method_cleaned,
                        notes=f"（標記已繳費補齊）方式：{method_cleaned}；原因：{reason_cleaned}",
                        operator=operator,
                    )
                    session.add(rec)
                    reg.paid_amount = total_amount
                reg.is_paid = _compute_is_paid(reg.paid_amount or 0, total_amount)
        else:
            # is_paid=False：一刀切全額沖帳會誤殺部分繳費者，收緊為「必須帶
            # confirm_refund_amount == current_paid 且 refund_reason ≥ 5 字」。
            current_paid = reg.paid_amount or 0
            if body.confirm_refund_amount is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"標記未繳費將沖帳全額已繳 NT${current_paid}，"
                        "請於 confirm_refund_amount 明確填寫同金額以二次確認"
                    ),
                )
            if body.confirm_refund_amount != current_paid:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"confirm_refund_amount NT${body.confirm_refund_amount} "
                        f"與當前已繳 NT${current_paid} 不符，請重新確認"
                    ),
                )
            reason_cleaned = require_refund_reason(body.refund_reason)
            # 大額沖帳需簽核權限（以「該 reg 累積退費 + 本次」判斷，封拆單繞過）
            require_approve_for_cumulative_refund(
                session,
                registration_id,
                current_paid,
                current_user,
                label="標記未繳費自動沖帳累積退費總額",
            )
            if current_paid > 0:
                rec = ActivityPaymentRecord(
                    registration_id=registration_id,
                    type="refund",
                    amount=current_paid,
                    payment_date=today,
                    payment_method=SYSTEM_RECONCILE_METHOD,
                    notes=f"（標記未繳費自動沖帳）原因：{reason_cleaned}",
                    operator=operator,
                )
                session.add(rec)
            reg.paid_amount = 0
            reg.is_paid = False

        status_str = "已繳費" if body.is_paid else "未繳費"
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "更新付款狀態",
            f"付款狀態更新為：{status_str}",
            operator,
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        _invalidate_finance_summary_cache()
        request.state.audit_summary = f"更新繳費狀態：{reg.student_name} → {status_str}"
        request.state.audit_changes = {
            "student_name": reg.student_name,
            "new_is_paid": bool(body.is_paid),
            "paid_amount_after": reg.paid_amount,
            "total_amount": total_amount,
        }
        return {"message": f"更新成功，狀態為：{status_str}"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/registrations/{registration_id}/remark")
async def update_remark(
    registration_id: int,
    body: RemarkUpdate,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """更新備註"""
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        reg.remark = body.remark
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "更新備註",
            f"備註更新為：{body.remark}",
            current_user.get("username", ""),
        )
        session.commit()
        return {"message": "備註更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/registrations/{registration_id}")
async def update_registration_basic(
    registration_id: int,
    body: AdminRegistrationBasicUpdate,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台編輯報名基本欄位（姓名、生日、班級、Email）。
    學期不可變更，若需更改請重新建立報名。"""
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        classroom = _require_active_classroom(session, body.class_)

        # 同學期內姓名+生日不得重複於另一筆
        dup = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id != registration_id,
                ActivityRegistration.student_name == body.name,
                ActivityRegistration.birthday == body.birthday,
                ActivityRegistration.school_year == reg.school_year,
                ActivityRegistration.semester == reg.semester,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if dup:
            raise HTTPException(
                status_code=400, detail="本學期已有另一筆相同姓名與生日的報名"
            )

        diffs: list[str] = []
        if reg.student_name != body.name:
            diffs.append(f"姓名：{reg.student_name} → {body.name}")
            reg.student_name = body.name
        if (reg.birthday or "") != body.birthday:
            diffs.append(f"生日：{reg.birthday or '—'} → {body.birthday}")
            reg.birthday = body.birthday
        if (reg.class_name or "") != classroom.name:
            diffs.append(f"班級：{reg.class_name or '—'} → {classroom.name}")
            reg.class_name = classroom.name
        new_email = body.email or None
        if (reg.email or None) != new_email:
            diffs.append(f"Email：{reg.email or '—'} → {new_email or '—'}")
            reg.email = new_email

        # 姓名+生日變更時重新匹配 student_id
        if any(d.startswith("姓名") or d.startswith("生日") for d in diffs):
            reg.student_id = _match_student_id(session, body.name, body.birthday)

        if diffs:
            activity_service.log_change(
                session,
                registration_id,
                reg.student_name,
                "編輯基本資料",
                "；".join(diffs),
                current_user.get("username", ""),
            )
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        return {"message": "基本資料更新成功", "changed": len(diffs)}
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/courses", status_code=201)
async def add_registration_course(
    registration_id: int,
    body: AddCourseRequest,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台為既有報名追加一筆課程（額滿時自動候補）。

    併發保護：鎖 reg 行，與 remove_registration_supply 對稱，避免與
    POS checkout / update_payment 並發時 is_paid 旗標短暫錯誤。
    """
    session = get_session()
    try:
        reg = _lock_registration(session, registration_id)
        if not reg:
            raise _not_found("報名資料")

        course = (
            session.query(ActivityCourse)
            .filter(
                ActivityCourse.id == body.course_id,
                ActivityCourse.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not course:
            raise _not_found("課程")

        # 不允許跨學期追加
        if course.school_year != reg.school_year or course.semester != reg.semester:
            raise HTTPException(status_code=400, detail="課程學期與報名學期不一致")

        # 不允許重複報名同課程
        exists = (
            session.query(RegistrationCourse)
            .filter(
                RegistrationCourse.registration_id == registration_id,
                RegistrationCourse.course_id == course.id,
            )
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="此報名已含該課程")

        # 佔容量 = enrolled + promoted_pending
        enrolled_count = (
            session.query(func.count(RegistrationCourse.id))
            .join(
                ActivityRegistration,
                RegistrationCourse.registration_id == ActivityRegistration.id,
            )
            .filter(
                RegistrationCourse.course_id == course.id,
                RegistrationCourse.status.in_(["enrolled", "promoted_pending"]),
                ActivityRegistration.is_active.is_(True),
            )
            .scalar()
            or 0
        )
        capacity = course.capacity if course.capacity is not None else 30
        if enrolled_count < capacity:
            status = "enrolled"
        elif course.allow_waitlist:
            status = "waitlist"
        else:
            raise HTTPException(
                status_code=400, detail=f"課程「{course.name}」已額滿且不開放候補"
            )

        paid_amount = reg.paid_amount or 0
        before_total = _calc_total_amount(session, registration_id)

        rc = RegistrationCourse(
            registration_id=registration_id,
            course_id=course.id,
            status=status,
            price_snapshot=course.price,
        )
        session.add(rc)
        session.flush()

        # 新增課程後可能把原本已繳清改為部分繳費；算出欠款供管理員即時追繳
        total_amount = _calc_total_amount(session, registration_id)
        reg.is_paid = _compute_is_paid(paid_amount, total_amount)
        outstanding_amount = max(0, total_amount - paid_amount)
        debt_delta = max(0, (total_amount - paid_amount) - (before_total - paid_amount))

        label = "候補" if status == "waitlist" else "正式"
        log_detail = f"課程「{course.name}」（{label}，價 ${course.price}）"
        if debt_delta > 0:
            log_detail += (
                f"，產生欠款 NT${debt_delta}（累計欠款 NT${outstanding_amount}）"
            )
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "新增課程",
            log_detail,
            current_user.get("username", ""),
        )

        session.commit()
        _invalidate_activity_dashboard_caches(session)
        return {
            "message": "課程新增成功" + ("（候補）" if status == "waitlist" else ""),
            "status": status,
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding_amount,
            "payment_status": _derive_payment_status(paid_amount, total_amount),
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/registrations/{registration_id}/supplies", status_code=201)
async def add_registration_supply(
    registration_id: int,
    body: AddSupplyRequest,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台為既有報名追加一筆用品。

    併發保護：鎖 reg 行，與 remove_registration_supply 對稱。
    """
    session = get_session()
    try:
        reg = _lock_registration(session, registration_id)
        if not reg:
            raise _not_found("報名資料")

        supply = (
            session.query(ActivitySupply)
            .filter(
                ActivitySupply.id == body.supply_id,
                ActivitySupply.is_active.is_(True),
            )
            .first()
        )
        if not supply:
            raise _not_found("用品")

        if supply.school_year != reg.school_year or supply.semester != reg.semester:
            raise HTTPException(status_code=400, detail="用品學期與報名學期不一致")

        paid_amount = reg.paid_amount or 0
        before_total = _calc_total_amount(session, registration_id)

        rs = RegistrationSupply(
            registration_id=registration_id,
            supply_id=supply.id,
            price_snapshot=supply.price,
        )
        session.add(rs)
        session.flush()

        total_amount = _calc_total_amount(session, registration_id)
        reg.is_paid = _compute_is_paid(paid_amount, total_amount)
        outstanding_amount = max(0, total_amount - paid_amount)
        debt_delta = max(0, (total_amount - paid_amount) - (before_total - paid_amount))

        log_detail = f"用品「{supply.name}」（價 ${supply.price}）"
        if debt_delta > 0:
            log_detail += (
                f"，產生欠款 NT${debt_delta}（累計欠款 NT${outstanding_amount}）"
            )
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "新增用品",
            log_detail,
            current_user.get("username", ""),
        )

        session.commit()
        _invalidate_activity_dashboard_caches(session)
        return {
            "message": "用品新增成功",
            "id": rs.id,
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding_amount,
            "payment_status": _derive_payment_status(paid_amount, total_amount),
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/registrations/{registration_id}/supplies/{supply_record_id}")
async def remove_registration_supply(
    registration_id: int,
    supply_record_id: int,
    force_refund: bool = Query(
        False,
        description="移除用品後若出現超繳，需顯式帶 true 才允許移除並自動寫退費沖帳紀錄",
    ),
    refund_reason: Optional[str] = Query(
        None,
        description="當 force_refund 觸發實際退費時必填（≥5 字），原因會寫入 notes 供稽核",
    ),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """後台移除已報名的單筆用品。

    併發保護：鎖住 reg 行，避免與其他端點（結帳、退課）同時改 is_paid。
    與 withdraw_course 對稱：若移除後 paid_amount > new_total，須顯式 force_refund。
    """
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        rs = (
            session.query(RegistrationSupply)
            .filter(
                RegistrationSupply.id == supply_record_id,
                RegistrationSupply.registration_id == registration_id,
            )
            .first()
        )
        if not rs:
            raise _not_found("用品記錄")

        supply = (
            session.query(ActivitySupply)
            .filter(ActivitySupply.id == rs.supply_id)
            .first()
        )
        supply_name = supply.name if supply else str(rs.supply_id)

        paid_amount = reg.paid_amount or 0
        before_total = _calc_total_amount(session, registration_id)
        # 估算移除後的應繳；真正的 refund_needed 在 flush 後用 new_total 重算
        estimated_after_total = before_total - int(rs.price_snapshot or 0)
        preview_refund = max(0, paid_amount - estimated_after_total)

        if preview_refund > 0 and not force_refund:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"移除用品後將產生超繳 NT${preview_refund}（已繳 NT${paid_amount}、"
                    f"移除後應繳 NT${estimated_after_total}），請先處理退費或於移除時"
                    f"指定 force_refund=true 自動沖帳"
                ),
            )

        # 與正式退費端點同套守衛：fail-fast，避免 ACTIVITY_WRITE 經自動沖帳繞過
        # require_refund_reason / require_approve_for_large_refund 簽核閾值
        cleaned_reason: Optional[str] = None
        if preview_refund > 0 and force_refund:
            cleaned_reason = require_refund_reason(refund_reason)
            require_approve_for_cumulative_refund(
                session,
                registration_id,
                preview_refund,
                current_user,
                label="移除用品自動沖帳累積退費總額",
            )

        session.delete(rs)
        session.flush()

        new_total = _calc_total_amount(session, registration_id)
        refund_needed = max(0, paid_amount - new_total)

        if refund_needed > 0 and force_refund:
            today = today_taipei()
            _require_daily_close_unlocked(session, today)
            session.add(
                ActivityPaymentRecord(
                    registration_id=registration_id,
                    type="refund",
                    amount=refund_needed,
                    payment_date=today,
                    payment_method=SYSTEM_RECONCILE_METHOD,
                    notes=(
                        f"（移除用品「{supply_name}」自動沖帳）原因：{cleaned_reason}"
                        if cleaned_reason
                        else f"（移除用品「{supply_name}」自動沖帳）"
                    ),
                    operator=current_user.get("username", ""),
                )
            )
            reg.paid_amount = max(0, paid_amount - refund_needed)

        reg.is_paid = _compute_is_paid(reg.paid_amount or 0, new_total)

        log_detail = f"用品「{supply_name}」已移除"
        if refund_needed > 0 and force_refund:
            log_detail += f"（自動沖帳退費 NT${refund_needed}）"
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "移除用品",
            log_detail,
            current_user.get("username", ""),
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session)
        final_paid = reg.paid_amount or 0
        return {
            "message": f"已移除用品「{supply_name}」",
            "total_amount": new_total,
            "paid_amount": final_paid,
            "refunded_amount": refund_needed if force_refund else 0,
            "payment_status": _derive_payment_status(final_paid, new_total),
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


def _desensitize_operator(operator: Optional[str], viewer_has_approve: bool) -> str:
    """對 operator 欄位去敏化：非簽核權限者只看得到首字 + ***。

    Why: 員工帳號暴露給過廣的閱讀者（ACTIVITY_READ）等同於社工輔助材料；
    但對於能執行簽核的主管/老闆仍需看完整帳號以便對帳追責。
    """
    if not operator:
        return ""
    if viewer_has_approve:
        return operator
    if operator == "system":
        return "system"
    # 保留首字，其餘遮蔽（例如 "fee_admin" → "f***"）
    return operator[0] + "***"


@router.get("/registrations/{registration_id}/payments")
async def get_registration_payments(
    registration_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_READ)),
):
    """取得報名的繳費／退費明細記錄（含 voided 軟刪紀錄，標示 is_voided）"""
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        records = (
            session.query(ActivityPaymentRecord)
            .filter(ActivityPaymentRecord.registration_id == registration_id)
            .order_by(ActivityPaymentRecord.created_at.asc())
            .all()
        )
        total_amount = _calc_total_amount(session, registration_id)
        paid_amount = reg.paid_amount or 0
        viewer_has_approve = has_payment_approve(current_user)
        return {
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "payment_status": _derive_payment_status(paid_amount, total_amount),
            "records": [
                {
                    "id": r.id,
                    "type": r.type,
                    "amount": r.amount,
                    "payment_date": (
                        r.payment_date.isoformat() if r.payment_date else None
                    ),
                    "payment_method": r.payment_method or "",
                    "notes": r.notes or "",
                    "operator": _desensitize_operator(r.operator, viewer_has_approve),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "is_voided": r.voided_at is not None,
                    "voided_at": (r.voided_at.isoformat() if r.voided_at else None),
                    "voided_by": _desensitize_operator(r.voided_by, viewer_has_approve),
                    "void_reason": r.void_reason or "",
                }
                for r in records
            ],
        }
    finally:
        session.close()


_IDEMPOTENCY_WINDOW_SECONDS = 600


def _lock_registration(session, registration_id: int):
    """對單筆 registration 取得行級鎖；SQLite（單元測試）自動降級為無鎖。"""
    query = session.query(ActivityRegistration).filter(
        ActivityRegistration.id == registration_id,
        ActivityRegistration.is_active.is_(True),
    )
    try:
        return query.with_for_update().first()
    except (CompileError, OperationalError, NotImplementedError):
        return query.first()


@router.post("/registrations/{registration_id}/payments", status_code=201)
async def add_registration_payment(
    registration_id: int,
    body: AddPaymentRequest,
    request: Request,
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """新增繳費或退費記錄

    冪等語意（2026-04-24 修正）：
    idempotency_key 於 DB 層永久全域唯一。同 key 必須對應同一 registration、
    同一 type、同一 amount；若上下文不符視為 key 誤用，回 409 避免錯帳到
    其他 registration（原本的 10 分鐘 window 過期後再重送會爆 500）。
    """
    session = get_session()
    try:
        # ── 冪等性重送檢查（先於任何寫入） ────────────────────────
        # 與 pos._find_idempotent_hit 對齊：排除 voided 紀錄。否則「key 命中但
        # 全 voided」會被當作合法 replay 回 200，但 DB 並無新紀錄、paid_amount
        # 反映 void 後（=0），員工以為已收實際永久漏收。Refs: 邏輯漏洞 audit
        # 2026-05-07 P0 (#7)。
        if body.idempotency_key:
            from .pos import _find_idempotent_hit, _has_any_record_for_key

            hit = _find_idempotent_hit(session, body.idempotency_key)
            if hit is None and _has_any_record_for_key(session, body.idempotency_key):
                # key 已用於 voided 紀錄；不可重複 replay 也不可作為新交易 key
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "idempotency_key 對應的紀錄已被作廢；請改用新 key "
                        "重新建立繳費/退費記錄"
                    ),
                )
            if hit is not None:
                # 上下文一致才 replay；不一致視為 key 誤用
                if (
                    hit.registration_id != registration_id
                    or hit.type != body.type
                    or hit.amount != body.amount
                ):
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"idempotency_key 已用於 registration {hit.registration_id} "
                            f"（{hit.type} NT${hit.amount}），不可重複用於本請求"
                        ),
                    )
                reg_hit = (
                    session.query(ActivityRegistration)
                    .filter(ActivityRegistration.id == hit.registration_id)
                    .first()
                )
                total_amount = _calc_total_amount(session, hit.registration_id)
                paid = (reg_hit.paid_amount if reg_hit else 0) or 0
                type_label = "繳費" if hit.type == "payment" else "退費"
                logger.info(
                    "add_registration_payment idempotent replay: key=%s reg=%s",
                    body.idempotency_key,
                    hit.registration_id,
                )
                return {
                    "message": f"{type_label}記錄新增成功",
                    "paid_amount": paid,
                    "payment_status": _derive_payment_status(paid, total_amount),
                }

        # 已簽核日守衛（payment_date 落在 daily-close 之日則拒絕）
        _require_daily_close_unlocked(session, body.payment_date)

        # ── 退費 reason 必填（schema 已強制；此處 cleaned 並覆寫）────
        # Pydantic 已在 schema 層強制 type=refund 時 notes ≥ MIN_REFUND_REASON_LENGTH；
        # 此處再檢一次（防 schema 日後被放寬）並處理 cleaned notes
        if body.type == "refund":
            cleaned_reason = require_refund_reason(body.notes)
            body.notes = cleaned_reason

        # 行級鎖住該 registration，防併發繳/退費 lost update
        reg = _lock_registration(session, registration_id)
        if not reg:
            raise _not_found("報名資料")

        # ── 累積退費簽核（必須在 _lock_registration 之後）─────────────
        # 鎖之後才查 prior_refunded，確保兩個併發小額退費不會各自看到相同舊累積值
        # 而各自通過簽核門檻；以同 registration 過去未作廢的退費 + 本次金額判斷，
        # 任一筆讓累積跨閾值即整筆需 ACTIVITY_PAYMENT_APPROVE。
        # Why: 舊版在 lock 前就算 prior_refunded，存在 race window；舊版亦有「只看本次
        # body.amount」的拆單問題。本次累積簽核同時封死兩條繞過路徑。
        if body.type == "refund":
            prior_refunded = (
                session.query(func.coalesce(func.sum(ActivityPaymentRecord.amount), 0))
                .filter(
                    ActivityPaymentRecord.registration_id == registration_id,
                    ActivityPaymentRecord.type == "refund",
                    ActivityPaymentRecord.voided_at.is_(None),
                )
                .scalar()
            ) or 0
            cumulative_refund = int(prior_refunded) + int(body.amount)
            require_approve_for_large_refund(
                cumulative_refund, current_user, label="活動累積退費總額"
            )

        operator = current_user.get("username", "")

        if body.type == "refund" and body.amount > (reg.paid_amount or 0):
            raise HTTPException(
                status_code=400,
                detail=f"退費金額 NT${body.amount} 超過已繳金額 NT${reg.paid_amount or 0}",
            )

        # 空報名守衛：與 POS checkout 對齊，避免對無應繳的殼報名寫入付款，產生孤兒金額。
        # 僅擋「空報名收款」，不擋超收（overpaid 是系統支援的四態之一，admin 可能需要手動處理）
        if body.type == "payment":
            current_total = _calc_total_amount(session, registration_id)
            if current_total <= 0:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"報名 {registration_id}（{reg.student_name}）無應繳金額，"
                        f"無法新增繳費記錄"
                    ),
                )

        rec = ActivityPaymentRecord(
            registration_id=registration_id,
            type=body.type,
            amount=body.amount,
            payment_date=body.payment_date,
            payment_method=body.payment_method,
            notes=body.notes,
            operator=operator,
            idempotency_key=body.idempotency_key,
        )
        session.add(rec)

        if body.type == "payment":
            reg.paid_amount = (reg.paid_amount or 0) + body.amount
        else:
            # max(0, ...) 防禦：即使驗證通過到執行之間狀態被搶改，也不會變負。
            reg.paid_amount = max(0, (reg.paid_amount or 0) - body.amount)

        total_amount = _calc_total_amount(session, registration_id)
        reg.is_paid = _compute_is_paid(reg.paid_amount or 0, total_amount)

        type_label = "繳費" if body.type == "payment" else "退費"
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            f"新增{type_label}記錄",
            f"{type_label} NT${body.amount}，繳費方式：{body.payment_method}",
            operator,
        )
        try:
            session.commit()
        except IntegrityError as e:
            # DB 層 UNIQUE 攔下並發同 idempotency_key 的第二筆：轉為 idempotent replay
            # 重要：必須驗證 (registration_id, type, amount) 一致，否則視為 key 誤用
            session.rollback()
            if body.idempotency_key and "idempotency_key" in str(e.orig).lower():
                hit = (
                    session.query(ActivityPaymentRecord)
                    .filter(
                        ActivityPaymentRecord.idempotency_key == body.idempotency_key
                    )
                    .order_by(ActivityPaymentRecord.id.asc())
                    .first()
                )
                if hit is not None:
                    if (
                        hit.registration_id != registration_id
                        or hit.type != body.type
                        or hit.amount != body.amount
                    ):
                        raise HTTPException(
                            status_code=409,
                            detail=(
                                f"idempotency_key 已用於 registration "
                                f"{hit.registration_id}（{hit.type} NT${hit.amount}），"
                                f"不可重複用於本請求"
                            ),
                        )
                    reg_hit = (
                        session.query(ActivityRegistration)
                        .filter(ActivityRegistration.id == hit.registration_id)
                        .first()
                    )
                    total_hit = _calc_total_amount(session, hit.registration_id)
                    paid_hit = (reg_hit.paid_amount if reg_hit else 0) or 0
                    type_label_hit = "繳費" if hit.type == "payment" else "退費"
                    logger.info(
                        "add_registration_payment idempotent replay via UNIQUE: key=%s reg=%s",
                        body.idempotency_key,
                        hit.registration_id,
                    )
                    return {
                        "message": f"{type_label_hit}記錄新增成功",
                        "paid_amount": paid_hit,
                        "payment_status": _derive_payment_status(paid_hit, total_hit),
                    }
            raise
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        _invalidate_finance_summary_cache()
        request.state.audit_summary = (
            f"新增{type_label}記錄：{reg.student_name} NT${body.amount}"
        )
        request.state.audit_changes = {
            "student_name": reg.student_name,
            "type": body.type,
            "amount": body.amount,
            "payment_method": body.payment_method,
            "payment_date": body.payment_date.isoformat(),
            "paid_amount_after": reg.paid_amount,
        }
        return {
            "message": f"{type_label}記錄新增成功",
            "paid_amount": reg.paid_amount,
            "payment_status": _derive_payment_status(reg.paid_amount, total_amount),
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/registrations/{registration_id}/payments/{payment_id}")
async def delete_registration_payment(
    registration_id: int,
    payment_id: int,
    request: Request,
    body: VoidPaymentRequest,
    current_user: dict = Depends(
        require_staff_permission(Permission.ACTIVITY_PAYMENT_APPROVE)
    ),
):
    """軟刪除（void）繳費記錄：原紀錄保留於 DB，設 voided_at/by/reason 並重算已繳金額。

    Why: 員工可能濫用「POS 收現金 → DELETE payment → paid_amount 歸零 → 私吞」；
    改軟刪後：
      1. 原 payment row 永不消失，稽核可追溯完整金流
      2. 需 ACTIVITY_PAYMENT_APPROVE（簽核權限）且強制填寫原因（≥5 字）
      3. paid_amount / daily snapshot 重算時以 voided_at IS NULL 為前提排除

    併發保護：鎖 reg 行，避免 GROUP BY 重算與 POS checkout 並發時，
    POS 的新付款在本端 commit 時被 paid_amount = 舊 sum 的 UPDATE 覆蓋（lost update）。
    """
    session = get_session()
    try:
        reg = _lock_registration(session, registration_id)
        if not reg:
            raise _not_found("報名資料")

        payment = (
            session.query(ActivityPaymentRecord)
            .filter(
                ActivityPaymentRecord.id == payment_id,
                ActivityPaymentRecord.registration_id == registration_id,
            )
            .first()
        )
        if not payment:
            raise _not_found("繳費記錄")

        # 已軟刪的紀錄不可重複 void，避免操作紀錄被洗成多次 void
        if payment.voided_at is not None:
            raise HTTPException(
                status_code=409,
                detail="此繳費記錄已於稍早被軟刪，不可重複操作",
            )

        # 若被刪除的付款日期已被日結簽核，拒絕刪除以免 snapshot 與 DB 失準
        _require_daily_close_unlocked(session, payment.payment_date)

        operator = current_user.get("username", "")
        now = datetime.now(TAIPEI_TZ).replace(tzinfo=None)
        payment.voided_at = now
        payment.voided_by = operator
        payment.void_reason = body.reason

        deleted_snapshot = {
            "type": payment.type,
            "amount": payment.amount,
            "payment_date": (
                payment.payment_date.isoformat() if payment.payment_date else None
            ),
        }

        session.flush()

        # 重算 paid_amount：以 voided_at IS NULL 為前提，排除軟刪紀錄
        totals = (
            session.query(
                ActivityPaymentRecord.type, func.sum(ActivityPaymentRecord.amount)
            )
            .filter(
                ActivityPaymentRecord.registration_id == registration_id,
                ActivityPaymentRecord.voided_at.is_(None),
            )
            .group_by(ActivityPaymentRecord.type)
            .all()
        )
        amount_map = {t: s for t, s in totals}
        new_paid = (amount_map.get("payment") or 0) - (amount_map.get("refund") or 0)
        reg.paid_amount = max(0, new_paid)

        total_amount = _calc_total_amount(session, registration_id)
        reg.is_paid = _compute_is_paid(reg.paid_amount or 0, total_amount)

        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "軟刪除繳費記錄",
            (
                f"void payment_id={payment_id}（{deleted_snapshot['type']} NT${deleted_snapshot['amount']}），"
                f"原因：{body.reason}，重新計算已繳 NT${reg.paid_amount}"
            ),
            operator,
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        _invalidate_finance_summary_cache()
        # URL 尾段為 payment_id，middleware 預設會抓成 entity_id；覆寫為 registration_id
        # 才能讓「該筆報名的所有稽核事件」查詢命中此筆。
        request.state.audit_entity_id = str(registration_id)
        request.state.audit_summary = (
            f"軟刪繳費記錄：{reg.student_name} payment_id={payment_id} "
            f"NT${deleted_snapshot['amount']}（{deleted_snapshot['type']}）原因：{body.reason}"
        )
        request.state.audit_changes = {
            "student_name": reg.student_name,
            "voided_payment_id": payment_id,
            "voided_type": deleted_snapshot["type"],
            "voided_amount": deleted_snapshot["amount"],
            "voided_payment_date": deleted_snapshot["payment_date"],
            "void_reason": body.reason,
            "paid_amount_after": reg.paid_amount,
        }
        return {
            "message": "記錄已軟刪（原紀錄保留供稽核）",
            "paid_amount": reg.paid_amount,
            "payment_status": _derive_payment_status(reg.paid_amount, total_amount),
            "voided_at": now.isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.put("/registrations/{registration_id}/waitlist")
async def promote_waitlist(
    registration_id: int,
    background_tasks: BackgroundTasks,
    course_id: int = Query(...),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """管理員手動將候補或 promoted_pending 直接升為正式報名（跳過 24h 確認窗）。"""
    session = get_session()
    try:
        student_name, course_name = activity_service.promote_waitlist(
            session, registration_id, course_id
        )

        activity_service.log_change(
            session,
            registration_id,
            student_name,
            "候補升正式",
            f"課程「{course_name}」由管理員手動升為正式",
            current_user.get("username", ""),
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session)
        line_svc = get_line_service()
        if line_svc is not None:
            # 管理員手動升正式：通知不帶 deadline（已確認）
            background_tasks.add_task(
                line_svc.notify_activity_waitlist_promoted,
                student_name,
                course_name,
                None,
            )
        return {"message": "成功升為正式報名"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.post("/waitlist/sweep-expired")
async def sweep_expired_waitlist_promotions(
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """管理員手動觸發候補轉正過期掃描（排程異常時備援）。"""
    session = get_session()
    try:
        result = activity_service.sweep_expired_pending_promotions(session)
        session.commit()
        _invalidate_activity_dashboard_caches(session, summary_only=True)
        logger.info(
            "手動觸發候補過期掃描：operator=%s expired=%s reminded=%s",
            current_user.get("username", ""),
            result["expired"],
            result["reminded"],
        )
        return {"message": "候補過期掃描完成", **result}
    except Exception as e:
        session.rollback()
        logger.error("候補過期掃描失敗：%s", e)
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/registrations/{registration_id}/courses/{course_id}")
async def withdraw_course(
    registration_id: int,
    course_id: int,
    request: Request,
    force_refund: bool = Query(
        False,
        description="退課後若出現超繳，需顯式帶 true 才允許退課並自動寫退費沖帳紀錄",
    ),
    refund_reason: Optional[str] = Query(
        None,
        description="當 force_refund 觸發實際退費時必填（≥5 字），原因會寫入 notes 供稽核",
    ),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """退出單一課程（含候補），若為正式報名則自動升位候補

    併發保護：鎖住 reg 行，避免兩個 DELETE 同時扣 paid_amount / 重複沖帳。
    第二個請求會在鎖釋放後發現 RC 已被刪而拿到 404。
    """
    session = get_session()
    try:
        reg = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .with_for_update()
            .first()
        )
        if not reg:
            raise _not_found("報名資料")

        from models.database import RegistrationCourse as RC

        rc = (
            session.query(RC)
            .filter(
                RC.registration_id == registration_id,
                RC.course_id == course_id,
            )
            .first()
        )
        if not rc:
            raise _not_found("課程報名項目")

        from models.database import ActivityCourse as AC

        course = session.query(AC).filter(AC.id == course_id).first()
        course_name = course.name if course else str(course_id)
        was_enrolled = rc.status == "enrolled"
        # enrolled 與 promoted_pending 都佔容量，刪除後都應嘗試遞補下一位候補
        was_occupying = rc.status in ("enrolled", "promoted_pending")

        # 先估算退課後的 total 用於 409 預檢；實際退費金額在 flush 後以 new_total 重算
        paid_amount = reg.paid_amount or 0
        # 估算退課後的 new_total：從目前 total 扣掉該 enrolled 項目的 price_snapshot
        # （candidate 列在 RegistrationCourse，候補 status 非 enrolled 不計入 total）
        before_total = _calc_total_amount(session, registration_id)
        if was_enrolled:
            estimated_after_total = before_total - int(rc.price_snapshot or 0)
        else:
            estimated_after_total = before_total
        preview_refund = max(0, paid_amount - estimated_after_total)

        if preview_refund > 0 and not force_refund:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"退課後將產生超繳 NT${preview_refund}（已繳 NT${paid_amount}、"
                    f"退課後應繳 NT${estimated_after_total}），請先處理退費或於退課時"
                    f"指定 force_refund=true 自動沖帳"
                ),
            )

        # 與正式退費端點同套守衛：fail-fast，避免 ACTIVITY_WRITE 經自動沖帳繞過
        # require_refund_reason / require_approve_for_large_refund 簽核閾值
        cleaned_reason: Optional[str] = None
        if preview_refund > 0 and force_refund:
            cleaned_reason = require_refund_reason(refund_reason)
            require_approve_for_cumulative_refund(
                session,
                registration_id,
                preview_refund,
                current_user,
                label="退課自動沖帳累積退費總額",
            )

        session.delete(rc)
        session.flush()

        # 清除該生在此課程所有場次的點名紀錄，避免統計把退課者算入，
        # 並防止未來重新報名時撞到 uq_activity_attendance_session_reg。
        session_ids_subq = (
            session.query(ActivitySession.id)
            .filter(ActivitySession.course_id == course_id)
            .subquery()
        )
        removed_attendance = (
            session.query(ActivityAttendance)
            .filter(
                ActivityAttendance.registration_id == registration_id,
                ActivityAttendance.session_id.in_(session_ids_subq),
            )
            .delete(synchronize_session=False)
        )

        if was_occupying:
            activity_service._auto_promote_first_waitlist(session, course_id)

        new_total = _calc_total_amount(session, registration_id)
        # 以 new_total 重算 refund_needed：避免 estimated_after_total 與實際 DB 狀態漂移
        # （例如 auto_promote 連動、或未來在 delete/flush 之間插入其他邏輯時自我校驗）
        refund_needed = max(0, paid_amount - new_total)

        # 若需要自動沖帳，寫 refund 紀錄並扣 paid_amount
        if refund_needed > 0 and force_refund:
            today = today_taipei()
            _require_daily_close_unlocked(session, today)
            session.add(
                ActivityPaymentRecord(
                    registration_id=registration_id,
                    type="refund",
                    amount=refund_needed,
                    payment_date=today,
                    payment_method=SYSTEM_RECONCILE_METHOD,
                    notes=(
                        f"（退課「{course_name}」自動沖帳）原因：{cleaned_reason}"
                        if cleaned_reason
                        else f"（退課「{course_name}」自動沖帳）"
                    ),
                    operator=current_user.get("username", ""),
                )
            )
            reg.paid_amount = max(0, paid_amount - refund_needed)

        reg.is_paid = _compute_is_paid(reg.paid_amount or 0, new_total)

        log_detail = f"退出課程「{course_name}」"
        if removed_attendance:
            log_detail += f"（同步清除 {removed_attendance} 筆舊點名紀錄）"
        if refund_needed > 0 and force_refund:
            log_detail += f"（自動沖帳退費 NT${refund_needed}）"
        activity_service.log_change(
            session,
            registration_id,
            reg.student_name,
            "退課",
            log_detail,
            current_user.get("username", ""),
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session)
        final_paid = reg.paid_amount or 0
        # URL 尾段為 course_id，覆寫為 registration_id 以便依報名 ID 彙整稽核事件
        request.state.audit_entity_id = str(registration_id)
        request.state.audit_summary = f"退課：{reg.student_name} 退出「{course_name}」"
        request.state.audit_changes = {
            "student_name": reg.student_name,
            "course_id": course_id,
            "course_name": course_name,
            "was_enrolled": was_enrolled,
            "refund_needed": refund_needed,
            "force_refund": force_refund,
            "paid_amount_after": final_paid,
            "total_amount_after": new_total,
            "removed_attendance_count": removed_attendance,
        }
        return {
            "message": f"已退出課程「{course_name}」",
            "total_amount": new_total,
            "paid_amount": final_paid,
            "refunded_amount": refund_needed if force_refund else 0,
            "payment_status": _derive_payment_status(final_paid, new_total),
        }
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()


@router.delete("/registrations/{registration_id}")
async def delete_registration(
    registration_id: int,
    request: Request,
    force_refund: bool = Query(
        False,
        description="若報名已有繳費金額，需顯式帶 true 才允許刪除並自動寫退費沖帳紀錄",
    ),
    refund_reason: Optional[str] = Query(
        None,
        description="當 force_refund 觸發實際退費時必填（≥5 字），原因會寫入 notes 供稽核",
    ),
    current_user: dict = Depends(require_staff_permission(Permission.ACTIVITY_WRITE)),
):
    """軟刪除報名"""
    session = get_session()
    try:
        reg_preview = (
            session.query(ActivityRegistration)
            .filter(
                ActivityRegistration.id == registration_id,
                ActivityRegistration.is_active.is_(True),
            )
            .first()
        )
        student_name = reg_preview.student_name if reg_preview else None
        paid_before = (reg_preview.paid_amount or 0) if reg_preview else 0

        # 與正式退費端點同套守衛：fail-fast，避免 ACTIVITY_WRITE 經自動沖帳繞過
        # require_refund_reason / require_approve_for_large_refund 簽核閾值
        cleaned_reason: Optional[str] = None
        if paid_before > 0 and force_refund:
            cleaned_reason = require_refund_reason(refund_reason)
            require_approve_for_cumulative_refund(
                session,
                registration_id,
                paid_before,
                current_user,
                label="刪除報名自動沖帳累積退費總額",
            )

        activity_service.delete_registration(
            session,
            registration_id,
            current_user.get("username", ""),
            force_refund=force_refund,
            refund_reason=cleaned_reason,
        )
        session.commit()
        _invalidate_activity_dashboard_caches(session)
        logger.warning(
            "課後才藝報名已刪除：id=%s operator=%s force_refund=%s",
            registration_id,
            current_user.get("username"),
            force_refund,
        )
        request.state.audit_summary = (
            f"刪除才藝報名：{student_name}" if student_name else "刪除才藝報名"
        )
        request.state.audit_changes = {
            "student_name": student_name,
            "paid_amount_before": paid_before,
            "force_refund": force_refund,
        }
        return {"message": "報名已刪除"}
    except ValueError as e:
        msg = str(e)
        # 找不到報名 → 404；尚有已繳金額 → 409（需呼叫端確認）
        if "找不到" in msg:
            raise HTTPException(status_code=404, detail=msg)
        raise HTTPException(status_code=409, detail=msg)
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e)
    finally:
        session.close()
