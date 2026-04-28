"""
Student attendance router — 學生每日出席紀錄
"""

import calendar
import logging
from datetime import datetime, date as date_type
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from utils.errors import raise_safe_500
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pydantic import BaseModel

from models.database import get_session, Student, StudentAttendance, Classroom
from services.student_attendance_report import (
    build_daily_classroom_overview,
    build_monthly_attendance_report,
    invalidate_student_attendance_report_caches,
)
from utils.auth import require_staff_permission
from utils.permissions import Permission
from utils.portfolio_access import (
    accessible_classroom_ids,
    assert_student_access,
    filter_student_ids_by_access,
    is_unrestricted,
)
from api.exports import (
    SafeWorksheet,
    _sanitize_excel_value,
    _to_response,
    THIN_BORDER,
    CENTER_ALIGN,
    TITLE_FONT,
    _export_rate_limit,
    _write_header_row,
    _auto_width,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["student-attendance"])

VALID_STATUSES = {"出席", "缺席", "病假", "事假", "遲到"}

# ============ Export Helpers ============

_STATUS_SHORT = {"出席": "出", "缺席": "缺", "病假": "病", "事假": "事", "遲到": "遲"}
_WEEKDAY_NAMES = ["一", "二", "三", "四", "五", "六", "日"]
_WEEKEND_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _fetch_class_data(session, classroom_id, year, month):
    """查詢班級月報資料。"""
    return build_monthly_attendance_report(session, classroom_id, year, month)


def _write_class_sheet(ws_raw, report_data, year, month):
    """將單班出席月報寫入 worksheet（橫向格式）。
    回傳 (summary_dict, workday_count, student_count)。
    """
    ws = SafeWorksheet(ws_raw)
    days = calendar.monthrange(year, month)[1]
    classroom_name = report_data["classroom_name"]
    workdays = report_data["school_days_count"]

    total_cols = 2 + days + 6  # 學號 + 姓名 + 每天 + 6 個統計欄

    # 第 1 列：標題（使用 raw ws，字串為我們自己產生，無注入風險）
    ws._ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws._ws.cell(
        row=1, column=1, value=f"班級：{classroom_name}  {year}年{month}月出席月報"
    )
    tc.font = TITLE_FONT
    tc.alignment = CENTER_ALIGN

    # 第 2 列：表頭
    headers = ["學號", "姓名"]
    for d in range(1, days + 1):
        wd = date_type(year, month, d).weekday()
        headers.append(f"{d}日({_WEEKDAY_NAMES[wd]})")
    headers += ["出席", "缺席", "病假", "事假", "遲到", "未點名"]
    _write_header_row(ws._ws, 2, headers)

    # 週末欄表頭改灰底
    for d in range(1, days + 1):
        if date_type(year, month, d).weekday() >= 5:
            ws._ws.cell(row=2, column=2 + d).fill = _WEEKEND_FILL

    class_summary = {"出席": 0, "缺席": 0, "病假": 0, "事假": 0, "遲到": 0, "未點名": 0}

    for row_idx, student in enumerate(report_data["students"], 3):
        stu_counts = {key: student[key] for key in class_summary.keys()}
        daily_map = {
            date_type.fromisoformat(entry["date"]).day: entry
            for entry in student["daily_records"]
        }

        # 學號、姓名：透過 SafeWorksheet 自動清理（防公式注入）
        c = ws.cell(row=row_idx, column=1, value=student["student_no"])
        c.border = THIN_BORDER
        c = ws.cell(row=row_idx, column=2, value=student["name"])
        c.border = THIN_BORDER

        for d in range(1, days + 1):
            col = 2 + d
            day_meta = report_data["calendar_days"][d - 1]
            daily_entry = daily_map.get(d)

            if day_meta["is_weekend"] or day_meta["is_holiday"]:
                holiday_label = day_meta["holiday_name"] or "-"
                cell = ws._ws.cell(row=row_idx, column=col, value=holiday_label)
                cell.fill = _WEEKEND_FILL
            else:
                status = (
                    daily_entry["status"]
                    if daily_entry and daily_entry["is_school_day"]
                    else None
                )
                if status:
                    cell = ws._ws.cell(
                        row=row_idx, column=col, value=_STATUS_SHORT.get(status, status)
                    )
                else:
                    cell = ws._ws.cell(row=row_idx, column=col, value="")
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        for i, key in enumerate(["出席", "缺席", "病假", "事假", "遲到", "未點名"], 1):
            cell = ws._ws.cell(row=row_idx, column=2 + days + i, value=stu_counts[key])
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            class_summary[key] += stu_counts[key]

    _auto_width(ws._ws)
    return class_summary, workdays, len(report_data["students"])


def _write_summary_sheet(ws_raw, year, month, summary_rows):
    """寫入全園摘要 sheet（每班一列）。"""
    ws = SafeWorksheet(ws_raw)
    headers = [
        "班級",
        "總人數",
        "應出勤天數",
        "出席",
        "缺席",
        "病假",
        "事假",
        "遲到",
        "未點名",
        "出席率",
    ]

    ws._ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws._ws.cell(row=1, column=1, value=f"{year}年{month}月 全園出席摘要")
    tc.font = TITLE_FONT
    tc.alignment = CENTER_ALIGN

    _write_header_row(ws._ws, 2, headers)

    for row_idx, row in enumerate(summary_rows, 3):
        total_possible = row["total_students"] * row["workdays"]
        rate = (
            f"{(row['出席'] + row['遲到']) / total_possible * 100:.1f}%"
            if total_possible > 0
            else "N/A"
        )
        values = [
            row["name"],
            row["total_students"],
            row["workdays"],
            row["出席"],
            row["缺席"],
            row["病假"],
            row["事假"],
            row["遲到"],
            row["未點名"],
            rate,
        ]
        for col, val in enumerate(values, 1):
            cell = ws._ws.cell(
                row=row_idx,
                column=col,
                value=_sanitize_excel_value(val) if isinstance(val, str) else val,
            )
            cell.border = THIN_BORDER

    _auto_width(ws._ws)


# ============ Pydantic Models ============


class AttendanceEntry(BaseModel):
    student_id: int
    status: str = "出席"
    remark: Optional[str] = None


class BatchSaveRequest(BaseModel):
    date: str
    entries: List[AttendanceEntry]


# ============ Routes ============


@router.get("/student-attendance/overview")
async def get_daily_attendance_overview(
    date: str = Query(..., description="YYYY-MM-DD"),
    school_year: Optional[int] = Query(
        None, ge=100, le=200, description="學年度（民國年）"
    ),
    semester: Optional[int] = Query(None, ge=1, le=2, description="學期 1=上 2=下"),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """取得指定日期的各班學生出席總覽。school_year/semester 不傳時使用目前學期。"""
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式錯誤，請使用 YYYY-MM-DD")

    session = get_session()
    try:
        return build_daily_classroom_overview(
            session, target_date, school_year=school_year, semester=semester
        )
    finally:
        session.close()


@router.get("/student-attendance")
async def get_daily_attendance(
    date: str = Query(..., description="YYYY-MM-DD"),
    classroom_id: int = Query(...),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """取得指定日期與班級的出席清單，未點名的學生也會回傳（status=None）"""
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式錯誤，請使用 YYYY-MM-DD")

    session = get_session()
    try:
        # 班級 scope：非管理角色僅可查詢自己班級
        if not is_unrestricted(current_user):
            allowed = accessible_classroom_ids(session, current_user)
            if classroom_id not in allowed:
                raise HTTPException(status_code=403, detail="您無權存取此班級")
        students = (
            session.query(Student)
            .filter(Student.classroom_id == classroom_id, Student.is_active == True)
            .order_by(Student.student_id)
            .all()
        )

        existing = {
            r.student_id: r
            for r in session.query(StudentAttendance)
            .filter(
                StudentAttendance.date == target_date,
                StudentAttendance.student_id.in_([s.id for s in students]),
            )
            .all()
        }

        result = []
        for s in students:
            rec = existing.get(s.id)
            result.append(
                {
                    "student_id": s.id,
                    "student_no": s.student_id,
                    "name": s.name,
                    "status": rec.status if rec else None,
                    "remark": rec.remark if rec else None,
                }
            )

        return {"date": date, "classroom_id": classroom_id, "records": result}
    finally:
        session.close()


@router.post("/student-attendance/batch")
async def batch_save_attendance(
    payload: BatchSaveRequest,
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_WRITE)),
):
    """批量儲存（upsert）一個日期的出席記錄"""
    try:
        target_date = datetime.strptime(payload.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式錯誤，請使用 YYYY-MM-DD")

    invalid = [e.status for e in payload.entries if e.status not in VALID_STATUSES]
    if invalid:
        raise HTTPException(status_code=400, detail=f"無效的出席狀態：{invalid}")

    user_id = current_user.get("id")
    session = get_session()
    try:
        # 班級 scope：批次內所有 student_id 必須通過存取檢查；任何一筆不通過整批 403
        if not is_unrestricted(current_user):
            candidate_ids = {e.student_id for e in payload.entries}
            allowed_ids = filter_student_ids_by_access(
                session, current_user, candidate_ids
            )
            if candidate_ids - allowed_ids:
                raise HTTPException(
                    status_code=403, detail="批次中包含您無權存取的學生"
                )

        existing = {
            r.student_id: r
            for r in session.query(StudentAttendance)
            .filter(
                StudentAttendance.date == target_date,
                StudentAttendance.student_id.in_(
                    [e.student_id for e in payload.entries]
                ),
            )
            .all()
        }

        for entry in payload.entries:
            if entry.student_id in existing:
                rec = existing[entry.student_id]
                rec.status = entry.status
                rec.remark = entry.remark
                rec.recorded_by = user_id
            else:
                rec = StudentAttendance(
                    student_id=entry.student_id,
                    date=target_date,
                    status=entry.status,
                    remark=entry.remark,
                    recorded_by=user_id,
                )
                session.add(rec)

        session.commit()
        invalidate_student_attendance_report_caches(session)
        logger.info(
            "學生出席批量儲存：date=%s count=%d operator=%s",
            payload.date,
            len(payload.entries),
            current_user.get("username"),
        )
        return {"message": "儲存成功", "saved": len(payload.entries)}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise_safe_500(e, context="儲存失敗")
    finally:
        session.close()


@router.get("/student-attendance/by-student")
async def get_attendance_by_student(
    student_id: int = Query(..., gt=0),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
    limit: int = Query(200, ge=1, le=1000),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """查詢單一學生的每日出席紀錄（含統計），供學生紀錄抽屜使用。"""
    df = None
    dt = None
    try:
        if date_from:
            df = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            dt = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式錯誤，請使用 YYYY-MM-DD")

    session = get_session()
    try:
        # 班級 scope：assert_student_access 同時處理 404 / 403
        student = assert_student_access(session, current_user, student_id)

        q = session.query(StudentAttendance).filter(
            StudentAttendance.student_id == student_id
        )
        if df:
            q = q.filter(StudentAttendance.date >= df)
        if dt:
            q = q.filter(StudentAttendance.date <= dt)
        rows = q.order_by(StudentAttendance.date.desc()).limit(limit).all()

        counts = {key: 0 for key in VALID_STATUSES}
        items = []
        for r in rows:
            items.append(
                {
                    "id": r.id,
                    "date": r.date.isoformat() if r.date else None,
                    "status": r.status,
                    "remark": r.remark,
                }
            )
            if r.status in counts:
                counts[r.status] += 1

        return {
            "student_id": student_id,
            "student_name": student.name,
            "items": items,
            "total": len(items),
            "counts": counts,
        }
    finally:
        session.close()


@router.get("/student-attendance/monthly")
async def get_monthly_summary(
    classroom_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """取得班級整月出席統計、出席率與連缺告警。"""
    session = get_session()
    try:
        # 班級 scope：非管理角色僅可查詢自己班級
        if not is_unrestricted(current_user):
            allowed = accessible_classroom_ids(session, current_user)
            if classroom_id not in allowed:
                raise HTTPException(status_code=403, detail="您無權存取此班級")
        return build_monthly_attendance_report(session, classroom_id, year, month)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    finally:
        session.close()


@router.get("/student-attendance/export")
def export_student_attendance(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    classroom_id: Optional[int] = Query(None),
    _rl=Depends(_export_rate_limit),
    current_user: dict = Depends(require_staff_permission(Permission.STUDENTS_READ)),
):
    """匯出班級（或全園）月出席 Excel。
    classroom_id 不傳時匯出全園（多 sheet + 摘要 sheet）。
    """
    # 班級 scope：
    # - 全園匯出（classroom_id = None）僅 admin/hr/supervisor 可用
    # - 單班匯出：非管理角色必須在自己 accessible 範圍內
    session = get_session()
    try:
        if not is_unrestricted(current_user):
            if classroom_id is None:
                raise HTTPException(
                    status_code=403, detail="僅管理角色可匯出全園出席月報"
                )
            allowed = accessible_classroom_ids(session, current_user)
            if classroom_id not in allowed:
                raise HTTPException(status_code=403, detail="您無權存取此班級")
        if classroom_id is not None:
            classrooms = (
                session.query(Classroom)
                .filter(Classroom.id == classroom_id, Classroom.is_active == True)
                .all()
            )
            if not classrooms:
                raise HTTPException(status_code=404, detail="班級不存在")
        else:
            classrooms = (
                session.query(Classroom)
                .filter(Classroom.is_active == True)
                .order_by(Classroom.name)
                .all()
            )

        wb = Workbook()

        if classroom_id is not None:
            cr = classrooms[0]
            report_data = _fetch_class_data(session, cr.id, year, month)
            ws_raw = wb.active
            ws_raw.title = cr.name[:31]
            _write_class_sheet(SafeWorksheet(ws_raw), report_data, year, month)
            filename = f"{year}年{month}月_{cr.name}_出席月報.xlsx"
        else:
            summary_ws = SafeWorksheet(wb.active)
            summary_ws.title = "全園摘要"
            summary_rows = []

            for cr in classrooms:
                report_data = _fetch_class_data(session, cr.id, year, month)
                sheet = SafeWorksheet(wb.create_sheet(title=cr.name[:31]))
                class_summary, workdays, total_students = _write_class_sheet(
                    sheet, report_data, year, month
                )
                summary_rows.append(
                    {
                        "name": cr.name,
                        "total_students": total_students,
                        "workdays": workdays,
                        **class_summary,
                    }
                )

            _write_summary_sheet(summary_ws, year, month, summary_rows)
            filename = f"{year}年{month}月_全園出席月報.xlsx"

        logger.info(
            "學生出席月報匯出：year=%d month=%d classroom_id=%s operator=%s",
            year,
            month,
            classroom_id,
            current_user.get("username"),
        )
        return _to_response(wb, filename)
    finally:
        session.close()
