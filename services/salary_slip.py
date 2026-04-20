"""
薪資單匯出服務 - PDF / Excel 產生
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Excel 公式注入防護集中於 utils.excel_utils
from utils.excel_utils import (
    SafeWorksheet,
    sanitize_excel_value as _sanitize_excel_value,
)


def _register_cjk_font():
    """註冊 CJK 字型（使用 reportlab 內建的 CID 字型）"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    try:
        pdfmetrics.getFont("STSong-Light")
    except KeyError:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


def generate_salary_pdf(record, employee, year: int, month: int) -> bytes:
    """
    產生單人薪資單 PDF

    Args:
        record: SalaryRecord ORM object
        employee: Employee ORM object
        year: 薪資年度
        month: 薪資月份

    Returns:
        PDF bytes
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate

    _register_cjk_font()
    font_name = "STSong-Light"
    styles = getSampleStyleSheet()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
    )

    elements = _build_salary_elements(record, employee, year, month, font_name, styles)
    doc.build(elements)
    return buffer.getvalue()


def _make_paragraph_styles(font_name: str, styles):
    """建立薪資單所需的三種 ParagraphStyle。"""
    from reportlab.lib.styles import ParagraphStyle

    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "ChineseSubtitle",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=12,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        "ChineseNormal", parent=styles["Normal"], fontName=font_name, fontSize=10
    )
    return title_style, subtitle_style, normal_style


def _build_employee_info_table(employee, font_name: str):
    """建立員工基本資料表格（姓名、職稱、員工編號、部門）。"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    job_title = ""
    if hasattr(employee, "job_title_rel") and employee.job_title_rel:
        job_title = employee.job_title_rel.name
    elif employee.title:
        job_title = employee.title

    info_data = [
        ["姓名", employee.name, "職稱", job_title],
        ["員工編號", employee.employee_id, "部門/班級", employee.position or ""],
    ]
    info_table = Table(info_data, colWidths=[60, 140, 60, 140])
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
                ("BACKGROUND", (2, 0), (2, -1), colors.Color(0.9, 0.9, 0.9)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return info_table


def _build_earnings_table(record, font_name: str, money_fmt):
    """建立應領項目表格（底薪、津貼、獎金明細）。"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    supervisor_dividend = getattr(record, "supervisor_dividend", 0) or 0
    bonus_separate = getattr(record, "bonus_separate", False)
    festival_bonus_val = record.festival_bonus or 0
    overtime_bonus_val = record.overtime_bonus or 0
    total_bonus = (
        (record.performance_bonus or 0)
        + (record.special_bonus or 0)
        + supervisor_dividend
    )

    earn_data = [
        ["應領項目", "", "金額"],
        ["底薪", "", money_fmt(record.base_salary)],
        ["績效獎金", "", money_fmt(record.performance_bonus)],
        ["特別獎金", "", money_fmt(record.special_bonus)],
        ["主管紅利", "", money_fmt(supervisor_dividend)],
        ["獎金小計", "", money_fmt(total_bonus)],
    ]
    if bonus_separate:
        festival_separate = festival_bonus_val + overtime_bonus_val
        earn_data.append(["節慶/超額獎金 (另行轉帳)", "", money_fmt(festival_separate)])
    earn_data.append(["月薪應發合計", "", money_fmt(record.gross_salary)])

    earn_table = Table(earn_data, colWidths=[120, 100, 120])
    earn_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.7)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.Color(0.9, 0.95, 1.0)),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return earn_table


def _build_deductions_table(record, font_name: str, money_fmt):
    """建立扣款項目表格（保險、考勤扣款明細）。"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    total_insurance = (
        (record.labor_insurance_employee or 0)
        + (record.health_insurance_employee or 0)
        + (record.pension_employee or 0)
    )
    total_attendance_deduction = (
        (record.late_deduction or 0)
        + (record.early_leave_deduction or 0)
        + (record.missing_punch_deduction or 0)
        + (record.leave_deduction or 0)
    )

    deduct_data = [
        ["扣款項目", "", "金額"],
        ["勞保費 (自付)", "", money_fmt(record.labor_insurance_employee)],
        ["健保費 (自付)", "", money_fmt(record.health_insurance_employee)],
        ["勞退自提", "", money_fmt(record.pension_employee)],
        ["保險小計", "", money_fmt(total_insurance)],
        ["遲到扣款", f"({record.late_count or 0}次)", money_fmt(record.late_deduction)],
        [
            "早退扣款",
            f"({record.early_leave_count or 0}次)",
            money_fmt(record.early_leave_deduction),
        ],
        [
            "未打卡扣款",
            f"({record.missing_punch_count or 0}次)",
            money_fmt(record.missing_punch_deduction),
        ],
        ["請假扣款", "", money_fmt(record.leave_deduction)],
        ["其他扣款", "", money_fmt(record.other_deduction)],
        ["考勤扣款小計", "", money_fmt(total_attendance_deduction)],
        ["扣款合計", "", money_fmt(record.total_deduction)],
    ]

    deduct_table = Table(deduct_data, colWidths=[120, 100, 120])
    deduct_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.7, 0.2, 0.2)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.Color(1.0, 0.9, 0.9)),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return deduct_table


def _build_net_salary_table(record, font_name: str, money_fmt):
    """建立實發金額匯總表格。"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    net_data = [["實發金額", money_fmt(record.net_salary)]]
    net_table = Table(net_data, colWidths=[220, 120])
    net_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 14),
                ("BACKGROUND", (0, 0), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return net_table


def _build_salary_elements(
    record, employee, year: int, month: int, font_name: str, styles
) -> list:
    """建構單人薪資單的 platypus elements 清單（供 generate_salary_pdf 與 generate_salary_all_pdf 共用）"""
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer

    title_style, subtitle_style, normal_style = _make_paragraph_styles(
        font_name, styles
    )

    def money(val):
        return f"{int(val):,}" if val else "0"

    elements = []

    # 標題區
    elements.append(Paragraph("薪資單", title_style))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(f"{year} 年 {month} 月", subtitle_style))
    elements.append(Spacer(1, 6 * mm))

    # 員工資料
    elements.append(_build_employee_info_table(employee, font_name))
    elements.append(Spacer(1, 6 * mm))

    # 應領項目
    elements.append(Paragraph("應領項目", normal_style))
    elements.append(Spacer(1, 2 * mm))
    elements.append(_build_earnings_table(record, font_name, money))
    elements.append(Spacer(1, 6 * mm))

    # 扣款項目
    elements.append(Paragraph("扣款項目", normal_style))
    elements.append(Spacer(1, 2 * mm))
    elements.append(_build_deductions_table(record, font_name, money))
    elements.append(Spacer(1, 8 * mm))

    # 實發金額
    elements.append(_build_net_salary_table(record, font_name, money))

    return elements


def generate_salary_all_pdf(records_with_employees, year: int, month: int) -> bytes:
    """
    產生全員薪資單合一 PDF（每人一頁）

    Args:
        records_with_employees: list of (SalaryRecord, Employee) tuples
        year: 薪資年度
        month: 薪資月份

    Returns:
        PDF bytes
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, SimpleDocTemplate

    _register_cjk_font()
    font_name = "STSong-Light"
    styles = getSampleStyleSheet()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
    )

    all_elements = []
    total = len(records_with_employees)
    for idx, (record, employee) in enumerate(records_with_employees):
        all_elements.extend(
            _build_salary_elements(record, employee, year, month, font_name, styles)
        )
        if idx < total - 1:
            all_elements.append(PageBreak())

    doc.build(all_elements)
    return buffer.getvalue()


def generate_salary_excel(records_with_employees, year: int, month: int) -> bytes:
    """
    產生全部員工薪資 Excel

    Args:
        records_with_employees: list of (SalaryRecord, Employee) tuples
        year: 薪資年度
        month: 薪資月份

    Returns:
        Excel bytes
    """
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = f"{year}年{month}月薪資表"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = "#,##0"

    # Title row
    ws.merge_cells("A1:P1")
    ws["A1"] = f"{year}年{month}月 薪資總表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "姓名",
        "員工編號",
        "職稱",
        "底薪",
        "節慶獎金(另轉)",
        "超額獎金(另轉)",
        "績效獎金",
        "主管紅利",
        "獨立獎金合計",
        "月薪應發",
        "勞保",
        "健保",
        "考勤扣款",
        "扣款合計",
        "實發金額",
        "編輯紀錄",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (record, employee) in enumerate(records_with_employees, 4):
        job_title = ""
        if hasattr(employee, "job_title_rel") and employee.job_title_rel:
            job_title = employee.job_title_rel.name
        elif employee.title:
            job_title = employee.title

        attendance_deduction = (
            (record.late_deduction or 0)
            + (record.early_leave_deduction or 0)
            + (record.missing_punch_deduction or 0)
            + (record.leave_deduction or 0)
        )

        supervisor_dividend = getattr(record, "supervisor_dividend", 0) or 0
        festival_bonus_val = record.festival_bonus or 0
        overtime_bonus_val = record.overtime_bonus or 0
        independent_bonus = festival_bonus_val + overtime_bonus_val

        values = [
            employee.name,
            employee.employee_id,
            job_title,
            record.base_salary or 0,
            festival_bonus_val,
            overtime_bonus_val,
            record.performance_bonus or 0,
            supervisor_dividend,
            independent_bonus,
            record.gross_salary or 0,
            record.labor_insurance_employee or 0,
            record.health_insurance_employee or 0,
            attendance_deduction,
            record.total_deduction or 0,
            record.net_salary or 0,
            record.remark or "",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=_sanitize_excel_value(value))
            cell.border = thin_border
            if isinstance(value, (int, float)) and col >= 4:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 12

    # Name column wider
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
