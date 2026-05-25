"""
銀行轉帳名冊 Excel 匯出 — 對齊會計慣用《義華薪資》Excel 排版。

四種類型：
- base：當月薪資（net_salary 扣除獨立轉帳獎金）
- festival：節慶獎金實發金額（發放月才有累積值）
- surplus：超額獎金實發金額（學期末）
- art_teacher：才藝/鐘點老師（employee_type='hourly'）淨薪

格式：
    Row1 B  園所名
    Row2 B  「{民國年}年{月}月 {類型}轉帳名冊」
    Row3 B  「帳號：{公司付款帳號}」
    Row4    [帳號] [戶名] [金額]
    Row5+   每員工一行（依員工編號排序）
    末-1    [   ] [合計] [sum]
    末      [   ] [經辦人:]
"""

from __future__ import annotations

import io
import logging
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from models.database import Employee, SalaryRecord, SystemConfig
from utils.excel_utils import SafeWorksheet
from utils.rounding import round_half_up

logger = logging.getLogger(__name__)

ROSTER_TYPES = ("base", "festival", "surplus", "art_teacher")

_TYPE_TITLE = {
    "base": "薪資",
    "festival": "節慶獎金",
    "surplus": "超額獎金",
    "art_teacher": "才藝老師",
}

# SystemConfig keys（不在這層 migration，未設定時 fallback 到 default）
CFG_PAYER_NAME = "bank.payer_name"
CFG_PAYER_ACCOUNT = "bank.payer_account"
DEFAULT_PAYER_NAME = "高雄市私立常春藤幼兒園"
DEFAULT_PAYER_ACCOUNT = "0727-940-008106"


def _read_config(session: Session, key: str, default: str) -> str:
    obj = session.query(SystemConfig).filter(SystemConfig.config_key == key).first()
    return obj.config_value if obj and obj.config_value else default


def _roc_year(year: int) -> int:
    return year - 1911


def _resolve_amount(record: SalaryRecord, roster_type: str) -> int:
    """依名冊類型決定該員工該月金額。回傳整數元（已四捨五入）。

    base：直接用 net_salary。
      net_salary = gross_salary - total_deduction，其中 gross_salary 公式
      （engine.py:1764-1770 / totals.py:21-30）為:
          base + perf + special + supervisor_dividend + birthday
          + meeting_overtime_pay + overtime_pay
      已不含 festival_bonus / overtime_bonus（這兩項走 festival / surplus 名冊獨立轉帳）。
      因此 base 名冊與 festival / surplus 名冊不會重複入帳，不可再扣 bonus_amount
      （含 supervisor_dividend）—— 否則主管紅利會漏付，festival/overtime 也會被
      重複扣一次。salary_slip.py:145-147 的「另行轉帳」也只取 festival + overtime。
    """
    if roster_type == "base":
        return int(round_half_up(float(record.net_salary or 0)))
    if roster_type == "festival":
        return int(round_half_up(float(record.festival_bonus or 0)))
    if roster_type == "surplus":
        return int(round_half_up(float(record.overtime_bonus or 0)))
    if roster_type == "art_teacher":
        # 才藝/鐘點老師整張薪資都走 net_salary（hourly 制無獎金分離議題）
        return int(round_half_up(float(record.net_salary or 0)))
    raise ValueError(f"unknown roster_type: {roster_type}")


def _query_rows(
    session: Session, year: int, month: int, roster_type: str
) -> list[tuple[Employee, int]]:
    """回傳 (employee, amount) 列表，已過濾 0 金額與無帳號者，按工號排序。"""
    q = (
        session.query(SalaryRecord, Employee)
        .join(Employee, SalaryRecord.employee_id == Employee.id)
        .filter(
            SalaryRecord.salary_year == year,
            SalaryRecord.salary_month == month,
            SalaryRecord.is_finalized == True,  # noqa: E712
            Employee.skip_payroll_transfer == False,  # noqa: E712
        )
    )
    if roster_type == "art_teacher":
        q = q.filter(Employee.employee_type == "hourly")
    else:
        q = q.filter(Employee.employee_type == "regular")

    rows: list[tuple[Employee, int]] = []
    skipped_no_account: list[str] = []
    for record, emp in q.all():
        amount = _resolve_amount(record, roster_type)
        if amount <= 0:
            continue
        if not (emp.bank_account and emp.bank_account.strip()):
            skipped_no_account.append(emp.name)
            continue
        rows.append((emp, amount))

    if skipped_no_account:
        logger.warning(
            "transfer roster %s %d/%02d: skipped %d employees with no bank_account: %s",
            roster_type,
            year,
            month,
            len(skipped_no_account),
            ", ".join(skipped_no_account),
        )

    rows.sort(key=lambda pair: pair[0].employee_id or "")
    return rows


def _write_workbook(
    title: str,
    payer_name: str,
    payer_account: str,
    rows: Iterable[tuple[Employee, int]],
) -> bytes:
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "轉帳名冊"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws["B1"] = payer_name
    ws["B1"].font = bold
    ws["B2"] = title
    ws["B2"].font = bold
    ws["B3"] = f"帳號：{payer_account}"

    ws["A4"] = "帳號"
    ws["B4"] = "戶名"
    ws["C4"] = "金額"
    for col in ("A4", "B4", "C4"):
        ws[col].font = bold
        ws[col].alignment = center

    row_idx = 5
    total = 0
    for emp, amount in rows:
        ws.cell(row=row_idx, column=1, value=emp.bank_account or "")
        ws.cell(row=row_idx, column=2, value=emp.bank_account_name or emp.name)
        ws.cell(row=row_idx, column=3, value=amount)
        total += amount
        row_idx += 1

    ws.cell(row=row_idx, column=2, value="合計").font = bold
    ws.cell(row=row_idx, column=3, value=total).font = bold
    row_idx += 1
    ws.cell(row=row_idx, column=2, value="經辦人:")

    # 欄寬
    raw_ws = wb.active
    raw_ws.column_dimensions["A"].width = 22
    raw_ws.column_dimensions["B"].width = 16
    raw_ws.column_dimensions["C"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_transfer_roster(
    session: Session, year: int, month: int, roster_type: str
) -> tuple[str, bytes]:
    """產生指定類型的轉帳名冊 xlsx。

    Returns:
        (filename, xlsx_bytes)
    Raises:
        ValueError：roster_type 不支援
    """
    if roster_type not in ROSTER_TYPES:
        raise ValueError(f"unsupported roster_type: {roster_type}")

    payer_name = _read_config(session, CFG_PAYER_NAME, DEFAULT_PAYER_NAME)
    payer_account = _read_config(session, CFG_PAYER_ACCOUNT, DEFAULT_PAYER_ACCOUNT)

    rows = _query_rows(session, year, month, roster_type)
    title = f"{_roc_year(year)}年{month:02d}月 {_TYPE_TITLE[roster_type]}轉帳名冊"

    xlsx_bytes = _write_workbook(title, payer_name, payer_account, rows)
    filename = f"transfer_roster_{roster_type}_{year}_{month:02d}.xlsx"
    return filename, xlsx_bytes
