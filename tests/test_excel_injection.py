"""
回歸測試：Excel 公式注入防護（Formula / DDE Injection）

現況：
    _sanitize_excel_value 已正確處理危險前綴，_write_data_row 也有呼叫它。
    但若未來新增報表功能直接用 ws.cell() 跳過 _write_data_row，
    防護就失效，員工姓名 / 請假原因等欄位若含 =cmd|'/C calc'!A0 將對
    財會端 Excel 開啟時執行任意指令。

修法：
    新增 SafeWorksheet 薄包裝器，在 .cell() 與 ws["A1"]=value 層級
    自動執行 sanitize，使防護掛在「worksheet 寫入」這一更底層，
    即使開發者直接呼叫 ws.cell() 也無法繞過。
"""
import pytest


class TestSanitizeExcelValue:
    """_sanitize_excel_value 現有函式的完整邊界條件測試"""

    def test_formula_equal_prefix_gets_quoted(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("=1+2") == "'=1+2"

    def test_plus_prefix_gets_quoted(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("+FOO") == "'+FOO"

    def test_minus_prefix_gets_quoted(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("-1") == "'-1"

    def test_at_prefix_gets_quoted(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("@SUM(A1)") == "'@SUM(A1)"

    def test_pipe_dde_prefix_gets_quoted(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("|cmd") == "'|cmd"

    def test_tab_bypass_stripped_then_formula_quoted(self):
        """\\t=cmd... 是繞過前綴偵測的常見手法，應先去除 tab 再加引號"""
        from api.exports import _sanitize_excel_value
        result = _sanitize_excel_value("\t=cmd|'/C calc'!A0")
        assert result == "'=cmd|'/C calc'!A0"

    def test_cr_lf_bypass_also_stripped(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("\r\n=evil") == "'=evil"

    def test_normal_chinese_name_unchanged(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("王小明") == "王小明"

    def test_empty_string_unchanged(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value("") == ""

    def test_non_string_int_unchanged(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value(30000) == 30000

    def test_non_string_float_unchanged(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value(3.14) == 3.14

    def test_none_unchanged(self):
        from api.exports import _sanitize_excel_value
        assert _sanitize_excel_value(None) is None

    def test_already_quoted_string_not_double_quoted(self):
        """已加了 ' 前綴的值再次清理時不應雙重加引號（冪等性）"""
        from api.exports import _sanitize_excel_value
        once = _sanitize_excel_value("=danger")
        twice = _sanitize_excel_value(once)
        assert twice == once  # 冪等


class TestSafeWorksheet:
    """SafeWorksheet 薄包裝器單元測試"""

    def _raw_ws(self):
        from openpyxl import Workbook
        return Workbook().active

    def test_cell_sanitizes_formula_injection(self):
        """直接呼叫 ws.cell() 寫入危險公式時，SafeWorksheet 自動加 ' 前綴"""
        from api.exports import SafeWorksheet
        ws = SafeWorksheet(self._raw_ws())
        cell = ws.cell(row=1, column=1, value="=cmd|'/C calc'!A0")
        assert cell.value.startswith("'")
        assert "=cmd" in cell.value

    def test_setitem_sanitizes_formula_injection(self):
        """ws['A1'] = value 語法也自動清理"""
        from api.exports import SafeWorksheet
        raw = self._raw_ws()
        ws = SafeWorksheet(raw)
        ws["A1"] = "=SUM(A1:A10)"
        assert raw["A1"].value.startswith("'")

    def test_clean_value_passes_through_unchanged(self):
        """正常姓名不受影響"""
        from api.exports import SafeWorksheet
        ws = SafeWorksheet(self._raw_ws())
        cell = ws.cell(row=1, column=1, value="陳美玲")
        assert cell.value == "陳美玲"

    def test_non_string_value_unchanged(self):
        """數字不被轉為字串"""
        from api.exports import SafeWorksheet
        ws = SafeWorksheet(self._raw_ws())
        cell = ws.cell(row=1, column=1, value=50000)
        assert cell.value == 50000

    def test_none_value_unchanged(self):
        from api.exports import SafeWorksheet
        ws = SafeWorksheet(self._raw_ws())
        cell = ws.cell(row=1, column=1, value=None)
        assert cell.value is None

    def test_proxies_title_attribute(self):
        """ws.title = '...' 正常傳遞給底層 worksheet"""
        from api.exports import SafeWorksheet
        raw = self._raw_ws()
        ws = SafeWorksheet(raw)
        ws.title = "員工名冊"
        assert raw.title == "員工名冊"

    def test_proxies_merge_cells(self):
        """ws.merge_cells() 正常傳遞，不拋出例外"""
        from api.exports import SafeWorksheet
        ws = SafeWorksheet(self._raw_ws())
        ws.merge_cells("A1:D1")  # must not raise

    def test_cell_returns_real_cell_for_style_setting(self):
        """cell() 回傳真實 Cell，font / fill / border 等可直接設定"""
        from api.exports import SafeWorksheet
        from openpyxl.styles import Font
        ws = SafeWorksheet(self._raw_ws())
        cell = ws.cell(row=1, column=1, value="測試")
        cell.font = Font(bold=True)  # must not raise
        assert cell.font.bold is True

    def test_getitem_returns_real_cell_for_style_setting(self):
        """ws['A1'] 取回真實 Cell，可設定 font 等樣式"""
        from api.exports import SafeWorksheet
        from openpyxl.styles import Alignment
        ws = SafeWorksheet(self._raw_ws())
        ws["A1"] = "標題"
        ws["A1"].alignment = Alignment(horizontal="center")  # must not raise

    def test_direct_cell_bypass_also_protected(self):
        """即使未來開發者跳過 _write_data_row 直接呼叫 ws.cell()，仍受保護"""
        from api.exports import SafeWorksheet
        raw = self._raw_ws()
        ws = SafeWorksheet(raw)

        # 模擬「忘記呼叫 _write_data_row」的開發者
        ws.cell(row=2, column=3, value="=HYPERLINK(\"http://evil.com\",\"click\")")

        stored = raw.cell(row=2, column=3).value
        assert stored.startswith("'"), f"期望 ' 前綴，實際儲存值：{stored!r}"

    def test_sanitize_is_idempotent_through_wrapper(self):
        """SafeWorksheet + _write_data_row 雙重清理不會產生多餘引號"""
        from api.exports import SafeWorksheet, _sanitize_excel_value
        raw = self._raw_ws()
        ws = SafeWorksheet(raw)

        # _write_data_row 先 sanitize 一次，SafeWorksheet 再 sanitize 一次
        pre_sanitized = _sanitize_excel_value("=evil")   # 模擬 _write_data_row 的預清理
        ws.cell(row=1, column=1, value=pre_sanitized)     # SafeWorksheet 再清理一次

        stored = raw.cell(row=1, column=1).value
        assert stored == "'=evil"    # 只有一個 ' 前綴
        assert not stored.startswith("''")  # 不應雙重引號
