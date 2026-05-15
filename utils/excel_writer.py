"""utils/excel_writer.py — Excel 匯出共用 header / data row 寫入器。

蒐集 api/exports.py / shifts / events / overtimes / leaves 五份完全相同的
header style + write helper（_SH_/_EV_/_OT_/_LV_ 前綴版本），統一在此。

公開 API：
- write_header_row(ws, row, headers) — 寫表頭（深藍底白字、置中、細線框）
- write_data_row(ws, row, values)    — 寫資料列（細線框、自動 sanitize 公式注入）

樣式常數亦 export 供需要客製化的呼叫端覆寫單格。
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.excel_utils import sanitize_excel_value

EXCEL_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
EXCEL_HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
EXCEL_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EXCEL_CENTER_ALIGN = Alignment(horizontal="center")


def write_header_row(ws, row, headers) -> None:
    """寫表頭：深藍底白字、置中、細線框。"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = EXCEL_HEADER_FONT
        cell.fill = EXCEL_HEADER_FILL
        cell.border = EXCEL_THIN_BORDER
        cell.alignment = EXCEL_CENTER_ALIGN


def write_data_row(ws, row, values) -> None:
    """寫資料列：細線框 + sanitize 公式注入（=, +, -, @ 開頭加前綴 ')。"""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=sanitize_excel_value(value))
        cell.border = EXCEL_THIN_BORDER
