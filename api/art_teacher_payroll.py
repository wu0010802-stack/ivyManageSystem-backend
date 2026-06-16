"""才藝老師薪資明細 CRUD + 清冊匯出。

每月每老師可有多筆給付（科目/班級/星期）；若該月有 entries，
salary engine 會以 sum(entries.total_amount) 覆寫 salary_record.hourly_total。
"""

import io
import logging
from datetime import datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import joinedload

from models.base import session_scope
from models.database import ArtTeacherPayrollEntry, Employee, SalaryRecord
from services.art_teacher_payroll import (
    generate_art_teacher_roster_xlsx,
    recompute_entry_amounts,
)
from services.finance.salary_access import has_full_salary_view
from services.salary.finalize_guard import assert_months_not_finalized
from services.salary.utils import mark_salary_stale
from utils.auth import require_staff_permission
from utils.excel_utils import SafeWorksheet
from utils.file_upload import read_upload_with_size_check, validate_file_signature
from utils.permissions import Permission, has_permission

router = APIRouter(prefix="/api", tags=["art-teacher-payroll"])

logger = logging.getLogger(__name__)


def _enforce_art_payroll_view(current_user: dict) -> None:
    """才藝薪資明細 viewer 守衛：admin/hr（full salary view）或持 SALARY_WRITE（會計，
    本就建立/管理才藝薪資）可看全部；僅持 SALARY_READ 者（如園長）不可越權看全所才藝
    老師逐筆給付（與 api/salary 其他端點的 self-or-full 行為一致）。"""
    if has_full_salary_view(current_user):
        return
    if has_permission(current_user.get("permission_names"), Permission.SALARY_WRITE):
        return
    raise HTTPException(
        status_code=403, detail="才藝薪資明細僅限 admin/hr 或具編輯權限者檢視"
    )


class EntryCreate(BaseModel):
    employee_id: int
    salary_year: int = Field(..., ge=2000, le=2100)
    salary_month: int = Field(..., ge=1, le=12)
    subject: str = Field(..., min_length=1, max_length=50)
    classroom_label: Optional[str] = Field(None, max_length=50)
    hours: float = Field(0, ge=0)
    hourly_rate: float = Field(0, ge=0)
    excess_amount: float = Field(0, ge=0)
    activity_bonus: float = Field(0, ge=0)
    note: Optional[str] = None


class EntryUpdate(BaseModel):
    subject: Optional[str] = Field(None, min_length=1, max_length=50)
    classroom_label: Optional[str] = Field(None, max_length=50)
    hours: Optional[float] = Field(None, ge=0)
    hourly_rate: Optional[float] = Field(None, ge=0)
    excess_amount: Optional[float] = Field(None, ge=0)
    activity_bonus: Optional[float] = Field(None, ge=0)
    note: Optional[str] = None


class EntryOut(BaseModel):
    id: int
    employee_id: int
    employee_name: Optional[str] = None
    salary_year: int
    salary_month: int
    subject: str
    classroom_label: Optional[str] = None
    hours: float
    hourly_rate: float
    base_amount: float
    excess_amount: float
    activity_bonus: float
    total_amount: float
    note: Optional[str] = None
    created_at: datetime
    updated_at: datetime


def _to_out(entry: ArtTeacherPayrollEntry) -> EntryOut:
    return EntryOut(
        id=entry.id,
        employee_id=entry.employee_id,
        employee_name=getattr(entry.employee, "name", None),
        salary_year=entry.salary_year,
        salary_month=entry.salary_month,
        subject=entry.subject,
        classroom_label=entry.classroom_label,
        hours=float(entry.hours or 0),
        hourly_rate=float(entry.hourly_rate or 0),
        base_amount=float(entry.base_amount or 0),
        excess_amount=float(entry.excess_amount or 0),
        activity_bonus=float(entry.activity_bonus or 0),
        total_amount=float(entry.total_amount or 0),
        note=entry.note,
        created_at=entry.created_at,
        updated_at=entry.updated_at,
    )


@router.get("/art-teacher-payroll")
def list_entries(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    employee_id: Optional[int] = Query(None),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
):
    """列出指定月份的才藝薪資明細，按工號 + entry id 排序。"""
    _enforce_art_payroll_view(current_user)
    with session_scope() as session:
        q = (
            session.query(ArtTeacherPayrollEntry)
            .options(joinedload(ArtTeacherPayrollEntry.employee))
            .filter(
                ArtTeacherPayrollEntry.salary_year == year,
                ArtTeacherPayrollEntry.salary_month == month,
            )
        )
        if employee_id is not None:
            q = q.filter(ArtTeacherPayrollEntry.employee_id == employee_id)
        rows = (
            q.join(Employee, Employee.id == ArtTeacherPayrollEntry.employee_id)
            .order_by(Employee.employee_id, ArtTeacherPayrollEntry.id)
            .all()
        )
        return {"items": [_to_out(e).model_dump() for e in rows]}


@router.post("/art-teacher-payroll")
def create_entry(
    payload: EntryCreate,
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_WRITE)),
):
    """新增單筆才藝薪資明細。"""
    with session_scope() as session:
        emp = session.query(Employee).get(payload.employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="員工不存在")
        if emp.employee_type != "hourly":
            raise HTTPException(
                status_code=400,
                detail=f"員工 {emp.name} 非時薪制（employee_type 須為 hourly）",
            )

        # 鐘點明細是薪資引擎 hourly_total 來源（engine.py），封存後禁改、未封存改動標 stale。
        assert_months_not_finalized(
            session,
            employee_id=payload.employee_id,
            months={(payload.salary_year, payload.salary_month)},
        )

        entry = ArtTeacherPayrollEntry(
            employee_id=payload.employee_id,
            salary_year=payload.salary_year,
            salary_month=payload.salary_month,
            subject=payload.subject,
            classroom_label=payload.classroom_label,
            hours=payload.hours,
            hourly_rate=payload.hourly_rate,
            excess_amount=payload.excess_amount,
            activity_bonus=payload.activity_bonus,
            note=payload.note,
            created_by=current_user.get("username"),
            updated_by=current_user.get("username"),
        )
        recompute_entry_amounts(entry)
        session.add(entry)
        session.flush()
        mark_salary_stale(
            session, payload.employee_id, payload.salary_year, payload.salary_month
        )
        entry.employee = emp
        result = _to_out(entry).model_dump()
        session.commit()
        return result


@router.put("/art-teacher-payroll/{entry_id}")
def update_entry(
    entry_id: int,
    payload: EntryUpdate,
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_WRITE)),
):
    """編輯才藝薪資明細（任何欄位變動後自動重算 base_amount / total_amount）。"""
    with session_scope() as session:
        entry = (
            session.query(ArtTeacherPayrollEntry)
            .options(joinedload(ArtTeacherPayrollEntry.employee))
            .filter(ArtTeacherPayrollEntry.id == entry_id)
            .first()
        )
        if not entry:
            raise HTTPException(status_code=404, detail="明細不存在")

        # 封存後禁改該月來源；未封存改動標 stale（entry 年月不可變，EntryUpdate 無此欄）。
        assert_months_not_finalized(
            session,
            employee_id=entry.employee_id,
            months={(entry.salary_year, entry.salary_month)},
        )

        update_data = payload.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(entry, field, value)
        recompute_entry_amounts(entry)
        entry.updated_by = current_user.get("username")
        session.flush()
        mark_salary_stale(
            session, entry.employee_id, entry.salary_year, entry.salary_month
        )
        result = _to_out(entry).model_dump()
        session.commit()
        return result


@router.delete("/art-teacher-payroll/{entry_id}")
def delete_entry(
    entry_id: int,
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_WRITE)),
):
    """刪除單筆明細。"""
    with session_scope() as session:
        entry = (
            session.query(ArtTeacherPayrollEntry)
            .filter(ArtTeacherPayrollEntry.id == entry_id)
            .first()
        )
        if not entry:
            raise HTTPException(status_code=404, detail="明細不存在")

        # 刪除也是改動來源：封存後禁刪，未封存則標 stale（先取年月再刪）。
        emp_id, yr, mo = entry.employee_id, entry.salary_year, entry.salary_month
        assert_months_not_finalized(session, employee_id=emp_id, months={(yr, mo)})
        session.delete(entry)
        session.flush()
        mark_salary_stale(session, emp_id, yr, mo)
        session.commit()
        return {"deleted": True, "id": entry_id}


IMPORT_TEMPLATE_HEADERS = [
    "員工姓名",
    "工號(選填)",
    "科目",
    "班級備註",
    "時數",
    "鐘點費",
    "超額",
    "加給活動",
    "備註",
]


@router.get("/art-teacher-payroll/import-template")
def download_import_template(
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
):
    """下載批次匯入範本 xlsx。"""
    wb = Workbook()
    ws = SafeWorksheet(wb.active)
    ws.title = "才藝薪資匯入範本"
    for col_idx, header in enumerate(IMPORT_TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    # 範例列
    ws.cell(row=2, column=1, value="歐瑞煌")
    ws.cell(row=2, column=3, value="美語")
    ws.cell(row=2, column=4, value="向.滿")
    ws.cell(row=2, column=5, value=25)
    ws.cell(row=2, column=6, value=550)
    ws.cell(row=3, column=1, value="鍾馨瑤")
    ws.cell(row=3, column=3, value="舞蹈")
    ws.cell(row=3, column=4, value="(二)")
    ws.cell(row=3, column=5, value=4)
    ws.cell(row=3, column=6, value=1000)
    ws.cell(row=3, column=7, value=200)

    raw_ws = wb.active
    for col, w in zip("ABCDEFGHI", (14, 12, 12, 14, 8, 10, 8, 10, 20)):
        raw_ws.column_dimensions[col].width = w

    buffer = BytesIO()
    wb.save(buffer)
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename*=UTF-8''"
                + quote("art_teacher_payroll_template.xlsx")
            )
        },
    )


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


@router.post("/art-teacher-payroll/batch-import")
async def batch_import(
    file: UploadFile = File(...),
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    replace_existing: bool = Query(
        False,
        description="True=匯入前先刪除該月所有 entries（整月覆寫，避免重複）",
    ),
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_WRITE)),
):
    """批次匯入才藝薪資明細。

    Excel 欄位：員工姓名 / 工號(選填) / 科目 / 班級備註 / 時數 / 鐘點費 / 超額 / 加給活動 / 備註

    回傳：{total, imported, skipped, errors:[{row, message}]}
    """
    content = await read_upload_with_size_check(file)
    validate_file_signature(content, ".xlsx")
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        logger.warning("才藝老師薪資 Excel 解析失敗", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail="無法解析 Excel 檔案，請確認檔案格式正確且未損壞",
        )

    ws = wb.active

    # R7-2：列數上限——避免超大表下游逐列處理耗盡資源（認證後 DoS）。read_only 串流
    # 因下方 ws[1] 隨機存取不適用，改以 max_row 拒絕超大檔（配合 10MB 上傳上限）。
    from utils.excel_io import MAX_IMPORT_ROWS

    if ws.max_row and ws.max_row > MAX_IMPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"匯入列數超過上限 {MAX_IMPORT_ROWS}，請分批匯入",
        )

    # 驗證表頭
    header_row = [_cell_str(c.value) for c in ws[1]]
    required = {"員工姓名", "科目", "時數", "鐘點費"}
    missing = required - set(header_row)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"範本缺少必填欄位：{', '.join(missing)}（請下載範本）",
        )
    col_idx = {h: i for i, h in enumerate(header_row)}

    results = {"total": 0, "imported": 0, "skipped": 0, "errors": []}

    with session_scope() as session:
        employees = (
            session.query(Employee).filter(Employee.employee_type == "hourly").all()
        )
        emp_by_name = {e.name: e for e in employees}
        emp_by_id = {str(e.employee_id): e for e in employees}

        # 月層封存守衛：批次匯入（尤其 replace_existing 為月寬刪除）會整體改動
        # 該月才藝薪資來源；若該月已有任一 hourly 員工薪資封存即拒絕（避免明細與
        # 已鎖定薪資紀錄/轉帳清冊/財報對不起來）。
        finalized_exists = (
            session.query(SalaryRecord.id)
            .join(Employee, Employee.id == SalaryRecord.employee_id)
            .filter(
                SalaryRecord.salary_year == year,
                SalaryRecord.salary_month == month,
                SalaryRecord.is_finalized == True,  # noqa: E712
                Employee.employee_type == "hourly",
            )
            .first()
        )
        if finalized_exists:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{year} 年 {month} 月已有才藝老師薪資封存，無法批次匯入。"
                    "請先至薪資管理頁面解除封存後再操作。"
                ),
            )

        # 受影響員工：匯入成功的列 + replace_existing 被刪除的舊 entries 員工，
        # 結束後全部標 needs_recalc，讓 finalize 完整性檢查擋下未重算的草稿。
        affected_emp_ids: set[int] = set()

        if replace_existing:
            affected_emp_ids.update(
                r[0]
                for r in session.query(ArtTeacherPayrollEntry.employee_id)
                .filter(
                    ArtTeacherPayrollEntry.salary_year == year,
                    ArtTeacherPayrollEntry.salary_month == month,
                )
                .distinct()
            )
            session.query(ArtTeacherPayrollEntry).filter(
                ArtTeacherPayrollEntry.salary_year == year,
                ArtTeacherPayrollEntry.salary_month == month,
            ).delete(synchronize_session=False)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            results["total"] += 1
            row_num = results["total"] + 1
            try:
                name = _cell_str(row[col_idx["員工姓名"]])
                if not name:
                    raise ValueError("員工姓名為空")
                emp = emp_by_name.get(name)
                if emp is None and "工號(選填)" in col_idx:
                    emp_id_str = _cell_str(row[col_idx["工號(選填)"]])
                    if emp_id_str:
                        emp = emp_by_id.get(emp_id_str)
                if emp is None:
                    raise ValueError(f"找不到 hourly 員工：{name}")

                subject = _cell_str(row[col_idx["科目"]])
                if not subject:
                    raise ValueError("科目為空")

                classroom = (
                    _cell_str(row[col_idx["班級備註"]]) if "班級備註" in col_idx else ""
                )
                hours = _cell_float(row[col_idx["時數"]])
                rate = _cell_float(row[col_idx["鐘點費"]])
                excess = _cell_float(row[col_idx["超額"]]) if "超額" in col_idx else 0
                activity = (
                    _cell_float(row[col_idx["加給活動"]])
                    if "加給活動" in col_idx
                    else 0
                )
                note = _cell_str(row[col_idx["備註"]]) if "備註" in col_idx else ""

                entry = ArtTeacherPayrollEntry(
                    employee_id=emp.id,
                    salary_year=year,
                    salary_month=month,
                    subject=subject,
                    classroom_label=classroom or None,
                    hours=hours,
                    hourly_rate=rate,
                    excess_amount=excess,
                    activity_bonus=activity,
                    note=note or None,
                    created_by=current_user.get("username"),
                    updated_by=current_user.get("username"),
                )
                recompute_entry_amounts(entry)
                session.add(entry)
                affected_emp_ids.add(emp.id)
                results["imported"] += 1
            except Exception as e:
                results["skipped"] += 1
                results["errors"].append({"row": row_num, "message": str(e)})

        session.flush()
        for emp_id in affected_emp_ids:
            mark_salary_stale(session, emp_id, year, month)

        session.commit()

    return results


@router.get("/art-teacher-payroll/{year}/{month}/roster")
def export_roster(
    year: int,
    month: int,
    current_user: dict = Depends(require_staff_permission(Permission.SALARY_READ)),
):
    """匯出才藝老師薪資清冊 xlsx（對齊《義華薪資》才藝老師 sheet）。"""
    _enforce_art_payroll_view(current_user)
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month 須介於 1~12")

    with session_scope() as session:
        filename, xlsx_bytes = generate_art_teacher_roster_xlsx(session, year, month)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (f"attachment; filename*=UTF-8''{quote(filename)}")
        },
    )
