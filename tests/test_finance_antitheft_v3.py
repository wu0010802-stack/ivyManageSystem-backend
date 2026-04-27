"""tests/test_finance_antitheft_v3.py — 2026-04-27 第三輪稽核 8 條守衛回歸。

涵蓋：
- #1 unfinalize_salary：reason 必填 ≥10 + ACTIVITY_PAYMENT_APPROVE 二人覆核
- #2 update_payment is_paid=True 補齊：method/reason 必填、拒絕系統補齊、大額簽核
- #3 學生離園自動沖帳：paid_amount > 0 時要求 ACTIVITY_PAYMENT_APPROVE
- #4 補打卡核准：核准後標 needs_recalc=True
- #5 force_overlap 假單：reason ≥10 + ACTIVITY_PAYMENT_APPROVE + ApprovalLog 痕跡
- #6 Course/Supply 價格：MAX_PAYMENT_AMOUNT 上限 + 高價需簽核
- #7 payment-report：voided 紀錄獨立標示
- #8 admin_waive：薪資端視為不扣 + 觸發 mark_salary_stale
"""

import os
import sys
from datetime import date, datetime, timedelta

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models.base as base_module
from api.activity import router as activity_router
from api.attendance import router as attendance_router
from api.auth import _account_failures, _ip_attempts
from api.auth import router as auth_router
from api.leaves import router as leaves_router
from api.punch_corrections import router as punch_corrections_router
from api.salary import router as salary_router
from api.students import router as students_router
from models.base import Base
from models.classroom import Classroom, Student
from models.database import (
    ActivityCourse,
    ActivityPaymentRecord,
    ActivityRegistration,
    ActivitySupply,
    Attendance,
    Employee,
    LeaveRecord,
    PunchCorrectionRequest,
    RegistrationCourse,
    SalaryRecord,
    User,
)
from utils.auth import hash_password
from utils.permissions import Permission

# ── Fixtures ────────────────────────────────────────────────────────────


@pytest.fixture
def v3_client(tmp_path):
    db_path = tmp_path / "antitheft_v3.sqlite"
    engine = create_engine(
        f"sqlite:///{db_path}", connect_args={"check_same_thread": False}
    )
    session_factory = sessionmaker(bind=engine)

    old_engine = base_module._engine
    old_session_factory = base_module._SessionFactory
    base_module._engine = engine
    base_module._SessionFactory = session_factory

    Base.metadata.create_all(engine)
    _ip_attempts.clear()
    _account_failures.clear()

    from api.activity import public as public_mod
    from api.activity import registrations as reg_mod
    from api.activity import pos as pos_mod

    for mod in (public_mod, reg_mod, pos_mod):
        for attr in dir(mod):
            obj = getattr(mod, attr)
            if hasattr(obj, "_timestamps"):
                obj._timestamps.clear()

    app = FastAPI()
    app.include_router(auth_router)
    app.include_router(salary_router)
    app.include_router(activity_router)
    app.include_router(students_router)
    app.include_router(leaves_router)
    app.include_router(attendance_router)
    app.include_router(punch_corrections_router)

    with TestClient(app) as client:
        yield client, session_factory

    _ip_attempts.clear()
    _account_failures.clear()
    base_module._engine = old_engine
    base_module._SessionFactory = old_session_factory
    engine.dispose()


def _make_user(
    session, *, username, permissions, employee_id=None, role="admin"
) -> User:
    user = User(
        username=username,
        password_hash=hash_password("Temp123456"),
        role=role,
        permissions=permissions,
        employee_id=employee_id,
        is_active=True,
        must_change_password=False,
    )
    session.add(user)
    session.flush()
    return user


def _login(client: TestClient, username: str):
    return client.post(
        "/api/auth/login", json={"username": username, "password": "Temp123456"}
    )


# ══════════════════════════════════════════════════════════════════════
# #1 unfinalize_salary 守衛
# ══════════════════════════════════════════════════════════════════════


def _seed_finalized_salary(session) -> int:
    emp = Employee(
        employee_id="E_unfin",
        name="封存測試",
        base_salary=30000,
        employee_type="regular",
        is_active=True,
    )
    session.add(emp)
    session.flush()
    rec = SalaryRecord(
        employee_id=emp.id,
        salary_year=2026,
        salary_month=3,
        base_salary=30000,
        gross_salary=30000,
        total_deduction=0,
        net_salary=30000,
        is_finalized=True,
        finalized_at=datetime(2026, 4, 5, 10, 0),
        finalized_by="hr_lead",
    )
    session.add(rec)
    session.flush()
    return rec.id


class TestUnfinalizeSalaryGuard:
    def test_no_body_rejected_422(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            rec_id = _seed_finalized_salary(s)
            _make_user(
                s,
                username="hr_admin",
                permissions=Permission.SALARY_READ
                | Permission.SALARY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "hr_admin").status_code == 200
        # 沒帶 body → 422
        res = client.request("DELETE", f"/api/salaries/{rec_id}/finalize")
        assert res.status_code == 422

    def test_short_reason_rejected_422(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            rec_id = _seed_finalized_salary(s)
            _make_user(
                s,
                username="hr_admin",
                permissions=Permission.SALARY_READ
                | Permission.SALARY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "hr_admin").status_code == 200
        res = client.request(
            "DELETE",
            f"/api/salaries/{rec_id}/finalize",
            json={"reason": "短"},
        )
        assert res.status_code == 422

    def test_no_finance_approve_rejected_403(self, v3_client):
        """有 SALARY_WRITE 與 admin/hr 角色但無 ACTIVITY_PAYMENT_APPROVE → 403"""
        client, sf = v3_client
        with sf() as s:
            rec_id = _seed_finalized_salary(s)
            _make_user(
                s,
                username="hr_only",
                permissions=Permission.SALARY_READ | Permission.SALARY_WRITE,
                role="hr",
            )
            s.commit()
        assert _login(client, "hr_only").status_code == 200
        res = client.request(
            "DELETE",
            f"/api/salaries/{rec_id}/finalize",
            json={"reason": "回補上游遲到資料後重新封存"},
        )
        assert res.status_code == 403
        assert "金流簽核" in res.json()["detail"]

    def test_unfinalize_succeeds_and_writes_audit(self, v3_client):
        """完整守衛通過：is_finalized=False、remark 記錄原因、cache invalidate"""
        client, sf = v3_client
        with sf() as s:
            rec_id = _seed_finalized_salary(s)
            _make_user(
                s,
                username="hr_admin",
                permissions=Permission.SALARY_READ
                | Permission.SALARY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "hr_admin").status_code == 200
        res = client.request(
            "DELETE",
            f"/api/salaries/{rec_id}/finalize",
            json={"reason": "回補上游遲到資料後重新封存"},
        )
        assert res.status_code == 200, res.text
        with sf() as s:
            rec = s.get(SalaryRecord, rec_id)
            assert rec.is_finalized is False
            assert "回補上游遲到資料後重新封存" in (rec.remark or "")
            assert "原封存：hr_lead" in (rec.remark or "")


# ══════════════════════════════════════════════════════════════════════
# #2 update_payment is_paid=True 補齊守衛
# ══════════════════════════════════════════════════════════════════════


def _seed_unpaid_registration(session, *, course_price=1500) -> int:
    cls = Classroom(name="補齊測試班", school_year=2025, semester=1)
    session.add(cls)
    session.flush()
    course = ActivityCourse(
        name="繪本課",
        price=course_price,
        capacity=10,
        school_year=2025,
        semester=1,
        is_active=True,
    )
    session.add(course)
    session.flush()
    reg = ActivityRegistration(
        student_name="補齊小朋友",
        parent_phone="0912345678",
        class_name=cls.name,
        classroom_id=cls.id,
        is_active=True,
        is_paid=False,
        paid_amount=0,
        school_year=2025,
        semester=1,
    )
    session.add(reg)
    session.flush()
    rc = RegistrationCourse(
        registration_id=reg.id,
        course_id=course.id,
        price_snapshot=course_price,
        status="enrolled",
    )
    session.add(rc)
    session.flush()
    return reg.id


class TestUpdatePaymentMarkPaidGuard:
    def test_missing_method_rejected_400(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            reg_id = _seed_unpaid_registration(s, course_price=500)
            _make_user(
                s,
                username="act_admin",
                permissions=Permission.ACTIVITY_READ
                | Permission.ACTIVITY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "act_admin").status_code == 200
        res = client.put(
            f"/api/activity/registrations/{reg_id}/payment",
            json={"is_paid": True},
        )
        assert res.status_code == 400
        assert "payment_method" in res.json()["detail"]

    def test_system_reconcile_method_rejected_400(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            reg_id = _seed_unpaid_registration(s, course_price=500)
            _make_user(
                s,
                username="act_admin",
                permissions=Permission.ACTIVITY_READ
                | Permission.ACTIVITY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "act_admin").status_code == 200
        res = client.put(
            f"/api/activity/registrations/{reg_id}/payment",
            json={
                "is_paid": True,
                "payment_method": "系統補齊",
                "payment_reason": "後台對帳補齊",
            },
        )
        assert res.status_code == 400
        assert "系統補齊" in res.json()["detail"]

    def test_large_shortfall_without_approve_rejected_403(self, v3_client):
        """shortfall > FINANCE_APPROVAL_THRESHOLD（1000）時要求金流簽核"""
        client, sf = v3_client
        with sf() as s:
            reg_id = _seed_unpaid_registration(s, course_price=2000)
            _make_user(
                s,
                username="act_writer",
                permissions=Permission.ACTIVITY_READ | Permission.ACTIVITY_WRITE,
                role="staff",
            )
            s.commit()
        assert _login(client, "act_writer").status_code == 200
        res = client.put(
            f"/api/activity/registrations/{reg_id}/payment",
            json={
                "is_paid": True,
                "payment_method": "現金",
                "payment_reason": "家長現金繳清",
            },
        )
        assert res.status_code == 403
        assert "審批閾值" in res.json()["detail"]

    def test_full_guard_passes_200(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            reg_id = _seed_unpaid_registration(s, course_price=500)
            _make_user(
                s,
                username="act_admin",
                permissions=Permission.ACTIVITY_READ
                | Permission.ACTIVITY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "act_admin").status_code == 200
        res = client.put(
            f"/api/activity/registrations/{reg_id}/payment",
            json={
                "is_paid": True,
                "payment_method": "現金",
                "payment_reason": "家長已現金繳清",
            },
        )
        assert res.status_code == 200, res.text
        with sf() as s:
            rec = (
                s.query(ActivityPaymentRecord)
                .filter(ActivityPaymentRecord.registration_id == reg_id)
                .first()
            )
            assert rec is not None
            assert rec.payment_method == "現金"


# ══════════════════════════════════════════════════════════════════════
# #6 Course/Supply 價格上限與大額簽核
# ══════════════════════════════════════════════════════════════════════


class TestActivityItemPriceGuard:
    def test_course_price_above_max_rejected_422(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            _make_user(
                s,
                username="act_admin",
                permissions=Permission.ACTIVITY_READ
                | Permission.ACTIVITY_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
        assert _login(client, "act_admin").status_code == 200
        res = client.post(
            "/api/activity/courses",
            json={"name": "高價課", "price": 1_000_000, "capacity": 10},
        )
        assert res.status_code == 422

    def test_course_price_high_without_approve_rejected_403(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            _make_user(
                s,
                username="act_writer",
                permissions=Permission.ACTIVITY_READ | Permission.ACTIVITY_WRITE,
                role="staff",
            )
            s.commit()
        assert _login(client, "act_writer").status_code == 200
        res = client.post(
            "/api/activity/courses",
            json={"name": "高價課", "price": 50_000, "capacity": 10},
        )
        assert res.status_code == 403
        assert "審批閾值" in res.json()["detail"]

    def test_supply_normal_price_allowed(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            _make_user(
                s,
                username="act_writer",
                permissions=Permission.ACTIVITY_READ | Permission.ACTIVITY_WRITE,
                role="staff",
            )
            s.commit()
        assert _login(client, "act_writer").status_code == 200
        res = client.post(
            "/api/activity/supplies",
            json={"name": "繪本", "price": 500},
        )
        assert res.status_code == 201, res.text


# ══════════════════════════════════════════════════════════════════════
# #4 補打卡核准觸發 mark_salary_stale
# ══════════════════════════════════════════════════════════════════════


class TestPunchCorrectionMarksSalaryStale:
    def test_approve_correction_marks_salary_record_stale(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            emp = Employee(
                employee_id="E_pc",
                name="補打卡測試",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            rec = SalaryRecord(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=3,
                base_salary=30000,
                gross_salary=30000,
                total_deduction=0,
                net_salary=30000,
                is_finalized=False,
                needs_recalc=False,
            )
            s.add(rec)
            s.flush()
            corr = PunchCorrectionRequest(
                employee_id=emp.id,
                attendance_date=date(2026, 3, 15),
                correction_type="punch_in",
                requested_punch_in=datetime(2026, 3, 15, 9, 0),
                reason="忘了打卡",
            )
            s.add(corr)
            s.flush()
            _make_user(
                s,
                username="pc_admin",
                permissions=Permission.ATTENDANCE_READ
                | Permission.ATTENDANCE_WRITE
                | Permission.APPROVALS,
                role="admin",
            )
            s.commit()
            corr_id = corr.id
            rec_id = rec.id

        assert _login(client, "pc_admin").status_code == 200
        res = client.put(
            f"/api/punch-corrections/{corr_id}/approve",
            json={"approved": True},
        )
        assert res.status_code == 200, res.text
        with sf() as s:
            rec = s.get(SalaryRecord, rec_id)
            assert rec.needs_recalc is True


# ══════════════════════════════════════════════════════════════════════
# #8 admin_waive 薪資端視為不扣 + mark_salary_stale
# ══════════════════════════════════════════════════════════════════════


class TestAdminWaiveSalaryRespect:
    def test_waived_attendance_excluded_from_late_count(self):
        """考勤端標 admin_waive 後，薪資端 _calc_attendance_stats 應排除該日。"""
        from services.salary_field_breakdown import _calc_attendance_stats

        class FakeAtt:
            def __init__(self, *, is_late=False, late_minutes=0, confirmed_action=None):
                self.is_late = is_late
                self.is_early_leave = False
                self.is_missing_punch_in = False
                self.is_missing_punch_out = False
                self.late_minutes = late_minutes
                self.early_leave_minutes = 0
                self.attendance_date = date(2026, 3, 15)
                self.confirmed_action = confirmed_action

        attendances = [
            FakeAtt(is_late=True, late_minutes=10),
            FakeAtt(
                is_late=True,
                late_minutes=20,
                confirmed_action="admin_waive",
            ),
            FakeAtt(is_late=True, late_minutes=5, confirmed_action="admin_accept"),
        ]
        stats = _calc_attendance_stats(attendances)
        # admin_waive 那筆完全排除：count=2、minutes=15
        assert stats["late_count"] == 2
        assert stats["total_late_min"] == 15

    def test_batch_confirm_waive_marks_salary_stale(self, v3_client):
        client, sf = v3_client
        with sf() as s:
            emp = Employee(
                employee_id="E_wv",
                name="豁免測試",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            att = Attendance(
                employee_id=emp.id,
                attendance_date=date(2026, 3, 10),
                is_late=True,
                late_minutes=15,
            )
            s.add(att)
            s.flush()
            rec = SalaryRecord(
                employee_id=emp.id,
                salary_year=2026,
                salary_month=3,
                base_salary=30000,
                gross_salary=30000,
                total_deduction=0,
                net_salary=30000,
                is_finalized=False,
                needs_recalc=False,
            )
            s.add(rec)
            s.flush()
            _make_user(
                s,
                username="att_admin",
                permissions=Permission.ATTENDANCE_READ | Permission.ATTENDANCE_WRITE,
                role="admin",
            )
            s.commit()
            att_id = att.id
            rec_id = rec.id

        assert _login(client, "att_admin").status_code == 200
        res = client.post(
            "/api/attendance/anomalies/batch-confirm",
            json={"attendance_ids": [att_id], "action": "admin_waive"},
        )
        assert res.status_code == 200, res.text
        with sf() as s:
            rec = s.get(SalaryRecord, rec_id)
            assert rec.needs_recalc is True

    def test_engine_load_attendance_excludes_waived_path(self, v3_client):
        """整合測試：engine._load_attendance_result 對 admin_waive 那筆完全排除。"""
        from datetime import date as _d, datetime as _dt
        from services.salary.engine import SalaryEngine

        client, sf = v3_client
        with sf() as s:
            emp = Employee(
                employee_id="E_eng_wv",
                name="引擎豁免測試",
                base_salary=30000,
                employee_type="regular",
                is_active=True,
            )
            s.add(emp)
            s.flush()
            # 一筆未豁免的遲到 + 一筆 admin_waive 的遲到
            s.add(
                Attendance(
                    employee_id=emp.id,
                    attendance_date=_d(2026, 3, 5),
                    is_late=True,
                    late_minutes=10,
                )
            )
            s.add(
                Attendance(
                    employee_id=emp.id,
                    attendance_date=_d(2026, 3, 12),
                    is_late=True,
                    late_minutes=20,
                    confirmed_action="admin_waive",
                )
            )
            s.commit()
            emp_id = emp.id

        engine = SalaryEngine()
        with sf() as s:
            emp = s.get(Employee, emp_id)
            emp_dict = {}
            result, _ = engine._load_attendance_result(
                s, emp, _d(2026, 3, 1), _d(2026, 3, 31), emp_dict
            )
        # admin_waive 那筆排除：late_count=1、total_late_minutes=10
        assert result.late_count == 1
        assert result.total_late_minutes == 10
        # late_details 也只剩未豁免的 10 分鐘
        assert emp_dict["_late_details"] == [10]


# ══════════════════════════════════════════════════════════════════════
# #7 payment-report：voided 紀錄獨立標示
# ══════════════════════════════════════════════════════════════════════


class TestPaymentReportVoidedHandling:
    def test_voided_payment_marked_in_export(self, v3_client):
        """匯出檔的繳費明細工作表會把 voided 紀錄類型標「（已作廢）」並列出作廢人/原因。"""
        import io
        from openpyxl import load_workbook

        client, sf = v3_client
        with sf() as s:
            reg_id = _seed_unpaid_registration(s, course_price=500)
            # 一筆有效繳費 + 一筆已作廢繳費
            valid = ActivityPaymentRecord(
                registration_id=reg_id,
                type="payment",
                amount=300,
                payment_date=date(2026, 4, 1),
                payment_method="現金",
                operator="staff_a",
            )
            voided = ActivityPaymentRecord(
                registration_id=reg_id,
                type="payment",
                amount=200,
                payment_date=date(2026, 4, 2),
                payment_method="現金",
                operator="staff_b",
                voided_at=datetime(2026, 4, 3, 12, 0),
                voided_by="admin",
                void_reason="重複輸入收據",
            )
            s.add(valid)
            s.add(voided)
            _make_user(
                s,
                username="act_admin",
                permissions=Permission.ACTIVITY_READ | Permission.ACTIVITY_WRITE,
                role="admin",
            )
            s.commit()

        assert _login(client, "act_admin").status_code == 200
        res = client.get("/api/activity/registrations/payment-report")
        assert res.status_code == 200
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb["繳費明細"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][2] == "類型" and "作廢狀態" in rows[0]
        # 找作廢列
        voided_rows = [r for r in rows[1:] if r[2] == "繳費（已作廢）"]
        valid_rows = [r for r in rows[1:] if r[2] == "繳費"]
        assert len(voided_rows) == 1
        assert len(valid_rows) == 1
        assert voided_rows[0][8] == "已作廢"  # 作廢狀態欄
        assert voided_rows[0][9] == "admin"  # 作廢人
        assert voided_rows[0][11] == "重複輸入收據"  # 作廢原因


# ══════════════════════════════════════════════════════════════════════
# #3 學生離園 paid_amount > 0 須具金流簽核
# ══════════════════════════════════════════════════════════════════════


class TestStudentDeactivateRefundGuard:
    def test_paid_registration_blocks_delete_without_approve(self, v3_client):
        """STUDENTS_WRITE 但無 ACTIVITY_PAYMENT_APPROVE → 刪除有付費報名的學生 403"""
        from utils.academic import resolve_current_academic_term

        _sy, _sem = resolve_current_academic_term()
        client, sf = v3_client
        with sf() as s:
            cls = Classroom(name="守衛班", school_year=_sy, semester=_sem)
            s.add(cls)
            s.flush()
            st = Student(
                student_id="S_PAID",
                name="付費小朋友",
                is_active=True,
                classroom_id=cls.id,
            )
            s.add(st)
            s.flush()
            reg = ActivityRegistration(
                student_id=st.id,
                student_name=st.name,
                parent_phone="0911000000",
                class_name=cls.name,
                classroom_id=cls.id,
                is_active=True,
                is_paid=True,
                paid_amount=1500,
                school_year=_sy,
                semester=_sem,
            )
            s.add(reg)
            _make_user(
                s,
                username="stu_writer",
                permissions=Permission.STUDENTS_READ | Permission.STUDENTS_WRITE,
                role="staff",
            )
            s.commit()
            student_id = st.id

        assert _login(client, "stu_writer").status_code == 200
        res = client.delete(f"/api/students/{student_id}")
        assert res.status_code == 403
        assert "金流簽核" in res.json()["detail"] or "簽核" in res.json()["detail"]

    def test_paid_registration_allows_delete_with_approve(self, v3_client):
        from utils.academic import resolve_current_academic_term

        _sy, _sem = resolve_current_academic_term()
        client, sf = v3_client
        with sf() as s:
            cls = Classroom(name="守衛班2", school_year=_sy, semester=_sem)
            s.add(cls)
            s.flush()
            st = Student(
                student_id="S_PAID2",
                name="付費二號",
                is_active=True,
                classroom_id=cls.id,
            )
            s.add(st)
            s.flush()
            reg = ActivityRegistration(
                student_id=st.id,
                student_name=st.name,
                parent_phone="0911000001",
                class_name=cls.name,
                classroom_id=cls.id,
                is_active=True,
                is_paid=True,
                paid_amount=500,
                school_year=_sy,
                semester=_sem,
            )
            s.add(reg)
            _make_user(
                s,
                username="stu_admin",
                permissions=Permission.STUDENTS_READ
                | Permission.STUDENTS_WRITE
                | Permission.ACTIVITY_PAYMENT_APPROVE,
                role="admin",
            )
            s.commit()
            student_id = st.id

        assert _login(client, "stu_admin").status_code == 200
        res = client.delete(f"/api/students/{student_id}")
        assert res.status_code == 200, res.text


# ══════════════════════════════════════════════════════════════════════
# #5 force_overlap 假單守衛（schema 層）
# ══════════════════════════════════════════════════════════════════════


class TestForceOverlapSchemaGuard:
    def test_force_overlap_requires_reason(self):
        """force_overlap=True 但無 reason → ValidationError"""
        from api.leaves import ApproveRequest

        with pytest.raises(Exception) as exc:
            ApproveRequest(approved=True, force_overlap=True)
        assert "force_overlap_reason" in str(exc.value) or "10" in str(exc.value)

    def test_force_overlap_short_reason_rejected(self):
        from api.leaves import ApproveRequest

        with pytest.raises(Exception) as exc:
            ApproveRequest(approved=True, force_overlap=True, force_overlap_reason="短")
        assert "10" in str(exc.value)

    def test_force_overlap_with_proper_reason_ok(self):
        from api.leaves import ApproveRequest

        req = ApproveRequest(
            approved=True,
            force_overlap=True,
            force_overlap_reason="園長確認補休重疊為合法行政安排",
        )
        assert req.force_overlap is True
        assert req.force_overlap_reason.startswith("園長")
